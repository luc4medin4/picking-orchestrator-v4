"""
Picking Orchestrator v4.0 — Beccacece Hnos SA
Streamlit unificado para automatización de picking (DPO 2.1 — Pilar Almacén)

Sucesor de app_v3_9_1.py (Planilla de Carga). No reemplaza la v3.9: se ejecuta
en paralelo. Reutiliza literalmente las funciones de la v3.9 para Tabs 1 y 2.

SPRINT 1 (este archivo):
  - Esqueleto con 6 tabs + sidebar (toggles dry-run y snapshot)
  - Tab 1: Planilla de Carga (reuso 100% v3.9 — generate_pdf)
  - Tab 2: Resumen por Camión (reuso 100% v3.9 — build_resumen_carga_pdf)
  - Tab 5: Extraíbles Sheets (Matriz Pall., Agregados AE, Reposición AE filtrada J<0, fx Picking)
  - Tabs 3, 4, 6: placeholders con scope documentado (Sprints 2-5)
  - Snapshot local provisional ./snapshots/YYYY-MM-DD/<run_id>/

ARQUITECTURA:
  - 1 solo archivo .py autocontenido (decisión Lucas)
  - Stack: streamlit, pandas, openpyxl, reportlab, loguru
  - MASTER (78 MB) leído con openpyxl(data_only=True, read_only=True)

EJECUCIÓN:
  pip install streamlit pandas openpyxl reportlab loguru
  streamlit run app_v4_0.py
"""
import io, math, hashlib, unicodedata, os, shutil, json
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from loguru import logger

# ─── VERSIÓN Y CONFIG GLOBAL ────────────────────────────────────────────────
APP_VERSION = "4.7.0"
SNAPSHOT_DIR = Path("./snapshots")

# Colores T2 (Sprint 3)
T2_COLORS = {
    "verde": "#22c55e",     # Pallet pura completa
    "amarillo": "#eab308",  # Pallet picking (parcial)
    "gris": "#9ca3af",      # Vacío
    "rojo": "#ef4444",      # Alerta (excedido)
}

# Setup logger único (no duplicar handlers en re-runs Streamlit)
if "logger_configured" not in st.session_state:
    logger.remove()
    logger.add(
        lambda m: None,  # Stdout silenciado en Streamlit (lo capturamos en sesión)
        level="INFO",
        format="{time:HH:mm:ss} | {level} | {message}",
    )
    st.session_state["logger_configured"] = True
    st.session_state["log_buffer"] = []

def log_event(level: str, msg: str):
    """Log a archivo (futuro) y a buffer de sesión (para mostrar en UI)."""
    ts = datetime.now().strftime("%H:%M:%S")
    entry = f"[{ts}] {level.upper()}: {msg}"
    st.session_state.setdefault("log_buffer", []).append(entry)
    logger.log(level.upper(), msg)

# ============================================================================
# ║              BLOQUE REUSADO LITERAL DE app_v3_9_1.py                     ║
# ║                                                                          ║
# ║  Funciones críticas para Tabs 1 y 2:                                     ║
# ║   - load_car, load_frescura, load_agr                                    ║
# ║   - generate_pdf (Planilla de Carga)                                     ║
# ║   - build_resumen_carga_pdf (Resumen por Camión)                         ║
# ║   + todas las helpers de dibujo del PDF (draw_*, header_*, etc)          ║
# ║                                                                          ║
# ║  ⚠ NO MODIFICAR — lógica validada en producción.                         ║
# ============================================================================

import io, math, hashlib, unicodedata
from datetime import datetime

import pandas as pd
import streamlit as st
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas

# ─── MEDIDAS FIJAS (puntos) ──────────────────────────────────────────────────
PAGE_W, PAGE_H = A4
MARGIN  = 12 * mm
CW      = PAGE_W - 2 * MARGIN

H_TITLE     = 26
H_TRANS     = 18
H_PARTIDA   = 16
H_CTRL      = 56
H_LEMA      = 18
H_CANCHA    = 20
H_THDR      = 14
H_ALM       = 16
H_ROW       = 14
H_TOT       = 14
H_ESPACIO   = 62
H_ROUTE     = 16
H_FOOT      = 28
GAP         = 2

# ─── COLORES ────────────────────────────────────────────────────────────────
DARK_BLUE  = colors.HexColor('#1a3a6b')
MED_BLUE   = colors.HexColor('#2e5fa3')
LIGHT_GRAY = colors.HexColor('#EFEFEF')
BG_GRAY    = colors.HexColor('#F4F4F4')
DARK_ROW   = colors.HexColor('#AAAAAA')
MED_ROW    = colors.HexColor('#D8D8D8')
AMBER      = colors.HexColor('#FFE0B2')
RED_ALERT  = colors.HexColor('#C0392B')
FOOT_GRAY  = colors.HexColor('#555555')
BORDER     = colors.HexColor('#AAAAAA')
HDR_BG     = colors.HexColor('#E8E8E8')
YELLOW_RTE = colors.HexColor('#FFF3A0')
WM_GRAY    = colors.HexColor('#888888')
ALM_BG     = colors.HexColor('#DDEEFF')

EXCLUDED_SKUS = {2730, 2731, 2776, 2780, 5192}
EXCLUDED_PATS = ['Q CERVEZAS', 'Q PLAS', 'BOT 1/1', 'BOT AMBAR', 'ARACELI']
CANCHA_ORDER  = ['CANCHA I', 'CANCHA II', 'CANCHA III', 'CANCHA IV', 'CANCHA V']
COL_W = [50, CW - 50 - 62 - 52 - 52, 62, 52, 52]

LEMAS = [
    "Cada bulto bien puesto es una entrega perfecta. Dale con todo!",
    "El picking preciso empieza con vos. Conta, verifica, confia.",
    "Hoy es dia de hacer historia en el deposito. A romperla!",
    "Rapido no es apurado. Rapido es preciso y sin errores.",
    "El mejor pickero no es el mas veloz, es el mas exacto.",
    "Tu equipo cuenta con vos. Cada bulto importa.",
    "Zapatos abrochados, espalda cuidada, mente enfocada. Arrancamos!",
    "Paleta perfecta = cliente feliz = empresa fuerte. Todo empieza aca.",
    "Cero errores de picking es el estandar. Vos podes.",
    "Si dudas del SKU, verifica. Un segundo de control evita una hora de correccion.",
    "La diferencia entre bueno y excelente esta en los detalles. Se excelente.",
    "Arrancar bien es terminar mejor. Que empiece el picking!",
    "Trabajo limpio, paleta ordenada, conciencia tranquila.",
    "Sos parte del corazon de la operacion. El deposito te necesita al 100%.",
    "Pensa antes de mover. El orden en la paleta es orden en la entrega.",
    "Tecnica correcta al levantar = cuerpo sano manana tambien.",
    "Verifica el SKU, verifica la cantidad, verifica el lote. Tres checks, cero errores!",
    "Cada noche de picking sin errores construye la reputacion de todos.",
    "El deposito funciona porque vos funcionas. Gracias por el compromiso.",
    "Paleta bien armada, camion bien cargado, cliente bien atendido.",
    "El cansancio es real. El orgullo del trabajo bien hecho tambien.",
    "Reporta lo que esta mal antes de moverlo. Un reporte a tiempo vale oro.",
    "Somos un equipo. Si uno falla, fallamos todos. Si uno gana, ganamos todos.",
    "Segui la planilla al pie de la letra. Esta hecha para que llegues sin errores.",
    "Esta noche vas a hacer 500 movimientos correctos. Demostralo!",
    "Seguridad primero: EPI puesto, pasillo libre, mente clara.",
    "Los mejores pickeros no nacen, se hacen en noches como esta.",
    "Picking nocturno = concentracion maxima. El dia descansa, vos construis.",
    "Cada producto en su lugar es una promesa cumplida al cliente.",
    "Ultima caja, mismo cuidado que la primera. Asi se hace el trabajo bien hecho.",
]

# ─── UTILIDADES ─────────────────────────────────────────────────────────────

def _norm(s):
    """Normaliza string: sin tildes, sin espacios extra, lowercase."""
    if s is None: return ''
    s = str(s).strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.lower()

def find_col(df, *candidates):
    """Busca columna en df por nombre normalizado (tolerante a tildes/case)."""
    norm_map = {_norm(c): c for c in df.columns}
    for cand in candidates:
        c = _norm(cand)
        if c in norm_map:
            return norm_map[c]
    return None

def to_int_sku(val):
    """Convierte SKU a int de forma robusta. Devuelve None si no se puede."""
    if val is None: return None
    if isinstance(val, float) and math.isnan(val): return None
    try:
        s = str(val).strip()
        if s in ('', 'nan', 'NaN', 'None'): return None
        return int(float(s))
    except (ValueError, TypeError):
        return None

def is_envase(sku, desc):
    if sku in EXCLUDED_SKUS: return True
    d = str(desc).upper()
    for p in EXCLUDED_PATS:
        if p in d: return True
    return d.strip() == 'RET'

def normalize_cancha(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 'SIN CANCHA'
    s = str(val).strip().upper()
    if s in ('', '0', 'NAN', 'NONE'): return 'SIN CANCHA'
    # Eliminar tildes
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    MAP = {'CANCHA I':'CANCHA I','CANCHA II':'CANCHA II','CANCHA III':'CANCHA III',
           'CANCHA IV':'CANCHA IV','CANCHA V':'CANCHA V',
           'MKPL':'CANCHA V','MERCH':'CANCHA V','MERCHANDISING':'CANCHA V',
           'MARKETPLACE':'CANCHA V'}
    # Match exacto primero
    if s in MAP: return MAP[s]
    # Match por prefijo (más específico primero)
    for k in sorted(MAP.keys(), key=len, reverse=True):
        if s.startswith(k): return MAP[k]
    return 'SIN CANCHA'

# ─── CARGA DE DATOS ─────────────────────────────────────────────────────────

def load_car(file, ddm=None):
    """
    v3.8 — Si se pasa ddm, se usa para decidir si los huérfanos del bloque azul
    son paletas enteras (no se agregan) o tienen fracción / no tienen BXP
    (se agregan al df). Si ddm es None, no se procesan huérfanos (compat retro).

    Retorna: (df, blue_audit) donde blue_audit es una lista de dicts con la
    trazabilidad de cada fila del bloque azul y su destino.
    """
    # ── Leer hoja completa sin header para manejar filas por índice ──────────
    raw = pd.read_excel(file, sheet_name=0, header=None)

    # Fila 0 = header (Excel fila 1)
    header = raw.iloc[0].tolist()

    # ── FILAS AZULES: Excel filas 2-41 → índices 1-40 ────────────────────────
    # Columna T = índice 19 (A=0 … T=19)
    blue_raw = raw.iloc[1:41].copy()
    blue_raw.columns = header

    blue_raw["Artículo"]   = pd.to_numeric(blue_raw["Artículo"], errors="coerce")
    blue_raw["Transporte"] = pd.to_numeric(blue_raw["Transporte"], errors="coerce")
    blue_raw["_bultos_azul"] = pd.to_numeric(blue_raw.iloc[:, 19], errors="coerce").fillna(0)
    blue_raw = blue_raw.dropna(subset=["Artículo", "Transporte"])
    blue_raw = blue_raw[blue_raw["Artículo"] > 0]
    blue_raw = blue_raw[blue_raw["_bultos_azul"] > 0]
    blue_raw["Artículo"]   = blue_raw["Artículo"].astype(int)
    blue_raw["Transporte"] = blue_raw["Transporte"].apply(lambda x: int(float(x)) if pd.notna(x) else x)
    # Excluir envases de las filas azules
    blue_raw = blue_raw[~blue_raw.apply(
        lambda r: is_envase(int(r["Artículo"]), str(r.get("Descripción Artículo", ""))), axis=1
    )]

    # Guardar descripción del primer registro azul por SKU (para huérfanos)
    blue_desc_map = (blue_raw.dropna(subset=["Descripción Artículo"])
                     .drop_duplicates(subset=["Artículo"], keep="first")
                     .set_index("Artículo")["Descripción Artículo"].to_dict())

    # Agregar bultos azules por Transporte + SKU
    blue_agg = (blue_raw.groupby(["Transporte", "Artículo"], as_index=False)
                .agg(_bultos_azul=("_bultos_azul", "sum")))
    blue_total_in = float(blue_agg["_bultos_azul"].sum()) if len(blue_agg) else 0.0

    # ── FILAS NORMALES: Excel fila 42 en adelante → índice 41+ ───────────────
    normal_raw = raw.iloc[41:].copy()
    normal_raw.columns = header

    df = normal_raw.copy()
    df["Artículo"]   = pd.to_numeric(df["Artículo"], errors="coerce")
    df["Transporte"] = pd.to_numeric(df["Transporte"], errors="coerce")
    df = df.dropna(subset=["Artículo", "Transporte"])
    df = df[df["Artículo"] > 0]
    df["Artículo"]   = df["Artículo"].astype(int)
    df["Transporte"] = df["Transporte"].apply(lambda x: int(float(x)) if pd.notna(x) else x)
    df["Bultos"]     = pd.to_numeric(df["Bultos"], errors="coerce").fillna(0)
    df["Unids"]      = pd.to_numeric(df["Unids"],  errors="coerce").fillna(0)
    df = df[~df.apply(lambda r: is_envase(r["Artículo"], r["Descripción Artículo"]), axis=1)]

    # Inicializar flag de huérfano sin DDM (siempre False por defecto)
    df["_orphan_no_ddm"] = False
    blue_audit = []

    # ── MERGE: sumar bultos azules sobre los normales ─────────────────────────
    blue_consumed_merge = 0.0
    blue_consumed_orphans = 0.0
    blue_discarded_pal_entera = 0.0

    if not blue_agg.empty:
        # Set de pares (T, SKU) que existen en filas normales
        normal_keys = set(zip(df["Transporte"], df["Artículo"]))

        # Partir blue_agg en dos: los que tienen match y los huérfanos
        blue_agg["_has_match"] = blue_agg.apply(
            lambda r: (r["Transporte"], r["Artículo"]) in normal_keys, axis=1)

        blue_match    = blue_agg[blue_agg["_has_match"]].copy()
        blue_orphans  = blue_agg[~blue_agg["_has_match"]].copy()

        # ── Caso 1: con match → merge clásico (sumar bultos azules) ───────────
        if not blue_match.empty:
            blue_match_clean = blue_match[["Transporte","Artículo","_bultos_azul"]]
            df = df.merge(blue_match_clean, on=["Transporte", "Artículo"], how="left")
            df["_bultos_azul"] = df["_bultos_azul"].fillna(0)
            df["Bultos"] = df["Bultos"] + df["_bultos_azul"]
            blue_consumed_merge = float(blue_match_clean["_bultos_azul"].sum())
            df = df.drop(columns=["_bultos_azul"])

            # Trazabilidad: filas sumadas
            for _, br in blue_match.iterrows():
                blue_audit.append({
                    "Transporte": int(br["Transporte"]),
                    "SKU":        int(br["Artículo"]),
                    "Descripción": str(blue_desc_map.get(int(br["Artículo"]), "")),
                    "Bultos azules": float(br["_bultos_azul"]),
                    "BXP": None,
                    "Acción": "Sumado a fila normal",
                })

        # ── Caso 2: huérfanos → decidir según DDM ─────────────────────────────
        if not blue_orphans.empty and ddm is not None:
            # Tomar info de reparto y descripción de transporte de cualquier
            # fila normal del mismo transporte (siempre existe)
            transp_info = (df.dropna(subset=["Número"])
                           .drop_duplicates(subset=["Transporte"], keep="first")
                           .set_index("Transporte")
                           [["Número","Fecha Mvto","Descripción Transporte","Depósito","Descripción Depósito"]]
                           .to_dict("index"))

            new_rows = []
            for _, br in blue_orphans.iterrows():
                t   = int(br["Transporte"]); sku = int(br["Artículo"])
                blt = float(br["_bultos_azul"])
                desc = str(blue_desc_map.get(sku, f"SKU {sku}"))

                d = ddm[ddm["sku"] == sku]
                bxp = float(d.iloc[0]["bxp"]) if len(d) and pd.notna(d.iloc[0]["bxp"]) else None

                if bxp and bxp > 0 and (blt % bxp == 0):
                    # Paleta entera exacta → NO agregar (no es picking)
                    blue_discarded_pal_entera += blt
                    blue_audit.append({
                        "Transporte": t, "SKU": sku, "Descripción": desc,
                        "Bultos azules": blt, "BXP": bxp,
                        "Acción": "Descartado (paleta entera)",
                    })
                    continue

                # Necesita aparecer en PDF: tiene fracción o no tiene BXP
                if t not in transp_info:
                    # No hay info de reparto → no se puede asignar a un reparto
                    # No debería pasar nunca pero registramos
                    blue_audit.append({
                        "Transporte": t, "SKU": sku, "Descripción": desc,
                        "Bultos azules": blt, "BXP": bxp,
                        "Acción": "ERROR: sin info de reparto",
                    })
                    continue

                info = transp_info[t]
                no_ddm = (bxp is None or bxp <= 0)

                new_row = {col: None for col in df.columns}
                new_row["Depósito"]              = info.get("Depósito")
                new_row["Descripción Depósito"]  = info.get("Descripción Depósito")
                new_row["Número"]                = info["Número"]
                new_row["Fecha Mvto"]            = info["Fecha Mvto"]
                new_row["Transporte"]            = t
                new_row["Descripción Transporte"]= info["Descripción Transporte"]
                new_row["Artículo"]              = sku
                new_row["Descripción Artículo"]  = desc
                new_row["Bultos"]                = blt
                new_row["Unids"]                 = 0
                new_row["_orphan_no_ddm"]        = no_ddm

                new_rows.append(new_row)
                blue_consumed_orphans += blt

                blue_audit.append({
                    "Transporte": t, "SKU": sku, "Descripción": desc,
                    "Bultos azules": blt, "BXP": bxp,
                    "Acción": ("Agregado como picking (SIN BXP — REVISAR)"
                               if no_ddm else "Agregado como picking (fracción)"),
                })

            if new_rows:
                # NOTA: el header del CAR puede tener columnas con nombres
                # duplicados (espacios non-breaking). Para evitar fallar en
                # pd.concat con InvalidIndexError, construimos el dataframe
                # nuevo respetando posiciones de columnas, no nombres.
                new_df = pd.DataFrame(new_rows)
                # Reindexar usando solo las columnas que SÍ existen en df
                # (asignación posicional por nombre único)
                aligned = pd.DataFrame(columns=df.columns)
                for nr in new_rows:
                    aligned_row = []
                    for col in df.columns:
                        # Si la columna está repetida tomamos el primer valor encontrado
                        aligned_row.append(nr.get(col, None))
                    aligned.loc[len(aligned)] = aligned_row
                df = pd.concat([df, aligned], ignore_index=True, sort=False)

        elif not blue_orphans.empty and ddm is None:
            # Sin DDM no podemos decidir → registrar pero no agregar (compat)
            for _, br in blue_orphans.iterrows():
                blue_audit.append({
                    "Transporte": int(br["Transporte"]),
                    "SKU":        int(br["Artículo"]),
                    "Descripción": str(blue_desc_map.get(int(br["Artículo"]), "")),
                    "Bultos azules": float(br["_bultos_azul"]),
                    "BXP": None,
                    "Acción": "PERDIDO (sin DDM para decidir)",
                })

    # ── Invariante de validación ─────────────────────────────────────────────
    blue_total_out = (blue_consumed_merge + blue_consumed_orphans
                      + blue_discarded_pal_entera)
    if blue_total_in > 0 and abs(blue_total_in - blue_total_out) > 0.01 and ddm is not None:
        raise ValueError(
            f"Invariante de bloque azul rota: entrada={blue_total_in:.0f} "
            f"bultos, salida={blue_total_out:.0f} bultos "
            f"(merge={blue_consumed_merge:.0f} + huérfanos={blue_consumed_orphans:.0f} "
            f"+ descartados pal entera={blue_discarded_pal_entera:.0f}). "
            f"Diferencia: {blue_total_in - blue_total_out:.0f} bultos perdidos."
        )

    return df, blue_audit

def load_frescura(file):
    """
    FIX v3.4 — Construcción correcta de cancha por SKU:
      1) Lee la hoja API completa.
      2) SKU está en columna A ('Artículo'). Cada SKU tiene un Nº de almacén
         (columna 'almacén' o similar).
      3) Lee el bloque I:M de la misma hoja API (tabla maestra de almacenes):
         I=Nº Almacén | J=Detalle | K=Calibre | L=Apilabilidad | M=Cancha
      4) Mergea SKU → Nº Almacén → Cancha.
    """
    xls = pd.ExcelFile(file)

    # ── HOJA API: lectura completa ──────────────────────────────────────────
    api_full = pd.read_excel(xls, sheet_name='API', header=0)

    # Columna SKU (col A)
    col_sku = find_col(api_full, 'Artículo', 'Articulo', 'SKU', 'Cód', 'Cod')
    if col_sku is None:
        raise ValueError("Hoja API: no se encontró la columna de SKU (Artículo).")

    # Columna almacén por SKU (entre A y H, la que asocia SKU con su almacén)
    col_alm_sku = find_col(api_full, 'almacén', 'almacen', 'Almacén', 'Almacen', 'Alm')
    if col_alm_sku is None:
        raise ValueError("Hoja API: no se encontró la columna 'almacén' que asocia SKU con almacén.")

    # Tabla SKU → Nº Almacén
    sku_alm = api_full[[col_sku, col_alm_sku]].copy()
    sku_alm.columns = ['sku_raw', 'alm_num_raw']
    sku_alm['sku'] = sku_alm['sku_raw'].apply(to_int_sku)
    sku_alm['alm_num'] = pd.to_numeric(sku_alm['alm_num_raw'], errors='coerce')
    sku_alm = sku_alm.dropna(subset=['sku'])
    sku_alm['sku'] = sku_alm['sku'].astype(int)
    # Regla de prevalencia: la última fila con almacén válido prevalece
    sku_alm = sku_alm.dropna(subset=['alm_num'])
    sku_alm['alm_num'] = sku_alm['alm_num'].astype(int)
    sku_alm = sku_alm.drop_duplicates(subset=['sku'], keep='last')[['sku', 'alm_num']]

    # ── BLOQUE I:M (tabla maestra de almacenes), header en fila 1 ───────────
    # usecols='I:M', header=0 → lee solo las cols I-M con su header propio.
    alm_master = pd.read_excel(
        xls, sheet_name='API', header=0, usecols='I:M'
    )
    # Mapeo posicional según definición del usuario:
    # I = Nº Almacén | J = Detalle | K = Calibre | L = Apilabilidad | M = Cancha
    if alm_master.shape[1] < 5:
        raise ValueError(
            f"Bloque I:M de la hoja API tiene {alm_master.shape[1]} columnas, se esperaban 5."
        )
    alm_master = alm_master.iloc[:, :5].copy()
    alm_master.columns = ['alm_num', 'alm_det', 'calibre', 'apil', 'cancha_raw']
    alm_master['alm_num'] = pd.to_numeric(alm_master['alm_num'], errors='coerce')
    alm_master = alm_master.dropna(subset=['alm_num'])
    alm_master['alm_num'] = alm_master['alm_num'].astype(int)
    # Prevalece la última definición del almacén
    alm_master = alm_master.drop_duplicates(subset=['alm_num'], keep='last')

    # ── MERGE: SKU → Almacén → Cancha ───────────────────────────────────────
    api = sku_alm.merge(alm_master, on='alm_num', how='left')
    api['cancha'] = api['cancha_raw'].apply(normalize_cancha)
    api = api.rename(columns={'alm_num': 'alm_id'})
    api['alm_det'] = api['alm_det'].fillna('sin datos').astype(str)
    api = api[['sku', 'cancha', 'alm_id', 'alm_det']]

    # ── HOJA DDM (Bultos por pallet) ────────────────────────────────────────
    ddm_full = pd.read_excel(xls, sheet_name='DDM')
    col_ddm_sku = find_col(ddm_full, 'ARTÍCULO', 'ARTICULO', 'Artículo', 'Articulo', 'SKU')
    col_bxp     = find_col(ddm_full, 'BULTOS X PALLET', 'BULTOS_X_PALLET', 'BXP')
    if not col_ddm_sku or not col_bxp:
        raise ValueError("Hoja DDM: no se encontraron las columnas ARTÍCULO y/o BULTOS X PALLET.")
    ddm = ddm_full[[col_ddm_sku, col_bxp]].copy()
    ddm.columns = ['sku', 'bxp']
    ddm['sku'] = ddm['sku'].apply(to_int_sku)
    ddm = ddm.dropna(subset=['sku'])
    ddm['sku'] = ddm['sku'].astype(int)
    ddm['bxp'] = pd.to_numeric(ddm['bxp'], errors='coerce')
    ddm = ddm.drop_duplicates(subset=['sku'], keep='last')

    # ── HOJA Frescura ───────────────────────────────────────────────────────
    fr_full = pd.read_excel(xls, sheet_name='Frescura')
    col_fr_sku    = find_col(fr_full, 'Cód', 'Cod', 'SKU', 'Artículo', 'Articulo')
    col_fr_fecha  = find_col(fr_full, 'FECHA DE VENC.', 'FECHA DE VENC', 'Fecha de Venc', 'Vencimiento')
    col_fr_status = find_col(fr_full, 'Status', 'Estado')
    if not all([col_fr_sku, col_fr_fecha, col_fr_status]):
        raise ValueError("Hoja Frescura: faltan columnas Cód / FECHA DE VENC. / Status.")
    fr = fr_full[[col_fr_sku, col_fr_fecha, col_fr_status]].copy()
    fr.columns = ['sku', 'fecha', 'status']
    fr['sku'] = fr['sku'].apply(to_int_sku)
    fr = fr.dropna(subset=['sku'])
    fr['sku'] = fr['sku'].astype(int)
    fr = fr.sort_values('fecha').drop_duplicates(subset=['sku'], keep='first')

    # Diagnóstico para debug
    diag = {
        'api_full_cols': list(api_full.columns),
        'col_sku_used': col_sku,
        'col_alm_sku_used': col_alm_sku,
        'alm_master_cols_raw': list(pd.read_excel(xls, sheet_name='API', header=0, usecols='I:M').columns),
        'sku_count_api': len(api),
        'alm_master_count': len(alm_master),
        'sku_sin_cancha': int((api['cancha'] == 'SIN CANCHA').sum()),
        'alm_master_preview': alm_master.head(20).to_dict('records'),
    }
    return api, ddm, fr, diag

def get_lema(transport, idx=0):
    h = int(hashlib.md5(str(transport).encode()).hexdigest(), 16)
    return LEMAS[(h + idx) % len(LEMAS)]

def pallet_reading(val):
    if not val or val == 0.0:
        return ('0.00', '0 paletas', 'Sin carga de picking en esta cancha')
    vm = f'{val:.2f}'; n = int(val); fr = val - n
    if val < 1.0:  return (vm, '1 paleta',    'Menos de una paleta')
    if val == 1.0: return (vm, '1 paleta',    '1 paleta exacta')
    if val < 1.5:  return (vm, '2 paletas',   'Mas de una paleta')
    if val == 1.5: return (vm, '2 paletas',   '1 paleta y media')
    if val < 2.0:  return (vm, '2 paletas',   'Casi 2 paletas')
    if fr == 0.0:  return (vm, f'{n} paletas',   f'{n} paletas exactas')
    if fr < 0.5:   return (vm, f'{n+1} paletas', f'Mas de {n} paletas')
    if fr == 0.5:  return (vm, f'{n+1} paletas', f'{n} paletas y media')
    return          (vm, f'{n+1} paletas', f'Casi {n+1} paletas')

def process_reparto(rep_df, api, ddm, fr):
    rows = []
    for _, r in rep_df.iterrows():
        sku = int(r['Artículo'])
        blts_raw = float(r['Bultos']); unids = float(r['Unids'])
        desc = str(r['Descripción Artículo'])
        orphan_no_ddm = bool(r.get('_orphan_no_ddm', False))

        a = api[api['sku'] == sku]
        if len(a):
            cancha     = str(a.iloc[0]['cancha'])
            alm_id_i   = int(a.iloc[0]['alm_id']) if pd.notna(a.iloc[0]['alm_id']) else None
            alm_det    = str(a.iloc[0]['alm_det'])
        else:
            cancha, alm_id_i, alm_det = 'SIN CANCHA', None, 'sin datos'

        d = ddm[ddm['sku'] == sku]
        bxp = float(d.iloc[0]['bxp']) if len(d) and pd.notna(d.iloc[0]['bxp']) else None

        if bxp and bxp > 0:
            pal_ent = int(blts_raw / bxp); blts_pick = blts_raw - pal_ent * bxp
            has_pal = pal_ent >= 1; miss_bxp = False
        else:
            pal_ent = 0; blts_pick = blts_raw; has_pal = False; miss_bxp = True

        skip_color = (sku == 7038) or (alm_id_i == 46)
        f = fr[fr['sku'] == sku]
        if len(f) and not skip_color:
            fv = f.iloc[0]['fecha']; st_ = str(f.iloc[0]['status']).strip().upper()
            if pd.isna(fv): fecha_s, row_st = '', 'NONE'
            else:
                fecha_s = pd.Timestamp(fv).strftime('%d/%m/%y')
                row_st  = st_ if st_ in ('RED','YELLOW','GREEN') else 'NONE'
        else:
            fecha_s, row_st = '', 'NONE'

        rows.append({'sku':sku,'desc':desc,'blts_raw':blts_raw,'unids':unids,
                     'blts_pick':blts_pick,'pal_ent':pal_ent,'has_pal':has_pal,
                     'miss_bxp':miss_bxp,'bxp':bxp,'cancha':cancha,
                     'alm_id':alm_id_i,'alm_det':alm_det,'fecha_s':fecha_s,'row_st':row_st,
                     'orphan_no_ddm':orphan_no_ddm})
    return rows

def build_alm_groups(c_rows):
    groups = {}; order = []
    for r in c_rows:
        key = (r['alm_id'], r['alm_det'])
        if key not in groups: groups[key] = []; order.append(key)
        groups[key].append(r)
    for key in order:
        groups[key].sort(key=lambda x: -x['blts_raw'])
    # Ordenar almacenes de menor a mayor por número de almacén
    order.sort(key=lambda k: (k[0] is None, k[0] if k[0] is not None else 0))
    return [(k[0], k[1], groups[k]) for k in order]

def compute_pall_value(rows, cancha):
    return sum(r['blts_pick']/r['bxp'] for r in rows
               if r['cancha'] == cancha and r['bxp'] and r['bxp'] > 0)

# ─── DRAWING HELPERS ─────────────────────────────────────────────────────────

def ry(y_top): return PAGE_H - y_top

def rfill(c, x, y_top, w, h, fill, stroke=None, lw=0.4):
    c.setFillColor(fill)
    if stroke:
        c.setStrokeColor(stroke); c.setLineWidth(lw)
        c.rect(x, ry(y_top+h), w, h, fill=1, stroke=1)
    else:
        c.rect(x, ry(y_top+h), w, h, fill=1, stroke=0)

def txt(c, x, y_top, s, font='Helvetica', sz=8, col=colors.black, align='left', mw=None):
    c.setFont(font, sz); c.setFillColor(col)
    if mw:
        while s and c.stringWidth(s, font, sz) > mw: s = s[:-1]
    fn = {'left': c.drawString, 'center': c.drawCentredString, 'right': c.drawRightString}[align]
    fn(x, ry(y_top), s)

def draw_watermark(c):
    c.saveState()
    c.setFillColor(WM_GRAY); c.setFillAlpha(0.07)
    c.setFont('Helvetica-Bold', 68)
    c.translate(PAGE_W/2, PAGE_H/2); c.rotate(35)
    c.drawCentredString(0, 0, 'BECCACECE HNOS')
    c.restoreState()

def draw_title_bar(c, y, numero, fecha_str):
    rfill(c, MARGIN, y, CW, H_TITLE, DARK_BLUE)
    label = f'PLANILLA DE CARGA  |  Reparto Nro: {numero}  |  Fecha: {fecha_str}'
    txt(c, MARGIN+CW/2, y+17, label, 'Helvetica-Bold', 10.5, colors.white, 'center', CW-16)
    return H_TITLE

def fmt_transport(t):
    """Elimina el .0 final del número de transporte (128.0 → 128)."""
    try:
        v = float(t)
        if v == int(v):
            return str(int(v))
    except (ValueError, TypeError):
        pass
    return str(t)

def draw_transport_line(c, y, transport, chofer):
    rfill(c, MARGIN, y, CW, H_TRANS, colors.HexColor('#EEF2F8'))
    txt(c, MARGIN+6,      y+12, f'Transporte: {fmt_transport(transport)} - {chofer}', 'Helvetica-Bold', 9)
    txt(c, MARGIN+CW-6,   y+12, 'Depósito: 001 - CASA CENTRAL',        'Helvetica-Bold', 9, align='right')
    return H_TRANS

def draw_partida_regreso(c, y):
    txt(c, MARGIN+6,     y+11, 'F. y H. Est. de Partida: ________ Hs.', 'Helvetica', 9)
    txt(c, MARGIN+CW-6,  y+11, 'F. y H. Est. de Regreso: ________ Hs.','Helvetica', 9, align='right')
    return H_PARTIDA

def draw_control_carga(c, y):
    hdr_h = 15; box_h = H_CTRL - hdr_h; half = CW/2
    rfill(c, MARGIN,        y, half, hdr_h, HDR_BG, BORDER)
    rfill(c, MARGIN+half,   y, half, hdr_h, HDR_BG, BORDER)
    txt(c, MARGIN+half/2,         y+10, 'CONTROL DE CARGA',    'Helvetica-Bold', 9, align='center')
    txt(c, MARGIN+half+half/2,    y+10, 'CONTROL DE DESCARGA', 'Helvetica-Bold', 9, align='center')
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)
    c.rect(MARGIN, ry(y+hdr_h+box_h), CW, box_h, fill=0, stroke=1)
    c.line(MARGIN+half, ry(y+hdr_h), MARGIN+half, ry(y+hdr_h+box_h))
    return H_CTRL

def draw_lema(c, y, lema):
    rfill(c, MARGIN, y, CW, H_LEMA, DARK_BLUE)
    txt(c, MARGIN+CW/2, y+12, f'"{lema}"', 'Helvetica-Oblique', 9, colors.white, 'center', CW-16)
    return H_LEMA

def draw_cancha_header(c, y, label):
    rfill(c, MARGIN, y, CW, H_CANCHA, MED_BLUE)
    txt(c, MARGIN+10, y+13, f'◀  {label}', 'Helvetica-Bold', 11, colors.white)
    return H_CANCHA

def draw_sin_cancha_header(c, y):
    rfill(c, MARGIN, y, CW, H_CANCHA, RED_ALERT)
    txt(c, MARGIN+CW/2, y+13, '⚠  SIN CANCHA ASIGNADA — REVISAR', 'Helvetica-Bold', 11, colors.white, 'center')
    return H_CANCHA

def draw_table_header(c, y):
    labels = ['SKU', 'Descripción', 'Venc.', 'Bultos', 'Unids']
    rfill(c, MARGIN, y, CW, H_THDR, HDR_BG, BORDER)
    x = MARGIN
    for lbl, w in zip(labels, COL_W):
        c.setStrokeColor(BORDER); c.setLineWidth(0.4)
        c.rect(x, ry(y+H_THDR), w, H_THDR, fill=0, stroke=1)
        txt(c, x+w/2, y+9.5, lbl, 'Helvetica-Bold', 8, align='center')
        x += w
    return H_THDR

def draw_alm_header(c, y, alm_id, alm_det):
    label = f'Almacén {alm_id} — {alm_det}' if alm_id is not None else 'Almacén: SIN DATOS'
    rfill(c, MARGIN, y, CW, H_ALM, ALM_BG, BORDER)
    txt(c, MARGIN+6, y+11, label, 'Helvetica-Bold', 9)
    return H_ALM

def draw_product_row(c, y, r):
    st_ = r['row_st']
    is_orphan_no_ddm = r.get('orphan_no_ddm', False)
    bg  = DARK_ROW if st_=='RED' else MED_ROW if st_=='YELLOW' else (
          AMBER if r.get('is_sin_cancha') else
          LIGHT_GRAY if r['has_pal'] else colors.white)

    rfill(c, MARGIN, y, CW, H_ROW, bg, BORDER)
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)

    # Marca de huérfano sin DDM: (*) rojo después del SKU
    sku_s = str(r['sku'])
    if is_orphan_no_ddm:
        sku_s = f"{r['sku']}(*)"
    elif r['miss_bxp']:
        sku_s = f"{r['sku']}(!)"

    bp     = r['blts_pick']; bp_s = str(int(bp)) if bp == int(bp) else f'{bp:.0f}'
    un_s   = str(int(r['unids']))
    ind    = {'RED':'■','YELLOW':'▲','GREEN':'○'}.get(st_,'')
    venc_s = f'{ind} {r["fecha_s"]}' if r['fecha_s'] else ''
    font   = 'Helvetica-Bold' if st_=='RED' else 'Helvetica'

    # Descripción con aviso embebido si es huérfano sin DDM
    desc_show = r['desc']
    if is_orphan_no_ddm:
        desc_show = f"{r['desc']}  ⚠ SIN BXP EN DDM — REVISAR"

    cells = [(sku_s,'center'),(desc_show,'left'),(venc_s,'center'),(bp_s,'center'),(un_s,'center')]
    x = MARGIN
    for i, ((s, align), w) in enumerate(zip(cells, COL_W)):
        c.rect(x, ry(y+H_ROW), w, H_ROW, fill=0, stroke=1)
        # Color rojo en SKU y descripción si es huérfano sin DDM
        col_text = RED_ALERT if (is_orphan_no_ddm and i in (0, 1)) else colors.black
        font_use = 'Helvetica-Bold' if (is_orphan_no_ddm and i in (0, 1)) else font
        if align == 'center':
            txt(c, x+w/2, y+9.5, s, font_use, 9, col=col_text, align='center', mw=w-3)
        else:
            txt(c, x+3,   y+9.5, s, font_use, 9, col=col_text, mw=w-5)
        # Subrayado SOLO en columna Bultos (índice 3)
        if r['has_pal'] and i == 3:
            tw = c.stringWidth(s, font, 9)
            cx = x+w/2
            c.setStrokeColor(colors.black); c.setLineWidth(0.6)
            c.line(cx-tw/2, ry(y+H_ROW)+1, cx+tw/2, ry(y+H_ROW)+1)
            c.setLineWidth(0.4)
        x += w
    return H_ROW

def draw_total_row(c, y, tot_b, tot_u):
    mw = COL_W[0]+COL_W[1]+COL_W[2]
    rfill(c, MARGIN, y, CW, H_TOT, colors.HexColor('#E8E8E8'), BORDER)
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)
    c.rect(MARGIN,    ry(y+H_TOT), mw,      H_TOT, fill=0, stroke=1)
    c.rect(MARGIN+mw, ry(y+H_TOT), COL_W[3],H_TOT, fill=0, stroke=1)
    c.rect(MARGIN+mw+COL_W[3], ry(y+H_TOT), COL_W[4], H_TOT, fill=0, stroke=1)
    txt(c, MARGIN+mw/2, y+9.5, 'T O T A L  A L M A C É N', 'Helvetica-Bold', 8.5, align='center')
    for val, xoff in [(tot_b, mw), (tot_u, mw+COL_W[3])]:
        s = str(int(val)) if val == int(val) else f'{val:.0f}'
        txt(c, MARGIN+xoff+COL_W[3 if xoff==mw else 4]/2, y+9.5, s, 'Helvetica-Bold', 8.5, align='center')
    return H_TOT

def draw_espacio_asignado(c, y, cancha, pv):
    rfill(c, MARGIN, y, CW, H_ESPACIO, BG_GRAY, BORDER)
    txt(c, MARGIN+CW/2, y+10, f'ESPACIO ASIGNADO — {cancha}', 'Helvetica-Bold', 9, align='center')
    col_w = CW/3
    hdrs  = ['Valor mínimo','Valor entero','Lectura']
    bx = MARGIN; by = y+14; rh = 13
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)
    for h in hdrs:
        rfill(c, bx, by, col_w, rh, HDR_BG, BORDER)
        txt(c, bx+col_w/2, by+9, h, 'Helvetica-Bold', 8, align='center')
        bx += col_w
    vm, ve, lec = pallet_reading(pv)
    bx = MARGIN; by2 = by+rh
    for v in [vm, ve, lec]:
        rfill(c, bx, by2, col_w, rh, colors.white, BORDER)
        txt(c, bx+col_w/2, by2+9, v, 'Helvetica', 8.5, align='center', mw=col_w-4)
        bx += col_w
    return H_ESPACIO

def draw_route_to(c, y, dest):
    rfill(c, MARGIN, y, CW, H_ROUTE, YELLOW_RTE, BORDER)
    txt(c, MARGIN+CW/2, y+11, f'▶  DIRIGIR PALETA A: {dest}', 'Helvetica-Bold', 9.5, align='center')
    return H_ROUTE

def draw_route_recv(c, y):
    rfill(c, MARGIN, y, CW, H_ROUTE, LIGHT_GRAY, BORDER)
    txt(c, MARGIN+CW/2, y+11, '◀  RECIBE PALETA DESDE: CANCHA I', 'Helvetica-Bold', 9.5, align='center')
    return H_ROUTE

def draw_footer(c, date_str, page_num):
    yb = PAGE_H - MARGIN + 2
    c.setStrokeColor(colors.HexColor('#CCCCCC')); c.setLineWidth(0.5)
    c.line(MARGIN, ry(yb-4), MARGIN+CW, ry(yb-4))
    txt(c, MARGIN+CW/2, yb+6, 'Almacen Digital 3.0  |  Beccacece Hnos SA  |  Distribuidor Oficial CMQ — Desde 1963',
        'Helvetica', 8, FOOT_GRAY, 'center')
    txt(c, MARGIN,    yb+16, date_str,           'Helvetica', 8, FOOT_GRAY)
    txt(c, MARGIN+CW, yb+16, f'Página: {page_num}','Helvetica', 8, FOOT_GRAY, 'right')

def header_height(is_first_rep_page: bool) -> float:
    h = H_TITLE + GAP + H_TRANS + GAP
    if is_first_rep_page:
        h += H_PARTIDA + GAP + H_CTRL + GAP
    h += H_LEMA + GAP + H_CANCHA + GAP + H_THDR
    return h

def footer_reserve(has_route: bool) -> float:
    return H_ESPACIO + (H_ROUTE if has_route else 0) + GAP + H_FOOT

def content_avail(is_first_rep_page: bool, has_route: bool) -> float:
    return PAGE_H - 2*MARGIN - header_height(is_first_rep_page) - footer_reserve(has_route)

ALM_GAP = 8   # espacio entre grupos de almacén (más legible)

def alm_group_height(n_rows: int) -> float:
    return ALM_GAP + H_ALM + n_rows*H_ROW + H_TOT

def draw_cancha_pages(c, reparto_rows, cancha, is_first_rep, numero, transport,
                       chofer, lema, fecha_str, date_str, pc,
                       pall_values, route_dest, receives_route):
    c_rows     = [r for r in reparto_rows if r['cancha'] == cancha]
    alm_groups = build_alm_groups(c_rows)

    remaining = list(alm_groups)
    pages     = []
    first_pg  = True

    while True:
        is_frp    = is_first_rep and first_pg
        avail     = content_avail(is_frp, has_route=False)
        used      = 0
        pg_groups = []

        for g in remaining[:]:
            gh = alm_group_height(len(g[2]))
            if used + gh > avail and pg_groups:
                break
            pg_groups.append(remaining.pop(0))
            used += gh

        pages.append((is_frp, pg_groups))
        first_pg = False
        if not remaining:
            break

    n_pages = len(pages)
    for pi, (is_frp, pg_groups) in enumerate(pages):
        is_last = (pi == n_pages - 1)
        has_rt  = is_last and (route_dest or receives_route)

        if pc['n'] > 0:
            c.showPage()
        pc['n'] += 1

        draw_watermark(c)
        y = MARGIN

        y += draw_title_bar(c, y, numero, fecha_str);  y += GAP
        y += draw_transport_line(c, y, transport, chofer); y += GAP

        if is_frp:
            y += draw_partida_regreso(c, y); y += GAP
            y += draw_control_carga(c, y);   y += GAP

        y += draw_lema(c, y, lema); y += GAP

        lbl = cancha if pi == 0 else f'{cancha}  (continuación {pi+1})'
        y += draw_cancha_header(c, y, lbl); y += GAP
        y += draw_table_header(c, y)

        if not pg_groups and not c_rows:
            txt(c, MARGIN+CW/2, y+13, f'— Sin carga de {cancha} para este camión —',
                'Helvetica-Oblique', 8, colors.HexColor('#888888'), 'center')
        else:
            for alm_id, alm_det, rows in pg_groups:
                y += ALM_GAP
                y += draw_alm_header(c, y, alm_id, alm_det)
                for r in rows:
                    y += draw_product_row(c, y, r)
                y += draw_total_row(c, y,
                                    sum(r['blts_pick'] for r in rows),
                                    sum(r['unids']     for r in rows))

        if is_last:
            pv = pall_values.get(cancha, 0.0)
            route_h   = H_ROUTE if has_rt else 0
            foot_y    = PAGE_H - MARGIN - H_FOOT - H_ESPACIO - route_h - GAP
            draw_espacio_asignado(c, foot_y, cancha, pv)
            if route_dest:
                draw_route_to(c, foot_y + H_ESPACIO + GAP, route_dest)
            elif receives_route:
                draw_route_recv(c, foot_y + H_ESPACIO + GAP)

        draw_footer(c, date_str, pc['n'])


def draw_sin_cancha_page(c, row, numero, transport, chofer, lema, fecha_str, date_str, pc):
    if pc['n'] > 0:
        c.showPage()
    pc['n'] += 1
    draw_watermark(c)
    y = MARGIN

    y += draw_title_bar(c, y, numero, fecha_str);  y += GAP
    y += draw_transport_line(c, y, transport, chofer); y += GAP
    y += draw_lema(c, y, lema); y += GAP
    y += draw_sin_cancha_header(c, y); y += GAP
    y += draw_alm_header(c, y, None, 'sin asignación en base de datos'); y += GAP
    y += draw_table_header(c, y)

    rc = dict(row); rc['is_sin_cancha'] = True
    y += draw_product_row(c, y, rc); y += 8

    rfill(c, MARGIN, y, CW, 22, colors.HexColor('#FFF3E0'), colors.HexColor('#FFB300'), lw=0.6)
    note = f"SKU {row['sku']} no tiene cancha asignada. Validar con Jefe de Almacén antes de cargar."
    txt(c, MARGIN+CW/2, y+14, note, 'Helvetica', 7.5, colors.HexColor('#C0392B'), 'center', CW-10)

    draw_footer(c, date_str, pc['n'])

# ─── GENERADOR PRINCIPAL ─────────────────────────────────────────────────────

def generate_pdf(car_df, api, ddm, fr):
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)
    pc  = {'n': 0}

    if 'Fecha Mvto' in car_df.columns and len(car_df):
        try:
            dt = pd.Timestamp(car_df['Fecha Mvto'].dropna().iloc[0])
            fecha_str = date_str = dt.strftime('%d/%m/%Y')
        except:
            fecha_str = date_str = datetime.today().strftime('%d/%m/%Y')
    else:
        fecha_str = date_str = datetime.today().strftime('%d/%m/%Y')

    stats = {'repartos':[], 'red':[], 'yellow':[], 'pallet_applied':[],
             'miss_bxp':[], 'no_fecha':[], 'sin_cancha_skus':[],
             'orphans_no_ddm':[],
             'total_pages':0, 'trace':[]}

    for rep_idx, (numero, rep_df) in enumerate(car_df.groupby('Número', sort=False)):
        transport = str(rep_df['Transporte'].iloc[0])
        chofer    = str(rep_df['Descripción Transporte'].iloc[0])
        lema      = get_lema(transport, rep_idx)
        stats['repartos'].append({'numero': numero, 'transport': transport, 'chofer': chofer})

        rows        = process_reparto(rep_df, api, ddm, fr)
        pall_values = {ch: compute_pall_value(rows, ch) for ch in CANCHA_ORDER}

        # Trazabilidad de cancha por SKU para el debug
        for r in rows:
            stats['trace'].append({
                'Reparto': numero, 'Camión': transport,
                'SKU': r['sku'], 'Descripción': r['desc'][:50],
                'Almacén': r['alm_id'], 'Detalle Almacén': r['alm_det'],
                'Cancha': r['cancha'], 'Bultos CAR': r['blts_raw'],
                'A pickear': r['blts_pick'],
                'Huérfano sin DDM': '⚠' if r.get('orphan_no_ddm') else '',
            })

        pv1 = pall_values['CANCHA I']; frac1 = pv1 - int(pv1)
        if frac1 > 0.001:
            pv2, pv4 = pall_values['CANCHA II'], pall_values['CANCHA IV']
            if pv2 == 0 and pv4 == 0:  route_dest = route_recv = None
            elif pv4 > pv2:             route_dest = route_recv = 'CANCHA IV'
            else:                       route_dest = route_recv = 'CANCHA II'
        else:
            route_dest = route_recv = None

        for ci, cancha in enumerate(CANCHA_ORDER):
            is_first_rep = (ci == 0)
            receives = (cancha == route_recv and not is_first_rep)
            draw_cancha_pages(c, rows, cancha, is_first_rep, numero, transport, chofer,
                              lema, fecha_str, date_str, pc, pall_values,
                              route_dest if is_first_rep else None, receives)

        for r in [r for r in rows if r['cancha'] == 'SIN CANCHA']:
            draw_sin_cancha_page(c, r, numero, transport, chofer, lema, fecha_str, date_str, pc)
            stats['sin_cancha_skus'].append(f"{r['sku']} — {r['desc']}")

        for r in rows:
            if r['row_st'] == 'RED':   stats['red'].append(r['sku'])
            if r['row_st'] == 'YELLOW':stats['yellow'].append(r['sku'])
            if r['has_pal']:           stats['pallet_applied'].append(r['sku'])
            if r['miss_bxp']:          stats['miss_bxp'].append(r['sku'])
            if not r['fecha_s']:       stats['no_fecha'].append(r['sku'])
            if r.get('orphan_no_ddm'): stats['orphans_no_ddm'].append(
                f"{r['sku']} — {r['desc']} (T{transport}, R{numero})")

    c.save(); buf.seek(0)
    stats['total_pages'] = pc['n']
    return buf.read(), stats

# ─── RESUMEN DE CARGA POR CAMIÓN (v3.9, hoja AGR) ───────────────────────────

def load_agr(file):
    """Lee la hoja AGR del CAR y devuelve DF consolidado por camión + SKU.

    Columnas resultantes: chofer | cod_producto | descripcion | bultos
    Una fila por (camión, SKU) con bultos sumados.
    Ignora Cliente / Razón social / Importes (no son relevantes para este doc).
    """
    file.seek(0)
    raw = pd.read_excel(file, sheet_name='AGR', header=None)

    # Buscar la fila de header (contiene 'Cliente' o 'Chofer')
    header_row = None
    for i in range(min(10, len(raw))):
        row_str = ' '.join(str(v) for v in raw.iloc[i].values if pd.notna(v))
        if 'Chofer' in row_str and 'Cantidad' in row_str:
            header_row = i
            break
    if header_row is None:
        raise ValueError("No se encontró el header de la hoja AGR (esperaba 'Chofer' y 'Cantidad').")

    df = pd.read_excel(file, sheet_name='AGR', header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_cod   = find_col(df, 'Cod Producto', 'CodProducto', 'Codigo Producto', 'Codigo', 'Código')
    col_desc  = find_col(df, 'DESCRIPCION', 'Descripcion', 'Descripción')
    col_cant  = find_col(df, 'Cantidad', 'Bultos')
    col_chof  = find_col(df, 'Chofer', 'Camion', 'Camión')

    if not all([col_cod, col_desc, col_cant, col_chof]):
        raise ValueError(
            f"Faltan columnas en hoja AGR. Encontradas: {list(df.columns)}. "
            f"Necesito: Cod Producto, Descripcion, Cantidad, Chofer."
        )

    df = df[[col_cod, col_desc, col_cant, col_chof]].copy()
    df.columns = ['cod_producto', 'descripcion', 'bultos', 'chofer']
    df = df.dropna(subset=['chofer', 'cod_producto', 'bultos'])

    df['chofer']       = df['chofer'].apply(to_int_sku)
    df['cod_producto'] = df['cod_producto'].apply(to_int_sku)
    df['bultos']       = pd.to_numeric(df['bultos'], errors='coerce').fillna(0).astype(int)
    df['descripcion']  = df['descripcion'].astype(str).str.strip()
    df = df[df['bultos'] > 0]

    # Consolidación: total por SKU por camión
    out = (df.groupby(['chofer', 'cod_producto', 'descripcion'], as_index=False)['bultos']
             .sum()
             .sort_values(['chofer', 'cod_producto']))
    return out


def _draw_resumen_header(c, fecha_str):
    """Cabecera del documento (título + subtítulo + línea separadora)."""
    margin = 15 * mm
    y_top = PAGE_H - margin

    c.setFillColor(colors.HexColor('#1a1a1a'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y_top - 14, 'Resumen de Carga por Camión')

    c.setFillColor(colors.HexColor('#666666'))
    c.setFont('Helvetica', 9)
    subt = f"Beccacece Hnos. S.A.   ·   La Reja, Moreno   ·   Fecha: {fecha_str}"
    c.drawString(margin, y_top - 28, subt)

    # Línea separadora
    c.setStrokeColor(colors.HexColor('#1a1a1a'))
    c.setLineWidth(0.8)
    c.line(margin, y_top - 36, PAGE_W - margin, y_top - 36)

    return y_top - 44  # y donde puede empezar el contenido (en coord top-down)


def _truck_block_height(n_rows):
    """Altura total de un bloque de camión (header + filas + spacer)."""
    return 7*mm + 6*mm + n_rows * 7*mm + 6*mm  # header + thead + filas + spacer


def _draw_truck_block(c, y_top, cam, rows):
    """Dibuja un bloque (header de camión + tabla) usando coords top-down.
    Devuelve el nuevo y_top (más abajo) después del bloque.
    """
    margin = 15 * mm
    total_w = PAGE_W - 2 * margin  # 180mm

    # Anchos de columna: Código 22 | Descripción 122 | Bultos 26 | Check 10 (mm)
    w_cod, w_desc, w_blt, w_chk = 22*mm, 122*mm, 26*mm, 10*mm

    # ── 1) Header del camión (barra negra) ──────────────────────────────────
    h_hdr = 7 * mm
    y_hdr = y_top - h_hdr
    c.setFillColor(colors.HexColor('#1a1a1a'))
    c.rect(margin, y_hdr, total_w, h_hdr, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin + 6, y_hdr + 2.3*mm, f'Camión {cam}')

    # ── 2) Header de columnas (banda gris clara) ────────────────────────────
    h_thd = 6 * mm
    y_thd = y_hdr - h_thd
    c.setFillColor(colors.HexColor('#f0f0f0'))
    c.rect(margin, y_thd, total_w, h_thd, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.setLineWidth(0.5)
    c.line(margin, y_thd, margin + total_w, y_thd)  # línea inferior

    c.setFillColor(colors.HexColor('#333333'))
    c.setFont('Helvetica-Bold', 8.5)
    # Centrado en la celda
    c.drawCentredString(margin + w_cod/2,                                y_thd + 1.7*mm, 'Código')
    c.drawString       (margin + w_cod + 6,                              y_thd + 1.7*mm, 'Descripción')
    c.drawCentredString(margin + w_cod + w_desc + w_blt/2,               y_thd + 1.7*mm, 'Bultos')
    c.drawCentredString(margin + w_cod + w_desc + w_blt + w_chk/2,       y_thd + 1.7*mm, 'Check')

    # ── 3) Filas ─────────────────────────────────────────────────────────────
    h_row = 7 * mm
    y_cur = y_thd
    c.setStrokeColor(colors.HexColor('#cccccc'))

    for i, r in enumerate(rows):
        y_cur -= h_row

        # Texto
        c.setFillColor(colors.HexColor('#1a1a1a'))
        c.setFont('Helvetica', 9)
        c.drawCentredString(margin + w_cod/2,                          y_cur + 2.2*mm, str(r['cod_producto']))
        c.drawString       (margin + w_cod + 6,                        y_cur + 2.2*mm, str(r['descripcion']))
        c.setFont('Helvetica-Bold', 9)
        c.drawCentredString(margin + w_cod + w_desc + w_blt/2,         y_cur + 2.2*mm, str(int(r['bultos'])))

        # Línea separadora entre filas (excepto la última, que la cierra el BOX)
        if i < len(rows) - 1:
            c.setStrokeColor(colors.HexColor('#e8e8e8'))
            c.setLineWidth(0.25)
            c.line(margin, y_cur, margin + total_w, y_cur)

        # Recuadro de check (casilla destacada con borde más marcado)
        chk_x = margin + w_cod + w_desc + w_blt
        # padding interno para que el cuadrado no toque los bordes de la fila
        pad_x, pad_y = 2.5*mm, 1.3*mm
        c.setStrokeColor(colors.HexColor('#666666'))
        c.setLineWidth(0.8)
        c.rect(chk_x + pad_x, y_cur + pad_y,
               w_chk - 2*pad_x, h_row - 2*pad_y,
               fill=0, stroke=1)

    # ── 4) Box exterior de toda la tabla ─────────────────────────────────────
    tabla_h = h_thd + len(rows) * h_row
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.setLineWidth(0.5)
    c.rect(margin, y_thd - len(rows) * h_row, total_w, tabla_h, fill=0, stroke=1)

    # Línea vertical antes de columna check
    chk_x = margin + w_cod + w_desc + w_blt
    c.line(chk_x, y_thd - len(rows)*h_row, chk_x, y_thd)

    # Devolver y_top siguiente (con un spacer de 6mm)
    return y_cur - 6 * mm


def _draw_resumen_footer(c, page_num, total_pages):
    """Pie de página: 'Página X de Y' centrado."""
    c.setFillColor(colors.HexColor('#666666'))
    c.setFont('Helvetica', 8)
    c.drawCentredString(PAGE_W / 2, 10 * mm, f'Página {page_num} de {total_pages}')


def build_resumen_carga_pdf(car_file):
    """Genera el PDF 'Resumen de Carga por Camión' a partir de la hoja AGR.

    Lógica:
      1) Lee hoja AGR y consolida bultos por (camión, SKU).
      2) Ordena camiones ascendentemente.
      3) Imprime en bloques compactos: varios camiones por página si entran.
      4) Ningún bloque se parte entre páginas.
      5) Pie de página 'Página X de Y'.
    """
    resumen = load_agr(car_file)
    if resumen.empty:
        raise ValueError("La hoja AGR no contiene datos válidos para generar el resumen.")

    margin = 15 * mm
    fecha_str = datetime.today().strftime('%d/%m/%Y')

    # ── PRIMERA PASADA: paginar (sin dibujar) para conocer total_pages ──────
    camiones = sorted(resumen['chofer'].unique())
    bloques = []  # lista de (cam, rows_list, alto)
    for cam in camiones:
        sub = resumen[resumen['chofer'] == cam].to_dict('records')
        h = _truck_block_height(len(sub))
        bloques.append((cam, sub, h))

    # Reserva para cabecera (solo en página 1) y pie (en todas)
    H_HEADER_PAGE1 = 44  # 'Resumen de Carga por Camión' + subt + línea + spacer
    H_FOOTER      = 18 * mm  # margen inferior efectivo donde reservamos espacio

    # Espacio disponible para contenido (top → bottom)
    def avail(is_first_page):
        top    = PAGE_H - margin
        bottom = H_FOOTER  # piso del contenido
        if is_first_page:
            top -= H_HEADER_PAGE1
        return top - bottom

    # Asignar bloques a páginas
    pages = []   # lista de lista de (cam, rows)
    current = []
    space_left = avail(True)
    is_first   = True
    for cam, rows, h in bloques:
        if h <= space_left:
            current.append((cam, rows))
            space_left -= h
        else:
            # Cerrar página actual y abrir nueva
            pages.append(current)
            is_first = False
            current = [(cam, rows)]
            space_left = avail(False) - h
    if current:
        pages.append(current)

    total_pages = len(pages)

    # ── SEGUNDA PASADA: dibujar ─────────────────────────────────────────────
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    c.setTitle('Resumen de Carga por Camión')

    for page_idx, page_content in enumerate(pages):
        is_first_page = (page_idx == 0)
        if is_first_page:
            y_top = _draw_resumen_header(c, fecha_str)
        else:
            y_top = PAGE_H - margin

        for cam, rows in page_content:
            y_top = _draw_truck_block(c, y_top, cam, rows)

        _draw_resumen_footer(c, page_idx + 1, total_pages)
        c.showPage()

    c.save()
    pdf_bytes = buf.getvalue()
    return pdf_bytes, {
        'total_pages': total_pages,
        'camiones': len(camiones),
        'filas_sku': len(resumen),
    }


# ============================================================================
# ║                                                                          ║
# ║        EXTENSIONES v4.0 — lectores MASTER, T2, extraíbles, UI            ║
# ║                                                                          ║
# ============================================================================

# ─── LECTORES MASTER (78 MB) ────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_master_matriz_pall(file_bytes: bytes) -> tuple[pd.DataFrame, datetime | None]:
    """Lee la hoja 'Matriz Pall.' del MASTER.

    Estructura confirmada por inspección:
      - Celda A1: fecha del día (datetime)
      - Fila 1 cols B-Y: headers
      - Datos desde fila 2

    Returns:
        df: DataFrame con columnas tipadas
        fecha: datetime de A1 (o None si no se pudo leer)
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Matriz Pall." not in wb.sheetnames:
        raise ValueError("Hoja 'Matriz Pall.' no encontrada en MASTER")
    ws = wb["Matriz Pall."]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("Hoja 'Matriz Pall.' vacía")

    # A1 = fecha del día; los headers de las columnas B-Y están en fila 1
    fecha_a1 = rows[0][0] if isinstance(rows[0][0], datetime) else None

    # Headers (col B en adelante)
    headers_raw = list(rows[0][1:25])  # B-Y
    headers = [
        "Camion", "Reparto",
        "UP_PICKING_AGR_FACT",
        "PICK_C1", "PICK_C2", "PICK_C3", "PICK_C4", "PICK_C5", "PICK_MKPL",
        "UP_AE_PALL_AGR_FACT",
        "AE_C1", "AE_C2", "AE_C3", "AE_C4", "AE_C5", "AE_MKPL",
        "KG_MATRIZ",
        "KG_C1", "KG_C2", "KG_C3", "KG_C4", "KG_C5", "KG_MKPL",
        "CLASIFICACION",
    ]

    # Datos: fila 2 en adelante, cols B-Y
    data = []
    for r in rows[1:]:
        if r[1] is None:  # Camion vacío → fin de datos
            break
        data.append(r[1:25])

    df = pd.DataFrame(data, columns=headers)
    # Tipos
    df["Camion"] = pd.to_numeric(df["Camion"], errors="coerce").astype("Int64")
    df["Reparto"] = df["Reparto"].astype(str).str.strip().str.upper()
    for col in df.columns:
        if col not in ("Camion", "Reparto"):
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    wb.close()
    return df, fecha_a1


@st.cache_data(show_spinner=False)
def load_master_matriz_picking_app(file_bytes: bytes) -> pd.DataFrame:
    """Lee la hoja 'Matriz Picking APP' del MASTER.

    Estructura: fila 1 = headers con A1 = fecha del día, 64 cols total.
    Para Sprint 1 (extraíble fx Picking) basta con devolver el bloque
    como DataFrame plano.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Matriz Picking APP" not in wb.sheetnames:
        raise ValueError("Hoja 'Matriz Picking APP' no encontrada en MASTER")
    ws = wb["Matriz Picking APP"]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return pd.DataFrame()

    headers_raw = list(rows[0])
    # Header A es la fecha → reemplazamos por 'FECHA'
    headers = ["FECHA"] + [str(h) if h is not None else f"COL_{i}"
                           for i, h in enumerate(headers_raw[1:], start=2)]
    # Recortar al ancho real
    n = len(headers)

    data = []
    for r in rows[1:]:
        if all(v is None for v in r[:n]):
            continue
        data.append(list(r[:n]))

    df = pd.DataFrame(data, columns=headers[:len(data[0]) if data else n])
    wb.close()
    return df


@st.cache_data(show_spinner=False)
def load_master_reposicion_ae(file_bytes: bytes) -> pd.DataFrame:
    """Lee la hoja 'Reposición AE' del MASTER.

    Columnas A-J:
        A: Fecha | B: Almacén | C: Descripción | D: Cancha | E: bxp
        F: Posiciones | G: En cancha | H: AGR | I: PICK | J: Reposición Pall
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Reposición AE" not in wb.sheetnames:
        raise ValueError("Hoja 'Reposición AE' no encontrada en MASTER")
    ws = wb["Reposición AE"]

    headers = ["Fecha", "Almacen", "Descripcion", "Cancha", "bxp",
               "Posiciones", "En_cancha", "AGR", "PICK", "Reposicion_Pall"]
    data = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if r[1] is None:  # Almacén vacío → fin
            break
        data.append(list(r[:10]))

    df = pd.DataFrame(data, columns=headers)
    df["Almacen"] = pd.to_numeric(df["Almacen"], errors="coerce").astype("Int64")
    for col in ("bxp", "Posiciones", "En_cancha", "AGR", "PICK", "Reposicion_Pall"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    wb.close()
    return df


@st.cache_data(show_spinner=False)
def load_master_agregados_ae(file_bytes: bytes) -> pd.DataFrame:
    """Placeholder para Tab 5 — Agregados AE.

    Por ahora devuelve un DataFrame derivado de Matriz Pall. con las cols
    de AE (K-Q). Cuando Lucas confirme la hoja exacta destino se ajusta.
    """
    df_pall, _ = load_master_matriz_pall(file_bytes)
    # Subset: Camion + Reparto + UP AE + 6 canchas AE
    cols = ["Camion", "Reparto", "UP_AE_PALL_AGR_FACT",
            "AE_C1", "AE_C2", "AE_C3", "AE_C4", "AE_C5", "AE_MKPL"]
    return df_pall[cols].copy()


# ─── LECTOR T2 Pall. Camiones (referencia estática) ─────────────────────────

@st.cache_data(show_spinner=False)
def load_pall_camiones(file_bytes: bytes) -> pd.DataFrame:
    """Lee la hoja 'Pall. Camiones' del T2_Status_Carga (datos maestros).

    Headers reales A-AF (32 cols). El archivo se carga una sola vez por
    sesión y queda en cache (referencia estática para Tab 3 — Sprint 3).
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Pall. Camiones" not in wb.sheetnames:
        raise ValueError("Hoja 'Pall. Camiones' no encontrada en T2_Status")
    ws = wb["Pall. Camiones"]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h).strip() if h else f"COL_{i}" for i, h in enumerate(rows[0])]
    # Renombre clave: TARA → CAPACIDAD MÁX (decisión cerrada handoff #9)
    # En esta hoja el header ya es 'Capacidad Maxima (kg)' en col AC, así que
    # el rename solo aplica al PDF rediseñado del Tab 3 — la fuente se respeta.

    data = []
    for r in rows[1:]:
        if r[1] is None:
            continue
        data.append(list(r[:len(headers)]))
    df = pd.DataFrame(data, columns=headers)
    df["Camion"] = pd.to_numeric(df["Camion"], errors="coerce").astype("Int64")
    wb.close()
    return df


# ─── EXTRAÍBLES SHEETS (TAB 5) ──────────────────────────────────────────────

def df_to_tsv(df: pd.DataFrame) -> bytes:
    """TSV: pegar directo en Google Sheets sin separadores ambiguos."""
    return df.to_csv(sep="\t", index=False, lineterminator="\n").encode("utf-8")

def df_to_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(sep=",", index=False, lineterminator="\n").encode("utf-8")

def df_to_xlsx(df: pd.DataFrame, sheet_name: str = "Hoja1") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df.to_excel(wr, index=False, sheet_name=sheet_name[:31])
    buf.seek(0)
    return buf.read()

def filter_reposicion_negativa(df: pd.DataFrame) -> pd.DataFrame:
    """Tab 5 — filtro decisión cerrada: solo filas con Reposicion_Pall < 0."""
    return df[df["Reposicion_Pall"] < 0].copy().reset_index(drop=True)


# ─── SNAPSHOT LOCAL (Sprint 5 → Drive) ──────────────────────────────────────

def make_run_id() -> str:
    """Run ID único por corrida (timestamp + hash corto)."""
    ts = datetime.now().strftime("%H%M%S")
    h = hashlib.sha1(ts.encode()).hexdigest()[:6]
    return f"{ts}_{h}"

def snapshot_local(run_id: str, files: dict[str, bytes], log_text: str):
    """Guarda los archivos y el log en ./snapshots/YYYY-MM-DD/<run_id>/."""
    fecha = datetime.today().strftime("%Y-%m-%d")
    base = SNAPSHOT_DIR / fecha / run_id
    base.mkdir(parents=True, exist_ok=True)
    for name, data in files.items():
        with open(base / name, "wb") as f:
            f.write(data)
    with open(base / "run.log", "w", encoding="utf-8") as f:
        f.write(log_text)
    return base


# ─── HELPERS UI ─────────────────────────────────────────────────────────────

def _stamp(prefix: str, ext: str) -> str:
    """Nombre de archivo con timestamp para descargas."""
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"

def _download_trio(df: pd.DataFrame, base_name: str, sheet_name: str, key_prefix: str):
    """Renderiza 3 botones de descarga (TSV / CSV / XLSX) en columnas."""
    c1, c2, c3 = st.columns(3)
    with c1:
        st.download_button(
            "⬇ TSV (pegar Sheets)",
            data=df_to_tsv(df),
            file_name=_stamp(base_name, "tsv"),
            mime="text/tab-separated-values",
            key=f"{key_prefix}_tsv",
            use_container_width=True,
        )
    with c2:
        st.download_button(
            "⬇ CSV",
            data=df_to_csv(df),
            file_name=_stamp(base_name, "csv"),
            mime="text/csv",
            key=f"{key_prefix}_csv",
            use_container_width=True,
        )
    with c3:
        st.download_button(
            "⬇ XLSX",
            data=df_to_xlsx(df, sheet_name),
            file_name=_stamp(base_name, "xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_xlsx",
            use_container_width=True,
        )


# ============================================================================
# ║                            UI STREAMLIT                                  ║
# ============================================================================

def render_sidebar():
    """Sidebar global con toggles y badges de versión."""
    with st.sidebar:
        st.markdown(f"### 🟦 Picking Orchestrator")
        st.caption(f"v{APP_VERSION}")
        st.caption("Beccacece Hnos SA — DPO 2.1")
        st.divider()

        st.session_state["dry_run"] = st.toggle(
            "🧪 Modo Dry-Run",
            value=st.session_state.get("dry_run", False),
            help="Procesa todo pero NO genera PDFs finales (solo preview en pantalla). "
                 "Recomendado durante la primera semana de validación.",
        )
        st.session_state["snapshot_on"] = st.toggle(
            "💾 Snapshot activo",
            value=st.session_state.get("snapshot_on", True),
            help=f"Guarda CAR + MASTER + PDFs + log en {SNAPSHOT_DIR}/<fecha>/<run_id>/ "
                 "cada corrida. Default ON. Local hasta Sprint 5 (después → Drive).",
        )

        st.divider()
        st.caption(f"📅 Hoy: {datetime.today().strftime('%d/%m/%Y')}")
        if st.session_state.get("dry_run"):
            st.warning("🧪 DRY-RUN ACTIVO\n\nNo se generan PDFs finales.")

        st.divider()
        with st.expander("🛠 Stack técnico", expanded=False):
            st.code(
                f"streamlit {st.__version__}\n"
                f"pandas {pd.__version__}\n"
                f"openpyxl {openpyxl.__version__}",
                language="text",
            )


def _file_uploader(label: str, types: list[str], key: str, required: bool = True):
    """Wrapper de st.file_uploader con label rico."""
    suffix = " *" if required else " (opcional)"
    return st.file_uploader(
        label + suffix, type=types, key=key, accept_multiple_files=False,
    )


# ── TAB 1 — Planilla de Carga (REUSO v3.9) ─────────────────────────────────

def render_tab_planilla():
    st.subheader("📦 Planilla de Carga")
    st.caption(
        "Genera el PDF multi-página de la planilla diaria de picking. "
        "Lógica idéntica a la app v3.9.1 validada en producción."
    )

    col_a, col_b = st.columns(2)
    with col_a:
        car_file = _file_uploader(
            "CAR.xlsx (export Chess)", ["xlsx"], key="t1_car"
        )
    with col_b:
        fr_file = _file_uploader(
            "Frescura 3.0.xlsx", ["xlsx"], key="t1_fr"
        )

    if not (car_file and fr_file):
        st.info("Subí ambos archivos para generar la Planilla.")
        return

    if st.button("🚀 Generar Planilla de Carga", type="primary", key="t1_gen"):
        try:
            with st.spinner("Cargando Frescura..."):
                api, ddm, fr, fr_diag = load_frescura(fr_file)
            log_event("info", f"Frescura cargada: API={len(api)} | DDM={len(ddm)} | FR={len(fr)}")

            with st.spinner("Cargando CAR..."):
                car_df, blue_audit = load_car(car_file, ddm=ddm)
            log_event("info", f"CAR cargado: {len(car_df)} filas | bloque azul={len(blue_audit)}")

            if st.session_state.get("dry_run"):
                st.success(f"✓ DRY-RUN OK — {len(car_df)} filas listas. PDF no generado.")
                with st.expander("Preview CAR (primeras 20 filas)"):
                    st.dataframe(car_df.head(20), use_container_width=True)
                return

            with st.spinner("Generando PDF (puede tardar)..."):
                pdf_bytes, stats = generate_pdf(car_df, api, ddm, fr)

            log_event("info", f"Planilla generada: {stats['total_pages']} páginas, "
                              f"{len(stats['repartos'])} repartos")

            st.success(
                f"✓ Planilla lista — {stats['total_pages']} páginas | "
                f"{len(stats['repartos'])} repartos"
            )

            st.download_button(
                "⬇ Descargar Planilla de Carga (PDF)",
                data=pdf_bytes,
                file_name=_stamp("Planilla_Carga", "pdf"),
                mime="application/pdf",
                type="primary",
                use_container_width=True,
            )

            with st.expander(f"📊 Alertas y diagnósticos"):
                a, b, c, d = st.columns(4)
                a.metric("🔴 RED", len(stats["red"]))
                b.metric("🟡 YELLOW", len(stats["yellow"]))
                c.metric("📦 c/Pallet", len(stats["pallet_applied"]))
                d.metric("⚠ Sin BXP", len(stats["miss_bxp"]))
                if stats["orphans_no_ddm"]:
                    st.warning("Huérfanos sin DDM:")
                    for o in stats["orphans_no_ddm"]:
                        st.write(f"- {o}")

            # Stash para snapshot (Sprint 1: solo en memoria)
            st.session_state["last_planilla_pdf"] = pdf_bytes

        except Exception as e:
            log_event("error", f"Error en Planilla: {e}")
            st.error(f"❌ Error generando Planilla: {e}")
            with st.expander("Stack trace"):
                import traceback
                st.code(traceback.format_exc())


# ── TAB 2 — Resumen por Camión (REUSO v3.9) ────────────────────────────────

def render_tab_resumen():
    st.subheader("📋 Resumen de Carga por Camión")
    st.caption(
        "Genera un PDF compacto con bultos consolidados por SKU por camión "
        "(hoja AGR del CAR). Útil para tildar al final de la carga."
    )

    car_file = _file_uploader(
        "CAR.xlsx (export Chess — debe tener hoja AGR)", ["xlsx"], key="t2_car"
    )

    if not car_file:
        st.info("Subí el CAR.xlsx con la hoja AGR para generar el Resumen.")
        return

    if st.button("🚀 Generar Resumen por Camión", type="primary", key="t2_gen"):
        try:
            if st.session_state.get("dry_run"):
                with st.spinner("Leyendo hoja AGR..."):
                    df_agr = load_agr(car_file)
                st.success(
                    f"✓ DRY-RUN OK — {len(df_agr)} filas SKU/camión, "
                    f"{df_agr['chofer'].nunique()} camiones."
                )
                st.dataframe(df_agr.head(30), use_container_width=True)
                return

            with st.spinner("Generando PDF resumen..."):
                pdf_bytes, stats = build_resumen_carga_pdf(car_file)

            log_event("info", f"Resumen generado: {stats['camiones']} camiones, "
                              f"{stats['filas_sku']} filas SKU, "
                              f"{stats['total_pages']} páginas")

            st.success(
                f"✓ Resumen listo — {stats['camiones']} camiones | "
                f"{stats['filas_sku']} filas SKU | {stats['total_pages']} páginas"
            )

            st.download_button(
                "⬇ Descargar Resumen por Camión (PDF)",
                data=pdf_bytes,
                file_name=_stamp("Resumen_Camiones", "pdf"),
                mime="application/pdf",
                type="primary",
                use_container_width=True,
            )

            st.session_state["last_resumen_pdf"] = pdf_bytes

        except Exception as e:
            log_event("error", f"Error en Resumen: {e}")
            st.error(f"❌ Error generando Resumen: {e}")
            with st.expander("Stack trace"):
                import traceback
                st.code(traceback.format_exc())


# ── TAB 3 — Camiones T2: Imprimir hojas desde Sheets ───────────────────────
#
#  LÓGICA:
#  1. Usuario sube CAR.xlsx → se extraen los números de camión con datos
#  2. Se construye la URL de impresión de la hoja correspondiente en el
#     T2 Status Carga Sheets (cada camión tiene su propia hoja: "101", "102", etc.)
#  3. Se abre cada URL en una nueva pestaña via JS → el browser lanza
#     el diálogo de impresión nativo (con el visual del Sheets intacto)
#
#  SHEETS ID: 1QCoMtpHcUaTITp9p9-1mo914HsaH0siq_bHL7nmJ_DA
#  URL de impresión por hoja: https://docs.google.com/spreadsheets/d/{ID}/export
#    → Alternativa más confiable: abrir la hoja con ?gid=XXX&rm=print
#    → Para abrir pestaña de impresión del browser: usar la URL de edit con
#      el parámetro rm=print aplicado al gid específico de cada hoja.
#
#  NOTA: el gid de cada hoja (101, 102, etc.) se puede hardcodear o leer
#  dinámicamente. En v4.8 se lee dinámicamente via gspread (opcional).
#  Por ahora la estrategia es abrir la URL de cada hoja en modo "edit" con
#  focus en esa hoja — el usuario imprime con Ctrl+P — O mejor aún: usar
#  la URL de exportación a PDF directa de Google Sheets por hoja (sin login
#  extra si ya está autenticado en el browser).
#
#  URL export PDF por hoja (requiere sesión Google activa en el browser):
#  https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=pdf
#    &gid={GID}&portrait=false&fitw=true&gridlines=false
#    &top_margin=0.25&bottom_margin=0.25&left_margin=0.25&right_margin=0.25
# ─────────────────────────────────────────────────────────────────────────────

# ID del Sheets T2 Status Carga (fijo)
_T2_SHEETS_ID = "1QCoMtpHcUaTITp9p9-1mo914HsaH0siq_bHL7nmJ_DA"

# Mapa camión → gid de la hoja en el Sheets.
# Se completa con los gid reales del Sheets (se pueden ver en la URL al hacer
# click en cada pestaña: ...#gid=XXXXXXX).
# Si un camión no está en este mapa se usa la URL de la hoja por nombre.
# IMPORTANTE: completar este dict con los gid reales del Sheets.
# Los valores actuales son PLACEHOLDERS — reemplazar con los gid correctos.
_T2_GID_MAP: dict[int, int] = {
    # camion: gid  ← obtener de la URL del Sheets al clickear cada pestaña
    101: 157072406,   # ejemplo real del screenshot
    # 102: XXXXXXX,
    # 103: XXXXXXX,
    # ... completar según el Sheets real
}


def _t2_build_print_url(sheet_id: str, gid: int) -> str:
    """
    Construye la URL de exportación PDF de una hoja específica del Sheets.
    Orientación horizontal, sin grillas, márgenes mínimos, ajustado a página.
    Requiere que el usuario esté autenticado en Google en el browser.
    """
    base = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export"
    params = (
        f"?format=pdf"
        f"&gid={gid}"
        f"&portrait=false"
        f"&fitw=true"
        f"&gridlines=false"
        f"&printtitle=false"
        f"&sheetnames=false"
        f"&pagenum=UNDEFINED"
        f"&attachment=false"
        f"&top_margin=0.25"
        f"&bottom_margin=0.25"
        f"&left_margin=0.25"
        f"&right_margin=0.25"
    )
    return base + params


def _t2_build_view_url(sheet_id: str, gid: int) -> str:
    """URL de visualización directa de la hoja (fallback si export falla)."""
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={gid}"


@st.cache_data(show_spinner=False)
def _t2_load_camiones_car(car_bytes: bytes) -> list[int]:
    """
    Lee el CAR.xlsx y devuelve la lista de números de camión que tienen
    datos (Bultos > 0), ordenados ascendentemente.
    Usa la misma lógica de load_car pero simplificada: solo extrae transportes.
    """
    raw = pd.read_excel(io.BytesIO(car_bytes), sheet_name=0, header=None)
    header = raw.iloc[0].tolist()

    # Filas normales (fila 42 en adelante → índice 41+)
    normal = raw.iloc[41:].copy()
    normal.columns = header

    col_trans = None
    col_bult  = None
    for c in header:
        nc = _norm(str(c))
        if nc in ("transporte",):
            col_trans = c
        if nc in ("bultos",):
            col_bult = c

    if col_trans is None:
        raise ValueError("No se encontró columna 'Transporte' en el CAR.")

    normal[col_trans] = pd.to_numeric(normal[col_trans], errors="coerce")
    if col_bult:
        normal[col_bult] = pd.to_numeric(normal[col_bult], errors="coerce").fillna(0)

    normal = normal.dropna(subset=[col_trans])
    normal = normal[normal[col_trans] > 0]
    normal[col_trans] = normal[col_trans].apply(lambda x: int(float(x)))

    if col_bult:
        normal = normal[normal[col_bult] > 0]

    camiones = sorted(normal[col_trans].unique().tolist())
    return camiones


def _t2_js_open_tabs(urls: list[str]) -> str:
    """
    Genera el HTML+JS que abre cada URL en una nueva pestaña del browser.
    Usa window.open() con un pequeño delay entre pestañas para evitar que
    el bloqueador de popups las cancele. Se avisa al usuario que debe
    permitir popups para este sitio.
    """
    js_lines = []
    for i, url in enumerate(urls):
        delay_ms = i * 600  # 600ms entre pestañas
        js_lines.append(
            f"setTimeout(function(){{ window.open('{url}', '_blank'); }}, {delay_ms});"
        )
    js_block = "\n    ".join(js_lines)

    html = f"""
    <div style="
        background: #f0f9ff;
        border: 1px solid #0ea5e9;
        border-radius: 8px;
        padding: 14px 18px;
        font-family: sans-serif;
        font-size: 13px;
        color: #0c4a6e;
        margin-bottom: 10px;
    ">
        ⚠️ <b>Si el browser bloquea las pestañas:</b> hacé click en el ícono de popup
        bloqueado en la barra de direcciones y seleccioná <i>"Permitir siempre"</i>.
        Luego presioná el botón nuevamente.
    </div>
    <button
        onclick="abrirHojas()"
        style="
            background: #1a3a6b;
            color: white;
            border: none;
            border-radius: 6px;
            padding: 10px 22px;
            font-size: 15px;
            font-weight: 600;
            cursor: pointer;
            letter-spacing: 0.3px;
        "
    >
        🖨️ Abrir {len(urls)} hoja(s) para imprimir
    </button>
    <script>
    function abrirHojas() {{
        {js_block}
    }}
    </script>
    """
    return html


def render_tab_t2():
    import streamlit.components.v1 as components

    st.subheader("🚛 Camiones T2 — Imprimir hojas")
    st.caption(
        "Subí el CAR del día → detecta los camiones con reparto → "
        "abre las hojas del T2 Status Carga listas para imprimir (una pestaña por camión)."
    )

    # ── Instrucciones colapsables ─────────────────────────────────────────────
    with st.expander("ℹ️ Cómo funciona", expanded=False):
        st.markdown(
            """
            1. **Subís el CAR.xlsx** del día (el mismo que usás en Tab 1).
            2. La app detecta automáticamente qué camiones tienen bultos asignados.
            3. Presionás **"Abrir hojas para imprimir"** → se abre una pestaña
               por camión directamente en el T2 Status Carga Sheets.
            4. Cada pestaña abre en modo **PDF export** → el browser muestra el
               diálogo de impresión con la hoja exacta como la ves en el Sheets.
            5. Imprimís o guardás como PDF normalmente.

            > **Requisito:** tenés que estar logueado en Google con la cuenta
            > que tiene acceso al Sheets. El Sheets no necesita ser público.

            > **Popups:** permitir popups para este sitio la primera vez.
            """
        )

    # ── GID Map status ────────────────────────────────────────────────────────
    gid_configurados = [c for c in _T2_GID_MAP]
    with st.expander(f"⚙️ GIDs configurados ({len(gid_configurados)} camiones)", expanded=False):
        st.caption(
            "El mapa camión→GID se completa en el código (`_T2_GID_MAP`). "
            "Si un camión del CAR no está en el mapa, se muestra como 'sin GID' y se omite. "
            "Para obtener el GID: abrí el Sheets → hacé click en la pestaña del camión → "
            "copiá el número después de `#gid=` en la URL."
        )
        if gid_configurados:
            gid_data = [{"Camión": c, "GID": _T2_GID_MAP[c]} for c in sorted(gid_configurados)]
            st.dataframe(pd.DataFrame(gid_data), use_container_width=True, hide_index=True)
        else:
            st.warning("No hay GIDs configurados. Completá `_T2_GID_MAP` en el código.")

    st.divider()

    # ── Upload CAR ────────────────────────────────────────────────────────────
    car_file = _file_uploader("CAR.xlsx del día", ["xlsx"], key="t3_car")

    if not car_file:
        st.info("Subí el CAR para detectar los camiones con reparto.")
        return

    # ── Detectar camiones ─────────────────────────────────────────────────────
    try:
        with st.spinner("Leyendo camiones del CAR..."):
            camiones_car = _t2_load_camiones_car(car_file.getvalue())
        log_event("info", f"T2: CAR leído → {len(camiones_car)} camiones: {camiones_car}")
    except Exception as e:
        st.error(f"❌ Error leyendo CAR: {e}")
        log_event("error", f"T2: Error leyendo CAR: {e}")
        return

    if not camiones_car:
        st.warning("No se encontraron camiones con bultos en el CAR.")
        return

    # ── Cruzar con GID map ────────────────────────────────────────────────────
    con_gid    = [c for c in camiones_car if c in _T2_GID_MAP]
    sin_gid    = [c for c in camiones_car if c not in _T2_GID_MAP]

    # Métricas
    col1, col2, col3 = st.columns(3)
    col1.metric("🚛 Camiones en CAR", len(camiones_car))
    col2.metric("✅ Con hoja configurada", len(con_gid))
    col3.metric("⚠️ Sin GID (se omiten)", len(sin_gid))

    if sin_gid:
        st.warning(
            f"Los siguientes camiones del CAR no tienen GID configurado y **no se imprimirán**: "
            f"{', '.join(str(c) for c in sin_gid)}. "
            f"Agregá su GID en `_T2_GID_MAP` del código."
        )

    if not con_gid:
        st.error(
            "Ningún camión del CAR tiene GID configurado. "
            "Completá `_T2_GID_MAP` con los GIDs del Sheets antes de continuar."
        )
        return

    # ── Tabla de camiones a imprimir ──────────────────────────────────────────
    st.markdown(f"#### 🖨️ Camiones a imprimir: **{len(con_gid)}**")

    rows_preview = []
    for c in sorted(con_gid):
        gid = _T2_GID_MAP[c]
        url = _t2_build_print_url(_T2_SHEETS_ID, gid)
        rows_preview.append({
            "Camión": c,
            "GID": gid,
            "URL (preview)": f"...gid={gid}&format=pdf",
        })

    st.dataframe(
        pd.DataFrame(rows_preview),
        use_container_width=True,
        hide_index=True,
    )

    # ── Opciones adicionales ──────────────────────────────────────────────────
    st.markdown("---")
    col_opt1, col_opt2 = st.columns(2)
    with col_opt1:
        modo = st.radio(
            "Modo de apertura",
            ["📄 Export PDF directo (recomendado)", "🔗 Abrir hoja en el Sheets (manual)"],
            key="t3_modo",
            help=(
                "Export PDF: el browser abre el PDF del Sheets directamente para imprimir. "
                "Hoja Sheets: abre la pestaña de edición del camión en el Sheets."
            ),
        )
    with col_opt2:
        camiones_sel = st.multiselect(
            "Filtrar camiones (dejar vacío = todos)",
            options=sorted(con_gid),
            default=[],
            key="t3_sel",
            help="Seleccioná solo los camiones que querés imprimir ahora.",
        )

    camiones_a_imprimir = camiones_sel if camiones_sel else sorted(con_gid)

    # ── Construir URLs ────────────────────────────────────────────────────────
    use_export = "Export PDF" in modo
    urls = []
    for c in camiones_a_imprimir:
        gid = _T2_GID_MAP[c]
        if use_export:
            urls.append(_t2_build_print_url(_T2_SHEETS_ID, gid))
        else:
            urls.append(_t2_build_view_url(_T2_SHEETS_ID, gid))

    st.markdown(f"**{len(urls)} hoja(s) seleccionadas:** {', '.join(str(c) for c in camiones_a_imprimir)}")

    # ── Botón JS para abrir pestañas ──────────────────────────────────────────
    html_btn = _t2_js_open_tabs(urls)
    components.html(html_btn, height=120)

    # ── Links individuales como fallback ──────────────────────────────────────
    with st.expander("🔗 Links individuales (fallback si el botón no funciona)", expanded=False):
        for c, url in zip(camiones_a_imprimir, urls):
            st.markdown(f"- **Camión {c}:** [Abrir hoja]({url})")

    log_event(
        "info",
        f"T2: {len(urls)} URLs generadas | modo={'export' if use_export else 'view'} | "
        f"camiones={camiones_a_imprimir}",
    )



# ── TAB 4 — Proyección Picking ×4 — v4.6 ──────────────────────────────────
#
#  CAMBIOS v4.7:
#  - Fix: hora estimada de fin calculada sobre BULTOS REALES por cancha
#    (no sobre UP). Fórmula: bultos_cancha / (vel_bult_h × personas) + inicio
#  - Fix: asignaciones entre canchas ahora afectan correctamente los bultos
#    de cada cancha para el cálculo de tiempo (nueva fn _t4_calcular_bultos_por_cancha)
#  - Los metrics "Bult CX" muestran delta cuando hay asignaciones activas
#
#  CAMBIOS v4.6:
#  - Tabla basada en BULTOS UP (col R Frescura DDM = UNIDAD PKT)
#  - Paletas enteras (floor), redondeo inteligente (≥0.97 → techo)
#  - Hora de inicio única (un solo input compartido para todas las canchas)
#  - Hora estimada de fin calculada por bultos/hora × personas
#  - Sin toggle "Sacar ceros" (siempre filtrado)
#  - Columnas ASIGN siempre en 0 por default
#  - Asignación de carga: dropdown para mover carga entre canchas
#  - PDF mejorado: sin encimamiento, bultos UP por cancha, info asignaciones
#  - Columna camiones correcta: 101=1, 106=2, etc.
# ─────────────────────────────────────────────────────────────────────────────

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Constantes Tab 4 ─────────────────────────────────────────────────────────

_T4_CANCHAS      = ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "MKPL"]
_T4_CANCHAS_PDF  = ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV"]

_T4_VEL_DEFAULT = {
    "CANCHA I":   370,
    "CANCHA II":  360,
    "CANCHA III": 390,
    "CANCHA IV":  400,
    "MKPL":       390,
}
_T4_HORA_INICIO_DEFAULT = {
    "CANCHA I":   (17, 30),
    "CANCHA II":  (17, 35),
    "CANCHA III": (17, 35),
    "CANCHA IV":  (17, 40),
    "MKPL":       (17, 45),
}
_T4_CAP_MAX = 9

_T4_CANCHA_COLORS = {
    "CANCHA I":   colors.HexColor("#1565C0"),
    "CANCHA II":  colors.HexColor("#2E7D32"),
    "CANCHA III": colors.HexColor("#E65100"),
    "CANCHA IV":  colors.HexColor("#6A1B9A"),
    "MKPL":       colors.HexColor("#00838F"),
}


def _t4_safe(v, default=0.0):
    if v is None:
        return default
    try:
        f = float(v)
        return default if (f != f) else f
    except (ValueError, TypeError):
        return default


def _t4_norm_cancha(val) -> str:
    if val is None:
        return "SIN CANCHA"
    s = str(val).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    if any(k in s for k in ("MKPL", "MARKETPLACE", "MERCH")):
        return "MKPL"
    for c in ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "CANCHA V"]:
        if s == c or s.startswith(c):
            return c
    return "SIN CANCHA"


@st.cache_data(show_spinner=False)
def _t4_load_car_proyeccion(car_bytes: bytes, fr_bytes: bytes) -> dict:
    """
    v4.6 — Proyección basada en BULTOS UP (UNIDAD PKT, col R de DDM Frescura).

    Lógica paletas enteras por camión × SKU:
      bxp      = col G DDM (bultos x pallet)
      upkt     = col R DDM (UNIDAD PKT = fracción de pallet por bulto = 1/bxp o similar)
      pall_raw = bultos_total × upkt
      REDONDEO: si pall_raw - floor(pall_raw) >= 0.97 → ceil (era casi entero)
      pall_ae  = floor_inteligente(pall_raw)
      pall_pick_frac = pall_raw - pall_ae        ← lo que queda para picking (fracción de pallet)
      bult_ae  = pall_ae / upkt                  ← bultos del AE
      bult_pick = bultos_total - bult_ae          ← bultos de picking
      up_pick   = pall_pick_frac                  ← fracción de paleta a pickear
    """
    import datetime as _dt
    import math as _math

    # ── 1. Cargar Frescura → api (cancha), ddm (bxp + upkt) ─────────────────
    api, ddm, fr, _diag = load_frescura(io.BytesIO(fr_bytes))

    # Agregar UNIDAD PKT (col R) si no está en ddm
    # load_frescura ya carga DDM sheet; necesitamos releer para col R
    try:
        wb_fr = openpyxl.load_workbook(io.BytesIO(fr_bytes), read_only=True, data_only=True)
        ws_ddm = wb_fr["DDM"]
        upkt_rows = []
        for row in ws_ddm.iter_rows(min_row=2, values_only=True):
            sku_  = row[2]   # col C
            bxp_  = row[6]   # col G
            upkt_ = row[17]  # col R
            if sku_ is not None:
                try:
                    upkt_rows.append({
                        "sku":  int(float(str(sku_))),
                        "bxp":  float(bxp_) if bxp_ not in (None, "") else 0.0,
                        "upkt": float(upkt_) if upkt_ not in (None, "") else 0.0,
                    })
                except (ValueError, TypeError):
                    pass
        wb_fr.close()
        df_upkt = pd.DataFrame(upkt_rows).drop_duplicates("sku")
    except Exception:
        df_upkt = pd.DataFrame(columns=["sku", "bxp", "upkt"])

    # ── 2. Leer Hoja1 del CAR ────────────────────────────────────────────────
    raw = pd.read_excel(io.BytesIO(car_bytes), sheet_name=0, header=None)
    header = raw.iloc[0].tolist()
    normal = raw.iloc[41:].copy()
    normal.columns = header

    normal["Artículo"]   = pd.to_numeric(normal["Artículo"],   errors="coerce")
    normal["Transporte"] = pd.to_numeric(normal["Transporte"], errors="coerce")
    normal["Bultos"]     = pd.to_numeric(normal["Bultos"],     errors="coerce").fillna(0)
    normal["Unids"]      = pd.to_numeric(normal["Unids"],      errors="coerce").fillna(0)
    normal = normal.dropna(subset=["Artículo", "Transporte"])
    normal = normal[normal["Artículo"] > 0]
    normal["Artículo"]   = normal["Artículo"].astype(int)
    normal["Transporte"] = normal["Transporte"].apply(lambda x: int(float(x)) if pd.notna(x) else x)

    # Fecha desde columna Fecha Mvto
    try:
        fecha = pd.to_datetime(normal["Fecha Mvto"].dropna().iloc[0]).date()
    except Exception:
        fecha = _dt.date.today()

    # ── 3. Excluir envases ───────────────────────────────────────────────────
    normal = normal[~normal.apply(
        lambda r: is_envase(int(r["Artículo"]), str(r.get("Descripción Artículo", ""))), axis=1
    )]

    # ── 4. Agregar por (Transporte, Artículo) ────────────────────────────────
    grp = (normal.groupby(["Transporte", "Artículo", "Descripción Artículo"], as_index=False)
           .agg(blt_raw=("Bultos", "sum"), unids=("Unids", "sum")))
    grp = grp[grp["blt_raw"] > 0]
    grp.rename(columns={"Transporte": "cam", "Artículo": "sku"}, inplace=True)

    # ── 5. Merge UPKT (DDM col R) ────────────────────────────────────────────
    grp = grp.merge(df_upkt[["sku", "bxp", "upkt"]], on="sku", how="left")

    # ── 6. Calcular picking vs AE usando UNIDAD PKT ──────────────────────────
    def _split_row(row):
        tot   = float(row["blt_raw"])
        bxp_  = float(row["bxp"])  if pd.notna(row["bxp"])  and row["bxp"]  > 0 else 0.0
        upkt_ = float(row["upkt"]) if pd.notna(row["upkt"]) and row["upkt"] > 0 else 0.0

        if upkt_ > 0:
            pall_raw = tot * upkt_          # paletas totales (con decimales)
            # Redondeo inteligente: si fracción ≥ 0.97, consideramos paleta completa
            frac = pall_raw - _math.floor(pall_raw)
            pall_ae = _math.ceil(pall_raw) if frac >= 0.97 else _math.floor(pall_raw)
            pall_pick_frac = max(0.0, pall_raw - pall_ae)  # fracción a pickear
            # Bultos correspondientes
            bult_ae   = (pall_ae / upkt_) if upkt_ > 0 else 0.0
            bult_ae   = min(bult_ae, tot)  # no puede superar el total
            bult_pick = max(0.0, tot - bult_ae)
            up_pick   = pall_pick_frac
        elif bxp_ > 0:
            # Fallback si upkt no disponible: usar bxp clásico
            pall_ae   = _math.floor(tot / bxp_)
            bult_ae   = pall_ae * bxp_
            bult_pick = tot - bult_ae
            up_pick   = bult_pick / bxp_ if bult_pick > 0 else 0.0
            pall_raw  = tot / bxp_
        else:
            pall_ae   = 0
            bult_ae   = 0.0
            bult_pick = tot
            up_pick   = 0.0
            pall_raw  = 0.0

        return pd.Series({
            "pall_ae":   float(pall_ae),
            "bult_ae":   bult_ae,
            "bult_pick": bult_pick,
            "up_pick":   up_pick,
        })

    split = grp.apply(_split_row, axis=1)
    grp = pd.concat([grp, split], axis=1)

    # ── 7. Merge cancha (API) ────────────────────────────────────────────────
    grp = grp.merge(api[["sku", "cancha"]], on="sku", how="left")
    grp["cancha_norm"] = grp["cancha"].fillna("SIN CANCHA").apply(_t4_norm_cancha)

    # ── 8. Pivotar por camión × cancha ───────────────────────────────────────
    pick_grp = grp.groupby(["cam", "cancha_norm"], as_index=False).agg(
        bult_pick=("bult_pick", "sum"),
        up_pick  =("up_pick",   "sum"),
    )

    ae_grp = grp.groupby("cam", as_index=False).agg(
        bult_ae=("bult_ae", "sum"),
        up_ae  =("pall_ae", "sum"),
    )

    pivot_bult = pick_grp.pivot_table(
        index="cam", columns="cancha_norm",
        values="bult_pick", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_bult.columns.name = None

    pivot_up = pick_grp.pivot_table(
        index="cam", columns="cancha_norm",
        values="up_pick", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_up.columns.name = None
    up_rename = {c: f"_up_{c}" for c in _T4_CANCHAS + ["SIN CANCHA"] if c in pivot_up.columns}
    pivot_up = pivot_up.rename(columns=up_rename)

    cam_df = pivot_bult.merge(ae_grp, on="cam", how="left")
    cam_df = cam_df.merge(pivot_up, on="cam", how="left")

    for c in _T4_CANCHAS:
        if c not in cam_df.columns:
            cam_df[c] = 0.0
        if f"_up_{c}" not in cam_df.columns:
            cam_df[f"_up_{c}"] = 0.0
        cam_df[f"_hl_{c}"] = 0.0

    cam_df["bult_ae"] = cam_df["bult_ae"].fillna(0.0)
    cam_df["up_ae"]   = cam_df["up_ae"].fillna(0.0)
    cam_df["hl_ae"]   = 0.0
    cam_df["kg"]      = 0.0

    cam_df["TOTAL_PICK"] = cam_df[[c for c in _T4_CANCHAS if c in cam_df.columns]].sum(axis=1)
    cam_df["TOTAL_AE"]   = cam_df["bult_ae"]
    cam_df["TOTAL_BULT"] = cam_df["TOTAL_PICK"] + cam_df["TOTAL_AE"]
    cam_df["TOTAL_PALL"] = cam_df[[f"_up_{c}" for c in _T4_CANCHAS]].sum(axis=1) + cam_df["up_ae"]
    cam_df["AE_PALL"]    = cam_df["up_ae"]

    cam_df = cam_df.rename(columns={"cam": "Camión"})
    cam_df = cam_df.sort_values("Camión").reset_index(drop=True)

    tot_pick = float(cam_df["TOTAL_PICK"].sum())
    tot_ae   = float(cam_df["TOTAL_AE"].sum())
    mix = (tot_pick / (tot_pick + tot_ae)) if (tot_pick + tot_ae) > 0 else 0.0

    return {
        "df":          cam_df,
        "fecha":       fecha,
        "fuente":      "CAR + Frescura (UP = UNIDAD PKT col R)",
        "mix_picking": mix,
        "tot_pick":    tot_pick,
        "tot_ae":      tot_ae,
    }


def _t4_calcular_pall_por_cancha(df: pd.DataFrame, asign_state: dict) -> dict:
    """
    Aplica ASIGN (reasignación entre canchas), calcula SUB, DESIGNADOS, TOTAL.
    asign_state = {(camion, cancha_origen): cancha_destino_str}
      - None / "" → sin asignación
      - "CANCHA X" → mueve los UP de esa línea a la cancha destino
    """
    sub        = {c: 0.0 for c in _T4_CANCHAS}
    designados = {c: 0.0 for c in _T4_CANCHAS}
    status_rows = {}
    asign_detail = {}  # {(cam, cancha_origen): cancha_destino}

    for _, row in df.iterrows():
        cam = int(row["Camión"])
        total_pall_cam = 0.0

        for c in _T4_CANCHAS:
            up_val    = float(row.get(f"_up_{c}", 0.0))
            asign_val = asign_state.get((cam, c))

            if asign_val and isinstance(asign_val, str) and asign_val.strip():
                dest = _t4_norm_cancha(asign_val)
                if dest in _T4_CANCHAS and dest != c:
                    designados[dest] += up_val
                    designados[c]    -= up_val
                    asign_detail[(cam, c)] = dest
            sub[c] += up_val
            total_pall_cam += up_val

        total_pall_cam += float(row.get("AE_PALL", 0.0))
        if total_pall_cam == 0:
            status = "CERO"
        elif total_pall_cam > _T4_CAP_MAX:
            status = f">{_T4_CAP_MAX} Pall"
        else:
            status = "OK"
        status_rows[cam] = {"status": status, "total_pall": total_pall_cam}

    return {
        "sub":          sub,
        "designados":   designados,
        "total":        {c: sub[c] + designados[c] for c in _T4_CANCHAS},
        "status_rows":  status_rows,
        "asign_detail": asign_detail,
    }


def _t4_hora_fin(bultos: float, vel: int, personas: int, inicio) -> object:
    """Calcula hora estimada de fin dado bultos, velocidad (bult/h) y personas."""
    import datetime as _dt
    inicio_dt = _dt.datetime.combine(_dt.date.today(), inicio)
    if vel <= 0 or personas <= 0 or bultos <= 0:
        return inicio_dt
    return inicio_dt + _dt.timedelta(hours=bultos / (vel * personas))


def _t4_calcular_bultos_por_cancha(df: pd.DataFrame, asign_state: dict) -> dict:
    """
    Calcula los BULTOS REALES de picking por cancha aplicando reasignaciones (ASIGN).

    La reasignación mueve los bultos de la cancha origen a la cancha destino.
    La proporción de bultos a mover se calcula como:
        bult_reasign = bult_cancha_origen * (up_origen_asignado / up_origen_total)
    Si up_origen_total == 0, se mueven todos los bultos de esa cancha.

    Retorna: {cancha: bultos_totales_con_asign}
    """
    sub_bult   = {c: 0.0 for c in _T4_CANCHAS}
    delta_bult = {c: 0.0 for c in _T4_CANCHAS}

    for _, row in df.iterrows():
        cam = int(row["Camión"])
        for c in _T4_CANCHAS:
            bult_val = float(row.get(c, 0.0))   # bultos de picking en esta cancha
            up_val   = float(row.get(f"_up_{c}", 0.0))  # fracción pallet de esta cancha
            sub_bult[c] += bult_val

            asign_val = asign_state.get((cam, c))
            if asign_val and isinstance(asign_val, str) and asign_val.strip():
                dest = _t4_norm_cancha(asign_val)
                if dest in _T4_CANCHAS and dest != c:
                    # Mover todos los bultos de esta cancha al destino
                    delta_bult[dest] += bult_val
                    delta_bult[c]    -= bult_val

    return {c: max(0.0, sub_bult[c] + delta_bult[c]) for c in _T4_CANCHAS}


def _t4_generar_pdf_x4(
    df, totales_bult, totales_pall, totales_hl, totales_kg,
    productividad, personas, inicio, fecha, mix_picking, fin_calc,
    asign_detail=None,
):
    """
    Genera PDF A4 portrait con una página por cancha (CI/CII/CIII/CIV).
    Muestra:
      - Bultos UP (fracción de paleta) a pickear por cancha y camión
      - Paletas puras (AE) en rojo
      - Total paletas por camión
      - Info de asignaciones (recibe / envía carga)
    """
    import datetime as _dt
    import math as _math

    if asign_detail is None:
        asign_detail = {}

    buf    = io.BytesIO()
    c_pdf  = rl_canvas.Canvas(buf, pagesize=A4)
    pw, ph = A4
    M      = 10 * mm

    # Colores
    HDR_BG_PDF   = colors.HexColor("#1a3a6b")
    SUB_BG_PDF   = colors.HexColor("#2e5fa3")
    TOT_BG_PDF   = colors.HexColor("#FF8C00")
    OK_BG        = colors.HexColor("#C8E6C9")
    OVER_BG      = colors.HexColor("#FFCDD2")
    ZERO_BG      = colors.HexColor("#F5F5F5")
    ALT_ROW      = colors.HexColor("#F0F4FF")
    WHITE        = colors.white
    GREEN_PICK   = colors.HexColor("#66BB6A")   # verde = bultos a pickear
    RED_AE       = colors.HexColor("#EF9A9A")   # rojo  = paletas puras AE
    YELLOW_FOCUS = colors.HexColor("#FFF9C4")
    ASIGN_BG     = colors.HexColor("#E3F2FD")   # azul claro = carga recibida

    FONT_B = "Helvetica-Bold"
    FONT_N = "Helvetica"

    # Columnas tabla: CAM | CI | CII | CIII | CIV | MKPL | AE | TOT | STS
    col_labels = ["CAM"] + [c.replace("CANCHA ", "C") for c in _T4_CANCHAS] + ["AE", "TOTAL", "STS"]
    inner_w    = pw - 2 * M
    cw = [28, 30, 30, 30, 30, 26, 26, 34, 28]
    cw[-1] = max(20, inner_w - sum(cw[:-1]))

    fin_global = fin_calc["fin_global_dt"]

    # Calcular status_rows una sola vez
    asign_st = st.session_state.get("t4_asign", {})
    calc      = _t4_calcular_pall_por_cancha(df, asign_st)
    status_rows = calc["status_rows"]

    def _draw_page(cancha_focus: str):
        # Filtrar camiones con picking > 0 en cualquier cancha
        rows_data = []
        for _, row in df.iterrows():
            if float(row.get("TOTAL_PICK", 0)) <= 0:
                continue
            cam = int(row["Camión"])
            sr  = status_rows.get(cam, {})
            palls = {c: float(row.get(f"_up_{c}", 0.0)) for c in _T4_CANCHAS}

            # Detectar asignaciones para este camión
            recibe_de = []
            envia_a   = []
            for (c_cam, c_origen), c_dest in asign_detail.items():
                if c_cam == cam:
                    if c_origen == cancha_focus:
                        envia_a.append(c_dest.replace("CANCHA ", "C"))
                    if c_dest == cancha_focus:
                        recibe_de.append(c_origen.replace("CANCHA ", "C"))

            rows_data.append({
                "cam":       cam,
                "palls":     palls,
                "ae":        float(row.get("AE_PALL", 0.0)),
                "total":     float(sr.get("total_pall", 0.0)),
                "status":    sr.get("status", "—"),
                "recibe_de": recibe_de,
                "envia_a":   envia_a,
                # bultos de picking en la cancha foco
                "bult_foco": float(row.get(cancha_focus, 0.0)),
            })

        y = ph - M

        # ── Header principal ───────────────────────────────────────────────
        hdr_h = 20
        c_pdf.setFillColor(HDR_BG_PDF)
        c_pdf.rect(M, y - hdr_h, inner_w, hdr_h, fill=1, stroke=0)
        c_pdf.setFillColor(WHITE)
        c_pdf.setFont(FONT_B, 8)
        fecha_s  = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)
        inicio_s = inicio.strftime("%H:%M")
        fin_s    = fin_global.strftime("%H:%M")
        hdr_txt  = (
            f"PICKING   {fecha_s}   "
            f"INICIO: {inicio_s}   FIN EST: {fin_s}   "
            f"MIX PICKING: {mix_picking*100:.1f}%"
        )
        c_pdf.drawCentredString(M + inner_w / 2, y - hdr_h + 6, hdr_txt)
        y -= hdr_h + 2

        # ── Subheader cancha ───────────────────────────────────────────────
        sub_h = 15
        c_pdf.setFillColor(SUB_BG_PDF)
        c_pdf.rect(M, y - sub_h, inner_w, sub_h, fill=1, stroke=0)
        c_pdf.setFillColor(WHITE)
        c_pdf.setFont(FONT_B, 9)
        c_pdf.drawCentredString(
            M + inner_w / 2, y - sub_h + 4,
            f"COPIA PARA: {cancha_focus}",
        )
        y -= sub_h + 3

        # ── Header tabla ───────────────────────────────────────────────────
        th = 12
        c_pdf.setFillColor(colors.HexColor("#E8EAF6"))
        c_pdf.rect(M, y - th, inner_w, th, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.setFont(FONT_B, 6.5)
        x = M
        for i, lbl in enumerate(col_labels):
            c_pdf.drawCentredString(x + cw[i] / 2, y - th + 4, lbl)
            x += cw[i]
        y -= th

        # ── Filas camiones ─────────────────────────────────────────────────
        rh = 11
        for ri, rd in enumerate(rows_data):
            bg = ALT_ROW if ri % 2 == 0 else WHITE
            c_pdf.setFillColor(bg)
            c_pdf.rect(M, y - rh, inner_w, rh, fill=1, stroke=0)

            st_val = rd["status"]
            st_bg  = OK_BG if st_val == "OK" else (OVER_BG if st_val.startswith(">") else ZERO_BG)

            x = M
            vals = [str(rd["cam"])] + [
                f"{rd['palls'][c]:.2f}" if rd['palls'][c] > 0 else "—"
                for c in _T4_CANCHAS
            ] + [
                f"{rd['ae']:.2f}" if rd['ae'] > 0 else "—",
                f"{rd['total']:.2f}",
                st_val,
            ]

            for i, v in enumerate(vals):
                col_name = _T4_CANCHAS[i-1] if 1 <= i <= len(_T4_CANCHAS) else None
                is_focus  = (col_name == cancha_focus)
                is_ae     = (i == len(_T4_CANCHAS) + 1)   # columna AE
                is_status = (i == len(vals) - 1)

                # Color de celda
                if is_focus:
                    cell_bg = YELLOW_FOCUS
                    if rd["recibe_de"]:
                        cell_bg = ASIGN_BG
                    c_pdf.setFillColor(cell_bg)
                    c_pdf.rect(x, y - rh, cw[i], rh, fill=1, stroke=0)
                if is_ae and rd["ae"] > 0:
                    c_pdf.setFillColor(RED_AE)
                    c_pdf.rect(x, y - rh, cw[i], rh, fill=1, stroke=0)
                if is_status:
                    c_pdf.setFillColor(st_bg)
                    c_pdf.rect(x, y - rh, cw[i], rh, fill=1, stroke=0)

                c_pdf.setFillColor(colors.black)
                fs = 7 if i == 0 else 6.5
                c_pdf.setFont(FONT_B if i == 0 else FONT_N, fs)
                c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 3, v)
                x += cw[i]

            # Nota de asignación al margen (si aplica)
            if rd["recibe_de"] or rd["envia_a"]:
                nota_parts = []
                if rd["recibe_de"]:
                    nota_parts.append(f"+{','.join(rd['recibe_de'])}")
                if rd["envia_a"]:
                    nota_parts.append(f"->{','.join(rd['envia_a'])}")
                nota = " | ".join(nota_parts)
                c_pdf.setFont(FONT_N, 5.5)
                c_pdf.setFillColor(colors.HexColor("#1565C0"))
                c_pdf.drawString(M + inner_w + 1, y - rh + 3, nota)
                c_pdf.setFillColor(colors.black)

            c_pdf.setStrokeColor(colors.HexColor("#DDDDDD"))
            c_pdf.line(M, y - rh, M + inner_w, y - rh)
            y -= rh

        # ── Fila TOTAL PALL x CANCHA ───────────────────────────────────────
        y -= 2
        c_pdf.setFillColor(TOT_BG_PDF)
        c_pdf.rect(M, y - rh, inner_w, rh, fill=1, stroke=0)
        c_pdf.setFillColor(WHITE)
        c_pdf.setFont(FONT_B, 6.5)
        x = M
        tot_vals = ["TOTAL PALL"] + [
            f"{totales_pall.get(c, 0):.2f}" for c in _T4_CANCHAS
        ] + ["", f"{sum(totales_pall.values()):.2f}", ""]
        for i, v in enumerate(tot_vals):
            c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 3, v)
            x += cw[i]
        y -= rh + 3

        # ── Sección métricas bultos/HL/KG ─────────────────────────────────
        mh = 10
        met_rows = [
            ("TOTAL BULTOS UP (Pick)", totales_bult, ":.1f"),
            ("TOTAL HL",               totales_hl,   ":.2f"),
            ("TOTAL KG",               totales_kg,   ":.0f"),
        ]
        for label, vals_dict, fmt in met_rows:
            c_pdf.setFillColor(colors.HexColor("#ECEFF1"))
            c_pdf.rect(M, y - mh, inner_w, mh, fill=1, stroke=0)
            c_pdf.setFillColor(colors.black)
            c_pdf.setFont(FONT_B, 6)
            c_pdf.drawString(M + 2, y - mh + 2, label)
            x = M + cw[0]
            for ci, cn in enumerate(_T4_CANCHAS):
                v = vals_dict.get(cn, 0.0)
                c_pdf.setFont(FONT_N, 6)
                c_pdf.drawCentredString(x + cw[ci+1]/2, y - mh + 2, f"{v:{fmt[1:]}}")
                x += cw[ci+1]
            y -= mh + 1

        # ── Leyenda de colores ─────────────────────────────────────────────
        y -= 3
        ley_h = 8
        c_pdf.setFillColor(GREEN_PICK)
        c_pdf.rect(M, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.setFont(FONT_N, 6)
        c_pdf.drawString(M + 10, y - ley_h + 2, "= Bultos a pickear (cancha foco)")

        c_pdf.setFillColor(RED_AE)
        c_pdf.rect(M + 100, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.drawString(M + 112, y - ley_h + 2, "= Paletas puras (AE)")

        c_pdf.setFillColor(ASIGN_BG)
        c_pdf.rect(M + 185, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.drawString(M + 197, y - ley_h + 2, "= Carga asignada recibida")
        y -= ley_h + 3

        # ── Pie: personas y hora fin ───────────────────────────────────────
        pie_rows = [
            ("# PERSONAS",    {c: str(personas.get(c, 1)) for c in _T4_CANCHAS}),
            ("HORA FIN EST.", {
                c: fin_calc["por_cancha"][c]["fin_dt"].strftime("%H:%M")
                   if fin_calc["por_cancha"][c]["bultos"] > 0 else "—"
                for c in _T4_CANCHAS
            }),
        ]
        for label, vals_dict in pie_rows:
            c_pdf.setFillColor(colors.HexColor("#1a3a6b"))
            c_pdf.rect(M, y - mh, inner_w, mh, fill=1, stroke=0)
            c_pdf.setFillColor(WHITE)
            c_pdf.setFont(FONT_B, 6)
            c_pdf.drawString(M + 2, y - mh + 2, label)
            x = M + cw[0]
            for ci, cn in enumerate(_T4_CANCHAS):
                is_this_focus = (cn == cancha_focus)
                if is_this_focus and label == "HORA FIN EST.":
                    c_pdf.setFillColor(colors.HexColor("#FFF176"))
                    c_pdf.rect(x, y - mh, cw[ci+1], mh, fill=1, stroke=0)
                    c_pdf.setFillColor(colors.HexColor("#1a3a6b"))
                else:
                    c_pdf.setFillColor(WHITE)
                c_pdf.setFont(FONT_B if is_this_focus else FONT_N, 6)
                c_pdf.drawCentredString(x + cw[ci+1]/2, y - mh + 2, vals_dict.get(cn, "—"))
                x += cw[ci+1]
            y -= mh + 1

        c_pdf.showPage()

    for cancha in _T4_CANCHAS_PDF:
        _draw_page(cancha)

    c_pdf.save()
    buf.seek(0)
    return buf.read()


def render_tab_proyeccion():
    import datetime as _dt

    st.subheader("📊 Proyección Picking ×4")
    st.caption(
        "Fuente: **CAR.xlsx + Frescura 3.0** — Calcula PICK vs AE por "
        "**UNIDAD PKT** (col R DDM) y cancha, sin necesitar el MASTER."
    )

    # ── Reutilizar uploads de Tab 1 si están disponibles ─────────────────────
    car_from_t1 = st.session_state.get("t4_car") or st.session_state.get("t1_car")
    fr_from_t1  = st.session_state.get("t4_fr")  or st.session_state.get("t1_fr")

    col_a, col_b = st.columns(2)
    with col_a:
        car_file = _file_uploader("CAR.xlsx (export Chess)", ["xlsx"], key="t4_car")
    with col_b:
        fr_file = _file_uploader("Frescura 3.0.xlsx",        ["xlsx"], key="t4_fr")

    car_use = car_file or car_from_t1
    fr_use  = fr_file  or fr_from_t1

    if not (car_use and fr_use):
        st.info("Subí el **CAR.xlsx** y la **Frescura 3.0** para activar la proyección.")
        return

    # ── Cargar datos ──────────────────────────────────────────────────────────
    pdata = None
    try:
        with st.spinner("Calculando proyección de picking desde CAR…"):
            pdata = _t4_load_car_proyeccion(
                car_bytes=car_use.getvalue(),
                fr_bytes=fr_use.getvalue(),
            )
        st.success(
            f"✓ Proyección lista — {len(pdata['df'])} camiones | "
            f"Fecha: {pdata['fecha']} | Fuente: {pdata['fuente']}"
        )
        log_event("info", f"Tab4 proyección v4.6: {len(pdata['df'])} camiones, {pdata['fecha']}")
    except Exception as e:
        st.error(f"❌ Error calculando proyección: {e}")
        with st.expander("Stack trace"):
            import traceback
            st.code(traceback.format_exc())
        return

    if pdata is None or pdata["df"].empty:
        st.warning("No hay datos de camiones disponibles.")
        return

    df_cam = pdata["df"].copy()

    # ── Hora de inicio única (compartida) ─────────────────────────────────────
    st.divider()
    col_hi1, col_hi2, col_hi3 = st.columns([1, 1, 4])
    with col_hi1:
        h_inicio = st.number_input("Hora inicio", 0, 23, value=17, step=1, key="t4_hora_h")
    with col_hi2:
        m_inicio = st.number_input("Min inicio", 0, 59, value=30, step=5, key="t4_hora_m")
    hora_inicio_global = _dt.time(int(h_inicio), int(m_inicio))

    # ── Productividad y personas por cancha ───────────────────────────────────
    with st.expander("⚙️ Productividad y personas por cancha", expanded=False):
        prod_cols = st.columns(5)
        vel_custom  = {}
        pers_custom = {}
        for i, cn in enumerate(_T4_CANCHAS):
            with prod_cols[i]:
                st.markdown(f"**{cn.replace('CANCHA ', 'C')}**")
                vel_custom[cn] = st.number_input(
                    f"Vel {cn} (bult/h)", 10, 2000, value=_T4_VEL_DEFAULT[cn], step=10,
                    key=f"t4_vel_{cn}",
                )
                pers_custom[cn] = st.number_input(
                    f"Pers {cn}", 1, 10, value=1, step=1,
                    key=f"t4_pers_{cn}",
                )

    # Hora inicio por cancha = hora_inicio_global para todas
    inicio_custom = {cn: hora_inicio_global for cn in _T4_CANCHAS}

    # ── Filtrar: siempre sin ceros ────────────────────────────────────────────
    df_display = df_cam[df_cam["TOTAL_PICK"] > 0].reset_index(drop=True)

    # ── ASIGN state ───────────────────────────────────────────────────────────
    if "t4_asign" not in st.session_state:
        st.session_state["t4_asign"] = {}

    # ── Calcular totales con ASIGN ────────────────────────────────────────────
    totales_calc   = _t4_calcular_pall_por_cancha(df_display, st.session_state["t4_asign"])
    totales_pall_c = totales_calc["total"]
    status_rows    = totales_calc["status_rows"]

    # ── Tabla principal ───────────────────────────────────────────────────────
    st.divider()
    st.subheader("📦 Pallets UP por camión")
    st.caption(
        "**UP** = fracción de paleta por bulto (UNIDAD PKT, col R Frescura DDM). "
        "Paletas enteras → AE. Fracción restante → picking."
    )

    # Número de camión = posición ordinal en la lista ordenada
    cam_list = sorted(df_display["Camión"].tolist())
    cam_to_num = {cam: i+1 for i, cam in enumerate(cam_list)}

    edit_rows = []
    for _, row in df_display.iterrows():
        cam = int(row["Camión"])
        sr  = status_rows.get(cam, {})
        num_cam = cam_to_num.get(cam, "—")
        rec = {"#": num_cam, "Camión": cam}
        for cn in _T4_CANCHAS:
            up_val = float(row.get(f"_up_{cn}", 0.0))
            rec[f"UP {cn.replace('CANCHA ', 'C')}"] = round(up_val, 3)
        rec["Bult Pick"] = round(float(row.get("TOTAL_PICK", 0.0)), 1)
        rec["AE Pall"]   = round(float(row.get("AE_PALL", 0.0)), 2)
        rec["TOT PALL"]  = round(float(sr.get("total_pall", 0.0)), 2)
        rec["STATUS"]    = sr.get("status", "—")
        edit_rows.append(rec)

    df_editor = pd.DataFrame(edit_rows)

    col_cfg = {
        "#":       st.column_config.NumberColumn("#", disabled=True, width="small"),
        "Camión":  st.column_config.NumberColumn("Camión", disabled=True),
        "Bult Pick": st.column_config.NumberColumn("Bult Pick", disabled=True, format="%.1f"),
        "AE Pall":   st.column_config.NumberColumn("AE Pall", disabled=True, format="%.2f"),
        "TOT PALL":  st.column_config.NumberColumn("TOT PALL", disabled=True, format="%.2f"),
        "STATUS":    st.column_config.TextColumn("STATUS", disabled=True),
    }
    for cn in _T4_CANCHAS:
        short = cn.replace("CANCHA ", "C")
        col_cfg[f"UP {short}"] = st.column_config.NumberColumn(f"UP {short}", disabled=True, format="%.3f")

    st.dataframe(df_editor, column_config=col_cfg, use_container_width=True, hide_index=True)

    # ── Filas SUB / DESIGNADOS / TOTAL ───────────────────────────────────────
    resumen_pall = pd.DataFrame({
        "Concepto": ["SUB Total PALL", "DESIGNADOS PALL", "TOTAL PALL x CANCHA"],
        **{cn.replace("CANCHA ", "C"): [
            round(totales_calc["sub"][cn], 2),
            round(totales_calc["designados"][cn], 2),
            round(totales_calc["total"][cn], 2),
        ] for cn in _T4_CANCHAS}
    })
    st.dataframe(
        resumen_pall.style.apply(
            lambda row: ["background-color: #FF8C00; color: white" if row.name == 2 else "" for _ in row],
            axis=1,
        ),
        use_container_width=True, hide_index=True,
    )

    # ── Asignación de carga ───────────────────────────────────────────────────
    st.divider()
    st.subheader("🔀 Asignar carga entre canchas")
    st.caption(
        "Seleccioná un camión, una cancha origen y la cancha destino. "
        "El tiempo estimado de fin se ajusta automáticamente."
    )

    asign_cols = st.columns([2, 2, 2, 1])
    cam_opts = [str(c) for c in cam_list]
    with asign_cols[0]:
        cam_sel = st.selectbox("Camión", ["—"] + cam_opts, key="t4_asign_cam")
    with asign_cols[1]:
        origen_sel = st.selectbox("Cancha origen", ["—"] + _T4_CANCHAS, key="t4_asign_origen")
    with asign_cols[2]:
        destino_opts = [c for c in _T4_CANCHAS if c != origen_sel]
        destino_sel  = st.selectbox("Cancha destino", ["—"] + destino_opts, key="t4_asign_destino")
    with asign_cols[3]:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("✅ Asignar", key="t4_asign_btn"):
            if cam_sel != "—" and origen_sel != "—" and destino_sel != "—":
                key_asign = (int(cam_sel), origen_sel)
                st.session_state["t4_asign"][key_asign] = destino_sel
                st.success(f"✓ Cam {cam_sel}: {origen_sel} → {destino_sel}")
                st.rerun()

    # Mostrar asignaciones activas y permitir eliminarlas
    if st.session_state["t4_asign"]:
        st.markdown("**Asignaciones activas:**")
        asign_items = [(k, v) for k, v in st.session_state["t4_asign"].items() if v]
        for (cam_a, can_a), dest_a in asign_items:
            col_x, col_txt = st.columns([1, 6])
            with col_x:
                if st.button("❌", key=f"del_asign_{cam_a}_{can_a}"):
                    del st.session_state["t4_asign"][(cam_a, can_a)]
                    st.rerun()
            with col_txt:
                up_val = 0.0
                row_m = df_display[df_display["Camión"] == cam_a]
                if not row_m.empty:
                    up_val = float(row_m.iloc[0].get(f"_up_{can_a}", 0.0))
                st.markdown(
                    f"Cam **{cam_a}** | {can_a} → **{dest_a}** "
                    f"({up_val:.3f} UP)"
                )

    # Recalcular con ASIGN actualizado
    totales_calc   = _t4_calcular_pall_por_cancha(df_display, st.session_state["t4_asign"])
    totales_pall_c = totales_calc["total"]

    # ── Totales bultos por cancha (base, sin asignaciones) ───────────────────
    totales_bult_base = {cn: float(df_display[cn].sum()) if cn in df_display.columns else 0.0 for cn in _T4_CANCHAS}
    # Totales bultos por cancha CON asignaciones aplicadas (para tiempo de fin)
    totales_bult = _t4_calcular_bultos_por_cancha(df_display, st.session_state["t4_asign"])
    totales_hl   = {cn: 0.0 for cn in _T4_CANCHAS}
    totales_kg   = {cn: 0.0 for cn in _T4_CANCHAS}

    with st.expander("📊 Totales por cancha", expanded=True):
        cols_met = st.columns(len(_T4_CANCHAS) + 1)
        for i, cn in enumerate(_T4_CANCHAS):
            short = cn.replace("CANCHA ", "C")
            delta_bult = totales_bult[cn] - totales_bult_base[cn]
            delta_str  = f"{delta_bult:+.0f} bult (asign)" if delta_bult != 0 else None
            cols_met[i].metric(f"Bult {short}", f"{totales_bult[cn]:.1f}", delta=delta_str)
        cols_met[-1].metric("TOT PICK", f"{pdata['tot_pick']:.0f}")

        cols_met2 = st.columns(len(_T4_CANCHAS) + 1)
        for i, cn in enumerate(_T4_CANCHAS):
            short = cn.replace("CANCHA ", "C")
            cols_met2[i].metric(f"PALL {short}", f"{totales_calc['total'].get(cn, 0):.2f}")
        cols_met2[-1].metric("TOT AE bult", f"{pdata['tot_ae']:.0f}")

    # ── Hora estimada de fin por cancha (sobre bultos con asignaciones) ───────
    fin_por_cancha = {}
    for cn in _T4_CANCHAS:
        bult  = totales_bult[cn]   # ya incluye reasignaciones
        fin_dt = _t4_hora_fin(bult, vel_custom[cn], pers_custom[cn], hora_inicio_global)
        fin_por_cancha[cn] = {"bultos": bult, "fin_dt": fin_dt}
    fin_global_dt = max(v["fin_dt"] for v in fin_por_cancha.values())
    fin_calc_dict = {"por_cancha": fin_por_cancha, "fin_global_dt": fin_global_dt}

    # ── Cabecera PICKING ──────────────────────────────────────────────────────
    st.divider()
    hdr_cols = st.columns([2, 1, 1, 1, 2])
    hdr_cols[0].markdown(f"**🕐 PICKING** — {pdata['fecha']}")
    hdr_cols[1].metric("Inicio", hora_inicio_global.strftime("%H:%M"))
    hdr_cols[2].metric("Fin est.", fin_global_dt.strftime("%H:%M"))
    hdr_cols[3].metric("Mix picking", f"{pdata['mix_picking']*100:.1f}%")
    hdr_cols[4].caption(f"Fuente: {pdata['fuente']}")

    fin_cols = st.columns(len(_T4_CANCHAS) + 1)
    for i, cn in enumerate(_T4_CANCHAS):
        short = cn.replace("CANCHA ", "C")
        fp = fin_por_cancha[cn]
        fin_s = fp["fin_dt"].strftime("%H:%M") if fp["bultos"] > 0 else "—"
        fin_cols[i].metric(
            f"FIN {short}",
            fin_s,
            f"{fp['bultos']:.1f} bult | {totales_pall_c.get(cn, 0):.2f} pall",
            delta_color="normal" if fp["bultos"] > 0 else "off",
        )
    fin_cols[-1].metric("🏁 FIN GLOBAL", fin_global_dt.strftime("%H:%M"), delta_color="off")

    # ── Generar PDF ×4 ────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📄 Generar PDF ×4 (una copia por cancha)")
    cp1, cp2 = st.columns([1, 2])
    with cp1:
        if st.button("📄 Generar PDF ×4", type="primary",
                     use_container_width=True, key="t4_pdf_btn"):
            with st.spinner("Generando PDF…"):
                try:
                    pdf_bytes = _t4_generar_pdf_x4(
                        df=df_display,
                        totales_bult=totales_bult,
                        totales_pall=totales_pall_c,
                        totales_hl=totales_hl,
                        totales_kg=totales_kg,
                        productividad=vel_custom,
                        personas=pers_custom,
                        inicio=hora_inicio_global,
                        fecha=pdata["fecha"],
                        mix_picking=pdata["mix_picking"],
                        fin_calc=fin_calc_dict,
                        asign_detail=totales_calc.get("asign_detail", {}),
                    )
                    fname = f"proyeccion_picking_{pdata['fecha']}.pdf"
                    st.download_button(
                        "⬇ Descargar PDF (4 páginas)",
                        data=pdf_bytes, file_name=fname,
                        mime="application/pdf", use_container_width=True,
                        key="t4_pdf_dl",
                    )
                    st.success(f"✓ PDF generado: {fname}")
                    log_event("info", f"PDF Proyección v4.6 generado: {fname}")
                except Exception as e:
                    st.error(f"❌ Error generando PDF: {e}")
                    with st.expander("Stack trace"):
                        import traceback
                        st.code(traceback.format_exc())
    with cp2:
        st.info(
            f"📋 **Resumen — {pdata['fecha']}**\n\n"
            f"- Fin estimado global: {fin_global_dt.strftime('%H:%M')}\n"
            f"- Hora inicio: {hora_inicio_global.strftime('%H:%M')}\n"
            f"- Camiones con picking: {len(df_display[df_display['TOTAL_PICK'] > 0])}\n"
            f"- Bultos picking total: {pdata['tot_pick']:.0f}\n"
            f"- Mix picking: {pdata['mix_picking']*100:.1f}%\n"
            f"- Fuente: {pdata['fuente']}"
        )

# ── TAB 5 — Extraíbles Sheets ──────────────────────────────────────────────

def render_tab_extraibles():
    st.subheader("📤 Extraíbles para Google Sheets")
    st.caption(
        "Bajá los bloques del MASTER en formato TSV/CSV/XLSX para pegar manual "
        "en el Ecosistema AD 3.0. Sin push automático en v4.0 (decisión #2)."
    )

    master = _file_uploader(
        "MASTER del día.xlsx (78 MB)", ["xlsx"], key="t5_master"
    )

    if not master:
        st.info("Subí el MASTER para acceder a los 4 bloques.")
        return

    file_bytes = master.getvalue()

    try:
        df_pall, fecha_a1 = load_master_matriz_pall(file_bytes)
        df_rep = load_master_reposicion_ae(file_bytes)
        df_pick = load_master_matriz_picking_app(file_bytes)
        df_agr = load_master_agregados_ae(file_bytes)
        log_event("info", f"MASTER cargado para extraíbles. Fecha A1: {fecha_a1}")
    except Exception as e:
        st.error(f"❌ Error leyendo MASTER: {e}")
        return

    fecha_str = fecha_a1.strftime("%d/%m/%Y") if fecha_a1 else "—"
    st.success(f"✓ MASTER cargado — Fecha en A1: **{fecha_str}**")

    st.divider()

    # ── Bloque 1: Matriz Pall. ──────────────────────────────────────────
    with st.expander("1️⃣  Matriz Pall. (29 camiones × 24 cols)", expanded=True):
        st.dataframe(df_pall, use_container_width=True, height=300)
        _download_trio(df_pall, "Matriz_Pall", "Matriz Pall.", "t5_pall")

    # ── Bloque 2: Agregados AE ──────────────────────────────────────────
    with st.expander("2️⃣  Agregados AE (subset Matriz Pall. — cols AE)"):
        st.caption(
            "⚠ Provisional: deriva de Matriz Pall. cols K-Q. "
            "Cuando confirmes hoja destino exacta en Ecosistema AD 3.0, se ajusta."
        )
        st.dataframe(df_agr, use_container_width=True, height=300)
        _download_trio(df_agr, "Agregados_AE", "Agregados AE", "t5_agr")

    # ── Bloque 3: Reposición AE (filtrado J<0) ──────────────────────────
    with st.expander("3️⃣  Reposición AE (solo Reposicion_Pall < 0)", expanded=True):
        df_rep_neg = filter_reposicion_negativa(df_rep)
        st.metric("Filas con reposición negativa", f"{len(df_rep_neg)} / {len(df_rep)}")
        if len(df_rep_neg) == 0:
            st.warning("No hay filas con Reposicion_Pall < 0. Nada para pegar hoy.")
        else:
            st.dataframe(df_rep_neg, use_container_width=True, height=300)
            _download_trio(df_rep_neg, "Reposicion_AE_neg", "Reposición AE", "t5_rep")

    # ── Bloque 4: (fx) Picking ──────────────────────────────────────────
    with st.expander("4️⃣  (fx) Picking — fila resumen Matriz Picking APP"):
        st.caption(
            "Vista plana de la hoja Matriz Picking APP. Para Sprint 1 se exporta "
            "completa; en Sprint 2 se filtra/recorta según destino exacto."
        )
        st.dataframe(df_pick.head(50), use_container_width=True, height=300)
        _download_trio(df_pick, "fx_Picking", "(fx) Picking", "t5_pick")


# ── TAB 6 — Validación + Log (PLACEHOLDER, Sprint 2) ───────────────────────

def render_tab_validacion():
    st.subheader("✅ Validación + Reglas + Log")
    st.caption(
        "Pre-checks + reglas de negocio + validación cruzada CAR↔MASTER. "
        "**Sprint 2** — pendiente."
    )

    st.info(
        "🔧 En desarrollo (Sprint 2).\n\n"
        "**Pre-checks previstos:**\n"
        "- CAR.xlsx modificado hoy\n"
        "- ANR (rechazos1.xlsx) modificado hoy (warning)\n"
        "- MASTER → Matriz Pall.!A1 == hoy\n"
        "- Suma TOTAL PALL > 0\n\n"
        "**Reglas:**\n"
        "- Camión TOTAL=0 → ocultar y loguear\n"
        "- Camión Reparto=NO → tabla checkbox, default OFF (decisión #1)\n"
        "- Cross-validate CAR↔MASTER (WARNING, no bloqueante — decisión #6)\n"
        "- Sanity checks bloqueantes pre-publicación"
    )

    st.divider()
    st.markdown("### 📜 Log de la corrida actual")
    log_lines = st.session_state.get("log_buffer", [])
    if not log_lines:
        st.caption("(sin eventos aún)")
    else:
        st.code("\n".join(log_lines[-200:]), language="text")
        st.download_button(
            "⬇ Descargar log completo",
            data="\n".join(log_lines).encode("utf-8"),
            file_name=_stamp("run_log", "txt"),
            mime="text/plain",
        )


# ── MAIN ────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title=f"Picking Orchestrator v{APP_VERSION}",
        page_icon="📦",
        layout="wide",
    )

    # Banner header
    st.markdown(
        f"""
        <div style="
            background: linear-gradient(135deg, #1a3a6b 0%, #2e5fa3 100%);
            padding: 14px 22px;
            border-radius: 8px;
            margin-bottom: 18px;
            color: white;
        ">
            <div style="font-size: 22px; font-weight: 600;">
                📦 Picking Orchestrator
            </div>
            <div style="font-size: 13px; opacity: 0.85; margin-top: 2px;">
                Beccacece Hnos SA — Distribuidor CMQ &nbsp;·&nbsp; DPO 2.1 — Pilar Almacén
                &nbsp;·&nbsp; v{APP_VERSION}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_sidebar()

    tabs = st.tabs([
        "📦 Planilla Carga",
        "📋 Resumen Camiones",
        "🚛 Camiones T2",
        "📊 Proyección Picking ×4",
        "📤 Extraíbles Sheets",
        "✅ Validación + Log",
    ])
    with tabs[0]: render_tab_planilla()
    with tabs[1]: render_tab_resumen()
    with tabs[2]: render_tab_t2()
    with tabs[3]: render_tab_proyeccion()
    with tabs[4]: render_tab_extraibles()
    with tabs[5]: render_tab_validacion()


if __name__ == "__main__":
    main()
