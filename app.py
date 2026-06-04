"""
Picking Orchestrator v4.70.0 — Beccacece Hnos SA

CAMBIOS v4.70.0 — 9 CORRECCIONES OPERATIVAS:

  1. TAB ARCHIVOS — alineacion simetrica de uploaders (6 columnas igual ancho,
     estado como badge coloreado debajo de cada uploader).

  2. PROYECCION PICKING — PDF x5: renombrado boton/seccion. Pagina resumen
     controladores separada del PDF x5 (boton propio "PDF Controladores").
     Numerales arabigos en canchas (C1/C2/C3/C4/MKPL+Merch). Resumen enriquecido:
     Top 5 SKUs por cancha, analisis de carga, recomendaciones SOPs.
     PDF x5 omite la hoja de resumen controladores.

  3. TABLERO RUTEADOR — fecha default = ultima fecha del ANR col S.

  4. TABLERO RUTEADOR — boton "Calcular tablero" explicito: los campos de carga
     del dia no disparan calculo automatico. El tablero calcula solo al presionar.

  5. TABLERO RUTEADOR — Licencias conductores: tabla editable persistente (como
     Productividad por cancha). Sin uploader. Datos no se borran con F5.

  6. FIX CRITICO UP/PALETAS — Tablero ruteador alineado con Proyeccion Picking:
     UP = bultos_eq / BXP (paletas reales, NO normalizado base 50).
     Paletas = floor(UP) por SKU sumadas (paletas AE completas).
     Picking UP = fraccion restante. TOT PALL = AE_pall + picking_up.

  7. TABLERO — Cap.(kg) reemplazada por semaforo de peso:
     verde si peso < capacidad, rojo si excede. Solo icono en tabla.

  8. TABLERO — Venc. Lic. y Lic. siempre completos (tomados de la tabla editable).
     Aplica tambien a PDF y Excel exportados.

  9. EXCEL/PDF TABLERO — formato como imagen de referencia: grafico dual
     (barras UP + linea PDV), tabla proporcional A4 landscape, sin Cap.(kg),
     con semaforo peso, licencias siempre visibles.

"""
# ── Picking Orchestrator v4.70.0 — Beccacece Hnos SA ─────────────────────────
_CHANGELOG = """

  1. FIX CRÍTICO UP — fuente única CAR (no más ANR para Bultos UP):
     - Los Bultos UP, HL y Peso se calculan EXCLUSIVAMENTE desde el CAR.
     - Lógica por fila: bultos_eq = Bultos + Unids / un_x_bulto (DDM col N).
       Si un_x_bulto = 0 o Unids = 0 → bultos_eq = Bultos.
     - HL = DDM[sku].hl × bultos_eq; Peso = DDM[sku].kg × bultos_eq.
     - El CAR contiene TODAS las líneas: VE (bloque azul hasta separador vacío)
       y CHESS (resto del archivo). Ambos bloques se procesan juntos.
     - PDV (pedidos) y Rechazos siguen leyéndose del ANR (CAR no tiene esa info).
     - El bloque VE ya no se procesa por separado — está incluido en el total CAR.
     - df_ddm_dict se extiende con 'un_bulto' (col N de DDM).

  2. TABLERO MENSUAL — subtab "📅 Histórico Mensual":
     - Upload persistente de Excel diarios generados por el Tablero del día.
     - Los archivos se acumulan: no se borran al subir nuevos.
     - Tabla resumen con todos los días cargados (fecha, camiones, PDV, UP, HL, kg,
       rechazos, drop size, paletas promedio, productividad ruteo).
     - Gráficos: UP total diario, PDV total diario, Drop Size diario.
     - Botones "Excel Mensual" y "PDF Mensual" para exportar el consolidado.
     - Estructura de datos en session_state["tr_mensual_data"] (lista de dicts).

CAMBIOS v4.67.0 — TABLERO RUTEADOR: FIX DDM + EXCEL/PDF COMPLETOS:

  1. FIX CRÍTICO df_ddm en Tablero Ruteador:
     - df_ddm NUNCA se guardaba en session_state → HL y Peso siempre = 0.
     - Al subir Frescura en Tab Archivos, ahora se parsea el DDM directamente
       (cols: ARTÍCULO=C, BULTOS X PALLET=G, VALOR HL=P, PESO KG=Q)
       y se guarda en ss["df_ddm"] con estructura {sku: {hl_unit, kg_unit, bxp}}.
     - Tablero Ruteador usa este dict directamente en lugar del DataFrame crudo.

  2. FIX Excel Tablero Ruteador — A4 landscape completo:
     - Fuente reducida a 8-9pt, márgenes comprimidos, todo en una hoja.
     - Header compacto con título + KPIs en banda coloreada.
     - Tabla de camiones con columnas ajustadas (ancho fijo proporcional A4).
     - Gráficos Bultos UP y HL por camión embebidos en la misma hoja.
     - Fila TOTAL en azul claro, alertas de licencias en rojo.

  3. FIX PDF Tablero Ruteador — layout A4 landscape limpio:
     - Columnas rebalanceadas para que todo entre sin decimación.
     - Gráficos adicionales: HL por camión y Peso(kg) por camión.
     - Bloque de análisis operativo: % carga por camión, varianza, top/bottom.

  4. PESO siempre visible en tabla Streamlit, Excel y PDF.

  5. FIX (fx) Picking fila destino:
     - Si fila 216 es correcta en la hoja (fx) Picking pero parece vacía,
       verificar que no haya desfase de filas vacías al inicio.
     - Ahora lee todas las filas no vacías de col A para buscar la fecha.



CAMBIOS v4.66.0 — FIX PASO 2 + EXCEL DESCARGABLE:

  1. FIX MATCH DE FECHA EN (fx) Picking:
     - Root cause: gspread devuelve las fechas de col A como string "d-mmm"
       (ej: "3-jun", "2-jun") porque el sheet usa fórmulas =Picking!Axxx
       que generan fechas serializadas de Sheets. Los formatos anteriores
       (dd/mm/yyyy, yyyy-mm-dd, etc.) nunca hacían match.
     - Fix: nueva función _normalizar_fecha_gspread() que convierte tanto
       fechas serializadas numéricas (serial Lotus) como strings "d-mmm"
       o "dd-mmm" al formato comparable con la fecha del día.
     - La búsqueda ahora genera el string "d-mmm" en español (ej: "3-jun")
       desde la fecha del día y lo compara directamente con col A.

  2. EXCEL DESCARGABLE SIEMPRE PRESENTE en Paso 2:
     - Nueva función _build_fx_picking_xlsx() que genera un .xlsx de una
       sola fila con: col A = fecha formateada, cols B:BL = los 63 valores
       calculados, con headers reales leídos desde el sheet (fila 1).
     - El Excel se ofrece para descarga SIEMPRE (independiente del envío
       a Sheets), con nombre fx_Picking_DD-MM-YYYY.xlsx.
     - Fallback: si no hay credenciales para leer headers, usa etiquetas
       genéricas (A=Fecha de Picking, B=PICKING, I=AE BULT, etc.).


Picking Orchestrator v4.64.0 — Beccacece Hnos SA

CAMBIOS v4.64.0 — 4 MEJORAS OPERATIVAS:

  1. NÚMEROS NATURALES en etiquetas de cancha (CI→C1, CII→C2, CIII→C3, CIV→C4):
     - Nueva función _cn_short() + dict _ROMAN_TO_NAT (display-only, keys internas sin cambio).
     - Aplicado en: PDF resumen controlador, col_cfg del data_editor (labels),
       metrics "Bult/PALL/FIN", semáforo alertas, caption reasignaciones,
       XLSX proyección picking, tablas HTML concepto.

  2. RECOMENDACIONES PARA CONTROLADOR en PDF resumen (basado en ALM-00013 v5):
     - Nueva sección visual de 10 items en dos columnas, fondo alternado.
     - Cubre: doble control, EPP, posición lateral al AE, filmación obligatoria,
       gestión de faltantes/sobrantes/dañados, condiciones de carga (patente/pallets/tara),
       registro en Errores Operativos, rearmado de paleta, checklist por camión.
     - Solo se dibuja si hay espacio suficiente sobre el footer.

  3. EXCEL CLASIFICACIÓN — botón de bajada junto al PDF:
     - Nueva función build_clasificacion_anr_xlsx(): misma visual que el PDF
       (header azul, tabla con header amarillo, filas alternadas, fila TOTAL azul,
       bloque Productividad Estimada en amarillo claro).
     - 3 columnas de download: PDF | Excel | TSV.

  4. FECHA AUTOMÁTICA EN TABLERO RUTEADOR desde el ANR cargado:
     - Lee la última fecha disponible en col FECHA/S del ANR antes de renderizar
       el date_input. Fallback a hoy() si el ANR no está cargado o falla la lectura.
     - Elimina el desfase cuando el ANR es del día anterior al actual.


Picking Orchestrator v4.63.0 — Beccacece Hnos SA

CAMBIOS v4.63.0 — FIX ALINEACIÓN TABLAS CONCEPTO (Proyección Picking):
  - 🐛 ROOT CAUSE: Las tablas "SUB Total PALL / DESIGNADOS / TOTAL PALL x CANCHA"
    y "TOTAL BULTOS / HL / KG / HORA EST. FIN" se renderizaban como st.dataframe
    separados. Streamlit distribuye el ancho de columnas de forma independiente
    en cada tabla, haciendo imposible la alineación pixel-perfecta con la tabla
    de reasignación principal (data_editor con 14 columnas).
  - ✅ FIX: Ambas tablas reemplazadas por HTML puro (st.markdown unsafe_allow_html).
    La columna Concepto ocupa 14% del ancho del contenedor (delgada).
    Cada cancha ocupa (86% / N_canchas) de forma proporcional.
    Esto replica exactamente la distribución visual del data_editor, donde las
    columnas CI/CII/CIII/CIV/MKPL aparecen en la misma posición horizontal.
  - ✅ Las dos tablas HTML se generan en un solo bloque envuelto en div oscuro,
    manteniendo el look & feel del tema dark de Streamlit.
  - ✅ TOTAL PALL x CANCHA mantiene el fondo naranja (#FF8C00) y texto blanco.
  - ✅ Sin cambios en lógica de cálculo — solo presentación.


Picking Orchestrator v4.62.0 — Beccacece Hnos SA

CAMBIOS v4.62.0 — FIX UI UPLOADERS TAB ARCHIVOS:
  - 🐛 FIX: El uploader de Frescura (y demás) mostraba "200MB per file • XLSX"
    y un área de drop más grande que el de CAR cuando había un archivo cargado.
  - ✅ CSS inyectado en render_tab_archivos() que oculta el texto "per file"
    y compacta el padding del dropzone, igualando visualmente todos los
    uploaders del tab independientemente de si tienen archivo cargado o no.


Picking Orchestrator v4.61.0 — Beccacece Hnos SA

CAMBIOS v4.56.0 — FIX HORARIOS CONGELADOS EN PROYECCIÓN PICKING:
  - 🐛 ROOT CAUSE: _rebuild_live_asign_from_editor_state() leía
    st.session_state["t4_main_editor"]["edited_rows"] que son DELTAS
    acumulativos (no estado completo). Al quitar una ASIGN → "", el delta
    decía "" pero t4_asign persistido seguía teniendo la key antigua.
    Resultado: horarios PRE-tabla congelados al reasignar.
  - ✅ FIX 1 — Eliminado _rebuild_live_asign_from_editor_state().
    El header PRE-tabla lee t4_asign directamente (fuente única y confiable).
  - ✅ FIX 2 — live_asign construido desde edited_df COMPLETO (todas las filas).
    Si ASIGN está vacío → key ausente (borrado limpio). Garantiza sincronía
    total entre lo que se ve en el editor y el estado persistido.
  - ✅ FIX 3 — Botón "🔄 Recalcular horarios" con st.rerun() forzado.
    Flujo correcto: editar ASIGN → Recalcular → rerun con t4_asign limpio
    → PRE-tabla muestra horarios actualizados sin lag ni congelamiento.
  - 📋 Caption dinámico: muestra reasignaciones activas al instante.


Picking Orchestrator v4.55.1 — Beccacece Hnos SA

CAMBIOS v4.54.0:
  - 🔧 HL y Peso — estrategia híbrida completa:
    El ANR de Chess tiene PESO KG y HL por fila pero algunos SKUs
    (merch, activos, etc.) los tienen en 0 aunque tienen bultos.
    NUEVA LÓGICA fila a fila:
      · PESO KG > 0 en ANR → usar valor real del ANR (Chess)
      · PESO KG = 0 en ANR → DDM[sku] × bultos como complemento
      · Ídem para HL
    Esto cubre tanto ANRs con valores reales como ANRs sin esas cols.
  - ✅ VE del CAR correctamente sumadas en HL y Peso:
    Las Ventas Especiales del CAR (hasta fila 40) se agregan siempre
    vía DDM × bultos sobre la suma del ANR, tanto en UP como en HL y
    Peso (kg). Antes solo sumaban UP.
  - ℹ️  Caption dinámico que indica exactamente qué fuentes se usaron
    para HL/Peso (ANR+DDM, solo ANR, solo DDM, o sin datos).

CAMBIOS v4.49.1:
  - 📂 Archivos: nuevo uploader "ANR -1.xlsx" (ANR del día anterior).
    Se guarda en session_state["cierre_anr_m1"]. Solo se usa en Cierre D+1.
    No afecta ninguna otra pestaña.
  - 💳 Cierre D+1: CTA CTE ahora se recalcula desde ANR -1 en lugar de
    reutilizar el valor del cierre del día. Si ANR -1 no está cargado,
    cae back al valor del cierre original (comportamiento anterior).
  - 📋 Estado de archivos: tabla actualizada con la nueva fila ANR -1.
  - ✅ Cierre del día: sin cambios — sigue usando ANR normal.

CAMBIOS v4.49.0:
  - 🗑️  Bloque PRE-tabla eliminado (Proyección Picking): se removió el bloque
    de horarios/semáforo duplicado que aparecía ANTES del data_editor. Solo
    queda el bloque POST-tabla (live_asign), que es el correcto y no sufre lag.
  - 📋 Tabs reordenados: "Tablero Ruteador" se mueve después de "Top SKUs"
    (nuevo orden: Archivos → Planilla → Resumen → Proyección → Clasificación →
    Camiones T2 → Top SKUs → Tablero Ruteador → Boletas → Cierre → Validación).
  - 💳 Cierre — detalle CTA CTE por chofer: se agrega sección expandible
    "Detalle CTA CTE por camión" que muestra, por cada chofer con CTA CTE > 0,
    los clientes con nombre + monto. Al sumar todos los clientes de todos los
    camiones, el resultado coincide exactamente con el total E).
  - 💳 Cierre PDF — sección CTA CTE detail: nueva página/bloque en el PDF
    con tabla de clientes CTA CTE por camión (nombre + monto).
  - 💳 Cierre Excel — hoja "CTA CTE Detalle": nueva hoja con el desglose
    cliente × camión, con totales por camión y total general.
  - 🔧 FIX desfasaje CTA CTE: `_build_cierre_df` ahora pasa `df_anr` completo
    a las funciones de export (PDF/Excel) para construir el detalle client-level.
    El cálculo del monto total no cambia — solo se expone el desglose.


CAMBIOS v4.46.0 — Tablero Ruteador:

CAMBIOS v4.46.0 — Tablero Ruteador:
  - 🔵 VE automático: Ventas Especiales se leen del CAR (Hoja1 filas 2-40),
    se elimina el editor manual del tablero. Se muestra en expander compacto.
  - 🧑 Chofer desde DESCRIPCIÓN TRANSPORTE del ANR (nombre real siempre disponible).
    Se capitaliza el nombre y se ignoran valores "SIN CHOFER ASIGNADO".
  - 🖼️  Header HTML del tablero corregido: ya no muestra código fuente visible.
  - 📄 PDF enriquecido: logo Beccacece, gráfico de barras Bultos UP por camión,
    tabla con Cap. kg + Peso + Rechazos, alertas destacadas de licencias vencidas,
    pie de página con versión y causa de demora.


CAMBIOS v4.43.0 — Tablero Ruteador:
  - 🔧 FIX: ANR ahora se lee desde "anr_df" (session_state, header=0)
    igual que el resto de los tabs. Antes buscaba "df_anr" que no existe.
  - 🔧 FIX: Ocupación bodega ahora tiene formula correcta del Excel:
    (UP_total + UP_pendientes - UP_rechazados) / (camiones_reparto + segundas_vueltas)
    en lugar de ser calculada como UP/camiones directamente.
  - 🔧 FIX: Fuera de zona — el input es CANTIDAD de PDV (ya estaba correcto),
    el porcentaje se calcula con formula: cantidad / pedidos_ruteados.
  - 📄 Exportación PDF del día: genera PDF envío-mail con logo, KPIs,
    tabla por camión, status semáforo y fecha, listo para adjuntar.
  - 📊 Exportación Excel del día mejorada: hoja KPIs, hoja Detalle,
    hoja Ventas Especiales.
  - 📅 Tab "📅 Histórico Mensual": permite subir los Excel diarios generados
    para conformar el tablero mensual acumulado. Almacena registros en
    session_state y muestra dashboard con:
    · Gráficos de tendencia (SLA %, Eficiencia, Drop Size, Util. Vehículos)
    · Heatmap días hábiles vs KPIs
    · Análisis de tendencia y alertas automáticas
    · Comparativo vs targets
    · Exportación del histórico mensual consolidado como Excel.
"""

import io, math, hashlib, unicodedata, os, shutil, json
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from loguru import logger

# Para Tab Boletas
try:
    from pypdf import PdfReader, PdfWriter
    import pdfplumber as _pdfplumber_lib
    _PYPDF_AVAILABLE = True
except ImportError:
    _PYPDF_AVAILABLE = False

# ─── VERSIÓN Y CONFIG GLOBAL ────────────────────────────────────────────────
APP_VERSION = "4.73.0"
SNAPSHOT_DIR = Path("./snapshots")

# Colores T2 (Sprint 3)
T2_COLORS = {
    "verde": "#22c55e",     # Pallet pura completa
    "amarillo": "#eab308",  # Pallet picking (parcial)
    "gris": "#9ca3af",      # Vacío
    "rojo": "#ef4444",      # Alerta (excedido)
}

# Setup logger único (no duplicar handlers en re-runs Streamlit)
try:
    if "logger_configured" not in st.session_state:
        logger.remove()
        logger.add(
            lambda m: None,  # Stdout silenciado en Streamlit (lo capturamos en sesión)
            level="INFO",
            format="{time:HH:mm:ss} | {level} | {message}",
        )
        st.session_state["logger_configured"] = True
        st.session_state["log_buffer"] = []
except Exception:
    # session_state aún no disponible en este punto del ciclo de vida: se
    # configurará perezosamente dentro de log_event(). No bloquear el arranque.
    pass

def log_event(level: str, msg: str):
    """Log a archivo (futuro) y a buffer de sesión (para mostrar en UI)."""
    ts = datetime.now().strftime("%H:%M:%S")
    entry = f"[{ts}] {level.upper()}: {msg}"
    st.session_state.setdefault("log_buffer", []).append(entry)
    logger.log(level.upper(), msg)

# ============================================================================
# ║              BLOQUE REUSADO LITERAL DE app_v3_9_1.py                     ║
# ║                                                                          ║
# ║  Funciones críticas para Tabs 1 y 2:                                     ║
# ║   - load_car, load_frescura, load_agr                                    ║
# ║   - generate_pdf (Planilla de Carga)                                     ║
# ║   - build_resumen_carga_pdf (Resumen por Camión)                         ║
# ║   + todas las helpers de dibujo del PDF (draw_*, header_*, etc)          ║
# ║                                                                          ║
# ║  ⚠ NO MODIFICAR — lógica validada en producción.                         ║
# ============================================================================

import io, math, hashlib, unicodedata
from datetime import datetime

import pandas as pd
import streamlit as st
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas

# ─── MEDIDAS FIJAS (puntos) ──────────────────────────────────────────────────
PAGE_W, PAGE_H = A4
MARGIN  = 12 * mm
CW      = PAGE_W - 2 * MARGIN

H_TITLE     = 26
H_TRANS     = 18
H_PARTIDA   = 16
H_CTRL      = 56
H_LEMA      = 18
H_CANCHA    = 20
H_THDR      = 14
H_ALM       = 16
H_ROW       = 14
H_TOT       = 14
H_ESPACIO   = 62
H_ROUTE     = 16
H_FOOT      = 28
GAP         = 2

# ─── COLORES ────────────────────────────────────────────────────────────────
DARK_BLUE  = colors.HexColor('#1a3a6b')
MED_BLUE   = colors.HexColor('#2e5fa3')
LIGHT_GRAY = colors.HexColor('#EFEFEF')
BG_GRAY    = colors.HexColor('#F4F4F4')
DARK_ROW   = colors.HexColor('#AAAAAA')
MED_ROW    = colors.HexColor('#D8D8D8')
AMBER      = colors.HexColor('#FFE0B2')
RED_ALERT  = colors.HexColor('#C0392B')
FOOT_GRAY  = colors.HexColor('#555555')
BORDER     = colors.HexColor('#AAAAAA')
HDR_BG     = colors.HexColor('#E8E8E8')
YELLOW_RTE = colors.HexColor('#FFF3A0')
WM_GRAY    = colors.HexColor('#888888')
ALM_BG     = colors.HexColor('#DDEEFF')

EXCLUDED_SKUS = {2730, 2731, 2776, 2780, 5192}
EXCLUDED_PATS = ['Q CERVEZAS', 'Q PLAS', 'BOT 1/1', 'BOT AMBAR', 'ARACELI']
CANCHA_ORDER  = ['CANCHA I', 'CANCHA II', 'CANCHA III', 'CANCHA IV', 'CANCHA V']
COL_W = [50, CW - 50 - 62 - 52 - 52, 62, 52, 52]

LEMAS = [
    "Cada bulto bien puesto es una entrega perfecta. Dale con todo!",
    "El picking preciso empieza con vos. Conta, verifica, confia.",
    "Hoy es dia de hacer historia en el deposito. A romperla!",
    "Rapido no es apurado. Rapido es preciso y sin errores.",
    "El mejor pickero no es el mas veloz, es el mas exacto.",
    "Tu equipo cuenta con vos. Cada bulto importa.",
    "Zapatos abrochados, espalda cuidada, mente enfocada. Arrancamos!",
    "Paleta perfecta = cliente feliz = empresa fuerte. Todo empieza aca.",
    "Cero errores de picking es el estandar. Vos podes.",
    "Si dudas del SKU, verifica. Un segundo de control evita una hora de correccion.",
    "La diferencia entre bueno y excelente esta en los detalles. Se excelente.",
    "Arrancar bien es terminar mejor. Que empiece el picking!",
    "Trabajo limpio, paleta ordenada, conciencia tranquila.",
    "Sos parte del corazon de la operacion. El deposito te necesita al 100%.",
    "Pensa antes de mover. El orden en la paleta es orden en la entrega.",
    "Tecnica correcta al levantar = cuerpo sano manana tambien.",
    "Verifica el SKU, verifica la cantidad, verifica el lote. Tres checks, cero errores!",
    "Cada noche de picking sin errores construye la reputacion de todos.",
    "El deposito funciona porque vos funcionas. Gracias por el compromiso.",
    "Paleta bien armada, camion bien cargado, cliente bien atendido.",
    "El cansancio es real. El orgullo del trabajo bien hecho tambien.",
    "Reporta lo que esta mal antes de moverlo. Un reporte a tiempo vale oro.",
    "Somos un equipo. Si uno falla, fallamos todos. Si uno gana, ganamos todos.",
    "Segui la planilla al pie de la letra. Esta hecha para que llegues sin errores.",
    "Esta noche vas a hacer 500 movimientos correctos. Demostralo!",
    "Seguridad primero: EPI puesto, pasillo libre, mente clara.",
    "Los mejores pickeros no nacen, se hacen en noches como esta.",
    "Picking nocturno = concentracion maxima. El dia descansa, vos construis.",
    "Cada producto en su lugar es una promesa cumplida al cliente.",
    "Ultima caja, mismo cuidado que la primera. Asi se hace el trabajo bien hecho.",
    "Un error de picking cuesta mas que diez movimientos correctos. Tomaste el camino del control.",
    "El deposito en orden es un equipo en orden. Vos sos parte de eso.",
    "No dejes pasar un faltante sin reportarlo. Una alerta a tiempo salva una entrega.",
    "Carga bien hecha desde el deposito = cero devoluciones = todos ganan.",
    "El mejor momento para revisar es antes de cerrar el pallet. Despues es tarde.",
    "Sos el ultimo control de calidad antes del cliente. Eso tiene un valor enorme.",
    "Hoy puede ser el dia con mas bultos del mes. Vamos con fuerza y precision.",
    "El cansancio de las 3 AM no puede mas que el orgullo de cero errores.",
    "Cada producto bien colocado es un compromiso cumplido con quien lo espera.",
    "La productividad no se mide en velocidad, se mide en entregas perfectas.",
    "Controlador: tu ojo critico es lo que separa lo bueno de lo excelente.",
    "El picking es un arte logistico. Esta noche, vos sos el artista.",
    "Conoces el deposito mejor que nadie. Esa experiencia vale oro para el equipo.",
    "Cero accidentes, cero errores, cero reclamos. El estandar DPO empieza con vos.",
    "Cada vez que hacemos esto bien, le ganamos terreno a la excelencia operativa.",
]

# ─── ENSEÑANZA / FRASE DEL DÍA ─────────────────────────────────────────────
# Una por día, rotación automática por número de día del año. Sin repetición.
ENSEÑANZAS_DIA = [
    "📖 «El éxito no es definitivo, el fracaso no es fatal: lo que cuenta es el coraje de continuar.» — Churchill",
    "🧠 «La calidad nunca es un accidente; siempre es el resultado de un esfuerzo inteligente.» — John Ruskin",
    "⚙️ «Si no puedes medir algo, no puedes mejorarlo.» — Peter Drucker | Control: medí tus bultos por hora hoy.",
    "🔁 «Pequeñas mejoras diarias llevan a resultados sorprendentes.» — Kaizen | Lean: 1% mejor cada turno.",
    "👁️ «El que mira hacia afuera sueña; el que mira hacia adentro despierta.» — Jung | Autocontrol: revisá tu postura al cargar.",
    "🎯 «Los planes son inútiles, pero la planificación es indispensable.» — Eisenhower | Revisá el orden de canchas antes de arrancar.",
    "🏆 «La excelencia no es un acto, es un hábito.» — Aristóteles | Hoy: un control extra por pallet.",
    "🔍 «En Dios confiamos. Todos los demás traigan datos.» — W. Edwards Deming | DPO: lo que no se registra no se mejora.",
    "🛡️ «La seguridad no es una prioridad — es un valor.» — DuPont | EPI completo antes de arrancar.",
    "📦 «El stock es como agua: fluye si el canal está limpio.» — FIFO/FEFO: mover lo más viejo primero.",
    "🤝 «El trabajo en equipo divide el esfuerzo y multiplica el éxito.» — Equidistribución de carga entre canchas.",
    "📋 «Lo que se escribe se recuerda; lo que se registra se mejora.» — OWD: cada observación cuenta.",
    "⏱️ «El tiempo de ciclo solo mejora si lo conocés.» — Midí tu tiempo de picking hoy y compará.",
    "💡 «Un problema bien definido está medio resuelto.» — 5 Porqués: antes de culpar, entender.",
    "🔧 «El mantenimiento preventivo es mil veces más barato que el correctivo.» — Checklist AE antes de operar.",
    "🌱 «La cultura se come a la estrategia en el desayuno.» — Peter Drucker | Cada hábito correcto construye cultura.",
    "📊 «Sin estándares no hay mejora.» — Taiichi Ohno | SOP: la planilla de carga es el estándar. Seguirla es respetarlo.",
    "🧩 «El todo es más que la suma de sus partes.» — Aristóteles | Picking, AE y control: los tres juntos hacen la diferencia.",
    "🚀 «La velocidad sin dirección solo genera accidentes.» — Rápido con precisión, no rápido con errores.",
    "🔄 «La mejora continua es mejor que la perfección tardía.» — Mark Twain | Un ajuste pequeño hoy vale más que un gran plan mañana.",
    "🧘 «La pausa de un segundo de control evita una hora de corrección.» — Verificá el SKU antes de palletizar.",
    "📍 «Si no sabés dónde estás parado, no podés saber hacia dónde vas.» — El DQI empieza con inventario real.",
    "🎓 «El experto fue alguna vez un principiante que no se rindió.» — Capacitá a alguien hoy, aunque sea una cosa.",
    "⚡ «La eficiencia es hacer bien las cosas; la eficacia es hacer las cosas correctas.» — Drucker",
    "📉 «El desperdicio más peligroso es el que no vemos.» — Lean | ¿Cuántos movimientos innecesarios hacés por turno?",
    "🏗️ «La estructura que soporta el negocio se construye bulto a bulto, noche a noche.» — Beccacece Hnos",
    "🧭 «La brújula del buen operador siempre apunta al cliente.» — Cada error de picking es un cliente insatisfecho.",
    "🔗 «La cadena es tan fuerte como su eslabón más débil.» — El eslabón esta noche sos vos. Sé el más fuerte.",
    "📌 «Organización es el arte de hacer que las cosas sucedan.» — Antes de empezar: cancha ordenada, planilla en mano.",
    "🌟 «El reconocimiento comienza por el autocontrol.» — Los mejores operadores se evalúan solos primero.",
]

def get_ensenanza_dia() -> str:
    """Retorna la enseñanza del día — rota automáticamente por día del año, sin repetición."""
    import datetime as _dte
    day_of_year = _dte.date.today().timetuple().tm_yday
    return ENSEÑANZAS_DIA[day_of_year % len(ENSEÑANZAS_DIA)]

# ─── UTILIDADES ─────────────────────────────────────────────────────────────

def _norm(s):
    """Normaliza string: sin tildes, sin espacios extra, lowercase."""
    if s is None: return ''
    s = str(s).strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.lower()

def find_col(df, *candidates):
    """Busca columna en df por nombre normalizado (tolerante a tildes/case)."""
    norm_map = {_norm(c): c for c in df.columns}
    for cand in candidates:
        c = _norm(cand)
        if c in norm_map:
            return norm_map[c]
    return None

def to_int_sku(val):
    """Convierte SKU a int de forma robusta. Devuelve None si no se puede."""
    if val is None: return None
    if isinstance(val, float) and math.isnan(val): return None
    try:
        s = str(val).strip()
        if s in ('', 'nan', 'NaN', 'None'): return None
        return int(float(s))
    except (ValueError, TypeError):
        return None

def is_envase(sku, desc):
    if sku in EXCLUDED_SKUS: return True
    d = str(desc).upper()
    for p in EXCLUDED_PATS:
        if p in d: return True
    return d.strip() == 'RET'

def normalize_cancha(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 'SIN CANCHA'
    s = str(val).strip().upper()
    if s in ('', '0', 'NAN', 'NONE'): return 'SIN CANCHA'
    # Eliminar tildes
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    MAP = {'CANCHA I':'CANCHA I','CANCHA II':'CANCHA II','CANCHA III':'CANCHA III',
           'CANCHA IV':'CANCHA IV','CANCHA V':'CANCHA V',
           'MKPL':'CANCHA V','MERCH':'CANCHA V','MERCHANDISING':'CANCHA V',
           'MARKETPLACE':'CANCHA V'}
    # Match exacto primero
    if s in MAP: return MAP[s]
    # Match por prefijo (más específico primero)
    for k in sorted(MAP.keys(), key=len, reverse=True):
        if s.startswith(k): return MAP[k]
    return 'SIN CANCHA'

# ─── CARGA DE DATOS ─────────────────────────────────────────────────────────

def load_car(file, ddm=None):
    """
    v3.8 — Si se pasa ddm, se usa para decidir si los huérfanos del bloque azul
    son paletas enteras (no se agregan) o tienen fracción / no tienen BXP
    (se agregan al df). Si ddm es None, no se procesan huérfanos (compat retro).

    Retorna: (df, blue_audit) donde blue_audit es una lista de dicts con la
    trazabilidad de cada fila del bloque azul y su destino.
    """
    # ── Leer hoja completa sin header para manejar filas por índice ──────────
    raw = pd.read_excel(file, sheet_name=0, header=None)

    # Fila 0 = header (Excel fila 1)
    header = raw.iloc[0].tolist()

    # ── FILAS AZULES: Excel filas 2-41 → índices 1-40 ────────────────────────
    # Columna T = índice 19 (A=0 … T=19)
    blue_raw = raw.iloc[1:41].copy()
    blue_raw.columns = header

    blue_raw["Artículo"]   = pd.to_numeric(blue_raw["Artículo"], errors="coerce")
    blue_raw["Transporte"] = pd.to_numeric(blue_raw["Transporte"], errors="coerce")
    blue_raw["_bultos_azul"] = pd.to_numeric(blue_raw.iloc[:, 19], errors="coerce").fillna(0)
    blue_raw = blue_raw.dropna(subset=["Artículo", "Transporte"])
    blue_raw = blue_raw[blue_raw["Artículo"] > 0]
    blue_raw = blue_raw[blue_raw["_bultos_azul"] > 0]
    blue_raw["Artículo"]   = blue_raw["Artículo"].astype(int)
    blue_raw["Transporte"] = blue_raw["Transporte"].apply(lambda x: int(float(x)) if pd.notna(x) else x)
    # Excluir envases de las filas azules
    blue_raw = blue_raw[~blue_raw.apply(
        lambda r: is_envase(int(r["Artículo"]), str(r.get("Descripción Artículo", ""))), axis=1
    )]

    # Guardar descripción del primer registro azul por SKU (para huérfanos)
    blue_desc_map = (blue_raw.dropna(subset=["Descripción Artículo"])
                     .drop_duplicates(subset=["Artículo"], keep="first")
                     .set_index("Artículo")["Descripción Artículo"].to_dict())

    # Agregar bultos azules por Transporte + SKU
    blue_agg = (blue_raw.groupby(["Transporte", "Artículo"], as_index=False)
                .agg(_bultos_azul=("_bultos_azul", "sum")))
    blue_total_in = float(blue_agg["_bultos_azul"].sum()) if len(blue_agg) else 0.0

    # ── FILAS NORMALES: Excel fila 42 en adelante → índice 41+ ───────────────
    normal_raw = raw.iloc[41:].copy()
    normal_raw.columns = header

    df = normal_raw.copy()
    df["Artículo"]   = pd.to_numeric(df["Artículo"], errors="coerce")
    df["Transporte"] = pd.to_numeric(df["Transporte"], errors="coerce")
    df = df.dropna(subset=["Artículo", "Transporte"])
    df = df[df["Artículo"] > 0]
    df["Artículo"]   = df["Artículo"].astype(int)
    df["Transporte"] = df["Transporte"].apply(lambda x: int(float(x)) if pd.notna(x) else x)
    df["Bultos"]     = pd.to_numeric(df["Bultos"], errors="coerce").fillna(0)
    df["Unids"]      = pd.to_numeric(df["Unids"],  errors="coerce").fillna(0)
    df = df[~df.apply(lambda r: is_envase(r["Artículo"], r["Descripción Artículo"]), axis=1)]

    # Inicializar flag de huérfano sin DDM (siempre False por defecto)
    df["_orphan_no_ddm"] = False
    blue_audit = []

    # ── MERGE: sumar bultos azules sobre los normales ─────────────────────────
    blue_consumed_merge = 0.0
    blue_consumed_orphans = 0.0
    blue_discarded_pal_entera = 0.0

    if not blue_agg.empty:
        # Set de pares (T, SKU) que existen en filas normales
        normal_keys = set(zip(df["Transporte"], df["Artículo"]))

        # Partir blue_agg en dos: los que tienen match y los huérfanos
        blue_agg["_has_match"] = blue_agg.apply(
            lambda r: (r["Transporte"], r["Artículo"]) in normal_keys, axis=1)

        blue_match    = blue_agg[blue_agg["_has_match"]].copy()
        blue_orphans  = blue_agg[~blue_agg["_has_match"]].copy()

        # ── Caso 1: con match → merge clásico (sumar bultos azules) ───────────
        if not blue_match.empty:
            blue_match_clean = blue_match[["Transporte","Artículo","_bultos_azul"]]
            df = df.merge(blue_match_clean, on=["Transporte", "Artículo"], how="left")
            df["_bultos_azul"] = df["_bultos_azul"].fillna(0)
            df["Bultos"] = df["Bultos"] + df["_bultos_azul"]
            blue_consumed_merge = float(blue_match_clean["_bultos_azul"].sum())
            df = df.drop(columns=["_bultos_azul"])

            # Trazabilidad: filas sumadas
            for _, br in blue_match.iterrows():
                blue_audit.append({
                    "Transporte": int(br["Transporte"]),
                    "SKU":        int(br["Artículo"]),
                    "Descripción": str(blue_desc_map.get(int(br["Artículo"]), "")),
                    "Bultos azules": float(br["_bultos_azul"]),
                    "BXP": None,
                    "Acción": "Sumado a fila normal",
                })

        # ── Caso 2: huérfanos → decidir según DDM ─────────────────────────────
        if not blue_orphans.empty and ddm is not None:
            # Tomar info de reparto y descripción de transporte de cualquier
            # fila normal del mismo transporte (siempre existe)
            transp_info = (df.dropna(subset=["Número"])
                           .drop_duplicates(subset=["Transporte"], keep="first")
                           .set_index("Transporte")
                           [["Número","Fecha Mvto","Descripción Transporte","Depósito","Descripción Depósito"]]
                           .to_dict("index"))

            new_rows = []
            for _, br in blue_orphans.iterrows():
                t   = int(br["Transporte"]); sku = int(br["Artículo"])
                blt = float(br["_bultos_azul"])
                desc = str(blue_desc_map.get(sku, f"SKU {sku}"))

                d = ddm[ddm["sku"] == sku]
                bxp = float(d.iloc[0]["bxp"]) if len(d) and pd.notna(d.iloc[0]["bxp"]) else None

                if bxp and bxp > 0 and (blt % bxp == 0):
                    # Paleta entera exacta → NO agregar (no es picking)
                    blue_discarded_pal_entera += blt
                    blue_audit.append({
                        "Transporte": t, "SKU": sku, "Descripción": desc,
                        "Bultos azules": blt, "BXP": bxp,
                        "Acción": "Descartado (paleta entera)",
                    })
                    continue

                # Necesita aparecer en PDF: tiene fracción o no tiene BXP
                if t not in transp_info:
                    # No hay info de reparto → no se puede asignar a un reparto
                    # No debería pasar nunca pero registramos
                    blue_audit.append({
                        "Transporte": t, "SKU": sku, "Descripción": desc,
                        "Bultos azules": blt, "BXP": bxp,
                        "Acción": "ERROR: sin info de reparto",
                    })
                    continue

                info = transp_info[t]
                no_ddm = (bxp is None or bxp <= 0)

                new_row = {col: None for col in df.columns}
                new_row["Depósito"]              = info.get("Depósito")
                new_row["Descripción Depósito"]  = info.get("Descripción Depósito")
                new_row["Número"]                = info["Número"]
                new_row["Fecha Mvto"]            = info["Fecha Mvto"]
                new_row["Transporte"]            = t
                new_row["Descripción Transporte"]= info["Descripción Transporte"]
                new_row["Artículo"]              = sku
                new_row["Descripción Artículo"]  = desc
                new_row["Bultos"]                = blt
                new_row["Unids"]                 = 0
                new_row["_orphan_no_ddm"]        = no_ddm

                new_rows.append(new_row)
                blue_consumed_orphans += blt

                blue_audit.append({
                    "Transporte": t, "SKU": sku, "Descripción": desc,
                    "Bultos azules": blt, "BXP": bxp,
                    "Acción": ("Agregado como picking (SIN BXP — REVISAR)"
                               if no_ddm else "Agregado como picking (fracción)"),
                })

            if new_rows:
                # NOTA: el header del CAR puede tener columnas con nombres
                # duplicados (espacios non-breaking). Para evitar fallar en
                # pd.concat con InvalidIndexError, construimos el dataframe
                # nuevo respetando posiciones de columnas, no nombres.
                new_df = pd.DataFrame(new_rows)
                # Reindexar usando solo las columnas que SÍ existen en df
                # (asignación posicional por nombre único)
                aligned = pd.DataFrame(columns=df.columns)
                for nr in new_rows:
                    aligned_row = []
                    for col in df.columns:
                        # Si la columna está repetida tomamos el primer valor encontrado
                        aligned_row.append(nr.get(col, None))
                    aligned.loc[len(aligned)] = aligned_row
                df = pd.concat([df, aligned], ignore_index=True, sort=False)

        elif not blue_orphans.empty and ddm is None:
            # Sin DDM no podemos decidir → registrar pero no agregar (compat)
            for _, br in blue_orphans.iterrows():
                blue_audit.append({
                    "Transporte": int(br["Transporte"]),
                    "SKU":        int(br["Artículo"]),
                    "Descripción": str(blue_desc_map.get(int(br["Artículo"]), "")),
                    "Bultos azules": float(br["_bultos_azul"]),
                    "BXP": None,
                    "Acción": "PERDIDO (sin DDM para decidir)",
                })

    # ── Invariante de validación ─────────────────────────────────────────────
    blue_total_out = (blue_consumed_merge + blue_consumed_orphans
                      + blue_discarded_pal_entera)
    if blue_total_in > 0 and abs(blue_total_in - blue_total_out) > 0.01 and ddm is not None:
        raise ValueError(
            f"Invariante de bloque azul rota: entrada={blue_total_in:.0f} "
            f"bultos, salida={blue_total_out:.0f} bultos "
            f"(merge={blue_consumed_merge:.0f} + huérfanos={blue_consumed_orphans:.0f} "
            f"+ descartados pal entera={blue_discarded_pal_entera:.0f}). "
            f"Diferencia: {blue_total_in - blue_total_out:.0f} bultos perdidos."
        )

    return df, blue_audit

def load_frescura(file):
    """
    FIX v3.4 — Construcción correcta de cancha por SKU:
      1) Lee la hoja API completa.
      2) SKU está en columna A ('Artículo'). Cada SKU tiene un Nº de almacén
         (columna 'almacén' o similar).
      3) Lee el bloque I:M de la misma hoja API (tabla maestra de almacenes):
         I=Nº Almacén | J=Detalle | K=Calibre | L=Apilabilidad | M=Cancha
      4) Mergea SKU → Nº Almacén → Cancha.
    """
    xls = pd.ExcelFile(file)

    # ── HOJA API: lectura completa ──────────────────────────────────────────
    api_full = pd.read_excel(xls, sheet_name='API', header=0)

    # Columna SKU (col A)
    col_sku = find_col(api_full, 'Artículo', 'Articulo', 'SKU', 'Cód', 'Cod')
    if col_sku is None:
        raise ValueError("Hoja API: no se encontró la columna de SKU (Artículo).")

    # Columna almacén por SKU (entre A y H, la que asocia SKU con su almacén)
    col_alm_sku = find_col(api_full, 'almacén', 'almacen', 'Almacén', 'Almacen', 'Alm')
    if col_alm_sku is None:
        raise ValueError("Hoja API: no se encontró la columna 'almacén' que asocia SKU con almacén.")

    # Tabla SKU → Nº Almacén
    sku_alm = api_full[[col_sku, col_alm_sku]].copy()
    sku_alm.columns = ['sku_raw', 'alm_num_raw']
    sku_alm['sku'] = sku_alm['sku_raw'].apply(to_int_sku)
    sku_alm['alm_num'] = pd.to_numeric(sku_alm['alm_num_raw'], errors='coerce')
    sku_alm = sku_alm.dropna(subset=['sku'])
    sku_alm['sku'] = sku_alm['sku'].astype(int)
    # Regla de prevalencia: la última fila con almacén válido prevalece
    sku_alm = sku_alm.dropna(subset=['alm_num'])
    sku_alm['alm_num'] = sku_alm['alm_num'].astype(int)
    sku_alm = sku_alm.drop_duplicates(subset=['sku'], keep='last')[['sku', 'alm_num']]

    # ── BLOQUE I:M (tabla maestra de almacenes), header en fila 1 ───────────
    # usecols='I:M', header=0 → lee solo las cols I-M con su header propio.
    alm_master = pd.read_excel(
        xls, sheet_name='API', header=0, usecols='I:M'
    )
    # Mapeo posicional según definición del usuario:
    # I = Nº Almacén | J = Detalle | K = Calibre | L = Apilabilidad | M = Cancha
    if alm_master.shape[1] < 5:
        raise ValueError(
            f"Bloque I:M de la hoja API tiene {alm_master.shape[1]} columnas, se esperaban 5."
        )
    alm_master = alm_master.iloc[:, :5].copy()
    alm_master.columns = ['alm_num', 'alm_det', 'calibre', 'apil', 'cancha_raw']
    alm_master['alm_num'] = pd.to_numeric(alm_master['alm_num'], errors='coerce')
    alm_master = alm_master.dropna(subset=['alm_num'])
    alm_master['alm_num'] = alm_master['alm_num'].astype(int)
    # Prevalece la última definición del almacén
    alm_master = alm_master.drop_duplicates(subset=['alm_num'], keep='last')

    # ── MERGE: SKU → Almacén → Cancha ───────────────────────────────────────
    api = sku_alm.merge(alm_master, on='alm_num', how='left')
    api['cancha'] = api['cancha_raw'].apply(normalize_cancha)
    api = api.rename(columns={'alm_num': 'alm_id'})
    api['alm_det'] = api['alm_det'].fillna('sin datos').astype(str)
    api = api[['sku', 'cancha', 'alm_id', 'alm_det']]

    # ── HOJA DDM (Bultos por pallet) ────────────────────────────────────────
    ddm_full = pd.read_excel(xls, sheet_name='DDM')
    col_ddm_sku = find_col(ddm_full, 'ARTÍCULO', 'ARTICULO', 'Artículo', 'Articulo', 'SKU')
    col_bxp     = find_col(ddm_full, 'BULTOS X PALLET', 'BULTOS_X_PALLET', 'BXP')
    if not col_ddm_sku or not col_bxp:
        raise ValueError("Hoja DDM: no se encontraron las columnas ARTÍCULO y/o BULTOS X PALLET.")
    ddm = ddm_full[[col_ddm_sku, col_bxp]].copy()
    ddm.columns = ['sku', 'bxp']
    ddm['sku'] = ddm['sku'].apply(to_int_sku)
    ddm = ddm.dropna(subset=['sku'])
    ddm['sku'] = ddm['sku'].astype(int)
    ddm['bxp'] = pd.to_numeric(ddm['bxp'], errors='coerce')
    ddm = ddm.drop_duplicates(subset=['sku'], keep='last')

    # ── HOJA Frescura ───────────────────────────────────────────────────────
    fr_full = pd.read_excel(xls, sheet_name='Frescura')
    col_fr_sku    = find_col(fr_full, 'Cód', 'Cod', 'SKU', 'Artículo', 'Articulo')
    col_fr_fecha  = find_col(fr_full, 'FECHA DE VENC.', 'FECHA DE VENC', 'Fecha de Venc', 'Vencimiento')
    col_fr_status = find_col(fr_full, 'Status', 'Estado')
    if not all([col_fr_sku, col_fr_fecha, col_fr_status]):
        raise ValueError("Hoja Frescura: faltan columnas Cód / FECHA DE VENC. / Status.")
    fr = fr_full[[col_fr_sku, col_fr_fecha, col_fr_status]].copy()
    fr.columns = ['sku', 'fecha', 'status']
    fr['sku'] = fr['sku'].apply(to_int_sku)
    fr = fr.dropna(subset=['sku'])
    fr['sku'] = fr['sku'].astype(int)
    fr = fr.sort_values('fecha').drop_duplicates(subset=['sku'], keep='first')

    # Diagnóstico para debug
    diag = {
        'api_full_cols': list(api_full.columns),
        'col_sku_used': col_sku,
        'col_alm_sku_used': col_alm_sku,
        'alm_master_cols_raw': list(pd.read_excel(xls, sheet_name='API', header=0, usecols='I:M').columns),
        'sku_count_api': len(api),
        'alm_master_count': len(alm_master),
        'sku_sin_cancha': int((api['cancha'] == 'SIN CANCHA').sum()),
        'alm_master_preview': alm_master.head(20).to_dict('records'),
    }
    return api, ddm, fr, diag

def get_lema(transport, idx=0):
    h = int(hashlib.md5(str(transport).encode()).hexdigest(), 16)
    return LEMAS[(h + idx) % len(LEMAS)]

def pallet_reading(val):
    if not val or val == 0.0:
        return ('0.00', '0 paletas', 'Sin carga de picking en esta cancha')
    vm = f'{val:.2f}'; n = int(val); fr = val - n
    if val < 1.0:  return (vm, '1 paleta',    'Menos de una paleta')
    if val == 1.0: return (vm, '1 paleta',    '1 paleta exacta')
    if val < 1.5:  return (vm, '2 paletas',   'Mas de una paleta')
    if val == 1.5: return (vm, '2 paletas',   '1 paleta y media')
    if val < 2.0:  return (vm, '2 paletas',   'Casi 2 paletas')
    if fr == 0.0:  return (vm, f'{n} paletas',   f'{n} paletas exactas')
    if fr < 0.5:   return (vm, f'{n+1} paletas', f'Mas de {n} paletas')
    if fr == 0.5:  return (vm, f'{n+1} paletas', f'{n} paletas y media')
    return          (vm, f'{n+1} paletas', f'Casi {n+1} paletas')

def process_reparto(rep_df, api, ddm, fr):
    rows = []
    for _, r in rep_df.iterrows():
        sku = int(r['Artículo'])
        blts_raw = float(r['Bultos']); unids = float(r['Unids'])
        desc = str(r['Descripción Artículo'])
        orphan_no_ddm = bool(r.get('_orphan_no_ddm', False))

        a = api[api['sku'] == sku]
        if len(a):
            cancha     = str(a.iloc[0]['cancha'])
            alm_id_i   = int(a.iloc[0]['alm_id']) if pd.notna(a.iloc[0]['alm_id']) else None
            alm_det    = str(a.iloc[0]['alm_det'])
        else:
            cancha, alm_id_i, alm_det = 'SIN CANCHA', None, 'sin datos'

        d = ddm[ddm['sku'] == sku]
        bxp = float(d.iloc[0]['bxp']) if len(d) and pd.notna(d.iloc[0]['bxp']) else None

        if bxp and bxp > 0:
            pal_ent = int(blts_raw / bxp); blts_pick = blts_raw - pal_ent * bxp
            has_pal = pal_ent >= 1; miss_bxp = False
        else:
            pal_ent = 0; blts_pick = blts_raw; has_pal = False; miss_bxp = True

        skip_color = (sku == 7038) or (alm_id_i == 46)
        f = fr[fr['sku'] == sku]
        if len(f) and not skip_color:
            fv = f.iloc[0]['fecha']; st_ = str(f.iloc[0]['status']).strip().upper()
            if pd.isna(fv): fecha_s, row_st = '', 'NONE'
            else:
                fecha_s = pd.Timestamp(fv).strftime('%d/%m/%y')
                row_st  = st_ if st_ in ('RED','YELLOW','GREEN') else 'NONE'
        else:
            fecha_s, row_st = '', 'NONE'

        rows.append({'sku':sku,'desc':desc,'blts_raw':blts_raw,'unids':unids,
                     'blts_pick':blts_pick,'pal_ent':pal_ent,'has_pal':has_pal,
                     'miss_bxp':miss_bxp,'bxp':bxp,'cancha':cancha,
                     'alm_id':alm_id_i,'alm_det':alm_det,'fecha_s':fecha_s,'row_st':row_st,
                     'orphan_no_ddm':orphan_no_ddm})
    return rows

def build_alm_groups(c_rows):
    groups = {}; order = []
    for r in c_rows:
        key = (r['alm_id'], r['alm_det'])
        if key not in groups: groups[key] = []; order.append(key)
        groups[key].append(r)
    for key in order:
        groups[key].sort(key=lambda x: -x['blts_raw'])
    # Ordenar almacenes de menor a mayor por número de almacén
    order.sort(key=lambda k: (k[0] is None, k[0] if k[0] is not None else 0))
    return [(k[0], k[1], groups[k]) for k in order]

def compute_pall_value(rows, cancha):
    return sum(r['blts_pick']/r['bxp'] for r in rows
               if r['cancha'] == cancha and r['bxp'] and r['bxp'] > 0)

# ─── DRAWING HELPERS ─────────────────────────────────────────────────────────

def ry(y_top): return PAGE_H - y_top

def rfill(c, x, y_top, w, h, fill, stroke=None, lw=0.4):
    c.setFillColor(fill)
    if stroke:
        c.setStrokeColor(stroke); c.setLineWidth(lw)
        c.rect(x, ry(y_top+h), w, h, fill=1, stroke=1)
    else:
        c.rect(x, ry(y_top+h), w, h, fill=1, stroke=0)

def txt(c, x, y_top, s, font='Helvetica', sz=8, col=colors.black, align='left', mw=None):
    c.setFont(font, sz); c.setFillColor(col)
    if mw:
        while s and c.stringWidth(s, font, sz) > mw: s = s[:-1]
    fn = {'left': c.drawString, 'center': c.drawCentredString, 'right': c.drawRightString}[align]
    fn(x, ry(y_top), s)

def draw_watermark(c):
    c.saveState()
    c.setFillColor(WM_GRAY); c.setFillAlpha(0.07)
    c.setFont('Helvetica-Bold', 68)
    c.translate(PAGE_W/2, PAGE_H/2); c.rotate(35)
    c.drawCentredString(0, 0, 'BECCACECE HNOS')
    c.restoreState()

def draw_title_bar(c, y, numero, fecha_str):
    rfill(c, MARGIN, y, CW, H_TITLE, DARK_BLUE)
    label = f'PLANILLA DE CARGA  |  Reparto Nro: {numero}  |  Fecha: {fecha_str}'
    txt(c, MARGIN+CW/2, y+17, label, 'Helvetica-Bold', 10.5, colors.white, 'center', CW-16)
    return H_TITLE

def fmt_transport(t):
    """Elimina el .0 final del número de transporte (128.0 → 128)."""
    try:
        v = float(t)
        if v == int(v):
            return str(int(v))
    except (ValueError, TypeError):
        pass
    return str(t)

def draw_transport_line(c, y, transport, chofer):
    rfill(c, MARGIN, y, CW, H_TRANS, colors.HexColor('#EEF2F8'))
    txt(c, MARGIN+6,      y+12, f'Transporte: {fmt_transport(transport)} - {chofer}', 'Helvetica-Bold', 9)
    txt(c, MARGIN+CW-6,   y+12, 'Depósito: 001 - CASA CENTRAL',        'Helvetica-Bold', 9, align='right')
    return H_TRANS

def draw_partida_regreso(c, y):
    txt(c, MARGIN+6,     y+11, 'F. y H. Est. de Partida: ________ Hs.', 'Helvetica', 9)
    txt(c, MARGIN+CW-6,  y+11, 'F. y H. Est. de Regreso: ________ Hs.','Helvetica', 9, align='right')
    return H_PARTIDA

def draw_control_carga(c, y):
    hdr_h = 15; box_h = H_CTRL - hdr_h; half = CW/2
    rfill(c, MARGIN,        y, half, hdr_h, HDR_BG, BORDER)
    rfill(c, MARGIN+half,   y, half, hdr_h, HDR_BG, BORDER)
    txt(c, MARGIN+half/2,         y+10, 'CONTROL DE CARGA',    'Helvetica-Bold', 9, align='center')
    txt(c, MARGIN+half+half/2,    y+10, 'CONTROL DE DESCARGA', 'Helvetica-Bold', 9, align='center')
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)
    c.rect(MARGIN, ry(y+hdr_h+box_h), CW, box_h, fill=0, stroke=1)
    c.line(MARGIN+half, ry(y+hdr_h), MARGIN+half, ry(y+hdr_h+box_h))
    return H_CTRL

def draw_lema(c, y, lema):
    rfill(c, MARGIN, y, CW, H_LEMA, DARK_BLUE)
    txt(c, MARGIN+CW/2, y+12, f'"{lema}"', 'Helvetica-Oblique', 9, colors.white, 'center', CW-16)
    return H_LEMA

def draw_cancha_header(c, y, label):
    rfill(c, MARGIN, y, CW, H_CANCHA, MED_BLUE)
    txt(c, MARGIN+10, y+13, f'◀  {label}', 'Helvetica-Bold', 11, colors.white)
    return H_CANCHA

def draw_sin_cancha_header(c, y):
    rfill(c, MARGIN, y, CW, H_CANCHA, RED_ALERT)
    txt(c, MARGIN+CW/2, y+13, '⚠  SIN CANCHA ASIGNADA — REVISAR', 'Helvetica-Bold', 11, colors.white, 'center')
    return H_CANCHA

def draw_table_header(c, y):
    labels = ['SKU', 'Descripción', 'Venc.', 'Bultos', 'Unids']
    rfill(c, MARGIN, y, CW, H_THDR, HDR_BG, BORDER)
    x = MARGIN
    for lbl, w in zip(labels, COL_W):
        c.setStrokeColor(BORDER); c.setLineWidth(0.4)
        c.rect(x, ry(y+H_THDR), w, H_THDR, fill=0, stroke=1)
        txt(c, x+w/2, y+9.5, lbl, 'Helvetica-Bold', 8, align='center')
        x += w
    return H_THDR

def draw_alm_header(c, y, alm_id, alm_det):
    label = f'Almacén {alm_id} — {alm_det}' if alm_id is not None else 'Almacén: SIN DATOS'
    rfill(c, MARGIN, y, CW, H_ALM, ALM_BG, BORDER)
    txt(c, MARGIN+6, y+11, label, 'Helvetica-Bold', 9)
    return H_ALM

def draw_product_row(c, y, r):
    st_ = r['row_st']
    is_orphan_no_ddm = r.get('orphan_no_ddm', False)
    bg  = DARK_ROW if st_=='RED' else MED_ROW if st_=='YELLOW' else (
          AMBER if r.get('is_sin_cancha') else
          LIGHT_GRAY if r['has_pal'] else colors.white)

    rfill(c, MARGIN, y, CW, H_ROW, bg, BORDER)
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)

    # Marca de huérfano sin DDM: (*) rojo después del SKU
    sku_s = str(r['sku'])
    if is_orphan_no_ddm:
        sku_s = f"{r['sku']}(*)"
    elif r['miss_bxp']:
        sku_s = f"{r['sku']}(!)"

    bp     = r['blts_pick']; bp_s = str(int(bp)) if bp == int(bp) else f'{bp:.0f}'
    un_s   = str(int(r['unids']))
    ind    = {'RED':'■','YELLOW':'▲','GREEN':'○'}.get(st_,'')
    venc_s = f'{ind} {r["fecha_s"]}' if r['fecha_s'] else ''
    font   = 'Helvetica-Bold' if st_=='RED' else 'Helvetica'

    # Descripción con aviso embebido si es huérfano sin DDM
    desc_show = r['desc']
    if is_orphan_no_ddm:
        desc_show = f"{r['desc']}  ⚠ SIN BXP EN DDM — REVISAR"

    cells = [(sku_s,'center'),(desc_show,'left'),(venc_s,'center'),(bp_s,'center'),(un_s,'center')]
    x = MARGIN
    for i, ((s, align), w) in enumerate(zip(cells, COL_W)):
        c.rect(x, ry(y+H_ROW), w, H_ROW, fill=0, stroke=1)
        # Color rojo en SKU y descripción si es huérfano sin DDM
        col_text = RED_ALERT if (is_orphan_no_ddm and i in (0, 1)) else colors.black
        font_use = 'Helvetica-Bold' if (is_orphan_no_ddm and i in (0, 1)) else font
        if align == 'center':
            txt(c, x+w/2, y+9.5, s, font_use, 9, col=col_text, align='center', mw=w-3)
        else:
            txt(c, x+3,   y+9.5, s, font_use, 9, col=col_text, mw=w-5)
        # Subrayado SOLO en columna Bultos (índice 3)
        if r['has_pal'] and i == 3:
            tw = c.stringWidth(s, font, 9)
            cx = x+w/2
            c.setStrokeColor(colors.black); c.setLineWidth(0.6)
            c.line(cx-tw/2, ry(y+H_ROW)+1, cx+tw/2, ry(y+H_ROW)+1)
            c.setLineWidth(0.4)
        x += w
    return H_ROW

def draw_total_row(c, y, tot_b, tot_u):
    mw = COL_W[0]+COL_W[1]+COL_W[2]
    rfill(c, MARGIN, y, CW, H_TOT, colors.HexColor('#E8E8E8'), BORDER)
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)
    c.rect(MARGIN,    ry(y+H_TOT), mw,      H_TOT, fill=0, stroke=1)
    c.rect(MARGIN+mw, ry(y+H_TOT), COL_W[3],H_TOT, fill=0, stroke=1)
    c.rect(MARGIN+mw+COL_W[3], ry(y+H_TOT), COL_W[4], H_TOT, fill=0, stroke=1)
    txt(c, MARGIN+mw/2, y+9.5, 'T O T A L  A L M A C É N', 'Helvetica-Bold', 8.5, align='center')
    for val, xoff in [(tot_b, mw), (tot_u, mw+COL_W[3])]:
        s = str(int(val)) if val == int(val) else f'{val:.0f}'
        txt(c, MARGIN+xoff+COL_W[3 if xoff==mw else 4]/2, y+9.5, s, 'Helvetica-Bold', 8.5, align='center')
    return H_TOT

def draw_espacio_asignado(c, y, cancha, pv):
    rfill(c, MARGIN, y, CW, H_ESPACIO, BG_GRAY, BORDER)
    txt(c, MARGIN+CW/2, y+10, f'ESPACIO ASIGNADO — {cancha}', 'Helvetica-Bold', 9, align='center')
    col_w = CW/3
    hdrs  = ['Valor mínimo','Valor entero','Lectura']
    bx = MARGIN; by = y+14; rh = 13
    c.setStrokeColor(BORDER); c.setLineWidth(0.4)
    for h in hdrs:
        rfill(c, bx, by, col_w, rh, HDR_BG, BORDER)
        txt(c, bx+col_w/2, by+9, h, 'Helvetica-Bold', 8, align='center')
        bx += col_w
    vm, ve, lec = pallet_reading(pv)
    bx = MARGIN; by2 = by+rh
    for v in [vm, ve, lec]:
        rfill(c, bx, by2, col_w, rh, colors.white, BORDER)
        txt(c, bx+col_w/2, by2+9, v, 'Helvetica', 8.5, align='center', mw=col_w-4)
        bx += col_w
    return H_ESPACIO

def draw_route_to(c, y, dest):
    rfill(c, MARGIN, y, CW, H_ROUTE, YELLOW_RTE, BORDER)
    txt(c, MARGIN+CW/2, y+11, f'▶  DIRIGIR PALETA A: {dest}', 'Helvetica-Bold', 9.5, align='center')
    return H_ROUTE

def draw_route_recv(c, y):
    rfill(c, MARGIN, y, CW, H_ROUTE, LIGHT_GRAY, BORDER)
    txt(c, MARGIN+CW/2, y+11, '◀  RECIBE PALETA DESDE: CANCHA I', 'Helvetica-Bold', 9.5, align='center')
    return H_ROUTE

def draw_footer(c, date_str, page_num):
    yb = PAGE_H - MARGIN + 2
    c.setStrokeColor(colors.HexColor('#CCCCCC')); c.setLineWidth(0.5)
    c.line(MARGIN, ry(yb-4), MARGIN+CW, ry(yb-4))
    txt(c, MARGIN+CW/2, yb+6, 'Almacen Digital 3.0  |  Beccacece Hnos SA  |  Distribuidor Oficial CMQ — Desde 1963',
        'Helvetica', 8, FOOT_GRAY, 'center')
    txt(c, MARGIN,    yb+16, date_str,           'Helvetica', 8, FOOT_GRAY)
    txt(c, MARGIN+CW, yb+16, f'Página: {page_num}','Helvetica', 8, FOOT_GRAY, 'right')

def header_height(is_first_rep_page: bool) -> float:
    h = H_TITLE + GAP + H_TRANS + GAP
    if is_first_rep_page:
        h += H_PARTIDA + GAP + H_CTRL + GAP
    h += H_LEMA + GAP + H_CANCHA + GAP + H_THDR
    return h

def footer_reserve(has_route: bool) -> float:
    return H_ESPACIO + (H_ROUTE if has_route else 0) + GAP + H_FOOT

def content_avail(is_first_rep_page: bool, has_route: bool) -> float:
    return PAGE_H - 2*MARGIN - header_height(is_first_rep_page) - footer_reserve(has_route)

ALM_GAP = 8   # espacio entre grupos de almacén (más legible)

def alm_group_height(n_rows: int) -> float:
    return ALM_GAP + H_ALM + n_rows*H_ROW + H_TOT

def draw_cancha_pages(c, reparto_rows, cancha, is_first_rep, numero, transport,
                       chofer, lema, fecha_str, date_str, pc,
                       pall_values, route_dest, receives_route):
    c_rows     = [r for r in reparto_rows if r['cancha'] == cancha]
    alm_groups = build_alm_groups(c_rows)

    remaining = list(alm_groups)
    pages     = []
    first_pg  = True

    while True:
        is_frp    = is_first_rep and first_pg
        avail     = content_avail(is_frp, has_route=False)
        used      = 0
        pg_groups = []

        for g in remaining[:]:
            gh = alm_group_height(len(g[2]))
            if used + gh > avail and pg_groups:
                break
            pg_groups.append(remaining.pop(0))
            used += gh

        pages.append((is_frp, pg_groups))
        first_pg = False
        if not remaining:
            break

    n_pages = len(pages)
    for pi, (is_frp, pg_groups) in enumerate(pages):
        is_last = (pi == n_pages - 1)
        has_rt  = is_last and (route_dest or receives_route)

        if pc['n'] > 0:
            c.showPage()
        pc['n'] += 1

        draw_watermark(c)
        y = MARGIN

        y += draw_title_bar(c, y, numero, fecha_str);  y += GAP
        y += draw_transport_line(c, y, transport, chofer); y += GAP

        if is_frp:
            y += draw_partida_regreso(c, y); y += GAP
            y += draw_control_carga(c, y);   y += GAP

        y += draw_lema(c, y, lema); y += GAP

        lbl = cancha if pi == 0 else f'{cancha}  (continuación {pi+1})'
        y += draw_cancha_header(c, y, lbl); y += GAP
        y += draw_table_header(c, y)

        if not pg_groups and not c_rows:
            txt(c, MARGIN+CW/2, y+13, f'— Sin carga de {cancha} para este camión —',
                'Helvetica-Oblique', 8, colors.HexColor('#888888'), 'center')
        else:
            for alm_id, alm_det, rows in pg_groups:
                y += ALM_GAP
                y += draw_alm_header(c, y, alm_id, alm_det)
                for r in rows:
                    y += draw_product_row(c, y, r)
                y += draw_total_row(c, y,
                                    sum(r['blts_pick'] for r in rows),
                                    sum(r['unids']     for r in rows))

        if is_last:
            pv = pall_values.get(cancha, 0.0)
            route_h   = H_ROUTE if has_rt else 0
            foot_y    = PAGE_H - MARGIN - H_FOOT - H_ESPACIO - route_h - GAP
            draw_espacio_asignado(c, foot_y, cancha, pv)
            if route_dest:
                draw_route_to(c, foot_y + H_ESPACIO + GAP, route_dest)
            elif receives_route:
                draw_route_recv(c, foot_y + H_ESPACIO + GAP)

        draw_footer(c, date_str, pc['n'])


def draw_sin_cancha_page(c, row, numero, transport, chofer, lema, fecha_str, date_str, pc):
    if pc['n'] > 0:
        c.showPage()
    pc['n'] += 1
    draw_watermark(c)
    y = MARGIN

    y += draw_title_bar(c, y, numero, fecha_str);  y += GAP
    y += draw_transport_line(c, y, transport, chofer); y += GAP
    y += draw_lema(c, y, lema); y += GAP
    y += draw_sin_cancha_header(c, y); y += GAP
    y += draw_alm_header(c, y, None, 'sin asignación en base de datos'); y += GAP
    y += draw_table_header(c, y)

    rc = dict(row); rc['is_sin_cancha'] = True
    y += draw_product_row(c, y, rc); y += 8

    rfill(c, MARGIN, y, CW, 22, colors.HexColor('#FFF3E0'), colors.HexColor('#FFB300'), lw=0.6)
    note = f"SKU {row['sku']} no tiene cancha asignada. Validar con Jefe de Almacén antes de cargar."
    txt(c, MARGIN+CW/2, y+14, note, 'Helvetica', 7.5, colors.HexColor('#C0392B'), 'center', CW-10)

    draw_footer(c, date_str, pc['n'])

# ─── GENERADOR PRINCIPAL ─────────────────────────────────────────────────────

def generate_pdf(car_df, api, ddm, fr):
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)
    pc  = {'n': 0}

    if 'Fecha Mvto' in car_df.columns and len(car_df):
        try:
            dt = pd.Timestamp(car_df['Fecha Mvto'].dropna().iloc[0])
            fecha_str = date_str = dt.strftime('%d/%m/%Y')
        except:
            fecha_str = date_str = datetime.today().strftime('%d/%m/%Y')
    else:
        fecha_str = date_str = datetime.today().strftime('%d/%m/%Y')

    stats = {'repartos':[], 'red':[], 'yellow':[], 'pallet_applied':[],
             'miss_bxp':[], 'no_fecha':[], 'sin_cancha_skus':[],
             'orphans_no_ddm':[],
             'total_pages':0, 'trace':[]}

    for rep_idx, (numero, rep_df) in enumerate(car_df.groupby('Número', sort=False)):
        transport = str(rep_df['Transporte'].iloc[0])
        chofer    = str(rep_df['Descripción Transporte'].iloc[0])
        lema      = get_lema(transport, rep_idx)
        stats['repartos'].append({'numero': numero, 'transport': transport, 'chofer': chofer})

        rows        = process_reparto(rep_df, api, ddm, fr)
        pall_values = {ch: compute_pall_value(rows, ch) for ch in CANCHA_ORDER}

        # Trazabilidad de cancha por SKU para el debug
        for r in rows:
            stats['trace'].append({
                'Reparto': numero, 'Camión': transport,
                'SKU': r['sku'], 'Descripción': r['desc'][:50],
                'Almacén': r['alm_id'], 'Detalle Almacén': r['alm_det'],
                'Cancha': r['cancha'], 'Bultos CAR': r['blts_raw'],
                'A pickear': r['blts_pick'],
                'Huérfano sin DDM': '⚠' if r.get('orphan_no_ddm') else '',
            })

        pv1 = pall_values['CANCHA I']; frac1 = pv1 - int(pv1)
        if frac1 > 0.001:
            pv2, pv4 = pall_values['CANCHA II'], pall_values['CANCHA IV']
            if pv2 == 0 and pv4 == 0:  route_dest = route_recv = None
            elif pv4 > pv2:             route_dest = route_recv = 'CANCHA IV'
            else:                       route_dest = route_recv = 'CANCHA II'
        else:
            route_dest = route_recv = None

        for ci, cancha in enumerate(CANCHA_ORDER):
            is_first_rep = (ci == 0)
            receives = (cancha == route_recv and not is_first_rep)
            draw_cancha_pages(c, rows, cancha, is_first_rep, numero, transport, chofer,
                              lema, fecha_str, date_str, pc, pall_values,
                              route_dest if is_first_rep else None, receives)

        for r in [r for r in rows if r['cancha'] == 'SIN CANCHA']:
            draw_sin_cancha_page(c, r, numero, transport, chofer, lema, fecha_str, date_str, pc)
            stats['sin_cancha_skus'].append(f"{r['sku']} — {r['desc']}")

        for r in rows:
            if r['row_st'] == 'RED':   stats['red'].append(r['sku'])
            if r['row_st'] == 'YELLOW':stats['yellow'].append(r['sku'])
            if r['has_pal']:           stats['pallet_applied'].append(r['sku'])
            if r['miss_bxp']:          stats['miss_bxp'].append(r['sku'])
            if not r['fecha_s']:       stats['no_fecha'].append(r['sku'])
            if r.get('orphan_no_ddm'): stats['orphans_no_ddm'].append(
                f"{r['sku']} — {r['desc']} (T{transport}, R{numero})")

    c.save(); buf.seek(0)
    stats['total_pages'] = pc['n']
    return buf.read(), stats

# ─── RESUMEN DE CARGA POR CAMIÓN (v3.9, hoja AGR) ───────────────────────────

def load_agr(file):
    """Lee la hoja AGR del CAR y devuelve DF consolidado por camión + SKU.

    Columnas resultantes: chofer | cod_producto | descripcion | bultos
    Una fila por (camión, SKU) con bultos sumados.
    Ignora Cliente / Razón social / Importes (no son relevantes para este doc).
    """
    file.seek(0)
    raw = pd.read_excel(file, sheet_name='AGR', header=None)

    # Buscar la fila de header (contiene 'Cliente' o 'Chofer')
    header_row = None
    for i in range(min(10, len(raw))):
        row_str = ' '.join(str(v) for v in raw.iloc[i].values if pd.notna(v))
        if 'Chofer' in row_str and 'Cantidad' in row_str:
            header_row = i
            break
    if header_row is None:
        raise ValueError("No se encontró el header de la hoja AGR (esperaba 'Chofer' y 'Cantidad').")

    df = pd.read_excel(file, sheet_name='AGR', header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_cod   = find_col(df, 'Cod Producto', 'CodProducto', 'Codigo Producto', 'Codigo', 'Código')
    col_desc  = find_col(df, 'DESCRIPCION', 'Descripcion', 'Descripción')
    col_cant  = find_col(df, 'Cantidad', 'Bultos')
    col_chof  = find_col(df, 'Chofer', 'Camion', 'Camión')

    if not all([col_cod, col_desc, col_cant, col_chof]):
        raise ValueError(
            f"Faltan columnas en hoja AGR. Encontradas: {list(df.columns)}. "
            f"Necesito: Cod Producto, Descripcion, Cantidad, Chofer."
        )

    df = df[[col_cod, col_desc, col_cant, col_chof]].copy()
    df.columns = ['cod_producto', 'descripcion', 'bultos', 'chofer']
    df = df.dropna(subset=['chofer', 'cod_producto', 'bultos'])

    df['chofer']       = df['chofer'].apply(to_int_sku)
    df['cod_producto'] = df['cod_producto'].apply(to_int_sku)
    df['bultos']       = pd.to_numeric(df['bultos'], errors='coerce').fillna(0).astype(int)
    df['descripcion']  = df['descripcion'].astype(str).str.strip()
    df = df[df['bultos'] > 0]

    # Consolidación: total por SKU por camión
    out = (df.groupby(['chofer', 'cod_producto', 'descripcion'], as_index=False)['bultos']
             .sum()
             .sort_values(['chofer', 'cod_producto']))
    return out


def _draw_resumen_header(c, fecha_str):
    """Cabecera del documento (título + subtítulo + línea separadora)."""
    margin = 15 * mm
    y_top = PAGE_H - margin

    c.setFillColor(colors.HexColor('#1a1a1a'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y_top - 14, 'Resumen de Carga por Camión')

    c.setFillColor(colors.HexColor('#666666'))
    c.setFont('Helvetica', 9)
    subt = f"Beccacece Hnos. S.A.   ·   La Reja, Moreno   ·   Fecha: {fecha_str}"
    c.drawString(margin, y_top - 28, subt)

    # Línea separadora
    c.setStrokeColor(colors.HexColor('#1a1a1a'))
    c.setLineWidth(0.8)
    c.line(margin, y_top - 36, PAGE_W - margin, y_top - 36)

    return y_top - 44  # y donde puede empezar el contenido (en coord top-down)


def _truck_block_height(n_rows):
    """Altura total de un bloque de camión (header + filas + spacer)."""
    return 7*mm + 6*mm + n_rows * 7*mm + 6*mm  # header + thead + filas + spacer


def _draw_truck_block(c, y_top, cam, rows):
    """Dibuja un bloque (header de camión + tabla) usando coords top-down.
    Devuelve el nuevo y_top (más abajo) después del bloque.
    """
    margin = 15 * mm
    total_w = PAGE_W - 2 * margin  # 180mm

    # Anchos de columna: Código 22 | Descripción 122 | Bultos 26 | Check 10 (mm)
    w_cod, w_desc, w_blt, w_chk = 22*mm, 122*mm, 26*mm, 10*mm

    # ── 1) Header del camión (barra negra) ──────────────────────────────────
    h_hdr = 7 * mm
    y_hdr = y_top - h_hdr
    c.setFillColor(colors.HexColor('#1a1a1a'))
    c.rect(margin, y_hdr, total_w, h_hdr, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin + 6, y_hdr + 2.3*mm, f'Camión {cam}')

    # ── 2) Header de columnas (banda gris clara) ────────────────────────────
    h_thd = 6 * mm
    y_thd = y_hdr - h_thd
    c.setFillColor(colors.HexColor('#f0f0f0'))
    c.rect(margin, y_thd, total_w, h_thd, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.setLineWidth(0.5)
    c.line(margin, y_thd, margin + total_w, y_thd)  # línea inferior

    c.setFillColor(colors.HexColor('#333333'))
    c.setFont('Helvetica-Bold', 8.5)
    # Centrado en la celda
    c.drawCentredString(margin + w_cod/2,                                y_thd + 1.7*mm, 'Código')
    c.drawString       (margin + w_cod + 6,                              y_thd + 1.7*mm, 'Descripción')
    c.drawCentredString(margin + w_cod + w_desc + w_blt/2,               y_thd + 1.7*mm, 'Bultos')
    c.drawCentredString(margin + w_cod + w_desc + w_blt + w_chk/2,       y_thd + 1.7*mm, 'Check')

    # ── 3) Filas ─────────────────────────────────────────────────────────────
    h_row = 7 * mm
    y_cur = y_thd
    c.setStrokeColor(colors.HexColor('#cccccc'))

    for i, r in enumerate(rows):
        y_cur -= h_row

        # Texto
        c.setFillColor(colors.HexColor('#1a1a1a'))
        c.setFont('Helvetica', 9)
        c.drawCentredString(margin + w_cod/2,                          y_cur + 2.2*mm, str(r['cod_producto']))
        c.drawString       (margin + w_cod + 6,                        y_cur + 2.2*mm, str(r['descripcion']))
        c.setFont('Helvetica-Bold', 9)
        c.drawCentredString(margin + w_cod + w_desc + w_blt/2,         y_cur + 2.2*mm, str(int(r['bultos'])))

        # Línea separadora entre filas (excepto la última, que la cierra el BOX)
        if i < len(rows) - 1:
            c.setStrokeColor(colors.HexColor('#e8e8e8'))
            c.setLineWidth(0.25)
            c.line(margin, y_cur, margin + total_w, y_cur)

        # Recuadro de check (casilla destacada con borde más marcado)
        chk_x = margin + w_cod + w_desc + w_blt
        # padding interno para que el cuadrado no toque los bordes de la fila
        pad_x, pad_y = 2.5*mm, 1.3*mm
        c.setStrokeColor(colors.HexColor('#666666'))
        c.setLineWidth(0.8)
        c.rect(chk_x + pad_x, y_cur + pad_y,
               w_chk - 2*pad_x, h_row - 2*pad_y,
               fill=0, stroke=1)

    # ── 4) Box exterior de toda la tabla ─────────────────────────────────────
    tabla_h = h_thd + len(rows) * h_row
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.setLineWidth(0.5)
    c.rect(margin, y_thd - len(rows) * h_row, total_w, tabla_h, fill=0, stroke=1)

    # Línea vertical antes de columna check
    chk_x = margin + w_cod + w_desc + w_blt
    c.line(chk_x, y_thd - len(rows)*h_row, chk_x, y_thd)

    # Devolver y_top siguiente (con un spacer de 6mm)
    return y_cur - 6 * mm


def _draw_resumen_footer(c, page_num, total_pages):
    """Pie de página: 'Página X de Y' centrado."""
    c.setFillColor(colors.HexColor('#666666'))
    c.setFont('Helvetica', 8)
    c.drawCentredString(PAGE_W / 2, 10 * mm, f'Página {page_num} de {total_pages}')


def build_resumen_carga_pdf(car_file):
    """Genera el PDF 'Resumen de Carga por Camión' a partir de la hoja AGR.

    Lógica:
      1) Lee hoja AGR y consolida bultos por (camión, SKU).
      2) Ordena camiones ascendentemente.
      3) Imprime en bloques compactos: varios camiones por página si entran.
      4) Ningún bloque se parte entre páginas.
      5) Pie de página 'Página X de Y'.
    """
    resumen = load_agr(car_file)
    if resumen.empty:
        raise ValueError("La hoja AGR no contiene datos válidos para generar el resumen.")

    margin = 15 * mm
    fecha_str = datetime.today().strftime('%d/%m/%Y')

    # ── PRIMERA PASADA: paginar (sin dibujar) para conocer total_pages ──────
    camiones = sorted(resumen['chofer'].unique())
    bloques = []  # lista de (cam, rows_list, alto)
    for cam in camiones:
        sub = resumen[resumen['chofer'] == cam].to_dict('records')
        h = _truck_block_height(len(sub))
        bloques.append((cam, sub, h))

    # Reserva para cabecera (solo en página 1) y pie (en todas)
    H_HEADER_PAGE1 = 44  # 'Resumen de Carga por Camión' + subt + línea + spacer
    H_FOOTER      = 18 * mm  # margen inferior efectivo donde reservamos espacio

    # Espacio disponible para contenido (top → bottom)
    def avail(is_first_page):
        top    = PAGE_H - margin
        bottom = H_FOOTER  # piso del contenido
        if is_first_page:
            top -= H_HEADER_PAGE1
        return top - bottom

    # Asignar bloques a páginas
    pages = []   # lista de lista de (cam, rows)
    current = []
    space_left = avail(True)
    is_first   = True
    for cam, rows, h in bloques:
        if h <= space_left:
            current.append((cam, rows))
            space_left -= h
        else:
            # Cerrar página actual y abrir nueva
            pages.append(current)
            is_first = False
            current = [(cam, rows)]
            space_left = avail(False) - h
    if current:
        pages.append(current)

    total_pages = len(pages)

    # ── SEGUNDA PASADA: dibujar ─────────────────────────────────────────────
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    c.setTitle('Resumen de Carga por Camión')

    for page_idx, page_content in enumerate(pages):
        is_first_page = (page_idx == 0)
        if is_first_page:
            y_top = _draw_resumen_header(c, fecha_str)
        else:
            y_top = PAGE_H - margin

        for cam, rows in page_content:
            y_top = _draw_truck_block(c, y_top, cam, rows)

        _draw_resumen_footer(c, page_idx + 1, total_pages)
        c.showPage()

    c.save()
    pdf_bytes = buf.getvalue()
    return pdf_bytes, {
        'total_pages': total_pages,
        'camiones': len(camiones),
        'filas_sku': len(resumen),
    }


# ============================================================================
# ║                                                                          ║
# ║        EXTENSIONES v4.0 — lectores MASTER, T2, extraíbles, UI            ║
# ║                                                                          ║
# ============================================================================

# ─── LECTORES MASTER (78 MB) ────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_master_matriz_pall(file_bytes: bytes) -> tuple[pd.DataFrame, datetime | None]:
    """Lee la hoja 'Matriz Pall.' del MASTER.

    Estructura confirmada por inspección:
      - Celda A1: fecha del día (datetime)
      - Fila 1 cols B-Y: headers
      - Datos desde fila 2

    Returns:
        df: DataFrame con columnas tipadas
        fecha: datetime de A1 (o None si no se pudo leer)
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Matriz Pall." not in wb.sheetnames:
        raise ValueError("Hoja 'Matriz Pall.' no encontrada en MASTER")
    ws = wb["Matriz Pall."]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("Hoja 'Matriz Pall.' vacía")

    # A1 = fecha del día; los headers de las columnas B-Y están en fila 1
    fecha_a1 = rows[0][0] if isinstance(rows[0][0], datetime) else None

    # Headers (col B en adelante)
    headers_raw = list(rows[0][1:25])  # B-Y
    headers = [
        "Camion", "Reparto",
        "UP_PICKING_AGR_FACT",
        "PICK_C1", "PICK_C2", "PICK_C3", "PICK_C4", "PICK_C5", "PICK_MKPL",
        "UP_AE_PALL_AGR_FACT",
        "AE_C1", "AE_C2", "AE_C3", "AE_C4", "AE_C5", "AE_MKPL",
        "KG_MATRIZ",
        "KG_C1", "KG_C2", "KG_C3", "KG_C4", "KG_C5", "KG_MKPL",
        "CLASIFICACION",
    ]

    # Datos: fila 2 en adelante, cols B-Y
    data = []
    for r in rows[1:]:
        if r[1] is None:  # Camion vacío → fin de datos
            break
        data.append(r[1:25])

    df = pd.DataFrame(data, columns=headers)
    # Tipos
    df["Camion"] = pd.to_numeric(df["Camion"], errors="coerce").astype("Int64")
    df["Reparto"] = df["Reparto"].astype(str).str.strip().str.upper()
    for col in df.columns:
        if col not in ("Camion", "Reparto"):
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    wb.close()
    return df, fecha_a1


@st.cache_data(show_spinner=False)
def load_master_matriz_picking_app(file_bytes: bytes) -> pd.DataFrame:
    """Lee la hoja 'Matriz Picking APP' del MASTER.

    Estructura: fila 1 = headers con A1 = fecha del día, 64 cols total.
    Para Sprint 1 (extraíble fx Picking) basta con devolver el bloque
    como DataFrame plano.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Matriz Picking APP" not in wb.sheetnames:
        raise ValueError("Hoja 'Matriz Picking APP' no encontrada en MASTER")
    ws = wb["Matriz Picking APP"]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return pd.DataFrame()

    headers_raw = list(rows[0])
    # Header A es la fecha → reemplazamos por 'FECHA'
    headers = ["FECHA"] + [str(h) if h is not None else f"COL_{i}"
                           for i, h in enumerate(headers_raw[1:], start=2)]
    # Recortar al ancho real
    n = len(headers)

    data = []
    for r in rows[1:]:
        if all(v is None for v in r[:n]):
            continue
        data.append(list(r[:n]))

    df = pd.DataFrame(data, columns=headers[:len(data[0]) if data else n])
    wb.close()
    return df


@st.cache_data(show_spinner=False)
def load_master_reposicion_ae(file_bytes: bytes) -> pd.DataFrame:
    """Lee la hoja 'Reposición AE' del MASTER.

    Columnas A-J:
        A: Fecha | B: Almacén | C: Descripción | D: Cancha | E: bxp
        F: Posiciones | G: En cancha | H: AGR | I: PICK | J: Reposición Pall
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Reposición AE" not in wb.sheetnames:
        raise ValueError("Hoja 'Reposición AE' no encontrada en MASTER")
    ws = wb["Reposición AE"]

    headers = ["Fecha", "Almacen", "Descripcion", "Cancha", "bxp",
               "Posiciones", "En_cancha", "AGR", "PICK", "Reposicion_Pall"]
    data = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if r[1] is None:  # Almacén vacío → fin
            break
        data.append(list(r[:10]))

    df = pd.DataFrame(data, columns=headers)
    df["Almacen"] = pd.to_numeric(df["Almacen"], errors="coerce").astype("Int64")
    for col in ("bxp", "Posiciones", "En_cancha", "AGR", "PICK", "Reposicion_Pall"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    wb.close()
    return df


@st.cache_data(show_spinner=False)
def load_master_agregados_ae(file_bytes: bytes) -> pd.DataFrame:
    """Placeholder para Tab 5 — Agregados AE.

    Por ahora devuelve un DataFrame derivado de Matriz Pall. con las cols
    de AE (K-Q). Cuando Lucas confirme la hoja exacta destino se ajusta.
    """
    df_pall, _ = load_master_matriz_pall(file_bytes)
    # Subset: Camion + Reparto + UP AE + 6 canchas AE
    cols = ["Camion", "Reparto", "UP_AE_PALL_AGR_FACT",
            "AE_C1", "AE_C2", "AE_C3", "AE_C4", "AE_C5", "AE_MKPL"]
    return df_pall[cols].copy()


# ─── LECTOR T2 Pall. Camiones (referencia estática) ─────────────────────────

@st.cache_data(show_spinner=False)
def load_pall_camiones(file_bytes: bytes) -> pd.DataFrame:
    """Lee la hoja 'Pall. Camiones' del T2_Status_Carga (datos maestros).

    Headers reales A-AF (32 cols). El archivo se carga una sola vez por
    sesión y queda en cache (referencia estática para Tab 3 — Sprint 3).
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True,
    )
    if "Pall. Camiones" not in wb.sheetnames:
        raise ValueError("Hoja 'Pall. Camiones' no encontrada en T2_Status")
    ws = wb["Pall. Camiones"]

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h).strip() if h else f"COL_{i}" for i, h in enumerate(rows[0])]
    # Renombre clave: TARA → CAPACIDAD MÁX (decisión cerrada handoff #9)
    # En esta hoja el header ya es 'Capacidad Maxima (kg)' en col AC, así que
    # el rename solo aplica al PDF rediseñado del Tab 3 — la fuente se respeta.

    data = []
    for r in rows[1:]:
        if r[1] is None:
            continue
        data.append(list(r[:len(headers)]))
    df = pd.DataFrame(data, columns=headers)
    df["Camion"] = pd.to_numeric(df["Camion"], errors="coerce").astype("Int64")
    wb.close()
    return df


# ─── EXTRAÍBLES SHEETS (TAB 5) ──────────────────────────────────────────────

def df_to_tsv(df: pd.DataFrame) -> bytes:
    """TSV: pegar directo en Google Sheets sin separadores ambiguos."""
    return df.to_csv(sep="\t", index=False, lineterminator="\n").encode("utf-8")

def df_to_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(sep=",", index=False, lineterminator="\n").encode("utf-8")

def df_to_xlsx(df: pd.DataFrame, sheet_name: str = "Hoja1") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df.to_excel(wr, index=False, sheet_name=sheet_name[:31])
    buf.seek(0)
    return buf.read()

def filter_reposicion_negativa(df: pd.DataFrame) -> pd.DataFrame:
    """Tab 5 — filtro decisión cerrada: solo filas con Reposicion_Pall < 0."""
    return df[df["Reposicion_Pall"] < 0].copy().reset_index(drop=True)


# ─── SNAPSHOT LOCAL (Sprint 5 → Drive) ──────────────────────────────────────

def make_run_id() -> str:
    """Run ID único por corrida (timestamp + hash corto)."""
    ts = datetime.now().strftime("%H%M%S")
    h = hashlib.sha1(ts.encode()).hexdigest()[:6]
    return f"{ts}_{h}"

def snapshot_local(run_id: str, files: dict[str, bytes], log_text: str):
    """Guarda los archivos y el log en ./snapshots/YYYY-MM-DD/<run_id>/."""
    fecha = datetime.today().strftime("%Y-%m-%d")
    base = SNAPSHOT_DIR / fecha / run_id
    base.mkdir(parents=True, exist_ok=True)
    for name, data in files.items():
        with open(base / name, "wb") as f:
            f.write(data)
    with open(base / "run.log", "w", encoding="utf-8") as f:
        f.write(log_text)
    return base


# ─── HELPERS UI ─────────────────────────────────────────────────────────────

def _stamp(prefix: str, ext: str) -> str:
    """Nombre de archivo con timestamp para descargas."""
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"

def _download_trio(df: pd.DataFrame, base_name: str, sheet_name: str, key_prefix: str):
    """Renderiza 3 botones de descarga (TSV / CSV / XLSX) en columnas."""
    c1, c2, c3 = st.columns(3)
    with c1:
        st.download_button(
            "⬇ TSV (pegar Sheets)",
            data=df_to_tsv(df),
            file_name=_stamp(base_name, "tsv"),
            mime="text/tab-separated-values",
            key=f"{key_prefix}_tsv",
            width="stretch",
        )
    with c2:
        st.download_button(
            "⬇ CSV",
            data=df_to_csv(df),
            file_name=_stamp(base_name, "csv"),
            mime="text/csv",
            key=f"{key_prefix}_csv",
            width="stretch",
        )
    with c3:
        st.download_button(
            "⬇ XLSX",
            data=df_to_xlsx(df, sheet_name),
            file_name=_stamp(base_name, "xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_xlsx",
            width="stretch",
        )


# ============================================================================
# ║                            UI STREAMLIT                                  ║
# ============================================================================

# ─── LOGO BECCACECE (base64 PNG con transparencia) ───────────────────────────
_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAfUAAAH1CAYAAADvSGcRAAAACXBIWXMAABcSAAAXEgFnn9JSAAAAGXRFWHRTb2Z0d2FyZQBBZG9iZSBJbWFnZVJlYWR5ccllPAABG4lJREFUeNrsvQl8VOW9N/5LyL5M9j1ACAESIhAUErBqwmZcagnQ3i6ihPftp7UuiLfvLaLeFvt3o8stolK9/95LUOzrbUsIVlRkMYpbgktAIUFICJCE7Mtknazv83vOnMlMMsuZmXNmy+/74TCTWc7MnPOc5/t8f6sXEAgE10Rmfi77v4BtWWyrYFsRnD1yWoHP2KX9DESn9u9d7LO66CQQCO4FLzoEBIJLEvpe9n+hkWeQbB+R6TN+w/7fYeJZXETkEbETCETqBAJBGUIXgYp9s52fsYnvxzxK2Oesk9EisJVt4dpHavmC4uyRy3TCCQQidQLB2cSLpJiiR36nZdrvwyCYvy3BdsWemb9Iq8SloJB9zj6FFimdWmvAaRpQBAKROoHgDDJfxEl8nNBFlGoJ8LId+w7TKthwie9AQvzAhs95n79XGmrZZ8xS0OqAxJ5Fip1AkAfedAgIBMkENVNL3ilGCRbVr0D6tqLQCkIHiYp+4m/ItYLQgf9WwSphy/H6kwVCB+3vLaLBRSAQqRMIjkaRBdIN56RvO7FvtfL1WVpzvbW/wVrssHHxIPX35GlfTyAQiNQJBIeodKkKF4m9RGtKt3b/KTYRrtRFhGAKt+UzUqxaPAi/vcTKzyiU5Rzhb8zM72DbGNu+stnKQCAQqRMIHg1rVDQSp7Wm8V02fi9p1gGBlO0hTmmLB4HQS8E6NwKiwE5C/xOIcQ3jn53FLRPCYoZAmBKgQDmCJypqsWBLp1Yxltidb43KzxYiPHvkSQn7NpcvLhWd2s97fsK+Z2oXDAUyHN0Kvh9TQW3jhJ5l4/5x34esPC+iVSBP9n0TCETqBIJTCd1U7nUnCJHph2zc71qw3pwswnxOubR8cWvJvRTE4jHWBcVJ3f/WSWluwu/YAbaZ90WUsv2usPLcWIquF2FfFD+BQKROIDhcoZdaeJVtud3SicOcwt1hsKiQV0E7A50wnuuOyjxcpv1KT9Oz/rzYn3NPIBCpEwgOIfVLElWiddXYBPKtlZkIU+xUtJ6MWhDy1rssnJc/gfXZAvJVyCMQXBQUKEfwBEJfawVJFloZEb1Dxm8aDoI5nAjdNPDYlGoXU8bOdRjbDtpA6ADyuyIIBFLqBIICpG6LedxyEJv8/m6CdIjd4kpBsG6gib8ArC/QM3nRYG/1OsHVIwZillI1PAKROoEgL6lLNb1PRCmYKu0qvQY7wb1gu1/ddDCgfJ3zCAQidQKRuk3pZvooAcOIcTEljuB5sCXCXkraHBE7gUidMCUJOEyPMDvt7tAlLeqdQNCHNRH21uTe29Zgh0CQERQoR3AkoaP5slY7SQrKGE3nQvEVW1FAB5ZgJXZJHK9ii1qpVptCOrQEInXCVCL0Ipgc5JQCQgnSr2xshEKkTrAWWRZLxwoZFaVgXawGjUWC00Hmd4IjCF1qn/BOEMp5SjWNUnQ6wR4YKwqEZI7pcnk2LxjsdSkRCETqBDdR6VJhOUJZ+kKBQHAkyK9OcCrI/E5wBKw1SxbxAiOmC5Cgmb6UCJ3gkqROIDgRPnQICA6ALelhBXzLzBe6rI2rcrEACYHgiqCFJoFIneDxSLFT5VMAEmHqKHUhTVMAmfIJVoJ86gTlYX9xGALBvRax1paOFVxNO7QLWH21b7zVLYFgAuRTJyhN6Ll0EAhTDIVWXiNYkrgCjNe1x7+LtFH5BAKROsEFVAuBMLWwQ1LNBaHjHObL7wLLvnjqQ0CQBDK/E4xNNjgh5Wn/qrDLr2dbBzUCwd3RCUJ622kT10UuCGme1ix6KV2OYBEUKEcwVA7GGldk5ottMLFpRZeVe6UgN8JUBCpv7AtfxK8pJOPxlq2FYFtGCF6XROoEUuoEyaT+lYXJphOEwjCHJO4Pa7rvoANLIMgCXByso8NAMAfyqRNEAt4kQT2E84klM/9PWlVvbn9owt9KB5ZAkFX9EwhmQeZ3gghrzORbuWLPzEcVXjTJJC8sEKQE/xDsxKJ5qRCuCjH5fEpiHKQkxRk8VlvfBLUNTSbf06nugdPna+jgEghuCDK/E0Qi7rCDhEtBqPiWorcRrMRMLQGHhwZDVvpsQZrp3UfkLV3olO9WeuqM7n5FVTV0dvcaPI4LhctmFgoEeU4DW0CvoMNAIFInSCF1KhDjINLO0qrribeeAFT5FUzlT7wl0pcFmImymA4DgUidQKTuQKBJnJM3U9hZevcJgrIXiR4VP5I9mfqtwNkjNGcTiNQJROpKIExrGkeTOPqus9JTibztIHskeCR6JPwP9Mz9BANQv3YCkTqBSF0O5DLyRuWtI/IJAWgEecEJvqqGEz7eJ0XPgbUiHrHy2sZMFbExkhg3UwtCkCvlvROpE4jUp4YKR+IWCdxZAWqEcaDZnhP8eYHop6ia79Sq9csSrmkk863azVQQbCE1iyFSJ3geoWOlq1Ii8YW6zd3M6IODdTA0VA/TvEMhIHD+lDlvSO64lRz/ZCopeWz+kme2uqPQAAbTSlMk7M/6rnIEInWCS5M65pUXTbWfjQFtBatudHklPsQIe2RUDQEBpsm6pXk3tDa/AGNae8vMWfshOCTH4r77+89Bc+ML/LUBgRkQIuE9rq7kS058ym89PNq+FoSWrIcmKPMCEMrQ5lmxLzTDb6aJ0DNAxWfcn5CFblD2Bc9MifrsqMYLVgokXrByuculkbW17ob2lt38PnLzjFlvciLv6iqGNu3jiOkzX4eg4BzTy3QrHCm9PeWg7jrGPuMY/zsh6XGIiSmEnu4yaGzczUmeb6GuT/Z4PnGRhhuXs1XVOpL3QFM9KnCs7sgNFmBffYgsIBCpE5xO5r8BfV/Z+MVdZJWPLDN/pieTOuaGI4EXFqxxuEldozkHoyPd/P4oU9rBIWssv0mPlMX32oJp00IlknqZwd8hIdn8tqvrKCf8nu5y9tcLkJK6B8LD17jVuecphWzbes86ruKR3EtOfMKJvktbPMdDkGfvoaIJlUid4FxCN9XONI9vQvnWHRbJfbwrm0dBNKsjmTuCyPt6jsLgYCW/PzKihqiYJ/j99panob+/TEfSqXMvKvcdeg3JWapfvQdJXW8hEah9X39/pcEiIygwQ+JCpg4u1z4D8fGbQBXmOuoeVTwu7HBDoA9eJHkPI3gCkTrBzQg9Fyz3J0/hit0cuQsKvcRTVumiIkdV5uhUMyT0zvZxs3lQ8BoIDMoBH58kg9ch4U+bppK8X2+t2tYMnDN43D8gQ6bvXQejo9067g7SqnRBwZfr7vv5JYGff7KkfTZe2wft7cf4hu+bnfYc9PZWwvBwNyQmbmLHROUS42XcTP9LHcHvO3SU5hcCkTrB4bCm85lI7hgFixGzpQaK3kOI3BmmdbPQhp/6+BoS4aCmkpO9KfT3TVDb2sC4iWZ4YwuDuPjHYURL0AOiypaA4OBsRroCgYtBcj16Jnn8KVL96cPDamhpKQYv9iYM2NNo6sHfPwkuXngUBtj9q1degJjYdTB37k6XGkciwe/a9nOdej/EbgkEInWCI2CLsg73FCJHbFq7hpO5GBAlNzT9ZTAwME5swSHrJxG0BT63Caqw9QLpTwh0i2GEjcQ+wBS7qKwnQt/cHhwsjYT9/JIhNe11YUHRf47/LSIsbDV0M3LHzw3RU/Dm0NFxlFsjxKOgCssGtbqck7vOAtBrecHR318HgYGGx7u65gXoaC+DGTMKITZ2tSLnXd9Ej9XtkNx3vVYyFWrWV9C0SqROcB5SpuKPRj85mtYdEbWuYYSu7hyPNvcPyJFM6vZAFbbBOGFrFXtQsHI+6kC9RYEY8S5YDwzJ3hzq614w+DshYRNcazD0/CQmFlrcz6nP10JIcAZ7/3q+Ia5eKYKhoW5o7zjFjkcSLF9WAr6+ypny0YWD4w03jKJHcvdg/3snTatE6gSC4hBT0LbeU+Bw87o1mWH6vnO8jwVgODmq1jMFLZCjn3+GVf50V0FgkPRCNvHxhdDYWMSUeQM3u/v7J3OlLh5Pb59QiItbb3Yf164Vc/97Z2c5dODWUQbh4Tn8MQFjbAGSpCihTwSOvaKnfwmdasE8v+u1g1SylkCkTiBIBfrKd9y/UTFVPjxcB/09xVqKEJQ4brYiRLWBbxPhy9S9rwMUvqsgPqGQb+3tR3UBcRGRq3nQHB5oKSq9rq7IYEEVHpHDVbr+SisxcZ3F/QwNqeHsuWdgVsomCAuTJ7BQ3zyPaXFFJUc9Jbguj2YdInWC81DqqRfhWm30utLV3UaG66G7azxa3dv7cbOk7u0dSqPOCkRGjuezp6fvgYGBOm6Gj401T8adnWXQ21s1biUISILQkAzo7jF8LClxvcXvcKl2H1ytOwhXrx6EyKhsmDfnQYiOls99IVYhxOC6XftLuHp3a9M8FrGi7m8eAW86BG4Hj/J/oYkdA98uHdkHJbt/45ByrVgIxkBR+5k3Mfv5z6dRZwcCApJhVurj/NYcGhuLDf5Gfzr6z+dnPAcBgYKLI1ECoSOQ0MWgxba2csUKYqN6R6tS56cHYO9Tv+RWJjdFIY1UzwDVfne/FTVWktvhCWTOA5E2FihiYm9v/gkMaoQIdkyvSphRrXuuu3O3gVKPjnvdLvO7s4ABXJ0T1CFWTquw0t9rbCGF/eGd0VoWif3q1SKuzm++6XMD33l7RxlX7pb86VfriuH0me385OL5VYWmQ+4thwxeg+b5ifv55NP7+WMzZ66H2Jhsm38DmuZ37NnvjqVppfVqN2zlmgeT27k+SRO180Dmd4JDIfrLxcpehMmEoE/Y+iSN95UI0JIyA2P2gbj4ytLeD2cLMzGAUS4LS3z8er5hWttE0o2MkLbwqmMqXR+pszYZ/N3XVwdHjq6ChPhVMHv2JoiJzoHe3jpouHacLwJqL5dAdHQ2rMh91abfgMeidO/v3NHvXsoIO88ksVtu5ZrCBYewjxV0NROpEyTOGVOVzAc1n+nue3upwMfPNrP4NJ8k8PHNgOGhSoebqjD/ubahSUfaIom7uqrTX0yY+665jNBEshdvs/QWBFIxMU/dGqhU6aBWV/IUOF+fUJg+3dBkf+WqQPrXGo9zIs9IfxAGBwWXjBikFxycZP+FqvW747hH5e4G5I5EXcFIuQiEapOdOgUvzDsFkuco7PxIfdqdAjK/uxsy8ztMrJI9gsz7e/fC4MD45BcW9Vfd/ZYGbVrbGPrBcyA85q8m96NvfucKcHq1Q3+3qLCRtEUi98BOYZIRpkfwaNbH+0rGT6B5vbHpGCf2iUr98NtLYGi4m7tlcFu5ogQ+PHkPDLLXCiZ7L7gj/7gsxD5xQecm5C6P6ie1TkqdYJHQc92F0G1V5kOaozA8WGZ8sPowdT0sNk6pc5nfKvbx1ifxKVCFzCpgZDguaiYubHCcZKWn6kheLqJH0/305MlBdVeuFnNCFxETnQ1dXYKq139MbkJH4GIG893dSLnbA+r8Rkrd4wl5LQgmLBzstdqtxKo0ksz8r1z9YrHXzK5u/wkMDY53NotKGFfYXa1Mfes9F5NUbUbxH4CR4XHiDwl7WLbfKPbp5kReVUMELjPQf69P8nIG7JV+sBY6u4QUOVTpOdkvQVhYOly+fBAuVu9j46sbli97CZISLZei7eiohIgI23PgPV65nz1C/EKk7pFkLrY3zTN1bQNGjALsYhdBl5l9YFOWQlf9mWhe3bXtPrsD4Lo77oNBptbNkjqfkc2TupwQCbzUiNIkOGahKBK8vSSPZvnqmn1w5cpBPsTy15wwvBgvF0PKTMtpc1Xn98EXXz0LC657EBayzd7x5abR8kTqROpTktT3WkHGJdqtVu8xVOYYbZriqmQuZ2paf8/z0NezW0fqqsjXwdd/mUNJXV+J40b9tl2X5O2pOohR8EFB1gfk1Vw6CJ+Vb+dKHzErZR3cuOxZWRaPW5972VNK0NYyUp9Fo5VI3RNVusc2S8CiMWhql9M8quk/wDdxwgxWPQE+vkKU+/DQOV3hGG9vle5xOSD21MaJlczp7gU01/P2qYzgle4RgIRedmq7LsgOER6eDnfeVmLwuitXj0F8XDb4+Vlfox7T4FC5u/k4pEA5InWPJPVcGO9h7jHI1ZbHdKke5lYCg9uod7Znqngkd67iZW7Ni2lv/zy8ikfJi6QewQh99cpXDcj7YvVB+Piz7fy529a8ahOx4/h08/KzO6gIDZG6J5K6R1R/058wkcyV6mOuNMQe2aiEqMuW50Ps8ockL9eY1Vfqvr6hsGrFqxCpFyx3seYgfPLpdsF7hKQfkQ4rc1+CkJAkm8fs1p0vu+PCs4CR+iEahUTqROouit/cv9FuvzkWjxnSlMHIcCU3q0/zUb6DGSqeokNHiciJ4GUj+Lr6YzxI7pbvvGgQ/d7eUQlvvbNOV55WJHVjav2z8mdhcdYD4C9RxaNbqPDxP7qTSR6ryn1AI49I3dNI3e3N72hqL3rql3b7zXvUz0Nf924+4Ly8MADuz+AXcKtiRE6mdYIpiCZ6zNSQ04VU/OYq6O2pF0zzICj5791RMkmlf/jRdvj2YglERqIvfp9kYkegr91NTPJE6kTqHknqiwDLLrqpskEyt0bVDA/VMfLeC0NDlbyamypsC4SGC/nhmv73oKv9FzpSDwrdwraHZf3OYq1tJHOKWCdIAQbZ8R7pa9fYnb3Rwwi99OQDTLFX8fK0+atfZcRtmMf+4cfb4cLFg4z4hak3MkJL7P7Sid1NTPJE6kTqHkvsblXWFfHwxgIe1W7tJIek3liXy0cVEndAQA5ExQmlXLEQTFtTrjDgvLCdaY5BCVh7VDma13e9VkJR6wS7gNkcSPD2VLXDYLr3P3wQlt6w3cDXjvjq9It8G4+c94I5swsg92bb0uEwY2Przldcddxj3Y1HaFQRqXsiqWP3iAJ3+KpolsQylvZMavWXs9iE1c0H1rRpoRA/fdxQ0dKQxWazbj7qvL1DISrediOGG3bAIrgJ8DrYek+BLOpdhIaR/d+LV/GKdWIqHPrj13/voN2LWjTJP7+/xNUOI+WpOwnT6BAojNi0dHCDzmoYCFey+zd2+841/R/CyEi9drU4CIEhG3hOOWJosIIp9hqu4n39FoNfwBp239+qCeyNdz6Agi2/5ZMYBb4RlAC6bo58/AW8/LfDUHWpjl8T8dGRdu3TZ5o/JCXeDPX1H3FiF8zur/LH7UGAvx/cdtMSod2raxVKCmdzH1vJV1tngsc4pNi0FLZ1svdqaDSSUndFpe7SEfDoU0R1LlfAkLrjeVB3jQfEhUf9DoIYsSPQBD82qra6ZSr6EAUT+0HylROcAgwYxewPeyPnUbGXnXoWcpZutypAzo1Vu/l8dSHuqBAEa2bKhGfRlLeVfPNE6kTqVqhz9J3LCU3/Z9DZ8RT4+WUwNT4fAoLWgI+NqWtYrhV95WRiJ7gK7G1YZNUCubueEX+oVUF0CBdMf6sFofx1hfZ+npbA80Ba+WsKuiNSJ1K3NDHZ6ztXEh7b4EJm5agP7FMuh/8XF1KdetYQaiNrHHL3PJiI1rZKePPwJkhIyIbb17zoKard9kUB+eeJ1F2I1F2qXSpG+GJVOEsT0ciIGgYGKqG/rww0A+cgafrLk59nqnx0tBtCVRuIzGUma3HBlcIWYGKcg7MXYUjwtVqCL9WeI3ERgLdT0TWiBLmf//YgvH/yMV1A3YLMe+Gm5dtt2hdGyBc+8UdPODdZVrWpJlInKEToLlN8xtq888uX7oY+Ruiibzw1rRR8/QQz+uWaXKFXOXvc3z8Dkma8RWRuBTCOAYka4xhE0pZLaTsbIsnjORUXAVPhvMpF7p989ix8ffZVXUU6vEVy3/ijY6AKta3ULKr2god/6+7ngUzwROqykjN2W0uxaqUovKfUFVS6LVXhWpp3QyvbvLQ550nT/wwhoYIf8WrtXTA4eE54jv09a45tLVA9ncxxokfiRoWdpUfkUxEiweM5R+KvqKrxSLO+PeTe3V0PBw6tA42mW0fqvn6hcNvqFyEpMdvgtZXnD0LGvHVW7R8DTR/Z+QqROpH6lCXymSD4wgsnzk9asi4x2bBAUOi7XIHQbQ2G6+0tgytMrYukHhFZCLHxT/Dnmq/9ik1AB3QqPiH5dQgIXDblyTxXS96cxFGFy9iS1hOBCrLifI2O6D2pd72t5I6+9H++s4kTe1RkOqy45RmIjjIsYnPs/e2M1EsgfV4BrFlhXeEaPM6YEuqGCyoidSJ1uwkdIzWlVIIr1RJ9rZbEs0BaRKfikwrmndvqh0Wf+beV1+tIPSg4B6bPfJ0/p+46AC2N29jjY/y5qJgnQBW+WZJaQ/+eJ5A5Hl88tiKBu2rQobtBVPFI8J7Q296WaPnzFw5yv3o+U+j6ke8ajRpK3toELa1VOn979pIHIGfJg1YvpvA6dLO+CGgpvUxXCJG6raTuNlXgTCnGkud/bbeP9uL5XBgerueV4QIC5+tIHQPn6i5/T0fqQSFrIC7hZbOTCJazdPfUtLXaPt0ikROUh9guVyR5d1Xy9macoIJ/570HQd3doCN0vI2OSocf/8C2qnRuZI6n6HcidbtJfcxdvzrWbd/16H2y7Gug/xwjdJUuQE4f7a27wT8gA/z8M8DXN9kkme/aX+K2RWPEbl44EbtrD3lPAxI7RnTjrTtWFMQFN2afWLsoRFJHlc797drZKX1uAdx84/ZJeezNLZX8NjYmQ5JlxA3M8VRHnkjdLkJ3y3apaA7ete0+hxTFkAIXbzZhEmLXLuy9TT5x91Hx7tZi15amSfUN5ZzYkdQXLbiXE/pEnKs6CEeOP8aL1ny/YJ8kYneD6HjypxOpTy1St6XUa2fnUehoOwAzZv4OpvnIV64SV/5I5u7kN0ezOpI4qnJPSCubikBiKmHEjiTvLgSPC3EkdgyokwqMekcYi3z/4KNn4aszrwpKHk3z0enwg3X7IEBiRbqtz73smsVqzh4hniJSt4vUMRWt012+rrX+886Oo1BX9xQMDtZzf/jMmTshMsr+4jHuVsEKjxt24SIiJ4J3hUU5uszsCbZ8650HofrS8fH8drb5MbWOpC5FrYvAzofYq92F3GUVjNQX06gmUreX2N2iB7ot/nMk9erqX4xHtQdlwLz0f9r1PdzF1C621CTT+tSBaKJHsnJ1H7wtJnkRR09sh3NVJTpSR5V+68pnrCJ0HYtWVUPe5l+5CrGXMlJfQSOZSN1eUt8Lk/PTXQp7n/qlzf7zb87kwuBQva54zLz0NyEwaL7V+8EJE1f1rqyG0MSJJI5kThHrUxtigyAkeVcN3LS28qMITHc7cGgTNLdWQUxUOvenW9sIZqK1A4ndBRZCFPluJaifujFgH2CAH7nqRf/ZX3fxHsq2YnikG3p6xkvAjo1pICzcugUCRrT/6N+edVn1gybN5x75X/Dyrx+CH92RZ3c/bIL7A8cAkuV9/3InpM+azhelTW0dLvUdNYND8D/vfsAXIMsWZkhW7T4+/jA37Q4YGRmEu+54kf9tD7BP+49uz4XG1g5nX+PYl30X9VYnpS6HWnc5E7xcvc81mjo4+3UeJ3R//ySIT9gi2a/uygVkSJUTbFXvrlhDwZZAOiWAsTJPss2ZX8FsT3YCkbpEUnepAjRI6KV7f2e1vw0JvKWlmEe7X3fduO+8rfUA+PknQ2hojlXqHC9wVzNdoq8cXRFKtcEkeD7Emgroe3e12BBbejdIRZe6Hj769CVYML8AZkzPNvk6PC6b2WLeWacHhLQ26tJGpG4XqbtMH3RL7VKHh9XQ1nYUGhqKYMaMLRAVJZjS667uhrr63TrfeWrqToiOsT7S3VXVOS50UMW4Sm4+wTOABIYLWFdyLdnqazeHz796FT7+7CUYGOgGlSoRNt9dDAEBpv3wWA+gYMuTzlrUI7EXmuy3IczZWN47hS8ABFSAED0/pcrLEqmbHiBrAZu2uACho8ndFHp6zsGZM3fDyEi3UIc9ajVkZAglW1uaD0B1zTad71ylyoH0jNet+nxX7MeMygXNkq5eb31oqA561MUQolpvsuoewXXhio2HsKYCkrs9FilU5++89xhcqTtlkAK3ZPE9sHqF+b7tLhAZX6udl/XTji313CjVLgimBLkTqbswqUvpsIYqvbw8V0fquN1wQylbcQsk8vmpxQbPLVxUCv7+lgnGFZs+4AIHVbk7NE8ZHVFDw9W7eD2A0VGAmPidEBa+ga4rIndZVLs9zZq+OXcQ3jn6uEENeby7JMsyqYvEXvj4H92tTG8nCE1hujx9vFL0uynEpmH0e56zPh5T1qQEyHh7+0NfXw309lZy0kb4+KggLEzwlQ8OtUJvT4XwnNYMHxZ+i4TV+DYoO1PlMmSOkxgSurvkl3d3/5Ud98PChMm2nu5j4OObxBZb8+nacjPgmBMXk9gT3tk+d4yQx8A+XHjbkgWDuetX68qZYm/Q/p0O/7LuFVh4nbSAPMwiwMj4dz/63OWyB8wggG0D0FLt8eVmvemSNQmnBclZm4OelFRo8Hdz8wHd/YQE4TlciUeEr4bwCPP7RV/i4u8/4BLBQmhmv3RkH3c/uFuxmJ6uIp3bQ7SHXavfxvvUE9wTSOoYrPo+23JdwFqE1RuzNtzPY16sxe23PsPH5XeWPQCbNxZDXKx1RWrQ/I/HAuNa3OkUToVxSuZ3Y8jM38T+L3L0x9rTlOXLL+/iaj08PAdiY9dDXNy4qbel5QD3p5szu7tSUwd38ZmbwuDgOWiq/x43uwOEgl9ADvSojwl/e6kgbV4p735HcG+4illeriC6pmahu5s1BO9mvdk74eyRCCL1qUfoa7WEHu7oCxNXvrbmV3d2lkFAQJLOl24NXKX9ohw1sF1Cpav3Qkfb05zEw8K3QGh4IdRfuRv6+yvZY14QHLIaZqT8WZhIG4vA1y8JIiMpgt+dyR19zM6+fmxtuzygUcPHn74Ep758jZvi/9c9xVbvA3+/K+b6TzxVU6HkrOeSupDeUAiGJhcMlqgAIRqyVhcNKTRxwddtBSeYaOwldHuA5vZHdr7i1FOFeeaozD0lNa29+T7o6xWUeXzym+DnPx9GRtRw6WIuDA11c2JPSd0PbW3F0NxczB7DOIhkSJn1GMTEELm7K1yhGYq19Syu1JXD2+89Bl1dDdxFh/EfNy1/AG5mmwcSewmb89d5+jj0TFLPzP+TlqBdHs4idDSbYRMWZ16E+NsxGNATisZ0d/4b9PcWwzSfDBgarOeleJHUZ6ReHFdEA+fgyqWNMDSM2QihnOCHhwG8pyVDb289e5y9fsZDMHv2Q8SQbgqxiI0zK7BJjY6/UH0cDv5TGGtiBDxou7v964O2xX64OLFvZaT+vKePQc+Lfhf84c95EqEPDakZQWhg2jR/WT4XA2tuu+8JOPLxF0777ZhvixMP+gGxzrQ7Y6B3L9te4RPiyEgrr6UPY158ogwMWs2IPoa/zofdtjT/nj83OjrIn4+MWg/pGfvZa0Khvf0kU+/l0NtXD3Fxq4kh3RA4lvO0bX0rztc4xSQvRsenJMaZnVuiIlM5sff2teoemzN7JWxY++KkIjRln78K0ZGzLNaUx+sZ5xcXTXe7D1qqKaXN7RCbVsT+j/cUQr9YvRu+/mYrj6KOjMix+3N5o4i7t7LJptkpvxtN7Ujmj/70hx5R0nV0VA3qjvsAvAa5ShdIHZWPl5bkWyA45Lv8tX29ZdDVWcyfQ1UUEJgBs9P+mz+nUmXxOvwtLcegs6uKx0eEqTKA4J7AsS2mwaHP3RkmeQxeQ4I1F0CXkLAQTn/9Nzb2QuF7d/wRbr7xQQNCx+C5N/7xczhztgRa22sgM+MOi5/rosSOpvdXpsLY80RSf9kTCB3V+Ucf5/NJfoypup6eSkhO+rFdah19frczhY4reWcAi+kgoXtSL/PurpdhWHOS56CHRR+AoOCfwKCmTCB3wEj4Gkb2SWzSnM/O4VFG7CcBvV5I/NNn7gI/v/HAxtDQDE7mzc3HoaHhOMQztR4QEEMM6cbgOe5MtaOCd0aUPBIrVoXEvHJjFrGQ4BgeHLcq71GDqPeBATW8f/I/4J/vPq5T8q1tl9hr0iE6ynIamwsSO6r0KVFRzhNJfYcnKHQk7+bmo+ziqte2Rx2EwaEWdgHaFki19bmXYfuuvU75zbna/F4561a7AjD4rb3lYfBmKj0k/N/B128xeHurICDou6AZ+BBGhwXV3ttzjBN8X89Jtlir50rdly0CEhKfmLRPJPb+/nroUlfBtcaTMGP6OtncLgTnQDTJF6xcDlWX6hxukscCMW+88wH/DsZaEKMZfqJZHQkdo+EnBl3h69LnrJL0uS5E7BgU/chUGW9E6g7Gu688DcsWSTOrBgYmM8VWrCtggmo9MQHriEvPccbAHex7vu/QMacsYISe5ls8snuauuuvMNB3jKekqSJ+r3vcy8sfAoO/CwP9H8LIsKBy+vvK2KKsXldhLjx8PYSEGq/sh/70xuZj0N19CYaHB5liv5mY0QOAhIom+XB2XXx2psqhFjM0/yOxYx/59NTpFl+flLAIvjrzP3z8CQuTUFh5y7/Cqtx/tepzXYTYS5hKPzRVxpknVpQrddUvhpXirMnBjozMgcTE9TAmPjCGhWSkR5YioWPzBWcUhsBAuNoj+5zeC1pJdLQVcYIODn148oXFFHtU7J/Be1roONmLG/svOMR8fMSS618CP99QuHBxHzQ3lxMjehDwmqj4xx6HV6VDYl/38G+5G86idSFABXfe9gy/PzdtJfx0UzHkLLnXps/FipBOrjxXO5XGlycqdSwac5srErotediq0Ay4fGUfREZkQ2bm7yA+/ruS3ifWbz9fW+dwdf7G77fDjgfucfuodrOE3n6AKfVi8PNPgujY3xtfMU9TsecXQ293sZADLG6o1CN/YuBPnwi0xnhP84emppPQ3dMAs1LWAcFzIAbSOUO14yJfSt346MhUTujZN2wy25JVCpxcK37HVPGneyqpYxeSR92R0Pv66uBc5dMQxRS66EfFyT2JqfWZMzdzc7x0Qv+Vwy8gVB7vvizdveDOuHzpF4yduyEiqhACg5aZfJ2PtuXqQH+ZjtWR1LFjmzlS55aaiCxobS2HpuZTEBSUBBER48dVra4Hf38qNevuwGvlR7fnOTz9DZs1WYqMR2AgnRzABb4Tib12KjRy0YkJj/tFQmu9Ilf5OmK7UHPASPevTm+DE6UroK6+GGovG359qWSOwEhXZ/Q7/tO2n/NgOE+KbDeF1tYDMKCp15LzZsuqLHILU+zjhOzlZVj1qa3tAJwqnwNnv9nIW+nq44brn2WvH4Ovv3lR99iJD7ZD0eur4Wo9meU9AXjN4LWD15AjgbnsBVue5KpdDvQPqOHT8lfh0uVyk9YJ/J2Y1upgZE2l8eSpXdp2uAqhoz/JErrUlVBXN15v+VJtEfT1W282R18Z+swcSejoK/vqHy95tO+8s+scI97xClt1V3cLvvSQ1ZIbs8QmaP3rE8KJMYK+oe5p/rBaXQZnzhgSe3BwMsTEZPOCNNU1B+GDj7bDtxdK+MLgzbcf5HW7CZ4BvIbwWnKk/xlN8SgC7CH2zq56KHnrMdi1Zw28c2wnfHrqNdMLXEbsJbt/zd10DkQBZOYvmirjyDP7qWPVoNg0bJ3qtCI0aIrGnGwpCApK5qTRN1DP/8ZqY8PD3RAfJ90Hj4S++Yk/OvQ3YgMJoS1qvEdfJEdP5ENd/RvQ1YUmdDU7V4dhGlsOx8X/HAIDDfujD/QdgJ6up8AvYA2PgtddaIz88e/+3pN8QeAfMB8Cg7Kg8drvobennD82OobNNdqgtfVDbTEawfTZcO04qLsuQe2V49DaVsVeJ+S5Dw1i/YI2mDN7FRA8A2Kv8sbWDodFjKM5HM3ipnLZzaHk8GNw6PDj0NhUBcMjg2xx6sXG7yXIWlgAgSb88Pgb0Z+P0fgOjCVYxjjhDcYNGiJ1d0VsWgA4KWAOV9rvvvyUVReISpUBl6++oYuO7u6uhKQkaelrjiZ0MRhu673rPDoYDtHReQ4uVO/jJD44VAddnScZQQPfEhK2sPMz7nMcHjrHCP1hdlsD/X1vgZ8/xkaMPx8QuJintmGuumagBgY1dWwC3KdV/dmQMmsnNDcdhIGBVmhoeBv6B1rYZHkSai8f1BE5En98XDZ0qRv4343N5+G6jAK7A5kIrgO8ptDXjWVesRqdI4hPJHYkW2vSTxubq+DylVP8vpf4P/sXEZYI05MWmV284ObAzBxUHvcxXmhkxH7ak8ePJ3dpywUnpLch4WEqly152RWnt3GfOpL63DkPQcrMQoukjv2cHdk8AhcsqM6d0VHOGaipPQBffPUo+Ptio4sx7S1T2mxbsnS8WQuWi+1q/S6MjtRz4sVmLmOAaW2vg6/fuJrXDJyDy5eEXuvYrY3fstenzt4PIaE5UFPzNFy5so83dxke9uId3PD+ELuPzV/mpN0LOUsfg4P/vJctAk/xxzLmrYM7858hNvRAYNArNklxlGq3tsEUVp7buWsZX2Ai5s1ZBcuW3gOzZmZLer8zLIww3q2zQns/Rbuh7z1c7zXIH1hedh8pdddQ6niSCh1N6EKwmGVzdEtrGXz9zdMQF3uLLtId1Tp28Fpy/R5uerdUSQwv9uf3lzjs92GMACp0Tze366Pq/D4e84BKPSJ8PgyPtPD7qNSjo9eDj4+w6MLKciPDFeCNz4k14Ec10NdzGPwDbjFo6oLo7y1jpC/UgA9VrYbYOCFIKiLiFqbej4JG08pJX1TnqNSXZT8LCzJ/JoyV0CSoPF/CPgPgWtN5WLL4HovNNgjuB0eb49EqgGZxVOzGqs9NhDjm4uPS4ftr/whLr/8hu06SJH8eLh6cUJwmQEvimLaSpyXzFO3j+q9JB/THoyvXjUz3ROoyAgkvL9t8PEZvXx188eU2OFv5B+jpreHEHR0tFCJBVY5kLsXk7ugWh5iWt+OBjR5vbp+IM9+8AIODrRASkgQ33fh3aG/7kKnjVvDGK2dMDeHha6CrYy/0dBdxog8IXg9hkUUwOlwHw0OVvLxvb89bEBA4Tuyo6rvVh3m3NmT1lNT/Mgi4i4y8Ga5dK2YLCI3O7D6NLR5uXPYfutcgqdc3lEMnN8N7QWTELIPa3QTPgb453hHmamuJPWVGNqSl3myzCwh/G7oZnNHRTuraii8AWqrdQrF7e/C1UODID8N0FCm1zfv66uFa43jJ1qpvX+D56dbAkYSO1geMyLWlcI4noLW9UmihGpDEF1uLF78OwcEZ/LH2tmKorbkbWpqe0hJvBoSECYVoVJG/h8Cg9fz+2Gg3NF27G0ZHhEj19tbduv2jL31ivnpAQDLMn/+ckPqmTX8bHlJDXb1hqd+cJQ/qXvPtxePEfh4OvAbxWnRE5Dhm0GBUPJr/rcFXZw7C3v2FPLXNGpQ8/2tnV52zhDx3iaD3TFLPzA9zpEpHs7TUlK4Ypsqjowz9TUjsrkjoeJFhfMBU8Z9PxLXGMl2tdrENKhL7woWvQ1CQQOw9PeOvUUX8znBBFMWInSl3ZGUk9saGu6G76wAMaiqFYBavMYiJ3WJ8nMSsgZkzNmlfJxB3c4th/m9yUjZMTxLG0rcXTxDrTQFwczW7Jh1BgFKJHf3qpR+9BLv2rIaSw49D7ZVTZtPajEFIdfuNo1PdrEU4kbrzCL3UUScAL65dVhaNuOH6nQZ/X71aLEmtY1CJowgdFyoYH+CJjVisVem4+ei5RNCPPn/+67yRy5jW3x0Ust4gIE5ERPTv+HMIJPOWpm3CE4yk/XyTzNaAnz37IVCp0nUNferrJ6vxZdkP6J6vvULFaKYCxCIueI26ArEfevsx+ICRemdXgy7yGu9XnrfOeoRFeKSmAROmCqln5m8CIaLRIRWExMA4a4kvOCgZZk4XJnpfn1C4LvNxnqtuidAdFSWKfc8xwn0qEzon9bZxUo+JNiRfJHb0g+tUepjpynKRjNiDQ9ePm9O1M19klHljEloFFi18jjd2wbeg66ajo9LgNajUw8IS+T6vXD1FM9oUIna8RvFadQSxF2z5rckCNbpGL7pcqjHe1c0WYMMrR1fW8zT4uIHqRt94HgjRifoQ0xJSYHI6gkOAq0pbiS89/SEICk6C2amFbNJWuQyh29p4xhOh7q7XRZ+Pt8oTgFXfsBUuxg1iCVgsJmMOUTG/4/vq7irWzX8hoZaPM2ZEZGY+Bp9/sZ2b65tayg1qwAsWgG6+v6ZmgfCxXCcikHLXPR47GKljAJ3S8wMGsaFiNyZiMFAOF5aozuNj0yFn6b2weKHtIU28i11VjUMDgSWgE84ecYv68a6bpy4EJZQYIXOXAK4mzfnRsWgJpkDZC0cROlodcJGS5+B2kK6Ml/4yF/x9x3he+r0/vmCoXrrKoKrybk7q02c8DlHRlmvAY6Bc/dW7YHCwnqnwDJgx65+Sv8uJ99dCa3sVxMethltueml84aGuh7/sW83z2KdNC+U57T29PXD76m1w0/J76SROEYhNnJQuEY3uRmPELrp+kOAtoaa2HL48XQI35twLifHpxhlU2zbayX3YDaZiRuqb3WEsuKb5PTN/Jgh+cZckdHOBcYODajj2/t3wztG1nNjtvVAdReh4oRKhj6OltVKn0tFFMmnS6SzTPR8UnCPtYkNz/aiar6QDJb5HhK+fir+vs7NqwgR5TGfS1wx2g4apdsRHn71GJ3EKAQPo8BpWOtAMSdZYrXgkc3OEjtYjJPIX/nM9/OXVzfDF6UPw8WemI+RdMHBuh7uMBVf1qeMBdMlIQ3OBce0d56DkrTxdlPLJT+7nJG/PytsRv6fiH3umbIS7MXSp6+CbymKdvzw8fHL+d0/POf6ct3fopPrv5pT66KhAuv7+1uWUi+WDsbFLT6/QI0CjUcPpb17VRcfPS1sJgQGh/D6aQuuvVdHJnGLE7ojIeCT2rTtfseo9L/3/G6D4TawRf143ns+dP6FzFRkDBs4VPfVLVzi0qNLdph+7q5K6Syp0XDWaCyDz8xN8mKL7tZdNvmWfb7P6c7DCkiNNaVOhXapUHDy8DV78y0r4vGKfjtSxU9rkBZyg1EPMRK8Paj6DHvXzMDIsZDb09R3Vpaj5+iZb9b30g+y++FIoCftp2bNscdHAH7t9zdPw/YIXuT8TW7XiZ5ytpNz1qQYxMl5pYkd/N6bXSsX8eZObDg0MdENNrfngTqz9gY2jnAiM3drqTmPAmy4D6di17T6zijaETf7Ls7Xpalpmr284BucvFEkfQeoe3uPYWb6xqYyz54/Cl2cO6shc3Boby5gi3q2zuvT318HQUDd/LsCESu/p/DfoarsbetW7obXxLkbue9nfT4n9LiwCA/EuXnyaf5YBsbPbq3XH4e8HV8GF6oNM8YfC3f9SDNfNF9xBy5bcwwPk8HWWJkyCZxN7rsLuNCT2Xa8dlPTa5dn36AZ+RHgi3HnrNvj1tk8hM91yh8Fdj97nzMI0qNK73On8u2aZ2Ni0QldT6+hHxzKpFtW8ajYMscm/rf00n4RDgpMgI/1nbKKNkfQ5y36yVfHgkLUrl9sVue+pKPviDahrOA2REUkQFKjidfixzvvwSDe0tZXD5cv/l71Kw83ora3HeJ33pKRCCAw0XOhpBt6DgZ4/CAs7XBiMahg5n4TRsUHdQiFEtcGsWr9c+weor9sH9fXFjOAHoaurku2jVZcXP6Dp5ilvd92xD2Kix035WIsbf0Nz6yVobW+AW1c8QCd2CgJLy2IWi9J11Y98/AWPvrfkvhMzMW7MuQcK7vwNzEheBL5W9CrAkrVFbBHhwFatIu6Dluomdzr3rkrq6E+/zVW+zsxELIrwa8l1z6OjsqCh8SS/zbv5v7mClwI0Z+FFovTi5I0/bJ9yNdyl4O1jf4S+vlZIjM+An97zJq+vfq0Re51rOLkDDEJnZzm0MELn7Ve9sf3qT3hZVwNrS+tmJkoEJa/LhtPWeQctqWMaHLZiNQa1ugwu1fyak/fIyCC0sgVFf38bD8oTSd3HJxRuv/VVNsYyjKj8Qag8f4K/flbKUr5IIUxNoPlaaWLHevQFTChYqhOPndtiog0VN/rUyz7/H07ylqwP6bOmw/+869CsMkxje8Tdzrmrmt9LXOrLMEI3p2rRLKsfEIe+9Vu+s4dtf9b52S2amF47qHheJhI6xgQQjKOuoUqnpBGZ6evhnh++D3Nnr59kkteVjw0z9KkP9H8Gw0P1ApEzHg+L/DPEJHwAAUGrhRdo/eJdnUVGv8PIsJqp9G38Nb6+oZy8xSA5sXIc3qzIfQmiIo0H26XMWAral0JHRwOd2CkOvOaVrj6HMUC4eJCKhsYq+Mehx+C3O5fDW0d2whcVlqd8J/jXK9zxfLsmqQuRhkWu8FUwH92caQkj3o8c3wilJ39h8LhUdc4XDcc/gUesjCYlQpcXbR11BoQuwt9fBbk37YTvLH9Ol8Kmf4u+b330dh8Q9oFNYIILwT/wVpjmkwyRMS9DRPROHi3P1TQj/s6OvZO+R2PjbhgarOcEvnDBfli+/H0IVWUYNHeJiEiH+DjT6UPY+hL9lvj6js56OrkExYldqDr3pMmqc/rK/MX/XM+2DfDl6UO6xz/6TFoDGCy24+KNX4jUzZ0/ECIPnQYMNDFXYOZi9QF499hGRuyV0NhUDqe/3m39UrCqGgoVzkUnQreM9o56AwU+EWmpG2BZ9nOTlHpb27h1ZWREDeouIRXOyzsJglX/brCP4NANEJvwOnhPE9LOOlp3g2ZgvJZBR/sBaGvdx5k7OfkhCA7J4H7zJTe8JtSA1yp2LBdrKVUynBE7vraaguUIDiJ2NPFbmsuMVTnEcX2t8TxX75Yglsd1EPLc8Ty7bj/1luouiE0bACf51jF97d2XnzZrdr9Sdwzqr53UKSjMT4+Py5Gs0oWqSdugqa2DCN0FSL38q4M8+C0qIgmuX7jeiAKeD8FBSWwBhz51L+5XH+ivhKQkodBUV8dfob/vQ/54BFPmqNAnXXA+MRAYdAtT9G+xRUA3qNVvMdXeCl2dh6G56QVtRH0GpM7eNf6eaf6QkHAntLScZEqnFYaG0bdfAykz7zQ9trrqoebyKTZ+k2DJ4gI6wQQOpX3s5y/V8XkNA9tMAYM5MUddxPWL1sJdtz1q0a8uAn334Wx+Vjr+iCM2rZZx0WkidfmIvUwbCe/wQjRv/H47LFtkvkAIEjgS+wCbaEUVdbX+KMxN+zGfiC3htvueUDSABS0N1PVIGtraGal/eZATclSkcVIXFPB8rp5b207y146NdXO1PTRUA+1tr4C3l4ZXi1OFP2z6okNiD74FerTE3t9XwRYHVbpAupkpu8DP33BBgOMpIjwL6hsOs0WAhpH2JV7pLiY6ywSpN8C5qhM85W3Z0h/SCSY4jNjLzlSZjYhPiE9npH4ccm74IWz84W5YdN0d3GVkDXBuLj11htekV1ytx6a9zLhI4y7n1x3y1Hc5+gMx5QsHvhSsvGUPD2gSLbZSVfrW516GD9igVArodyp5/tc0g0kEqvNRM+Z3fWATnsjIbBgdFV7b3nEMrjXshuFhIeI9JLTQ4uf5+8+H6JgnJj0eHJLN3m+8oE1YWAakz3tI518v/+JZthipNL740HZua2g8TyeXMAlovcN5Tils3fmy2XatD/2sGFblPmBX0yGsNueAMrIoKGshMz+XSF0+ODQSnleNs6I0YUhIMty0TCg4kz53E9x525sWI94xMO75/cr9LCosYwOpRyZrCd0L+vstl/ZduGAnb7860cfu5a2CoJBbJX1maNgGCI8o1FaZE6rARUSuN/uetNmbICYmW+fyKTv1rAlSTzJo80ogGCNFpYLOMHAOU3QtBc7ZA6yEucMBrWe1xF7KiP19tj3MCV7oIEqkbhOESPhORw50S2T45endTCGNBzjNmL4GNnzvfVh6wxMW9690YBwRuu1ITkjnxNzQZDlgJygwGWbMKJxE6gGB1jVqiYzeAr5+SboUtIjIDRbfs/SG53TuHgzQbLhWPuk1aM70gjEgTieYZCqFS8pKCZyTCuxj8OEnr/LYFwOLwD3rFK+cp4c8ECzHpZyTMvPHjGwi8TuN9N2lTKxD8gUtmd01g2o4fORu+Or0C3C89H7+t75itwRcteLqVakSsJZq0xPMIykxQ0fOHZ11Fl8/Y3ohb3eqn+Lm529du91p01Sc2LlyV62W9J7g4CTIzHxQp8S/OWc8HUh8/t+fXgYv/GchHH5vD51kglFixwJbSgAL00gtJTsRJz99FYr++hA8wcbvf7y0AQ69vRNOfVliVIi5UDc3kfgrtO3DidRNIEvpD5Bidi9+83u8whiiu7cePvrEumYt2NlIqeAUsX0qdVuzHXNmZetI/dLlcouvx4C55OTNBkrdPyDD6s/F9yD5YtS7VMybc6/OBN/YWG6W1LGkbPWlU/D20T08IJBAmEjsWGBLKWLEGhzm/OumcPT9PXC26gRv/CJanHAcT4QDzfDWIAUEk73DFbvrk7oQoKB49Ds2a7GkcG/IeojfinFUl68egwvVByTtv6jkqKIV43BBQoRuHxZmrtYp7prLZZLeE59gWG1uZNj6VrtYiAYRYMWCAFuwisQ9OKSG7m4jZK1XsEYk+LZ2qjBHMKKaFO7HXrDlt1b712enLJ30mDFS54LJsWZ4yeslthUSqU/GDqU/AAcDNj+whLlpG2BOmjaQScvs5Z8/bbEQCKaPYDSoUtjLCF1qtD7BNLCJy8L5q3iwHHZr6+iybIIPDEiG4OAMnQl+cNB6JdzZUcRv0RQvFdglTr8da3eP4ec2Nlfq1I3+68jJTjBH7ChulACmnlnrX0+dtVR7jYVCZsZKWHvHNvjXBw+YEWY/d8nD6ugPdO089cx8rKOpeOUMawLLEuJzoL5eKAISGpIEt676L3Zr3p+OdZEvNzQr8t2ldo8jWMbV+kp2Xruh+nI5b9Zytb4M5qTezBS0ebIdHGyFbnWZrsmLpWA3zUAZ7/SG+erNjb+C3u5j2rKyOeyzpPnkUalfqj0oWBZGMSp+HYSGjuf6vnv0SWhpu8Sbv4gLDkzBWzB/JcTHUplNgmliV6qwCxamwX1bqv8hIjZ6FmQtuIOR+aOweOEdMHP6IlCFRpt8vdhQRslUYRtQAS3Vh0ipZ+ZvYtslR5gufnP/Ru6TkQp/3qzlOZgzez2s/e6bEBVpfhLesWe/Yn50DOyjanHy4ZPPi+Hw0Rd1pvRrTVXwl/1r4cSHT0NLa6XJ92FTF/E9vb1ljOSNK/zRUTW0XLsLWhvvhsb670Hd5VzoUReDGPo+NGSdyhdN6/ifSOj4PQ++9SBcqD6hVedjXOmIJvi6BspbJ5gHmrKVKieL86FU/3pgoAoSE9Kt++4bCxQL+rMRtY7+QMcb44TAATRJ5E14JkW75Tnqq+DJr31vn3IWALZiXMFUuhKg1DUFSP1UMez/+3bw82OLN98x4ZZtfr7CbUhwKCTEpUNyYg5X5WipCQ5O5qr72/O/0L0+Ono9JCb/btL+uzv/P0biRYzcvbhqFtWzuPn6ZUDK7H9K+q4tLWVwonQTDA0D32Kis6GxuYotKrr538PDXvz2u7c9Dae+PATfVp9ijwH8aMPTsGwJlY0lSFDtG+5XRJDg3FVxwLpMDExlQ3/6xUvlUF1zCjZvfAGSTBC+kvOuLYcRzh5xaJlZHweS+UwQ/OOFrnK0lVS5YvqaEsBgFiz/SoQu89V33Wp47W/bdap72ZJNkDYrm6nfKmhsKoNmRqSY/dDWXs5IH7Tkj7djnPhFM3dXZzGvCqcKGzfDDw+dg/6eIr6KVoVvAV//HBjoL4Nu9QEY1AgKXaOphIvn8yA2/nH2XstKCVW4F5PfqMCvNZWzz/DS+c5nTF8Kq3IfhbjYDPj8q0O6x6OotzpBqihhoiElf5PsKbi4UEDFLiViHfPT9/3fhxipN3AXFa+kzP67WFNuktTzli7kVkxMp3P2IXQ0oTuO1IV8vVJwQg13U8CTnqdgtCQOWqXqEiOhW+MyIEgDBsotylwF5749zsn5WmMl3LnmcZiXhgQrZD5cvnIUmlvLeFMXjFofNdFnHX3lQ5rPwD8gmT1QBwO9gpl9mk8ShIYLeelYqCY8cgtvwdrWshtGh7u5Cf7q5fshLHw9JCQ9bjJ4rrpmn2F0Oy4WVIkwPXk1pM8tgNiYcb9l7ZVT45XlKFCOIBFiDjvGBMlN7E+y+bGAzcGWMnaQuDs6DTM2cAijas/9zr0m34cBf6jYlaoJIhFbnfGhjvKpl7gSoaPSVSrKUzT/KFUGFvu757le6obHYNmS9bpysTWXyycVoZk5Yw0svf4JuOv2UliRewiSk9Zz4tUVoBkdL0TT210MPV27QdNfrCPTwODJQXThEZshZfYHTN1j8RmhXGxHRzFUX9g4qV87VzqnH2WLiuM6Qo+Pz4Y1K1+Eu394HG75znYDQhcnQdGnHhmRSCeZIN16pWBEvNQysqliapsVC1IUPebaZjsAu5yh0hHKR78LKv1RVxqoj/70h4qlgPG2g/c9ocgKEa0Lux69DwjKASPDz5w7Br29rdxvPshINWOucVN4QEAMxMetYcT+YwgKSoX+/kretQ0j4LE1qze/FSLi+XzE+Dos6mVGrpM7+Hl7+4Mq7Lvg65sEfb1lMDo6yFR7K3R1nYSIiDv580NDavj660eg4drbfOHgMy0UFlz3S8hZ+iSEhxmPaG9sqoRTX/5N8N+zreDOR+kkE6wmdiW6umHL6QHNoNk2rYiOTsE9tXRxAdy68n748YZneDS8JaD4wfogTlDrFYzQnbaiUJ7UY9PQ8VHoKgMUg+OwUEsARjQpsWD4038rkg6C3/vdl59S7HsTxhEXkwqnvhLasLa0VcHiBevMprVhW9SQkPmQkLAZ/PySQDOAkfIiuQsbkrp/0HoICLzL7GcHBM5n+7oZ1J2HGQkP8nS5jo6TbJFRA5VVj0KXuopbEUJD0yEnZw9bUJgvLVvXcAa+Pvc2f88spniWUm91gg1AEYSNqJCI5QS2aUUzvJiOZgyzZ2XzcYu3kVbGhGAL2P959wNHH67boKW6yVnnynuqDU4MzlAqwEwps7sQGPdrCoxzEOak5sAtN96rM6O/9/4zkt8bHb0B5sz7J0RFbzHs3uaF5WClpQkhsc/NKOVlY3Ex0N9XCdeu7eO913E/KlU63Lj8NabOLef7Xr5aris6kzZrKZ1cgs3AWB4lKs4pFVAsLkYcXGmuyFlmd0eSeoWrDEopleMqzxdD+RcvQE3tUVB310net5LR7rgQoRKwDl5qr3gIEuOESnFVF47BmbPFkt+LPvbo2C2QNPOf4D0tSdef3cdvmVX7mJ22H/z9k8YD4pDQmULPXvoarzsvBVeuntL51DMzVtGJJdgM9FNb05ZaKsRoeKXg4EpzJc4+T8qb31uqNRCbhnnp6c7+sZjCZilq/PQ3RWzbBxdr3oavz+5jE3oxT2NKS73T7Pue+6+/KZJCQX5058DX1x9mJGcxMj/MSHkQGhrLYHbKLRAcHCN5Hz4+MUxlH4Wx0Xrw88uAwODN1q24vf25Wm9rK+Z57NO8VZCVhURv/Du0tVdCZ+cl6Oquhy51PSP0cvjytOBPj4tLh1W5D9CJJdiF9NTpXMCg2VxOYBW4wrVrFLFGomlfiZgAozh75MfOPkeOylPfAQ4o92qJHKVEjbe0CZXDxKYtWFNbFWLej4MVkp5UYKUp+v8JzkFifAb89J79sO+Njbx87N9KNsIP1+2HmGjpjVe8vFSckPHWFoSG5nC1PjxcD7NSH4eAgPGSxNjI5fLlg4y8j0N9Q7lB0Rm+6eWto1LHgCPss04g2KV8HxXSxeQmSawNjyl0ihDQ/RsVbaglUoErnB/H+NQFH8Mupw5EiWkZra1VOkIX2T0qyvwkji1VlbIskB/duUiIy4DCH++HsNAk3gKy+J8bobWtUvL7UaGPIrlPs51Mw8NXC61ZA8b3ceHCC3Ds+Eo4/fUz0NJSbtCwRb8zm/h3Qnw6ETpB1rlJbv86qnWMVrcW9Q1VUP6Feas3Wmh/o3x7VpdI23ZcQ5eW6iMQm4ZqPd7RPxLrGEvpwobImLcOkhOz2QQ4m0+GGqaG0uesh2gTxI6DUIngOByAUr8zQVmEBMfA3Nmr4ULNMRjQtELtlcMQHBRtse4/Auu99/cdBh8fFQQEbbDp8/t6K0CtZkp8sBuiY+6Es2cfhUu1rzJVPsitABjZzlPcfEMhJnoR97djSp6QxuYFCzMLoOC7z9CJJMgGNGljJo7cmT5oAbjvX+40m+WDJWPfKH4cTnz4F/h7yW/hk/K/wdfnTkBaqvno+Kx5qfDy3w6DZnBIOVKPTStiXNc1NUgdEZv2LgjpbQGO+khcTb7x++2SFa+/v4oTelJiDqTPXQ/XL/q5SUIXc9LlHiRYG/mNP2ynmcNFUPzWNnjrve0wNCykqY2BhqmDY4xoKxmJZjE1bi7dLQa6u14Bby81BIXaFhvR043lZMsZUV+ChmvFvEwtJ/JRtCJlQ0bGQ3DD9Y/BksW/hLlz1rEFhD/UXBKq4mVmrIO7bidCJ8gP7LaGrkfsviYXcC61lLve3tkAh97eCeruNoNKiVh9LmXGIpPvw4XCANu/wl3cwh3dlW0iHJvSdvbIZXBAf3R9YFUhc8FxAxq1zfvGiE25CxvgIoQ6r7kWVtz8EPj7hfJqcWK1Oby9UncM3npnBZz8+G749sJuqKnZDZ2dZYzsGQkzIu7twe2oUG1upBsGevfa9Pm4H9G8rtHU66W1vQo33/QazEpZByHBgkL5tPwZeP/Dx/gkd11GAdyZT4ROUA5KVJtDy6e5Tm6mar6jgrfIBxsLFEnL00Mh7zLqRDinErTQVjXFESq99sg+syr97yUboe5aOUxPzOYBUMnaW1Wo+R7pOOgWf1/+aGIsA+vk8oYEIyj7oghOfPiM0LVNv4ub2NiFd3Ob8Ji22YvY6Q23YNXvwN8KM/zIiBq+OXODNghOCH5Lnv4QzJr1kMHrBgfVcOqLZ6Hq24PstV6QlloAq1c8SyeOoDyJXHeb7PvE9GNzQXN/fHE9NFw7D6mzlkJEeCI3uy/IWAVJiZaTrHa9dhAeUSgOSv9juIA9e8ThpnhnkfqfwAHF7tEvbakT0Et/uQEGh7p1gUW4zU5ZzRSO+daA2ORAbjOOpYFMcC5e//tG3g1NJPA5qashKEgFg5o68PUBrs4nkzoS/piO1BHe0zLA13+NtuOUl65ADW/2ojIk/I72A3Dl8qMwPIKvUUFGxn4IDsmYROgnSu+F5tYqGB4CNtGtg7xbiNAJygNdkBE3fl+Rfe996pcm44r6+9W837qtSLl1k2INtyagRLvVglA+VnGS93HSWChRmtRRpaOpxRy61HWgGRQIXX95E20h2h3LJcpN6NzsTulrLo0N39sDr76xFgYG6jkJq3vq4ablOyf51LEJC9aBxwXi0OA5NrTU3Kc+Noq14dUwpKlkRFyp7aPuBSN6PdWxxGx07BM6ld7c9AIfmj4+oTBnzmsQFGw4Nru6KuGDk/dAX3+PUGBm/r2Qs/QxOlkEh6BIwTSxHdpObsYsrfYQOt83E3ubn/ijIw5RAeinc2fm87UQCOlvAuELbmnZMM0pIyE2DX+Uop0lsGmLpUYBGo2aqat6Nnlq2CTbrVPqOUseMmt+V6Jhi5JNZggyrYB9/GF6Ug58exEL0mgYYbdCU/NJSJlxJ6//LgKLxvj7JzOyT2aTz2IICFwG/oG5PPo9IOgnEKx6mJeM9fVbzF6rglFG3qOj3fy9A/0VbCzW8cebG/8AfX2nmbIPhbQ5+xmhG0bb19bug/LPH9BGwXtBTs6zsCDzZ3SiCA5R6BhJjsSrVDQ5zrEY3Ca1KyWmtl2+chq+PPMuvHtsD+TcYFrUYYVOJzV7QWCgeApSCRe3sWmYHSZbgXpnmd9zQeivrphKt+RLN0bwdQ1l0NpWBVkLNvEoeFOrR7kLzWC0e8WBPTRTuAlaWivhzXc2AjDVjeb12OgMWJG7H/x8bVcPPeoD0NbyFCPobkG1jwkqHu+jkp8z9xAEBgmk3tNTCefPP82+xynua/eCULjpO69BeHgGnRyC7MD4oYqqGqhtaOIpZ1idzUGma0lz+cWacnjpL5uF+iJadxbiwZ/thTmppvsdOMi3LtnowRT7Zjl25CxSfxgULEYjxZdu6+o0JX+T7Ku7r/7xEtV2dzNgAZq337ubTSLdOmJflr0HgoKSbd5n/ZW7YGCgUmeKF0ldJPbAoGxeMQ5T2oa1VeNUqhxYcsNLkmvBEwimgGQtEjcSucNKq1oA1hkxlxFUf60K/vDChnFS57cAP/nBU2bVulLzuR3IkqMZjON96pn5YaBwWpspXzqmr6EiD1PZNvHu2l8i+wB4mH1XInT3A8Zd3Jm/H0pP3g+Dg/XQ2VUJJz/+HmQt2glxsdYVDULze1fHbhgarNQ1XxEwpisPhzc93UIpWC/2h69vKMxLfxySEtfTySBYLU4qGFmLxI33Fc7dtgtY3hVFmqnUZDHFzQvAoBpoe0eD2f2i+sdMoycVbCZjJZC47CZ1xyp1gdBL+YrECau6r88VwzvHHoUA/1CYkZwDsTEZMD0pm/tJpaxis75/v6ykboubgOBawMjz4x9shN7eSl0KW1xMDkyfXghRUcbJfWS4DoaHzsGgpoxtldDfX66nzkHP/G4YSDcyArzrW2TUBkhM3MSr1BEI5uBM07kj1fpf/yEEh0aGJ8FsXlkuEaIk9F53MbWOKXBPug+pCwn5qNBTlPyYS4wkTa3o3j66Db6pPGjQ5AL/++UD31rcL7ZVlbshgLmUDYJ74VzlbrhY/YKQl66Xux6myuDxGb4+Y+AzTc2Gm6DGx0l7gpl9bAK5i8+xxzGtDQPm0OROIEwUHa5oOpcTSrkplYiTshFbGak/b+9OlDG/C4o8S0vgWVqzQorSRwRXc+aqxzW1jDfiGNNaNmOj0yVdMHITupTe7gT3wfyMLRAbm8OryvX0lOlyz/v6KpkyF/LaxcI13t5s7I1NXl77+WewWxUM9J+DsVG9VEutXdELnBUEQ3AVuJvpXFbG2/mKInU8sOWri5C6LE1E5CF1gcQLteSd57STfo/5vPQ5qWsggKkmjF7G/HScV6W00dyhwAnfte3nNEN5GKKjctj2OvQP1EF721HeLhX95FiYBpU63vr4YD14oQANtmOd5juf3/oHjKtvddcBaLq2TfCl65G5F7H6lIKnmM7lAi5e8DhITXFD1DVU8fKxdQ3nuUl+2ZLJHIFCEAWhA1qzmkORXPnq9k8RglkdI9md2nbO2mpszUy1NzNyD1MlmfWp44U0S+ZSvpb8QwRCbXUur/Oub35H33ps3EOQmLiFDpAnnespYDp39Dz/u90boJ4RObeY4QPsv9ybNsL3v/eoyQWUEmW/pRpgQIh8l4XU7VPqAqEXuYRpxkL1uInAIDncHK3SMTiOVDrBEgKDcmBQU6yV6l5kfvcATGXTuTPVuhgVjyRvCuirxwWDk87HVjmrytlO6oLJfZcrnOiZiXGKVGNTwpeOKRQU7U6wBF/f5Enmd4L7gEznygGFliW1jmlunMRFRvcSTPGWhKETSB0JfZ+cO7RHqeeBk03u40RZoNjgkXvxoURRHMLUgJi1QXAdkOncOWodj7W5SHhMZ0tLXcr96JjmJt43BxSGOEc7cPFVJEe0u5yknuUKJxjN2Ri96A4qnQidYBWJ6xO5VnH0dJcBJNCxcTTIdO5a2PVaidm4pNtW3w+2NIRFgeig0rG1oFBTM3tIvdQVTm7ByhsVMWdjXWA5gfXdKYWNYDWrj+kRvHiHoChEczmZzl0XlqrM2QoUiA4idcV6rdtO6mePfACZ+Ri151QTvBKmd1yVy91ScNej99GVSLARWqegF3G6rFIJFTf6vvUUOJnO3QfoHpU7iwgFogPS22rl9qPLpdQ5V4HCddzNAaMVLVUYOvr+09DUWslOVhKEhSXzSRHT2BbMN10zW+4a7/g9rYnWJBBEAtfPTw8Ozobk6Tvp2NiwSBeJWwxeI9O5Z6h1zCSS21KLQlFhUi9VcudykDpKZaf416X40htbKuFqXTlc0SsNO3N6tllSxz67sq4oyZdOsBKRUVv4RrBytpxgOkcid6EuXASZgQJM7vkVhSK6SxW02tS6LqmjTyAzv1C78nCoGZ4HyNnoox4bA7OELqf/jFQ6gaDArEimcwIIsU9KiCbkFhfqte5QpQ68/2tmfp6jiR0D5CSTOBj6ItH8bm6QkEonEFwDZDonmANaYVCIWSvw+vrVcKH6FAQGqmDu7KWTSV3ZgDm0bj+p1M7li7vJzF8LMhWklwJrO/Z0dtWxAVCv86kb66mO5roVm38lq0pXogEBgeCJINM5wRagqbziwB6Lr0MSP3P2OHzLbuu01eUwf/2RX+w1rtYV6MxpQOxnjxxyTaU+rtgPMWLfAQ4InMOTaG0LvvCwZL6ZA/nSCQTlQaZzgpzAsSOldGzpR68xUj8BYxOI3iTrrlyuJKkXcQs3WrpdltQFYn9S62NPUfIkKpHvLXexGfKlE6Y6yHROcBRQkFmab+fMXgpnzp3Q9VEQyR2VuzETvMIV5tBVXaEVwrvkzFlXop/6DlC4yYulqPd9b2yEgAAVxMdm6Ap2zJu9GuJiTTdwkTsvXYkqdwSCq4JM5wRnQkp6G5raRSQlzoM57G8k+mR235xaf36/ol5l5MutjNyRM0v5ZifBK0Hqih6BtewgW8pLrL1azsn824vHdGlsuJkldRlN77i6o+pxBE8Emc4JLqvWGbFjwyxTSE5Mhy0/28sJPShQJU2csXlcYVIXVftWEMvGZubXchUvcGmJtSQvP6kLaW6K/XpJUe/64e7a+2PmViHHP5HVxEK+dIK7g0znBHcD1oM3R+qIOXpm9rr6Kvi25hRcuHgKFl63EpYvnVyd1AE568aQot3wC3VyE70VjV983O3EoTnE7GTUVWeU482lsclpesf8eUvfkUBwJZDpnOAJuKwdv5Z86/v/9jic/uYE9PV3C/zACKJvoNsoqYtq3Yk566jidzFiz2LEvtk5pJ6Zv0ipX4c1eS2Z3jHC/Te/+hYGBtS8mpwo2ONijJvecTI7dOJT2b4j9UsnuCrIdE7wdEgJmEMy72ckrh8sh2rdtJC80RUK0RQybi2VUjNeqUA5p6h0fWCgXMr0HIuvKznxibxHngLkCE4Gmc4JUxU4n3eqzQfM8Sj4sycMHgsMDIWr9VUwPSl90uuxE5wTTPBGNSPbHEzqmfmbQPADyA5u1l51o+z7RT+MnJYEuVsBEgjmQKZzAmEcOPZLTnxqNlAZo94DA0I5uWNE/Fy2JRshcwOx5hplYyX1WJGH1DPzw0AMzVdMpZsm9I7OOujsqodZM3Os2idOgHIGyFHEO0EpkOmcQJAGjJEyNxdjFPzvf2udy9VFTPC1ypO64D8v1G6K1n03Z3p//6MX4KuvDwrpa9otZUY2FP5ov/mTL3MaGxWbIdgLMp0TCPbhA631Sk6rqYuY4EvlJ/XM/JkgmNfztJvDGriYI8wOptINMAbmc9j0VnRyAXvwEghWXaFkOicQFAH61i2lt1ktLFfd6GxS3yEfqQuKfJeWyB0OKQVnDAhdQpsazE2XcwKlADmCKYjmcn0FrlDpSQKBAIIVVnZSZzz05J79zvpJWEr2sjykLgS/FTnzBFkqOLN4wXqYNSMHaq+WCdXj2H/xZqrHCSs5+dLYpKTaETwfoulcX4GT6ZxAcDxQUcttgsdCNArWgjcHTGV7ROqLfSwQeq6zCV1cIZnD9QvXa+89JHmfcqayUYDc1INI3PrBa2Q6JxBcB4qY4JWvBT+Z0K3MKLOk1Lc6+8RgcILcKlhO0zsFyHk2yHROILgn7DHBt7XXQ1Tk5CqkONc7kNSLpFaRs4bUnR79pURuupymdyoJ6xkg0zmBoBxmxQ/ArDgNhIcMQ1jwMH+sf9AbGtv9oOpqIDR2+Mn+mVJN8J+dKuE14NvaG3g9eKw4t+i6lXDf5t0O4SNTNGULoUsh9U5wYIS7o0hTTtO73OYdgvIg0zmBoDyQwFdldUHGjD4I8BPSkbDO+pg2MwlvM6b3Q97CLrjU5A/vfh4BTTKTuxQT/GefH4ILjNTH9LKmxLrwxoCB23KWFjeCWhDSxG2CJVIvdaZaxypyGJwg94Qu1wSOrgGqIOe6INM5geAcfCdTDXfmdHKS5Dw5Zj7LOIWp+MLVTVB0LE5WYpdigo+KTGSkrv1DWxD+QrXpWvBogleY1HfY01PdEqkXOZPUJbVZtXbldpwC5DwNZDonEFwHP7ilDa6f06tT5BOB2UkTn0MuRTW/aXUzvPJ2PHT1ylPsFE3wOD+Yi8uKihj3nUdGJPK/kxPnmeUlBavL1Upp2mI7qZ89coh3hnFSfroSAWjy+tNvpCvYwSDTOcEVkJqogdiIYXY7CLMSBuHpV93bYhfoNwrJMYM6Nd2v8Yb6VusV8w1zetjWqysXMpHXjT2mjwD2PXIXdMGbn0XJOuebE2A5S9bCwsyVFuu/66wKSXFKpraV2rsDKcuhAu0HZTl6oMlN6kgGcp0IMr0rCzKdE1wJa5Z2w8LZAzCLkTgSOWJMz7QcHDACvQPT3Oo3LZvfA3OSBmBO8gBEho6M/ybRr6zxgtIKFXxwWsWD2iwBCfmuZR2G5C3eMcLmOsWu9xzezUrtgyNfRIBmyFuW34l+dXOkbizKHTu2IaabIHrkpn0yViQ1UOqKkzra9jPz8xxN7OZIs39ADdcaK3WjINBfBQnxGZJOrlwg07s8INM5wR2w/Lo+uJFtYwBGzcpI9F/XBLrFb1k0uw9+kNsOkaoRna/b2G8K9BuD27O7IHehGv7yTixUNwSY3W/mzD42F48ZNa3jAuFERRh8eTEEBtgCIWt2D9y+tAMCfI3p9jFIiRuA83VB8khfCfPJ6W+Ow/sn9/NUNoyCF7/Vn//4jaNJ3W6Olea4GCd2TNDLc6ZKf2XfPVBzuUxo3KJd7aXOzIaf3mO5fB+Z3p0LMp0TXAFpyRqIjxxmt4OwaE4/XKzzg5cORJt9T02DH9y4oI8zoL5PWBSZs9yE1Dflt8Ly+b3jRC5BQSNRP1jQBC+WxJkl9syUfsODooe3yyM4oYuoqA7hyv72JZ2TVT1DfMQQI3V5fjPOMTj3mLP81jWcH4+A10NfvxqCAlVG5v/lsFmZU5THu54qGChnSOwAK9gH7gU7wu3tJfWE+HRO6gbKXaOWpAjlUoBkejcPMp0TXA0P/aAVFs8Z4ESur0zxdlHagEVSP3ORkVm+6efjIoZd/hisul4NyzN7DUzdY1a8/3/f3gy//1sidHQbpw30y5uCPqGL+KxSBcvTuyE8ZETx344B0uZIfe7spfD2UZh0YOrqz8PctKWTXo+Bdwp1bcMUcuyzYvOawZYQw61aE4GipnhTJwBN7RPR2FQliwlGKsj0Pr5QItM5wR0wZ/ogpLHNGIuhMo2LHIKmdl+T729q9zH6PnFxIPrZ5UR02DAsSe+DjBkDEBQwqluEfHE+CE6eCYE+jXSfc5RqGP5lRYfVEen6ChrN8eu+0w7//W6s2c8yprxNofJqECzP6J70vpmxGlmPpaX5PzAglN8GBYZCUmI6REUkQmREEk93MwVMt1aoa1shE8+gVPEZ44o9Mx+VeoVSF6C50rAR4UlwU84mdhJUuhHkJcNJlWPB4ckg0znBnXHhqh9X6mNgXKWigjdL6h2+ZtXtgtkDsn3XIP9RuPe2drhlUY+OyPUtC+kzB6Dg5k54dn88XGmWFqF+1/JOk98fHztf5w8X6gIgMnQYcjJ6TRL1dbP6YXbigGkz/IQPEP9MiByEa+2TvytWk+OkPuF9AzIFyYmwlNqGke8v/f4bq3lAIb+6SOwp/FZidzZ7lDoS+2n2gUWgkBneXMGZJVnrFVmpSQWmMshdEMeVQKZzgieisc0XzLEykvrHZ4LN7gNN8AuNkbd2n7ERQ9Dc4WuZtANGYFb8IMyfNQAx4cNQ9E4U9A0IJDaTPf7rTY1cmY+ZUbnB7PntGxslE3vWnD6TCvpwWRgc/my8cCiS+92r20wq71sWqo2SekcP0onG6Puwstz+EzGT3lPbFGD0cDZJOI62cICcZV4dIO7y8BBp08qRb0uk+NrtyfAvUYrU5T5YuEKTy0ziKbXeyXROmFJKvc448Ylm57Qky+be6no/HakbM1ejX90cqW++ow1WXt+jI2wxHe7cpQAorQjlhP6bwkau1MdMfE/DxcEobPl+M/yfPclmv/f0mEG2zzGjgrofo9K/MnRpllWFMLXeIxwTI8p7QUo/RDBFP9G3Xt3gDzfM7TW6eJo/sx8Wp/XAV0Z869fafSE+cogRvD9fGDSyv6tkinxXktQVzlefSO55/J5I8GYK1NhD6oqZ36WQ+kdlRXDu/DHdgMcxhGltd655XDGV7qDVmSKDeaICJ9M5YSqh4ttAo+paRHyU5UC3pg4f49Hi2ofQBG8uAj44cJRvEyPnl2b0cVK/v6CVK/AxC35o/a8QEzYMNy3ogY++Nl0xbXrsoMkd1bX48UIzE/F2eThsWddk8n0LUvrgw68NFwM11wLMutG/f3M7v51I7EXvxclubjc1D8qNrPRUR1syBYLPzN8Bgmn+AzlJXZFGL1jvXUpk+dmqY3DpSrkurU3clD6ZDuzSY/3EpUfcogIn0zmBICpCH57KZox40PxuCTX1fiaVPgIry5lDS4eP0fchqd+xTA0pCYMGhVqM+r+NFGy5danaLKlHhQ2bXMyYwsX6AGhTTxNy2Y28b+m8nkmkjir782+DeUU5U5/3/ZvaeSrbp+fG3+sIQkdIKRkrAovP9Pd3w7fVp9itGgIDVfDd/PuNijyF68CbNBQgtfH4tgmq3R5Sz3OWSp+IMQtjVS5Sz3URlS6ay/WD1xSKwiQQPAaNbT6QwEjd1DyxKK0fTl80rbSr6837ruMizav9by4FwA9NKO/C29sNvldlbQCcqgqC3gFvHjCXMXPApGKfETfII+Vbu0xP58Zy66UQe46q12hwYVL0kFET/NEvwjip67924ufdmd0JCRFDUPxxlMPHgBQT/K4/b+Zkrm8xmTN7qQml7vT4qiJuktcLprON1DE5XkhtU8CcIe0gpabkwKXL5brQdzwBAwOT89Xl9Kc7w/ROpnMCQR589W0gj4A3pYTRBH/6oun3YxnYnn5vCAmY7PPGnVlKa2vu8AGTzKz3ZT6oCIE/l4znzX94OgQev6fRKLGL77t+Th+897lK+sGQwOwXkNQnRsLrvS8tcQBOnQ+ZpNaPfhkGa67vMvs7r0/rhfjIQfjvd+McptSlkjo2dZmIuvoql+EEI0AufsRepb5LK/+dptTX5D7ENykn0ZlWBKkg0zmBoCwuXjWvtKX41Wv0guWMYVaCBi5d8zf6XEunL/SyRYGYc26MZ8uZOtcndBHvlquMk7r2fajWreVyS6nAqNRNvQ/vJ0YZ/8xjjNSxZGxC5JBZxY7P33/XNfjr+zHQKHMfdZPzrASBJ9aC1/++/QOm+6srVITGKi2s/4d1pJ6ZPxOE0Po8xb4dO0COPomOBJnOCQTn4Fqbj1GGE03TWXP6YR9EmN3HmeoAWJg2MKlcrAhU66ZIHXHpmh9clzpgJBhOkLH73o00+j4sOIPFZjAy3hiiw4bNErkpzJ1uOuq/XTStm0huR6VuCq8cjoOH112DCFPV4rT7xGpym/Ob4J1TEbx0rNKQkuGDpvZlS9byFqxJifN4mVhjVeV0nKVcERqbII3UBTIv1Mr8cKW+DKYHSAlisNbcIue+rFHrZDonEFxIqdf5m30+Lcmy2jVWWU6f9yz51UVSBzDNl6ZwudFPp9Ynvi99pu3FbwLZQqHfRHW6C/X+JtP90K9uCti05dWjMfCzO5smNW2Z+N15pbob2yE8eARKz4QpPg4szeNYMnbu7KXShWh6Kuw75NShXSKN1AW/eYHeprwNId28Sn/v/Rf4bUNTJWg03XxoYDGaGxaZLkiDRCoXntyzH8JDg2HrPesmfQaZzgkE1wfmq2OkuzEyDQnClLMR6O2fZhWp6yv2BanmybWl8/+xdx3wUVTb+ySEVNIhpAGhB0JTOuqjE8VCKHYR8O/TZ0FR7BV99oqC2BXE+kQIiNIRVHqRHlogkJ6QuumU/Ofc2dnM7t57ZzaZ3Q3kfr/fZDfJzuzszOx89zv3nO/QM+AVxEXWWL3GitRzZVKnZcArap2WLKfYybLWwz7qaDhDQ2a+N3SWSJ1l+4oh+KwCeugcHeQ+kxT75FH5TH939TEY3rtEet05SN7s3AQ6vF8bOZXq5mQ5tAqczyf1uiQ4p6ry+hyc1RvmWi4EeamFjnEDuSfPaGX8yJufwiyJ3HFfjUzCExAQcA2pdzY3daGqdel/e4+xM+D3pfI7sUWEait1ntJHUsesdxrOUMhezbNtI2qopJ5OcZxTr9eGQ+oFNtnttryOtrIsUleI/cPkKPj3NblkDp23LcIBHcolxX4OftzQymkJdOiSaSTcnCw31dZlzsuG0NuZpXwfd+yd1nx6x7gBkJq23epvRcUZLjt5CnCgINzXBATqh67tqqF/t0qIbnmOPEdknWkOh9O84Y9dAeS5s2BrF2vnAR/DJ3UElrZ1jK6hZsBj+D3A9zzJlKfh4Ek/NqN7yDaxLKBS540IMFlu9zH7AYEt0dvavqJSZyHzjDdzPURMyxo4kMZ3f8NQ/OcrWsOkqwqgW9tKzXavca2rYcroXJi/pjVUO4HYjYzeKnBTshwSul3g38tGobuN0PUodVpuSVFxpktPnoCAgOMI9D8Pk68pgTuuLoHAgAuW72/dd7oKxl0F8MTkAli/0x/e/ralU8j9n6OSIr2W/X/sta4FDMF35JSvddDorY6lbej57qjSTznF7mXOy4AvKPXirocd3FhQZ8DTBgYdo/XN5SOxf7uuFVw7oAiGJJg0X4/91JMGF8BPf7Yy/BpwBvmiYZoLSb2YRei2Sn2GOwldOTA8jLv6GVKL3rH9QH0jMhEaFxBwO0b0K4dX/pMHQf4X6INzu9dXQL9uGXD3q9Fw5LSPwUqdnwGvq6wtyxuG9Kywy4BXiK69DlJH8rY3g6kl4XceUHUrme62gpeXAX803YeZ6c7LgJcHBWZnOUoGPIbfHcFv20Mhq7A5TLyykGu5i0/i21TC0J4lsHG/8clzjiY96xGkLnKWQ3v2JF7nNnVsY5g7v/h63NpiorrpJnSh1AUE3I9x/yqFD2fmQFDABYfWw9d/8WwWRHMyrOtH6s3BVOlpIRBb9O6krTw1neU05tUPnvTl/r97HHsf7JLoVB+CV6t+ppRf6BTGUeuFJva6YYHnHY+WHG8Bc5dFQmWNdtNsJPXWIcb3qjeaG4wuxWZgPuFpjVasno3lyx8X3drQ7WEGuigfExCoHwYkVMKA7nVLfYDz5a/dl1/vfQiUlP3L9+QZ/tmOZ7BJ2YPMi9cNJPB5r06VMPnqInjstjz4+LEMmHVXHnU9BVrOcupkOfV6ylNWaB6hhOBp6yHi29IHBAXmeXXLejZ8qhWC9wD6egi9IXirwVWhN3y0LIp0ZKP17FC/X2LfYieQurFRXD39ShoaXJDIfJqjrVeL3XkTMfqgpImSMgEB3YhpdRamXldMCLxbezMp1VpHR9duD4AFv4XA9kN+urb52n15lhs0K+Ru9T/bNp/S7/26VcENV5XCsr+CDPusx9K9iV0sy+nsgYkF0MLvApk3J13V1McCH2s9uJ+rZ0c+yeUVc1SztNEIDqlrJcuxQvBH0n3heihhHvsusVXMDHjs5MY8ZyCH4Ouje4vLvODLla1hwpUFJNTOegNMnGsXUQWn8nwbLT84uawNuVl3Wbn66koGF9WjuyJ8sUFkpwsI6FDD5+GhmwslQi+x3EhrGTfwUf3LYaS0LPgtGF5fwE9g6i8NDuLb1TD556NfQmHpxkAwVXiS5Ln7JhYxifL2xBJDSd02A94WV/aqUBE45WU6RikRoWeZvdXTaI5zqm06FH632QSL1E/nMYhZRwa8OvxOy4DXmlfvEFlFmr9grXqRqRnsVrVexbI1tIkdP6QALutUble3r7xN7w7lhpK6nuqlrTuSYcvOpZYBXZvorsxubQgnZsDP1qPQ7Ukd27dhGzc3za0brtQzhVIXENAi9O9ezoTu7Wv0uZqZ77BTrpXvLzxiTxpqsic/8+/zJEL/+Jc6O9SPF8vPFWK3RVdpcNC1bbVhSXPHGOF3mhmMHiKnrYfz6ixSR2AIHlu10uxi4zgtXE9LSp23n6zwOzrGVVZ7gJ9PLfWj8cLvlrI220iK+ddQBqnfOTofEtpVWg2QUqUBze7j9q6hSzaHE+MZVOX2x7yWZMMbDeQIHu8UFGXBsdQdFlLH59jshUXqRruh2ghu3bCdU0elvsctSt3g8IUIvwsI8PHdf2VCt7qJ2t7tGUBi58218/6XLCl0W3y7MhhKy9kpPsP7Gpcfs+eon+bnpM0he6iIV2s9zRB8kRdzFwJ8LzCbviDQLpa167xkOZoJjYI2EXzSzMhvzhzssJR6jzj7ayA6jL1/S8ytWGmXXWtnkLoGR2Rk2XdmU5q90OA0E5qDq/bWn9RliT/M1cTezoEkucKiDMPCKwICFwsG9aiER24thBfvzocfX80kywvS87uuL4bYCMdveBhytyL0euDBGwuZ/4tpxVZ+tPpzU0UzWPZnIHOdft0qDT2e2Wa7Vw+G8uZGK3Ss1z6Kf2zTGMlyCuJ0mNDQkt5wMMBq+oLz6rzBTOdY9kDEKgPeo54HXVrP16cW2kfS36e43AvScn04IyxjoTVFW1lpcuj90ULcKZA9ZHTDizIqQGK/TNrQi9LjLFfcsFghkIrKUsjMTiEucqknt5PHsNAYeO7R9ZphFQGBix1I1kjkiYPKIaiFvWHL4J6V5JcX7j4Di9YFwstftpTUbjNd2552HTsvNjPPCxZvCCLh+QnDTHVmMWAdgcUMeUywy7RRcfHtqqn3Za0Q//pdAWR+3ZYoMWyLCXNGAuvVo8LOae6Tst+pGd6QIw0EUjN9YO8xX9JXfd5jWcz1tBq7HDjpCzdz3g9J/VAafQ45X8suVlLrh0/br6vMq9Nq64kKDToHx1jXxBlv6NWhkplcqOcYKs8v71wOJ3Ponw2T58AcgnekTWx9gDbfPMy472ur3wsKMyVOYhvnODFZDiPoC+pP6nXk/pJE7BjL3wBO9oBnlbNt3LwAVq6fK/u8my9GdJBDtR4WGlvvsIqAQGPHjSNL4cV/n4HgAH2GLZNGmoiav+f1KDh0kj/3PHpAGakDp20zI98LbnisDVHOiLXbW8DCl9iujfFx1XakHuhgTbqCnSl+XLbAmnWjXOb+OepHMuBZdrF7JOJetS2QlL8hkdOAxN7C9wLVLlarrM02/G7FYh58pc50llPsYiPopM4Mv5vXwwz4rYccmBfWYPbULB/pOFTbvaavROrr/gmWCVwF3+YXIL5theOjhnrCUXMyDL2H14PHDMAsR0idX6cux/LjwMnheJZSR7MZ2h3N1v9dKHWBS43Q33skD4JbOEaOsa3PwWdPZ0uEfV6D1Mup92fEnJ/CLISOwPK1JX+ww+Ld4mq493x9f6zDjkO+zJegV7xROJ7ON5DZe9yPkDqL0BEnNExo2kexndryi/mDk1YcAxutDHieXSy6w9FOBT7vwgm/q8vaPHSezqIydt7AHSPySWKcAmziMv6KAtKGlfaZ0vJ8DP+eafFERuZhmP3xNHgfl3nT4D1pWb5qnsM8ZgRFSgL7a2NIXSZ2jIdNBSfWsbNGOJ3aD7A8J1mWITGQED+SPHJPllDqAm7AkF4VVkt9gCH19x+pv+FKbMQ5ePsh/voxEec4N297spn/ewjzLk4LtWuReFfOOjwlbuS8enYB3ZlNmafu01n7vfal+tqtp4aWWj9wwpcxf18LCZyyNrSKrahm37p5drH/HGM3XwkPPs90lqus8dQ1KLNV6izTmujws/D4jVnw3G3p8MrU0zBzUpbc7MX25eYne0/4G/591WqPvffgejh2YgfJeleW31bP464T7Kx5deTghMQleubXvXRtDhV7QuJscNIcO2uE4+8XBFePeBDCQ2MIwfNC7lZhFWEPK+ACtGl9Fu4ZXwxXSASe0JFu2LJicwB8tiQEtuzXd1N675Fcy72sHtVV5J9jBpbDwB6VsO0A3SQGw/QsbD9ov87hNB+SmU6zeg2qR6hd8YCnknq+l0vO3fEMvvLDbm1aoPVWV58frXl1LGvroeq/bnte0VmO2Vs9R+6tTlsvvh17QLD5QAsY1Zc9Lzy4exn8ttV+tjW2ZQ2HvOnTAQdP4TVfyD0GqMy1ouzFZc0kUndOuRhyhZFz4bgtJyZp49x6msTFM0gJer2Veh1mO2tPeVmD14ycDgMun0Al9EJG29ViYQ8r4ERgePvDx3Jg18I0uFci9R4d2Te8awaXw+K3MmHm7QWa28Wwe5vWbCKY9XlL6HFrexhydzv4eV0gVzlhVrwu2KyPiW96yZ4FUzn/ttKfU+62I4X9PkaG3xFKvTrtELaQBh4BfucdJnW18u7ZgZ/cpyZsIzPgeWo9Pd9bUvrNmOuNuKwU/CjZ84qCpylvljc8dmZbsyuYuR7v2CmorvGAH53QqU0PV3hoRJwc5TKj6BLQAz4hsYgks8vt0utJ6nIYfoMz9lLvSOn4yW2wav0cmPfVZHju9X7w/sdJQqkLuJzQk9/JgFvHmBy6GTx2R6EmsV89uJy5sZckQv9qWQjJbs/Iaw6PfdDaQuy0+w2q9RiNUjfaeqzQfEqaj9V6qNyR6NfusL+JHT7lw32/aEq5G2baY3h9BKUeXbnZR7cytlYZSZ13r+4Uy1fr+1L5A52IUG2lzjtOPFI/w8iAV4DJciys3RXEXM/fpxZuHGqtrrEOfWC3Mub7Hc9iO739dSCQdGXTuu5of6s660F6qucWeTvt+8zjirFj7oeP3j4A896Rl4/fPQDPPLrIEC4ziNxnmZX7i+qwvBdcJFgpkfkqJRNeNY+Faj0sxFrFi0YuAs4k9J4cZc6LjSOx/7QmiJAyDaREjQGLMlfh5S9aEvIOtgmBK7uAYfZf1jd3+DNSicBM3kjkmfledhnvtmCF6xE4D4/d27CWHYkcST6mpU15Wa3zz6etXaztqcMQ/N5jfOLGjm3oEU/LgMfwe4DveSivaqaf1FUZ8O10KHXW+pgst5sxf775YAu4YUgxcZej2b4O6l5OiPxohi9R7UjoJExOcZPLlI5hEaeLG6r1T5e3Ju5yHSKr7d+P1n5V+tvu4wGwckcosZF1JhyN6raJiW+MtyYk9ySJ2LGDW4mjpB5n9N701un5jnPqq1Tfd2VkR2rX+9SRush8F0BgjXds67MwxEyUykAQCfVAqg8cPOF4Nu29E4qhZ6catjm6Djx2RwHMeC+Sur+Y7U7bVHqeF7X+HP+2elsA3DiCHjVAwv9lvX7PdPwomM2+llJcgvPqh9P0H7PDp7xhQPcq6uGJl97j1f/kW3N3re6xkWH456ikMK/lKfVqzW1gCL4jJyGuA6e3eoVE9ljaxurKxlP6zLI24GfAI9Aydu3uILh+MK/BSzV0lhaFyFm2tBv3al9fSOyf/dYa+nYuI+Vs7SPt69CxDWt2oTeknPaTFn9iROMKGM0XLmrBSn1rwEi6ROz6j1xC4lBnkLpev9yYyG5WZO7nGwgdMXnOJhNeZL43bSK/64ZiSBxURuam1f4G6gWxZZ8fvP1tGGzepy+BDRXsfyaw56nTc7zgR0mFY/LcLaPZofmbpf/RSJ03l56Ry1bFq7eySb17+2qHown1caejkro0AEBSdxlD10upe1GPi2LOEhmuPYd/IstbGjhWWK2n/sjtOaSOQFJH8rY3g6nlht8RmAWvzJ3bHmJeBjzi180h0KdTBbSxndLQUNC1qv8dz/KBHUf0J7DtOtaCLIiosBrw9ZYjOSwjGpeQusF84UT/d73EPkMfqcvx+tnu3FvsjjN+7DOEzB3JhBe49BEoEe4L/5cPN440yYM+zkRprfnmO6R3JST3yYSH3m5NyFgLY68oZyrpA6nekPR4rEVN40Dhg5m5TA4dLJEANxveJhwc1IKdsLX1gB+zcUosY36chMbVGeiq9bvFVRtyTjINzmI/ctpbNqcxlNSbgwkNZPwuUMcdvTtpu9ilatSqt9aYVz940tcqA95uYBZXxXWWsyJv1YfgKXUFX69oCY/dlGPX5EXPGCzjTHP48veIeh97VOUCToEOpS4T+gbzKMD4PXDABH/okCmarxEtV5sWkIQ+fTqLKF2FsHVDev2cx3PhwAkfEpLnkvqQMqbgffvbcKvwOM6b3zy6FIb0os+R499ZpE67oSZwPNrxfdNzvaBNBN3ylFbalnLSBwYm0PetWwP94C1K/VT9zEJwMIBlbUjguA18blR3NhrQMa5PZzqpepB58bOQW8iOlNAMaNSKXdNZziYD3laxtwrhh+CxrI1l+4od22jOchZizveGd/4XSYjd16eWeQxsQ++YTzF3SWRd7fpFDKPLz5zW1MUBte6lQei9AdPnnUToAgJ94ystwhp7aztyA8eM6e//m2FR0OrIoULa+IvlfwzSf/W+fBj3GD/y06NjNZN5N+21V5BYm24h9YaGnzUGKpgn0IahymmJbxl5XhKps3cNB0opaQ0j0sNp3txBSymea+k1SNxI5PhotBLXg2Pp3sQuluVpjhnwPFLPNbdXZZ1irW5trAx4ZaMRHFLXSpbTCsErxP70F7Fw07BCGNy9nDvARGOiDXuCYNvhFuLG1Xixx4tB5lj7NgtkJzkBAcPQpU01XHdVKQy/vJxkP9tWM5RJN/s/dgXAvF/CNH2+35qeZ/Ew92CLcS6hE+Xcu5LMhadz5q7bRp7jqmVbrNzCvvEh2b/7nWPHbVDPSti6n056GIIfbGMoo9yQsRPbmm02SiuvOXf8gKVwDSV1tJpFsibnmEJ4OICb9kqM269H2wx4WyCpb9rHrz3ed9wXetHI27zNiNCzzN7qadk+3HBNd46znJZdrB5SR2Di3IJVLWH5lhDSqc3SW91spITEj0uhyUvcwBo/5nvaKfOExD/wWnMVoRudLSjC740TLfzOw9sPZsN3L6fDbaNLiJFIrYp4LepbIulxQ02w6I10UvbEApqkjFJ7mNvckDEk/d73YXDj0zHwyHutub26EfdykuCQ8OuDzfuMU568ftcZuY7dbC3z8Azw/NwdgdW8up3JzblGcV0qBjR2pGje304x+jLgbddTo7WOenWWXWwcp4Xr6Vxv7nQTht8dAXrDY0MXdJQjy7YQ+F1a9p3wv6QJ3WjO6O2+DPg96DTnqSL0F0Fu3DLMlXuhlS2YkZUC23Ythu8WPQVLfntVsONFiM6SOk9+6xQMvbzcjoNrGfIaE7levS8PRvSjt0ecMNzEXP/gCW+45uG28P4P4ZK69Yef1wXBpKesw+uWUh3z71f0ruCo9LMsbuKSvnqeHtcrKfMkRL9yi73yS89lW44iYinZ8Rhax1r07h2qmSuywu88sObbHcWOQ/xBTX+De6TX6y541I9+wM3QkwGfmuVNvTCUP2mF4G07tqk3FYD90X3ZtrpoF8u6NvUkywloDHQKMyUOOmwopzmN0M3c7WUm9K+hkYXaf1s7BzZsWgBVVSZLaZKfXyCMv/ZZcaVdZAp93uOZEOhf1w/c7u5j8zf1/OZrErGPnu5n1TnMjnjU60vPZ/8QbhcSx7r0n9YEwi1jTJZQvGW1WuAbymipaEboHskbIwRI5Ph/XnjfzpDGJhyMZXqoyDGjHUPxWHpmlxxHGeF0pyS+UY1jVO9nVFmbVgZ817hqri0sDdh+FWFUC1ZEtqS0I8Po0wRarnIIrWS59lH8baRJSn1g9wq79RTweqvjvDqSN66Hr8HfMSyPKv50rsgwbyh++fVN2HdgPfj6BkJsdDx07tgfeiWMaGwmNMmEv2XXV4nUExKnQCOdO6+sNFmFlyolgs/MTqlryUr7ggjzmQYB5/8G9yiH3p0qif813uzwbpcn3fhyJUWxeV8AbNqvfyT6/F15FkLnJbOpFbT6nOO6d44tgY8WhdHfoBZHfLVWvLZ6G33/MOMc68StQpaq90e1vmmvzm5QqrIzVo05lrbprYNXIgwJHWroGfDS39+bkWfN3Q0wbMGMeGVgZLser4ubI2CZ1SjvF6Ph5d61bbXFcQ47u/XvVkWuj2V/BsILn0cYds1jvXpU2DnmscPvArZiZSp1rbI2jcYuB076ws2c48Qj9V82hpDlTImY764vCGf0Zwy2zSoduedo6g55Ob4DHrn/60ax69KCzV2Wqv+IV8Isd+4Vq+1q5w4DYAWFCAqLMrmkfsoJ5jORYdWk9y8aJkSHVxPrxKjwGvCXHtWJXsVlXpBd4A2pWX5wIM0fikzNL5oLu2eHSrgtsZAk/FiMWqDOsCUqXFZIiQNNRJm8+W1rbq9pctykdf7Vp5xJ6EdOecPG3S2Ip/e4f5mYyWxJQ0vtSL2WofBTTrJvsFpz3CxyJkTPYUx1eL4hQCWP5O0KHJKOEy/MPkD6nyNNXKikrlHWFq9RE//E5DMSqduHro32gP/nqB/JgGfZxWIIfu9x9vpoA1uG9e6+F6h2sZplbUVeVEbHgXRecXNIy2Ff04LMDWBGBmdUVJZKfJNlN8jq0rE/d3uYJ7bR+bldctdUszq3JfU4t5I6o+1qlw4DidFM544DoE10N2I401n6mzPh0/y8ROA1ENe6SiLxsxAqEXn7KPnL7mk+q2rCs0Vo4DnSzahH+wpIurKAuC0l/x0OVTXNGvVFfevoQrh9TBHYsaYH/deOsTXw/sOZ8MgHMVxiv3lUMXObG3YHwGMfRln+lX2mOfxnAr1NIyo1vJFnqcLGmTgvnGC/6Vher3BO1jeub1W25gBY9eiOAqcIqE1d6gkc4GxltF/VyoDHsraGkjpi+yFfpl1s/+78uWYscSPq3ObvNKJvCI6ne2sMTnU4y0kD3V6cufP2UdVwMpv+XcmXiBt7q2PCHBI8huNPSkReUdW47xtNAWNH30/UOqpzjBwjYmO6ctdxwZz6VF7r1UY9zHt71k6Xvt/A+FIY2rvISqE2pP55QHwZdIqpgq9WtIasMz6N8hjfOqpQIvUi6megfnYz0IVr9oxMuOWFdlBeSb/5XNalkjY+IPhsibXy/kT6/QZzkw/aIAD/riZ1ueTKZDX/ThLsNPp7o1q/oncl9Tz25JD66RwvaMtQ8vUdDNiTev3mQDHJDgcsWPKGChyfHzrpo6HUfVxiF7tuRwDXLhaT5Xa4oT5djWwNu9g+nSthAYRyt7Ev1Rd6oQNdLX1eHNU6i9QRL3wZJRi0kcHfLwjCQqPJguSOIhMJHufW3QguoSuknuZutd5YQJoIqIhEnUgFLMLj1D/ja1C53zYyHz5aEgWVBih2f5/zdVmt5vc9fKp+N8UeHSrhFonQ7T6fzmMQIBH7pOElsOB3+nx3lzY19PWl50fT7W9w360MISFXUB1XZbUB3Sutsqm3SSqSNTc/qEeFpFD9GWrdy+pjqgcEPHLG0LhC6rbchOY3mGVOq1d3jJzpyXLK+5WUe8KhEz6E/NORuKXnLCWuhUMn+QMIvWVteF6wvBDn4fHx6XnWkTetDPgR/cqZpI4GNa7A8Qz+AAi7tWnBtrc6/o4Lkj3mouw/4StusBchVqyZBwVFWeR+hKQeIxH6oH7jYHCYWzwWkrUIXSH12eBmX3ceCooy4PiJ7ZCRnQKpJ7dDZs5hePGxdU7xfi8p87JXlhzC40FNZjEta2D8VQXw/br6Jfe0iaiGK3ua4PIu5dDKXPPq6QF10wHS011H/GHlthBIcYDg/++6M8zPg3+uqPQk20dVzjoGk4YXU0kdO1zxBgQ0oOkMkjotO90WmISFIXjbUi8kqxIOuWKZGSbL0QZjPHI+cNwHruCE2XFA4EhSHFM9a/z/pmeMuZlgNQEOcGJb0b3DB/Zgf9YHbyyEUf3LiKWsrSK1JXXFMY5Vl46eBG8ubEn9n1YinZHAenXMdKcFLzBhNMDvPDMiRZT6cV/4JDmMJM1hKL5chM4vGrB6qqvn1PG6wGS5Y6k7NOfUnYgZel7kJTH/B5CQ2AcaYQb8sRPb4IPP7rRzHcvMPkwl9YaaCJzK8+OGnNXh2mKTF9kX7Dts9RI1Wag2MqBbGazcHgqFDiTP+Umq/K6xeRKZV9RNB9AIT3rSL74C+nergBVbg2Hhqpaa28Y5vvaMBB5M+vnv/EjYnyoPEGbemgeJA0upChpveGjQcdxmbp1VwsbjdyxT4vXhtsWc/4URhZiS5k3miPW4oOG8tR2hq3aqZ6dqagb8/lS285eHQaSOwOkBnKOnkQuv33p9gBnwseZ6f0ftYhWPeNv1ULlvt1Hn3/weDE9PKaAdOuJHcP/EQuIgaIvh/ezzC/C8lZYZr+CR1DvH1jAz4JHweb3V0S42+c9gwZAXIVg91Y+f2EH9YsRGd3XHbqZJXH1KH6mTO92qaRKxQ2MjdiXL3ZYIjp/cBj27j3LKe+YWeZNkOds3Vp7OXRoDOYV1Nzpf7/OQ0K4Crh9SQLLirQjXhjeuHlgE36/Vp9bbtKqGx27NIuYT2uGAul+vGVRCwuKfJPPfZ3hfertOJLzkv4IthI5494cI6N25kmSz0wY8fbpU2pE6b3d5ZVfo/a42JUGSPyKpvRSKl/jiDUEOnVucJ/53UjH3NTjfTiN1Yg7D2Xmj5tW1jGHwM/AS/hwl9YnD2W1iWaTO8nVnxgw3BsIDk4qIWyAN908sIuf525Uh5HesRceIjUXd2xzznYeND2Wr7WJzzniReXY0psFHbPqSmtk4c2IEnIdeCSPhoXu+lsTlDqtkOTfNqafpfWHdHUQmdixiny8tIY3hoMqJCjFQWJxpUVcxkfFObbuKZWmE1Bnz6MVl1jddzGzfdSwQDp7yhxkTM8kcOmtAgElzehX6zFuySMmcFRkyogC2vw7tY5JufAFk4Sl11vhATegKFq4Mhcdvz6MyNQnP6yR05Qma0pRRwpmfLA6DnRKpY+YzlrzZms7UF/93QxHMvL2QhNh5+8hyh9OqX2/b2phQMc+gBoHucY6Seow56c024503H4/nCm14aQMnq+kJm4EOlqnZKnU8h0js6DdQy4h2PHVnATx5Z4HlRFA6yVqA0zRGA10H/zniSwi8rFKEzgVkoNlMZ1W4HR3mwjXm07FT20vG74rupmrWDIVF7AmJcSDH7mc4m9zbMWrU1RjUd7x8cDsMIModid6ZyC3ygfg2FUzCRMJPy7G/GSK5/29DK/jPDdlMxsBQfVjgWc0Q/P1JORZrSL0JebZz1/cl5cH02W2ZZTGkhzMji59G6pv3BzDfGFW8Lqj2r0vbGth9xP59kMyN7NaFyva9GbkwuFclNexvO53AI+f0HC9mLfuQ3saExrfs5392rGNfvZVD+u2rZetY6ZG4z0nP8bPN/jEMPvjROsSNznKHJNXdnZEUN3pgOemE58jAijV18tGiUEgaapLVegM61uH8/LK/jL8HIJHvOebeLHyBxo/wMLc1IQqRuHmcrdEMDfaTU1jMfnDVS+aRwR5n7iWrRl2NsaOmkwVr1J1N6CQMV+Rt4R/1o/KkfSRbbZ/I9oOsM952PGalmlrxM2ljW1VDlzZV1iuqnqPZRPJfobBqRzCUV3ky3cUwbD92UAmfYIHtTmYLTBJKVWcJ12qvVst5TZe21U4/l2MGlsGKD04TW1Vd+1cLdeVuFOy38XK3RX0bv1h9/WzL2jys3y+hPf+4vXD3GbJMGmmSP7elIoD+ub7+NYT6fgqmXlfCVP40pDBC8zgweObjhrnAYYj+kfcjBbMINFXMlohdM3GDnXGCk/IHV10Gcji+ySC3yNueS1WshKY0POw8GsglTMyE52Hk5SVMMtx91B8en9dOIvUw+H5NS5g5ty2cKfayf6H596GX0edL/X3Pk5ewCB2tYmmwtcrkEbrJXI7Eeg/Fw9tZePiWAvj8mWzdSXfq/WPNjx9IZc+reoAx8+oY2laInTZwGKSRLIdz8h5UBX2e+vo12wK4HeymXldM5tat/nYtOy/BxKk8WL8zQCL2VvU6LiZpH+9+NZrkXAgINFHEgY5KNe00Upxrd7Jib0woKW/ObZKBLnM8oFLnEWZMS/6Nv3ObKiZh/rDOOqu9oroZfPZrBJWb8LFVyDnoHmdPAloNJlhe1Vh3iz7Z2F96wYoweOHzSPj3622Is5wtjlNUvfpXrGF3Ft5+KAdm3FJIPY/4gI1WajkjEmbXteM+9kzroT0YcBQ8O1vMCejOUeus+XZaYxdFQX+9PIQ5iMBB0bcvZcKzU/Nh+k2FsH5eml0Nu3q97Rp16Us3BhFix4EEr8pR/b+lfwXCNY+0FYQuIIDJ7NgenaPY9TrK4fz6hqZy1E7l+kIcJcyOPMALvyMwBG+1gs3cvJ8PWzliglx4EJ1QMexeUGJ/w0bjGezMhE0faOV43dtXwqE06xst1WtatZ6s1O1vzks2hpBFLzBzONqm1lh5D2eF39+angMTR9AjFJjB/n+vRJN676zfjzGPAZa1rdhsb/W4P5VPKj066PtMg3tWkLl5HDwgab74mbV6xXn1fyeVcNav1KxptyVIXlTl61+DYZqkyG2jGpayM+nvGIZXD1Rp21yyIVDX/iz9M4iY0ky+pgRG9C+31KNbtik92XHQF9bvCiDq3siObAIClwCGAWbDJyTOoJnR6CP1g6s2ShtAtd6nKRwxS4Y7JYMdH7HBi7qszRZFJq+6LHg6v1MRS5tvN6/Ia9zw195AidQLqG9EU+roNc0CKtjWocZkcudIN2M1qasT+rC+HZvEZBcYd8OeMLxUJnSKYdDBk95w0zOxlsxtDHH3YLRbZSluzEzHnujBjJA+T6m/fG8+DOlVQd7TNkpgS+qrtraQ34eRqf9/44rhy2X0wVV97F1RrT8xJwI+eSqnfgdeOtYmaX/n/hymexUkajSdwQWT8eLbyecCE+EEiQsIaAJvAPMlXp4FsnXsRuUfjrg4GK7UXdDJpl7ADHiqtDH/LUQjBF9o8qKt5njWb62+1dBFrpaxfkJclcPv0bOjMZnc2P2KdwyiDHQMQ4vS5+/Kp34+tFd9dHakVSmWunTMNoeBR84HOMlybSLPMeeu5e3WUNejEfH/1gZSFTd5n4hz8H832M9r43uPGVhOJV0tW9g121vAL+sD2VKf97uE1+a3pPdp1zmoQKtYXAShCzQmoE3svoPrSClbI0Uc4eaExPfrQ+puqV3PyEohdn2uJXVGBrxCRmH8+eAikxeTPCur2YecFnpX3tufE7Y/netjv6MqomoVYk8aGIKvZST0RRik1I+le1urUpsD2TfeOIe0h24utC+XMj9HVWsbrkZnORqh4xNeWdvvm/g10qzseWJewwBtDv+LpSFcEn3k1kKS3a8eGPz4aqas7imv1xOuf2Jua5nYPRz4fuZ5weQXY2DJxiAQELiUgIYzK9bOgy++eRhmvZEIL7w+Bj6b/xDpqd4IMUMidtLk3ZEubcNctXdI4lt2LoE//l5AjGfuvPF1GNRvgsuODpa12c1Pq0K6UWH8uVNswWpn52rG8Uw/XaRuW5feJoI/kDiVI8+r2w4IPEBOmLMNuSOpI3nTPh8rUc5RKKF1lsNelEEZ8Bi+nYDOaLXopVtrN4b4ihKqRsXNs4u9oncF1XCGakKjsou9ZkgZdT5eGUToJ8vmRK3fNMpENWxB8v7yuRy50Yt5SoBn2IJZ7nrwpETsa7YHwHPTzti1sVVvMyPfCxYsD4HFGwINMwgSEGhMQDty9cWPPvCFhVkw7KrJ3PWKS8vctcuYRJemj9Sx6N1Fndzm//gUIXRLwxLp4Y9NC3SRekigMU5T1WebQXWNJ/j60P3LIzWUenR4jfVqqvX3n2A7kxWUmhvKSOTkIZGTLRliIl1ldTOmWldI3XY9WalbDyawDSQxoGEAQ/A0ExpHYJcBb5Of0KWNMclyoweUMx32Vm8NoDZo0SJZlskMDgZO57LbsF4zREc/dK3MNTPe/S4MEgeVQ1ALflkeUeec7R064Q2rt+nv8bx2ewuyYCkbusrZKnPss17fULuAwEVD6pJSp31XtRq67Dlywq2KXZvU5dT5+a7ao/SsFDtiwjAIhuFjo7tx1+0T39FQtU4y4ClkwStru6JHCfh6X6CS2faUFlw3OSR1nnNcW0mtH0mnE21+sXWHOfVuo1K3e71GBnz76OoGkzoCQ/DoHkd7I+rf64H4OLblLSvszLNaxfV58+o/rg6CJyYXMgn25tGl8NMa63D0kF6OTzXgPr74WUt4/9G8eh8bbH7y2Iet67Uu+r7raZIjIHAp4upR90OnDv0lcj9C/N/TMw+T3uqNHCGeOgh9A7hwPn3kVVOs7q5onj/x+meIB7wrcSrXj3qzVx7jIu1v0n07m+D6wQXU7eFc+sodofXbmVpFPVZz9tfHoYQ8VOq227d8PgMz4I/Z9E233TUj1LrSMYznD0ADrR681hKpYO/XD6v488ev3JcPCaryNkxgQ6JnEi/HsAU9yWlJc3oJ/ZbnYhwqfxMQaGro07UD9e/oYIpNXYZeeYf0OALumfIhWRo7vDiE3huwKbuLwu6WA9xjFCxfPRe6dhoAIySCb6Ohzp2F4nIvashYeYJqPc382vYSwaNCT4irYBPB+lZQpOH5npHvzZTOpNmIeV49PPgsUe3tIquhrUT0qMQt8+m0/u8UHDzpR1e35icdoo0JjZPuVzYe81b5CS3PwtH0BpJOrc3H1tm/Xd0RzbZZDk+pY+b8D6sD4dYxJmokHdV68tsZ8KOk1nG+Gwkdw/W0OW/8v9ZUwKPvtyaveenfZ7iRe/X/fl4fCC9/0ZI7YBAQEJAUaxB/agpbsP7y65uW+wyq99iorjD8qsnu9IJ3gNRldY5WdFPdsUM4Onrt2fVuPzAlZV72/KD65bJOJrhcUuZR4TWWlqsshT5/VWtIzdIOZdvNl9uQ4WWdy+HKJ1MtvdWV9/T04Pd/p9WqI5RkOdq+s3qtOwosa6tVkY7tjmIIfuM/2tvp162SlK3hPC8mZn20SF9N9MAe7LA3zo/fPNpETWpEYkaFzSLFt74Jh7FDypm15Pj3e8cXa0YObMP0LHy5NARWbQmAR28rlOfZbWvlpR3fss8PVm0NgNXbAgxrzyog0NSB4Xf1gPl46g44Ji3DNRLm3IQNXjaEjglx88GF4fY9h1M158KPpm6DbbuWgJ9fIEy6/lmX7NepPD/q/LSF9CKrZDLkKEIkC6xZ79G+nHRo23/Sn3Rz437WDF/oamMVq7yFbVmbHRkx2sWywMuAx3aq6CyXV9QwcsA5dWAoaPy1b9dK+Jyx7sv35EL/7pWkr7ZlICP9fYdNFzcr73JVBjz+REtVFjmj+uVlwKOzHKvlKqr1t74Jg1fvP2Mlk/UqafylxORJkuH0Aon60dny/DiWsCnZ6RhxECTuGAYllJOBK1om4/fqZJY3nMj2hj92B4qDI2B9Dzu5w+4LjHPrjVClo3nFVC8Vob8o/Zzl8r0w0bOFrcraijLJjddfB6kP7d/LMFMbrFe3ZLoz3OU8OISO+xzdsoZ0ZsPXJdV4wqodofDnPnajnQJ0jmtTlwGvGxTCVK5DloMchuB5GfA4r95QUseWlvW1i42WyDyaYlAzoLu1+j6c5kPahNJOCCraSSNK4atf7fMZ2mj0QMeac14f9U+XhJIw/S2JJu0DYcP2GHa/6+WoeofHkcQFkTuOwT3K4Z4bCsmUVa35e1orjZV7tK8iv988ohje+DYC0nJEHoKAjAnXPUlC8JiwjQq9oDCL5HrxhKobgI6v6Cx3ystM6FPcQeg8/G/Za7BVInUihs036soqE2zdtRgG9XVNzTraxRJSZyhgD8YfWOoPw/RJVxVIJF8NP6yjt6HMyEf1WGYfEmf0Prd9b2zHejrXm7jMYQe3/JLmkJJGD/2fzPLhjg9IWduJhmfA/3PETyJnk/2hMtvFYrIcbV4d+6r371ZJPQbqPt+k3aft8VGdnxm3FsLWA/5WCWOJg8rgpXvyuWMkPa1Up78jtwK1I3aOZEcjmqkvRTtcuy7QMIzub4JHbzkjE7l5OZDqC62kwSupEKmVjZeevD0PZn4UDRVVIh+hKWCYJAR5QAJXkzgKzspKk8NC1YnqfLa5XTqBl0To7aARtlcdeeUUQupq+PkGcg+m0UC72Pg2FdT6Zx7hamFAfBlUSap9yV8tKaTurSnGFUI/fNqXNHlBEsc6dXysqNZ/I8or9uIOGiIcNKHpGCOr7tRMa7L6a08AXHuFiTkqujy+UneynLJafFwNaQqC2HbQjx5BMT9Btb5i9mlYtD6QqNvEwWXQo0ON3cDJdjqjbWt9nx+JfdM+f3h8cgF3nQOp3vBZcqjueXQBNi7rUgljh5ikxypp4C2bPZVWeErXmj/MXRROIkRqBPieh3uT6soQv1sdAkv/CpIGwfLrerSvhKfuyAN/3wuE2IdfVga/bQkWB1rADpj3hYubIZO5TOhW3Z+8wE0JcQrSMnMBKLX8bWK6kdERGgDg4wiJ5HsljNI8mEYZ0CByVHaxNLKoqvYkr8ku9CZz5QpJJMSVU/umq0l/aO9SSQUH2DnM8TLgZZXnDXMXR0JBacNDr2nZPvRBiY4M+A4SgV/Rsxwiw89Kyzm4rHMlWae80hOuf9y6RIQkyzEc9hDXXVkKP64JoX9uoEQ/bBQwKvY1OwJkExrO6GvSSJNVgiFv0IBPWJavNGDtOi7oKGfJnDfv5+nc5qR8Tu03L1B/PD8tTyL0MivFjQj0uwBXDyqDK3pVwN2vx0BuYd3xHjOgjOSJ4GuR0L9fYz0dc+CkH3y4qCUhdjLw7lYhSF2gsWIWjczVpD7MraSelcv83003PAN+Eok7UtaGSXdL128xSKl7U2/2ClkcOu0PS/62D6Ov2x0K3SViv3lYvpzcxkhmu3pAEcxdYk3qmAFfIQ0WAnwvUNUjPhpB6Jab2Qlfy7y67f514GTATxhaDIkDTZbkNWU1DKe38DtvpZTw+YotgURZ0RL6urapgcu7VsLuI9bHItpsI0sbEJjKrS0WFvwWUkfqtvkFtXwif+fbMHjsjkLqdAYS9IFU/WFytIil2cQKGIMX7sqDa4eYzPPhHvD7lhaQfaY5BEjX3DUSoeP3Bsn7qcn58MgH0VZRJOU6XbODfn62p9QJgrioGnGwmwjiolsbuj0iVJ2nzpPUHdlowDtjWmM92F06DnRbnTqipLy5tSi0UYg8Z7lDaQHwvw2tuIYonWOrJEVvr4Yz8rwtZGS7WtsIY282tvXqtmB1bFNC7LWUeeNOsfb7+PvmIOoxUI7PzNvypQFBXYczfD68bzl1+4jDp6yJFq1L125XbsoedqeMdh7xYeWWAHj3u3DySENwi/PirmcwcLD2yn/yYNUHp2D/96lk2fddKvzvtXS44Sq2SQ8O/K670kRO79HTPjDlvzHw6vwI+Gp5KMz5uSXc9FwbyCn0IoOzPp2rLNNBCLmfQS25MvQkf6Zle4sT1VRIPcZYUj+V5TRS1yR0hdQ3uPOAGm1+b/So61SuL/XvSAhY1sYlTInYT2T5MskMMaCbfY7AsQw/O3JSo2sb47qbHWAkwin7h6U/LFLnEaYtMAT/zxE/YL0U69W/fzkd7kkqJHOfP72SLnddoxF6Gv2G++TcCKsWo7zaAfwf1n3PeE++XlZKig/nvH9cHQhvLwyDcY/FwGV3xHGz3wXqcVcaWgpr5pyG8UNNcqmi6n/YU/2Ve/Ph5Xvotri3jalrN/vfr1vZuRViROjr30LheIY3/C2dN1TsFmiUnyKuH1JiidYcTPMVJ0ugMWG+HkJHeEkvXAAJiTOk533csadGm98bPeoqNpvQsEraIsOqJXXADs/+tT9YUgxVdoSiJKN1jrEfGDA94M0roqMcpGtnpXdrV0mauWAoMS6yGjbuCZSWIDuljvPgGDa3I75auQTo82UtHTpmLDL98tcwSW1lWh9O1XGNDj8nJzN58A11kv+k51Xg3PodL8TAW9Pz6tqSUhz2Sio84f3vw+DLZXXzqpi8JhLYHAdWIUwcLofDcQpEi9Bfuy/fEn1Zt9MfjpySB4fxbatheD/ZkXHcv0zSufSEt7+1vu76xsvfFfQ+OMZIrFy5NZAstvhkSTh53xOUig9MohvRtwz+77pC8pq0HG9YvllcC00B7S6e0PssvS9U0p+HmRV7H3GarYEZ8FBbRu8dKv0tpMU5LqmnZvlSCV15gjXsttBrF8vCM5MzoXv7Kks5oDLvfSrHB2hDvW2HAmBkPxM1ix8zgWkd21qHnaVb0nLUEKr1n9YGw82jS7hmOUwPAOmPmWe8YOEKNnkgsd/3ZhQMTKiAiSNKiUEL+rCnnPQh7UKxtI3VuU3AcaBj37N3nSHnhkfq0a3OwlNT6voiTH83EtbvtJ7ywNA7KnXEHVeXwHcrgyHrTHNzJKdaGkDIiW4b/3E8Gda2IgOBTYvuub5Quj6qrBLuDp70JaY0WlNTAhc/jBaBac4KvR9cdcoxUscsuoTEYebRwAxXHlSjRzYsc/76k7o3l2yiwmrg8Gn2TQaz4rPOeFtnw9tsBOfVM8/4qEjdh9jL+vteoBJe29Y6fNnV72F+3i6Kvt7yzcGSUjEx3eluG1MIT39s7Z7Uu1MVc9DBg6zWK+UObVoDApu/YcnSQ+9F6Tpv2w76k0XAuVBXHKCFL6ur26h+Zmtb6ZwuXBFsR+iIZX8FkZD8fyYUkd/x8YXPIswRAfpUzL/6lMNV0qKUteElg9M8K7e2kAbb2nPnPTtW2eWuXDekFK4dXArrd7eAjxa3Eif5EoaeaqkvFj4EMVHx0Kl9f4iJ7toYytl0KXUwp8c/IpH7fDOxT3XFDjiSVIBF/+gwhwd1+JVT6CcpyNjMYyxZY5W0yaSuTbBoFYukzhKy+D81qSOOpvtCn85yONI2FI82srze6odP+UH3OPuQP60FKxlYZfsQdYJZ8DQF3Uu68T0/NRs+XdqShOpHDzARIw8aoeNqOKfJAs57PvBODMx7PNPSepU6YLLZB5xHf/6z1iRcK9A4gCSOoXdyujz4pN4/oS4PZN1O9o106Z+BFlK/4SoTvP1tuMVkSH09YIXFvMezSVKmurQNlz7SgHPS8BKYsygcVm1j276i+dLanS1gf6qvNHj3IpnzA7pXkBp1BD6ilfLPf4SKk32JQsuiHJ3k9h9aD/sPrrcM/qKjusLYUfdDrx4j7V7vNDc59JPRqda9KDJ/r/RzmnmeXVlC3HngkczXS2SOhF5VZSJ2sQP7jmeOmHpLan2vQXP11WebQXWNJ/gqvus2ZBPSQtugJKvAB3q0r2CGnMOC7LdxNMOPkLpiF2u7WnzbKvjnWABTpNO4kUXqiB/XhsEr92Qx/z+oRwUM7nnaEsqvu8vWWlnS7jmGc/T80DYS+50vt4W7byiEW0YXk4Q41oDnyClv+G5VCFFxAo0HU68rhodvtu4pHxPBvr6CGErb7rtypjlk5XsRi2BEv/gq+GO39XWOqv3jJ7Kgc5sa8v3IPuMlredFriHMelecCrGsDQeYtNC7gtn/s1biOBWFPvDTxsqfDRW7IPWmi8zsw2Bzu5P+dgTCGL7vTnSTm0FEd71I3Vq5v2RW7rgMc9bebtixj2nVt+fAWvjmp6ckVWqyzBGjXeyGTQuk0dJ0l6n1uMgqqr96VJh2iVmqOgPe1p8de8ZTytr2pvrDjcMKmPXV2LGNRepK2F6PR73ldJ/0I+FqJG/bAQHLkpY2RvnlD/2GHV8sC4Mf1wRD33g5HI/zpmWVnqTuGG/UOw/7kecCjQM4fz5NIvNB0qMSDq9VjSIHSmp8Dmd9XrMbNTBJzhKValdtIXVlfXQnxMx2bOSD1xDmaSiIDDsLr96bCx3NZZX//Xcu3DarrUOfE6ejUKW3i6wh6h27HB5KE/PrlyK0LGKR1GnXLcv73ehqLitSRy6WRXc9Sb2O3FHyD5c2+DW4wX0OVXlFldm0RJV9XlCUyVwHy9o2GrgPp3L9ZFKnKGB8jIushLQc9pe+yORFVdMKGfr52KuYwtLmcCzDF7q0oc9dD+lpgqV/h1KNaDDrnTZVoFV7O2dRBHSIySCNXKiETvns6n+t2hoIm/Y5NqBC1b7xnxa62q8KmFVvwHnSQQ7h6pK7MQPL7cgcyRWjLTERbK/87Yf8YECCfC1jQx7F4tcWmE2PpW215gE8ttyFJQC7DvtZiF1JmPtiWagVoZMBeGFzeHh2FPz4cjoESMSPbodYr85T69T9TfEnVSO1teJ6a8q4bdKrMHTIZMiQyD3T3NDFjzOnbnQ1lw2SJR7uw3KSU+Cpe3MHV00DuROMU5Q6C2hA06XDAMvvnaXnD9/zDUy+8Q02qRtd1lbuZW1ZavMkVCMEX2RqbkeK6ies+8aWg4FAW1F5/fSJOWRu3eqmO6AY2rWuoW44v5g/hkMf7Ne/iSRqmbpjnGOAc5dvfdda3AWcDLRAPbH0BPz6XiaMvcKljSNIi9c12wLggx/D4LbnYuBf97aDThM6kcoCol5anSOkTMO6HXVRpcnXlDBfh1nvanRtWxcJwxC7QuwYyflpbQhzoIiJcupjJn83TPDYrfnw1v3Zuj+zh4e45pqyUkfESKp8YN8kmHD9U/DkjF/goXu/ZnOF85Q6oTbCwQmJvY0hdSUE4AZcN+ZB6J0wEmbc+420LJRIfiD39UZnwJeYa9VZJit65tVZZMjD1kOBcitW2gCgVi5te+muDBh3ZSEkXVUIT92eCXeMLqBvuhZ0hRBPZvvAPW+2Jd2ruPtv/gWV0YtfRMLbgtBdgv1oWWvO8u7Zsdql752Z1xzufSMKPvgpjDTRycyXB6tKQx0SJWpPn45CB0DSTc/cYGf+C1kktK7GfRMK4b6JReQ1ig2w2oDo6Om6SBO28+Xh7311g4hOZme5Xp0qCbH36lRlnehJwfDL627OFVWe4sIT0IW9zlXqCrFvkIj9YdYLvBzaHDraJCSmmTdsGLQyBola7zhQ9/aMnlM/lefHnZ9uH1kJfwA7mSYqvJo7L11oYp+G5VtCYeo1eVbJaGq0DD4nEXqRRVXQwu6yCvck5jN6gIr92c9ioEeHShjVz0Tq1DGMqWxy33FfYuKxZnsg1cxDwHnA+vrTuV7QLvKcQw1nnK3gyXfDnAG//SB98PjMvAhY/FYGeY5d9ha/kUGIHgm8a7saOawOcgY8fl/QhIYQu7nNLuZYDL28wuH9U4yVNu8PgDH9y8i+3j6mWPrd39KlTY3pE/OJPwPuA0a3RG/1SxNDdah0R+BklW5FcYANXRISk0C2ji2pP6nLMJzUjc4YHGbwyUJgvXpkWI0dKcuucvxkuX5dTDQut/zteIYfV60P7mFyzBqWksz2zcpwh/tDo4XsgRMiQaixIT2nOSF1PMfY893d3d8UxY6IbcWeV0e1Pv3tSHj1vjyi1mvN5K5EHhBLNwbCWwvD4XZVGB5D8EjoG3YHwMzbZAObLm21E1SVQXSOWdUjqadmepNGRVjX/tHMLPh2VQhsOSDnJWBJ5/VXlEKP9lVwwbxDcxe3FBfcJQpejTomyL0zZyJ0xNr0qHiydOzQH8JDY9ji1Pkq3Y7qzKp9mJrY60PqcUbv2UbOnLoaSp36H5sWwKTrn4FBfScwXxssnbASAwcLaBdLyJvCzL7eF6TlPDGaoan0K3uWUgnXUtedyfeZ/mRpa3j0pixo29p6UKE3O33V1mD4c68oCbtUsGmvH1zZRx7ktY1kkzom1N0zvhhuHl0qkb+sPEvKPInPPXamw97yRkBP+F0B1qhvn94WkoaaYES/uu8nhvaTJYW+M8WPGpFCYCXErsO+cHnXKqLqr5UI+LdN9Ov66kF14fW9x+v27+0fWsHbD2RDgE8tafLy6C1n4MIFc2MhHFxckB8x5P7l8jDhKncJg1ejfvzEdvKYenIHqVXHKxKviwfv+Ro6d+jvbqVu9TFAdoO9TPmDY5NF8gR9nDP2jHdAkMyXr54Dz70+An5bMxcqK03w+9q59T5h9VPqPsymLMBQ693blcO91/GTcrantIBCE//miiYz7/0vGtLz7JuV8LJzMeT+/k+RsHCVUBuXErA/uzJ2Y4Xg0Rp31zdp8PjkQmkwWJfzEdziAtw0qhTWzE0nrzEKSiOdgQnaESUMpaPV77T/xliW5z6NsBA6Uefm+XYcrO48XPf3z5LDLIT/6C0F0LmN/We4RiL0awbL9xNM+vx7X12FwIlMH7j/7RjYp84ZsZmmQie5R+ZEwx//BIqL7VImdU7uVRaWsnnYXx8sQneTUq8j9oTEF+ur1Gc5a6/wgLDC5u9+PNlSL6gcaCxp27prMVOt4wnTGwHQgxyWXaz5CdarY1kbzq9HhddA3y4miJYeWdmzuFpVtaSatusztkBif/enaBjVtwRG9SuRa9EZBeg4D7h6RzD8tTcQKqqFx/mlgME9K4giR3Wt7u9OS5bDkHzyOxkQHHABiiVl/vwnrWDF5gDy+91JxUS9I7l/9Xw2DJxmzBgdlbai0mNanbUKydcHOMeOKC231h27SBjeH4ZeVkHmyhe+mAHLNwUSFY8D3Mu6VBIDGkv4fFG4nRlSblFzeGJeFESEniVqvWeHKjiR5U0GAGK6qemAVyUVFhoDoSHRUFhUZ8jViUPoCCc2c9GDqYC+Mg6RekIi+rImOWuPyAFhHLORV02Bb/73tBUhYpkbb37D6LI2xQMegJ6MNrxPEVw7sID8rjRRoTK5Kmye/Fe4XbmbFrH/ujkM1u4KJlaxige8ZaBT4gWHT/vBmRJh2HIpAI2AHr2tEIb0qgRPT/k8ozHL58kh8jw0zqlH2s9hPz65gJB2ickTRt7f1hKexyS7Fz+THdT+LZE7huRvlFT7z2sbPjWD9rCjzDXs6CzHInXs1Iahd6xFn/pyNJlntwXWpqMHPJL0zhT7qalZX0TAp09lQedYuZ792iFldlax+JV4fWErrk0s9lXHxbZZkUATUeqcaG7iyAfIUllZSgTl8RM7wdeXH7lJy3Irqcc5ptRlQp/vzD3iHZAh/SeQsHuhpM6xTv3a0Q9qZsMbHX4vKW/O9Sb389a2wVQi5ajQl24Kh51H6hfeQ3LfczyALAKXJiaNKIV3Z+TaDQ5RbWNIXbmgena0nvbBefRbx8itUH9cE0Sdb3/3uzBC6oTcxxUbQuqH0rwtGfAYgmdlwCOICY20g5gwh8Ru8XaXEN3yLLxyb10/dbQItgXWod/+Yhv49w2FMPYKE0SFn7dMQ2FS3MY9/vDzumBdDV0EmiZ66yx7Pn5yByxZ/gbERHUjDV2OndjBDMEbGRluCLRJXa6Hm+3sHdmgcUCm3Pw6edRb2mZ0rTriVK6vnbOcmt95pK8ks6Fl7E9/tHJIoQs0HG0iqsmCJYDd4iotUzmogDGykZLmKw2yAhyuEHAEaId7+9XFxFENydnDU/a2n/iUtY3p6AFl8PZDueQaws50Xy0LgS+WhhClffXgMnjizgLo2anGcnn16FhtCcn3UIXjMeRuNTjuVWFpRYth7aCAWul5jSGfTTGgQWBZGwvJG4PgwUlFxNsdM99XzzkNyRsCSQQiuuU5GCepeIWg/9jlbzXPbovPl4WRRUDAYVmrM5KLSXJFxVkkDI+NXZROgB++ccDqdW5KklNjjzapJyQGm9V5kiv2SOugOFKnjsBa9XbRrR3qAqe5j2VedoStx18dL4RDp/zh7/3BcCJbhPpchV4dy6FPp3Lik4+WoYTEzUSuPEdi9fSogmF9TNLzfFi0IRR+3xJsOLm/eHcujLvKZBlIkOul1lzSZYPn/y/fcm3d/UoU6QGvADPXN+/zgw2fnIZ2UXJZm5rUMeyu4JYxpWT+HEm8jTlZztJmlDzK3Xm6t6+GQycbVoutDrfz7GIRD74TScxnsKwNG73cObbErv3p+p3+8PynEeIiFnAK9EZys3IO2/0tPDTanlHdlySnYBaf1GVC3wByurxL4AwnHhyNGUnqcgZ8mWbfb/y98qwnnMz2hZM5fnAozR+KyoQydzXuvT7XTOC1lnN05LQvpJv71eMpaxdZLZFaFWkQgi+7cXgRDOxWDi9+Fc0l9gC/88SdrFNsNUSFnyNZ03uPYR/vQBIeVuO6K0rh+ivkkDi6on26JIx0I0NCi7Gp60aVS+aTpefbDvhZEboCVOxYyz33iVxSvtKWMq/uQUi9TvXiBjFpDskfFzSMQdfALfuNG2Ti/g7oUQnd4/jqH+fRR09vS8gcy9pwfh13OFPaJ/SEV5e2CQg4Aw57mai6utAS5pzWclUfpsLBVUvZpO4GQlfA69amxpadiyEzKwUm3fCs5okzcp4jl5EBX1XjSbLjMfsd1Xx2oTfkFAoXKndDaYiD6vjrFa1gy0H23PE1A4thkkTomFGNjTyevD0HXvySnog5qn8p3De+gDidqVX/v/qUw13XF8JD78XA8Yy68z/0MjmBDEPMd78WA2WqOeQdNuRlKQmTLi4e4WKtuoIre1fCWwrhl3laLk58zWZp2SQpe5xbd7ZJTUa+FwxQDU5YvdXlY9EMPloURhYBAVcDm37pwQN3z5cjUVmHyfx6hvRII3U3Zb6nmQndqn8ZTanPdgehWw4Mp2pgy47FsHzNXCgqziQ3rl4Jo7hheaPn1Qlx5/pCSbkXUe34e65E4FVnRdlYY0RGvrelyx2tm50aK7aFwKFTfvDWfRnk2kqQ1Pvwy0vhj93WA4FR/Uph5i35hMxzC5vDmh1ysiOWUuGCin/OzEy48Zl2FsXewmx/ij/UhE6D2ut86wE2qROCpmTA71eVuwW3OA9vfxvusuNN5tWHm+w+h4BAYwIakzlaHYVNXWIY7VaJUndt+D0ZcGpcpc7V8LRR6e3ADe1VtUIYZwoz4JlXR8CCn54mGfBKRNHVBjTVEnkvXBsNy7ZEwLbDwaQlqyD0xguFyDEEHdtK22jlVI4PLFgRbglZ3zSiyOr/WNf8yC3yfPem/f4w+eV28M2KMLI88kEMzF0km/wgsT988xnLemXm/uBI7u89nA1d2/L3Rbm+ca6bB6LWpRejuQxmvROlXt7MouJ7dKwhyXGuwi9/BMINM9tA50mduNnvAgLuhNG8gHBB5nuamZtDJDIfzyJ0e1J3UVKco6OdlmGxlpuduktZRnYKFBRlsEMs0mgsOFCUfTVlpS7Dg3S004Pft4ZYWtRiU4+4qDpiHXeVbK+MJiXvfm+fxLXojxA4li6/59jBJoncZaL9YU2I5cIddnk5/PhqOvz92Qn46rlMuOOaYqs2pKSTmTkFQIvU1VD6qyM+XVxXBrZgVjaV2HEQMPP2Atj2dZplQNBQYEidF3IXEGgM0DPFu/S3N+Dr76aTOnVNtnVu6B1rT2dIJN5eWhZo9VKnkXqIW0mdk2wwuP94y3M0nZl80+vwzqyd0vPYBp9AgUuX1JVBYHjQWd3r7UgJsKj1gd3r/Ml7dqgkbItd6mjdvTBpTplLx9Uv6yrPj+8+4gf3vhlDeoAr/0TvcjRZeWryGVj8ZjrEm21R1aVhYwaWcwk3qMV5ZcxiU8rWAn5YLU8LYOnckrczYd28U/DYHQXwwcwcWPvRKTiy6IRE6oUkK/7qweXiYhEQpG4Gms38tWUhHEhZD6+8MwY2bvqm3rxlAKFjs5YPHFnJi7IRtwEbsOCohzbfga5yew+sJY+D+09wKNSydP0WcSU3QaBJD7rstQo5Z5lb1wNs4nHtELnMSq3U20fLDmZoM4rtaHt3wpa0Z0n2O86ny+VxdZ4EndvUwF/m6lEk9usfi4MubatheN9yQuIj+8tkivXZcx/PhlEPxpHSsG0H/YmbHM5L33V9Mcz+0X5eHNV3jw7y/uDI3NYu9qF3IqG0rBncM6GYDCIwFN+jQ6GlcQkCa9U/Tw6FlVtENEvA/cCoKt6vkXQxkQ15IHndZpi/dI2hzbm0cq2W/f5G3T2kygRLf38Ttu9OhtsmvkadV3fifDq2Vd3r6Eq2pD4f5Ho3tyl2HPXQSN3fLwiee3RpvUZlL4nvS5NFYakXMZxBkkW1rpUwh8DyNAUBvvYJX0lXlcD4oSVynbtHrVXd+bFMb0iV1Dom0f2+xd4x8OhpH7IgMOz+9fOZ0C2uhpSxIdFjudeHP4URUseNzri10GxAE2pF6AtelBsFped4kXr1Hp3sQ/XPfdIKPl0SAreMLoXBveoarWBG/IETPqTmXUDAHRhqJu4+8R0IkSPRorcI7f496/47oM+kBwwpT+7NeB81oqPiSXc2hFLllJV9BPz86A6gG5wzn77BNqu9fqSO8fqExBngZEtYLqlLo56kkUNcFmoRuLRxNMPPotJjI2p0kfqhNHqSl6LA8UueV+gFeUVexDcce3QjiadmOjafjHPQ63e2kEi9kGw7yJwxjq1MMels0ggTebMX7z4Dd48rJk1TsCa9beQ5MpDAuXPcn/9MKrazi1WAWfKuzIAXELAlURRpCnHjo6OZ50jCSOzTnnu34SpdR5LcuLFPQY9uIyD5tzcImSMSR95PmrywhKgTkFzfFe1L2nAyPiGx2EzsLlfszhj14IW11/2OPwJuQEGpl4WQ0SZ2rw6//FYhZ6kOgftP+ELvTvIA4YF3Yqnz6o4C1bryXuqOZE/MiQRTeTO464Zi8j+0d8Usd0/zSz6TCP35T1uRuXSsQz+dI8yNBNwHWujcSEFlVIMuvfvUsf0AmPngYvhz8zewfVcyDB0ymUnoRk4NqDdtHKnLxL5UIvY4wKw7OY0+zlUXhzNKA/BEClJvmqjLgJdIvZW+bPJ2kTUWZyHs4qXgZJYP9Ooom9mM6m+CpX/pH/NGhZ+FoZeXw7F0H8g64yX7nP+rFMYPleu6MevdtmPZf79qBV8vDyHNXTB8juF+NKRZtbUFHDwhv1ZxiBMQcBX0hs6NhFHe6o4ONP415E6yMJn3sNN4ZRjSoXGkLhM7ps6/RJaExN4gl7u5hOBx9GNkLSGeyA++TRbfxiaIzHwfMz97QHjQOV3r9I8vt6j7AyfqWn8u+ztYLmuT/n5HYhEJvZ/IsidULGW7cWQJ/LUnwJINjwlyM289Y0mmU3vAZ+V7wbMf05UI9k/HRLnZP4pzKeBaGBE6NwrJBiQ7tzNHEDTvGdmHISYqXtc2NzivPn2YmX8NJHVrgscMvL1mgh8KTnadwwNlNKkLNGG1nucNbVvXkDl1LbQMOQv9JFJHsq2o8oTtKXXhelTt368JhdvHFJEGMfMey4DkP4Nh84EAEiLHTPg+XSqJXSyWrKE6f22BfBNB4xm0icV581q5j4qkzL3h17+CYOmfgVbtRwUEXAlnh86N4gRX8ECWROizP54IHdr1hxvGPqVJ7k4sZxtGuLYeyXJeDr+V/CaXSW/4NTjJfc6RkMbR1G0QHhbDrVfH0JCYV2+6wHl1NJ9BddylTSUcTacnwvn7nIeZN+dYMt5/22zfre2HNWEQGXYORvc3EXLGLPgJw4qtPODxORrUYKhdwW+bgsgiIOBOuCN03lBgmbMRme96SH3ZCrmcLTVtB7z/0US4cvBkSBxxP/j52X93cUrAyZwyX+LZPnoMZxpG6nXkPk16Q3AGsWuNyioqS2HLziXwx98LoLA4E64d9SBcO3q65gkVpN5ElXq+D/TpJLuqySY09qTetnU13Dsuj8yno0o/le0Nv20Jpm7v/Z8iYMvBAJicWAidYs3qH33dJWW/77gfs1ubgICr0JhC5w1F8vrNxkhfDVLf+U8ynEjbUfcHaYAum9Csg+ceW+OU6IEG4kBOWB/vGlKvI/ZhYPA8O47KWCY08398ihA6KiJzO2jYumuJJqljmZyYV2+awG5thHcl4g0Ptp5Xv6xLOfSTlqv6mORrCq+/HG94eT6/9erWAwFkERBwJy6G0LmzRZ4e6JlP73dZEmkW9qdE5FWVJosbZf/Lk5y2XzqQJHHsEpC7selS7F4GvCky5QxnnMipMaOZ/1eXHGGTl70H10LvhFENCr0IXJpQytoQIy8vgfi2lSTEjqpcCZkj0HRm1fZgWLxRtAMVaHy4GEPnDQWGuI1wBE0aMVjX60YPfwD69kkiYXi0iQ0LiYbEEQ+4k9RlYkczGhTQOoi9cZN6kj2powc8KnWF2cPCYmDEVVOgc4cBmtscJ51YYRnb9FBY2tySAe/nc8HSY10xkzmV6w2b9gfCX3sDoaJahMwF3ItLKXTeGFS6o6IOTWam3jYHUk9uB1/fIOZgw8XTuX30ErsRpB7nypPZteNA0tCla6cBMLjfeG4/ddqJFaTeNHEs3Zcky2WeaQ5VNZ6Qnu8DGXk+cDrPGwpKhHGLgOvRFELnDVaMBt2v6+NSigY0zh5s1IvYMVHdyaQ+zBl7z5tXf+3Z9Va/Ywb8tt1LYOJ1zxCPeHYIZgg88uan4pvSBDF7UbQ4CAJuQ1MMnRtD6g1PkhunM/TujghCvYg9IXEKcX51CqknJGJ6sNN6sPPm1W0z4HFaFEPwg/qyO7jhAAETJowojxAQEBCwhQidG0jo6zYbYsGqN/qB2e9rN3wE112N3u8jGyupI2ZJi5NI3ckd3XCURptXR0J/5Pn+Vhnw+AMJnkfqslofLLLgBQQEGgQROneFSjco9D5CO/ReWVUKy1e+QVqtfvPDQ9Ahrj/cNP5VahMXjCC7uTw6ThLU7SS1fspYUscQgBMS5PSMhjDE3qXjADh2YrtV0w209ysoyuAa0YjSNgEBAUcgQufuIvWGh96VyIkW1m6YB5XVJsvvaD7z+ntj4PprnrTzfnezSq8jdgADSV32gp/t7L3G0AvLB35Iv/GE1NUZ8IP6jufOqSNwNI2jbCd11hEQELhIIULnjYjQDQq90yK9NKAyP3h4HRQVZdX90QOoSr1RkDrHPtZxUpcV+nzXjda2UEl9cP8JcFQidUcz4IlaHzEEFixdI745AgJNEO1U4XK1AhdoTCrdmNC73imRhPiRhNj/3rIQ1kiqHcUi/k6bWzfK4a4B4Apqj8ZM6Mroec8v8wwfBY5/+GXxzREQuMSBofM+NgpchM4bP0IGT2ywUsfBW9rqBQ6vh2Zmv654A66/5ik7pY4qffi0J9x5aLDPOrdWXb9Sl0Pu8139CTAhgVXaVl/gvLoIwQsIXDpQh87VyWsCF6FKNyj0nlTPUjYk8im3zWHum5tQTBT6wVWa7VgdCb/Pd9en0bKMrd8JFyF4AYGLDSJ0fuljvkH3Zb3z6Y5ykavpD2TX1vnGer8nJI4DJ/ZP1xy5MUrbGjqKE6QuINB4IULnTQ9Geb337mr8YM9FpWxpIM+Zb5BIfG99NqBXqSe580TjScaTbeQXWoTgBQQaB0ToXOBiUOlOTpDD8Dp2Ylva0A3pJfU4d59sDHvo9e9FcxqEVnmbCMELCLgOInQuoIXZC43xENEynNm9Nxk2bfsG7r5zPvj5BukbcCQ7jSs0k9+cQepp7j7ZWOKgRepI5ugqt2HzAhh+xRQYO4rfY33G5CRB6gICToAInQvUR7gZYeFNzII4kZ7snMPw+2rZPe7tD8cQYo+OjOcToPNC78VGErojpD6fhAbcSuoY+pjJJPP1EpkjoVdJJwrbaa5YNxcG9h3PdZfDm43wghcQqD9E6FzAKBilhKeOY4fe0Q72u5+nE55A4OPczybCxBteJX3U+fzjnOCEkYSun9TRvSYh0a3EjnPfWE5AU+s/L3sNtu5aQoruPcyV99g/G4n9jklv8C+ApNHw0rxvxTdKQIADEToXcCYwZ8qIqCnmSfFK2TDUPmTAZEmpv2n190VLn4UoSa2zFLsTQ+8bjN6g/pK2g6umScSeBnITFzepdXoI/rrRD8LWnUusrHQ6tx8AAy8fr2tUJ0hdQKAOInQu4HKVbtA0KM6la12rQwbeCSEhMfCLROSVZsV+xaDJTEJ3ctb7HveRukzsL6kUO8YqXFrmhiGQ4tJ77U5aeFgsDOo3HrZJah3br46VSL5LB33WsXjjwn67Sw2yJRQQuFggQucCjQVGJcjpzXrv3nUkhN4ZA4uWPQO+PkFwXeJTXN5xIpBDNxq5QY8GrS33U+9jXoaBC0rfvn5lJvXEFRRmQEFRppUP/LZdi0mv9WtG8hPmMLQy7bl3xTdL4JKECJ0LNGYYZdtdH1tYnGNH8DLg+0y835lKHefUH2k8pG5P8u1ADs9PddYRQFWd/OGL3NfsPbgWfln+GhRJhI5z7NPv/kZS8HzlboTXsICAuyFC5wIXG4ZNewI2GuDU9v6T98KMyfwp15Qj6yAkOIbMn+sBht7bk5YnTgNmv8e5I/tdH+Sm7dPMIXqMp4QYfQR4RjSo1r/5+Wk4jn3WPeqS5laum6tJ6ji3LvqsC1wsEKFzgUsBSJobDbJe5WW9K4T+/aKHwNcnECZc/yp06zpSW0YvXOLsQxBiFsKGqXUvp+ymnC2PofgNztg8JlXQRmR+fkGQkZVi+b3WHIookBQ7lr3xzGhwe4LUBRobROhc4FLGLIOSlKdIhM6LSGXnHoYly58lfFBVbYLvfn4IRvzrfml5gLvdZNfkWs2Q+HKDEW5yziP1OmJHljR8nh3nwGmkjqQ94sop8PvauYTQW4bEwNhRD8LAvhM0tykS5gTcDRE6F2hqKn2Bi2xhiyVhZ4v1f84jRjS338TuyOZCD5P5El8Oq6/fu2tI3XxcnEHqmLSw53AqVbEMl0gda9YxGx5d5bSsYq0uDGm0J0hdwNlQmk2oFbgInQs0NRgV2sbvE36PeMBQO5axff3tVFLGhoodhV/7dv3drdIVYBh+gxHE7uHU3UxIHApOCsE/fEcSzH7qP4ZvN27MFOEwJ2AI1KFztQIXEGjqwLyouMQphiQnsyqiaMBs968kYs/JOQJ9eo2DiTe8xty/0CGT3BLAAKwma0DinLOVeoizNozz6s4g9Vn33yHK2wQcArpYKeFydfKagIAAQ6V/m2wIoeN3z5GObFi69sDdi2Hz9m/gsl5JXH5xE+JAbr06rbGSutPq1vGCwLl1R06oVrIc2eERg2GGaMkqwIAInQvwkJ6xlnQA27LjD3jw379AbEy8OCgUFWxU6F2rhI2FIQPu5IvGZLc2+poKCYmzzNVkjYjUZWMap5rR4GhKD6mjCc2K9XMhNioe7r5jHj+0ENSCXCjCOrZpQ4TOBRxBYcEvcPjIXDidmQV5ZzzBy8sDMrIPC1J3skqfcYc2xVRVlcLWHQth2FUP6Nou5ms50WzGEUH8QWNT6rPAieF3BNY3YgYlSykdO7ENvv35aeIqhzXraEZzXPpbJx0164LUmwZE6FygIaiu2gq52U9CqSkTzp3zAK9mSOi10KyZB2RJpC7gPJWux+cdkfzbM3D46HpIOboOpt6u3T/dKMvaBqLe3OkcUpcT5Ga4ZNQnXSC0ufWtOxfDwkVPW3VuQ6yUFPuDGqSON3esexS91i8tiNC5gJEoOvNfOFs1Hy6c9wBP6R7j5WVemslLVo4gdWepdKIa779D8zUr174Oh4+tJ89zcg/D+x+NgWl3zIeo1vHMQcfFft83ntQTEnuDXMrmEmAIHk+u7YitV8Io8Fv+mqVvLiIsJAYGXD5e9wUjSP3ihAidC9QXVdWlsHzlG+DvHwNjR7PDtSkpT0Az+AVCg2TR4OmpkHktCb2jWi8uyRIH1EkqHUWXngF5cYmqPt1DNp7Bsrbx19Ed5eY3nnt+WuMgdXke3Sn2sCyQPuvrt9jNrWNCHNasr1g7F/x8A2Hidc/oMqERav3igQidCxiJmhqJ0FdMgaOpR6CgyAN6dh8JbShz4lt3vAKFBYulwaP8u6dnrUTsHtCsmbVSL6QYngiV7jqVjrhl4lzY8PdHsOGvjxRel+/vjPr0RhJ63wMHVy2o78pGK3U8InGuPgJoNUhLmCPmM75BEpmPd8iERsGMyUmC1BsJROhcwJk4d64Udu6cDDXVh6FFgCdU1wD8sOhZeOLhX6xeV1SSATt2L4C2EqHXnPWAC7W1JPSOSr2ZWa038zKr9Wa14sAqsjMz1+UqXcGwKx+AkKBoSeC9AVVVZST8TptXd7GDHAsboIEJ5saRegJpZTPMHUcBT8SGHfvsVJqi1usLJBG07jSq4YCANkToXMAdyEx/EmovHIIW/hKhV9dK5A5wpjgFflv9EVw7pi4Mv/OfJVBzDuDsOQ9pqYXzF2RCV4hdrdS9vMRxVQsvo1S6I2XMlnt5r/EQ2TpeGpRlMefTZ7u390cyWRqg0J2h1Ge584jgKNAZoVcM8wyf9oT4VhoMEToXaCwoLvwQqirXSCLAQ1LotUSl41JV4wGbty2E4VdNtkT6Uo5th7OSQq85KxH/WYDzEsE3b64o9Vozmctz6i0CgsTBBWM93lFk1fc+Edm6G1lowDI2F4m3NLMax8c9ePmRPikGwhhSl/uox7nzwkHPdl55W32BF5BQ6w2DCJ0LNFacO3sIKss+AO/mHnD+vETE/mAmdYXgS+HPTQvh6lGyWj9/HgiZnz1Xt/j4yMlyHkjszeqUeo9uI8UBRmVtoEOn3rl0x0Wh01V6MhG+BjRscZVSj2sMFw+GeOa/OlOodTdBhM4F3I1tO5PB1zcQevfQR6i5Oa+At0TAzSVlfU4idj9fVNhgUexV0rL3QLKF1Dt3GADbdm2Tif2sTPCYfKWE4OWyNnlO/arBDzT584HTokYJIpxLd0Y0z8hIAgXFOK4xqq2qK0m9UQBPzOwn7zW8XSVeSCITvg4idC7QWLF520eQlZMFi5fHwNUj74fB/dk5R6WlhyAzeztEt5JD6Dg/jko8wA/IvHp1NZCloCiT1JxHR8aT7o9/b51DCP2sWbFjOpynR61E7B6WZLl+lz8AQUExQqU/ezGo9CXOPAQzXEnoRpL6nsZyEWGyg6Mnv7KyFPw0suObat26CJ0LXCwoKEyRCDUDWoYBVFRlwJJfn4VV6+bB+OuepCr3o6mLobgEwN8XICJMVuvnMQx/QVHrdWF4LHdDtAyLhRvHvQF/b3mSJMxhBnzthVqLUsfwe9s246Fbt+lN/nwgWRqVTe5oxrtuGV1a5sza9A1GJL65h9SxTVxC4mxwkYuc1oWEfsB61DqS+Vc/PCB9qQNh2m18T/hLvW5dhM4FLnYUF22XSPcCVFZ5QEUlNnCqlZYM+O7nh2DnPyPg1kmvWpW2ns44DOck0kZS9/ethcAAIIR+/rz8t2pVGB7JWkHf3hPI/48cfUVS66Vw7oJM6N4+MRAe8RCEhk0U50Iiy1kGWm07TaUbWDtPgVvS6Y0Lvx9c9YhE7OBuYscTpEetZ2anwEdfTiYOQzgptuqPOZA4fLrmhZW8fvNF3cFNhM4FLlVcOH8IWoVJRF6lELqHNHBH1V4LGZnr4ItvMuHfU+pqlMvKgfzfTyJoPz8P8PWpldS6ROzNZXIPsCTNSUq+pXUZVLeuE8hSXn4ImnmWQvPmsdDMK1acBDNmvPmpYffJhyWRplel5+SlwM9LpsON4+dAZEQ3zYGHk0PvbolgGzunLhP7fDOx42RWiDs+lJZa37BpAST//hohc0+zxdDqP+ZCj/hREBPVjavWL6YObkNV4XIROhe41IGh9/DQWvCvrpXI2qzWq2Ryx+eVlSmwKHk6TL5FjoievxAk3djNSt2vljyGh+L8uqzWkeQxG97Xtxv4+NCn5wICuosDb3t/3bHPsIgmihC9Kj1XIvRvf5pKrMEX/jgVJt88n1nC5gKVjugjLRtd/j0wfItyyv40sliXuvUxPx9mfu4WtY491Veun0OSW5DPyaP0Y8yIB7mEbhmBSoMFHDQ0JrWOofM+8R3sFLiAQFNCM88MQs7ezWX1jc8rqsxq3azcTabtkJG1HWKjB0B851Fw5PhaWamTRc5+9/dDtY7GMh7keVQM28AqL38bGQCcOw8Q2CIGQoJFcpyRyXEoovQmPiOhV1fLvT7w8RuJ2O+8hU7sLlDp4GyeY8HDLWddJvvZ4MR+6zjCS1u1gHpBYEvWuV/eSWpL/f0C4dYJb0DPbqPI/yqrSjVb881PXgPTDKy9dOQzKcStTl4TEBAAyD7dkWSjYwtUJFlSR37WA6qqwaLWkeRjYybAkEGvk3Ve/2AcnK1JgeiIWohuXUseI1uB+V7gAbUe3aBDp1+t3uf8+VI4kz8f0tIWQGGxCUrLpMGCtJSaPEgIPqFbElzWO0lS/U2P4HEe3ahIJoqVtNXsPLMdu7+Bnt2TwNd8v953YAn8uvJZqDW78+Kjj08gPHTvGstrnLGfHBQTIYs5Z5c8qdeROw6B5ztr8y9KSp0VusEQ/I5/FsNtE9+wKPSlK16FnXuWwCP3JUNYCH9+LG7MFKf6BOMFTULmquQ1o0v1BAQaOw6krINFy96Efw2eDIP6j+P2cMhJ70hKypDQzyGho4I+Jz+vMZM7EvXZc0EwZvROsg5G7j6ZPxnOnzskETpYiD0MW1M1GwhhrT6BZs3q3vPChVLIzbwdyspToKZGHiQQQlcWk4eZ5AF6dB9PyurCw5oGuWO9d/vEKYZtb8kHL0DSyCHU//226hnYfyiZ5DrcduN8C2nvlYh9+YpnSQQWSb1XjyQYN/Y1O5UeJ+2ni6KtaDjzkksjVm69CvJT90JEJ3w2zBmbR+u//9x0Lfj6eNuTcts+cMWAWyEosBXJgv984d2w58Dv0g2hBk6c2g59elwLzb182HEVSSk7KxMey8i2fj8bbhk7zELstM8gIHCpY/+hnyA9/W84fHQT/Lnpf1AtMWlsdFdo3tz+u5mZvVYi73wSfselmdnhrZmn3EENk+Dwa+TVrBpiYx8i6+B2Bve/FQICYuFCLWbBt4LgkIHQMuI5CA59GDw9rd/HVPwWXDi7xlzC5kFq0pt7y+VwXio3OXy/rOzDsGnbUmkwUQ1dOg645M9V0sMvGyZ0MB/ojUfuov7v9/9n70vgoyrP9d8h+76QBZNAFgIkAUkASXBDkCBuLRBspZWW0Nt7Ra0B760FxVvRKxXs7V9Epdre1qBoqZUQXEECBlEhYQtrwpIFSAIh+0z29f+933fOLMksZ/Yl38PvYyYzc5Y5c855vuddv34ezhJCR0Xa3t4A5VXfQcqkB8Cd3K8xOC4oKAouXj4AqVoIHbHxbx/D3u+P2+qwzCEcl0+4zmadYtzsfiZEJGKE4EoyvC296u6eXkqG+kzUGAX/v1sXwc3GSnqSoEm+jZwoSOjj4zJ0K3VCtDhpuFBZbfFDsuN/n4OkhLH8js7BSf3sn4jSbgBPT9RePVBeeRQO/vAx+PuHEXLXjEi/cvUQNDRVUJM7flpJ7m5ibXahOAx5LTwiR2NZJIPEhIdh3NglEBIyn3xGu6Xu7LlfUdJGHzxbL1mnjK0Tt8UqyqkaugwO9sCl8mNwrOQAxI5NhaDAMJf8ndA//e7HX1psfflbfg9jsOCAFlwq3w8368uUf7d3NEBFpSaxx46dCRm3/XLYsqjSlz77KuUGG2IW4bkdhNi7bbGxUXY/G5i/ocSaJxv+kLpQfDIPOrsUGq8lxKXDXbMMm5E2r1lJ/dwWn9pxPzkHB1MdslIIDx2AsNBBmoOOw89bDru/eB7+vv1pamUT4eWdAdU3ZGQAfaytk0GzHE3mRDF4AQ2CwxEakmHSvtTdLKLrrSHrbWhmr/l4sXVi5Ly/H1l3sGo/1UdnRyn8+e/ZcORovsv9Rmh2t2ROOqaw6Qv0ffC+P8CUyZrhWEjyV6uPKv+OHafdMmKDiHdtwIC5zSNHqTO1nkZnM1aAIbUeF5MGZZcOgaKtgar0BXN/A0sXb9RreheBPu4usn5LN3uxVqEFDg5ngkJRBO2KnbSEK6aZicobn+NrTc2VUHX1NExPZTf4wMAEOHRkB1FuPUIlOFYRrocMJHZcxouMwOAc8PA0PhVtoB/guyPboE/01fczk763mmpHiwDbX80WrPg+qnZ5mwJum7bIpX4nNLtbymKJIil/y4sG3Y0Tx8+DVnmNUrE/dP8GSEl6UO8ydlLpKmKPSCwkav3KSCH1tWDFpjD6fOvoU4sdS4j98iFK5nfM/JnRqhqj4S05+8OgOG5+5xjp6Oosgp6ufdQPjkMkd/RhM4IfhLb2GqKQoyEyIpmaXv39x8P3xV9SIleWee0VyL1XBl4+6C//b43tYDR7S8u3UFPzJa0F39JSS7YTSO4XmkF5mKt+8PAuIgDkNMJejLRHvx1aAjwEP75sFPO1e6r52rEO/N23Pwf3Zz7nUr+Rpc3uO/74HKQlS0vHnZg4D7q65XD/vBdhfPzdBj+/9vW/w8FjZ+x5uOIIqVu9bKzMIc6MyQvQmGXVQjVY4tUaHdwQWGzBkl3c9KXjcXCMFLQ2b4Gmhi0skp0OVQQ7KwMro9Hn7u7R8POf7lcud7G8CD7Kew66u6sh0A/zx1kJ2NTJWTBl8jryeRVZoyXgxvVXoKODpaY1y8loGQXNrTIICUmHO2c9Rd1xIqprS2Hre78gJC6H0cGDZAAteCM+R187NoXBSURffyCZRMyHwKAsCArKcLnfB8XSHHLfs5SgweC4wvdeM3xeEIUeZGSzHEtH5pul2K3cftX+Sp2ltS219mZOXaiAbELs1iBKDJrDkwa3YQmgeais8hosfWAOv7NzjFj09x2Bgb4iqn5HqQW5oUoXlTuq9cEBBdxySyYhVJZgPjo0BubcuRwmJWZC0sSHyMiCaVOfhrjYJRrR7Ejo7XIyGR/sphGyLFKe9UWXkX8trbVQfHw3VF4tI6rwLmrVw2yZ22cuJRMMLzrBqGuooWpdJgsEX79p4OObAUHBSyAw+HGIGPM/hMzng7e3a5aPvf/xF+BK7U2LCZk972wweH/GyPdvDm6EuLi7wN8vXPL6sWytpe7PZqKMqPUi11bqkxdUgo36sUudCZoCa+Q+vr7mcVpRiYPD1XCx/ChNTdOXd97U+AZ0tW1hbVGF3HNq9hbz0NWUe3z88zBuXLbk7WMjlrNn7iHqWk5N55hz3tXN1qdol9HysU0tRLW3MtXe3RMIK3+VOyzifqRi9cZ34I3tlgv601dTBIFm9t2fP02D4TCzAfPTlz6SO8xFog2WtqSaCavnrds3+n3yglW2InQEBrQVWjioTQTOMHNfsax5HyNK0cTFweFK6OqSw45PnobXtjwCFy8f1U28fTE00ryphZE6qnMxgp2VdUWz+iCEBA2Cu7vcqH24cm0fVF9XwI16GZkUMP831nrHoDc01+M60aQeGjxAI9rR3P761iUuGb1uLPAeaklCx7ochoKDrxEyv1ZzVKlEMUBuxyfZlOyl3EdHEuxnfp+8IBXs0JqOBs09+pBV1o3BbZbMXUcz/JFTZbD0gXt48RkXxrWaMjjw7f9B8sS7jFoOq6H19nVrLcTiyDhXuoPcpPeT81sBh4t3g6JNAXGxU4d9D1TjZ879Q4hkZxHsGEnuJeSDq+eeBwdngbeP9Ij28xf2Q+XVYuEvGZ0wIKHLZMz0LhauoQ2fxMZP5M0Tpw9AYEA0xESPTMWOFslZj622aAT5nndf0ZmTLmJ0aAI57lFwueKA8rW2jgaovX4Kbp2s25qJQcyWnIBYANlQX27VsrH2I/WIxD3k/zG23mxdYzMEB/jBrNRko5f9fO8GGBuTpjfd7f47b4N3Pv7CYic97u+Nhmad5RI5nBtIzH/dlgNVV/dDb28NOb9mgrubNJL++4fPwrffvw0Txs80yr9ob5w4+Z8w0K9gPmw3VM2nCHl/D+OGFGfx9g6npZxbFQro6RUj2WVUtYvR5TRdjBBwSNgLMMotUPI+4PoOH9vFyonSIaO+el+B2JHJR7mpiF0mDA1iH4Gm+PtXvmDRgltodpcaOxQRnkzODxWxe3sFwI8f/F+d5z5LYdvoSM23NsO5vf90TaXOguNW2uvIHjldpjPFTRf+tXsNFJ/YQcY/YOL42RDgr/1EwnUmxY+Ff+6xXMc9DPCIi+Kd11wR+V++DpfLD0Bw4CBtLDJKVg9RUZkGlzt8NB++/eH/wNtTDtevf0mU7l3KQDFDqK4tI6Rkn8pm7e2lUHfjL5Qs3d0GlcFpWO7z6MmvaECa+r55eUVD0Ykv1XLOh5C7O+adZ4GP3yNG7UdoSAycOl8M9Y01ygYgSOSYd+7jIxI5yzt3c1eRukx4/SQh9siIJDISRsy5imZsS5bGRrM7Vs80SgsKxI6+dfSpR4brFmdYDnb3gcMOYeAgY62tasDbh9SxFq6deq0jUEV3kbvE/XfdJunzHxNCP3FqFzXDYW34M+e/gKkpD+ns5mZpMzwC/Vi4v4bMVBy2Vdnytga9wV76gGb3HTvXkZvUICF1oMQuAyT2alqqVN923/n7SqIqu+gyAX7dIG/9AmLHLR1Wq3woGptqYOPrP4L6hmKIj03X2SfcWmhqeBf6eksIGQ+Cu1ib3U1VnOV4yVcwdfIDhFjZfoWNHk+us2g4cmL/kIIy7NHDIwWix76h9XsrFKUgV1SAXF5LVHgAzWNXx60pmVB1rQJqb1RSYqdDIHY/H2beB9EUr0bso0cnw+3pv4bkSfMkFalyBeD9x9KdKaWY3XURe9rUR8m5rzurALORFq962ZaHqApYZdQqtZErqPMV1o54V4fto9+ZL73EEU7Uk5+8bVD9HivJIyp9LT1SzN8GMCN1MTzy4036p2ZWiIbHmS1G7/P8dccABnn9ZVs2RIYNECKaCePjZlKV7esrzbWz+Z1suHmzmOU50zFAg7MCAwYhJmYdBIeu0Lrc+zvWwbmyfPLZAY1lseZ1QuJ2ja5i2pYtvbCLll2NGuMPc2e/CmMiM212zGqrHwOFvIiavzFqva1dRqPNsdNZWzuLPA8PT4fsn2vW6LhA7onf/rANLlUU0JzzxPgkSJmUSUY2LRSjjtbm98hx3QYtrTVkyGj0Og5P7xRIm/ILmJaq6YPF/g/llQXQ0lJEA+RCgoiSD2J91WnE/UAgjHJLIdtJBm+fZDLhiBlR5zkSZNojT1r0XmYo2t1cYP78QSsFRasBSTufkPZuR/q93O2wzTRH+fKYu2goxe3QEc2byy2RSTA9NQv++NZcWPaTreRv7TdwMRrekrNFmmtPZstYRpHDMSDOiuvqigH6j4C8ZQsEB0cRNZcFAUEriNrTTrCY0lVdfZSQMdAIbjaAEgn6duUtr0BA4Hzan3voROLk6XwakY3L+AvL+vkCDPSdpwVbQsNe0KnSj5XkQ1gI+o6x0pkcykqfBE/3VyF09BLbHK+BI+DtJajfUYNUFXvSErCEdD1l1E/e2lI8rMDIpPEZdBiCvPlZaJfngRuZgfv7aprNm1vPw9cHnofvjnwAWT/6A0QLPnFsvczaLz/NT2gtWJTzksXFiTUJPX//D9YmdDSnLyJkftARfy/bm98jErHw8RxH+PLYJtBQ0By2YL1Qfoh2bosakwS//OlW+PBfTxElXkPN8GFhCRA+WrvaRzM8Kvai02UW22fRpM+bvtgfjU21cIIQLBKrnx9r6IHkimQ52E/UaOdH0NXtRZT7tOE3ni82QldXBTWfiyMoYJCSLY3E9iIKsbcUfPw0yfargq3UrBysZrLH4evDloOBk0RNZgybDCC++Hor1NefZtvCZQPYNvt69oG3dzR4eaVY9Xj19lZDV0cui1yXsXKqbmKtdHeVGR6D4IKDk2F0qHHBrI2N+6C+7o90YoSmc+oTl7H1qZM79nn49od/0cpyceNS+YmsB9nr/mTRNqVYZKbwvT9azdqI91usRW/F4LgWyl/n9hY56m9mjzz1Qkc6ABj8oa+LG/rNVz++G+7KWE4J/YOPn4Sm1hr6HnZ3++iTp6C5RbfvHGeksVGRFt3nl8g+42yUw3GUukzLwEpn165ugKPHn4DuHs182ssV++kkgA0yIfAdpB2/xCppSHb1DUXQ3n5euQz60k+fy6efVS3LcraVjU4IOTbUb9G6r8dO7AYfH/Z5XyHX29uTkV5r4+8I4e606rHq7KiBm41I7kLOOeaFe7EJjNjhDN0IoWT09Rofj3Lxch7NO29uZd/JW1i/nw/LOw8NVh8DRLW/Cts/XkePK8dwYF13SwbG0XWuWUkrcEoB5qBLyUMfej+3VE93Hci1dplX5yN1ZrKocpQDgDO6bAkBID9asA7eJ4R+vY6pbjFiNuvhjRASrNvHhjNS7A1s8Rk02WdemMZ81N0shQuXdkFp2Ztw5vQyqCx/DGqvPQaNdT+Hxps/h+u1W+BGnf4bm0w2hOEFiOlStbX74MDBNcrXK6qKqaldNLuLCp8SujsjdiyIgn7gyit5yuWu3yiDAF9hEiCSuq9gvnZnxI6+amxKIlec19iXU2f3w8CgnJI5JXQfoOoeJwJiD/CWxv+Bjo7zVjvW5NtC7U0ZHa0KpqaReEVyx/1C6wHGFaDaNvq3rFfAdULqOJpaVeun69ZK7IM0vmDLuys4sQ9VXkdPwzOb3rXoOhfeeztkL5ov6bPVNcXw4T8Xw5HityWvH++HNshJz3X0385eFeUWCWYMhwCmPUipNBcVmQyDan8/OP95mD41y+ByGIyHJV8tPRlZlPOyXisDh36gSnvrr0tg7/7n4dTpt+BmfTG0thZDW1sRUZVFMNBbBM1Nb5D3noDt/5wBZ87n6SR0mUyLfBeYHc+Zi+UFcOI0Wx5NzOpKG4na03NQUNus/zf2AceBDUREiMsFqA0v5XKDdB9aFECV6tVrBRr7WUsmBJTQfYCqdWrm9wKB0JllANuCXry4wWrHOygoBa4TQr9+Eyixq6t2b29WzY1WdPMbpGRsLDq6A+EGWfcNug0ZNDYPIXZv5qoYSuxtbaXw5l84sauTI/rRLQm0VkqtuHnwu1fhk93LycS0Fk6eeh8uVxRIEzrr/mT9g+PgKt1+pM4OTBywxvEOQe54QhgiyJ8s3AQzpi6md+np5PHO9GzJ68ca7jhTtSTQzIRRnpzYTcOUlHtZvW+iijvJY2c30JQpWme8X0aULUtt8vdhfvI9+9fA7j3DSU/DBK9NrQv8dOY8uznFjk1Xmd2RwLyZ0qaq2Q3JFVi3sFYWHT6U1MVlfYeY3TF6HJfBuuWXKoo19uPqtaOC6V1VYtVTaAuKA3uDY+T5lWtF0NhkPbXu5Z1BCf26OIiqbm1juerqqj04xPiI/MlJWYJSB6Vib2gCoTWqSrHrIva3OLHTewneCy3tk0ZrpTY/en1DKeR9+ktC4DXK18Tn4pm/t+B5g2Z4NLvbpGHL5AVBjv4b2q+iXH15Nxl7ydgEEYlo0sgXTBvqA1PfbgDLabdqXjuexFJy1ycnzYeQ4GjIvCfH6G1gtbkdXx206AXDK86ZjjERCbD/4AcwStZDFa+X0P0Ln4sBW0iaSPSdAvGXXjwFY6PTaSew0aHRUHBwq5Jk/dCM7gNKxT2KBmWhqRmJFu9SEXBbGrPsNDYUgLdnPeB9DsmGqlVPtq36JkZG+DgAMTDtVrYMVtAqr/gLC44LYAQt9hrHiUg9UaYNZJn6RrZc6hSVFens+V0w0F9DA+NwWdV22fdWdJBlm2V0+UGIIN/ROq1CfX2j4bviXSxVrF+mMYESjz92OgsMHl6bCvueN9T/A65c/QecL3ufTlTa2hW0uhvmoOPv6ecXDUdL9iunWYPCRAsnQDjxwT/wb7HMrBg8h5H2KUkPku+d6nRldy1J6CgSLE2OaKXUdn8qPv4WfL3/t1SRNzSWQvIklmoYN+4uuHj5S+juVihrgzQ1V0LSxAe1rh9T7tAdacnStXpwhHDWBU7qhgm+lYwrWkYRGbvJeIMQfxWwqHlva+0GRqljVLmhQI6oMckmrR+rzeH63/34S4vuN16EeGJzYjdhMtfaANdqTmkQOpqlxWhsVNBIOJ1U0cvoY0OTAmaksv4BdTfLoL29ggWv+apS0iipU5M2U87oQ3bziIGZAqm7u3dBT9e3LGodg8U8mfm8QSRmYYyNyYTkCbOFZbygpbmAqNl68PMVCN2LqW30IYuEXk8mBBPHL4HxcSpi7uisgebmYhoxj8VulMsLRRUbW9jySOyhIRnkxmodUg8NjoEOMkE6W1pMybyvX5PcMSI9Muq9YamAHW074XrNr0DeWgAdHWXQ0V5DiKCM3Pz3w5FjH5ObfiPERE2F+Ng0mJ66mFpgOru6QdHeQIkdJ1j422DrVjRP4sQhICAZoqOXQurU/4I7Mv6L9k0fqYSOWPnymxaNdEegdXLz2uETNFTjX33NUgjxvEdi9/IMgDGRafQ8Dw9LJhO3fKVi9/ULh8nJ2mu8Y7S7JQt9GUAX5SRO6hYh/lOE2N8B1nvdaqodfevYd91aDVSwghKWfLV0+UJeStY0REbEwzeHtqsIXXikdcXdZErF3tUjtvlEn20j3Hv3f9DlA/zD4MLFfEqSIql7CT2/6c1KTakjiWCKJMLPLwG6Og+B26h6Suj4eTRDKwm9EaC9MwgeXbhRo2KdN5kBdHcVKCcCuJ8dnaAkc3Ey8MiPNZfDCUZN7S6q1IP8MfKcETou396pInScVISHZ2hMCCyNiQkZ9KaN1dyamhuUxB4TnQWJE94jN3XNcrdtin3QLl9Fbu7dWtPU+vu74Wr1KZqmFhoSTyYkabSU89TJP4PUKTkwPiEHIsesgoCgVeAfuAqCQlfRnuchIQ9DUFAarTE/0oEmd0tHumM+OtbU0HYvFSsZ1tSquvTdqDsFExMfpO8FBkbThj9NzRVw1+3/CZlz12vdBkboW1okGUAc5SG0NHNStwixd5MDiib5bKspN4lmeH0ou7iPNhlw11FCEokXlbWlzVw4UeDEbhyQ+Ooba6GuvowRuocqT1zMm0bC7RMqoKFS7+3zgrtnscBHdMX09SlALi+huep+viqlLpI6KnUMYFtw78vKTAksa+rnnwoDfSWEoOpp1DqSKpI5knJbRyDcd+86Qk6a5OrvnwyDA0SVDJTS/cSgugY1Mkdynz51OcxIfVhjOTRRX7+eB77erdRNIE5ecB+xtSkjdDbSp2dTU7Y1geu/M/1nMPeup2HWzBxIu5UQb+R8reVeS079CgYH5fTYugmlW3GSMjT/vI+QO9aPb5ErIGXSXfzklgjsZPbSny3bnhTz0ZHQ46J19+yKjkqHiqr9ZFLaQP/u7++B7m45jI9n8RS3RKbC5KTFEBd7t9bl8R669NlXbWV2V86rBbV+0FF/TzenOwPRLM8K2Fitw5tUM7w2XK8rhdwdv4Dyqm9hStJDOokd17/nu2PUJ25pYk8jM2QsfMMhDdG3JMHB77ez1DA1xY4mePdRMiVB02A6Quxho9NoUJaI2LF3U/N2b3cpM78Ly+IyVKWTMS4mC25L05yLoiL183+YKE8vaGyugRv1ZHLQFkjI7WFY/NDrMCFBu1rGSnNkugCD/SU0Qr5ezeyePj0HHr7vt1qXQ2LvaPuCBsqJpne0PDQ2q1T6oCwGfrrQ+jWzG5tK4UTJ/8He/S/Bl19vJI9baQvWpCHtZ+WKajh64i0abIhEjscXJ1rYHlU2ivnGNZutAFReOQ3nyo7CrZPvHdHmdKmEbuma7ogdf3wO5qQbLuwTGZEK50pZ47L0256C2Xc+r3Z9eOntTYD+/yu1N+1x2ObQODArt1AdOaSOiEjE+o6zrLkJU8zwSOjvfbSMzjix69RlPcSO60VrQO7ufRafae75/hhv/mKkWke/+ZVrR5UpYiKxi13EPDxYAxE0wU9JyRlW7SwmOhMCA5OIoq8gxFLPgrIIOroCISrqcci4TXvpVhkhdE+vWRAWvgLGj88hRPQ4TBg/X2ezIOX54zMbvHyW0IA4T69pZPtLYPYd62ggpy6gyd/HJxq6OwuUwYCY/iYq9PbOQPjl0r8R8reeObqhsQiKjz0HR4//P6i9fho6OhTMp94HcLnyNJkQBUJ8bKrGNYX9FwaFWGgkdla1jx48aop3c9dsj4oDq/2VnP0epqc+wIndxoSOdd1XPvqQpM/6+YZDAJlsIqFPTHxI8jYw2t2SnTBNQBoh9W2O+LvKnPJsnLwAi5+vt/ZmMMhDap11rCq39e8LqR9ITG3y9g6A7J9tp402dAFzQqc98pTF952VY3yNm+KNwFf7XoULF9+H8NEDZAxCWAgLKMP8ZjTHYzBaR88smJb2ocF1dXcdIZO5GK3lWu2Nrs4i6Gp/DxTyfVB9Q0ZHcGgW3DbtaQgJss7+YstVzIG/fqOYWjs6OmXUStCBLg3yHIPn8HXMLFi18j2NZV/64wwI9JNDZBhRdmGDdGD9enqcMYCxhy2LVpHmVs0R4J8Mv/mP90zupOeqwPsOKl1Lp64Zc880Z9+tcc80SbE7YP13Z1Xqq8n/SdbeDEZUSvVRf7TzSWhoqlR2csPHXy/7F4yJ1B8pb63AOVT/mD7HFbt0TBh/N0QTUsGJWWdnpVK1u7nJaCewkNGPQ1zca5LWhYSuq5kLdinraN8HMNBAfci6PmctuHvEgLfvwxAUsgrGjs2Byck5MD7OsHXAVNTd2AKXLz0Dbe21LMq9D6PdZTStjT4KSh2fj4lMghlEXavDyyucpqmxCn0ywcKhmaY2SkuaGo1pkGN2QyVR7A/yE9zKhK4vMM5SYGl3a6xZ2904OGAkvLuTnpc26/S2etM7kvzrD2auo8TeinXhyc1k0YMbqVKXAiydaI0Sh3ji48XLFbt0oFLE0dsrp+oSiQLN75gCpY6+Pjl0dpRSE7avn7RI8dNn34KG+lwICWylVdMwnQ3R1jUf4hNeG9ZCdChoCddBBetu5pVi88mAsejtqYZrV58gx7GM+r9FE7lIvPhIh2yQPLL88bl3De/edfttWYTAA+Gr/Rvgxs1qGByQ0QDBgYFBiBjNCstQqldOqDEAcBSMjZkHkZGZMEVHKhQndMtBDIyT2qgFU9owxsP4+/G71q7tbpxSd0A4n/l98oJ7wMZNYXAGWrJzq8HPdXbJqU/99pnZcOVqEZRdLoDspdsNqnURWJrR0opdvOA4sVsGtdeL4cLFN2Ggv0hpnvf3DQBf/yzwC1yls5d5wTfPQUNDHowJR9PxAPj6skYjmI5WWS2Dnv4MuD9Tu1m/uqYAKio2QIBvNasTj5MBcuX2Dq6AqKgcvZMBLNjS3V1KiC8ZRrnZbhLQ3rYP6q7/Dnp6FFSZi/3TqZm9U/2RmeEDA5PhrllPQUrSPL3rrbxSBHX1RTR90E1WDV6e1RASFEh7naPil42KBk/PZDIJS+Enq40IHXHyk7cl3V8UhMwPfPscffzJ4l16A+GGAptYWbKVtUVwbq/Dcagzkvo39pghrVq2SGsRBW04fDQX9hz4Az24Pj4B8OiirZKKeVirohMndsvgi73PQ01tnuDXHWCtUn1YyVW5AqBJkQLpMz8bthx2Aysv3wZjIgYJqQ8IrVJZLXKMWEdSx/Hwgu3DzhOcDNTXC5OB0AFasQ5JHUva4jLgNh/m3v3nYdusqS2Amuo3wcvjPC2Mg8sMyqLBy/e/IXT0fKO+901CorT/uRA0OMqN9UKnr3kMtxi0t+2Ehro1rGocGb2EbHt7hap8XQKRdwKNNxg3bjFMSlxM85I5nJPQ33vlvyQ1arlwcRd8f+RV6t5CX8rECYtg3pxXJW0D09fSHnnScczuDkzqzmV+ZyrdLiYPNI2jGd5Q1baTp/Pgq/1/UNYA7+pSwN5vNsDjyz81uA00XVHifeQpi5uYuCnePOz89Hm4cnUXJXQkSG+1ZihoCsbCMTW1pXD6XC5MnZytXK68shiOnviAkDLrpoZR9Iwg2XKdglLFcaG8WIPUv/3+bbh4OZ8uK0ar03SuUaBcpqq6AJInlmpYgw4cfA4a0SoQMUD3FZdBIm6R18CZsicgPmETTBi/ROd3xcYzlVUF0NBUBB5ucjaRQMuCUGCH/c1K4iKweIybxyyqkj3cA6BdkctSzoQetPRBMLXjI1YOG5+wHFKSn+YnlpMTOpaAldp5zdMrgFpulCR/KR/i4+ZBQpzhOv9oxXQ4QnegbqPqGOVk5+dme24c6wvjjFEXSi/ug7wv1rI/hCYekRFJsPxR6YUdxFatqKwtDZHYectW4/DtDx/AudJ8SpC+QstST6GJiocbq+iGRWZwnD6v2VGqoHCrWqU6VflZJPaePpV6xRGkplZrr5eR7b4t5M6rTQaIQsacbZHUcahv87M9z8Pl8l2sWpzaMgjx8/sP/kHnd931xQb45NO11HXU2algvutBEHzYoPn3oIq0e7qKoKdzJxm5SjKnpkAxaA2YTz0iPBPmzjnACd0FCH35wvm0UZU+KNQatcTHZkJcrOBekYkT0OdpwRl9WL3xHds0azEeuZzUzVPpC8GGAXK6SFFfS0JUWckTMpWEHhwUTX3q3kZGFaOSRkXNid3+aGyqIST4NiVyX8H0jcVbRJWOPzV2VcOBar29Q7XspfKjcPVasVpBG6zTrlLctOysMGSyQJiaolIsn+7ZqJoMeLAa9OJyvX2qQjjqk4FzZfvJ5GMX3Z6nMnKfDVTT4jLNrQq4XFk07Ltih7IDh94fTt56iF0MUBscHFQWghlK5ixwDWBc3DqYMmWrwYBADucg9NwN+lupHj/5Fuz4JBOqrqgmnXNn/wEC/KOUf6PVpltNvQ8F+tFt0CPdNEI/t/clR9wxZ1Lq2Y6wEzhjxJmjNmBK0M8f2Qppty6mHbWWLt5qNKGrE/vmNSut8h04sUvHh/9aRwhJLqj0QU2V7s6UL1XpCqbUJyWqTJGF33+gJHRPoc66m9tw5Yzj9pnLlbnUF4XJgJdAzrQIjtDzHEd3t9g1jo3E+HS63Od7N2rUsHdXW6anR3MZJGRtE5j+QU0C79dD7GJLWUxTwwnK4ICg3NXSOmUy9rlWIsbOnH0LSsveJPvC+5ZbnXHy99FcbmsQOgYOb17zuM73u8nvm/fpYjhR8ja7Dr57XqnYMTDu3nuYH33qlF/CT7N26YyCx/tT9gt/svWhwzLkhaA9GLtKUOeYn77CUX97Z/KpL3KUHcGZI5KuLl9S1sObiBqqNruQh7h+a1R9wosdL3qpQS4jERcvE3KtPkr96L6CT9nHBwRCHxRqu6uGh0c0pE9frFS950r30wA3L6FCnYc7KAPOkAjFoLFJE7Ig8x6VObr4eL5SbYvEzhQ3i3rvUJsMzLptOW0DW3QsH9rbaiAgQrN97FAfPA60CgytKY8YG51MJ6MDA3IVcWtT7MLf/cKoa2DlW/G5uxqZI5DQ2ztl5D0ZjcS/cPEt2jYVfeppU3/DTzIrEfoKK5EhEjpaEQ2lruHETZzQoR+98NBz8KMH36fvRd2SDo89WqA3pc1afd11qm5s/X1u725X+P2do/jM5AWx5P/VjrRLWEZWX2EXSxXywMmDNYrTiOBNYPRM3t5ZQYhaTiZnQMYgBJMRIDZD8WK9z282yujo6Y+Bny7aCqEhbCJ3rOQruFR+gLY5ZQPYsl5MsWOKV4siCBISnoVZM59VbhMnA598+jL4eXezvum4bICq5zotgSr0PY+OyoIH57MUn48+WUf+b2Bd2DSWGaQBati0RWz6MvuO30JsjHZPFjZauXDpENlQN6t97y6DyPBkCA1OgIjwWRAROZ+sP4NMTmTg5R0N12pqaCqZp4eMNbJxZzdyRvwsda1XKDYjtlft6emm3bnOl+2CkJBkk/KVORyX0N3dvGgZ5UuXVWbztrZa8PAMgMiINKVi1wdrtIHVNncA7Pp5bu8mR++R7opKPdjRdoj511+Gkk/ellxwwRzFXlVbBy9t3W6V9eNNAE1dUlP2RopKb2uvYSpdNL17q1Q6EqVCUOgJ8ag6cwiJqm5Up88d0AiQw+5vonJmqWAZMGPGa+DtrWnNOXX2APT3KZhKFxrMuKv502m+dzdQpTtjGqsnj2Zz7OtOrQKCD95Dhw8+eWIW3D0rW/eNe3ImUfEH4EZdKXUHROopcXz12k5oai0GDP0QVTy1yKNKHxRM8oOs2ptMiHxXf97WVgN5u5fDlJRfwr33PMdPOnPvE1ZonyoC43vQhy71XnfLmHSYnvYUHD/5tng6SM6fxnaq1voeQ7DeVdS5OpzDp35u7ylH3C1MO1tkRjGEupul0s++J5fR4BSrqdLt+fSmwMHw5b63ldHu4qOX16AQIIem5kAICFoBc2YXEqX9ggahIyqqitX824NKf7qnZzQEjv4zxIz7aBihIzC4TmV2B1VwnZAS1j8QA6lTP1QSOuLIsd3KCYQ2H7zYB35KytPw0H2bDH53tDLFx2boJXQ0r5ac2qBhlkfzO969aQlX/2zw9slQpbOJvnbhuXoDluMl78NfcrNo8SYOE+SmYKq2JqGbkgo7Pe03RLEngb9/FGQtzINbJy83uAxaQJ/Z9K4tDlsV4ZU3XPF8cCafei44SLCcOg6SkxAD54xVuZ/tWQNnzu+ChxdshKmTs6QdACHa1FoXL64XFbsUE5srA5UvdmyLGK2m0jHq3ScGgoLTISBgPnj56J5gXaspg95ehbKVK47RozNhzJhsgyVlK6qOagS7iWlpmP8dGPICREQPzy8/c26/Xh/8KLcUyMjYBAH+lquydvL0FjJZUFCLgEjoIrn7B+aAb8AqCAnDqnHnoe7GNqitzSNELtMgc/U67Tfry2DrX5fAzx55E6JuSQIO6YRurYJV5hC6iPn3vkXOxwBJlePw3qMvu8gKfOKScJ6GLhGJheT/+8GKfdRNBfZfN8YvfepsHhw6/Ba94V6qKKDRoglxsyUti8VvMFfeWhcx9nfHPu+zUpNGbCMYbNsZShhpYuJMmJA4D5In/QeMHfc0REQSsvKbD+4e4w1YYCqp6X7q5HlErfw73DrlJQgLWwIenjEGJxOY1y764PER/dm3RP8Wwse8QbY7nJTRB7/7y03Uhz7UB4+ZF4Ehz0LEmNcI2VuunSr2OD9w8BmaO+/hwfrNewrD1zcZgkZvUX7WwyMcgkPmQ3hEFrS1lUI7OS4saE5G0+zYc+Zrb+tQENX+FUwcfxcEBoRxxpZAgtjc5EJVtUMSOgLJXFvraW2TE3Rn2rA/eiHUlx90xfPCucrETl4QBCxgDhV7nFaTiqrKD0Zk2NQXL6X+MRL6Z3vWqqptyViBmn977FOjtmVNc5t4QWODBqyix2EbHD6aD0XHPoCE2FsIsc2E2HHzqHVA7439zH746F85hPwH4BYyIiMGISoyCqJjngZ///lWqfe+Z/8aqLqSR33pdPgPCs8HITb+Q/D2maV7wlNXAMdOrAVFmwK6MM++m/Wox+csb19Go/NfXLOXt0vVA8zfxnQva0WH27qsNFob0OppQ6x31DzzkUXqwwk+TUnm5/Ze0fKZWGECsNoWBG/oQrhxsxT++v5CVUcpGUYbJ8FjP9k+zCfrCMSOwDKQhqpGcdgPX3y9FQ4XvwXTp94GkxJnwrixmRAQkGy17XV1yeHd3Lng4S4fRuoYjR891nBr2p5eOXz3/VNwreaoBpl3daH/X0YDAZdmbYDbZy7iP7AWYCCZNf3OtiZ0W9zHtCDXkXPNzYGb0+55fXk3GVeE0arjM63UxBKRuANYzXirmu6xh/mRU2Ww9IF7tPYU9vcLh+DAaLhwuYASOqs4969hBWpu1peCn59hc6m1TfEITCvBbaBit2afZA7TgIr+3tlPQfKkxRARnkF7j1sTR0/+Ay5V7Fea29GXz8zwAImJfwY3d8OTUzc3L4iPy6Km+PrGUiHdTQZ9A6BMfUuIT4f42FT+A6sBTdSY6rXpbx+7DKFjCt5Lf95uj8MZTLjBJQPl3EbE1YDkHpG4R1Dt3tbclOiT1kXsYyKSKbFfrS4iauRvhNg1zatnzufBJ5/+Glrl1TBhvOFod7HBjDVNVzhpwO+EefkjOYCOA2vDvwjd3Q3Mly4QOpI7qvSIyCVGrQutCv7+0VBZtZ/51/tVvvafLPw9N7+rASfW9698waq526YSOmZCnDn7Vzj0w3OQOH4xzVOXSugrXrBbxk0wjdNCUchJ3amJHZX6LGtvCon9RkOzzo5uSOx3pD9Olbs6kOh3ff4UVfH1jWVQ31AKcbF3Gww0QRVtzQI14nfK3b0PkuLHQlLCWH6XHYFobK6Grwr+xIjcUxUch2PSpBfAx9v4Copho5MhICAarlUfhe6eHtqC9SeLNkDcOK7SRaD/HAndmkFkWFgGY2iMJXRsHrTvwL9DTe130NOrAB+fMIgIN9yiA4P8lj77KrVu2hFxhBe2cVJ3ZkQkdoGN0uJQ3eLs2lCrViVp1pfCv/L/jSiVHmWaT0dHPcxIy5YUPWrtynMIvAD/ueegciLBMbJw5NguuFB+aBihBwXGQEryCyavNzwsGdJv+3eYfedTZLL7S0IKCfxgC8B02Wdee9eq5CdWiouLNt47iUR+4dI/lX+3tlbC5OTlBgndWo1mjCb1iMR8Qux1rnTOjOKXjfWAwR8Y1CIFH36yjOb9isDuRY8uNi6ADivPYS13a3R3UwdWtsOLUl8bWg7XQ3unUBO+X7OxS3h4Jj84FgZeW2lLnrR6hzKppV8pgffI4fSZt+D7I6rqf2Mi08HfT1XmF1M5L13Wfc8TU9ccqDf6alc7dzipWxkYpYq+I0NY8qOt4OUVoCL0rO3UFGkskNit1bZVHejDT3vkSWoa5BgZwEpzWM9dvZkLjnExWfzgWBB4TeG1Ze0e4lihUiqhl13YBvmfzYPTZ9+CispdcKOuWPkebcyjlkd18tRbOgkdxQBW4nQguFyKxUgjdbvUkMdgEEPEHhuTAT9/ZDsl9Afmb4TjJbnwzzxU78aXzhT7scdGRVr1e+Fse/Gql6mJEC9YDtfGHTMXQ9SYJLV2rDKIj82C4OAUfnAsALyG8FrCa8raSlbshy418BXN7FglUQQqdhEYHIdqHXk9MWEx3H3nqzoJ3doTFZM4YfIClwrgkI2oq2bygvfAjqVmvyFEK8UX/eW+NXC2dBf1qwcRtb7owa0QEZ5s0k3CVhcSmvHwJsG7vbk2OjvlcP5iAU1HCwuNhpQkrtItAdo7fN2fbHKtmlJ7Ak3vqNSxjSpt2kP+mz/vfWp+R6ByR2LHbAatcjjnJavG+5gJ7I/uMtXlRlqgXC5YOaVNH/IP/KC3XSviuyNvwrGSbcpguZ5uBQ2Ui4+dbfT2MKUOU+vKKq/Bhcpqq343jI5/9+Mv2RXCg+hcFh4eXkStJ0Ps2Awa4MZhPtZv3Q4/e3YjvYasDYy5WfnoQ8YThZsX9A/0QN3NYo3XMS0RgWTu6ak9/gcnK2KArYMi15VS20aO+X3yAgzJtGsLVzSpoXLGWbk2XCzfR0ldHTHR6TD37nUmbxPNa5iqYs0Ob+rAIDoM8NH1HTk4OFTqHK8Va7VUVgfG2GAZa4y5MRVJE39JH9G8G0kU+vgEw2rfTtXijEWVK51XI0epM5Vu92YwmJqy46uDWhU7VpGruPItdHQ0UJUeEZ4Ejyz8m6SUNkPA1Dprp7xx1c7B4XjqnEW4/9Hs+hKo1v38oiF50nJIvfU3Ok3tTkboha7WgnVk+NRZDXiHmo3pqt6EgXH/+GQZJfWlS0yrCW9IHdgyR5T72jk4NK8/W/nOEQvvvR1yX9EfEIf+8trr+6Gyahf09srhvkzz0+ichNBLgPnTWzmpOx+pryL/b3a03dJH7N1kYFEPa4AWxcl5yaaRqKuWLYL1Ty7jZWY5RiQwaBXVubXzztXxIrne8JozhH37F0FLS5kqAC5zF4QEmx4vYQdCbyEjXxgtaq/PAdb0Kw5Uzb+YOmefzXU1Qh9JpP46OGiRAVs3UFC/yaze9K5NLz5Msdu85nHJVfY4OFwBmHeO15qt8rPxnoLqXOp1dunyNig59aqS1ONiF0P6zFedhdCRnLNdkZw5qesn9W+EWZtDwl7EjrB2G0dtuGfmVHrTiYuO5Fcgh8sCLWLY89yWfcJNcXdheuKXe+dRQh8cxAyHAFi88KgzEDr6w+fyM00TIyNQLiIRTS+zHHX39AXPWRuzUpNh0b23w57vjtvMz46KRTRDppGbEG/pyuFKQCvYxr99TIvI2LJ6Gma47Pjjc0bXcMdUtJaWUlAoKqnMw/4Tfr7RRpng7eRDn6Oz7fYIxkhJaStx9B0U090KzZjVdwu+eGOBs/qST96mQTW2BKbyxC1YLqmMLgeHMwDPZTynbZGmpg4sKCOlQlx7RzWcPrMB9n+zUOP16CiWbx4clATpt71K/p7n6ISOKv0KP+OGY6SY34MEYo9zht3FAhHG5pMime/6nEXNL3poO3ipRc0rFNUQECAt6M4e5ngEmg03r13JU+A4nBI4GccSr7Yug4pxKvlbfi/J3H7m7Aa4XP4+fY5m9hnTX4XYcaqKgGiG9/OT3m8CLRLoXrBTpbjNhNSf4WfecIwM83t9eTdEJBYCM8Hrs01h5OQRYNGRJcLfWLDGplXo8CLBnHKpfjEk9DxC6I2NZdDR2QDny/4B42Jmg69vOHzz7Roo/H4tBPhHQ9howzW67WGOR2C+Ls728eaYlpRgczcEB4cpYH3BN1Jlboucc3WguR0JXaq5faC/B6prvlT+jbXc1UldV0U4XYSOlsWDx87Y69DvIff1g/wMHKlKXVO1o90pTU21VwmjhMz8TulYBlPi1oONK9JJrdH8+d4nobKqgP6aYnnZebM3QmNTKZw5jyVnaVwrzL17I0xMXCL5orV1dPzQGxam4/BgOg5HBAbBYYqaPa4PDKzdvGalSdXhPv/yNkrmg+yWAAvu2w9+vsalzjpIcxZMR1vBz0RO6uZMBlIFBW9TYhe7KenCkWNvQvFxVlpWJPS0KcuJKk+Gbw6tpX+z9wYhPjYT7rv3z0ZtH9Nx0MRmr/7HnNw5OJmrYG7miGiCF0k9adJvICX5aaMsE7YsnqNvVwipT+Nn5HC48UMgEfXldRCReANs3H8XLx68kaCvWVuUeDgh7yvXvqVmdyTw5ImLYUrKz4h6/7UaoQMl+fvm/Rnc3YwrOYulJZc+MAdKyH7Yow8yfn+MlMdjgO4IXryGw15kjpYrbKNsL0LDYjLGtEvVBm+vcKis2kHT1mJiHoKoWzLBz0+aUherUV6pvekIP8kYWvqbR79zpW4Bxd4MdmgMw+o3v6b1gqY+9c+W0a5ZmXM20df2H1wDFy/vos/DRifBjx/QDJ4zBRhEhyrFXqqdK3eOkabMxWvfkqWWa68XQFhYOnh6SL8fYFT/6k3v2PXa14LVrla3nZO6fUh9l63VupSLW0xlUyfu02dz4djJN+Gxn34zjNAvXt4Jo0OTyUgxah/sUVBDGzD9DuMNeLQ8hzWAAZs4ibV3D3CppV7VceNGAfT2yWFsjGV63SOho4XCAcFN8JzULULqLwILmrMLLFF9Dgn92x/WgpdnADx433ajid1RVDsCfYzZRL2b01KSg0OdwHKJKrf3pNUUdX7+/Aa4VrOLBsJhdbgF9x0FDw/zrHNO0JglmJeI1QT3qRuLiERU6XarTofV57CtqTEpb+pobDoPewqYvx1TXC5c2gGhRLEHBxm3Lkx9W/nTh6Cs8hpcqKy228+Bfn5UU3gzbiETDF6hjsNYiBXgRAKzR+yI+qR94zO/IoT+W6PTOmuvfwktLaeUSs3fLwGCgpJNPib3r3zB7pYKCcDUNl6EhpO6WaS+FhygiA1ebPTCu+s2ycs0EEL/Ys8y6B/oYZHy5DVU69NSc8ij8TN6JE8MokMiPXK6zK6qHbeN6moTuTmjiwB97jzXnUMf0MRO+5r/biM9d+xtdUKXUv6WF426pofi+nUhD11g9uioh4xeBwbELcp5GYrINe0EKCGkXsTPZhW4+d1Y2ClQThfQ/Jz/xu8lRcQWHloDF8vzVITuZbr5XdvMfvP2fJuXx9QHNGGi3x2L6fCoeQ7xPEXzOlp2HCAti8KS3Qu/3ifkoZPn7m4BcP+CY0ZPdLAts4MFxOnDeji39yV+ZnNSN5XQsXBNvqPtljGlIgu/WwOXLudRQn9ogWUIXR2OEkinDjRpLrr3Dup354F1IxNYbyH/wGGH8w9jINzqZYuMnnTK5aVQW5tHyTs5aZ3y9WPHn4S6uv30dR/vaLhtxtuSTfD2KhFtJuYQUueV5Tipm0zqDt2XXWqVqcPFr9DId6nV5Uy9idqyh7QxE6BFQuQ8T4tzbaAZGRU5krmjnYdoasfr1ZRz8PDhhSBXMNM4FpGZd68qIO5GXQF0dtRA6Oh0CAqURub2rh5pBrCMdxwPlOOkbg6pnwRWYtZhsYrM+rExiiVRXrkTxsebNgFAnyUqAEc056F5HidBqOI5wXMit9WkEqPazbEYnT27FmpqdylJfVra2xAZmWny8XKACnFVwmDK24jbCze9c1I3l9QHnWE39RWqMRY/FK2Byqo8SIjPgtvTNylfr67dBxFhGZKaQDiDEuAEz4ncmjCnXvtQ3LxZACUlT1ETO5J6dNRiuPXWjUavx94loAHrtzNivjLkPpsqkHu2DhHVIizHC89wUjeL0GPVZpMOD7yJYCStOYrg1JktcObcm6zcLBkJcYzYK6t2QvHxtRASnAx3ztoqucykI/rbdakpNNFj4BL3wTsmkJAwqMuRiVy8DtHVY4rfXBd6e+XwzTczlaTu7h4AmfOMC4jDNrFYftmOyCakvE3CfTdoGLFzHzon9ZFI6iJMqUiFuFyxEw4XrVUSOj5GRqTDjLR1cODgMujDAhcylhI35+7tEBwsPeBOTCVydHIXb8pI7Kjg8ZGrePupcTxvcDhB7jQFusLw2jOHzJHA6xsKyAQ6HXx8VJPnkyVPEsW+nxJ6RHgmJCb+RuN9fRNrjG63s7mdm805qTsMsQ86426jaRlVu1RCamurhrxP52oQeihR5TOmrYNDPzxBCV18z5OQ+lwjSd0ZyV39WCK5i4OnylkHSD4iieNwZDU+FJboT9DcXATV1dugvr6AqvEJE56HceOyle83NRVBX58CIiKk+9IdwNwugleB46TuMKT+DRgXyOFQihNbNkrNhb1RVwSFh56AXnLjCA1Jhvlzt8O33z9BVEMxO2lEQp+9nagI89LinJHch5I8phNyJW++Ei8pq3A6ErckmYtQKErh6LGFtNwrkrq/fxJkZHxq0rocLKalkBD6XH7Gc1J3FFK/B1hPdUliA+9VwgBhMmD3CQHeeLDQhRSF2dR8Hg5+9yTMv3c7+PvFQE+PHPYXLoOW1lKq1Bdk7jab0F2F3NUnT+okj9X2uJofTuBI3vTxQoVT/96WJnN1HD48Fzo7a6jvHP+7444DkkzsQ481VodzoElSPiH1xfwq4KTuSMSOBWg2g/ZSsVXAitNsHhbRyZbFoI/1YOdcd3PSapDY0ac+acJyiI8znOaGPkGEMY0lXIHchxK9UsmTY483/5EQgIcm9CpCJiKJ43NX+k2xkZA59Q66uqrB21s3SV+6tAGuXdumJPXk5I0QFSW98xpeQ45U4VEA3huf4UTCSd1RVTtGZQYL6r1KK5FrX3aVMDGwKywRyGOI0L87vIyeZHfevt3ojlFIBps/yHfGohiSJ1dICKKaFx+difBF4lYn8BahDr8rwtxo9kuX1kB7eym0t5WBp1c0zJz5jc7PUhP80YX0OZrgx45dDhMnrpN03ThA7rnOOTs3v3NSd9VJgUP45y1RDEMXDv3wGDQ1FtOzDKtbTUvdCEFBxpvrkTCwgA3W7HaimtRmI1Ug+WBB6SPUnyNE5W/Ru64aIaM/tkQgB/XnSBwj7bdAMjc3z/zsmWXQKi9mvnLy9113XdT7+RMnl0F4WCaEkSHF9O4oLZH1Xc6E1OM5AXBSd0VSXw6s+IJDwNKq/fjJNXD1mtA8RmggkxC/XFmHemyM8RXqxGYcqN6dMZDK1rhHz0QNJ0r8GBrGQqGksKUmvZUVG6D2+jYlqd966wfkmsgwe704ycJgOCexkMRJtmpycFJ3IlJ3uLx3S6n2U2degfKKbSpCJ2NcTBaMHZsFh4uWwSjyd+qtGyEmxvTa86gmUZU4S84yh/MArwNU5OgzN9cKgqZ2Pz9VDfba2lyoqvwDNacjqU+YsBEiI7PM2oaD+s71YREh9d38TLMOeD91e6G+vBUiEtc70i6hyQ7916iIZ01Nov3STUFoaBrU3TwE3d0NjNAJmY8nKr2o+N9gYLCHfqbuZgH09ckhPHy2aVN9crPFXu5440XrAirPkWQO5rCOKt+4+lfwzos5ZtUgaKjfCRUVa6Gy8vdw8+YOCAxMBy8vZjofGOiG+pu7lJ/19o6GkBDTrgGxTeo/9zhdgbUycv/jVeG4UndJte6wxWyMzWsfih4Mkvt+GSX1W6esg+KjT0Bfv4KecPiap0cAzMrYTm54lkuJc9T2mhwjQ5WLuHEjF65e2aA0sSeM30Qmr0yN40S2rPRJqt69vKIhMCgD/P2TjVo/TrpRndu5zKs54BHwnNRdktAxva3F0XcT/bJI7qbc8JDY8QQrr8yFCxffVPrWsWgNEnpQYIpV9hlvekjuaJ530AhgDjtPWLHsL5K5NQJE5fIiQtzLlKQeHZMDMTFPW2zi6iBV4cz6GjxXnZO6K5K6MYVs7A6sIW9OU4qr13ZCySlWS37mjK0wZsx8m+w3muXzD/xAu3hxgudELjbqsSZQjZ84PoM+R995QGAGpKRsN/s8doZmSBLBa79zUndJUn8RWCEapwGaKjFC3tS0HiR2POHGjpUeINfZWW10FS19N0aMnke1wwl+ZBG5pev0d3ach5aWAujpqYHu7mqYOOlDjfdPHJtO3U1I6p5eMTBt2jcmbccFTO3aIK1DGwcndScj9feA9Qt2OqBJHsndFkVSvv/hHkrqExLXQUCA5cz1XMFzRW4OrlatgcbGPGW1twmTtpPzU5WadvXKK/Q9L88Y8PVLhsBA49PW8NxcvekdRzO1Y9nrfOExThhzQHvfc63zFGApbbyhCyd1lyN1p20OI8JaNa9FVFRsgcqqLUpffOKEdTA2ZoXFtyP64JHkMaKYR9E7F7AwDBI4Erl6YR5rov5mLlRf20CfoxqPjdsIo8OWWGTdeA5iv3MHm2wiGS/S2cucpeiup59hVTaNXwcHJ3VO6o6hjMwpm6kLCsV5KCr+sZLQ8b/k5E1wy5gltGY2Ql/dbHNvqji4md4xgW4gsfUtErk9Gua0KYrg0sVlSlK/JeppMnLMWqeD+83TCBmfknBfCxKIfY6akq8SxnpedIaTOid1JyJ3VO1I8JbAseOPQUtLEX2OxWpCQjJg2jTmtyw59RjIW4sgNjaHDqvKE6Li1ft6c5K3D4mnJSUIJH6HQ7S37e+Tw+nTM1StUQPSh/nVjSFz9Js7cCpmLiHjFfxM5KTOoZ/UXwc7d2uz1g3YnGA6BAbHHTu+DLq6q+kJ6uERAOkzP6O+9YaGfXDu3BP0zEWyxxzfadM+s9n3E0kea6Djo6s2LrEn0JwudrWzZI/6rs7zcOO6YDInIzg4C0JCTTeZnz41nZC7Avz806k/3Vil7kRBcNJUOgcn9RFO6samtLUIny9Re20RSA9QcSpyxw5vV6/lQmXlFkhIyIGE+BzoJeqo6Mg9rIiNYJa3hVo3BKy7TYle6BHO1bzxKlydxK2F9rYiuFK5DMSKT+ERT0NEJDt3mpt2Qm9PjfK9yDGGzylU627ugUbvB5L5ZkLkWEfBKeI3zu3lPMFJnUMise8SiFkfkMhzdaaAsMkBTvWDXZHc1VPaSst+B3U38uhZiyeuj080TJ/+GbibcGO1BdGLPcSR8Ed6AxWxpzy2l6XtZgUStyX0kXpl+WPk/WLle7emXrL49p2OzMX7D2+V6lRw54fArsgGFjGqzQyPRL3ZYKQovj95wRyB/B2O2JHIVrzwJ2pmNIXc1XPU3d0C6U1XnIlOnPiaZELHXGKx/rYtQAlMSyS2SPDqvcddqY0ppjuK7WHjhvSLdyQlM7Q+s4cnnhvFWt8boWTOwZU6hxmKHSNGRTN6i0n+KycpZmOucm9rOw9lZWuoLz0p6TXJhH761BwICcmEyDErNPKJHU3dI8mLpD+0l7k9zfqxan3b07T0ebe16taHq5WPQW9vDS0M4+uXDrHxLIANfeoVlxfS54NDlPrNui1kvKks7Ro/fjs5x8w7T5wgAE7SaUnuR9P4TZqTOod9JgYtzrK71kqF03qTv/IK1NXlKtvARkXlaA1q6u+Xg5tboFMcP3ECMNQKYCq0kbIjEbUxKDs3QUncvr4qUqdEW/GY8r3AwPkwOizb4qSOv83mD/Jdp7EQ96lzUuewG7E7XZockjt2yEKCt1aq0vHj02BALbhuytTCYab4pqadUHNtDYSEZhFFvwT8/DP4+WRjNDVsAUXrTnD3iAFP72QIDskGDw/jXSb6SF0XLEHqWNcAzewumBERx/PLnQfcp+5aqHK2HUYfI6b04MB+1kjullSI9fU7adqRTHCW+gdmaPWtNzflKR9bmvMgZuwmQipLoKdHSKvzjOFnl5XR2VkEfX01dODz0NG2y2pAM7xoijcGYjVCNLO7cCBkHBmc1Dmpc3BSNw67yc0Rh+h3t0S1sOZmTRPo6LCsYZ/p6a6mkdEIJH83twBK6Ii6G69Au6IAAgIzISh4CfgHzOdnmY0gukJ6e6tB3ppHJmPJMGpUIHgRFa/PTeLhEU196oj+foV1LjTBX46lhXnwGwcndQ5rzqidHmLE/GqhOcfqXywyuaZ3bOwL4OebAg0NOylhh2mpz11fn0urgskEZ5RI6KjS2+QF1GbfRoi9s6MIEied5GeZldDVUaR8jgQuorurlJrmKcjvFBqWA6PDdavqoOAs8aMWBzZZwU5/I6zoUBU/Ozmpc3BSNxuogDDYCAdWGUPTvLHqHU3t0TE5dHS0n9f6GfSna6r5bLb95jxlCh2SfmCQ4epj1689DJ6EkHz85oO39ywY5SSBd5ZAT/d5cowL6HNvnwxwd4+m/nHJUMsnQ0WuJPXuUo2P+fjq93WHRVjWbI+qHNPRkMydUJUXqt0bTLk/tHB/Oid1Dg6LA9O51NU7psQZ63v39dPeunXSpM8IsaMvfSd4esbQgWhseE+pDpFwQkKzDZJaT08p9JLR3pZHiS3ilo/oe80Nz4IXIXlv3/kahOVSpE6+d2vzFmFCRI7X6HUQGMxKhrc0bQGFfCch+hhK9sGjczSC4NDEHkJe6+46T4Ma0cQuYqBfbvPvIvrKkcydtELgemB1LlQtTlkntWxhSCX4fH734aTOwWET9Y6+dyR3jJ43J3Lek6j5Mbfk0IEmd0Rn53noH1AoTfK+fhkGg+U62vdpmPF9fJn/vbPja+hsz4OujjxobQIICMoB/6BV+gmy62tCZjXg6bNE6ySgv68a5M0rycRhCXj5Dv9MZ/tOsr2dZBJB3vexzUSiv69myHFNGfZ+n+Dr9g9cokHq+FxXYFzPEKWOv4W1gBHsqMgxtsOJob1WO1PcL9HB6lpg0atgvSrdCWpfcGhiFD8EHM4K9L2/tHU7xC9YDmlLnqSqCk2l5kBU6T4+KZCUcgIix6wjZB4NQUFZBpdtb9unFPbqpN7dVaTh33VzN2yS7u7IhQ7FK9BycxooCHn391cPIbp90N9bCu3yV6DxxjRK4OrAv3u7iwjx/w4ars+G3p7zOrfVRyYITfUroU3+Hn1uKvp6qzVyZEeNClDtT2fREMJPdpjzCIk8e92fIPj2JbB41cvOTujrJRWvOrf3JUGtI2lXafkEvjaHm96dDzxP3ZUgrZa8ywP976jgbdmmEwvXXK2cRi8oVOoYgR017lv63o3q2eT9GuV7kTEn9SpnJPDWhnuUswOZLABCIks0PtN0czYMCMoYPxYS/hm4e6QoVXxj3T3KnGtcPiK6RPdkhJC5vGUD29wgsyQEBBvvl66rfQy6O1UTmNjxl5Xv3ah5jBL7oPBm/ITLRq+/o6MI3MhEwcs7xSJEjuZ1F4tebwGWU95qwr0jVUO1GypPzeGw4OZ310KhCaSeC8xvVkJn5eziniPM4IOd8SCgD/SZTe/SIRI8+t9NjaCXisCgbKrK0acuqnRUvph3LROm0O4eyQZN4T2dOzVqkKMJXlMRn6emeeVF7J6sJHRRpavP2n389Af4dXdrqmhv30wVSzS9Qta3Dzy9M+h38iKPuvZfXZnL1J5rKHWyQ54epql0X1/Tze6ijxyr7rlwGlq+SYTOSJy3VuWkzuGAyDWCjNlnh5rX2MV9ipB7rjBJSHPmAyISPFWOUZE0eh4JftG8Oyy6HcybHh3+3wLpqkzYqJo9PJOhr6dUIEzDee5I6urwGkLqXe3vqRh7EKjfXON9kdR1vD+M1DsLNMjYw1N9grCP+sI72vLoiBp3Qud6wse8Qx8xsG2oGT84NIcGEmKcwihZgE1+e3TFIIEjkTu5SV3yV+a3QA5ufnc1TF6wXCBsfRd+tiTzGqsnX+Wsil0fsDyt2L/bFmZ6JLkeouJR8brr8akPDMihQ/476O1mQXduRIUHhX2uSfpdX9OgO/Sr07ztMZrmfAyS6yTE39dbSlV8aOTnOreHpN3c8AR9PkhVfRaEjH5Nuc911XOU7+HkJDLqM4f9TVGNI4EzNX54JLa6ncPN5hyc1F2T2BcKil3V+U3Vl323ketyiu5v5gJVvEjyOGzli9cF9Kv3du0jyjlwmFJXtwKgKd7L5z6d7+N6PL1m6VbpXUegrfVNMkFg5vGg0E3g68+2h4Fzrc0blH4A9LMHBuc41O8mkjiOEVYQhpM6Byd1DpMmCE7V/c1VSd7aQAuBaEkQVX/TzZXQiaZ5gdTDoz4FT88Uu+2jqMSxHS0ncU7qHJzUOUwn9pPg5L51cyGa6zHYTiR6VwemwWERHTTRI+lHjbNtiVxsYUpJvKyCPnfSIjCc1Dk4qXM4HKnzVDktwMh6JPm0JOGR/G3t3vD2VPLWLGDDyLucBrehEucqnJM6h2ng0e8ckkQTJ/XhQOWIY5talAKa7dFUT032wnNXIHtLETqSNxJ3VW0dJXFU4SMwoG0oWoRrbI4F1sPBSZ2Dg8NSQILCMVRpovke1bxI9EMfXYad5G1UaYuP6gTOW5QOn+MAlmpVzxFndSKyhWFM1kkLzzXnQHDzO4dhjJAIeHsiVo3cRX99sDARGPrcHkBSbhFIWSRq9dfxNa64jQJmoqzQc81h85XNIN1Cpn99HJzUOTg4qTsugvSQvFRzv6imdZE4V9bWmyMRAp4m8dpbLpC7IdUex+u0c3BS55BK6pj3bmoLRlaClpka44D5DbP5QeUYwTAuoI2p9vV6rhssJrWNH1YOTuocxtxUqoxcKhe0laFl60sVSD6YH1yOEYYqck3Em3EdLhImx3HCZDmXK3QOTuocptxQvgFp0bkt9MZjSIlwYucYmeC+bw6rgvdT55CK9RI+gySdJsm0yCJ11/PDyjHilDoHByd1DruDEXU2aM+FrQLm15trpCkwF3huLcfIQiE/BBzWBDe/cxgHlV8vWHmTMqeK1eQF7wEPnOMYOeBV3zisCl58hsNYxY5K/A0LKxdO6hyOiipgGRwtwkR2DpjXB6GKH1IOTuocrn7T5OBwRGD2xkvDXp284B5g8SBzjFxfC49U57A2uE+dg5M6B8dwZGsldASazzF+BEu8GodCflg5OKlzuDbMVy65oMrdTQNmyucTBQ5zUCipmMu5vW8I557UYM/N/NByWBs8UI7D/jCtX3uhoKa0FbcJEm6g2fzgcpiANKOao7CaC/nCxFLfRGEuP7QcXKlzjAxlZBzW602fO7e3VSjwUcgPLYeRKDG62xn7fJoeJV4IvHUxB1fqHCNIqRtThlZ6nWvTyttyjGys1+lLl3bOBYHKHSROEnbzw8rBSZ1jpBH766A/8KgFWO/pbUaul+fBuzYKQVVueA6Yl24GwPPIOZwcPKWNwzFwbu8zhICDdRBwIejynxtGPid1l0QJsB4DV4ZM4jDdbLMZ5F7FDy0HV+ocHJZT7PeAZv5vvtE+zuHrbAbzGsfgjT4XePtYx1HnhoLODFt+tP/OpnZQ4+DgpM7BYbOJgtQOc0Oh2+TPIp5zwXxzL4fxv0kcDYY0/LsvF34jqeAd1DicHjz6nWNkKDvjUQIstUm7D59ZD+YAN9faGqslETr7jfC3MyY3nOeRc3BS5+BwAuSbQOhzDPrwGbnwVCXtanq9oKhlwFwW2RaYALUYHSiJsRrS1Hqu2W4eDg5O6hwcNgC7WZcYQUjZRqjBU2CciVffxCONkiAjwmCBGK3VmrbQiGNi7IQojqaFiZMifGRknGbmsco18fdfAbrbBov7vJpfKByc1Dk4nAdSbtotgkI3VrGtN1PVYhT3Yo3tsgI6LwEz8VuKfFsEcgsWivdMU5s8WIrQ5+icEKmKApm6vUIzJnbbBIvBZrXjWQViYxapkzgODgcHD5TjGDnQHxFdJZDrKRPXbUqpW2mTCFZEpwTMi+DXngKm2kYqqPK9TYX08qrG1w9A03sIP4k5OLhS5+AQ1dozAqm3DCHW9WBsve/hyDXJeiBlm4yIs83YtyowFCOgCvxrMfn7G3P8mGI3xgKRz09gDg6u1Dk4QI8yBYsFRxlfkjafmtyN24apqXnSq6QZnwYmIs7o4kDGWQfieC9yDg6u1Dk4dCtTS0Y7M8KRSoZVJirv9SYq6INGfI9tYLzvutAkwpVuHVjNCZ2Dg5M6B4etsVoCQYmBccYHZjFyNsYa0GLiRMDYCcd6k4+Yith1meI3C33LOTg4OKlzcNhU/bcaUJ6mRtebSribTVTQV4wg6kKzG6Awq8k04bvlC5aCXOFYPcNPLA4O6eA+dQ4OS4P5ijeDyv/dIpDVaoukTkmLHC8RiNKa28HvlcZN4xwcnNQ5ODhMJ9sgQc2m6SR0S+Ve6yZ2S1gdODg4LAw3fgg4OJwM9eXdEJG4gzzzJmPWkHdzyVhqsWIq9eW7ybZwAoER6mPIKCNjDxn3c4XOweF4+P8CDADHIsSgKG488wAAAABJRU5ErkJggg=="

def render_sidebar():
    """Sidebar global con toggles y badges de versión."""
    with st.sidebar:
        st.markdown(f"### 🟦 Picking Orchestrator")
        st.caption(f"v{APP_VERSION}")
        st.caption("Beccacece Hnos SA — DPO 2.1")
        st.divider()

        st.session_state["dry_run"] = st.toggle(
            "🧪 Modo Dry-Run",
            value=st.session_state.get("dry_run", False),
            help="Procesa todo pero NO genera PDFs finales (solo preview en pantalla). "
                 "Recomendado durante la primera semana de validación.",
        )
        st.session_state["snapshot_on"] = st.toggle(
            "💾 Snapshot activo",
            value=st.session_state.get("snapshot_on", True),
            help=f"Guarda CAR + MASTER + PDFs + log en {SNAPSHOT_DIR}/<fecha>/<run_id>/ "
                 "cada corrida. Default ON. Local hasta Sprint 5 (después → Drive).",
        )

        st.divider()
        st.caption(f"📅 Hoy: {datetime.today().strftime('%d/%m/%Y')}")
        if st.session_state.get("dry_run"):
            st.warning("🧪 DRY-RUN ACTIVO\n\nNo se generan PDFs finales.")

        st.divider()
        with st.expander("🛠 Stack técnico", expanded=False):
            st.code(
                f"streamlit {st.__version__}\n"
                f"pandas {pd.__version__}\n"
                f"openpyxl {openpyxl.__version__}",
                language="text",
            )


def _file_uploader(label: str, types: list[str], key: str, required: bool = True):
    """Wrapper de st.file_uploader con label rico."""
    suffix = " *" if required else " (opcional)"
    return st.file_uploader(
        label + suffix, type=types, key=key, accept_multiple_files=False,
    )


# ── TAB 1 — Planilla de Carga (REUSO v3.9) ─────────────────────────────────


# ════════════════════════════════════════════════════════════════════════════
# TAB 0 — ARCHIVOS (carga centralizada de CAR, Frescura y ANR)
# ════════════════════════════════════════════════════════════════════════════

def render_tab_archivos():
    """
    Tab central de carga de archivos. CAR.xlsx, Frescura 3.0.xlsx, ANR.xlsx y SR.xlsx
    se suben aquí una sola vez. Todas las demás pestañas consumen los archivos
    desde session_state (sin uploaders propios redundantes).
    v4.73: rediseño visual — badges de estado, descripción por archivo, ANR requerido,
           SR D+1 y ANR -1 sin ❌ cuando no están cargados.
    """
    # ── CSS v4.73 ──────────────────────────────────────────────────────────
    st.markdown(
        """
        <style>
        /* Ocultar texto auxiliar de tamaño de archivo */
        [data-testid="stFileUploader"] small,
        [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzoneInstructions"] small {
            display: none !important;
        }
        /* Dropzone compacto y uniforme */
        [data-testid="stFileUploaderDropzone"] {
            padding: 0.5rem 0.4rem !important;
            min-height: 56px !important;
            max-height: 56px !important;
        }
        [data-testid="stFileUploaderDropzoneInstructions"] {
            gap: 0.15rem !important;
        }
        /* Alertas compactas */
        [data-testid="stColumn"] .stAlert {
            padding: 3px 8px !important;
            font-size: 11.5px !important;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            border-radius: 6px !important;
        }
        /* Bloque de tarjeta por uploader */
        .arch-card {
            background: #1a1a2e;
            border: 1px solid #2d2d4e;
            border-radius: 10px;
            padding: 10px 10px 6px 10px;
            margin-bottom: 4px;
        }
        /* Título del archivo grande */
        .arch-title {
            font-size: 15px !important;
            font-weight: 700 !important;
            letter-spacing: 0.3px;
            margin-bottom: 2px !important;
            color: #e8e8f0;
        }
        /* Badge de tipo */
        .arch-badge {
            display: inline-block;
            font-size: 10px;
            font-weight: 600;
            padding: 1px 7px;
            border-radius: 20px;
            margin-bottom: 5px;
            letter-spacing: 0.4px;
        }
        .badge-req   { background:#1a3a5c; color:#60b4ff; border:1px solid #2a5a8c; }
        .badge-opt   { background:#1e2e1e; color:#6dbf6d; border:1px solid #2e4e2e; }
        .badge-miss  { background:#3a1a1a; color:#ff7070; border:1px solid #5a2a2a; }
        /* Descripción */
        .arch-desc {
            font-size: 10.5px;
            color: #8888aa;
            margin-top: 4px;
            line-height: 1.35;
        }
        /* Tabla de estado */
        .status-table { width:100%; border-collapse:collapse; font-size:12.5px; margin-top:4px; }
        .status-table th {
            background:#1a1a2e; color:#8888aa; font-weight:600;
            padding:5px 8px; text-align:left; border-bottom:1px solid #2d2d4e;
        }
        .status-table td { padding:5px 8px; border-bottom:1px solid #1e1e32; vertical-align:top; }
        .status-table tr:last-child td { border-bottom:none; }
        .dot-ok  { color:#4caf50; font-size:14px; }
        .dot-warn{ color:#ff9800; font-size:14px; }
        .dot-off { color:#555577; font-size:14px; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        "<p style='font-size:13px;color:#8888aa;margin-bottom:12px'>"
        "Subí los archivos fuente una sola vez. Quedan disponibles para "
        "<strong>todas las pestañas</strong> de la sesión.</p>",
        unsafe_allow_html=True,
    )

    col_a, col_b, col_c, col_d, col_e, col_f = st.columns(6)

    # ── COL A — CAR ────────────────────────────────────────────────────────
    with col_a:
        st.markdown(
            "<div class='arch-card'>"
            "<div class='arch-title'>🗂️ CAR</div>"
            "<span class='arch-badge badge-miss'>Requerido</span>"
            "<div class='arch-desc'>Planilla de Carga · Resumen · Camiones T2 · Proyección</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        car_file = st.file_uploader(
            "CAR.xlsx",
            type=["xlsx"],
            key="arch_car",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        if car_file:
            st.session_state["t1_car"] = car_file
            st.session_state["t2_car"] = car_file
            st.session_state["t3_car"] = car_file
            st.session_state["t4_car"] = car_file
            st.session_state["tc_car"] = car_file
            st.success(f"✅ {car_file.name}")
        elif st.session_state.get("t1_car"):
            st.info(f"📎 {st.session_state['t1_car'].name}")
        else:
            st.warning("Sin archivo")

    # ── COL B — FRESCURA ───────────────────────────────────────────────────
    with col_b:
        st.markdown(
            "<div class='arch-card'>"
            "<div class='arch-title'>🌿 Frescura</div>"
            "<span class='arch-badge badge-miss'>Requerido</span>"
            "<div class='arch-desc'>Planilla de Carga · Proyección Picking · Clasificación</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        fr_file = st.file_uploader(
            "Frescura 3.0.xlsx",
            type=["xlsx"],
            key="arch_fr",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        if fr_file:
            st.session_state["t1_fr"] = fr_file
            st.session_state["t4_fr"] = fr_file
            st.session_state["tc_fr"] = fr_file
            # v4.67: Parsear DDM inmediatamente
            try:
                import openpyxl as _ox67
                _wb67 = _ox67.load_workbook(fr_file, read_only=True, data_only=True)
                fr_file.seek(0)
                _ws67 = _wb67["DDM"]
                _rows67 = list(_ws67.iter_rows(min_row=1, values_only=True))
                if _rows67:
                    _hdr67 = [str(v).strip().upper() if v is not None else "" for v in _rows67[0]]
                    def _ci67(names):
                        for n in names:
                            if n in _hdr67:
                                return _hdr67.index(n)
                        return -1
                    _ic   = _ci67(["ARTÍCULO", "ARTICULO"])
                    _ig   = _ci67(["BULTOS X PALLET"])
                    _ihl  = _ci67(["VALOR HL"])
                    _ikg  = _ci67(["PESO KG"])
                    _ican = _ci67(["CAN", "CANCHA"])
                    _iun  = _ci67(["UNIDADES", "UN BULTO", "UN/BULTO", "UN X BULTO"])
                    if _ic  < 0: _ic  = 2
                    if _ig  < 0: _ig  = 6
                    if _iun < 0: _iun = 13
                    if _ihl < 0: _ihl = 15
                    if _ikg < 0: _ikg = 16
                    if _ican < 0: _ican = 14
                    _ddm_dict67 = {}
                    for _r67 in _rows67[1:]:
                        try:
                            _s = _r67[_ic]
                            if _s is None:
                                continue
                            _sid = int(float(str(_s)))
                            _ddm_dict67[_sid] = {
                                "bxp":      float(_r67[_ig])   if _r67[_ig]   else 50.0,
                                "hl_unit":  float(_r67[_ihl])  if _r67[_ihl]  else 0.0,
                                "kg_unit":  float(_r67[_ikg])  if _r67[_ikg]  else 0.0,
                                "un_bulto": float(_r67[_iun])  if _r67[_iun]  else 0.0,
                                "can":      str(_r67[_ican]).strip() if _r67[_ican] else "",
                            }
                        except Exception:
                            pass
                    st.session_state["df_ddm_dict"] = _ddm_dict67
                _wb67.close()
            except Exception:
                st.session_state["df_ddm_dict"] = {}
            st.success(
                f"✅ {fr_file.name[:14]}… · "
                f"{len(st.session_state.get('df_ddm_dict', {}))} SKUs"
            )
        elif st.session_state.get("t1_fr"):
            st.info(f"📎 {st.session_state['t1_fr'].name[:16]}…")
        else:
            st.warning("Sin archivo")

    # ── COL C — ANR ────────────────────────────────────────────────────────
    with col_c:
        st.markdown(
            "<div class='arch-card'>"
            "<div class='arch-title'>📊 ANR</div>"
            "<span class='arch-badge badge-miss'>Requerido</span>"
            "<div class='arch-desc'>Clasificación · Top SKUs · Top Clientes · Cierre · 🖨️ Boletas</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        anr_file = st.file_uploader(
            "ANR.xlsx",
            type=["xlsx"],
            key="arch_anr",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        if anr_file:
            st.session_state["tc_anr"] = anr_file
            try:
                anr_file.seek(0)
                xl_anr = pd.ExcelFile(anr_file)
                _sheet_anr = xl_anr.sheet_names[0]
                for _s in xl_anr.sheet_names:
                    if _s.upper() == "BASE":
                        _sheet_anr = _s
                        break
                anr_file.seek(0)
                _df_anr_raw = pd.read_excel(anr_file, sheet_name=_sheet_anr, header=0)
                st.session_state["anr_df"] = _df_anr_raw
                anr_file.seek(0)
            except Exception:
                st.session_state["anr_df"] = None
            st.success(f"✅ {anr_file.name}")
        elif st.session_state.get("tc_anr"):
            st.info(f"📎 {st.session_state['tc_anr'].name}")
        else:
            st.warning("Sin archivo")

    # ── COL D — SR ─────────────────────────────────────────────────────────
    with col_d:
        st.markdown(
            "<div class='arch-card'>"
            "<div class='arch-title'>💰 SR</div>"
            "<span class='arch-badge badge-req'>Req. Cierre</span>"
            "<div class='arch-desc'>Cierre del día · Totales por camión desde Chess</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        sr_file = st.file_uploader(
            "SR.xlsx",
            type=["xlsx"],
            key="arch_sr",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        if sr_file:
            st.session_state["cierre_sr"] = sr_file
            st.success(f"✅ {sr_file.name}")
        elif st.session_state.get("cierre_sr"):
            st.info(f"📎 {st.session_state['cierre_sr'].name}")
        else:
            st.warning("Sin archivo")

    # ── COL E — SR D+1 ─────────────────────────────────────────────────────
    with col_e:
        st.markdown(
            "<div class='arch-card'>"
            "<div class='arch-title'>📅 SR D+1</div>"
            "<span class='arch-badge badge-opt'>Opcional D+1</span>"
            "<div class='arch-desc'>Cierre D+1 · TotVal real con rechazos incluidos</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        sr_d1_file = st.file_uploader(
            "SR D+1.xlsx",
            type=["xlsx"],
            key="arch_sr_d1",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        if sr_d1_file:
            st.session_state["cierre_sr_d1"] = sr_d1_file
            st.success(f"✅ {sr_d1_file.name}")
        elif st.session_state.get("cierre_sr_d1"):
            st.info(f"📎 {st.session_state['cierre_sr_d1'].name}")
        else:
            st.info("Disponible D+1")

    # ── COL F — ANR -1 ─────────────────────────────────────────────────────
    with col_f:
        st.markdown(
            "<div class='arch-card'>"
            "<div class='arch-title'>📊 ANR -1</div>"
            "<span class='arch-badge badge-opt'>Opcional D+1</span>"
            "<div class='arch-desc'>Cierre D+1 · CTA CTE del día anterior (exclusivo)</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        anr_m1_file = st.file_uploader(
            "ANR -1.xlsx",
            type=["xlsx"],
            key="arch_anr_m1",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        if anr_m1_file:
            st.session_state["cierre_anr_m1"] = anr_m1_file
            st.success(f"✅ {anr_m1_file.name}")
        elif st.session_state.get("cierre_anr_m1"):
            st.info(f"📎 {st.session_state['cierre_anr_m1'].name}")
        else:
            st.info("Disponible D+1")

    # ── TABLA DE ESTADO ────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("##### 📋 Estado de archivos cargados")

    _status_rows = [
        ("CAR.xlsx",          "t1_car",       "req",  "Planilla de Carga · Resumen · Camiones T2 · Proyección"),
        ("Frescura 3.0.xlsx", "t1_fr",        "req",  "Planilla de Carga · Proyección Picking · Clasificación"),
        ("ANR.xlsx",          "tc_anr",       "req",  "Clasificación · Top SKUs · Top Clientes · Cierre · 🖨️ Boletas"),
        ("SR.xlsx",           "cierre_sr",    "req",  "💰 Cierre — totales por camión desde Chess"),
        ("SR D+1.xlsx",       "cierre_sr_d1", "opt",  "📅 Cierre D+1 — TotVal real con rechazos"),
        ("ANR -1.xlsx",       "cierre_anr_m1","opt",  "📅 Cierre D+1 — CTA CTE día anterior (exclusivo)"),
    ]

    _table_html = (
        "<table class='status-table'>"
        "<tr>"
        "<th style='width:20px'></th>"
        "<th>Archivo</th>"
        "<th>Nombre cargado</th>"
        "<th>Alimenta</th>"
        "</tr>"
    )
    for _nombre, _key, _tipo, _desc in _status_rows:
        _f = st.session_state.get(_key)
        if _f:
            _dot   = "<span class='dot-ok'>●</span>"
            _label = f"<span style='color:#c8ffc8'>{_f.name}</span>"
        elif _tipo == "opt":
            _dot   = "<span class='dot-off'>●</span>"
            _label = "<span style='color:#555577'>— (D+1)</span>"
        else:
            _dot   = "<span class='dot-warn'>●</span>"
            _label = "<span style='color:#ff9898'>Sin cargar</span>"
        _table_html += (
            f"<tr>"
            f"<td>{_dot}</td>"
            f"<td><strong>{_nombre}</strong></td>"
            f"<td>{_label}</td>"
            f"<td style='color:#8888aa'>{_desc}</td>"
            f"</tr>"
        )
    _table_html += "</table>"
    st.markdown(_table_html, unsafe_allow_html=True)

    st.markdown(
        "<p style='font-size:11px;color:#555577;margin-top:8px'>"
        "Los archivos persisten durante toda la sesión. "
        "Si cambia el día, recargá los archivos aquí.</p>",
        unsafe_allow_html=True,
    )


def render_tab_planilla():
    st.subheader("📦 Planilla de Carga")
    st.caption(
        "Genera el PDF multi-página de la planilla diaria de picking. "
        "Lógica idéntica a la app v3.9.1 validada en producción. "
        "**Los archivos se cargan en la pestaña 📁 Archivos.**"
    )

    car_file = st.session_state.get("t1_car")
    fr_file  = st.session_state.get("t1_fr")

    if not car_file or not fr_file:
        st.info("⬅️ Subí **CAR.xlsx** y **Frescura 3.0.xlsx** en la pestaña **📁 Archivos** para continuar.")
        return

    if st.button("🚀 Generar Planilla de Carga", type="primary", key="t1_gen"):
        try:
            with st.spinner("Cargando Frescura..."):
                api, ddm, fr, fr_diag = load_frescura(fr_file)
            log_event("info", f"Frescura cargada: API={len(api)} | DDM={len(ddm)} | FR={len(fr)}")

            with st.spinner("Cargando CAR..."):
                car_df, blue_audit = load_car(car_file, ddm=ddm)
            log_event("info", f"CAR cargado: {len(car_df)} filas | bloque azul={len(blue_audit)}")

            if st.session_state.get("dry_run"):
                st.success(f"✓ DRY-RUN OK — {len(car_df)} filas listas. PDF no generado.")
                with st.expander("Preview CAR (primeras 20 filas)"):
                    st.dataframe(car_df.head(20), width="stretch")
                return

            with st.spinner("Generando PDF (puede tardar)..."):
                pdf_bytes, stats = generate_pdf(car_df, api, ddm, fr)

            log_event("info", f"Planilla generada: {stats['total_pages']} páginas, "
                              f"{len(stats['repartos'])} repartos")

            st.success(
                f"✓ Planilla lista — {stats['total_pages']} páginas | "
                f"{len(stats['repartos'])} repartos"
            )

            st.download_button(
                "⬇ Descargar Planilla de Carga (PDF)",
                data=pdf_bytes,
                file_name=_stamp("Planilla_Carga", "pdf"),
                mime="application/pdf",
                type="primary",
                width="stretch",
            )

            with st.expander(f"📊 Alertas y diagnósticos"):
                a, b, c, d = st.columns(4)
                a.metric("🔴 RED", len(stats["red"]))
                b.metric("🟡 YELLOW", len(stats["yellow"]))
                c.metric("📦 c/Pallet", len(stats["pallet_applied"]))
                d.metric("⚠ Sin BXP", len(stats["miss_bxp"]))
                if stats["orphans_no_ddm"]:
                    st.warning("Huérfanos sin DDM:")
                    for o in stats["orphans_no_ddm"]:
                        st.write(f"- {o}")

            # Stash para snapshot (Sprint 1: solo en memoria)
            st.session_state["last_planilla_pdf"] = pdf_bytes

        except Exception as e:
            log_event("error", f"Error en Planilla: {e}")
            st.error(f"❌ Error generando Planilla: {e}")
            with st.expander("Stack trace"):
                import traceback
                st.code(traceback.format_exc())


# ── TAB 2 — Resumen por Camión (REUSO v3.9) ────────────────────────────────

def render_tab_resumen():
    st.subheader("📋 Resumen de Carga por Camión")
    st.caption(
        "Genera un PDF compacto con bultos consolidados por SKU por camión "
        "(hoja AGR del CAR). **Se genera automáticamente** al subir el CAR "
        "en la pestaña 📦 Planilla de Carga."
    )

    # Auto-reusar del Tab Archivos / Planilla
    car_use = st.session_state.get("t1_car") or st.session_state.get("t2_car")

    if not car_use:
        st.info("⬅️ Subí el **CAR.xlsx** en la pestaña **📁 Archivos** para generar el resumen automáticamente.")
        return

    if st.session_state.get("dry_run"):
        try:
            df_agr = load_agr(car_use)
            st.success(
                f"✓ DRY-RUN OK — {len(df_agr)} filas SKU/camión, "
                f"{df_agr['chofer'].nunique()} camiones."
            )
            st.dataframe(df_agr.head(30), width="stretch")
        except Exception as e:
            st.error(f"❌ Error leyendo AGR: {e}")
        return

    # ── Generación automática del PDF ────────────────────────────────────────
    # Cacheamos el resultado en session_state usando hash del file para no
    # regenerar el PDF en cada re-run de Streamlit.
    car_hash = hashlib.md5(car_use.getvalue()).hexdigest()
    cache_key = f"resumen_pdf_{car_hash}"

    if cache_key not in st.session_state:
        try:
            with st.spinner("Generando PDF resumen automáticamente..."):
                pdf_bytes, stats = build_resumen_carga_pdf(car_use)
            st.session_state[cache_key] = (pdf_bytes, stats)
            log_event("info",
                      f"Resumen auto-generado: {stats['camiones']} cam, "
                      f"{stats['filas_sku']} filas SKU, {stats['total_pages']} págs")
        except Exception as e:
            log_event("error", f"Error en Resumen: {e}")
            st.error(f"❌ Error generando Resumen: {e}")
            with st.expander("Stack trace"):
                import traceback
                st.code(traceback.format_exc())
            return

    pdf_bytes, stats = st.session_state[cache_key]

    st.success(
        f"✓ Resumen listo — {stats['camiones']} camiones | "
        f"{stats['filas_sku']} filas SKU | {stats['total_pages']} páginas"
    )

    st.download_button(
        "⬇ Descargar Resumen por Camión (PDF)",
        data=pdf_bytes,
        file_name=_stamp("Resumen_Camiones", "pdf"),
        mime="application/pdf",
        type="primary",
        width="stretch",
        key="resumen_dl",
    )

    st.session_state["last_resumen_pdf"] = pdf_bytes


# ── TAB 3 — Camiones T2: PDF vía Apps Script Web App (v4.18.1) ─────────────
#
#  CAMBIOS v4.18.1:
#  - Eliminada dependencia de pypdf. El merge ahora lo hace un Apps Script
#    Web App deployado en el propio Google Sheet T2 Status Carga.
#  - Eliminada detección via CAR.xlsx. El Apps Script detecta directamente
#    qué camiones tienen reparto leyendo F3 de cada hoja numérica del Sheet
#    (F3 > 0 = camión con carga). Ya no hace falta CAR para esta pestaña.
#  - PDF embebido (iframe) + botón de descarga. Listo para Ctrl+P.
#
#  ARQUITECTURA:
#    Streamlit (botón) → GET https://script.google.com/.../exec
#                     → Apps Script:
#                         1. Lee F3 de cada hoja 101-129
#                         2. Crea copia temporal del Sheet
#                         3. Oculta hojas sin reparto
#                         4. Exporta PDF (Carta, 1 hoja por camión)
#                         5. Borra la copia temporal
#                     → JSON con PDF en base64
#    Streamlit muestra preview + botón descarga
#
#  SETUP INICIAL: ver T2_README.md (deploy del Apps Script + pegar URL acá).
# ─────────────────────────────────────────────────────────────────────────────

# URL del Apps Script Web App. Pegar la URL /exec del deploy aquí.
# Formato: https://script.google.com/macros/s/AKfycb.../exec
_T2_WEB_APP_URL = "https://script.google.com/macros/s/AKfycbwcFO8gby7kZvEhiWoUg_lBOL8Xhsqzph7YdycJkNcZoZsvCxqg1BoBeVe6CIFae2Gt/exec"

_T2_REQUEST_TIMEOUT = 120  # segundos (makeCopy + export tarda ~15-30s)


def _t2_rotate_to_landscape(pdf_bytes: bytes) -> bytes:
    """
    Rota todas las páginas del PDF 90° (portrait → landscape).
    Intenta pypdf primero; si no está instalado, cae a pikepdf;
    si tampoco está, devuelve el PDF sin rotar (sin crashear la app).
    """
    # Intento 1: pypdf (puede no estar disponible en Streamlit Cloud)
    if _PYPDF_AVAILABLE:
        try:
            from pypdf import PdfReader, PdfWriter
            reader = PdfReader(io.BytesIO(pdf_bytes))
            writer = PdfWriter()
            for page in reader.pages:
                page.rotate(90)
                writer.add_page(page)
            buf = io.BytesIO()
            writer.write(buf)
            buf.seek(0)
            return buf.read()
        except Exception:
            pass  # Cae al siguiente método

    # Intento 2: pikepdf (disponible en Streamlit Cloud por defecto)
    try:
        import pikepdf
        with pikepdf.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                page.rotate(90, relative=True)
            buf = io.BytesIO()
            pdf.save(buf)
            buf.seek(0)
            return buf.read()
    except ImportError:
        pass  # Tampoco está pikepdf
    except Exception:
        pass

    # Fallback: devolver sin rotar para no romper la descarga
    log_event("WARNING", "No se pudo rotar el PDF (pypdf y pikepdf no disponibles). Se descarga en portrait.")
    return pdf_bytes


def _t2_fetch_pdf_from_apps_script() -> tuple[bytes | None, str, list[str], str | None]:
    """
    Llama al Apps Script Web App y devuelve:
        (pdf_bytes, filename, trucks, error_msg)
    Si hay error, pdf_bytes es None y error_msg trae el detalle.
    """
    import base64
    import requests

    try:
        resp = requests.get(_T2_WEB_APP_URL, timeout=_T2_REQUEST_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
    except requests.Timeout:
        return None, "", [], "Timeout: el Apps Script tardó más de 2 minutos."
    except requests.RequestException as e:
        return None, "", [], f"Error de red llamando al Apps Script: {e}"
    except ValueError:
        return None, "", [], "Respuesta inválida del Apps Script (no es JSON)."

    if not data.get("success"):
        return None, "", [], f"Apps Script devolvió error: {data.get('error', 'desconocido')}"

    try:
        pdf_bytes = base64.b64decode(data["pdf_base64"])
    except Exception as e:
        return None, "", [], f"Error decodificando PDF: {e}"

    return (
        pdf_bytes,
        data.get("filename", "T2_Camiones.pdf"),
        data.get("trucks", []),
        None,
    )


def render_tab_t2():
    import base64

    st.subheader("🚛 Camiones T2 — PDF único para imprimir")
    st.caption(
        "v4.18.1 — Detecta automáticamente los camiones con reparto en el "
        "**Sheet T2 Status Carga** (F3 > 0) y genera un PDF combinado "
        "(1 hoja por camión, tamaño Carta) listo para imprimir."
    )

    with st.expander("ℹ️ Cómo funciona", expanded=False):
        st.markdown(
            """
            1. Al apretar **Generar PDF**, Streamlit llama a un Apps Script
               deployado dentro del propio Sheet T2 Status Carga.
            2. El Apps Script lee la celda **F3** de cada hoja numérica
               (101–129). Si F3 > 0, ese camión tiene reparto y se incluye.
            3. Crea una copia temporal del Sheet, oculta las hojas que no van,
               exporta como PDF (Carta, con header de fecha/título) y borra
               la copia temporal.
            4. Devuelve el PDF a Streamlit, que lo muestra embebido + botón
               de descarga. Ctrl+P sobre el preview imprime directo.

            > **Ya no hace falta subir el CAR** para esta pestaña — la detección
            > es directa desde el Sheet, que se mantiene actualizado con la
            > planilla de inputs.
            >
            > **Setup inicial:** ver `T2_README.md` para el deploy del Apps Script.
            """
        )

    st.divider()

    # ── Validar config ────────────────────────────────────────────────────────
    if _T2_WEB_APP_URL.startswith("PEGAR"):
        st.error(
            "⚠️ Falta configurar `_T2_WEB_APP_URL` en el código. "
            "Hacé el deploy del Apps Script (ver T2_README.md) y pegá la URL "
            "`/exec` en la constante al inicio de esta sección."
        )
        return

    # ── Botones de acción ─────────────────────────────────────────────────────
    col_gen, col_clear = st.columns([3, 1])
    with col_gen:
        gen_clicked = st.button(
            "🖨️  Generar PDF Camiones",
            type="primary",
            width="stretch",
            key="t3_btn_generate",
        )
    with col_clear:
        clear_clicked = st.button(
            "🗑️ Limpiar",
            width="stretch",
            key="t3_btn_clear",
        )

    if clear_clicked:
        for k in ("t3_pdf_bytes", "t3_pdf_filename", "t3_pdf_trucks"):
            st.session_state.pop(k, None)
        st.rerun()

    if gen_clicked:
        with st.spinner("Generando PDF... (15–30 segundos)"):
            pdf_bytes, filename, trucks, error = _t2_fetch_pdf_from_apps_script()

        if error:
            st.error(f"❌ {error}")
            log_event("error", f"T2 v4.28.0: {error}")
        elif pdf_bytes:
            # Nota: la rotación landscape se maneja en el Apps Script (parámetro
            # `fitw=true&size=7` en la URL de export). No se rota acá para evitar
            # distorsión del contenido (rotar un PDF portrait ≠ landscape real).
            st.session_state["t3_pdf_bytes"]    = pdf_bytes
            st.session_state["t3_pdf_filename"] = filename
            st.session_state["t3_pdf_trucks"]   = trucks
            log_event(
                "info",
                f"T2 v4.28.0: PDF generado (landscape nativo) | {len(trucks)} camiones | "
                f"{len(pdf_bytes)//1024} KB",
            )

    # ── Render del PDF cacheado ───────────────────────────────────────────────
    pdf_cached = st.session_state.get("t3_pdf_bytes")
    if pdf_cached:
        trucks   = st.session_state.get("t3_pdf_trucks", [])
        filename = st.session_state.get("t3_pdf_filename", "T2_Camiones.pdf")

        st.success(
            f"✅ {len(trucks)} camión(es) listos para imprimir: "
            f"**{', '.join(trucks)}** ({len(pdf_cached)//1024} KB)"
        )

        st.download_button(
            f"⬇  Descargar {filename}",
            data=pdf_cached,
            file_name=filename,
            mime="application/pdf",
            type="primary",
            width="stretch",
            key="t3_pdf_dl",
        )

        # Preview embebido — Ctrl+P imprime directo
        b64 = base64.b64encode(pdf_cached).decode("utf-8")
        st.markdown(
            f"""
            <iframe
                src="data:application/pdf;base64,{b64}"
                width="100%"
                height="900"
                style="border: 1px solid #d0d0d0; border-radius: 8px; margin-top: 12px;">
            </iframe>
            """,
            unsafe_allow_html=True,
        )

# ── TAB 4 — Proyección Picking ×4 — v4.6 ──────────────────────────────────
#
#  CAMBIOS v4.7:
#  - Fix: hora estimada de fin calculada sobre BULTOS REALES por cancha
#    (no sobre UP). Fórmula: bultos_cancha / (vel_bult_h × personas) + inicio
#  - Fix: asignaciones entre canchas ahora afectan correctamente los bultos
#    de cada cancha para el cálculo de tiempo (nueva fn _t4_calcular_bultos_por_cancha)
#  - Los metrics "Bult CX" muestran delta cuando hay asignaciones activas
#
#  CAMBIOS v4.6:
#  - Tabla basada en BULTOS UP (col R Frescura DDM = UNIDAD PKT)
#  - Paletas enteras (floor), redondeo inteligente (≥0.97 → techo)
#  - Hora de inicio única (un solo input compartido para todas las canchas)
#  - Hora estimada de fin calculada por bultos/hora × personas
#  - Sin toggle "Sacar ceros" (siempre filtrado)
#  - Columnas ASIGN siempre en 0 por default
#  - Asignación de carga: dropdown para mover carga entre canchas
#  - PDF mejorado: sin encimamiento, bultos UP por cancha, info asignaciones
#  - Columna camiones correcta: 101=1, 106=2, etc.
# ─────────────────────────────────────────────────────────────────────────────

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Constantes Tab 4 ─────────────────────────────────────────────────────────

_T4_CANCHAS      = ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "MKPL"]
_T4_CANCHAS_PDF  = ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "MKPL"]   # v4.36: incluye MKPL

# v4.64.0 — Etiquetas con números naturales en lugar de romanos (display only)
_ROMAN_TO_NAT = {
    "CANCHA I":   "C1",
    "CANCHA II":  "C2",
    "CANCHA III": "C3",
    "CANCHA IV":  "C4",
    "MKPL":       "MKPL+Merch",
}

def _cn_short(cn: str) -> str:
    """Convierte 'CANCHA I' -> 'C1', 'CANCHA II' -> 'C2', etc. (display only)."""
    return _ROMAN_TO_NAT.get(cn, cn.replace("CANCHA ", "C"))

_T4_VEL_DEFAULT = {
    "CANCHA I":   370,
    "CANCHA II":  360,
    "CANCHA III": 390,
    "CANCHA IV":  400,
    "MKPL":       390,
}
_T4_HORA_INICIO_DEFAULT = {
    "CANCHA I":   (17, 30),
    "CANCHA II":  (17, 35),
    "CANCHA III": (17, 35),
    "CANCHA IV":  (17, 40),
    "MKPL":       (17, 45),
}
_T4_CAP_MAX = 9

_T4_CANCHA_COLORS = {
    "CANCHA I":   colors.HexColor("#1565C0"),
    "CANCHA II":  colors.HexColor("#2E7D32"),
    "CANCHA III": colors.HexColor("#E65100"),
    "CANCHA IV":  colors.HexColor("#6A1B9A"),
    "MKPL":       colors.HexColor("#00838F"),
}


def _t4_safe(v, default=0.0):
    if v is None:
        return default
    try:
        f = float(v)
        return default if (f != f) else f
    except (ValueError, TypeError):
        return default


def _t4_norm_cancha(val) -> str:
    """
    v4.33.1 — Normaliza nombre de cancha. FIX: el orden de match anterior
    usaba startswith() con "CANCHA I" primero → "CANCHA II/III/IV" caían
    todos a CANCHA I por prefijo. Ahora se prueba IGUALDAD primero, luego
    canónicos cortos (CI/CII/.../C4/4/etc), y como fallback un regex de
    número romano. Cancha V (legacy) → MKPL.
    """
    if val is None:
        return "SIN CANCHA"
    s = str(val).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Colapsar espacios múltiples a uno solo
    import re as _re_ws
    s = _re_ws.sub(r'\s+', ' ', s)
    if not s:
        return "SIN CANCHA"

    # MKPL / Marketplace / Cancha V legacy
    if any(k in s for k in ("MKPL", "MARKETPLACE", "MERCH")):
        return "MKPL"

    # 1) Igualdad exacta con nombres completos
    for c in ("CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV"):
        if s == c:
            return c
    if s == "CANCHA V":
        return "MKPL"

    # 2) Canónicos cortos y numéricos
    canon_short = {
        "CI": "CANCHA I",   "C1": "CANCHA I",   "1": "CANCHA I",
        "CII": "CANCHA II", "C2": "CANCHA II",  "2": "CANCHA II",
        "CIII": "CANCHA III","C3": "CANCHA III","3": "CANCHA III",
        "CIV": "CANCHA IV", "C4": "CANCHA IV",  "4": "CANCHA IV",
        "CV": "MKPL",       "C5": "MKPL",       "5": "MKPL",
    }
    if s in canon_short:
        return canon_short[s]

    # 3) Regex "CANCHA <romano>" — orden de más largo a más corto evita
    #    que "CANCHA III" matchee con "CANCHA I".
    import re as _re
    m = _re.search(r'CANCHA\s*(IV|III|II|V|I)\b', s)
    if m:
        roman = m.group(1)
        rmap = {"I": "CANCHA I", "II": "CANCHA II", "III": "CANCHA III",
                "IV": "CANCHA IV", "V": "MKPL"}
        return rmap[roman]

    # 4) Regex numérico "C<n>" o solo "<n>"
    m2 = _re.search(r'\bC?\s*([1-5])\b', s)
    if m2:
        nmap = {"1": "CANCHA I", "2": "CANCHA II", "3": "CANCHA III",
                "4": "CANCHA IV", "5": "MKPL"}
        return nmap[m2.group(1)]

    return "SIN CANCHA"


@st.cache_data(show_spinner=False)
def _t4_extract_cv(cell_val):
    """
    v4.8 — Extrae el valor cacheado (COMPUTED_VALUE) de fórmulas IMPORTRANGE
    de Google Sheets exportadas como .xlsx.
    Formato: =IFERROR(__xludf.DUMMYFUNCTION(...), "VALOR_CACHEADO")
    Si el cell_val no es fórmula, lo devuelve tal cual.
    """
    import re as _re
    if cell_val is None:
        return None
    if not isinstance(cell_val, str):
        return cell_val
    s = cell_val
    m = _re.search(r',\s*"([^"]*)"\s*\)\s*$', s)
    if m:
        v = m.group(1)
        return None if v == "" else v
    m2 = _re.search(r',\s*([\d.]+)\s*\)\s*$', s)
    if m2:
        try:
            return float(m2.group(1))
        except Exception:
            return m2.group(1)
    return cell_val


def _t4_find_ddm_col(headers: list, *candidates) -> int:
    """
    v4.33 — Busca columna en headers DDM por nombre (normalizado, case/acentos
    insensitive). Devuelve índice 0-based o -1.
    """
    def _nm(s):
        if s is None:
            return ""
        s = str(s).strip().upper()
        return ''.join(c for c in unicodedata.normalize('NFD', s)
                       if unicodedata.category(c) != 'Mn')
    norm_h = [_nm(h) for h in headers]
    for cand in candidates:
        c = _nm(cand)
        for i, h in enumerate(norm_h):
            if h == c:
                return i
        for i, h in enumerate(norm_h):
            if c in h:
                return i
    return -1


def _t4_load_car_proyeccion(car_bytes: bytes, fr_bytes: bytes) -> dict:
    """
    v4.33 — Proyección Picking corregida.

    Lógica de UP por SKU × camión:
        bultos_eq = Bultos + Unids / unidades_por_bulto    (sueltas → bultos)
        UP_total  = bultos_eq / BXP                        (BXP = bultos por pallet)
        AE_SKU    = floor(UP_total)                        ← paletas enteras
        PICK_SKU  = UP_total - floor(UP_total)             ← fracción a cancha

    Lecturas:
      - CAR Hoja1: K=Transporte (camión), R=Artículo (SKU),
                   T=Bultos, V=Unids (sueltas).
                   Detección dinámica de fila vacía separadora →
                   VE (antes) + CHESS (después), ambas procesadas con flag 'fuente'.
      - DDM (hoja Frescura): columnas buscadas por NOMBRE de header
          BXP        ← 'BULTOS X PAL' / 'BXP'           (fallback col F = idx 5)
          un_bulto   ← 'UNIDADES' / 'UN BULTO' / 'UN'    (fallback col M = idx 12)
          can        ← 'CAN' / 'CANCHA'                 (fallback col O = idx 14)

    Retorna mismo schema que v4.32 para compatibilidad con UI/PDF:
      {df, fecha, fuente, mix_picking, tot_pick, tot_ae}
    """
    import datetime as _dt
    import math as _math

    # ── 1. DDM desde Frescura — búsqueda por header ──────────────────────────
    try:
        wb_fr = openpyxl.load_workbook(io.BytesIO(fr_bytes), read_only=True)
        ws_ddm = wb_fr["DDM"]
        rows_iter = ws_ddm.iter_rows(values_only=True)
        headers_ddm = list(next(rows_iter))

        idx_sku  = _t4_find_ddm_col(headers_ddm, "ARTICULO", "ARTÍCULO", "SKU", "CODIGO", "CÓDIGO")
        if idx_sku < 0:
            idx_sku = 2  # fallback col C
        idx_bxp  = _t4_find_ddm_col(headers_ddm, "BULTOS X PAL", "BULTOS X PALLET", "BXP", "BULTOS/PAL")
        if idx_bxp < 0:
            idx_bxp = 5  # fallback col F
        idx_un   = _t4_find_ddm_col(headers_ddm, "UNIDADES", "UN BULTO", "UN/BULTO", "UN X BULTO", "UN")
        if idx_un < 0:
            idx_un = 12  # fallback col M
        idx_can  = _t4_find_ddm_col(headers_ddm, "CAN", "CANCHA")
        if idx_can < 0:
            idx_can = 14  # fallback col O
        # v4.35: HL/bulto (col P, idx 15) y KG/bulto (col Q, idx 16)
        idx_hl   = _t4_find_ddm_col(headers_ddm, "VALOR HL", "HL")
        if idx_hl < 0:
            idx_hl = 15  # fallback col P
        idx_kg   = _t4_find_ddm_col(headers_ddm, "PESO KG", "PESO", "KG")
        if idx_kg < 0:
            idx_kg = 16  # fallback col Q

        ddm_rows = []
        max_idx = max(idx_sku, idx_bxp, idx_un, idx_can, idx_hl, idx_kg)
        for row in rows_iter:
            if row is None or len(row) <= max_idx:
                continue
            sku_raw = _t4_extract_cv(row[idx_sku])
            bxp_raw = _t4_extract_cv(row[idx_bxp])
            un_raw  = _t4_extract_cv(row[idx_un])
            can_raw = _t4_extract_cv(row[idx_can])
            hl_raw  = _t4_extract_cv(row[idx_hl])
            kg_raw  = _t4_extract_cv(row[idx_kg])
            if sku_raw is None:
                continue
            try:
                sku_int = int(float(str(sku_raw)))
                bxp_val = float(bxp_raw) if bxp_raw not in (None, "", "None") else 0.0
                un_val  = float(un_raw)  if un_raw  not in (None, "", "None") else 0.0
                can_str = str(can_raw).strip() if can_raw not in (None, "", "None") else "SIN CANCHA"
                hl_val  = float(hl_raw)  if hl_raw  not in (None, "", "None") else 0.0
                kg_val  = float(kg_raw)  if kg_raw  not in (None, "", "None") else 0.0
                upkt_val = (1.0 / bxp_val) if bxp_val > 0 else 0.0
                ddm_rows.append({
                    "sku":      sku_int,
                    "bxp":      bxp_val,
                    "un_bulto": un_val,
                    "can":      can_str,
                    "upkt":     upkt_val,
                    "hl_unit":  hl_val,   # HL por bulto
                    "kg_unit":  kg_val,   # KG por bulto
                })
            except (ValueError, TypeError):
                pass
        wb_fr.close()
        df_ddm = pd.DataFrame(ddm_rows).drop_duplicates("sku")
    except Exception:
        df_ddm = pd.DataFrame(columns=["sku", "bxp", "un_bulto", "can", "upkt", "hl_unit", "kg_unit"])

    # ── 2. CAR Hoja1 — detección dinámica VE / CHESS ─────────────────────────
    raw = pd.read_excel(io.BytesIO(car_bytes), sheet_name=0, header=None)
    if len(raw) < 2:
        raise ValueError("CAR vacío o sin filas")
    header = raw.iloc[0].tolist()

    # Localizar primera fila completamente vacía (separador VE/CHESS).
    split_idx = None
    for i in range(1, len(raw)):
        row_vals = raw.iloc[i].tolist()
        if all(v is None or (isinstance(v, float) and pd.isna(v))
               or (isinstance(v, str) and v.strip() == "")
               for v in row_vals):
            split_idx = i
            break

    body = raw.iloc[1:].copy()
    body.columns = header
    if split_idx is not None:
        # offset por header eliminado
        sep_local = split_idx - 1
        body = body.reset_index(drop=True)
        body["fuente"] = ""
        body.loc[:sep_local - 1, "fuente"] = "VE"
        body.loc[sep_local + 1:, "fuente"] = "CHESS"
        body = body.drop(index=sep_local)
    else:
        body["fuente"] = "CHESS"
    body = body.reset_index(drop=True)

    # Normalizar tipos y filtrar (mismas reglas que v4.32 + ambas fuentes)
    body["Artículo"]   = pd.to_numeric(body["Artículo"],   errors="coerce")
    body["Transporte"] = pd.to_numeric(body["Transporte"], errors="coerce")
    body["Bultos"]     = pd.to_numeric(body["Bultos"],     errors="coerce").fillna(0)
    body["Unids"]      = pd.to_numeric(body["Unids"],      errors="coerce").fillna(0)
    body = body.dropna(subset=["Artículo", "Transporte"])
    body = body[body["Artículo"] > 0]
    body["Artículo"]   = body["Artículo"].astype(int)
    body["Transporte"] = body["Transporte"].apply(lambda x: int(float(x)) if pd.notna(x) else x)

    # Fecha
    try:
        fecha = pd.to_datetime(body["Fecha Mvto"].dropna().iloc[0]).date()
    except Exception:
        fecha = _dt.date.today()

    # ── 3. Excluir envases ───────────────────────────────────────────────────
    body = body[~body.apply(
        lambda r: is_envase(int(r["Artículo"]), str(r.get("Descripción Artículo", ""))), axis=1
    )]

    # ── 4. Agregar por (Transporte, Artículo) — sumando bultos Y sueltas ─────
    grp = (body.groupby(["Transporte", "Artículo", "Descripción Artículo"], as_index=False)
           .agg(blt_raw=("Bultos", "sum"), unids=("Unids", "sum")))
    grp = grp[(grp["blt_raw"] > 0) | (grp["unids"] > 0)]
    grp.rename(columns={"Transporte": "cam", "Artículo": "sku"}, inplace=True)

    # ── 5. Merge DDM (bxp, un_bulto, can, hl_unit, kg_unit) ──────────────────
    if not df_ddm.empty:
        grp = grp.merge(
            df_ddm[["sku", "bxp", "un_bulto", "can", "upkt", "hl_unit", "kg_unit"]],
            on="sku", how="left",
        )
    else:
        grp["bxp"] = 0.0; grp["un_bulto"] = 0.0; grp["can"] = "SIN CANCHA"
        grp["upkt"] = 0.0; grp["hl_unit"] = 0.0; grp["kg_unit"] = 0.0

    grp["bxp"]      = grp["bxp"].fillna(0.0)
    grp["un_bulto"] = grp["un_bulto"].fillna(0.0)
    grp["upkt"]     = grp["upkt"].fillna(0.0)
    grp["hl_unit"]  = grp["hl_unit"].fillna(0.0)
    grp["kg_unit"]  = grp["kg_unit"].fillna(0.0)
    grp["can"]      = grp["can"].fillna("SIN CANCHA")

    # ── 6. Conversión sueltas → bultos_eq, luego split AE/Picking FLOOR ──────
    def _split_row(row):
        bultos    = float(row["blt_raw"])
        sueltas   = float(row["unids"])
        bxp_      = float(row["bxp"])
        un_bulto_ = float(row["un_bulto"])
        hl_u_     = float(row["hl_unit"])
        kg_u_     = float(row["kg_unit"])

        # 6a. Convertir sueltas a bultos equivalentes
        if un_bulto_ > 0:
            bultos_eq = bultos + (sueltas / un_bulto_)
        else:
            bultos_eq = bultos

        # 6b. UP total y split FLOOR estricto
        if bxp_ > 0:
            up_total = bultos_eq / bxp_
            pall_ae  = _math.floor(up_total)
            up_pick  = max(0.0, up_total - pall_ae)
            bult_ae  = float(pall_ae) * bxp_
            bult_ae  = min(bult_ae, bultos_eq)
            bult_pick = max(0.0, bultos_eq - bult_ae)
        else:
            pall_ae   = 0
            bult_ae   = 0.0
            bult_pick = bultos_eq
            up_pick   = 0.0

        # 6c. HL y KG por bulto (DDM col P y Q)
        hl_pick = bult_pick * hl_u_
        kg_pick = bult_pick * kg_u_
        hl_ae_v = bult_ae   * hl_u_
        kg_ae_v = bult_ae   * kg_u_

        return pd.Series({
            "bultos_eq": bultos_eq,
            "pall_ae":   float(pall_ae),
            "bult_ae":   bult_ae,
            "bult_pick": bult_pick,
            "up_pick":   up_pick,
            "hl_pick":   hl_pick,
            "kg_pick":   kg_pick,
            "hl_ae":     hl_ae_v,
            "kg_ae":     kg_ae_v,
        })

    split = grp.apply(_split_row, axis=1)
    grp = pd.concat([grp, split], axis=1)

    # ── 7. Normalizar cancha ─────────────────────────────────────────────────
    grp["cancha_norm"] = grp["can"].apply(_t4_norm_cancha)

    # ── 8. Pivotar por camión × cancha (igual que v4.32) ─────────────────────
    pick_grp = grp.groupby(["cam", "cancha_norm"], as_index=False).agg(
        bult_pick=("bult_pick", "sum"),
        up_pick  =("up_pick",   "sum"),
        hl_pick  =("hl_pick",   "sum"),
        kg_pick  =("kg_pick",   "sum"),
    )
    ae_grp = grp.groupby("cam", as_index=False).agg(
        bult_ae=("bult_ae", "sum"),
        up_ae  =("pall_ae", "sum"),
        hl_ae  =("hl_ae",   "sum"),
        kg_ae  =("kg_ae",   "sum"),
    )

    pivot_bult = pick_grp.pivot_table(
        index="cam", columns="cancha_norm",
        values="bult_pick", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_bult.columns.name = None

    pivot_up = pick_grp.pivot_table(
        index="cam", columns="cancha_norm",
        values="up_pick", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_up.columns.name = None
    up_rename = {c: f"_up_{c}" for c in _T4_CANCHAS + ["SIN CANCHA"] if c in pivot_up.columns}
    pivot_up = pivot_up.rename(columns=up_rename)

    pivot_hl = pick_grp.pivot_table(
        index="cam", columns="cancha_norm",
        values="hl_pick", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_hl.columns.name = None
    hl_rename = {c: f"_hl_{c}" for c in _T4_CANCHAS + ["SIN CANCHA"] if c in pivot_hl.columns}
    pivot_hl = pivot_hl.rename(columns=hl_rename)

    pivot_kg = pick_grp.pivot_table(
        index="cam", columns="cancha_norm",
        values="kg_pick", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_kg.columns.name = None
    kg_rename = {c: f"_kg_{c}" for c in _T4_CANCHAS + ["SIN CANCHA"] if c in pivot_kg.columns}
    pivot_kg = pivot_kg.rename(columns=kg_rename)

    # Pivot pall_ae por cancha (pallets AE completos por cancha) — v4.38
    pivot_pall_ae = grp.groupby(["cam", "cancha_norm"], as_index=False).agg(
        pall_ae_c=("pall_ae", "sum"),
    ).pivot_table(
        index="cam", columns="cancha_norm",
        values="pall_ae_c", aggfunc="sum", fill_value=0.0,
    ).reset_index()
    pivot_pall_ae.columns.name = None
    pall_ae_rename = {c: f"_pall_ae_{c}" for c in _T4_CANCHAS if c in pivot_pall_ae.columns}
    pivot_pall_ae = pivot_pall_ae.rename(columns=pall_ae_rename)

    cam_df = pivot_bult.merge(ae_grp, on="cam", how="left")
    cam_df = cam_df.merge(pivot_up, on="cam", how="left")
    cam_df = cam_df.merge(pivot_hl, on="cam", how="left")
    cam_df = cam_df.merge(pivot_kg, on="cam", how="left")
    cam_df = cam_df.merge(pivot_pall_ae, on="cam", how="left")

    for c in _T4_CANCHAS:
        if c not in cam_df.columns:
            cam_df[c] = 0.0
        if f"_up_{c}" not in cam_df.columns:
            cam_df[f"_up_{c}"] = 0.0
        if f"_hl_{c}" not in cam_df.columns:
            cam_df[f"_hl_{c}"] = 0.0
        if f"_kg_{c}" not in cam_df.columns:
            cam_df[f"_kg_{c}"] = 0.0
        if f"_pall_ae_{c}" not in cam_df.columns:
            cam_df[f"_pall_ae_{c}"] = 0.0
        # Fill NaN producto del merge
        cam_df[c]               = cam_df[c].fillna(0.0)
        cam_df[f"_up_{c}"]      = cam_df[f"_up_{c}"].fillna(0.0)
        cam_df[f"_hl_{c}"]      = cam_df[f"_hl_{c}"].fillna(0.0)
        cam_df[f"_kg_{c}"]      = cam_df[f"_kg_{c}"].fillna(0.0)
        cam_df[f"_pall_ae_{c}"] = cam_df[f"_pall_ae_{c}"].fillna(0.0)

    cam_df["bult_ae"] = cam_df["bult_ae"].fillna(0.0)
    cam_df["up_ae"]   = cam_df["up_ae"].fillna(0.0)
    cam_df["hl_ae"]   = cam_df["hl_ae"].fillna(0.0)
    cam_df["kg_ae"]   = cam_df["kg_ae"].fillna(0.0)
    # Compatibilidad legacy con código que esperaba "kg" simple
    cam_df["kg"]      = cam_df["kg_ae"] + sum(cam_df[f"_kg_{c}"] for c in _T4_CANCHAS)

    cam_df["TOTAL_PICK"] = cam_df[[c for c in _T4_CANCHAS if c in cam_df.columns]].sum(axis=1)
    cam_df["TOTAL_AE"]   = cam_df["bult_ae"]
    cam_df["TOTAL_BULT"] = cam_df["TOTAL_PICK"] + cam_df["TOTAL_AE"]
    cam_df["TOTAL_PALL"] = cam_df[[f"_up_{c}" for c in _T4_CANCHAS]].sum(axis=1) + cam_df["up_ae"]
    cam_df["AE_PALL"]    = cam_df["up_ae"]

    cam_df = cam_df.rename(columns={"cam": "Camión"})
    cam_df = cam_df.sort_values("Camión").reset_index(drop=True)

    tot_pick = float(cam_df["TOTAL_PICK"].sum())
    tot_ae   = float(cam_df["TOTAL_AE"].sum())
    mix = (tot_pick / (tot_pick + tot_ae)) if (tot_pick + tot_ae) > 0 else 0.0

    # Stats VE/CHESS para auditoría
    fuentes = body["fuente"].value_counts().to_dict() if "fuente" in body.columns else {}

    # v4.65 — Bultos AE por cancha (para Paso 2 fx Picking)
    # Se calcula desde grp ANTES del pivot, que es la única forma de obtenerlo
    # con la lógica floor() correcta aplicada por cam×SKU.
    bult_ae_grp = grp.groupby("cancha_norm", as_index=False).agg(bult_ae=("bult_ae", "sum"))
    bult_ae_por_cancha = {c: 0.0 for c in _T4_CANCHAS}
    for _, _r in bult_ae_grp.iterrows():
        if _r["cancha_norm"] in _T4_CANCHAS:
            bult_ae_por_cancha[_r["cancha_norm"]] = float(_r["bult_ae"])

    # v4.65 — DDM cancha lookup (para Paso 2: mapear SKU ANR → cancha)
    ddm_can_lookup = {}
    if not df_ddm.empty:
        for _sku_k, _sku_v in df_ddm.set_index("sku").to_dict("index").items():
            ddm_can_lookup[_sku_k] = _sku_v.get("can", "SIN CANCHA")

    return {
        "df":               cam_df,
        "fecha":            fecha,
        "fuente":           f"CAR + Frescura DDM · v4.33 · VE={fuentes.get('VE',0)} líneas · CHESS={fuentes.get('CHESS',0)} líneas",
        "mix_picking":      mix,
        "tot_pick":         tot_pick,
        "tot_ae":           tot_ae,
        "bult_ae_por_cancha": bult_ae_por_cancha,   # v4.65: AE bultos por cancha (pre-asign)
        "ddm_can_lookup":   ddm_can_lookup,           # v4.65: SKU→cancha para ANR lookup
    }


def _t4_calcular_pall_por_cancha(df: pd.DataFrame, asign_state: dict) -> dict:
    """
    Aplica ASIGN (reasignación entre canchas), calcula SUB, DESIGNADOS, TOTAL.
    asign_state = {(camion, cancha_origen): cancha_destino_str}
      - None / "" → sin asignación
      - "CANCHA X" → mueve los UP de esa línea a la cancha destino
    """
    sub        = {c: 0.0 for c in _T4_CANCHAS}
    designados = {c: 0.0 for c in _T4_CANCHAS}
    status_rows = {}
    asign_detail = {}  # {(cam, cancha_origen): cancha_destino}

    for _, row in df.iterrows():
        cam = int(row["Camión"])
        total_pall_cam = 0.0

        for c in _T4_CANCHAS:
            up_val    = float(row.get(f"_up_{c}", 0.0))
            asign_val = asign_state.get((cam, c))

            if asign_val and isinstance(asign_val, str) and asign_val.strip():
                dest = _t4_norm_cancha(asign_val)
                if dest in _T4_CANCHAS and dest != c:
                    designados[dest] += up_val
                    designados[c]    -= up_val
                    asign_detail[(cam, c)] = dest
            sub[c] += up_val
            total_pall_cam += up_val

        total_pall_cam += float(row.get("AE_PALL", 0.0))
        if total_pall_cam == 0:
            status = "CERO"
        elif total_pall_cam > _T4_CAP_MAX:
            status = f">{_T4_CAP_MAX} Pall"
        else:
            status = "OK"
        status_rows[cam] = {"status": status, "total_pall": total_pall_cam}

    return {
        "sub":          sub,
        "designados":   designados,
        "total":        {c: sub[c] + designados[c] for c in _T4_CANCHAS},
        "status_rows":  status_rows,
        "asign_detail": asign_detail,
    }


def _t4_hora_fin(bultos: float, vel: int, personas: int, inicio) -> object:
    """Calcula hora estimada de fin dado bultos, velocidad (bult/h) y personas."""
    import datetime as _dt
    inicio_dt = _dt.datetime.combine(_dt.date.today(), inicio)
    if vel <= 0 or personas <= 0 or bultos <= 0:
        return inicio_dt
    return inicio_dt + _dt.timedelta(hours=bultos / (vel * personas))


def _t4_calcular_metricas_por_cancha(df: pd.DataFrame, asign_state: dict) -> dict:
    """
    v4.35 — Calcula BULTOS, HL y KG reales de picking por cancha aplicando
    reasignaciones (ASIGN). La reasignación mueve el bloque cancha-camión
    completo (UP + bultos + HL + KG) a la cancha destino.

    Retorna:
        {
            "bultos": {cancha: float},   # bultos reales picking post-asign
            "hl":     {cancha: float},   # HL post-asign
            "kg":     {cancha: float},   # KG post-asign
        }
    """
    sub = {
        "bultos": {c: 0.0 for c in _T4_CANCHAS},
        "hl":     {c: 0.0 for c in _T4_CANCHAS},
        "kg":     {c: 0.0 for c in _T4_CANCHAS},
    }
    delta = {
        "bultos": {c: 0.0 for c in _T4_CANCHAS},
        "hl":     {c: 0.0 for c in _T4_CANCHAS},
        "kg":     {c: 0.0 for c in _T4_CANCHAS},
    }

    for _, row in df.iterrows():
        cam = int(row["Camión"])
        for c in _T4_CANCHAS:
            b_val  = float(row.get(c,           0.0))
            hl_val = float(row.get(f"_hl_{c}",  0.0))
            kg_val = float(row.get(f"_kg_{c}",  0.0))

            sub["bultos"][c] += b_val
            sub["hl"][c]     += hl_val
            sub["kg"][c]     += kg_val

            asign_val = asign_state.get((cam, c))
            if asign_val and isinstance(asign_val, str) and asign_val.strip():
                dest = _t4_norm_cancha(asign_val)
                if dest in _T4_CANCHAS and dest != c:
                    delta["bultos"][dest] += b_val
                    delta["bultos"][c]    -= b_val
                    delta["hl"][dest]     += hl_val
                    delta["hl"][c]        -= hl_val
                    delta["kg"][dest]     += kg_val
                    delta["kg"][c]        -= kg_val

    return {
        "bultos": {c: max(0.0, sub["bultos"][c] + delta["bultos"][c]) for c in _T4_CANCHAS},
        "hl":     {c: max(0.0, sub["hl"][c]     + delta["hl"][c])     for c in _T4_CANCHAS},
        "kg":     {c: max(0.0, sub["kg"][c]     + delta["kg"][c])     for c in _T4_CANCHAS},
    }


def _t4_calcular_bultos_por_cancha(df: pd.DataFrame, asign_state: dict) -> dict:
    """v4.35 — Wrapper de compatibilidad. Retorna {cancha: bultos_post_asign}."""
    return _t4_calcular_metricas_por_cancha(df, asign_state)["bultos"]


# ── HELPER: dibuja la página Resumen Controlador en un canvas existente ───────

def _draw_controlador_page(
    c, PW, PH, M, fin_global_dt,
    ctrl_params, mix_picking, fecha,
    DARK_BLUE, MED_BLUE, LIGHT, ALT, GREEN_OK, RED_WARN, WHITE, YELLOW_HL,
    FONT_B, FONT_N,
):
    """
    v4.41 — Dibuja la página Resumen Controlador sobre el canvas `c`.
    Llamado desde _t4_generar_pdf_x4 (última página del PDF ×5).
    También es usado por _t4_generar_pdf_controlador para el PDF standalone.
    """
    import datetime as _dt

    fin_por_cancha  = ctrl_params.get("fin_por_cancha", {})
    totales_bult    = ctrl_params.get("totales_bult", {})
    totales_pall_c  = ctrl_params.get("totales_pall_c", {})
    totales_hl      = ctrl_params.get("totales_hl", {})
    totales_kg      = ctrl_params.get("totales_kg", {})
    n_camiones      = ctrl_params.get("n_camiones", 0)
    tot_pick        = ctrl_params.get("tot_pick", 0)
    top_skus_lines  = ctrl_params.get("top_skus_lines", "")
    alertas         = ctrl_params.get("alertas", [])

    fecha_s = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)
    fin_g_s = fin_global_dt.strftime("%H:%M") if fin_global_dt else "—"

    usable = PW - 2 * M
    y = PH - M

    # ── HEADER ────────────────────────────────────────────────────────────────
    hdr_h = 40
    c.setFillColor(DARK_BLUE)
    c.rect(M, y - hdr_h, usable, hdr_h, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont(FONT_B, 16)
    c.drawString(M + 10, y - 22, f"RESUMEN CONTROLADOR — PICKING  {fecha_s}")
    c.setFont(FONT_N, 10)
    c.drawString(M + 10, y - 36, f"Beccacece Hnos SA  ·  Fin estimado global: {fin_g_s}  ·  Mix picking: {mix_picking*100:.1f}%")
    y -= hdr_h + 8

    # ── TABLA HORARIOS POR CANCHA ─────────────────────────────────────────────
    c.setFont(FONT_B, 11)
    c.setFillColor(DARK_BLUE)
    # v4.41: sin emoji (incompatible con ReportLab)
    c.drawString(M, y - 2, "Horarios de finalizacion por cancha")
    y -= 16

    cn_labels = [_cn_short(cn) for cn in _T4_CANCHAS]  # v4.64: números naturales
    # v4.41: primera columna del bloque horarios más ancha para evitar superposición
    # La tabla de horarios no tiene col-label izquierdo — usamos todo el ancho
    col_w_cn  = usable / len(_T4_CANCHAS)

    # Header cancha
    c.setFillColor(MED_BLUE)
    c.rect(M, y - 18, usable, 18, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont(FONT_B, 10)
    for i, lbl in enumerate(cn_labels):
        c.drawCentredString(M + col_w_cn * i + col_w_cn / 2, y - 13, lbl)
    y -= 18

    # Fila INICIO — v4.41: columna etiqueta ampliada, no choca con valores
    lbl_w = 55   # ancho de la celda etiqueta en la tabla de filas
    data_w = (usable - lbl_w) / len(_T4_CANCHAS)   # ancho por columna de datos

    def _draw_fila_cancha(y_pos, row_h, label, font_label, font_val, fsize_val,
                          bg_color, get_val_fn, hl_focus=False):
        c.setFillColor(bg_color)
        c.rect(M, y_pos - row_h, usable, row_h, fill=1, stroke=0)
        c.setFillColor(DARK_BLUE)
        c.setFont(font_label, 8)
        c.drawString(M + 3, y_pos - row_h + (row_h - 8) // 2 + 2, label)
        for i2, cn in enumerate(_T4_CANCHAS):
            val = get_val_fn(cn)
            is_f = hl_focus  # para HORA FIN resaltar toda la fila
            x_val = M + lbl_w + data_w * i2
            if is_f:
                c.setFillColor(YELLOW_HL)
                c.rect(x_val, y_pos - row_h, data_w, row_h, fill=1, stroke=0)
                c.setFillColor(DARK_BLUE)
            c.setFont(font_val, fsize_val)
            c.drawCentredString(x_val + data_w / 2, y_pos - row_h + (row_h - fsize_val) // 2 + 1, val)
        return y_pos - row_h

    y = _draw_fila_cancha(
        y, 16, "INICIO", FONT_B, FONT_N, 9, LIGHT,
        lambda cn: fin_por_cancha[cn]["inicio"].strftime("%H:%M") if fin_por_cancha[cn]["bultos"] > 0 else "—",
    )

    # Fila HORA FIN — resaltada
    y = _draw_fila_cancha(
        y, 18, "HORA FIN EST.", FONT_B, FONT_B, 11, YELLOW_HL,
        lambda cn: fin_por_cancha[cn]["fin_dt"].strftime("%H:%M") if fin_por_cancha[cn]["bultos"] > 0 else "—",
        hl_focus=True,
    )

    y = _draw_fila_cancha(
        y, 14, "BULTOS", FONT_B, FONT_N, 9, ALT,
        lambda cn: f"{totales_bult.get(cn, 0.0):.0f}" if totales_bult.get(cn, 0.0) > 0 else "—",
    )

    y = _draw_fila_cancha(
        y, 14, "PALLETS UP", FONT_B, FONT_N, 9, LIGHT,
        lambda cn: f"{totales_pall_c.get(cn, 0.0):.2f}" if totales_pall_c.get(cn, 0.0) > 0 else "—",
    )

    # Borde exterior tabla canchas
    tabla_cn_h = 18 + 16 + 18 + 14 + 14
    c.setStrokeColor(DARK_BLUE)
    c.setLineWidth(0.8)
    c.rect(M, y, usable, tabla_cn_h, fill=0, stroke=1)
    y -= 12

    # ── TOTALES OPERATIVOS + ALERTAS (dos columnas) ────────────────────────────
    # v4.41: primera columna más ancha (55%) para evitar superposición de texto
    col_left_w  = usable * 0.55
    col_right_w = usable * 0.41
    col_gap     = usable * 0.04

    y_sec = y

    # columna izquierda: totales
    c.setFont(FONT_B, 11)
    c.setFillColor(DARK_BLUE)
    c.drawString(M, y_sec - 2, "Totales operativos")
    y_sec -= 14

    tot_rows = [
        ("Camiones con picking",  f"{n_camiones}"),
        ("Bultos totales picking", f"{tot_pick:.0f}"),
        ("Bultos MKPL",           f"{totales_bult.get('MKPL', 0.0):.0f}"),
        ("HL total picking",      f"{sum(totales_hl.values()):.1f}"),
        ("KG total picking",      f"{sum(totales_kg.values()):.0f}"),
        ("Mix picking",           f"{mix_picking*100:.1f}%"),
    ]
    for ti, (label, valor) in enumerate(tot_rows):
        bg_r = ALT if ti % 2 == 0 else LIGHT
        c.setFillColor(bg_r)
        c.rect(M, y_sec - 13, col_left_w, 13, fill=1, stroke=0)
        c.setFillColor(DARK_BLUE)
        c.setFont(FONT_N, 9)
        c.drawString(M + 4, y_sec - 9, label)
        c.setFont(FONT_B, 9)
        c.drawRightString(M + col_left_w - 4, y_sec - 9, valor)
        y_sec -= 13

    # columna derecha: alertas y top SKUs
    y_r = y
    c.setFont(FONT_B, 11)
    c.setFillColor(DARK_BLUE)
    c.drawString(M + col_left_w + col_gap, y_r - 2, "Alertas de cancha")
    y_r -= 14

    if alertas:
        for alerta in alertas:
            c.setFillColor(colors.HexColor("#FFEBEE"))
            c.rect(M + col_left_w + col_gap, y_r - 13, col_right_w, 13, fill=1, stroke=0)
            c.setFillColor(RED_WARN)
            c.setFont(FONT_N, 9)
            alerta_clean = alerta.replace("🔴 ", "").replace("⚠️ ", "")
            c.drawString(M + col_left_w + col_gap + 4, y_r - 9, alerta_clean)
            y_r -= 13
    else:
        c.setFillColor(colors.HexColor("#E8F5E9"))
        c.rect(M + col_left_w + col_gap, y_r - 13, col_right_w, 13, fill=1, stroke=0)
        c.setFillColor(GREEN_OK)
        c.setFont(FONT_N, 9)
        c.drawString(M + col_left_w + col_gap + 4, y_r - 9, "Sin alertas — todas las canchas en tiempo")
        y_r -= 13

    y_r -= 8
    c.setFont(FONT_B, 11)
    c.setFillColor(DARK_BLUE)
    c.drawString(M + col_left_w + col_gap, y_r - 2, "Top 3 SKUs del dia")
    y_r -= 14

    if top_skus_lines:
        lines = [l.strip().lstrip("•").strip() for l in top_skus_lines.split("\n") if l.strip()]
        for ti2, line in enumerate(lines[:3]):
            bg_r2 = ALT if ti2 % 2 == 0 else LIGHT
            line_clean = line.replace("**", "")
            c.setFillColor(bg_r2)
            c.rect(M + col_left_w + col_gap, y_r - 14, col_right_w, 14, fill=1, stroke=0)
            c.setFillColor(DARK_BLUE)
            c.setFont(FONT_N, 8)
            c.drawString(M + col_left_w + col_gap + 4, y_r - 10, line_clean[:60])
            y_r -= 14
    else:
        c.setFont(FONT_N, 9)
        c.setFillColor(DARK_BLUE)
        c.drawString(M + col_left_w + col_gap + 4, y_r - 9, "(no disponible)")

    # ── SECCIÓN RECOMENDACIONES CONTROLADOR (v4.64) ───────────────────────────
    # Basado en ALM-00013 SOP Verificación de Cargas T2 v5
    y_rec = min(y_sec, y_r) - 14
    if y_rec > M + 100:   # solo si queda espacio suficiente
        c.setFont(FONT_B, 10)
        c.setFillColor(DARK_BLUE)
        c.drawString(M, y_rec - 2, "Recomendaciones operativas para Controlador de Deposito")
        y_rec -= 12

        # Línea separadora
        c.setStrokeColor(MED_BLUE)
        c.setLineWidth(0.6)
        c.line(M, y_rec, M + usable, y_rec)
        y_rec -= 6

        _rec_items = [
            # (icono_texto, descripcion)
            ("DOBLE CTRL",  "1° ctrl: SKU + bultos vs planilla. 2° ctrl (Seguridad): bultos picking + completas."),
            ("EPP",         "Casco + chaleco reflectivo + zapatos de seguridad + guantes. TODOS en darsena."),
            ("POSICION",    "Ubicarse siempre LATERAL a la paleta. Nunca frente al AE. Unas en posicion baja."),
            ("FILMACION",   "Filmar TODAS las paletas con tablet. Material retenido hasta retorno T2."),
            ("FALTANTES",   "Faltante: detener carga hasta resolver. Sin stock: informar ventas con firma supervisor."),
            ("DAÑOS",       "Producto roto o con riesgo de derrame: segregar de inmediato. NUNCA cargar."),
            ("SISTEMA",     "Verificar: patente habilitada + pallets <= limite + peso vs tara ANTES de cargar."),
            ("ERRORES",     "Registrar todo error en herramienta Errores Operativos. Recurrente: 5 Porques / Ishikawa."),
            ("REARMADO",    "Error detectado: paleta a cancha, rearmar, re-presentar al equipo de control."),
            ("CHECKLIST",   "Completar UN checklist por camion. Archivar en carpeta del turno o AppSheet."),
        ]
        # Dos columnas de recomendaciones
        rec_col_w = (usable - 8) / 2
        rec_row_h = 11
        for ri, (tag, desc) in enumerate(_rec_items):
            col_offset = rec_col_w + 8 if ri % 2 == 1 else 0
            row_y = y_rec - (ri // 2) * (rec_row_h + 1)
            if row_y < M + 20:
                break
            bg_rec = ALT if (ri // 2) % 2 == 0 else LIGHT
            c.setFillColor(bg_rec)
            c.rect(M + col_offset, row_y - rec_row_h, rec_col_w, rec_row_h, fill=1, stroke=0)
            # Tag en azul fuerte
            c.setFillColor(MED_BLUE)
            c.setFont(FONT_B, 7)
            c.drawString(M + col_offset + 3, row_y - rec_row_h + 3, tag)
            # Descripcion
            c.setFillColor(DARK_BLUE)
            c.setFont(FONT_N, 7)
            # Truncar si no entra
            _tag_w = c.stringWidth(tag + "  ", FONT_B, 7)
            c.drawString(M + col_offset + 3 + _tag_w, row_y - rec_row_h + 3, desc[:75])

    # ── FOOTER ────────────────────────────────────────────────────────────────
    c.setFillColor(DARK_BLUE)
    c.rect(M, M, usable, 12, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont(FONT_N, 7)
    from datetime import datetime as _dtm
    c.drawString(M + 4, M + 3, f"Picking Orchestrator v{APP_VERSION}  ·  {_dtm.now().strftime('%d/%m/%Y %H:%M')}")
    c.drawRightString(M + usable - 4, M + 3, "Beccacece Hnos SA — DPO 2.1 Pilar Almacen")

    c.showPage()


def _t4_generar_pdf_x4(
    df, totales_bult, totales_pall, totales_hl, totales_kg,
    productividad, personas, inicio, fecha, mix_picking, fin_calc,
    asign_detail=None,
    # v4.41: parámetros opcionales para incluir hoja Controlador
    ctrl_params=None,
):
    """
    Genera PDF A4 LANDSCAPE con una página por cancha (CI/CII/CIII/CIV/MKPL).
    v4.41: modo horizontal — mejor uso del espacio, columnas balanceadas, STS proporcional.
    Si se provee ctrl_params, agrega una página extra con el Resumen del Controlador.
    Muestra:
      - Bultos UP (fracción de paleta) a pickear por cancha y camión
      - Paletas puras (AE) en rojo
      - Total paletas por camión
      - Info de asignaciones (recibe / envía carga)
    """
    import datetime as _dt
    import math as _math

    if asign_detail is None:
        asign_detail = {}

    buf    = io.BytesIO()
    # ── v4.37: PDF LANDSCAPE (A4 horizontal) ──────────────────────────────────
    LS = landscape(A4)
    c_pdf  = rl_canvas.Canvas(buf, pagesize=LS)
    pw, ph = LS          # pw ≈ 841 pt, ph ≈ 595 pt
    M      = 10 * mm

    # Colores
    HDR_BG_PDF   = colors.HexColor("#1a3a6b")
    SUB_BG_PDF   = colors.HexColor("#2e5fa3")
    TOT_BG_PDF   = colors.HexColor("#FF8C00")
    OK_BG        = colors.HexColor("#C8E6C9")
    OVER_BG      = colors.HexColor("#FFCDD2")
    ZERO_BG      = colors.HexColor("#F5F5F5")
    ALT_ROW      = colors.HexColor("#F0F4FF")
    WHITE        = colors.white
    GREEN_PICK   = colors.HexColor("#66BB6A")   # verde = bultos a pickear
    RED_AE       = colors.HexColor("#EF9A9A")   # rojo  = paletas puras AE
    YELLOW_FOCUS = colors.HexColor("#FFE57F")   # v4.37: amarillo más vibrante para cancha foco
    ASIGN_BG     = colors.HexColor("#E3F2FD")   # azul claro = carga recibida (leyenda)
    # v4.41: colores diferenciados para reasignaciones
    ASIGN_RECIBE = colors.HexColor("#B2DFDB")   # verde agua = esta cancha RECIBE carga
    ASIGN_ENVIA  = colors.HexColor("#FFE0B2")   # naranja claro = esta cancha ENVÍA carga
    ASIGN_RECIBE_TEXT = colors.HexColor("#00695C")   # verde oscuro para texto "+CX"
    ASIGN_ENVIA_TEXT  = colors.HexColor("#BF360C")   # naranja oscuro para texto "→CX"

    FONT_B = "Helvetica-Bold"
    FONT_N = "Helvetica"

    # Columnas tabla: CAM | CI | CII | CIII | CIV | MKPL | AE | TOT | STS
    # v4.39: widths ajustados — sin espacio vacío sobrante; fuente más grande
    col_labels = ["CAM"] + [_cn_short(c) for c in _T4_CANCHAS] + ["AE", "TOTAL", "STS"]  # v4.64: números naturales
    inner_w    = pw - 2 * M
    # CAM=38, 5 canchas proporcionales, AE=42, TOTAL=50, STS=38
    _cw_fixed = [38, 46, 46, 46, 46, 42, 42, 50, 38]
    _used = sum(_cw_fixed)
    _slack = max(0, inner_w - _used)
    # Distribuir el slack entre las 5 canchas (índices 1-5) equitativamente
    _bonus = int(_slack / 5)
    cw = [_cw_fixed[0]] + [_cw_fixed[i] + _bonus for i in range(1, 6)] + _cw_fixed[6:]
    # Si aun hay residuo, sumarlo a TOTAL (idx 7)
    _residuo = inner_w - sum(cw)
    cw[7] = max(_cw_fixed[7], cw[7] + int(_residuo))

    fin_global = fin_calc["fin_global_dt"]

    # Calcular status_rows una sola vez
    asign_st = st.session_state.get("t4_asign", {})
    calc      = _t4_calcular_pall_por_cancha(df, asign_st)
    status_rows = calc["status_rows"]

    def _draw_page(cancha_focus: str):
        # Filtrar camiones con picking > 0 en cualquier cancha
        rows_data = []
        for _, row in df.iterrows():
            if float(row.get("TOTAL_PICK", 0)) <= 0:
                continue
            cam = int(row["Camión"])
            sr  = status_rows.get(cam, {})
            palls = {c: float(row.get(f"_up_{c}", 0.0)) for c in _T4_CANCHAS}

            # Detectar asignaciones para este camión
            recibe_de = []
            envia_a   = []
            for (c_cam, c_origen), c_dest in asign_detail.items():
                if c_cam == cam:
                    if c_origen == cancha_focus:
                        envia_a.append(c_dest.replace("CANCHA ", "C"))
                    if c_dest == cancha_focus:
                        recibe_de.append(c_origen.replace("CANCHA ", "C"))

            rows_data.append({
                "cam":       cam,
                "palls":     palls,
                "ae":        float(row.get("AE_PALL", 0.0)),
                "total":     float(sr.get("total_pall", 0.0)),
                "status":    sr.get("status", "—"),
                "recibe_de": recibe_de,
                "envia_a":   envia_a,
                # bultos de picking en la cancha foco
                "bult_foco": float(row.get(cancha_focus, 0.0)),
            })

        y = ph - M

        # ── Header principal ───────────────────────────────────────────────
        hdr_h = 32
        c_pdf.setFillColor(HDR_BG_PDF)
        c_pdf.rect(M, y - hdr_h, inner_w, hdr_h, fill=1, stroke=0)
        c_pdf.setFillColor(WHITE)
        fecha_s  = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)
        fin_s    = fin_global.strftime("%H:%M")

        # Línea 1: PICKING + fecha + FIN GLOBAL + MIX
        c_pdf.setFont(FONT_B, 11)
        line1 = (
            f"PICKING   {fecha_s}   "
            f"FIN GLOBAL: {fin_s}   "
            f"MIX PICKING: {mix_picking*100:.1f}%"
        )
        c_pdf.drawCentredString(M + inner_w / 2, y - 12, line1)

        # Línea 2: INI / FIN por cancha
        c_pdf.setFont(FONT_N, 8)
        partes = []
        por_cn = fin_calc.get("por_cancha", {})
        for cn in _T4_CANCHAS:
            short = cn.replace("CANCHA ", "C")
            fp = por_cn.get(cn, {})
            ini_v = fp.get("inicio")
            fin_v = fp.get("fin_dt")
            bult_v = fp.get("bultos", 0)
            ini_str = ini_v.strftime("%H:%M") if ini_v else "—"
            fin_str = fin_v.strftime("%H:%M") if (fin_v and bult_v > 0) else "—"
            partes.append(f"{short}: {ini_str}→{fin_str}")
        c_pdf.drawCentredString(M + inner_w / 2, y - 24, "   |   ".join(partes))
        y -= hdr_h + 2

        # ── Subheader cancha ───────────────────────────────────────────────
        sub_h = 15
        c_pdf.setFillColor(SUB_BG_PDF)
        c_pdf.rect(M, y - sub_h, inner_w, sub_h, fill=1, stroke=0)
        c_pdf.setFillColor(WHITE)
        c_pdf.setFont(FONT_B, 11)
        c_pdf.drawCentredString(
            M + inner_w / 2, y - sub_h + 4,
            f"COPIA PARA: {cancha_focus}",
        )
        y -= sub_h + 3

        # ── Header tabla ───────────────────────────────────────────────────
        th = 15
        c_pdf.setFillColor(colors.HexColor("#E8EAF6"))
        c_pdf.rect(M, y - th, inner_w, th, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.setFont(FONT_B, 8)
        x = M
        for i, lbl in enumerate(col_labels):
            c_pdf.drawCentredString(x + cw[i] / 2, y - th + 5, lbl)
            x += cw[i]
        y -= th

        # ── Filas camiones ─────────────────────────────────────────────────
        rh = 15
        for ri, rd in enumerate(rows_data):
            bg = ALT_ROW if ri % 2 == 0 else WHITE
            c_pdf.setFillColor(bg)
            c_pdf.rect(M, y - rh, inner_w, rh, fill=1, stroke=0)

            st_val = rd["status"]
            st_bg  = OK_BG if st_val == "OK" else (OVER_BG if st_val.startswith(">") else ZERO_BG)

            x = M
            vals = [str(rd["cam"])] + [
                f"{rd['palls'][c]:.2f}" if rd['palls'][c] > 0 else "—"
                for c in _T4_CANCHAS
            ] + [
                f"{rd['ae']:.2f}" if rd['ae'] > 0 else "—",
                f"{rd['total']:.2f}",
                st_val,
            ]

            for i, v in enumerate(vals):
                col_name = _T4_CANCHAS[i-1] if 1 <= i <= len(_T4_CANCHAS) else None
                is_focus  = (col_name == cancha_focus)
                is_ae     = (i == len(_T4_CANCHAS) + 1)   # columna AE
                is_status = (i == len(vals) - 1)

                # Color de celda
                if is_focus:
                    # v4.41: color diferenciado según si la cancha ENVÍA o RECIBE
                    if rd["envia_a"] and rd["recibe_de"]:
                        cell_bg = ASIGN_RECIBE  # mixto: priorizar recibe (verde agua)
                    elif rd["envia_a"]:
                        cell_bg = ASIGN_ENVIA   # naranja claro
                    elif rd["recibe_de"]:
                        cell_bg = ASIGN_RECIBE  # verde agua
                    else:
                        cell_bg = YELLOW_FOCUS
                    c_pdf.setFillColor(cell_bg)
                    c_pdf.rect(x, y - rh, cw[i], rh, fill=1, stroke=0)
                if is_ae and rd["ae"] > 0:
                    c_pdf.setFillColor(RED_AE)
                    c_pdf.rect(x, y - rh, cw[i], rh, fill=1, stroke=0)
                if is_status:
                    c_pdf.setFillColor(st_bg)
                    c_pdf.rect(x, y - rh, cw[i], rh, fill=1, stroke=0)

                c_pdf.setFillColor(colors.black)
                fs = 9 if i == 0 else 8
                c_pdf.setFont(FONT_B if i == 0 else FONT_N, fs)

                # v4.41: nota de asignación inline en la celda foco — más visible
                if is_focus and (rd["recibe_de"] or rd["envia_a"]):
                    # Líneas separadas para RECIBE (+CX) y ENVÍA (→CX), fuente grande bold
                    nota_parts = []
                    if rd["recibe_de"]:
                        nota_parts.append((f"+{','.join(rd['recibe_de'])}", ASIGN_RECIBE_TEXT))
                    if rd["envia_a"]:
                        nota_parts.append((f"\u2192{','.join(rd['envia_a'])}", ASIGN_ENVIA_TEXT))

                    if len(nota_parts) == 1:
                        # Solo un tipo: valor arriba, nota abajo
                        c_pdf.setFillColor(colors.black)
                        c_pdf.setFont(FONT_B, 8)
                        c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 9, v)
                        c_pdf.setFont(FONT_B, 8)
                        c_pdf.setFillColor(nota_parts[0][1])
                        c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 2, nota_parts[0][0])
                        c_pdf.setFillColor(colors.black)
                    else:
                        # Dos tipos: valor pequeño + dos notas
                        c_pdf.setFillColor(colors.black)
                        c_pdf.setFont(FONT_N, 7)
                        c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 12, v)
                        c_pdf.setFont(FONT_B, 7)
                        c_pdf.setFillColor(nota_parts[0][1])
                        c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 6, nota_parts[0][0])
                        c_pdf.setFillColor(nota_parts[1][1])
                        c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 1, nota_parts[1][0])
                        c_pdf.setFillColor(colors.black)
                else:
                    c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 5, v)
                x += cw[i]

            c_pdf.setStrokeColor(colors.HexColor("#DDDDDD"))
            c_pdf.line(M, y - rh, M + inner_w, y - rh)
            y -= rh

        # ── Fila TOTAL PALL x CANCHA ───────────────────────────────────────
        y -= 2
        c_pdf.setFillColor(TOT_BG_PDF)
        c_pdf.rect(M, y - rh, inner_w, rh, fill=1, stroke=0)
        c_pdf.setFillColor(WHITE)
        c_pdf.setFont(FONT_B, 8)
        x = M
        tot_vals = ["TOTAL PALL"] + [
            f"{totales_pall.get(c, 0):.2f}" for c in _T4_CANCHAS
        ] + ["", f"{sum(totales_pall.values()):.2f}", ""]
        for i, v in enumerate(tot_vals):
            c_pdf.drawCentredString(x + cw[i] / 2, y - rh + 5, v)
            x += cw[i]
        y -= rh + 3

        # ── Sección métricas bultos/HL/KG ─────────────────────────────────
        mh = 12
        met_rows = [
            ("TOTAL BULTOS UP (Pick)", totales_bult, ":.1f"),
            ("TOTAL HL",               totales_hl,   ":.2f"),
            ("TOTAL KG",               totales_kg,   ":.0f"),
        ]
        for label, vals_dict, fmt in met_rows:
            c_pdf.setFillColor(colors.HexColor("#ECEFF1"))
            c_pdf.rect(M, y - mh, inner_w, mh, fill=1, stroke=0)
            c_pdf.setFillColor(colors.black)
            c_pdf.setFont(FONT_B, 7)
            c_pdf.drawString(M + 2, y - mh + 3, label)
            x = M + cw[0]
            for ci, cn in enumerate(_T4_CANCHAS):
                v = vals_dict.get(cn, 0.0)
                c_pdf.setFont(FONT_N, 7)
                c_pdf.drawCentredString(x + cw[ci+1]/2, y - mh + 3, f"{v:{fmt[1:]}}")
                x += cw[ci+1]
            y -= mh + 1

        # ── Pie rows data ──────────────────────────────────────────────────
        pie_rows = [
            ("# PERSONAS",    {c: str(personas.get(c, 1)) for c in _T4_CANCHAS}),
            ("HORA FIN EST.", {
                c: fin_calc["por_cancha"][c]["fin_dt"].strftime("%H:%M")
                   if fin_calc["por_cancha"][c]["bultos"] > 0 else "—"
                for c in _T4_CANCHAS
            }),
        ]

        # ── Leyenda de colores ─────────────────────────────────────────────
        y -= 3
        ley_h = 9
        c_pdf.setFillColor(GREEN_PICK)
        c_pdf.rect(M, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.setFont(FONT_N, 7)
        c_pdf.drawString(M + 10, y - ley_h + 2, "= Bultos a pickear (cancha foco)")

        c_pdf.setFillColor(RED_AE)
        c_pdf.rect(M + 120, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(colors.black)
        c_pdf.drawString(M + 132, y - ley_h + 2, "= Paletas puras (AE)")

        # v4.41: leyenda diferenciada para reasignaciones
        c_pdf.setFillColor(ASIGN_RECIBE)
        c_pdf.rect(M + 215, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(ASIGN_RECIBE_TEXT)
        c_pdf.setFont(FONT_B, 7)
        c_pdf.drawString(M + 227, y - ley_h + 2, "+CX = Recibe carga de CX")
        c_pdf.setFillColor(ASIGN_ENVIA)
        c_pdf.rect(M + 370, y - ley_h, 8, ley_h, fill=1, stroke=0)
        c_pdf.setFillColor(ASIGN_ENVIA_TEXT)
        c_pdf.drawString(M + 382, y - ley_h + 2, "\u2192CX = Envia carga a CX")
        c_pdf.setFillColor(colors.black)
        c_pdf.setFont(FONT_N, 7)
        y -= ley_h + 3

        # ── Pie: personas y hora fin ───────────────────────────────────────
        for label, vals_dict in pie_rows:
            c_pdf.setFillColor(colors.HexColor("#1a3a6b"))
            c_pdf.rect(M, y - mh, inner_w, mh, fill=1, stroke=0)
            c_pdf.setFillColor(WHITE)
            c_pdf.setFont(FONT_B, 7)
            c_pdf.drawString(M + 2, y - mh + 3, label)
            x = M + cw[0]
            for ci, cn in enumerate(_T4_CANCHAS):
                is_this_focus = (cn == cancha_focus)
                if is_this_focus and label == "HORA FIN EST.":
                    c_pdf.setFillColor(colors.HexColor("#FFF176"))
                    c_pdf.rect(x, y - mh, cw[ci+1], mh, fill=1, stroke=0)
                    c_pdf.setFillColor(colors.HexColor("#1a3a6b"))
                else:
                    c_pdf.setFillColor(WHITE)
                c_pdf.setFont(FONT_B if is_this_focus else FONT_N, 7)
                c_pdf.drawCentredString(x + cw[ci+1]/2, y - mh + 3, vals_dict.get(cn, "—"))
                x += cw[ci+1]
            y -= mh + 1

        c_pdf.showPage()

    for cancha in _T4_CANCHAS_PDF:
        _draw_page(cancha)

    # v4.41: agregar página Resumen Controlador si se proveen parámetros
    if ctrl_params:
        _draw_controlador_page(
            c_pdf, pw, ph, M, fin_global,
            ctrl_params, mix_picking, fecha,
            DARK_BLUE=HDR_BG_PDF,
            MED_BLUE=SUB_BG_PDF,
            LIGHT=colors.HexColor("#F0F4FA"),
            ALT=colors.HexColor("#E8EEF8"),
            GREEN_OK=colors.HexColor("#1a7a4a"),
            RED_WARN=colors.HexColor("#b91c1c"),
            WHITE=WHITE,
            YELLOW_HL=colors.HexColor("#FFF176"),
            FONT_B=FONT_B,
            FONT_N=FONT_N,
        )

    c_pdf.save()
    buf.seek(0)
    return buf.read()


def _push_matriz_pall_to_sheets(df_display, pdata, credentials_json: dict) -> int:
    """
    Inserta una fila por camión en la hoja 'Matriz Pall' del Sheets maestro.
    Detecta la última fila con datos y escribe a continuación (nunca sobreescribe).

    Parámetros
    ----------
    df_display       : DataFrame ya calculado en render_tab_proyeccion()
    pdata            : dict con al menos pdata['fecha'] (date o str)
    credentials_json : dict con el service account JSON (viene de st.secrets)

    Retorna
    -------
    int : cantidad de filas insertadas
    """
    import gspread
    from google.oauth2.service_account import Credentials

    SPREADSHEET_ID = "1QCoMtpHcUaTITp9p9-1mo914HsaH0siq_bHL7nmJ_DA"
    SHEET_NAME     = "Matriz Pall."
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds  = Credentials.from_service_account_info(credentials_json, scopes=SCOPES)
    client = gspread.authorize(creds)
    sheet  = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

    # Orden de canchas idéntico al XLSX exportable
    CANCHAS = _T4_CANCHAS  # ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "MKPL"]

    fecha_str = (
        pdata["fecha"].strftime("%d/%m/%Y")
        if hasattr(pdata["fecha"], "strftime")
        else str(pdata["fecha"])
    )

    rows_to_append = []
    for _, row in df_display.iterrows():
        camion = row.get("Camión", row.get("Camion", ""))
        try:
            camion = int(camion)
        except (ValueError, TypeError):
            pass

        import math as _math_push

        # ── UP picking por cancha (fracción de pallet) ──
        up_ci   = round(float(row.get("_up_CANCHA I",   0.0)), 3)
        up_cii  = round(float(row.get("_up_CANCHA II",  0.0)), 3)
        up_ciii = round(float(row.get("_up_CANCHA III", 0.0)), 3)
        up_civ  = round(float(row.get("_up_CANCHA IV",  0.0)), 3)
        up_mkpl = round(float(row.get("_up_MKPL",       0.0)), 3)
        pick_matr_up = round(up_ci + up_cii + up_ciii + up_civ + up_mkpl, 3)

        # ── Pall completas AE por cancha (del pipeline upstream) ──
        pall_ci   = int(row.get("_pall_ae_CANCHA I",   0))
        pall_cii  = int(row.get("_pall_ae_CANCHA II",  0))
        pall_ciii = int(row.get("_pall_ae_CANCHA III", 0))
        pall_civ  = int(row.get("_pall_ae_CANCHA IV",  0))
        pall_mkpl = int(row.get("_pall_ae_MKPL",       0))
        # Col K: pall AE total (suma de todas las canchas activas)
        ae_pall      = round(float(row.get("AE_PALL", 0.0)), 2)
        pall_matr_up = pall_ci + pall_cii + pall_ciii + pall_civ + pall_mkpl

        # ── KG por cancha ──
        kg_ci   = round(float(row.get("_kg_CANCHA I",   0.0)), 1)
        kg_cii  = round(float(row.get("_kg_CANCHA II",  0.0)), 1)
        kg_ciii = round(float(row.get("_kg_CANCHA III", 0.0)), 1)
        kg_civ  = round(float(row.get("_kg_CANCHA IV",  0.0)), 1)
        kg_mkpl = round(float(row.get("_kg_MKPL",       0.0)), 1)
        kg_matr = round(kg_ci + kg_cii + kg_ciii + kg_civ + kg_mkpl, 1)

        # ── Armado de fila según estructura del Sheet ──
        # A  Fecha
        # B  Camión
        # C  REPA RTO → siempre "SI"
        # D  PICK MATR IZ UP  (suma UP picking todas las canchas activas)
        # E  CANCHA I   UP picking
        # F  CANCHA II  UP picking
        # G  CANCHA III UP picking
        # H  CANCHA IV  UP picking
        # I  CANCHA V   (inactiva) → 0
        # J  MKPL       UP picking
        # K  PALL MATR IZ UP  (pall AE total = suma pall_ae por cancha)
        # L  CANCHA I   pall AE completas (del pipeline _pall_ae_)
        # M  CANCHA II  pall AE completas
        # N  CANCHA III pall AE completas
        # O  CANCHA IV  pall AE completas
        # P  CANCHA V   (inactiva) → 0
        # Q  MKPL       pall AE completas
        # R  KG MATR IZ total (suma KG todas las canchas activas)
        # S  CANCHA I   KG
        # T  CANCHA II  KG
        # U  CANCHA III KG
        # V  CANCHA IV  KG
        # W  CANCHA V   (inactiva) → 0
        # X  MKPL       KG
        # (nada después de X)
        new_row = [
            fecha_str,    # A
            camion,       # B
            "SI",         # C — REPA RTO siempre SI
            pick_matr_up, # D — PICK MATR IZ UP
            up_ci,        # E — CANCHA I   UP picking
            up_cii,       # F — CANCHA II  UP picking
            up_ciii,      # G — CANCHA III UP picking
            up_civ,       # H — CANCHA IV  UP picking
            0,            # I — CANCHA V   (inactiva)
            up_mkpl,      # J — MKPL       UP picking
            pall_matr_up, # K — PALL MATR IZ UP (AE total)
            pall_ci,      # L — CANCHA I   pall completas
            pall_cii,     # M — CANCHA II  pall completas
            pall_ciii,    # N — CANCHA III pall completas
            pall_civ,     # O — CANCHA IV  pall completas
            0,            # P — CANCHA V   (inactiva)
            pall_mkpl,    # Q — MKPL       pall completas
            kg_matr,      # R — KG MATR IZ total
            kg_ci,        # S — CANCHA I   KG
            kg_cii,       # T — CANCHA II  KG
            kg_ciii,      # U — CANCHA III KG
            kg_civ,       # V — CANCHA IV  KG
            0,            # W — CANCHA V   (inactiva)
            kg_mkpl,      # X — MKPL       KG
        ]
        rows_to_append.append(new_row)

    if not rows_to_append:
        return 0

    # append_rows escribe al final del rango detectado por Sheets
    # sin insertar ni desplazar filas → fórmulas en AT+ quedan intactas
    sheet.append_rows(
        rows_to_append,
        value_input_option="USER_ENTERED",
        insert_data_option="INSERT_ROWS",
        table_range="A1",
    )
    return len(rows_to_append)


def render_tab_proyeccion():
    """
    Tab 4 — Proyección Picking ×5 (v4.56.0).

    CAMBIOS v4.56.0 — FIX HORARIOS CONGELADOS TRAS REASIGNACIÓN:
      - 🐛 ROOT CAUSE: _rebuild_live_asign_from_editor_state() usaba
        st.session_state["t4_main_editor"]["edited_rows"], que son DELTAS
        acumulativos desde la primera edición, no el estado completo del widget.
        Al borrar una asignación (→ ""), el delta decía "" pero la key seguía
        en t4_asign persistido → el header PRE-tabla mostraba la hora vieja.
      - ✅ FIX 1: Eliminado _rebuild_live_asign_from_editor_state. El header
        PRE-tabla ahora lee t4_asign directamente (estado del ciclo anterior,
        que es siempre correcto porque fue escrito desde edited_df completo).
      - ✅ FIX 2: live_asign se construye iterando edited_df completo (todas
        las filas, no solo edited_rows). Si ASIGN está vacío → la key NO se
        agrega (borrado limpio). Garantiza que quitar una asignación la elimine.
      - ✅ FIX 3: Botón "🔄 Recalcular horarios" que hace st.rerun() forzado.
        Flujo: editar ASIGN → presionar Recalcular → rerun → live_asign limpio
        en t4_asign → PRE-tabla muestra horarios correctos inmediatamente.
      - 📋 Caption dinámico debajo del botón: muestra todas las reasignaciones
        activas en formato "Cam X CII→CIV" para visibilidad inmediata.
    """
    import datetime as _dt

    st.subheader("📊 Proyección Picking ×5")
    st.caption(
        "Fuente: **CAR.xlsx (VE + CHESS) + Frescura 3.0 (DDM)** — Calcula PICK vs AE por "
        "**UP** (`bultos_eq / BXP`) por cancha, sumando sueltas con UNIDADES x BULTO. v4.36"
    )

    # ── Reutilizar uploads de Archivos / Tab 1 ───────────────────────────────
    car_use = st.session_state.get("t1_car") or st.session_state.get("t4_car")
    fr_use  = st.session_state.get("t1_fr")  or st.session_state.get("t4_fr")

    if not (car_use and fr_use):
        st.info("⬅️ Subí **CAR.xlsx** y **Frescura 3.0** en la pestaña **📁 Archivos** para activar la proyección.")
        return

    # ── Cargar datos (CACHEADO — solo recalcula si cambian los bytes) ─────────
    pdata = None
    try:
        with st.spinner("Calculando proyección de picking desde CAR…"):
            pdata = _t4_load_car_proyeccion(
                car_bytes=car_use.getvalue(),
                fr_bytes=fr_use.getvalue(),
            )
        st.success(
            f"✓ Proyección lista — {len(pdata['df'])} camiones | "
            f"Fecha: {pdata['fecha']} | Fuente: {pdata['fuente']}"
        )
        log_event("info", f"Tab4 proyección v4.36: {len(pdata['df'])} camiones, {pdata['fecha']}")
    except Exception as e:
        st.error(f"❌ Error calculando proyección: {e}")
        with st.expander("Stack trace"):
            import traceback
            st.code(traceback.format_exc())
        return

    if pdata is None or pdata["df"].empty:
        st.warning("No hay datos de camiones disponibles.")
        return

    df_cam = pdata["df"].copy()

    # ── Hora de inicio única (compartida) ─────────────────────────────────────
    st.divider()
    col_hi1, col_hi2, col_hi3 = st.columns([1, 1, 4])
    with col_hi1:
        h_inicio = st.number_input("Hora inicio", 0, 23, value=17, step=1, key="t4_hora_h")
    with col_hi2:
        m_inicio = st.number_input("Min inicio", 0, 59, value=30, step=5, key="t4_hora_m")
    hora_inicio_global = _dt.time(int(h_inicio), int(m_inicio))

    # ── Productividad y personas por cancha ───────────────────────────────────
    with st.expander("⚙️ Productividad y personas por cancha", expanded=False):
        prod_cols = st.columns(5)
        vel_custom  = {}
        pers_custom = {}
        for i, cn in enumerate(_T4_CANCHAS):
            with prod_cols[i]:
                st.markdown(f"**{cn.replace('CANCHA ', 'C')}**")
                vel_custom[cn] = st.number_input(
                    f"Vel {cn} (bult/h)", 10, 2000, value=_T4_VEL_DEFAULT[cn], step=10,
                    key=f"t4_vel_{cn}",
                )
                pers_custom[cn] = st.number_input(
                    f"Pers {cn}", 1, 10, value=1, step=1,
                    key=f"t4_pers_{cn}",
                )

    # Hora inicio por cancha: escalonado +5 min entre canchas
    _OFFSET_MIN = 5
    inicio_custom = {}
    base_dt = _dt.datetime.combine(_dt.date.today(), hora_inicio_global)
    for i, cn in enumerate(_T4_CANCHAS):
        inicio_custom[cn] = (base_dt + _dt.timedelta(minutes=i * _OFFSET_MIN)).time()

    # ── Filtrar: siempre sin ceros ────────────────────────────────────────────
    df_display_base = df_cam[df_cam["TOTAL_PICK"] > 0].reset_index(drop=True)

    # ── v4.71: Exclusión de camiones ─────────────────────────────────────────
    # Misma metodología que el Tablero Ruteador. Los camiones excluidos se
    # omiten de TODOS los cálculos, PDFs, Excel y exportaciones de esta sección.
    _all_cams = sorted(df_display_base["Camión"].astype(int).tolist())
    with st.expander("🚫 Excluir camiones de la proyección (opcional)", expanded=False):
        st.caption(
            "Los camiones excluidos no aparecen en la tabla, ni en PDFs ni en Excel de esta sección. "
            "Usalo para salidas parciales, camiones sin reparto o situaciones especiales."
        )
        _cams_excluir = st.multiselect(
            "Camiones a excluir:",
            options=_all_cams,
            default=st.session_state.get("t4_cams_excluir", []),
            key="t4_cams_excluir_widget",
            format_func=lambda x: f"Cam {x}",
        )
        st.session_state["t4_cams_excluir"] = _cams_excluir
        if _cams_excluir:
            st.warning(f"⚠️ Excluidos: {', '.join(str(c) for c in _cams_excluir)}")

    _cams_excluir_set = set(st.session_state.get("t4_cams_excluir", []))
    if _cams_excluir_set:
        df_display = df_display_base[~df_display_base["Camión"].astype(int).isin(_cams_excluir_set)].reset_index(drop=True)
    else:
        df_display = df_display_base

    # ── v4.55.2: Botón "Enviar a Sheets" PROMINENTE al inicio ─────────────────
    # Se muestra antes de la tabla para recordar al usuario que debe enviarlo
    # al principio, antes de reasignar o ajustar cargas, para que la Matriz Pall
    # refleje el estado real del CAR sin modificaciones manuales.
    with st.container(border=True):
        _sheets_top_cols = st.columns([3, 1])
        with _sheets_top_cols[0]:
            st.markdown(
                "#### 📤 Paso 1 — Enviar proyección a Google Sheets",
            )
            st.caption(
                "**Hacé esto primero**, antes de reasignar cargas entre canchas. "
                "El botón escribe la distribución original del CAR en la *Matriz Pall* del Sheets maestro, "
                "que es la base para el seguimiento diario de productividad y el histórico de pallets por cancha. "
                "Si enviás después de reasignar, los valores quedarán con los ajustes manuales y no con el dato fuente real.",
            )
        with _sheets_top_cols[1]:
            if st.button(
                "📤 Enviar a Sheets",
                use_container_width=True,
                type="primary",
                key="t4_sheets_top_btn",
                help="Inserta las filas de hoy en la hoja 'Matriz Pall' del Sheets maestro (dato fuente, sin reasignaciones)",
            ):
                with st.spinner("Enviando a Google Sheets…"):
                    try:
                        _creds_json_top = dict(st.secrets["gcp_service_account"])
                        _n_rows_top = _push_matriz_pall_to_sheets(df_display, pdata, _creds_json_top)
                        st.success(f"✅ {_n_rows_top} fila(s) enviadas a la Matriz Pall")
                        log_event("info", f"Sheets Matriz Pall (top): {_n_rows_top} filas ({pdata['fecha']})")
                    except KeyError:
                        st.error(
                            "❌ Falta  en . "
                            "Agregá el JSON de la service account en .",
                        )
                    except Exception as _ex_top:
                        st.error(f"❌ Error al escribir en Sheets: {_ex_top}")
                        with st.expander("Stack trace"):
                            import traceback
                            st.code(traceback.format_exc())

    st.divider()

    # ── ASIGN state ── v4.56.0 ───────────────────────────────────────────────
    # SOURCE OF TRUTH: t4_asign se escribe SIEMPRE desde edited_df en POST-tabla.
    # Es el estado completo (no deltas). El bloque PRE-tabla lo consume directo.
    # Se eliminó _rebuild_live_asign_from_editor_state porque los edited_rows de
    # Streamlit son *deltas acumulativos desde la primera edición*, no el estado
    # completo del widget. Eso causaba que al borrar una asignación (→ "") la key
    # quedaba en t4_asign → horarios congelados y sin reflejo de la reasignación.
    if "t4_asign" not in st.session_state:
        st.session_state["t4_asign"] = {}

    _current_asign = dict(st.session_state["t4_asign"])

    # ── Calcular totales con ASIGN actual (consistente con el header) ─────────
    totales_calc_pre = _t4_calcular_pall_por_cancha(df_display, _current_asign)
    status_rows_pre  = totales_calc_pre["status_rows"]

    # ── Tabla principal con ASIGN integrado ──────────────────────────────────
    st.subheader("📦 Pallets UP por camión")

    # ── v4.55.1: Horarios de salida estimados ENCIMA de la tabla (sin lag) ────
    # Usa _current_asign que ya incorpora los cambios del editor en este ciclo.
    _pre_met = _t4_calcular_metricas_por_cancha(df_display, _current_asign)
    _pre_bult = _pre_met["bultos"]
    _pre_fin_por_cancha = {}
    for _cn in _T4_CANCHAS:
        _bult_cn = _pre_bult[_cn]
        _fin_cn  = _t4_hora_fin(_bult_cn, vel_custom[_cn], pers_custom[_cn], inicio_custom[_cn])
        _pre_fin_por_cancha[_cn] = {"bultos": _bult_cn, "fin_dt": _fin_cn, "inicio": inicio_custom[_cn]}
    _pre_fin_global = max(v["fin_dt"] for v in _pre_fin_por_cancha.values())

    def _semaforo_delta_pre(fin_dt, fin_global_dt, bultos):
        """
        Semáforo comparativo entre cancha y FIN GLOBAL.
        delta_min > 0 → cancha termina ANTES que el global (holgura, puede recibir más)
        delta_min < 0 → cancha termina DESPUÉS (sobrecargada, debería ceder carga)
        delta_min ≈ 0 → emparejada con el global
        """
        if bultos <= 0:
            return "off", "sin carga"
        delta_min = (fin_global_dt - fin_dt).total_seconds() / 60.0
        if delta_min >= 15:
            # Termina mucho antes → tiene holgura, puede absorber más carga
            return "normal", f"✅ {delta_min:.0f}' holgura — puede recibir carga"
        elif delta_min >= 5:
            return "off", f"🟡 {delta_min:.0f}' holgura"
        elif delta_min >= -5:
            # Emparejada: muy cerca del fin global → bien balanceada
            return "off", "🎯 al tope del global"
        elif delta_min >= -15:
            # Sobrecargada: termina más tarde → es la cancha crítica
            return "inverse", f"⚠ {-delta_min:.0f}' sobre el global — redistribuir carga"
        else:
            return "inverse", f"🔴 {-delta_min:.0f}' sobre el global — redistribuir urgente"

    _pre_totales_pall_c = totales_calc_pre["total"]

    _hdr_top = st.columns([2, 1, 1, 1, 2])
    _hdr_top[0].markdown(f"**🕐 PICKING** — {pdata['fecha']}")
    _hdr_top[1].metric("Inicio (CI)", hora_inicio_global.strftime("%H:%M"))
    _hdr_top[2].metric("Fin global", _pre_fin_global.strftime("%H:%M"))
    _hdr_top[3].metric("Mix picking", f"{pdata['mix_picking']*100:.1f}%")
    _hdr_top[4].caption(f"Fuente: {pdata['fuente']}")

    st.markdown("##### 🏁 Horario estimado fin por cancha")
    _fin_cols_top = st.columns(len(_T4_CANCHAS) + 1)
    for _i, _cn in enumerate(_T4_CANCHAS):
        _short_cn = _cn.replace("CANCHA ", "C")
        _fp_pre   = _pre_fin_por_cancha[_cn]
        _fin_s    = _fp_pre["fin_dt"].strftime("%H:%M") if _fp_pre["bultos"] > 0 else "—"
        _ini_s    = _fp_pre["inicio"].strftime("%H:%M")
        _dc, _dm  = _semaforo_delta_pre(_fp_pre["fin_dt"], _pre_fin_global, _fp_pre["bultos"])
        _sub_top  = (
            f"{_dm} | inicio {_ini_s} → {_fp_pre['bultos']:.0f} bult"
        )
        _fin_cols_top[_i].metric(f"FIN {_short_cn}", _fin_s, _sub_top, delta_color=_dc)
    _fin_cols_top[-1].metric("🏁 FIN GLOBAL", _pre_fin_global.strftime("%H:%M"),
                             f"{sum(v['bultos'] for v in _pre_fin_por_cancha.values()):.0f} bult tot",
                             delta_color="off")
    st.divider()

    st.caption(
        "**UP CX** = fracción de pallet de picking por cancha (usá esto para reasignar). "
        "**TOT UP** = suma UP picking de la fila. **→CX** = reasignar esa carga a otra cancha. "
        "TOT PALL = AE + picking total. Barras rojas = UP picking · barra verde = PALL total."
    )

    cam_list = sorted(df_display["Camión"].tolist())
    cam_to_num = {cam: i+1 for i, cam in enumerate(cam_list)}

    # Opciones de reasignación
    _ASIGN_OPTS = ["", "CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "MKPL"]

    edit_rows = []
    max_up_cn = 0.0   # escala dinámica para ProgressColumn UP por cancha
    max_up_tot = 0.0  # escala para columna TOT UP
    for _, row in df_display.iterrows():
        cam = int(row["Camión"])
        sr  = status_rows_pre.get(cam, {})
        rec = {"#": cam_to_num.get(cam, "—"), "Camión": cam}
        tot_up_fila = 0.0
        for cn in _T4_CANCHAS:
            short = cn.replace("CANCHA ", "C")
            # v4.58: columnas muestran UP picking por cancha (fracción de pallet).
            # Es el valor correcto para reasignar: indica cuántos pallets parciales
            # tiene esa cancha → se puede comparar entre canchas para balancear.
            up_cn = round(float(row.get(f"_up_{cn}", 0.0)), 3)
            max_up_cn = max(max_up_cn, up_cn)
            tot_up_fila += up_cn
            rec[f"UP {short}"]    = up_cn
            cur_asign = _current_asign.get((cam, cn), "")
            rec[f"ASIGN {short}"] = cur_asign if cur_asign else ""
        tot_up_fila = round(tot_up_fila, 3)
        max_up_tot = max(max_up_tot, tot_up_fila)
        rec["TOT UP"]  = tot_up_fila   # suma UP picking de todas las canchas
        rec["AE Pall"] = round(float(row.get("AE_PALL", 0.0)), 1)
        rec["TOT PALL"] = round(float(sr.get("total_pall", 0.0)), 2)
        raw_status = sr.get("status", "—")
        if raw_status == "OK":
            rec["STATUS"] = "✅ OK"
        elif raw_status.startswith(">"):
            rec["STATUS"] = f"⚠️ {raw_status}"
        else:
            rec["STATUS"] = raw_status
        edit_rows.append(rec)

    df_editor_in = pd.DataFrame(edit_rows)

    # ── column_config v4.58 ──────────────────────────────────────────────────
    # UP CI/CII/…: ProgressColumn rojo (escala dinámica, máx UP de la operación)
    # TOT UP: suma UP picking fila → referencia para reasignar
    # TOT PALL: ProgressColumn verde (AE + pick, escala fija 0–12)
    up_max_scale  = max(2.0, max_up_cn  * 1.1)
    upt_max_scale = max(5.0, max_up_tot * 1.1)

    col_cfg = {
        "#":        st.column_config.NumberColumn("#", disabled=True, width="small"),
        "Camión":   st.column_config.NumberColumn("Cam.", disabled=True, width="small"),
        "TOT UP":   st.column_config.ProgressColumn(
            "TOT UP",
            help="Suma de UP picking de todas las canchas para este camión.",
            format="%.2f", min_value=0.0, max_value=upt_max_scale,
            width="small",
        ),
        "AE Pall":  st.column_config.NumberColumn(
            "AE Pall", disabled=True, format="%.1f", width="small",
        ),
        "TOT PALL": st.column_config.ProgressColumn(
            "TOT PALL",
            help="Pallets totales por camión — AE + picking (barra verde, escala 0–12).",
            format="%.2f", min_value=0.0, max_value=12.0,
            width="small",
        ),
        "STATUS":   st.column_config.TextColumn("OK?", disabled=True, width="small"),
    }
    for cn in _T4_CANCHAS:
        short = cn.replace("CANCHA ", "C")
        disp  = _cn_short(cn)                    # v4.64: label natural (C1, C2…)
        col_cfg[f"UP {short}"] = st.column_config.ProgressColumn(
            f"UP {disp}",
            help=f"UP picking en {cn} — fracción de pallet. Usá esto para reasignar.",
            format="%.2f" if cn != "MKPL" else "%.3f",
            min_value=0.0,
            max_value=up_max_scale,
            width="small",
        )
        col_cfg[f"ASIGN {short}"] = st.column_config.SelectboxColumn(
            f"→{disp}", options=_ASIGN_OPTS, default="", required=False,
            width="small",
        )

    ordered_cols = ["#", "Camión"]
    for cn in _T4_CANCHAS:
        short = cn.replace("CANCHA ", "C")
        ordered_cols += [f"UP {short}", f"ASIGN {short}"]
    ordered_cols += ["TOT UP", "AE Pall", "TOT PALL", "STATUS"]
    df_editor_in = df_editor_in[ordered_cols]

    edited_df = st.data_editor(
        df_editor_in,
        column_config=col_cfg,
        use_container_width=True,
        hide_index=True,
        key="t4_main_editor",
        num_rows="fixed",
    )

    # ── v4.56.0: construir live_asign desde edited_df COMPLETO ─────────────
    # Lee TODAS las filas del editor (no solo edited_rows / deltas).
    # Esto garantiza que borrar una asignación (→ "") la elimine correctamente
    # del estado. El resultado se persiste en t4_asign como source of truth.
    live_asign = {}
    for _, erow in edited_df.iterrows():
        cam_e = int(erow["Camión"])
        for cn in _T4_CANCHAS:
            short = cn.replace("CANCHA ", "C")
            new_val = (erow.get(f"ASIGN {short}", "") or "").strip()
            if new_val:
                live_asign[(cam_e, cn)] = new_val
            # Si está vacío → simplemente no se agrega: la key queda ausente
            # (borrado correcto, sin "congelar" asignaciones previas)
    # Persistir estado completo — PRE-tabla lo leerá en el próximo rerun
    st.session_state["t4_asign"] = live_asign

    # ── v4.56.0: Botón "Recalcular horarios" ────────────────────────────────
    # Al hacer cambios en ASIGN, live_asign ya está correcto en este ciclo
    # para los cálculos POST-tabla. El header PRE-tabla (encima del editor)
    # se actualiza al ciclo SIGUIENTE. El botón fuerza ese rerun inmediatamente.
    # Botón en la primera columna; caption en la segunda.
    # Las tablas de concepto debajo son HTML con anchos fijos (v4.63.0).
    _rcol1, _rcol2, _rcol3, _rcol4, _rcol5, _rcol6 = st.columns([2, 1, 1, 1, 1, 1])
    with _rcol1:
        if st.button(
            "🔄 Recalcular horarios",
            key="t4_recalc_btn",
            type="secondary",
            use_container_width=True,
            help="Aplica las reasignaciones y actualiza los horarios estimados por cancha.",
        ):
            # t4_asign ya fue persistido con live_asign arriba → rerun lo lee
            st.rerun()
    # El caption ocupa las 5 columnas de cancha, alineado con ellas
    with _rcol2:
        if live_asign:
            _asign_count = len(live_asign)
            _asign_desc = " | ".join(
                f"Cam {k[0]} {_cn_short(k[1])}→{_cn_short(v)}"
                for k, v in sorted(live_asign.items())
            )
            st.caption(f"✏️ {_asign_count} reasignación(es) activa(s): {_asign_desc}")
        else:
            st.caption("Sin reasignaciones activas.")

    # ── Recalcular totales con live_asign (UNA sola vez) ─────────────────────
    totales_calc   = _t4_calcular_pall_por_cancha(df_display, live_asign)
    totales_pall_c = totales_calc["total"]

    # ── Filas SUB / DESIGNADOS / TOTAL ───────────────────────────────────────
    # v4.63.0 — Tablas de concepto renderizadas como HTML para alineación
    # pixel-perfecta con las columnas CI/CII/CIII/CIV/MKPL de la tabla
    # de reasignación. La columna Concepto es delgada (130px); cada cancha
    # ocupa 1fr del espacio restante, replicando la distribución visual del
    # data_editor principal.
    _cancha_cols_pall = [_cn_short(cn) for cn in _T4_CANCHAS]  # v4.64: números naturales

    def _html_concepto_tables(
        canchas_short,
        pall_sub, pall_des, pall_tot,
        bult_vals, hl_vals, kg_vals, fin_vals,
    ):
        """Genera dos tablas HTML alineadas con la tabla de reasignación."""
        # Anchos: col Concepto fija 130px, canchas igual distribución con flex
        # Usamos tabla con table-layout:fixed y porcentajes para que respeten
        # el ancho del contenedor igual que use_container_width=True.
        # Concepto = 14% del ancho total (≈130px en pantalla típica ~930px).
        # Cada cancha = (86% / N_canchas).
        N = len(canchas_short)
        conc_pct = 14
        each_pct = round((100 - conc_pct) / N, 2)

        def _td(val, bg="", color="", bold=False, align="right"):
            style = f"padding:4px 8px;text-align:{align};font-size:12px;"
            if bg:    style += f"background:{bg};"
            if color: style += f"color:{color};"
            if bold:  style += "font-weight:600;"
            return f'<td style="{style}">{val}</td>'

        def _th(val, align="right"):
            style = (
                f"padding:4px 8px;text-align:{align};font-size:11px;"
                "background:#1e3a5f;color:#fff;font-weight:600;"
            )
            return f'<th style="{style}">{val}</th>'

        tbl_style = (
            "width:100%;table-layout:fixed;border-collapse:collapse;"
            "margin-bottom:4px;"
        )
        col_tags = (
            f'<col style="width:{conc_pct}%">'
            + "".join(f'<col style="width:{each_pct}%">' for _ in canchas_short)
        )

        # ── Tabla 1: PALL ────────────────────────────────────────────────────
        hdr = _th("Concepto", align="left") + "".join(_th(c) for c in canchas_short)
        rows_pall = [
            ("SUB Total PALL",     pall_sub, "",        "",      False),
            ("DESIGNADOS PALL",    pall_des, "",        "",      False),
            ("TOTAL PALL x CANCHA",pall_tot, "#FF8C00", "#fff",  True),
        ]
        body1 = ""
        for label, vals, bg, fg, bold in rows_pall:
            cells = _td(label, bg=bg, color=fg, bold=bold, align="left")
            for v in vals:
                cells += _td(f"{v:.2f}", bg=bg, color=fg, bold=bold)
            body1 += f'<tr style="border-bottom:1px solid #333;">{cells}</tr>'

        table1 = (
            f'<table style="{tbl_style}">'
            f"<colgroup>{col_tags}</colgroup>"
            f"<thead><tr>{hdr}</tr></thead>"
            f"<tbody>{body1}</tbody>"
            f"</table>"
        )

        # ── Tabla 2: BULTOS / HL / KG / FIN ─────────────────────────────────
        hdr2 = _th("Concepto", align="left") + "".join(_th(c) for c in canchas_short)
        rows_extra = [
            ("TOTAL BULTOS",   [f"{v:.1f}" for v in bult_vals]),
            ("TOTAL HL",       [f"{v:.2f}" for v in hl_vals]),
            ("TOTAL KG",       [f"{v:.0f}" for v in kg_vals]),
            ("HORA EST. FIN",  fin_vals),
        ]
        body2 = ""
        for label, vals in rows_extra:
            cells = _td(label, align="left")
            for v in vals:
                cells += _td(v)
            body2 += f'<tr style="border-bottom:1px solid #333;">{cells}</tr>'

        table2 = (
            f'<table style="{tbl_style}">'
            f"<colgroup>{col_tags}</colgroup>"
            f"<thead><tr>{hdr2}</tr></thead>"
            f"<tbody>{body2}</tbody>"
            f"</table>"
        )

        wrapper_style = (
            "font-family:sans-serif;background:#0e1117;border-radius:6px;"
            "padding:6px 10px 2px 10px;margin-bottom:6px;"
        )
        return f'<div style="{wrapper_style}">{table1}{table2}</div>'

    # Calcular valores para tabla 1 (PALL)
    pall_sub_vals = [round(totales_calc["sub"][cn],       2) for cn in _T4_CANCHAS]
    pall_des_vals = [round(totales_calc["designados"][cn], 2) for cn in _T4_CANCHAS]
    pall_tot_vals = [round(totales_calc["total"][cn],      2) for cn in _T4_CANCHAS]

    # ── Totales bultos/HL/KG por cancha (post-asign) ─────────────────────────
    totales_bult_base = {cn: float(df_display[cn].sum()) if cn in df_display.columns else 0.0 for cn in _T4_CANCHAS}
    _met_post = _t4_calcular_metricas_por_cancha(df_display, live_asign)
    totales_bult = _met_post["bultos"]
    totales_hl   = _met_post["hl"]
    totales_kg   = _met_post["kg"]

    # ── Hora estimada de fin POR CANCHA (sobre bultos reales post-asign) ──────
    fin_por_cancha = {}
    for cn in _T4_CANCHAS:
        bult   = totales_bult[cn]
        fin_dt = _t4_hora_fin(bult, vel_custom[cn], pers_custom[cn], inicio_custom[cn])
        fin_por_cancha[cn] = {"bultos": bult, "fin_dt": fin_dt, "inicio": inicio_custom[cn]}

    bult_list = [totales_bult[cn] for cn in _T4_CANCHAS]
    hl_list   = [totales_hl[cn]   for cn in _T4_CANCHAS]
    kg_list   = [totales_kg[cn]   for cn in _T4_CANCHAS]
    fin_list  = [
        fin_por_cancha[cn]["fin_dt"].strftime("%H:%M") if totales_bult[cn] > 0 else "—"
        for cn in _T4_CANCHAS
    ]

    st.markdown(
        _html_concepto_tables(
            _cancha_cols_pall,
            pall_sub_vals, pall_des_vals, pall_tot_vals,
            bult_list, hl_list, kg_list, fin_list,
        ),
        unsafe_allow_html=True,
    )

    with st.expander("📊 Totales por cancha", expanded=True):
        cols_met = st.columns(len(_T4_CANCHAS) + 1)
        for i, cn in enumerate(_T4_CANCHAS):
            disp = _cn_short(cn)  # v4.64: números naturales
            delta_bult = totales_bult[cn] - totales_bult_base[cn]
            delta_str  = f"{delta_bult:+.0f} bult (asign)" if delta_bult != 0 else None
            cols_met[i].metric(f"Bult {disp}", f"{totales_bult[cn]:.1f}", delta=delta_str)
        cols_met[-1].metric("TOT PICK", f"{pdata['tot_pick']:.0f}")

        cols_met2 = st.columns(len(_T4_CANCHAS) + 1)
        for i, cn in enumerate(_T4_CANCHAS):
            disp = _cn_short(cn)  # v4.64: números naturales
            cols_met2[i].metric(f"PALL {disp}", f"{totales_calc['total'].get(cn, 0):.2f}")
        cols_met2[-1].metric("TOT AE bult", f"{pdata['tot_ae']:.0f}")

    # ── fin_global_dt ─────────────────────────────────────────────────────────
    fin_global_dt = max(v["fin_dt"] for v in fin_por_cancha.values())
    fin_calc_dict = {"por_cancha": fin_por_cancha, "fin_global_dt": fin_global_dt}

    # ── v4.36: Semáforo de cancha — detecta desfase vs FIN GLOBAL ─────────────
    # Verde: ±5min del FIN GLOBAL (parejas) | Ámbar: 5–15min | Rojo: >15min
    # Sirve para identificar canchas atrasadas (rojo) o con holgura (verde
    # con tiempo sobrante) para decidir reasignar.
    def _semaforo_delta(fin_dt, fin_global_dt, bultos):
        """Semáforo POST-tabla (idéntica lógica al PRE)."""
        if bultos <= 0:
            return "off", "sin carga"
        delta_min = (fin_global_dt - fin_dt).total_seconds() / 60.0
        if delta_min >= 15:
            return "normal", f"✅ {delta_min:.0f}' holgura — puede recibir carga"
        elif delta_min >= 5:
            return "off", f"🟡 {delta_min:.0f}' holgura"
        elif delta_min >= -5:
            return "off", "🎯 al tope del global"
        elif delta_min >= -15:
            return "inverse", f"⚠ {-delta_min:.0f}' sobre el global — redistribuir carga"
        else:
            return "inverse", f"🔴 {-delta_min:.0f}' sobre el global — redistribuir urgente"

    # ── Cabecera PICKING (detalle) ────────────────────────────────────────────
    st.divider()
    hdr_cols = st.columns([2, 1, 1, 1, 2])
    hdr_cols[0].markdown(f"**🕐 PICKING** — {pdata['fecha']}")
    hdr_cols[1].metric("Inicio (CI)", hora_inicio_global.strftime("%H:%M"))
    hdr_cols[2].metric("Fin global", fin_global_dt.strftime("%H:%M"))
    hdr_cols[3].metric("Mix picking", f"{pdata['mix_picking']*100:.1f}%")
    hdr_cols[4].caption(f"Fuente: {pdata['fuente']}")

    st.markdown("##### 🏁 Horario fin por cancha — semáforo de desfase")
    fin_cols = st.columns(len(_T4_CANCHAS) + 1)
    for i, cn in enumerate(_T4_CANCHAS):
        disp = _cn_short(cn)  # v4.64: números naturales
        fp = fin_por_cancha[cn]
        fin_s = fp["fin_dt"].strftime("%H:%M") if fp["bultos"] > 0 else "—"
        ini_s = fp["inicio"].strftime("%H:%M")
        delta_color, delta_msg = _semaforo_delta(fp["fin_dt"], fin_global_dt, fp["bultos"])
        sub_line = (
            f"{delta_msg} | inicio {ini_s} → {fp['bultos']:.0f} bult"
        )
        fin_cols[i].metric(f"FIN {disp}", fin_s, sub_line, delta_color=delta_color)
    fin_cols[-1].metric("🏁 FIN GLOBAL", fin_global_dt.strftime("%H:%M"),
                        f"{sum(v['bultos'] for v in fin_por_cancha.values()):.0f} bult tot",
                        delta_color="off")

    # ── Generar PDF ×5 + Excel ────────────────────────────────────────────────
    st.divider()
    st.subheader("📄 PDF Pickeros — Resumen por cancha (C1/C2/C3/C4/MKPL+Merch)")

    # v4.41: calcular ctrl_params ANTES de los columns, para pasarlo al PDF ×5
    import datetime as _dt_res
    _top_skus_lines = ""
    try:
        _car_b = st.session_state.get("t1_car") or st.session_state.get("t4_car")
        _fr_b  = st.session_state.get("t1_fr")
        if _car_b and _fr_b:
            _df_top = _build_top10_skus(_car_b.getvalue(), _fr_b.getvalue())
            if not _df_top.empty:
                _top3 = _df_top.nlargest(3, "BULTOS")
                _lines = []
                for _, _sk in _top3.iterrows():
                    _cod  = int(_sk["CODIGO"]) if not pd.isna(_sk["CODIGO"]) else "—"
                    _desc = str(_sk.get("DESCRIPCION", "")).strip()[:30] or "—"
                    _blt  = int(_sk["BULTOS"])
                    _lines.append(f"  • **{_cod}** {_desc} — {_blt} bult")
                _top_skus_lines = "\n".join(_lines)
    except Exception:
        _top_skus_lines = "  *(no disponible)*"

    _alertas = []
    for _cn in _T4_CANCHAS:
        _fp = fin_por_cancha[_cn]
        if _fp["bultos"] > 0:
            _delta = (fin_global_dt - _fp["fin_dt"]).total_seconds() / 60.0
            if _delta < -15:
                _short_cn = _cn_short(_cn)  # v4.64: números naturales
                _alertas.append(f"🔴 {_short_cn} atrasada {abs(_delta):.0f}' vs global")

    _ctrl_params_for_pdf = {
        "fin_por_cancha":  fin_por_cancha,
        "totales_bult":    totales_bult,
        "totales_pall_c":  totales_pall_c,
        "totales_hl":      totales_hl,
        "totales_kg":      totales_kg,
        "n_camiones":      len(df_display[df_display["TOTAL_PICK"] > 0]),
        "tot_pick":        pdata["tot_pick"],
        "top_skus_lines":  _top_skus_lines,
        "alertas":         _alertas,
    }

    cp1, cp2 = st.columns([1, 1])
    # v4.71: layout vertical — pickeros arriba, controladores abajo

    # ── v4.71: Enseñanza/frase del día ──────────────────────────────────────
    _ensenanza = get_ensenanza_dia()
    st.info(f"💡 **Enseñanza del día** — {_ensenanza}")

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN A: PDF PICKEROS
    # ─────────────────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("#### 📄 PDF Pickeros — Resumen por cancha (C1/C2/C3/C4/MKPL+Merch)")
        if st.button("📄 Resumen Pickeros — C1/C2/C3/C4/MKPL+Merch (sin paletas completas)", type="primary",
                     use_container_width=True, key="t4_pdf_btn"):
            with st.spinner("Generando PDF…"):
                try:
                    pdf_bytes = _t4_generar_pdf_x4(
                        df=df_display,
                        totales_bult=totales_bult,
                        totales_pall=totales_pall_c,
                        totales_hl=totales_hl,
                        totales_kg=totales_kg,
                        productividad=vel_custom,
                        personas=pers_custom,
                        inicio=hora_inicio_global,
                        fecha=pdata["fecha"],
                        mix_picking=pdata["mix_picking"],
                        fin_calc=fin_calc_dict,
                        asign_detail=totales_calc.get("asign_detail", {}),
                        ctrl_params=None,   # v4.70: hoja controlador se genera por separado
                    )
                    fname = f"proyeccion_picking_{pdata['fecha']}.pdf"
                    st.download_button(
                        "⬇ Descargar PDF Pickeros (C1/C2/C3/C4/MKPL+Merch — LANDSCAPE)",
                        data=pdf_bytes, file_name=fname,
                        mime="application/pdf", use_container_width=True,
                        key="t4_pdf_dl",
                    )
                    st.success(f"✓ PDF generado: {fname}")
                    log_event("info", f"PDF Proyección v4.41 generado: {fname}")
                except Exception as e:
                    st.error(f"❌ Error generando PDF: {e}")
                    with st.expander("Stack trace"):
                        import traceback
                        st.code(traceback.format_exc())

        # ── Botón XLSX proyección ──────────────────────────────────────────────
        st.markdown("---")
        try:
            import datetime as _dt_xlsx
            # Construir DataFrame exportable de proyección
            _export_rows = []
            for _, _row in df_display.iterrows():
                _exp = {"Camión": int(_row["Camión"])}
                for _cn in _T4_CANCHAS:
                    _short = _cn.replace("CANCHA ", "C")
                    _disp  = _cn_short(_cn)       # v4.64: label natural
                    _exp[f"UP {_disp}"]   = round(float(_row.get(f"_up_{_cn}", 0.0)), 3)
                    _exp[f"Bult {_disp}"] = round(float(_row.get(_cn, 0.0)), 1)
                _exp["AE Pall"]    = round(float(_row.get("AE_PALL", 0.0)), 2)
                _exp["TOT PALL"]   = round(float(_row.get("TOTAL_PALL", 0.0)), 2)
                _exp["Bult Pick"]  = round(float(_row.get("TOTAL_PICK", 0.0)), 1)
                _export_rows.append(_exp)
            df_xlsx_exp = pd.DataFrame(_export_rows)
            # Fila totales
            _tot_row = {"Camión": "TOTAL"}
            for _cn in _T4_CANCHAS:
                _disp = _cn_short(_cn)            # v4.64: label natural
                _tot_row[f"UP {_disp}"]   = round(totales_pall_c.get(_cn, 0), 3)
                _tot_row[f"Bult {_disp}"] = round(totales_bult.get(_cn, 0), 1)
            _tot_row["AE Pall"]   = ""
            _tot_row["TOT PALL"]  = round(sum(totales_pall_c.values()), 2)
            _tot_row["Bult Pick"] = round(pdata["tot_pick"], 1)
            df_xlsx_exp = pd.concat([df_xlsx_exp, pd.DataFrame([_tot_row])], ignore_index=True)

            xlsx_bytes = df_to_xlsx(df_xlsx_exp, sheet_name="Proyección")
            fname_xlsx = f"proyeccion_picking_{pdata['fecha']}.xlsx"
            st.download_button(
                "📊 Descargar Excel (XLSX)",
                data=xlsx_bytes, file_name=fname_xlsx,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="t4_xlsx_dl",
            )
        except Exception as _ex_xlsx:
            st.warning(f"⚠️ XLSX no disponible: {_ex_xlsx}")

    # ─────────────────────────────────────────────────────────────────────────
    # SECCIÓN B: RESUMEN CONTROLADOR + PDF CONTROLADORES
    # ─────────────────────────────────────────────────────────────────────────
    with st.container(border=True):
        # ── v4.37: Resumen para el Controlador ───────────────────────────────
        # v4.41: _top_skus_lines y _alertas ya computados antes del bloque columns

        # Construir horarios por cancha
        _horas_lines = []
        for _cn in _T4_CANCHAS:
            _fp = fin_por_cancha[_cn]
            _short_cn = _cn.replace("CANCHA ", "C")
            if _fp["bultos"] > 0:
                _ini_s = _fp["inicio"].strftime("%H:%M")
                _fin_s = _fp["fin_dt"].strftime("%H:%M")
                _blt_s = f"{_fp['bultos']:.0f} bult"
                _horas_lines.append(f"  • **{_short_cn}**: {_ini_s} → {_fin_s} ({_blt_s})")

        _fecha_str_res = pdata["fecha"].strftime("%d/%m/%Y") if hasattr(pdata["fecha"], "strftime") else str(pdata["fecha"])

        # ── v4.71: Top 3 por CANCHA ──────────────────────────────────────────
        _top3_por_cancha_lines = []
        try:
            _car_b71 = st.session_state.get("t1_car") or st.session_state.get("t4_car")
            _fr_b71  = st.session_state.get("t1_fr")
            if _car_b71 and _fr_b71:
                _df_top71 = _build_top10_skus(_car_b71.getvalue(), _fr_b71.getvalue())
                if not _df_top71.empty and "CANCHA" in _df_top71.columns:
                    for _cn71 in _T4_CANCHAS:
                        _cn71_short = _cn71.replace("CANCHA ", "C")
                        _df_cn = _df_top71[_df_top71["CANCHA"].str.upper().str.contains(
                            _cn71.upper().replace("CANCHA ", "").strip(), na=False)]
                        if _df_cn.empty:
                            continue
                        _t3 = _df_cn.nlargest(3, "BULTOS")
                        _items = []
                        for _, _sk71 in _t3.iterrows():
                            _cod71  = int(_sk71["CODIGO"]) if not pd.isna(_sk71["CODIGO"]) else "—"
                            _desc71 = str(_sk71.get("DESCRIPCION","")).strip()[:28] or "—"
                            _blt71  = int(_sk71["BULTOS"])
                            _items.append(f"{_cod71} {_desc71} ({_blt71}b)")
                        if _items:
                            _top3_por_cancha_lines.append(f"  • **{_cn71_short}**: " + " | ".join(_items))
        except Exception:
            pass

        # ── v4.71: Info de Frescura (FEFO críticos por cancha) ──────────────
        _frescura_lines = []
        try:
            _fr_b71f = st.session_state.get("t1_fr")
            _ddm_d71 = st.session_state.get("df_ddm_dict", {})
            if _fr_b71f and _ddm_d71:
                import openpyxl as _ox71
                _wb71 = _ox71.load_workbook(_fr_b71f, read_only=True, data_only=True)
                _fr_b71f.seek(0)
                if "Frescura" in _wb71.sheetnames:
                    _wsfr71 = _wb71["Frescura"]
                    _rfr71 = list(_wsfr71.iter_rows(min_row=1, values_only=True))
                    if _rfr71:
                        _hfr71 = [str(v).strip().upper() if v else "" for v in _rfr71[0]]
                        _isku71  = next((i for i, h in enumerate(_hfr71) if "ALM" in h or "CÓD" in h or "COD" in h), 1)
                        _idate71 = next((i for i, h in enumerate(_hfr71) if "FECHA" in h or "VENC" in h), 5)
                        _ican71  = next((i for i, h in enumerate(_hfr71) if "CAN" in h), -1)
                        _istk71  = next((i for i, h in enumerate(_hfr71) if "STOCK" in h and "TOTAL" in h), 14)
                        _can_skus = {}
                        for _rr71 in _rfr71[1:]:
                            try:
                                _sku71 = int(float(str(_rr71[_isku71]))) if _rr71[_isku71] else 0
                                if not _sku71: continue
                                _fecha71 = _rr71[_idate71]
                                _can71   = str(_rr71[_ican71]).strip() if _ican71 >= 0 and _rr71[_ican71] else ""
                                _stk71   = float(_rr71[_istk71]) if _istk71 < len(_rr71) and _rr71[_istk71] else 0
                                if not _can71: continue
                                if _sku71 not in _can_skus:
                                    _can_skus[_sku71] = {"can": _can71, "fecha": _fecha71, "stk": _stk71}
                            except Exception:
                                pass
                        # Filtrar los de fecha más próxima (FEFO)
                        import datetime as _dt71
                        _today71 = _dt71.date.today()
                        _venc_prox = []
                        for _s71, _v71 in _can_skus.items():
                            try:
                                _f71 = _v71["fecha"]
                                if isinstance(_f71, (int, float)):
                                    _f71 = _dt71.date(1899,12,30) + _dt71.timedelta(days=int(_f71))
                                elif hasattr(_f71, "date"):
                                    _f71 = _f71.date()
                                elif isinstance(_f71, str):
                                    _f71 = pd.to_datetime(_f71, dayfirst=True, errors="coerce").date()
                                if _f71 and not pd.isna(_f71) and _f71 >= _today71:
                                    _dias = (_f71 - _today71).days
                                    _venc_prox.append((_dias, _s71, _v71["can"], _f71, _v71["stk"]))
                            except Exception:
                                pass
                        _venc_prox.sort()
                        for _dias71, _sku71, _can71, _f71, _stk71 in _venc_prox[:5]:
                            _alert_icon = "🔴" if _dias71 <= 10 else ("🟡" if _dias71 <= 30 else "🟢")
                            _frescura_lines.append(
                                f"  {_alert_icon} SKU **{_sku71}** ({_can71}) — vence {_f71.strftime('%d/%m/%Y')} en {_dias71}d · stock {_stk71:.0f}b"
                            )
                _wb71.close()
        except Exception:
            pass

        _resumen_md = (
            f"### 📋 Resumen Controlador — {_fecha_str_res}\n\n"
            f"**🕐 Horarios de finalización por cancha:**\n"
            + "\n".join(_horas_lines) +
            f"\n\n**🏁 Fin estimado global: `{fin_global_dt.strftime('%H:%M')}`**\n\n"
            f"---\n"
            f"**📦 Totales operativos:**\n"
            f"  • Camiones con picking: **{len(df_display[df_display['TOTAL_PICK'] > 0])}**\n"
            f"  • Bultos totales: **{pdata['tot_pick']:.0f}**\n"
            f"  • Bultos MKPL: **{totales_bult.get('MKPL', 0):.0f}**\n"
            f"  • HL total picking: **{sum(totales_hl.values()):.1f}**\n"
            f"  • KG total picking: **{sum(totales_kg.values()):.0f}**\n"
            f"  • Mix picking: **{pdata['mix_picking']*100:.1f}%**\n\n"
        )
        if _top_skus_lines:
            _resumen_md += f"**🏆 Top 3 SKUs del día (por bultos):**\n{_top_skus_lines}\n\n"
        if _top3_por_cancha_lines:
            _resumen_md += f"**📍 Top 3 por cancha:**\n" + "\n".join(_top3_por_cancha_lines) + "\n\n"
        if _frescura_lines:
            _resumen_md += f"**📅 FEFO críticos (próximos a vencer):**\n" + "\n".join(_frescura_lines) + "\n\n"
        if _alertas:
            _resumen_md += "**⚠️ Alertas:**\n" + "\n".join(f"  {a}" for a in _alertas) + "\n"
        _resumen_md += f"\n---\n💡 *{_ensenanza}*"

        st.info(_resumen_md)

        st.divider()
        st.markdown("#### 📋 PDF Controladores — Resumen operativo enriquecido")
        st.caption("Incluye: Top 5 SKUs por cancha, análisis de carga, horarios, recomendaciones SOPs, FEFO y enseñanza del día.")

        if st.button("📄 PDF Controladores", type="secondary",
                     use_container_width=True, key="t4_pdf_ctrl_btn",
                     help="Genera PDF landscape con el resumen completo para imprimir junto a las copias de canchas"):
            with st.spinner("Generando PDF Controlador…"):
                try:
                    _pdf_ctrl = _t4_generar_pdf_controlador(
                        fecha=pdata["fecha"],
                        fin_por_cancha=fin_por_cancha,
                        fin_global_dt=fin_global_dt,
                        totales_bult=totales_bult,
                        totales_hl=totales_hl,
                        totales_kg=totales_kg,
                        totales_pall_c=totales_pall_c,
                        mix_picking=pdata["mix_picking"],
                        n_camiones=len(df_display[df_display["TOTAL_PICK"] > 0]),
                        tot_pick=pdata["tot_pick"],
                        top_skus_lines=_top_skus_lines,
                        alertas=_alertas,
                    )
                    _fecha_ctrl = pdata["fecha"].strftime("%d-%m-%Y") if hasattr(pdata["fecha"], "strftime") else str(pdata["fecha"])
                    _fname_ctrl = f"{_fecha_ctrl}_resumen_controlador_picking.pdf"
                    st.download_button(
                        "⬇ Descargar PDF Controlador",
                        data=_pdf_ctrl, file_name=_fname_ctrl,
                        mime="application/pdf", use_container_width=True,
                        key="t4_pdf_ctrl_dl",
                    )
                    st.success(f"✓ PDF Controlador: {_fname_ctrl}")
                except Exception as _e_ctrl:
                    st.error(f"❌ Error generando PDF Controlador: {_e_ctrl}")
                    with st.expander("Stack trace"):
                        import traceback
                        st.code(traceback.format_exc())

    # ── Paso 2 — Enviar resumen (fx) Picking a Google Sheets ─────────────────
    # v4.65: Se ubica al FINAL de la sección, después de generar los PDFs,
    # para que los operadores siempre generen primero los PDFs de cancha
    # (y apliquen las reasignaciones si las hay) y LUEGO registren el dato.
    # El aviso en rojo recuerda ajustar los bultos si hubo reasignaciones.
    st.divider()
    with st.container(border=True):
        st.markdown("#### 📥 Paso 2 — Registrar picking en (fx) Picking (Inputs Sheet)")

        # Aviso prominente sobre reasignaciones
        st.error(
            "⚠️ **IMPORTANTE — ANTES DE ENVIAR:** "
            "Si reasignaste cargas entre canchas en la tabla de arriba, los bultos "
            "enviados aquí ya reflejan esa reasignación. "
            "Si además moviste bultos **físicamente entre canchas** (por ejemplo, "
            "un autoelevador llevó una carga de C1 a C3), "
            "debés ajustar manualmente esa cantidad en la fila del sheet "
            "**DESPUÉS** de que este botón haya completado la escritura.",
            icon="🔴",
        )

        # Preview de los valores que se van a escribir
        with st.expander("👁️ Vista previa de valores a enviar", expanded=False):
            _T4_C_p2 = _T4_CANCHAS
            _prev_cols = st.columns(len(_T4_C_p2) + 1)
            _prev_cols[0].markdown("**Métrica**")
            for _pi, _pc in enumerate(_T4_C_p2):
                _prev_cols[_pi + 1].markdown(f"**{_cn_short(_pc)}**")

            def _prev_row(label, vals_dict, fmt=".1f"):
                _row_c = st.columns(len(_T4_C_p2) + 1)
                _row_c[0].caption(label)
                for _pi, _pc in enumerate(_T4_C_p2):
                    _row_c[_pi + 1].caption(f"{vals_dict.get(_pc, 0):{fmt}}")

            _prev_row("PICK BULTOS", totales_bult, ".1f")
            _prev_row("AE BULTOS",   pdata.get("bult_ae_por_cancha", {}), ".1f")
            _prev_row("PESO KG",     totales_kg, ".0f")
            _prev_row("PICKING HL",  totales_hl, ".3f")

            # Preview HL venta (necesita ANR)
            _anr_df_prev = st.session_state.get("anr_df")
            if _anr_df_prev is not None:
                _hl_venta_prev = _t4_compute_hl_venta_por_cancha(
                    _anr_df_prev,
                    pdata.get("ddm_can_lookup", {}),
                )
                _prev_row("HL VENTA (ANR)", _hl_venta_prev, ".3f")
            else:
                st.caption("*(HL VENTA no disponible — subí ANR.xlsx en Archivos)*")

        # Verificar si ANR está cargado
        _anr_disponible = st.session_state.get("anr_df") is not None

        if not _anr_disponible:
            st.warning(
                "⬅️ Para incluir TOTAL HL VENTA subí **ANR.xlsx** en la pestaña 📁 Archivos. "
                "Si no está cargado, esa columna quedará en 0."
            )

        _p2_col1, _p2_col2 = st.columns([1, 2])
        with _p2_col1:
            if st.button(
                "📥 Enviar a (fx) Picking",
                type="primary",
                use_container_width=True,
                key="t4_paso2_btn",
                help="Busca la fila de hoy en la hoja (fx) Picking y completa las columnas B:BL",
            ):
                with st.spinner("Calculando y enviando a (fx) Picking…"):
                    try:
                        _creds_p2 = dict(st.secrets["gcp_service_account"])

                        # HL venta por cancha desde ANR
                        _anr_df_p2 = st.session_state.get("anr_df")
                        _hl_venta_p2 = _t4_compute_hl_venta_por_cancha(
                            _anr_df_p2,
                            pdata.get("ddm_can_lookup", {}),
                        )

                        _resultado_p2 = _push_fx_picking_to_sheets(
                            totales_bult        = totales_bult,
                            totales_hl          = totales_hl,
                            totales_kg          = totales_kg,
                            bult_ae_por_cancha  = pdata.get("bult_ae_por_cancha", {}),
                            hl_venta_por_cancha = _hl_venta_p2,
                            tot_pick            = pdata["tot_pick"],
                            tot_ae              = pdata["tot_ae"],
                            fecha               = pdata["fecha"],
                            credentials_json    = _creds_p2,
                        )
                        st.success(f"✅ {_resultado_p2}")
                        log_event(
                            "info",
                            f"Paso 2 (fx) Picking: {_resultado_p2} | "
                            f"pick={pdata['tot_pick']:.0f} bult | "
                            f"ae={pdata['tot_ae']:.0f} bult | "
                            f"hl_pick={sum(totales_hl.values()):.2f} | "
                            f"hl_venta={sum(_hl_venta_p2.values()):.2f}",
                        )
                    except KeyError:
                        st.error(
                            "❌ Falta `gcp_service_account` en `st.secrets`. "
                            "Agregá el JSON de la service account en `.streamlit/secrets.toml`."
                        )
                    except Exception as _ex_p2:
                        st.error(f"❌ Error al escribir en (fx) Picking: {_ex_p2}")
                        with st.expander("Stack trace"):
                            import traceback
                            st.code(traceback.format_exc())

        with _p2_col2:
            _fecha_str_p2 = (
                pdata["fecha"].strftime("%d/%m/%Y")
                if hasattr(pdata["fecha"], "strftime")
                else str(pdata["fecha"])
            )
            st.info(
                f"**¿Qué hace este botón?**\n\n"
                f"Busca la fila del **{_fecha_str_p2}** en la hoja `(fx) Picking` "
                f"del Inputs Sheet y completa las columnas **B:BL** con los datos "
                f"calculados de esta sesión (picking + AE + HL + KG + HL venta del ANR). "
                f"Las columnas de matrices quedan en 0 (se calculan en el sheet). "
                f"No inserta filas nuevas, no rompe fórmulas."
            )

        # ── Excel descargable — siempre disponible ─────────────────────────
        st.markdown("##### 📊 Excel de respaldo")
        _anr_df_xl  = st.session_state.get("anr_df")
        _hl_venta_xl = _t4_compute_hl_venta_por_cancha(
            _anr_df_xl,
            pdata.get("ddm_can_lookup", {}),
        )
        # Intentar obtener credenciales para leer headers reales
        _creds_xl: dict | None = None
        try:
            _creds_xl = dict(st.secrets["gcp_service_account"])
        except Exception:
            pass

        _fecha_xl = pdata["fecha"]
        _fname_xl = (
            f"fx_Picking_{_fecha_xl.strftime('%d-%m-%Y')}.xlsx"
            if hasattr(_fecha_xl, "strftime")
            else "fx_Picking.xlsx"
        )
        try:
            _xlsx_bytes = _build_fx_picking_xlsx(
                totales_bult        = totales_bult,
                totales_hl          = totales_hl,
                totales_kg          = totales_kg,
                bult_ae_por_cancha  = pdata.get("bult_ae_por_cancha", {}),
                hl_venta_por_cancha = _hl_venta_xl,
                fecha               = _fecha_xl,
                credentials_json    = _creds_xl,
            )
            st.download_button(
                label          = f"⬇️ Descargar {_fname_xl}",
                data           = _xlsx_bytes,
                file_name      = _fname_xl,
                mime           = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width = True,
                key            = "t4_paso2_xlsx_dl",
                help           = "Una sola fila con fecha en col A y datos B:BL — "
                                 "mismos headers que el sheet. "
                                 "Útil si Sheets no está disponible o como respaldo.",
            )
        except Exception as _ex_xl:
            st.warning(f"No se pudo generar el Excel: {_ex_xl}")

    # ── Paso 3 — Exportar líneas AE Pallets completos por SKU/Camión ─────────
    st.divider()
    with st.container(border=True):
        st.markdown("#### 🏗️ Paso 3 — Paletas completas AE (por SKU × Camión)")
        st.caption(
            "Exporta el detalle de paletas completas que carga el autoelevador por SKU por camión. "
            "Fuente: hoja AGR del CAR — `floor(bultos / BXP)` por SKU × camión. "
            "Los camiones excluidos arriba figuran con 0 en todas sus columnas."
        )

        try:
            _car_p3 = st.session_state.get("t1_car") or st.session_state.get("t4_car")
            _fr_p3  = st.session_state.get("t1_fr")
            _ddm_p3 = st.session_state.get("df_ddm_dict", {})

            if not (_car_p3 and _fr_p3):
                st.info("Necesitás CAR.xlsx y Frescura 3.0 cargados.")
            else:
                # Leer hoja AGR del CAR
                _car_p3.seek(0)
                _xl_p3 = pd.ExcelFile(_car_p3)
                _agr_sheet = next((s for s in _xl_p3.sheet_names if s.upper() == "AGR"), None)
                if _agr_sheet is None:
                    st.warning("No se encontró hoja AGR en el CAR.")
                else:
                    # ── Detección inteligente de header (igual que load_agr) ──
                    _car_p3.seek(0)
                    _raw_p3 = pd.read_excel(_car_p3, sheet_name=_agr_sheet, header=None)
                    _car_p3.seek(0)
                    _hdr_row_p3 = None
                    for _hi in range(min(15, len(_raw_p3))):
                        _row_s = " ".join(str(v) for v in _raw_p3.iloc[_hi].values if pd.notna(v))
                        if ("chofer" in _row_s.lower() or "camion" in _row_s.lower()) and \
                           ("cantidad" in _row_s.lower() or "bulto" in _row_s.lower()):
                            _hdr_row_p3 = _hi
                            break
                    if _hdr_row_p3 is None:
                        # fallback: buscar cualquier fila con cod producto / descripcion
                        for _hi in range(min(15, len(_raw_p3))):
                            _row_s = " ".join(str(v) for v in _raw_p3.iloc[_hi].values if pd.notna(v))
                            if "descrip" in _row_s.lower() and ("cod" in _row_s.lower() or "art" in _row_s.lower()):
                                _hdr_row_p3 = _hi
                                break
                    if _hdr_row_p3 is None:
                        _hdr_row_p3 = 0

                    _car_p3.seek(0)
                    _df_agr_p3 = pd.read_excel(_car_p3, sheet_name=_agr_sheet, header=_hdr_row_p3)
                    _df_agr_p3.columns = [str(c).strip() for c in _df_agr_p3.columns]
                    _car_p3.seek(0)

                    # Identificar columnas clave
                    _cols_p3 = list(_df_agr_p3.columns)
                    def _fc3(*names):
                        for n in names:
                            for c in _cols_p3:
                                if n.lower() in c.lower():
                                    return c
                        return None

                    _col_sku3   = _fc3("cod producto", "codproducto", "codigo producto", "almac", "artículo", "articulo", "codigo", "sku")
                    _col_desc3  = _fc3("descrip")
                    _col_bult3  = _fc3("cantidad", "bultos")
                    _col_cam3   = _fc3("chofer", "camion", "cam")

                    if not all([_col_sku3, _col_bult3, _col_cam3]):
                        st.warning(f"No se pudieron identificar columnas AGR. Cols: {_cols_p3}")
                    else:
                        _df_agr_p3 = _df_agr_p3.dropna(subset=[_col_sku3, _col_bult3, _col_cam3])

                        # Obtener lista de todos los camiones con picking (incluyendo excluidos)
                        _all_cams_p3 = sorted(df_display_base["Camión"].astype(int).tolist())
                        _excluidos_p3 = set(st.session_state.get("t4_cams_excluir", []))

                        # Construir pivot: SKU × Camión → paletas completas
                        _p3_rows = []
                        _skus_sin_ddm = []
                        for _, _rp3 in _df_agr_p3.iterrows():
                            try:
                                _sku3  = int(float(str(_rp3[_col_sku3])))
                                _bult3 = float(_rp3[_col_bult3])
                                _cam3  = int(float(str(_rp3[_col_cam3])))
                                _bxp3  = _ddm_p3.get(_sku3, {}).get("bxp", 0)
                                if _bxp3 <= 0:
                                    if _sku3 not in _skus_sin_ddm:
                                        _skus_sin_ddm.append(_sku3)
                                    continue
                                _pall_comp = int(_bult3 // _bxp3)
                                if _pall_comp <= 0:
                                    continue
                                _p3_rows.append({
                                    "SKU": _sku3,
                                    "DESCRIPCION": str(_rp3[_col_desc3]).strip()[:40] if _col_desc3 else str(_sku3),
                                    "CAMION": _cam3,
                                    "BULTOS": _bult3,
                                    "BXP": _bxp3,
                                    "PALL_COMP": _pall_comp,
                                })
                            except Exception:
                                pass

                        if not _p3_rows:
                            st.info("No se encontraron paletas completas en el AGR.")
                        else:
                            _df_p3 = pd.DataFrame(_p3_rows)
                            # Pivot: SKU/Desc como filas, camiones como columnas
                            _pivot_p3 = _df_p3.pivot_table(
                                index=["SKU","DESCRIPCION"],
                                columns="CAMION",
                                values="PALL_COMP",
                                aggfunc="sum",
                                fill_value=0,
                            ).reset_index()

                            # Asegurar que todos los camiones con picking están presentes
                            for _cam_all in _all_cams_p3:
                                if _cam_all not in _pivot_p3.columns:
                                    _pivot_p3[_cam_all] = 0

                            # Aplicar exclusión: camiones excluidos → 0
                            for _cam_exc in _excluidos_p3:
                                if _cam_exc in _pivot_p3.columns:
                                    _pivot_p3[_cam_exc] = 0

                            # Ordenar columnas: SKU, DESC, camiones ordenados, TOTAL
                            _cam_cols_p3 = sorted([c for c in _pivot_p3.columns
                                                   if isinstance(c, (int, float))])
                            _pivot_p3["TOTAL"] = _pivot_p3[_cam_cols_p3].sum(axis=1)

                            # Agregar FEFO y STOCK desde Frescura DDM
                            _frescura_fefo = {}
                            _frescura_stk  = {}
                            try:
                                _fr_p3.seek(0)
                                import openpyxl as _ox_p3
                                _wb_p3 = _ox_p3.load_workbook(_fr_p3, read_only=True, data_only=True)
                                _fr_p3.seek(0)
                                if "Frescura" in _wb_p3.sheetnames:
                                    _ws_p3 = _wb_p3["Frescura"]
                                    _rws_p3 = list(_ws_p3.iter_rows(min_row=1, values_only=True))
                                    if _rws_p3:
                                        _h_p3 = [str(v).strip().upper() if v else "" for v in _rws_p3[0]]
                                        _isku_p3  = next((i for i,h in enumerate(_h_p3) if "ALM" in h or "COD" in h), 1)
                                        _idate_p3 = next((i for i,h in enumerate(_h_p3) if "FECHA" in h), 5)
                                        _istk_p3  = next((i for i,h in enumerate(_h_p3) if "STOCK" in h and "TOTAL" in h), 14)
                                        for _rp3f in _rws_p3[1:]:
                                            try:
                                                _s3 = int(float(str(_rp3f[_isku_p3]))) if _rp3f[_isku_p3] else 0
                                                if not _s3: continue
                                                _f3 = _rp3f[_idate_p3]
                                                if hasattr(_f3, "strftime"):
                                                    _f3_str = _f3.strftime("%d/%m/%Y")
                                                elif _f3:
                                                    _f3_str = str(_f3)
                                                else:
                                                    _f3_str = ""
                                                _stk3 = float(_rp3f[_istk_p3]) if _istk_p3 < len(_rp3f) and _rp3f[_istk_p3] else 0
                                                if _s3 not in _frescura_fefo:
                                                    _frescura_fefo[_s3] = _f3_str
                                                    _frescura_stk[_s3]  = _stk3
                                            except Exception:
                                                pass
                                _wb_p3.close()
                            except Exception:
                                pass

                            _pivot_p3["F.E.FO"] = _pivot_p3["SKU"].map(lambda s: _frescura_fefo.get(int(s), "—"))
                            _pivot_p3["STOCK"] = _pivot_p3["SKU"].map(lambda s: _frescura_stk.get(int(s), 0))

                            # Agregar Bs x p (BXP) desde DDM
                            _pivot_p3["Bs x p"] = _pivot_p3["SKU"].map(
                                lambda s: _ddm_p3.get(int(s), {}).get("bxp", 0))

                            # Reordenar: Cod | Descripcion | Bs x p | camiones... | TOTALES | F.E.FO | STOCK
                            _final_cols_p3 = ["SKU", "DESCRIPCION", "Bs x p"] + _cam_cols_p3 + ["TOTAL", "F.E.FO", "STOCK"]
                            _pivot_p3 = _pivot_p3[[c for c in _final_cols_p3 if c in _pivot_p3.columns]]
                            # Filtrar solo filas con TOTAL > 0
                            _pivot_p3 = _pivot_p3[_pivot_p3["TOTAL"] > 0].reset_index(drop=True)

                            # Renombrar columnas para display/export (igual imagen referencia)
                            _rename_p3 = {"SKU": "Cod", "DESCRIPCION": "Descripcion", "TOTAL": "TOTALES"}
                            _pivot_p3_disp = _pivot_p3.rename(columns=_rename_p3)

                            if _skus_sin_ddm:
                                st.warning(f"⚠️ SKUs sin BXP en DDM (excluidos): {_skus_sin_ddm}")

                            st.dataframe(_pivot_p3_disp, use_container_width=True, hide_index=True)

                            # Exportar Excel
                            try:
                                import io as _io_p3, openpyxl as _opx_p3
                                from openpyxl.styles import (PatternFill as _PF3, Font as _FO3,
                                                             Alignment as _AL3)
                                _wb3 = _opx_p3.Workbook()
                                _ws3 = _wb3.active; _ws3.title = "AE_Pallets"
                                # Header con nombres renombrados
                                _hdrs3 = list(_pivot_p3_disp.columns)
                                for ci3, h3 in enumerate(_hdrs3, 1):
                                    _c3 = _ws3.cell(1, ci3, str(h3))
                                    _c3.fill = _PF3("solid", fgColor="1F3864")
                                    _c3.font = _FO3(bold=True, color="FFFFFF", size=10)
                                    _c3.alignment = _AL3(horizontal="center")
                                # Datos
                                for ri3, row3 in _pivot_p3_disp.iterrows():
                                    for ci3, v3 in enumerate(row3, 1):
                                        _c3d = _ws3.cell(ri3+2, ci3, v3)
                                        _c3d.alignment = _AL3(horizontal="center")
                                        if ri3 % 2 == 0:
                                            _c3d.fill = _PF3("solid", fgColor="EEF2FF")
                                # Autowidth
                                for _col3 in _ws3.columns:
                                    _ws3.column_dimensions[_col3[0].column_letter].width = 12
                                _ws3.column_dimensions["B"].width = 30  # descripcion
                                _buf3 = _io_p3.BytesIO()
                                _wb3.save(_buf3); _buf3.seek(0)
                                _fecha_p3 = pdata["fecha"].strftime("%d-%m-%Y") if hasattr(pdata["fecha"],"strftime") else "hoy"
                                st.download_button(
                                    "⬇️ Descargar Excel AE Pallets",
                                    data=_buf3.getvalue(),
                                    file_name=f"{_fecha_p3}_AE_pallets_bkcc.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key="t4_paso3_dl",
                                )
                            except Exception as _ep3x:
                                st.warning(f"⚠️ No se pudo generar Excel: {_ep3x}")


        except Exception as _ep3:
            st.error(f"❌ Error en Paso 3: {_ep3}")
            with st.expander("Stack trace"):
                import traceback
                st.code(traceback.format_exc())

    # ── Paso 4 — Reposición de canchas de picking ─────────────────────────────
    st.divider()
    with st.container(border=True):
        st.markdown("#### 🔄 Paso 4 — Reposición de canchas de picking")
        st.caption(
            "Calcula qué SKUs necesitan reposición en cancha. "
            "Fuente: Frescura (posiciones × BXP = bultos en cancha) vs picking total (ANR + sub-pallet AGR). "
            "Reposición negativa = hay que traer bultos del depósito. "
            "**⚠️ Clave de seguridad: revisá antes de que arranque el picking.**"
        )

        # Posiciones especiales por SKU (sobreescriben pos=1 del DDM)
        _POSICIONES_ESPECIALES = {7038: 7, 7634: 2, 31434: 2, 19019: 2}

        try:
            _car_p4  = st.session_state.get("t1_car") or st.session_state.get("t4_car")
            _anr_p4  = st.session_state.get("tc_anr")
            _fr_p4   = st.session_state.get("t1_fr")
            _ddm_p4  = st.session_state.get("df_ddm_dict", {})
            _anr_df4 = st.session_state.get("anr_df")

            if not (_car_p4 and _fr_p4):
                st.info("Necesitás CAR.xlsx y Frescura 3.0 cargados. ANR opcional pero recomendado.")
            else:
                # ── 1. PICK desde ANR (hoja BASE, picking normal) ───────────
                _pick_anr = {}  # {sku: bultos}
                if _anr_df4 is not None and not _anr_df4.empty:
                    _ac = [str(c).strip() for c in _anr_df4.columns]
                    def _fc4(*nms):
                        for n in nms:
                            for i, c in enumerate(_ac):
                                if n.lower() in c.lower(): return _anr_df4.columns[i]
                        return None
                    _col_sku4  = _fc4("almac","artíc","cod")
                    _col_bult4 = _fc4("bultos","cantidad","cant")
                    _col_sc4   = _fc4("sin cargo","sincarg")
                    if _col_sku4 and _col_bult4:
                        _df4 = _anr_df4.copy()
                        if _col_sc4:
                            _df4 = _df4[~_df4[_col_sc4].astype(str).str.upper().str.strip().isin(["SI","SÍ","S","1","TRUE"])]
                        for _, _ra4 in _df4.iterrows():
                            try:
                                _s4 = int(float(str(_ra4[_col_sku4])))
                                _b4 = float(_ra4[_col_bult4])
                                _pick_anr[_s4] = _pick_anr.get(_s4, 0) + _b4
                            except Exception:
                                pass

                # ── 2. Sub-pallet desde AGR (CAR) ──────────────────────────
                _pick_agr_sub = {}  # {sku: bultos sub-pallet}
                try:
                    _car_p4.seek(0)
                    _xl_p4 = pd.ExcelFile(_car_p4)
                    _agr4 = next((s for s in _xl_p4.sheet_names if s.upper() == "AGR"), None)
                    if _agr4:
                        _car_p4.seek(0)
                        _df_agr4 = pd.read_excel(_car_p4, sheet_name=_agr4, header=0)
                        _car_p4.seek(0)
                        _cols4 = [str(c).strip() for c in _df_agr4.columns]
                        def _fca4(*nms):
                            for n in nms:
                                for i, c in enumerate(_cols4):
                                    if n.lower() in c.lower(): return _df_agr4.columns[i]
                            return None
                        _csku4  = _fca4("cod","almac","artíc")
                        _cblt4  = _fca4("cantidad","bultos")
                        _ccam4  = _fca4("chofer","camion")
                        if _csku4 and _cblt4:
                            for _, _rr4 in _df_agr4.dropna(subset=[_csku4,_cblt4]).iterrows():
                                try:
                                    _s4  = int(float(str(_rr4[_csku4])))
                                    _b4  = float(_rr4[_cblt4])
                                    _bxp4 = _ddm_p4.get(_s4, {}).get("bxp", 0)
                                    if _bxp4 > 0:
                                        _sub4 = _b4 % _bxp4  # solo fracción sub-pallet
                                        if _sub4 > 0:
                                            _pick_agr_sub[_s4] = _pick_agr_sub.get(_s4, 0) + _sub4
                                except Exception:
                                    pass
                except Exception:
                    pass

                # ── 3. Leer posiciones y cancha desde Frescura DDM ──────────
                # Usamos df_ddm_dict que ya está en session_state
                # La cancha la tomamos de col O (ya en df_ddm_dict["can"])
                # Posiciones = 1 por defecto (1 pallet en cancha) salvo especiales
                # En cancha = posiciones × BXP

                # ── 4. Construir tabla reposición ───────────────────────────
                _repos_rows = []
                import datetime as _dt_p4
                for _sku4, _ddm_v4 in _ddm_p4.items():
                    _bxp4  = _ddm_v4.get("bxp", 0)
                    _can4  = _ddm_v4.get("can", "")
                    if _bxp4 <= 0 or not _can4:
                        continue
                    # Posiciones en cancha
                    _pos4 = _POSICIONES_ESPECIALES.get(_sku4, 1)
                    _en_cancha4 = _pos4 * _bxp4
                    # PICK total = ANR + sub-pallet AGR
                    _pick_tot4 = _pick_anr.get(_sku4, 0) + _pick_agr_sub.get(_sku4, 0)
                    if _pick_tot4 <= 0:
                        continue  # si no hay picking, no hay reposición
                    _repos4 = _en_cancha4 - _pick_tot4
                    # Solo mostrar los que necesitan reposición (negativo)
                    _repos_rows.append({
                        "Almacén": _sku4,
                        "Descripción": "",  # se podría enriquecer con ANR
                        "Cancha": _can4,
                        "BXP": int(_bxp4),
                        "Posiciones": _pos4,
                        "En cancha (bult)": int(_en_cancha4),
                        "PICK (bult)": round(_pick_tot4, 2),
                        "Reposición": round(_repos4, 2),
                        "¿Reponer?": "⚠️ SÍ" if _repos4 < 0 else "✅ OK",
                    })

                if not _repos_rows:
                    st.info("No hay datos suficientes para calcular reposiciones. Verificá que ANR esté cargado.")
                else:
                    _df_repos = pd.DataFrame(_repos_rows)
                    _df_repos_neg = _df_repos[_df_repos["Reposición"] < 0].sort_values("Reposición")
                    _df_repos_ok  = _df_repos[_df_repos["Reposición"] >= 0]

                    _r4c1, _r4c2, _r4c3 = st.columns(3)
                    _r4c1.metric("🔄 SKUs a reponer", len(_df_repos_neg))
                    _r4c2.metric("✅ SKUs OK", len(_df_repos_ok))
                    _r4c3.metric("📦 Total SKUs picking", len(_df_repos))

                    if not _df_repos_neg.empty:
                        st.error(f"⚠️ **{len(_df_repos_neg)} SKU(s) necesitan reposición en cancha antes del picking:**")
                        st.dataframe(_df_repos_neg, use_container_width=True, hide_index=True)
                    else:
                        st.success("✅ Todas las canchas tienen stock suficiente para el picking de hoy.")

                    with st.expander("Ver todos los SKUs (incluyendo OK)", expanded=False):
                        st.dataframe(_df_repos.sort_values("Reposición"), use_container_width=True, hide_index=True)

                    # Exportar Excel reposición
                    try:
                        import io as _io_r4, openpyxl as _opx_r4
                        from openpyxl.styles import PatternFill as _PFr4, Font as _FOr4, Alignment as _ALr4
                        _wbr4 = _opx_r4.Workbook()
                        _wsr4 = _wbr4.active; _wsr4.title = "Reposición Cancha"
                        _hdrs_r4 = list(_df_repos_neg.columns) if not _df_repos_neg.empty else list(_df_repos.columns)
                        _df_exp_r4 = _df_repos_neg if not _df_repos_neg.empty else _df_repos
                        for ci_r4, h_r4 in enumerate(_hdrs_r4, 1):
                            _cr4 = _wsr4.cell(1, ci_r4, str(h_r4))
                            _cr4.fill = _PFr4("solid", fgColor="1F3864")
                            _cr4.font = _FOr4(bold=True, color="FFFFFF", size=10)
                            _cr4.alignment = _ALr4(horizontal="center")
                        for ri_r4, row_r4 in _df_exp_r4.iterrows():
                            for ci_r4, v_r4 in enumerate(row_r4, 1):
                                _cr4d = _wsr4.cell(ri_r4 + 2, ci_r4, v_r4)
                                _cr4d.alignment = _ALr4(horizontal="center")
                                if str(v_r4) == "⚠️ SÍ":
                                    _cr4d.fill = _PFr4("solid", fgColor="FFE0E0")
                        for _colr4 in _wsr4.columns:
                            _wsr4.column_dimensions[_colr4[0].column_letter].width = 14
                        _bufr4 = _io_r4.BytesIO(); _wbr4.save(_bufr4); _bufr4.seek(0)
                        _fecha_r4 = pdata["fecha"].strftime("%d-%m-%Y") if hasattr(pdata["fecha"],"strftime") else "hoy"
                        st.download_button(
                            "⬇️ Descargar Excel Reposición Cancha",
                            data=_bufr4.getvalue(),
                            file_name=f"{_fecha_r4}_reposicion_cancha_bkcc.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="t4_paso4_dl",
                        )
                    except Exception as _er4x:
                        st.warning(f"⚠️ No se pudo generar Excel reposición: {_er4x}")

        except Exception as _ep4:
            st.error(f"❌ Error en Paso 4: {_ep4}")
            with st.expander("Stack trace"):
                import traceback
                st.code(traceback.format_exc())


# ── HELPER: (fx) Picking — Paso 2 ─────────────────────────────────────────
# v4.65 — Calcula HL venta por cancha desde ANR (excluyendo SIN CARGO)
# y escribe la fila resumen en la hoja (fx) Picking del Inputs Sheets.

def _t4_compute_hl_venta_por_cancha(
    anr_df: "pd.DataFrame",
    ddm_can_lookup: dict,
) -> dict:
    """
    v4.65 — Suma TOTAL HL VENTA por cancha desde el ANR (hoja BASE).

    Lógica:
      - Col O (idx 14 = UM / HL) contiene los HL de cada línea de venta.
      - Se excluyen filas con SIN CARGO = 'SI' (col G, idx 6).
      - Cada SKU (col H, idx 7) se mapea a su cancha via ddm_can_lookup.
      - La suma por cancha es TOTAL HL VENTA (no es el HL del picking).

    Retorna {cancha: float} para T4_CANCHAS + 'SIN CANCHA'.
    """
    result = {c: 0.0 for c in _T4_CANCHAS}

    if anr_df is None or anr_df.empty:
        return result

    # Nombres de columna flexibles — buscar por nombre o por posición
    def _col(df, *names):
        for n in names:
            if n in df.columns:
                return df[n]
        return None

    col_sku      = _col(anr_df, "ARTÍCULO", "ARTICULO", "H")
    col_hl       = _col(anr_df, "UNIDAD DE MEDIDA", "UM", "O")
    col_sin_cargo = _col(anr_df, "SIN CARGO", "G")

    if col_hl is None:
        # Fallback posicional (col O = idx 14 en BASE, que puede tener header en row 2)
        if len(anr_df.columns) > 14:
            col_hl = anr_df.iloc[:, 14]
        else:
            return result

    if col_sku is None and len(anr_df.columns) > 7:
        col_sku = anr_df.iloc[:, 7]

    if col_sin_cargo is None and len(anr_df.columns) > 6:
        col_sin_cargo = anr_df.iloc[:, 6]

    for i in range(len(anr_df)):
        try:
            # Excluir SIN CARGO
            sc = str(col_sin_cargo.iloc[i]).strip().upper() if col_sin_cargo is not None else ""
            if sc == "SI":
                continue

            hl_val = col_hl.iloc[i]
            hl_f   = float(hl_val) if hl_val not in (None, "", "None") else 0.0
            if hl_f <= 0:
                continue

            sku_raw = col_sku.iloc[i] if col_sku is not None else None
            sku_int = int(float(str(sku_raw))) if sku_raw not in (None, "", "None") else 0
            can_raw = ddm_can_lookup.get(sku_int, "SIN CANCHA")
            can_norm = _t4_norm_cancha(can_raw)

            if can_norm in _T4_CANCHAS:
                result[can_norm] += hl_f
        except (ValueError, TypeError, IndexError):
            pass

    return result


def _push_fx_picking_to_sheets(
    *,
    totales_bult:      dict,
    totales_hl:        dict,
    totales_kg:        dict,
    bult_ae_por_cancha: dict,
    hl_venta_por_cancha: dict,
    tot_pick:          float,
    tot_ae:            float,
    fecha,
    credentials_json:  dict,
) -> str:
    """
    v4.65 — Escribe UNA SOLA FILA en la hoja '(fx) Picking' del Inputs Sheet,
    buscando la fila del día actual por fecha en col A y completando B:BL.

    NO inserta filas nuevas — solo hace update() sobre la fila ya generada
    automáticamente por la fórmula del sheet. Así nunca rompe fórmulas.

    Columnas B..BL (63 valores, 0-indexados en la fila):
      B  = PICKING BULTOS total   (sum todas las canchas pick)
      C..G = CI CII CIII CIV MKPL (pick)
      H  = Comodín 1 = 0
      I  = AE BULTOS total        (sum AE todas las canchas)
      J..N = CI CII CIII CIV MKPL (AE)
      O  = Comodín 1 = 0
      P  = PESO KG total           (pick + AE)
      Q..U = CI CII CIII CIV MKPL (KG pick, AE no tiene columna propia)
      V  = Comodín 1 = 0
      W  = PICKING HL total        (sum HL pick todas las canchas)
      X..AB = CI CII CIII CIV MKPL (HL pick)
      AC = Comodín 1 = 0
      AD = TOTAL HL VENTA total    (sum ANR todas las canchas)
      AE..AI = CI CII CIII CIV MKPL (HL venta ANR)
      AJ = Comodín 1 = 0
      AK..BL = MATRICES → todas en 0 (44 ceros)
    """
    import gspread
    from google.oauth2.service_account import Credentials
    import datetime as _dt

    SPREADSHEET_ID = "1OjIhtpzwV-MpnBWu09lKyguit2iKR-nCFG9g5MOd-mM"
    SHEET_NAME     = "(fx) Picking"
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds  = Credentials.from_service_account_info(credentials_json, scopes=SCOPES)
    client = gspread.authorize(creds)
    sheet  = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

    # Fecha a buscar en col A
    if hasattr(fecha, "strftime"):
        fecha_dt = fecha
    elif hasattr(fecha, "year"):
        fecha_dt = fecha
    else:
        fecha_dt = _dt.date.fromisoformat(str(fecha))

    # ── Normalizar fecha del día al formato "d-mmm" que usa el sheet ──────
    # gspread devuelve las celdas con fórmulas =Picking!Axxx como strings
    # con formato "d-mmm" en español (ej: "3-jun", "15-may", "1-ene").
    # Generamos ese string y también variantes comunes como fallback.
    _MESES_ES = {
        1: "ene", 2: "feb", 3: "mar", 4: "abr",
        5: "may", 6: "jun", 7: "jul", 8: "ago",
        9: "sep", 10: "oct", 11: "nov", 12: "dic",
    }
    _dia  = fecha_dt.day if hasattr(fecha_dt, "day") else _dt.date.today().day
    _mes  = fecha_dt.month if hasattr(fecha_dt, "month") else _dt.date.today().month
    _anio = fecha_dt.year if hasattr(fecha_dt, "year") else _dt.date.today().year

    # Formato principal que devuelve gspread: "3-jun"
    _fmt_dmmm    = f"{_dia}-{_MESES_ES[_mes]}"       # "3-jun"
    _fmt_ddmmm   = f"{_dia:02d}-{_MESES_ES[_mes]}"   # "03-jun"

    fecha_targets = {
        _fmt_dmmm,
        _fmt_ddmmm,
        # Fallbacks adicionales por si el formato del sheet varía
        f"{_dia:02d}/{_mes:02d}/{_anio}",             # "03/06/2026"
        f"{_dia}/{_mes}/{_anio}",                      # "3/6/2026"
        str(fecha_dt),                                  # "2026-06-03"
        f"{_dia:02d}/{_mes:02d}/{str(_anio)[2:]}",     # "03/06/26"
    }

    # Leer col A completa para encontrar la fila del día
    col_a = sheet.col_values(1)  # 1-indexed, returns list desde row 1
    target_row = None
    _debug_vals = []
    for i, val in enumerate(col_a):
        val_clean = str(val).strip().lower()
        if val_clean:  # ignorar celdas vacías
            _debug_vals.append(f"row{i+1}={val_clean!r}")
        # Comparar en minúsculas para "3-jun" == "3-JUN" etc.
        if val_clean in {t.lower() for t in fecha_targets}:
            target_row = i + 1  # gspread usa 1-indexed
            break

    if target_row is None:
        raise ValueError(
            f"No se encontró la fecha {_fmt_dmmm!r} (ni variantes) en la col A "
            f"de la hoja '(fx) Picking'. "
            f"Verificá que la fórmula del sheet ya generó la fila de hoy antes de "
            f"ejecutar este paso.\n"
            f"Valores buscados: {sorted(fecha_targets)}\n"
            f"Últimas 10 celdas con dato: {_debug_vals[-10:]}"
        )

    # ── Construir los 63 valores (B..BL) ───────────────────────────────────
    C = _T4_CANCHAS  # ["CANCHA I", "CANCHA II", "CANCHA III", "CANCHA IV", "MKPL"]

    def _r(v, dec=3):
        try: return round(float(v), dec)
        except: return 0.0

    # Totales globales
    pick_total = sum(_r(totales_bult.get(c, 0)) for c in C)
    ae_total   = sum(_r(bult_ae_por_cancha.get(c, 0)) for c in C)
    kg_total   = sum(_r(totales_kg.get(c, 0)) for c in C)     # KG es pick only (AE incluido)
    hl_pick_total  = sum(_r(totales_hl.get(c, 0)) for c in C)
    hl_venta_total = sum(_r(hl_venta_por_cancha.get(c, 0)) for c in C)

    row_values = [
        # B — PICKING BULTOS
        _r(pick_total, 2),
        # C..G — CI CII CIII CIV MKPL (pick bultos)
        _r(totales_bult.get("CANCHA I",   0), 2),
        _r(totales_bult.get("CANCHA II",  0), 2),
        _r(totales_bult.get("CANCHA III", 0), 2),
        _r(totales_bult.get("CANCHA IV",  0), 2),
        _r(totales_bult.get("MKPL",       0), 2),
        # H — Comodín 1 = 0
        0,
        # I — AE BULTOS total
        _r(ae_total, 2),
        # J..N — CI CII CIII CIV MKPL (AE bultos)
        _r(bult_ae_por_cancha.get("CANCHA I",   0), 2),
        _r(bult_ae_por_cancha.get("CANCHA II",  0), 2),
        _r(bult_ae_por_cancha.get("CANCHA III", 0), 2),
        _r(bult_ae_por_cancha.get("CANCHA IV",  0), 2),
        _r(bult_ae_por_cancha.get("MKPL",       0), 2),
        # O — Comodín 1 = 0
        0,
        # P — PESO KG total
        _r(kg_total, 2),
        # Q..U — CI CII CIII CIV MKPL (KG)
        _r(totales_kg.get("CANCHA I",   0), 2),
        _r(totales_kg.get("CANCHA II",  0), 2),
        _r(totales_kg.get("CANCHA III", 0), 2),
        _r(totales_kg.get("CANCHA IV",  0), 2),
        _r(totales_kg.get("MKPL",       0), 2),
        # V — Comodín 1 = 0
        0,
        # W — PICKING HL total
        _r(hl_pick_total, 3),
        # X..AB — CI CII CIII CIV MKPL (HL pick)
        _r(totales_hl.get("CANCHA I",   0), 3),
        _r(totales_hl.get("CANCHA II",  0), 3),
        _r(totales_hl.get("CANCHA III", 0), 3),
        _r(totales_hl.get("CANCHA IV",  0), 3),
        _r(totales_hl.get("MKPL",       0), 3),
        # AC — Comodín 1 = 0
        0,
        # AD — TOTAL HL VENTA total
        _r(hl_venta_total, 3),
        # AE..AI — CI CII CIII CIV MKPL (HL venta ANR)
        _r(hl_venta_por_cancha.get("CANCHA I",   0), 3),
        _r(hl_venta_por_cancha.get("CANCHA II",  0), 3),
        _r(hl_venta_por_cancha.get("CANCHA III", 0), 3),
        _r(hl_venta_por_cancha.get("CANCHA IV",  0), 3),
        _r(hl_venta_por_cancha.get("MKPL",       0), 3),
        # AJ — Comodín 1 = 0
        0,
        # AK..BL — MATRICES (44 ceros: 4 bloques × 7 cols × … = 28 restantes)
        # Bloque MATRIZ PICKING HL AGREGADOS: total + CI CII CIII CIV MKPL + Comodín
        0, 0, 0, 0, 0, 0, 0,
        # Bloque MATRIZ TODO HL AGREGADOS: total + CI CII CIII CIV MKPL + Comodín
        0, 0, 0, 0, 0, 0, 0,
        # Bloque MATRIZ TODO BULTOS PICKING (1): total + CI CII CIII CIV MKPL + Comodín
        0, 0, 0, 0, 0, 0, 0,
        # Bloque MATRIZ TODO BULTOS PICKING (2): total + CI CII CIII CIV MKPL + Comodín
        0, 0, 0, 0, 0, 0, 0,
    ]

    assert len(row_values) == 63, f"Error interno: se esperaban 63 valores, hay {len(row_values)}"

    # ── Escribir en la fila encontrada, columnas B:BL ──────────────────────
    # Rango: B{target_row}:BL{target_row}
    range_notation = f"B{target_row}:BL{target_row}"
    sheet.update(
        range_name=range_notation,
        values=[row_values],
        value_input_option="USER_ENTERED",
    )

    return f"Fila {target_row} actualizada (fecha {_fmt_dmmm!r} encontrada en col A, {len(row_values)} valores escritos en B:BL)"


def _build_fx_picking_xlsx(
    *,
    totales_bult:       dict,
    totales_hl:         dict,
    totales_kg:         dict,
    bult_ae_por_cancha: dict,
    hl_venta_por_cancha: dict,
    fecha,
    credentials_json:   dict | None = None,
) -> bytes:
    """
    v4.66 — Genera un .xlsx con UNA fila de datos de (fx) Picking:
      - Col A: fecha formateada "d-mmm" (mismo formato que el sheet)
      - Cols B:BL: los 63 valores calculados
    Los headers se leen desde la fila 1 del sheet si hay credenciales,
    o se usan etiquetas genéricas como fallback.
    """
    import io as _io
    import datetime as _dt2

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        # Fallback sin estilos
        import openpyxl

    # ── Fecha ─────────────────────────────────────────────────────────────
    if hasattr(fecha, "day"):
        fecha_dt = fecha
    else:
        fecha_dt = _dt2.date.fromisoformat(str(fecha))

    _MESES_ES2 = {
        1: "ene", 2: "feb", 3: "mar", 4: "abr",
        5: "may", 6: "jun", 7: "jul", 8: "ago",
        9: "sep", 10: "oct", 11: "nov", 12: "dic",
    }
    _fecha_str = f"{fecha_dt.day}-{_MESES_ES2[fecha_dt.month]}"

    # ── Construir los 63 valores (misma lógica que _push_fx_picking_to_sheets) ──
    C = _T4_CANCHAS

    def _rv(v, dec=3):
        try: return round(float(v), dec)
        except: return 0.0

    pick_total     = sum(_rv(totales_bult.get(c, 0)) for c in C)
    ae_total       = sum(_rv(bult_ae_por_cancha.get(c, 0)) for c in C)
    kg_total       = sum(_rv(totales_kg.get(c, 0)) for c in C)
    hl_pick_total  = sum(_rv(totales_hl.get(c, 0)) for c in C)
    hl_venta_total = sum(_rv(hl_venta_por_cancha.get(c, 0)) for c in C)

    row_values = [
        _rv(pick_total, 2),
        _rv(totales_bult.get("CANCHA I",   0), 2),
        _rv(totales_bult.get("CANCHA II",  0), 2),
        _rv(totales_bult.get("CANCHA III", 0), 2),
        _rv(totales_bult.get("CANCHA IV",  0), 2),
        _rv(totales_bult.get("MKPL",       0), 2),
        0,
        _rv(ae_total, 2),
        _rv(bult_ae_por_cancha.get("CANCHA I",   0), 2),
        _rv(bult_ae_por_cancha.get("CANCHA II",  0), 2),
        _rv(bult_ae_por_cancha.get("CANCHA III", 0), 2),
        _rv(bult_ae_por_cancha.get("CANCHA IV",  0), 2),
        _rv(bult_ae_por_cancha.get("MKPL",       0), 2),
        0,
        _rv(kg_total, 2),
        _rv(totales_kg.get("CANCHA I",   0), 2),
        _rv(totales_kg.get("CANCHA II",  0), 2),
        _rv(totales_kg.get("CANCHA III", 0), 2),
        _rv(totales_kg.get("CANCHA IV",  0), 2),
        _rv(totales_kg.get("MKPL",       0), 2),
        0,
        _rv(hl_pick_total, 3),
        _rv(totales_hl.get("CANCHA I",   0), 3),
        _rv(totales_hl.get("CANCHA II",  0), 3),
        _rv(totales_hl.get("CANCHA III", 0), 3),
        _rv(totales_hl.get("CANCHA IV",  0), 3),
        _rv(totales_hl.get("MKPL",       0), 3),
        0,
        _rv(hl_venta_total, 3),
        _rv(hl_venta_por_cancha.get("CANCHA I",   0), 3),
        _rv(hl_venta_por_cancha.get("CANCHA II",  0), 3),
        _rv(hl_venta_por_cancha.get("CANCHA III", 0), 3),
        _rv(hl_venta_por_cancha.get("CANCHA IV",  0), 3),
        _rv(hl_venta_por_cancha.get("MKPL",       0), 3),
        0,
        0, 0, 0, 0, 0, 0, 0,
        0, 0, 0, 0, 0, 0, 0,
        0, 0, 0, 0, 0, 0, 0,
        0, 0, 0, 0, 0, 0, 0,
    ]
    assert len(row_values) == 63

    # ── Intentar leer headers reales desde el sheet ────────────────────────
    headers_sheet: list[str] = []
    if credentials_json:
        try:
            import gspread as _gs
            from google.oauth2.service_account import Credentials as _Creds
            _creds = _Creds.from_service_account_info(
                credentials_json,
                scopes=[
                    "https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive",
                ],
            )
            _cl  = _gs.authorize(_creds)
            _sh  = _cl.open_by_key("1OjIhtpzwV-MpnBWu09lKyguit2iKR-nCFG9g5MOd-mM")
            _ws  = _sh.worksheet("(fx) Picking")
            # Leer fila 1 completa (A:BL = 64 columnas)
            _row1 = _ws.row_values(1)
            headers_sheet = _row1[:64]  # A + B..BL
        except Exception:
            headers_sheet = []

    # Fallback si no pudimos leer headers
    if not headers_sheet or len(headers_sheet) < 64:
        headers_sheet = (
            ["Fecha de Picking", "PICKING", "CI pick", "CII pick", "CIII pick", "CIV pick", "MKPL pick",
             "Comodín1", "AE BULT", "CI AE", "CII AE", "CIII AE", "CIV AE", "MKPL AE",
             "Comodín2", "TOTAL HL", "CI HL", "CII HL", "CIII HL", "CIV HL", "MKPL HL",
             "Comodín3", "PICKING HL", "CI HL pick", "CII HL pick", "CIII HL pick", "CIV HL pick", "MKPL HL pick",
             "Comodín4", "TOTAL HL VENTA", "CI HL vta", "CII HL vta", "CIII HL vta", "CIV HL vta", "MKPL HL vta",
             "Comodín5"]
            + [f"MATRIZ_{i}" for i in range(28)]
        )
        # Rellenar hasta 64
        while len(headers_sheet) < 64:
            headers_sheet.append(f"Col_{len(headers_sheet)+1}")

    # ── Armar el workbook ──────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws2 = wb.active
    ws2.title = "(fx) Picking"

    # Estilos
    try:
        hdr_fill  = PatternFill("solid", fgColor="1F3864")
        hdr_font  = Font(color="FFFFFF", bold=True, size=9)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dat_align = Alignment(horizontal="center", vertical="center")
        a_fill    = PatternFill("solid", fgColor="FFF2CC")  # amarillo claro para col A
    except Exception:
        hdr_fill = hdr_font = hdr_align = dat_align = a_fill = None

    # Fila 1: headers
    full_row = [headers_sheet[0]] + headers_sheet[1:64]
    for col_idx, hdr in enumerate(full_row, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=hdr)
        if hdr_fill:
            cell.fill    = hdr_fill if col_idx > 1 else a_fill
            cell.font    = hdr_font if col_idx > 1 else Font(bold=True, size=9)
            cell.alignment = hdr_align
        ws2.row_dimensions[1].height = 30

    # Fila 2: datos
    data_row = [_fecha_str] + row_values
    for col_idx, val in enumerate(data_row, start=1):
        cell = ws2.cell(row=2, column=col_idx, value=val)
        if dat_align:
            cell.alignment = dat_align

    # Anchos de columna
    ws2.column_dimensions["A"].width = 14
    for col_letter in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        ws2.column_dimensions[col_letter].width = 10
    # BK, BL (columnas >26)
    for extra in range(27, 65):
        col_letter = openpyxl.utils.get_column_letter(extra)
        ws2.column_dimensions[col_letter].width = 10

    buf2 = _io.BytesIO()
    wb.save(buf2)
    return buf2.getvalue()


# ── HELPER: PDF Controlador de Depósito ────────────────────────────────────

def _t4_generar_pdf_controlador(
    fecha, fin_por_cancha, fin_global_dt,
    totales_bult, totales_hl, totales_kg, totales_pall_c,
    mix_picking, n_camiones, tot_pick,
    top_skus_lines="", alertas=None,
) -> bytes:
    """
    v4.41 — Genera PDF landscape A4 con el Resumen para el Controlador de Depósito.
    Una sola página, lista para imprimir junto a las 5 copias de canchas.
    Reutiliza _draw_controlador_page (el mismo helper que usa el PDF ×5).
    """
    if alertas is None:
        alertas = []

    buf   = io.BytesIO()
    LS    = landscape(A4)
    c     = rl_canvas.Canvas(buf, pagesize=LS)
    PW, PH = LS        # ~841 × 595 pt
    M     = 14 * mm

    ctrl_params = {
        "fin_por_cancha":  fin_por_cancha,
        "totales_bult":    totales_bult,
        "totales_pall_c":  totales_pall_c,
        "totales_hl":      totales_hl,
        "totales_kg":      totales_kg,
        "n_camiones":      n_camiones,
        "tot_pick":        tot_pick,
        "top_skus_lines":  top_skus_lines,
        "alertas":         alertas,
    }

    _draw_controlador_page(
        c, PW, PH, M, fin_global_dt,
        ctrl_params, mix_picking, fecha,
        DARK_BLUE=colors.HexColor("#1a3a6b"),
        MED_BLUE=colors.HexColor("#2e5fa3"),
        LIGHT=colors.HexColor("#F0F4FA"),
        ALT=colors.HexColor("#E8EEF8"),
        GREEN_OK=colors.HexColor("#1a7a4a"),
        RED_WARN=colors.HexColor("#b91c1c"),
        WHITE=colors.white,
        YELLOW_HL=colors.HexColor("#FFF176"),
        FONT_B="Helvetica-Bold",
        FONT_N="Helvetica",
    )

    c.save()
    buf.seek(0)
    return buf.read()


# ── HELPER: Top 10 SKUs ────────────────────────────────────────────────────

def _build_top10_skus(car_bytes: bytes, fr_bytes: bytes) -> pd.DataFrame:
    """
    Calcula el Top 10 SKUs más vendidos (por Bultos) del CAR del día,
    cruzando con la DDM de Frescura para obtener VALOR HL por bulto.

    Devuelve DataFrame con columnas:
      #  | CÓDIGO | DESCRIPCIÓN | BULTOS | HL | U_PAQUETE | RUBRO | ABC
    """
    import openpyxl as _ox

    # ── 1. Leer CAR ──────────────────────────────────────────────────────────
    wb_car = _ox.load_workbook(io.BytesIO(car_bytes), read_only=True, data_only=True)
    ws_car = wb_car.active
    rows_car = list(ws_car.iter_rows(values_only=True))
    wb_car.close()

    # Detectar fila header (buscar columna "Artículo" o "Código")
    header_idx = 0
    for i, row in enumerate(rows_car[:10]):
        vals = [str(v).strip().lower() if v else "" for v in row]
        if any("artículo" in v or "articulo" in v for v in vals):
            header_idx = i
            break
    headers_car = [str(v).strip() if v else f"col_{j}" for j, v in enumerate(rows_car[header_idx])]

    df_car = pd.DataFrame(rows_car[header_idx + 1:], columns=headers_car)
    df_car = df_car.dropna(how="all")

    # Mapear columnas clave
    col_art  = next((c for c in df_car.columns if "artículo" in c.lower() or "articulo" in c.lower()), None)
    col_desc = next((c for c in df_car.columns if "descripción artículo" in c.lower() or "descripcion articulo" in c.lower()), None)
    col_bul  = next((c for c in df_car.columns if "bultos" in c.lower()), None)
    col_uni  = next((c for c in df_car.columns if "unids" in c.lower() or "u paquete" in c.lower() or "unidades" in c.lower()), None)
    col_anu  = next((c for c in df_car.columns if "anulado" in c.lower()), None)

    if col_art is None or col_bul is None:
        raise ValueError(f"No se encontraron columnas clave en CAR. Columnas: {list(df_car.columns)}")

    # Filtrar anulados
    if col_anu:
        df_car = df_car[df_car[col_anu].astype(str).str.lower() != "yes"]

    df_car[col_art] = pd.to_numeric(df_car[col_art], errors="coerce")
    df_car[col_bul] = pd.to_numeric(df_car[col_bul], errors="coerce").fillna(0)
    if col_uni:
        df_car[col_uni] = pd.to_numeric(df_car[col_uni], errors="coerce").fillna(0)

    # Filtrar SKUs reales (>= 1 bulto, códigos > 0)
    df_car = df_car[(df_car[col_bul] > 0) & (df_car[col_art] > 0)]

    # Agrupar por artículo
    agg_dict = {col_bul: "sum"}
    if col_uni:
        agg_dict[col_uni] = "sum"
    if col_desc:
        agg_dict[col_desc] = "first"

    df_grp = df_car.groupby(col_art).agg(agg_dict).reset_index()
    df_grp.rename(columns={col_art: "CODIGO", col_bul: "BULTOS"}, inplace=True)
    if col_desc:
        df_grp.rename(columns={col_desc: "DESCRIPCION"}, inplace=True)
    else:
        df_grp["DESCRIPCION"] = df_grp["CODIGO"].astype(str)
    if col_uni:
        df_grp.rename(columns={col_uni: "U_PAQUETE"}, inplace=True)
    else:
        df_grp["U_PAQUETE"] = 0.0

    # ── 2. Leer DDM de Frescura para VALOR HL, RUBRO, ABC ────────────────────
    wb_fr = _ox.load_workbook(io.BytesIO(fr_bytes), read_only=True, data_only=True)

    # Encontrar hoja DDM
    ddm_sheet = None
    for sname in wb_fr.sheetnames:
        if sname.strip().upper() == "DDM":
            ddm_sheet = sname
            break
    if ddm_sheet is None:
        wb_fr.close()
        df_grp["HL"]    = 0.0
        df_grp["RUBRO"] = ""
        df_grp["ABC"]   = ""
    else:
        ws_ddm = wb_fr[ddm_sheet]
        rows_ddm = list(ws_ddm.iter_rows(values_only=True))
        wb_fr.close()

        # Extraer valores computados de IMPORTRANGE (están en el campo como COMPUTED_VALUE)
        def _extract_val(cell_val):
            """Extrae el COMPUTED_VALUE de las fórmulas IFERROR/IMPORTRANGE del xlsx offline."""
            if cell_val is None:
                return None
            s = str(cell_val)
            marker = "COMPUTED_VALUE\"\"\"),"
            idx = s.find(marker)
            if idx >= 0:
                rest = s[idx + len(marker):]
                # Puede ser número o string entre comillas
                rest = rest.rstrip(")")
                # Número
                try:
                    return float(rest)
                except ValueError:
                    pass
                # String: quitar comillas
                rest = rest.strip('"').strip("'")
                return rest if rest else None
            # Valor directo (no fórmula)
            return cell_val

        # Header de DDM
        hdr_ddm_raw = rows_ddm[0]
        hdr_ddm = [str(_extract_val(v)).strip() if _extract_val(v) else f"c{i}" for i, v in enumerate(hdr_ddm_raw)]

        ddm_data = []
        for row in rows_ddm[1:]:
            extracted = [_extract_val(v) for v in row]
            if any(v is not None for v in extracted):
                ddm_data.append(extracted)

        df_ddm = pd.DataFrame(ddm_data, columns=hdr_ddm)
        df_ddm = df_ddm.dropna(how="all")

        # Mapear columnas DDM
        col_d_art  = next((c for c in df_ddm.columns if c.strip().upper() in ("ARTÍCULO", "ARTICULO")), None)
        col_d_hl   = next((c for c in df_ddm.columns if "VALOR HL" in c.upper()), None)
        col_d_rub  = next((c for c in df_ddm.columns if "RUBRO" in c.upper()), None)
        col_d_abc  = next((c for c in df_ddm.columns if c.strip().upper() == "ABC"), None)
        col_d_desc = next((c for c in df_ddm.columns if "DESCRIPCIÓN" in c.upper() or "DESCRIPCION" in c.upper()), None)

        if col_d_art:
            df_ddm[col_d_art] = pd.to_numeric(df_ddm[col_d_art], errors="coerce")
            if col_d_hl:
                df_ddm[col_d_hl] = pd.to_numeric(df_ddm[col_d_hl], errors="coerce").fillna(0)

            ddm_lookup = df_ddm[df_ddm[col_d_art].notna()].copy()
            ddm_lookup.rename(columns={col_d_art: "CODIGO"}, inplace=True)

            keep_cols = ["CODIGO"]
            rename_map = {}
            if col_d_hl:
                keep_cols.append(col_d_hl); rename_map[col_d_hl] = "VALOR_HL"
            if col_d_rub:
                keep_cols.append(col_d_rub); rename_map[col_d_rub] = "RUBRO"
            if col_d_abc:
                keep_cols.append(col_d_abc); rename_map[col_d_abc] = "ABC"
            if col_d_desc and "DESCRIPCION" not in df_grp.columns:
                keep_cols.append(col_d_desc); rename_map[col_d_desc] = "DESC_DDM"

            ddm_lookup = ddm_lookup[keep_cols].rename(columns=rename_map)
            ddm_lookup = ddm_lookup.drop_duplicates(subset=["CODIGO"])

            df_grp = df_grp.merge(ddm_lookup, on="CODIGO", how="left")
        else:
            df_grp["VALOR_HL"] = 0.0
            df_grp["RUBRO"]    = ""
            df_grp["ABC"]      = ""

    # ── 3. Calcular HL totales ─────────────────────────────────────────────
    if "VALOR_HL" in df_grp.columns:
        df_grp["VALOR_HL"] = pd.to_numeric(df_grp["VALOR_HL"], errors="coerce").fillna(0)
        df_grp["HL"] = (df_grp["BULTOS"] * df_grp["VALOR_HL"]).round(2)
    else:
        df_grp["HL"] = 0.0

    # ── 3.b. FILTRO ALMACENES (DDM) ──────────────────────────────────────────
    # Solo contar SKUs que existen en la DDM (= productos de almacén).
    # Excluye automáticamente envases/esqueletos (ej. 2776 BOT ARACELI,
    # 2731 Q CERVEZAS, 2730 Q PLAS, 2780 BOT AMBAR) que no son productos
    # almacenables. Si la DDM no se pudo leer, se aplica fallback por
    # lista negra hardcodeada.
    if "VALOR_HL" in df_grp.columns and ddm_sheet is not None:
        # DDM cargada exitosamente → universo válido = SKUs presentes en DDM
        # Un SKU está "en DDM" si tiene VALOR_HL no-NaN tras el merge
        # (puede ser 0, pero existir). Para distinguir 0-en-DDM vs ausente,
        # creamos un set de SKUs DDM explícitamente.
        try:
            ddm_skus = set(ddm_lookup["CODIGO"].dropna().astype(int).tolist())
            df_grp["__en_ddm"] = df_grp["CODIGO"].astype(int).isin(ddm_skus)
            df_grp = df_grp[df_grp["__en_ddm"]].drop(columns="__en_ddm")
        except Exception:
            # Fallback: lista negra clásica
            df_grp = df_grp[~df_grp["CODIGO"].astype(int).isin(EXCLUDED_SKUS)]
    else:
        df_grp = df_grp[~df_grp["CODIGO"].astype(int).isin(EXCLUDED_SKUS)]

    # ── 4. Top 10 por Bultos ───────────────────────────────────────────────
    df_top = (
        df_grp
        .sort_values("BULTOS", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    df_top.index += 1  # Rankeo 1..10

    # Limpiar descripción
    df_top["DESCRIPCION"] = df_top["DESCRIPCION"].astype(str).str.strip()

    # Columnas finales
    cols_out = ["CODIGO", "DESCRIPCION", "BULTOS", "HL"]
    if "U_PAQUETE" in df_top.columns:
        cols_out.append("U_PAQUETE")
    if "RUBRO" in df_top.columns:
        cols_out.append("RUBRO")
    if "ABC" in df_top.columns:
        cols_out.append("ABC")

    return df_top[cols_out].copy()


def _render_top10_section(pdata: dict, car_bytes: bytes, fr_bytes: bytes):
    """
    Renderiza la sección 'Top 10 SKUs más vendidos' dentro de la Tab Proyección.
    Se llama automáticamente si CAR y Frescura están disponibles.
    """
    st.divider()
    st.subheader("🏆 Top 10 SKUs — Día de Carga")
    st.caption(
        "Ranking por **Bultos** del CAR del día · "
        "HL calculado con **VALOR HL × Bultos** (DDM Frescura) · "
        "Tabla lista para copiar en Sheets."
    )

    try:
        df_top = _build_top10_skus(car_bytes, fr_bytes)
    except Exception as e:
        st.error(f"❌ Error construyendo Top 10: {e}")
        import traceback
        with st.expander("Stack trace"):
            st.code(traceback.format_exc())
        return

    if df_top.empty:
        st.warning("No se encontraron SKUs con bultos > 0.")
        return

    fecha_str = pdata.get("fecha", "")
    tot_bultos = df_top["BULTOS"].sum()
    tot_hl     = df_top["HL"].sum() if "HL" in df_top.columns else 0

    # ── Métricas resumen ─────────────────────────────────────────────────────
    mc1, mc2, mc3 = st.columns(3)
    mc1.metric("SKUs únicos en Top 10",  len(df_top))
    mc2.metric("Bultos acumulados Top 10", f"{tot_bultos:,.0f}")
    mc3.metric("HL acumulados Top 10",    f"{tot_hl:,.2f}")

    # ── Tabla con formato condicional ────────────────────────────────────────
    # Preparar display DF
    df_display = df_top.copy()
    df_display.insert(0, "#", range(1, len(df_display) + 1))
    df_display["BULTOS"] = df_display["BULTOS"].map(lambda x: f"{x:,.0f}")
    df_display["HL"]     = df_display["HL"].map(lambda x: f"{x:,.2f}" if x else "—")
    if "U_PAQUETE" in df_display.columns:
        df_display["U_PAQUETE"] = df_display["U_PAQUETE"].map(
            lambda x: f"{x:,.2f}" if pd.notna(x) and x != 0 else "—"
        )
    df_display["CODIGO"] = df_display["CODIGO"].astype(int).astype(str)

    # Renombrar para Sheets
    rename_display = {
        "#":           "#",
        "CODIGO":      "Código",
        "DESCRIPCION": "Descripción",
        "BULTOS":      "Bultos",
        "HL":          "HL",
        "U_PAQUETE":   "U. Paquete",
        "RUBRO":       "Rubro",
        "ABC":         "ABC",
    }
    df_display.rename(columns={k: v for k, v in rename_display.items() if k in df_display.columns}, inplace=True)

    # Estilo condicional sutil: ABC coloreado
    def _style_abc(val):
        colors_abc = {"A": "background-color:#d4edda;color:#155724;font-weight:bold",
                      "B": "background-color:#fff3cd;color:#856404",
                      "C": "background-color:#f8d7da;color:#721c24"}
        return colors_abc.get(str(val).strip().upper(), "")

    try:
        styler = df_display.style.hide(axis="index")
        # Styler.map (pandas >= 2.1). Si no existe el método o falla, mostrar sin estilo.
        if "ABC" in df_display.columns and not df_display["ABC"].isna().all() \
           and hasattr(styler, "map"):
            styler = styler.map(_style_abc, subset=["ABC"])
        st.dataframe(styler, width="stretch", height=420)
    except Exception:
        # Fallback ultra-robusto: tabla plana sin formato condicional
        st.dataframe(df_display, width="stretch", height=420, hide_index=True)

    # ── Exportar como TSV (listo para pegar en Sheets) ───────────────────────
    tsv_raw = df_display.to_csv(sep="\t", index=False)
    st.download_button(
        label="📋 Descargar TSV (pegar en Sheets)",
        data=tsv_raw.encode("utf-8"),
        file_name=f"top10_skus_{fecha_str}.tsv",
        mime="text/tab-separated-values",
        width="stretch",
        key="top10_tsv_dl",
    )

    # ── Nota metodológica ───────────────────────────────────────────────────
    with st.expander("ℹ️ Metodología"):
        st.markdown(
            "- **Bultos**: suma de todos los remitos del CAR del día (excluye anulados)\n"
            "- **HL**: `Bultos × VALOR HL` — factor tomado de la hoja DDM de Frescura 3.0\n"
            "- **U. Paquete**: unidades individuales del paquete (col `Unids` del CAR)\n"
            "- **Rubro / ABC**: clasificación de la DDM (base maestra CMQ)\n"
            "- SKUs sin VALOR HL en DDM aparecen con `0.00` en HL\n"
            "- El ranking es exclusivo del **día de carga actual** — no es histórico"
        )


# ── TAB 5 — Extraíbles Sheets ──────────────────────────────────────────────

def render_tab_extraibles():
    st.subheader("📤 Extraíbles para Google Sheets")
    st.caption(
        "Bajá los bloques del MASTER en formato TSV/CSV/XLSX para pegar manual "
        "en el Ecosistema AD 3.0. Sin push automático en v4.0 (decisión #2)."
    )

    master = _file_uploader(
        "MASTER del día.xlsx (78 MB)", ["xlsx"], key="t5_master"
    )

    if not master:
        st.info("Subí el MASTER para acceder a los 4 bloques.")
        return

    file_bytes = master.getvalue()

    try:
        df_pall, fecha_a1 = load_master_matriz_pall(file_bytes)
        df_rep = load_master_reposicion_ae(file_bytes)
        df_pick = load_master_matriz_picking_app(file_bytes)
        df_agr = load_master_agregados_ae(file_bytes)
        log_event("info", f"MASTER cargado para extraíbles. Fecha A1: {fecha_a1}")
    except Exception as e:
        st.error(f"❌ Error leyendo MASTER: {e}")
        return

    fecha_str = fecha_a1.strftime("%d/%m/%Y") if fecha_a1 else "—"
    st.success(f"✓ MASTER cargado — Fecha en A1: **{fecha_str}**")

    st.divider()

    # ── Bloque 1: Matriz Pall. ──────────────────────────────────────────
    with st.expander("1️⃣  Matriz Pall. (29 camiones × 24 cols)", expanded=True):
        st.dataframe(df_pall, width="stretch", height=300)
        _download_trio(df_pall, "Matriz_Pall", "Matriz Pall.", "t5_pall")

    # ── Bloque 2: Agregados AE ──────────────────────────────────────────
    with st.expander("2️⃣  Agregados AE (subset Matriz Pall. — cols AE)"):
        st.caption(
            "⚠ Provisional: deriva de Matriz Pall. cols K-Q. "
            "Cuando confirmes hoja destino exacta en Ecosistema AD 3.0, se ajusta."
        )
        st.dataframe(df_agr, width="stretch", height=300)
        _download_trio(df_agr, "Agregados_AE", "Agregados AE", "t5_agr")

    # ── Bloque 3: Reposición AE (filtrado J<0) ──────────────────────────
    with st.expander("3️⃣  Reposición AE (solo Reposicion_Pall < 0)", expanded=True):
        df_rep_neg = filter_reposicion_negativa(df_rep)
        st.metric("Filas con reposición negativa", f"{len(df_rep_neg)} / {len(df_rep)}")
        if len(df_rep_neg) == 0:
            st.warning("No hay filas con Reposicion_Pall < 0. Nada para pegar hoy.")
        else:
            st.dataframe(df_rep_neg, width="stretch", height=300)
            _download_trio(df_rep_neg, "Reposicion_AE_neg", "Reposición AE", "t5_rep")

    # ── Bloque 4: (fx) Picking ──────────────────────────────────────────
    with st.expander("4️⃣  (fx) Picking — fila resumen Matriz Picking APP"):
        st.caption(
            "Vista plana de la hoja Matriz Picking APP. Para Sprint 1 se exporta "
            "completa; en Sprint 2 se filtra/recorta según destino exacto."
        )
        st.dataframe(df_pick.head(50), width="stretch", height=300)
        _download_trio(df_pick, "fx_Picking", "(fx) Picking", "t5_pick")


# ── TAB 6 — Validación + Log (PLACEHOLDER, Sprint 2) ───────────────────────

def render_tab_validacion():
    st.subheader("✅ Validación + Reglas + Log")
    st.caption(
        "Pre-checks + reglas de negocio + validación cruzada CAR↔MASTER. "
        "**Sprint 2** — pendiente."
    )

    st.info(
        "🔧 En desarrollo (Sprint 2).\n\n"
        "**Pre-checks previstos:**\n"
        "- CAR.xlsx modificado hoy\n"
        "- ANR (rechazos1.xlsx) modificado hoy (warning)\n"
        "- MASTER → Matriz Pall.!A1 == hoy\n"
        "- Suma TOTAL PALL > 0\n\n"
        "**Reglas:**\n"
        "- Camión TOTAL=0 → ocultar y loguear\n"
        "- Camión Reparto=NO → tabla checkbox, default OFF (decisión #1)\n"
        "- Cross-validate CAR↔MASTER (WARNING, no bloqueante — decisión #6)\n"
        "- Sanity checks bloqueantes pre-publicación"
    )

    st.divider()
    st.markdown("### 📜 Log de la corrida actual")
    log_lines = st.session_state.get("log_buffer", [])
    if not log_lines:
        st.caption("(sin eventos aún)")
    else:
        st.code("\n".join(log_lines[-200:]), language="text")
        st.download_button(
            "⬇ Descargar log completo",
            data="\n".join(log_lines).encode("utf-8"),
            file_name=_stamp("run_log", "txt"),
            mime="text/plain",
        )


# ── MAIN ────────────────────────────────────────────────────────────────────

# ── TAB — CLASIFICACIÓN (Retornables: Almacén 1 y 3) ───────────────────────

def _build_clasificacion_retornables(car_bytes: bytes, fr_bytes: bytes) -> tuple:
    """
    Cantidad total por camión de los SKUs de Almacén 1 y 3 (retornables).

    Lógica v4.11:
      - CAR (Hoja1): filas con Depósito 1 o 3, no anuladas, SKU en DDM (= almacén).
      - Bultos: suma directa de col 'Bultos' (cajones cerrados).
      - Cajones extra desde unidades sueltas:
          · Almacén 1 → 12 un. por cajón (todo X12 1000)
          · Almacén 3 → 24 un. por cajón (todo X24 340)
          · cajones_extra = floor(sum(Unids del almacén) / 12 o 24)
      - Total camión = bultos_enteros + cajones_extra (sumando ambos almacenes).
      - Pallets equivalentes = total_camion / 50 (BXP estándar retornable).

    Excluye automáticamente envases/esqueletos (2776, 2731, 2730, 2780, etc.)
    porque no están en DDM.

    Devuelve (df_por_camion, dict_totales).
    Columnas: Camión | Bultos Retornables | Pallets Equivalentes
    """
    import openpyxl as _ox
    import math as _math

    BXP_RETORNABLE = 50  # estándar Quilmes retornable
    UN_POR_CAJON = {1: 12, 3: 24}  # Alm 1 = X12, Alm 3 = X24

    # ── 1. CAR ───────────────────────────────────────────────────────────────
    wb_car = _ox.load_workbook(io.BytesIO(car_bytes), read_only=True, data_only=True)
    ws_car = wb_car["Hoja1"] if "Hoja1" in wb_car.sheetnames else wb_car.active
    rows = list(ws_car.iter_rows(values_only=True))
    wb_car.close()
    if not rows:
        raise ValueError("CAR vacío.")
    hdr = [str(v).strip() if v is not None else f"c{i}" for i, v in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=hdr).dropna(how="all")

    col_dep  = find_col(df, "Depósito", "Deposito")
    col_tra  = find_col(df, "Transporte", "Camión", "Camion")
    col_art  = find_col(df, "Artículo", "Articulo")
    col_bul  = find_col(df, "Bultos")
    col_uni  = find_col(df, "Unids", "Unidades")
    col_anu  = find_col(df, "Anulado")
    if col_dep is None or col_tra is None or col_bul is None:
        raise ValueError(f"CAR: faltan columnas clave (Depósito/Transporte/Bultos). Detectadas: {list(df.columns)}")

    df[col_dep] = pd.to_numeric(df[col_dep], errors="coerce")
    df[col_bul] = pd.to_numeric(df[col_bul], errors="coerce").fillna(0)
    if col_uni:
        df[col_uni] = pd.to_numeric(df[col_uni], errors="coerce").fillna(0)
    if col_art:
        df[col_art] = pd.to_numeric(df[col_art], errors="coerce")
    if col_anu:
        df = df[df[col_anu].astype(str).str.lower() != "yes"]

    # ── 2. DDM → universo SKUs almacén ───────────────────────────────────────
    ddm_skus = set()
    try:
        wb_fr = _ox.load_workbook(io.BytesIO(fr_bytes), read_only=True, data_only=True)
        if "DDM" in wb_fr.sheetnames:
            wd = wb_fr["DDM"]
            rd = list(wd.iter_rows(values_only=True))
            wb_fr.close()
            hd = [str(v).strip() if v is not None else f"c{i}" for i, v in enumerate(rd[0])]
            ddm_df = pd.DataFrame(rd[1:], columns=hd).dropna(how="all")
            c_art = find_col(ddm_df, "ARTÍCULO", "ARTICULO", "Artículo")
            if c_art:
                ddm_df[c_art] = pd.to_numeric(ddm_df[c_art], errors="coerce")
                ddm_skus = set(ddm_df[ddm_df[c_art].notna()][c_art].astype(int).tolist())
        else:
            wb_fr.close()
    except Exception:
        ddm_skus = set()

    # ── 3. Filtro retornables (Depósito 1 o 3) + DDM ─────────────────────────
    df_ret = df[df[col_dep].isin([1, 3])].copy()
    if df_ret.empty:
        return pd.DataFrame(columns=["Camión", "Bultos Retornables", "Pallets Equivalentes"]), \
               {"bultos": 0, "pallets": 0.0, "camiones": 0}

    if ddm_skus and col_art:
        # Filtrar SKUs presentes en DDM (excluye envases/esqueletos)
        df_ret[col_art] = df_ret[col_art].astype("Float64")
        df_ret = df_ret[df_ret[col_art].notna()]
        df_ret = df_ret[df_ret[col_art].astype(int).isin(ddm_skus)]
    else:
        # Fallback: lista negra clásica
        if col_art:
            df_ret = df_ret[~df_ret[col_art].isin(EXCLUDED_SKUS)]

    # Mantener solo filas que aporten algo: bultos>0 o unids>0
    if col_uni:
        df_ret = df_ret[(df_ret[col_bul] > 0) | (df_ret[col_uni] > 0)].copy()
    else:
        df_ret = df_ret[df_ret[col_bul] > 0].copy()

    if df_ret.empty:
        return pd.DataFrame(columns=["Camión", "Bultos Retornables", "Pallets Equivalentes"]), \
               {"bultos": 0, "pallets": 0.0, "camiones": 0}

    # ── 4. Consolidar por camión y almacén ───────────────────────────────────
    # Agregamos por (camión, depósito) para aplicar la regla de 12/24 un.
    agg_dict = {col_bul: "sum"}
    if col_uni:
        agg_dict[col_uni] = "sum"

    grp = df_ret.groupby([col_tra, col_dep]).agg(agg_dict).reset_index()
    grp.rename(columns={col_bul: "BULTOS", col_uni: "UNIDS" if col_uni else "_un"},
               inplace=True)
    if "UNIDS" not in grp.columns:
        grp["UNIDS"] = 0.0

    # Cajones extra desde unidades sueltas
    def _cajones_extra(row):
        upk = UN_POR_CAJON.get(int(row[col_dep]), 12)
        return int(_math.floor(row["UNIDS"] / upk)) if row["UNIDS"] > 0 else 0

    grp["CAJ_EXTRA"] = grp.apply(_cajones_extra, axis=1)
    grp["TOTAL_BULTOS"] = grp["BULTOS"] + grp["CAJ_EXTRA"]

    # Sumar Alm 1 + Alm 3 por camión
    cam = grp.groupby(col_tra).agg(
        BULTOS=("TOTAL_BULTOS", "sum"),
        BULTOS_ENT=("BULTOS", "sum"),
        UNIDS=("UNIDS", "sum"),
        CAJ_EXTRA=("CAJ_EXTRA", "sum"),
    ).reset_index()
    cam[col_tra] = pd.to_numeric(cam[col_tra], errors="coerce")
    cam = cam.sort_values(col_tra)

    # Pallets equivalentes: total_bultos / 50
    cam["PALLETS"] = (cam["BULTOS"] / BXP_RETORNABLE).round(2)

    out = pd.DataFrame({
        "Camión":               cam[col_tra].apply(lambda x: str(int(x)) if pd.notna(x) else "—"),
        "Bultos Retornables":   cam["BULTOS"].astype(float),
        "Pallets Equivalentes": cam["PALLETS"].astype(float),
        "_bultos_enteros":      cam["BULTOS_ENT"].astype(float),
        "_unids":               cam["UNIDS"].astype(float),
        "_cajones_extra":       cam["CAJ_EXTRA"].astype(float),
    })

    tot = {
        "bultos":         float(out["Bultos Retornables"].sum()),
        "pallets":        round(float(out["Pallets Equivalentes"].sum()), 2),
        "camiones":       int(len(out)),
        "bultos_enteros": float(out["_bultos_enteros"].sum()),
        "unids":          float(out["_unids"].sum()),
        "cajones_extra":  float(out["_cajones_extra"].sum()),
    }
    return out, tot


def _build_clasificacion_anr(anr_bytes: bytes, fr_bytes: bytes) -> tuple:
    """
    v4.16 — Clasificación POR CAMIÓN (envases vacíos + esqueletos a retornar).

    Cambio de enfoque vs v4.15: ya no importa el SKU individual sino cuánto
    debe retornar cada camión. Misma lógica de filtrado y cálculo, solo cambia
    la agrupación final.

    Lógica:
      - ANR (hoja BASE, header en fila 2 / índice 1):
          col H=7  ARTÍCULO | col K=10 BULTOS | col AP=41 TRANSPORTE (camión)
      - Frescura (hoja DDM, header en fila 1 / índice 0):
          A=0 ALMACÉN | C=2 ARTÍCULO | D=3 DESCRIPCIÓN | G=6 BULTOS X PALLET
      - Filtra DDM a ALMACÉN ∈ {1, 3} → mapa SKU → (BXP, ALMACÉN).
      - Por cada fila ANR de SKU retornable: pallets_fila = bultos_fila / BXP_SKU.
      - Agrupa por CAMIÓN: Σ bultos, Σ pallets.
      - Productividad: ENTERO(total_pallets / target_x_operario) + 1.

    Devuelve (df_por_camion, dict_totales) con columnas:
        CAMION | BULTOS | PALLETS
    """
    import openpyxl as _ox
    import math as _math

    # ── 1. DDM → mapa SKU para almacenes 1 y 3 ───────────────────────────
    wb_fr = _ox.load_workbook(io.BytesIO(fr_bytes), read_only=True, data_only=True)
    if "DDM" not in wb_fr.sheetnames:
        wb_fr.close()
        raise ValueError("La Frescura no contiene la hoja 'DDM'.")
    wd = wb_fr["DDM"]
    rd = list(wd.iter_rows(values_only=True))
    wb_fr.close()
    if len(rd) < 2:
        raise ValueError("DDM vacía.")

    # Cols: A=0 ALMACÉN | C=2 ARTÍCULO | D=3 DESCRIPCIÓN | G=6 BULTOS X PALLET
    sku_map: dict[int, dict] = {}
    for row in rd[1:]:
        if not row:
            continue
        try:
            alm = row[0]
            art = row[2]
            desc = row[3] if len(row) > 3 else ""
            bxp = row[6] if len(row) > 6 else None
            if alm is None or art is None:
                continue
            alm_i = int(alm)
            if alm_i not in (1, 3):
                continue
            art_i = int(art)
            try:
                bxp_f = float(bxp) if bxp is not None else 50.0
            except (ValueError, TypeError):
                bxp_f = 50.0
            if bxp_f <= 0:
                bxp_f = 50.0
            # Si un SKU aparece duplicado (ej. en alm 1 y 3), nos quedamos con el primero
            if art_i not in sku_map:
                sku_map[art_i] = {
                    "ALMACEN":     alm_i,
                    "DESCRIPCION": str(desc) if desc else "",
                    "BXP":         bxp_f,
                }
        except (ValueError, TypeError):
            continue

    if not sku_map:
        return pd.DataFrame(columns=["CAMION", "BULTOS", "PALLETS"]), \
               {"bultos": 0.0, "pallets": 0.0, "skus": 0, "camiones": 0}

    # ── 2. ANR BASE → bultos+pallets POR CAMIÓN ──────────────────────────
    # Header en fila 2 (índice 1) → header=1 con pandas.
    df_anr = pd.read_excel(io.BytesIO(anr_bytes), sheet_name="BASE", header=1)
    if df_anr.empty or df_anr.shape[1] < 42:  # necesitamos hasta AP=41 inclusive
        return pd.DataFrame(columns=["CAMION", "BULTOS", "PALLETS"]), \
               {"bultos": 0.0, "pallets": 0.0, "skus": 0, "camiones": 0}

    # H=7 ARTÍCULO, K=10 BULTOS, AP=41 TRANSPORTE (camión)
    sub = df_anr.iloc[:, [7, 10, 41]].copy()
    sub.columns = ["ARTICULO", "BULTOS", "CAMION"]
    sub["ARTICULO"] = pd.to_numeric(sub["ARTICULO"], errors="coerce")
    sub["BULTOS"]   = pd.to_numeric(sub["BULTOS"],   errors="coerce").fillna(0)
    sub["CAMION"]   = pd.to_numeric(sub["CAMION"],   errors="coerce")
    sub = sub[sub["ARTICULO"].notna() & (sub["BULTOS"] > 0) & sub["CAMION"].notna()].copy()
    sub["ARTICULO"] = sub["ARTICULO"].astype(int)
    sub["CAMION"]   = sub["CAMION"].astype(int)

    # Filtrar SOLO SKUs retornables (alm 1 o 3 según DDM)
    sub = sub[sub["ARTICULO"].isin(sku_map.keys())]
    if sub.empty:
        return pd.DataFrame(columns=["CAMION", "BULTOS", "PALLETS"]), \
               {"bultos": 0.0, "pallets": 0.0, "skus": 0, "camiones": 0}

    # Pallets por fila (bultos_fila / BXP del SKU)
    sub["BXP"]     = sub["ARTICULO"].map(lambda s: sku_map[s]["BXP"])
    sub["PALLETS"] = sub.apply(
        lambda r: r["BULTOS"] / r["BXP"] if r["BXP"] > 0 else 0, axis=1
    )

    # Agrupar por CAMIÓN
    out = sub.groupby("CAMION", as_index=False).agg(
        BULTOS=("BULTOS", "sum"),
        PALLETS=("PALLETS", "sum"),
    )
    out["PALLETS"] = out["PALLETS"].round(4)
    out = out.sort_values("CAMION").reset_index(drop=True)

    tot = {
        "bultos":   float(out["BULTOS"].sum()),
        "pallets":  round(float(out["PALLETS"].sum()), 2),
        "skus":     int(sub["ARTICULO"].nunique()),
        "camiones": int(out["CAMION"].nunique()),
    }
    return out, tot


def build_clasificacion_anr_pdf(df_cl: pd.DataFrame, tot: dict, fecha_str: str = "") -> bytes:
    """
    v4.16 — PDF de Clasificación POR CAMIÓN (Almacén 1 y 3).
    Muestra: CAMIÓN | BULTOS | PALLETS  +  caja de Productividad estimada.
    """
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib import colors as _col
    from reportlab.lib.units import mm as _mm
    from reportlab.pdfgen import canvas as _cv
    import io as _io

    df = df_cl[df_cl["BULTOS"] > 0].copy() if not df_cl.empty else df_cl
    PW, PH = _A4
    M = 12 * _mm

    DARK       = _col.HexColor("#1a3a6b")
    YELLOW_HDR = _col.HexColor("#f8d772")
    BAND_LIGHT = _col.HexColor("#f4f6fa")
    BAND_WHITE = _col.white
    BORDER     = _col.HexColor("#8c8c8c")
    TOTAL_BG   = _col.HexColor("#1a3a6b")

    bio = _io.BytesIO()
    c = _cv.Canvas(bio, pagesize=_A4)

    # Header
    HDR_H = 22 * _mm
    c.setFillColor(DARK)
    c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)
    c.setFillColor(_col.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(M, PH - 11 * _mm, "CLASIFICACIÓN POR CAMIÓN — Almacenes 1 y 3")
    c.setFont("Helvetica", 10)
    c.drawString(M, PH - 17 * _mm, "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén · Envases vacíos + esqueletos a retornar")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(PW - M, PH - 11 * _mm, f"Fecha: {fecha_str}")
    c.setFont("Helvetica", 9)
    c.drawRightString(PW - M, PH - 17 * _mm, f"Camiones: {tot.get('camiones', 0)}")

    # Tabla — 3 columnas centradas, ancho moderado
    y = PH - HDR_H - 10 * _mm
    table_w = (PW - 2 * M) * 0.62  # estrecha, centrada
    table_x = (PW - table_w) / 2
    col_widths = [
        table_w * 0.28,  # Camión
        table_w * 0.36,  # Bultos
        table_w * 0.36,  # Pallets
    ]
    col_x = [table_x]
    for w in col_widths[:-1]:
        col_x.append(col_x[-1] + w)

    THDR_H = 8 * _mm
    c.setFillColor(YELLOW_HDR)
    c.rect(table_x, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
    c.setStrokeColor(BORDER)
    c.setFillColor(DARK)
    c.setFont("Helvetica-Bold", 10)
    for i, h in enumerate(["CAMIÓN", "BULTOS", "PALLETS"]):
        c.drawCentredString(col_x[i] + col_widths[i] / 2, y - THDR_H + 2.5 * _mm, h)
    y -= THDR_H

    ROW_H = 7 * _mm
    if df.empty:
        c.setFillColor(_col.HexColor("#777777"))
        c.setFont("Helvetica-Oblique", 10)
        c.drawCentredString(PW / 2, y - 10 * _mm, "Sin camiones con retornables hoy.")
        y -= 20 * _mm
    else:
        for i, row in df.reset_index(drop=True).iterrows():
            if y < 70 * _mm:
                c.showPage()
                y = PH - 20 * _mm
                c.setFillColor(YELLOW_HDR)
                c.rect(table_x, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
                c.setFillColor(DARK)
                c.setFont("Helvetica-Bold", 10)
                for j, h in enumerate(["CAMIÓN", "BULTOS", "PALLETS"]):
                    c.drawCentredString(col_x[j] + col_widths[j] / 2, y - THDR_H + 2.5 * _mm, h)
                y -= THDR_H

            bg = BAND_LIGHT if i % 2 == 0 else BAND_WHITE
            c.setFillColor(bg)
            c.rect(table_x, y - ROW_H, table_w, ROW_H, fill=1, stroke=0)
            c.setStrokeColor(BORDER)
            c.setLineWidth(0.3)
            c.rect(table_x, y - ROW_H, table_w, ROW_H, fill=0, stroke=1)
            for cx in col_x[1:]:
                c.line(cx, y - ROW_H, cx, y)

            c.setFillColor(_col.black)
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(col_x[0] + col_widths[0] / 2, y - ROW_H + 2.2 * _mm,
                                str(int(row["CAMION"])))
            c.setFont("Helvetica", 10)
            c.drawRightString(col_x[1] + col_widths[1] - 3 * _mm, y - ROW_H + 2.2 * _mm,
                              f"{float(row['BULTOS']):,.0f}".replace(",", "."))
            c.setFont("Helvetica-Bold", 10)
            c.drawRightString(col_x[2] + col_widths[2] - 3 * _mm, y - ROW_H + 2.2 * _mm,
                              f"{float(row['PALLETS']):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            y -= ROW_H

        # Fila TOTAL
        TOT_H = 9 * _mm
        c.setFillColor(TOTAL_BG)
        c.rect(table_x, y - TOT_H, table_w, TOT_H, fill=1, stroke=1)
        c.setFillColor(_col.white)
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(col_x[0] + col_widths[0] / 2, y - TOT_H + 3 * _mm, "TOTAL")
        c.drawRightString(col_x[1] + col_widths[1] - 3 * _mm, y - TOT_H + 3 * _mm,
                          f"{float(tot.get('bultos', 0)):,.0f}".replace(",", "."))
        c.drawRightString(col_x[2] + col_widths[2] - 3 * _mm, y - TOT_H + 3 * _mm,
                          f"{float(tot.get('pallets', 0)):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        y -= TOT_H

    # Caja de Productividad Estimada
    prod = tot.get("productividad") or {}
    if prod:
        y -= 8 * _mm
        BOX_H = 30 * _mm
        box_w = PW - 2 * M
        c.setFillColor(_col.HexColor("#fff8e1"))
        c.setStrokeColor(YELLOW_HDR)
        c.setLineWidth(1.2)
        c.rect(M, y - BOX_H, box_w, BOX_H, fill=1, stroke=1)
        c.setFillColor(DARK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(M + 4 * _mm, y - 6 * _mm, "⏱  PRODUCTIVIDAD ESTIMADA DE CLASIFICACIÓN")

        cell_w = box_w / 5
        y_label = y - 12 * _mm
        y_value = y - 19 * _mm
        def _fmt_pal(v):
            return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        labels = [
            ("Pallets a clasificar", _fmt_pal(prod.get('pallets_total', 0))),
            ("Target × operario",    f"{prod.get('target_oper', 0)}"),
            ("Pallets / hora",       f"{prod.get('pall_hora', 0)}"),
            ("Personal necesario",   f"{prod.get('personal_nec', 0)}"),
            ("Duración estimada",    f"{prod.get('duracion', '00:00')}"),
        ]
        for i, (lbl, val) in enumerate(labels):
            cx = M + i * cell_w + cell_w / 2
            c.setFillColor(_col.HexColor("#555555"))
            c.setFont("Helvetica", 8.5)
            c.drawCentredString(cx, y_label, lbl)
            c.setFillColor(DARK)
            c.setFont("Helvetica-Bold", 14 if i in (3, 4) else 12)
            c.drawCentredString(cx, y_value, val)

        c.setFillColor(_col.HexColor("#444444"))
        c.setFont("Helvetica-Oblique", 8)
        concl = (f"Próximo turno: {prod.get('personal_nec',0)} operarios · "
                 f"{prod.get('pallets_total',0):.2f} pall · duración estimada {prod.get('duracion','00:00')} hs.")
        c.drawCentredString(M + box_w/2, y - 26 * _mm, concl)
        y -= BOX_H

    # Footer
    c.setFillColor(_col.HexColor("#555555"))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(M, 10 * _mm,
                 "Fuente: ANR (BASE: H=Artículo, K=Bultos, AP=Transporte) + Frescura DDM (alm 1/3, BXP).")
    c.drawRightString(PW - M, 10 * _mm,
                      f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.save()
    return bio.getvalue()


def build_clasificacion_anr_xlsx(df_cl: pd.DataFrame, tot: dict, fecha_str: str = "") -> bytes:
    """
    v4.64.0 — Excel de Clasificación POR CAMIÓN (Almacén 1 y 3).
    Misma visual y diseño que el PDF: header azul, tabla amarilla, filas alternadas,
    fila TOTAL azul, bloque de Productividad Estimada.
    """
    import io as _io
    import openpyxl as _xl
    from openpyxl.styles import (
        Font as _Font, PatternFill as _Fill, Alignment as _Align,
        Border as _Border, Side as _Side,
    )
    from openpyxl.utils import get_column_letter as _gcl

    df = df_cl[df_cl["BULTOS"] > 0].copy() if not df_cl.empty else df_cl

    DARK_HEX   = "1A3A6B"
    YELLOW_HEX = "F8D772"
    LIGHT_HEX  = "F4F6FA"
    WHITE_HEX  = "FFFFFF"
    GRAY_HEX   = "555555"

    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Clasificacion"

    thin = _Side(style="thin", color="8C8C8C")
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    center = _Align(horizontal="center", vertical="center", wrap_text=True)
    right  = _Align(horizontal="right",  vertical="center")
    left   = _Align(horizontal="left",   vertical="center")

    # ── HEADER (fila 1-3) ─────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = "CLASIFICACIÓN POR CAMIÓN — Almacenes 1 y 3"
    ws["A1"].font      = _Font(name="Calibri", bold=True, size=14, color=WHITE_HEX)
    ws["A1"].fill      = _Fill("solid", fgColor=DARK_HEX)
    ws["A1"].alignment = left
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:C2")
    ws["A2"] = "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén · Envases vacíos + esqueletos a retornar"
    ws["A2"].font      = _Font(name="Calibri", size=9, color=WHITE_HEX)
    ws["A2"].fill      = _Fill("solid", fgColor=DARK_HEX)
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 16

    # Fecha y cantidad de camiones en A3:C3
    ws["A3"] = f"Fecha: {fecha_str}"
    ws["A3"].font      = _Font(name="Calibri", bold=True, size=10, color=WHITE_HEX)
    ws["A3"].fill      = _Fill("solid", fgColor=DARK_HEX)
    ws["A3"].alignment = left
    ws["C3"] = f"Camiones: {tot.get('camiones', 0)}"
    ws["C3"].font      = _Font(name="Calibri", size=9, color=WHITE_HEX)
    ws["C3"].fill      = _Fill("solid", fgColor=DARK_HEX)
    ws["C3"].alignment = right
    for col in ["A", "B", "C"]:
        ws[f"{col}3"].fill = _Fill("solid", fgColor=DARK_HEX)
    ws.row_dimensions[3].height = 14

    # ── HEADER TABLA (fila 5) ─────────────────────────────────────────────────
    headers = ["CAMIÓN", "BULTOS", "PALLETS"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font      = _Font(name="Calibri", bold=True, size=10, color=DARK_HEX)
        cell.fill      = _Fill("solid", fgColor=YELLOW_HEX)
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[5].height = 18

    # ── FILAS DE DATOS ────────────────────────────────────────────────────────
    row_num = 6
    for i, row in df.reset_index(drop=True).iterrows():
        bg = LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        # Camión
        c1 = ws.cell(row=row_num, column=1, value=int(row["CAMION"]))
        c1.font      = _Font(name="Calibri", bold=True, size=10)
        c1.fill      = _Fill("solid", fgColor=bg)
        c1.alignment = center
        c1.border    = border
        # Bultos
        c2 = ws.cell(row=row_num, column=2, value=float(row["BULTOS"]))
        c2.font      = _Font(name="Calibri", size=10)
        c2.fill      = _Fill("solid", fgColor=bg)
        c2.alignment = right
        c2.border    = border
        c2.number_format = "#,##0"
        # Pallets
        c3 = ws.cell(row=row_num, column=3, value=float(row["PALLETS"]))
        c3.font      = _Font(name="Calibri", bold=True, size=10)
        c3.fill      = _Fill("solid", fgColor=bg)
        c3.alignment = right
        c3.border    = border
        c3.number_format = "#,##0.00"
        ws.row_dimensions[row_num].height = 16
        row_num += 1

    # ── FILA TOTAL ────────────────────────────────────────────────────────────
    t1 = ws.cell(row=row_num, column=1, value="TOTAL")
    t1.font      = _Font(name="Calibri", bold=True, size=11, color=WHITE_HEX)
    t1.fill      = _Fill("solid", fgColor=DARK_HEX)
    t1.alignment = center
    t1.border    = border
    t2 = ws.cell(row=row_num, column=2, value=float(tot.get("bultos", 0)))
    t2.font      = _Font(name="Calibri", bold=True, size=11, color=WHITE_HEX)
    t2.fill      = _Fill("solid", fgColor=DARK_HEX)
    t2.alignment = right
    t2.border    = border
    t2.number_format = "#,##0"
    t3 = ws.cell(row=row_num, column=3, value=float(tot.get("pallets", 0)))
    t3.font      = _Font(name="Calibri", bold=True, size=11, color=WHITE_HEX)
    t3.fill      = _Fill("solid", fgColor=DARK_HEX)
    t3.alignment = right
    t3.border    = border
    t3.number_format = "#,##0.00"
    ws.row_dimensions[row_num].height = 18
    row_num += 2

    # ── BLOQUE PRODUCTIVIDAD ESTIMADA ─────────────────────────────────────────
    prod = tot.get("productividad") or {}
    if prod:
        ws.merge_cells(f"A{row_num}:C{row_num}")
        ws[f"A{row_num}"] = "PRODUCTIVIDAD ESTIMADA DE CLASIFICACIÓN"
        ws[f"A{row_num}"].font      = _Font(name="Calibri", bold=True, size=11, color=DARK_HEX)
        ws[f"A{row_num}"].fill      = _Fill("solid", fgColor="FFF8E1")
        ws[f"A{row_num}"].alignment = center
        ws[f"A{row_num}"].border    = _Border(
            left=_Side(style="medium", color=YELLOW_HEX),
            right=_Side(style="medium", color=YELLOW_HEX),
            top=_Side(style="medium", color=YELLOW_HEX),
            bottom=_Side(style="thin", color=YELLOW_HEX),
        )
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        def _fmt_pal2(v):
            return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        prod_data = [
            ("Pallets a clasificar",  _fmt_pal2(prod.get("pallets_total", 0))),
            ("Target x operario",     str(prod.get("target_oper", 0))),
            ("Pallets / hora",        str(prod.get("pall_hora", 0))),
            ("Personal necesario",    str(prod.get("personal_nec", 0))),
            ("Duracion estimada",     str(prod.get("duracion", "00:00"))),
        ]
        for pi, (label, val) in enumerate(prod_data):
            lbl_cell = ws.cell(row=row_num, column=1, value=label)
            lbl_cell.font      = _Font(name="Calibri", size=9, color=GRAY_HEX)
            lbl_cell.fill      = _Fill("solid", fgColor="FFF8E1")
            lbl_cell.alignment = left
            # Valor en col 2-3 merged
            ws.merge_cells(f"B{row_num}:C{row_num}")
            val_cell = ws.cell(row=row_num, column=2, value=val)
            val_cell.font      = _Font(name="Calibri", bold=True, size=12, color=DARK_HEX)
            val_cell.fill      = _Fill("solid", fgColor="FFF8E1")
            val_cell.alignment = center
            ws.row_dimensions[row_num].height = 16
            row_num += 1

        # Conclusión
        ws.merge_cells(f"A{row_num}:C{row_num}")
        concl_txt = (
            f"Proximo turno: {prod.get('personal_nec',0)} operarios · "
            f"{prod.get('pallets_total',0):.2f} pall · "
            f"duracion estimada {prod.get('duracion','00:00')} hs."
        )
        ws[f"A{row_num}"] = concl_txt
        ws[f"A{row_num}"].font      = _Font(name="Calibri", italic=True, size=8, color=GRAY_HEX)
        ws[f"A{row_num}"].fill      = _Fill("solid", fgColor="FFF8E1")
        ws[f"A{row_num}"].alignment = center
        ws[f"A{row_num}"].border    = _Border(
            left=_Side(style="medium", color=YELLOW_HEX),
            right=_Side(style="medium", color=YELLOW_HEX),
            bottom=_Side(style="medium", color=YELLOW_HEX),
        )
        ws.row_dimensions[row_num].height = 14

    # ── ANCHOS DE COLUMNA ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # ── FOOTER INFO (en A al final) ───────────────────────────────────────────
    row_num += 2
    ws.merge_cells(f"A{row_num}:C{row_num}")
    ws[f"A{row_num}"] = (
        f"Fuente: ANR + Frescura DDM (alm 1/3, BXP) — "
        f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws[f"A{row_num}"].font      = _Font(name="Calibri", italic=True, size=7, color=GRAY_HEX)
    ws[f"A{row_num}"].alignment = left

    bio = _io.BytesIO()
    wb.save(bio)
    return bio.getvalue()



    """
    PDF a color vertical A4 con la clasificación por camión.
    Diseño inspirado en la planilla del Sheets (header azul, filas alternadas,
    semáforo en pallets, total destacado). Omite camiones con 0 carga.
    Muestra desglose: bultos enteros + unids sueltas + cajones extra + total + pallets.
    """
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib import colors as _col
    from reportlab.lib.units import mm as _mm
    from reportlab.pdfgen import canvas as _cv
    import io as _io

    # Filtrar camiones con 0 bultos (omitir camiones sin carga)
    df = df_cl[df_cl["Bultos Retornables"] > 0].copy()
    if df.empty:
        bio = _io.BytesIO()
        c = _cv.Canvas(bio, pagesize=_A4)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40 * _mm, 250 * _mm, "Sin camiones con retornables hoy")
        c.save()
        return bio.getvalue()

    PW, PH = _A4
    M = 10 * _mm

    DARK = _col.HexColor("#1a3a6b")
    YELLOW_HDR = _col.HexColor("#f8d772")
    BAND_LIGHT = _col.HexColor("#f4f6fa")
    BAND_WHITE = _col.white
    RED_BG  = _col.HexColor("#f5b7b1")
    GREEN_BG = _col.HexColor("#a9dfbf")
    AMBER_BG = _col.HexColor("#fad7a0")
    BORDER  = _col.HexColor("#8c8c8c")
    TOTAL_BG = _col.HexColor("#1a3a6b")

    bio = _io.BytesIO()
    c = _cv.Canvas(bio, pagesize=_A4)

    # ─ Header banner ─
    HDR_H = 22 * _mm
    c.setFillColor(DARK)
    c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)
    c.setFillColor(_col.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(M, PH - 11 * _mm, "CLASIFICACIÓN — Retornables por Camión")
    c.setFont("Helvetica", 10)
    c.drawString(M, PH - 17 * _mm, "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén · Alm 1 (x12) + Alm 3 (x24)")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(PW - M, PH - 11 * _mm, f"Fecha: {fecha_str}")
    c.setFont("Helvetica", 9)
    c.drawRightString(PW - M, PH - 17 * _mm, f"Camiones con carga: {len(df)}")

    # ─ Tabla ─
    y = PH - HDR_H - 8 * _mm
    table_w = PW - 2 * M

    # Columnas: Camión | Enteros | Un.Sueltas | Caj.Extra | TOTAL | Pallets | Estado
    col_widths = [
        table_w * 0.10,   # Camión
        table_w * 0.14,   # Bultos enteros
        table_w * 0.14,   # Un. sueltas
        table_w * 0.14,   # Cajones extra
        table_w * 0.16,   # Total bultos
        table_w * 0.16,   # Pallets
        table_w * 0.16,   # Estado
    ]
    col_x = [M]
    for w in col_widths[:-1]:
        col_x.append(col_x[-1] + w)

    # Header tabla (amarillo)
    THDR_H = 9 * _mm
    c.setFillColor(YELLOW_HDR)
    c.rect(M, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
    c.setStrokeColor(BORDER)
    c.setFillColor(DARK)
    headers = ["CAMIÓN", "BULTOS\nENTEROS", "UN.\nSUELTAS", "CAJONES\nEXTRA", "TOTAL\nBULTOS", "PALLETS\n(÷50)", "ESTADO"]
    c.setFont("Helvetica-Bold", 8.5)
    for i, h in enumerate(headers):
        lines = h.split("\n")
        cy = y - THDR_H + 5.5 * _mm if len(lines) == 2 else y - THDR_H + 3 * _mm
        for j, ln in enumerate(lines):
            c.drawCentredString(col_x[i] + col_widths[i] / 2, cy - j * 3.2 * _mm, ln)
    y -= THDR_H

    # Estadística para semáforo: percentiles de pallets
    pall_vals = sorted(df["Pallets Equivalentes"].values)
    n = len(pall_vals)
    p_high = pall_vals[int(n * 0.66)] if n > 0 else 0
    p_low  = pall_vals[int(n * 0.33)] if n > 0 else 0

    ROW_H = 7.5 * _mm

    for i, row in df.reset_index(drop=True).iterrows():
        bg = BAND_LIGHT if i % 2 == 0 else BAND_WHITE
        c.setFillColor(bg)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=1, stroke=0)

        pal = float(row["Pallets Equivalentes"])
        if pal >= p_high and pal > 0:
            est_bg, est_txt, est_color = GREEN_BG, "ALTA", _col.HexColor("#0e6b3c")
        elif pal <= p_low and pal > 0:
            est_bg, est_txt, est_color = RED_BG, "BAJA", _col.HexColor("#8b1a0e")
        else:
            est_bg, est_txt, est_color = AMBER_BG, "MEDIA", _col.HexColor("#7d5a00")

        # Pintar bg semáforo en Pallets y Estado
        c.setFillColor(est_bg)
        c.rect(col_x[5], y - ROW_H, col_widths[5], ROW_H, fill=1, stroke=0)
        c.rect(col_x[6], y - ROW_H, col_widths[6], ROW_H, fill=1, stroke=0)

        # Bordes
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.3)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=0, stroke=1)
        for cx in col_x[1:]:
            c.line(cx, y - ROW_H, cx, y)

        # Datos
        bul_ent  = float(row.get("_bultos_enteros", 0))
        unids    = float(row.get("_unids", 0))
        caj_ext  = float(row.get("_cajones_extra", 0))
        tot_bul  = float(row["Bultos Retornables"])

        c.setFillColor(_col.black)
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(col_x[0] + col_widths[0] / 2, y - ROW_H + 2.5 * _mm, str(row["Camión"]))
        c.setFont("Helvetica", 9.5)
        c.drawCentredString(col_x[1] + col_widths[1] / 2, y - ROW_H + 2.5 * _mm,
                            f"{bul_ent:,.0f}".replace(",", "."))
        c.drawCentredString(col_x[2] + col_widths[2] / 2, y - ROW_H + 2.5 * _mm,
                            f"{unids:,.0f}".replace(",", ".") if unids > 0 else "—")
        c.drawCentredString(col_x[3] + col_widths[3] / 2, y - ROW_H + 2.5 * _mm,
                            f"+{caj_ext:,.0f}".replace(",", ".") if caj_ext > 0 else "—")
        c.setFont("Helvetica-Bold", 10.5)
        c.drawCentredString(col_x[4] + col_widths[4] / 2, y - ROW_H + 2.5 * _mm,
                            f"{tot_bul:,.0f}".replace(",", "."))
        c.drawCentredString(col_x[5] + col_widths[5] / 2, y - ROW_H + 2.5 * _mm,
                            f"{pal:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        c.setFillColor(est_color)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(col_x[6] + col_widths[6] / 2, y - ROW_H + 2.5 * _mm, est_txt)

        y -= ROW_H

        if y < 35 * _mm:
            c.showPage()
            c.setFillColor(DARK)
            c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)
            c.setFillColor(_col.white)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(M, PH - 11 * _mm, "CLASIFICACIÓN — Retornables por Camión (cont.)")
            c.setFont("Helvetica", 10)
            c.drawString(M, PH - 17 * _mm, "Beccacece Hnos SA · DPO 2.1")
            c.drawRightString(PW - M, PH - 11 * _mm, f"Fecha: {fecha_str}")
            y = PH - HDR_H - 8 * _mm
            c.setFillColor(YELLOW_HDR)
            c.rect(M, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
            c.setFillColor(DARK)
            c.setFont("Helvetica-Bold", 8.5)
            for i2, h in enumerate(headers):
                lines = h.split("\n")
                cy = y - THDR_H + 5.5 * _mm if len(lines) == 2 else y - THDR_H + 3 * _mm
                for j, ln in enumerate(lines):
                    c.drawCentredString(col_x[i2] + col_widths[i2] / 2, cy - j * 3.2 * _mm, ln)
            y -= THDR_H

    # ─ Fila TOTAL ─
    TOT_H = 10 * _mm
    c.setFillColor(TOTAL_BG)
    c.rect(M, y - TOT_H, table_w, TOT_H, fill=1, stroke=1)
    c.setFillColor(_col.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(col_x[0] + col_widths[0] / 2, y - TOT_H + 3.5 * _mm, "TOTAL")
    c.drawCentredString(col_x[1] + col_widths[1] / 2, y - TOT_H + 3.5 * _mm,
                        f"{float(tot.get('bultos_enteros', 0)):,.0f}".replace(",", "."))
    un_t = float(tot.get('unids', 0))
    cj_t = float(tot.get('cajones_extra', 0))
    c.drawCentredString(col_x[2] + col_widths[2] / 2, y - TOT_H + 3.5 * _mm,
                        f"{un_t:,.0f}".replace(",", ".") if un_t > 0 else "—")
    c.drawCentredString(col_x[3] + col_widths[3] / 2, y - TOT_H + 3.5 * _mm,
                        f"+{cj_t:,.0f}".replace(",", ".") if cj_t > 0 else "—")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(col_x[4] + col_widths[4] / 2, y - TOT_H + 3.5 * _mm,
                        f"{float(tot['bultos']):,.0f}".replace(",", "."))
    c.drawCentredString(col_x[5] + col_widths[5] / 2, y - TOT_H + 3.5 * _mm,
                        f"{tot['pallets']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(col_x[6] + col_widths[6] / 2, y - TOT_H + 3.5 * _mm, f"{tot['camiones']} cam.")
    y -= TOT_H

    # ─ Caja de Productividad Estimada ─────────────────────────────────────
    prod = tot.get("productividad") or {}
    if prod:
        y -= 6 * _mm
        BOX_H = 30 * _mm
        c.setFillColor(_col.HexColor("#fff8e1"))  # crema suave
        c.setStrokeColor(YELLOW_HDR)
        c.setLineWidth(1.2)
        c.rect(M, y - BOX_H, table_w, BOX_H, fill=1, stroke=1)

        # Título caja
        c.setFillColor(DARK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(M + 4 * _mm, y - 6 * _mm, "⏱  PRODUCTIVIDAD ESTIMADA DE CLASIFICACIÓN")

        # Grid 5 columnas
        cell_w = table_w / 5
        y_label = y - 12 * _mm
        y_value = y - 19 * _mm
        labels = [
            ("Pallets a clasificar", f"{prod.get('pallets_total', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
            ("Target × operario",    f"{prod.get('target_oper', 0)}"),
            ("Pallets / hora",       f"{prod.get('pall_hora', 0)}"),
            ("Personal necesario",   f"{prod.get('personal_nec', 0)}"),
            ("Duración estimada",    f"{prod.get('duracion', '00:00')}"),
        ]
        for i, (lbl, val) in enumerate(labels):
            cx = M + i * cell_w + cell_w / 2
            c.setFillColor(_col.HexColor("#555555"))
            c.setFont("Helvetica", 8.5)
            c.drawCentredString(cx, y_label, lbl)
            c.setFillColor(DARK)
            c.setFont("Helvetica-Bold", 14 if i in (3, 4) else 12)
            c.drawCentredString(cx, y_value, val)

        # Conclusión
        c.setFillColor(_col.HexColor("#444444"))
        c.setFont("Helvetica-Oblique", 8)
        concl = (f"Próximo turno: {prod.get('personal_nec',0)} operarios · "
                 f"{prod.get('pallets_total',0):.2f} pall · duración estimada {prod.get('duracion','00:00')} hs.")
        c.drawCentredString(M + table_w/2, y - 26 * _mm, concl)
        y -= BOX_H

    # ─ Footer ─
    c.setFillColor(_col.HexColor("#555555"))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(M, 10 * _mm,
                 "Total bultos = Enteros + Cajones extra (un. sueltas ÷12 Alm1 / ÷24 Alm3). "
                 "Pallets = Total ÷ 50. Filtros: anulados off, SKUs DDM.")
    c.drawRightString(PW - M, 10 * _mm,
                      f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.save()
    return bio.getvalue()


def build_top_skus_pdf(df_top: pd.DataFrame, fecha_str: str = "") -> bytes:
    """
    PDF a color vertical A4 con el ranking de SKUs por bultos. Diseño tipo
    tabla con header azul, ABC coloreado (verde/amarillo/rojo), totales.
    """
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib import colors as _col
    from reportlab.lib.units import mm as _mm
    from reportlab.pdfgen import canvas as _cv
    import io as _io

    PW, PH = _A4
    M = 12 * _mm

    DARK = _col.HexColor("#1a3a6b")
    MED  = _col.HexColor("#2e5fa3")
    GOLD = _col.HexColor("#f8d772")
    BAND_LIGHT = _col.HexColor("#f4f6fa")
    BAND_WHITE = _col.white
    BORDER = _col.HexColor("#8c8c8c")
    ABC_A = _col.HexColor("#a9dfbf")
    ABC_B = _col.HexColor("#fad7a0")
    ABC_C = _col.HexColor("#f5b7b1")
    TOTAL_BG = _col.HexColor("#1a3a6b")

    bio = _io.BytesIO()
    c = _cv.Canvas(bio, pagesize=_A4)

    # ─ Header banner ─
    HDR_H = 22 * _mm
    c.setFillColor(DARK)
    c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(M, PH - 11 * _mm, "TOP SKUs — Día de Carga")
    c.setFillColor(_col.white)
    c.setFont("Helvetica", 10)
    c.drawString(M, PH - 17 * _mm, "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(PW - M, PH - 11 * _mm, f"Fecha: {fecha_str}")
    tot_bultos = float(df_top["BULTOS"].sum()) if "BULTOS" in df_top.columns else 0
    tot_hl     = float(df_top["HL"].sum()) if "HL" in df_top.columns else 0
    c.setFont("Helvetica", 9)
    c.drawRightString(PW - M, PH - 17 * _mm,
                      f"Bultos: {tot_bultos:,.0f}  ·  HL: {tot_hl:,.2f}".replace(",", "."))

    # ─ Tabla ─
    y = PH - HDR_H - 8 * _mm
    table_w = PW - 2 * M
    # Columnas: # | Código | Descripción | Bultos | HL | Rubro  (sin ABC, sin "#" duplicada)
    col_widths = [
        table_w * 0.06,   # #
        table_w * 0.12,   # Código
        table_w * 0.44,   # Descripción
        table_w * 0.13,   # Bultos
        table_w * 0.12,   # HL
        table_w * 0.13,   # Rubro
    ]
    col_x = [M]
    for w in col_widths[:-1]:
        col_x.append(col_x[-1] + w)

    # Header tabla
    THDR_H = 8 * _mm
    c.setFillColor(GOLD)
    c.rect(M, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
    c.setStrokeColor(BORDER)
    c.setFillColor(DARK)
    c.setFont("Helvetica-Bold", 9.5)
    headers = ["#", "CÓDIGO", "DESCRIPCIÓN", "BULTOS", "HL", "RUBRO"]
    for i, h in enumerate(headers):
        c.drawCentredString(col_x[i] + col_widths[i] / 2, y - THDR_H + 2.5 * _mm, h)
    y -= THDR_H

    ROW_H = 8 * _mm

    for i, row in df_top.reset_index(drop=True).iterrows():
        bg = BAND_LIGHT if i % 2 == 0 else BAND_WHITE
        c.setFillColor(bg)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=1, stroke=0)

        # Bordes
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.3)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=0, stroke=1)
        for cx in col_x[1:]:
            c.line(cx, y - ROW_H, cx, y)

        # Texto
        c.setFillColor(_col.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(col_x[0] + col_widths[0] / 2, y - ROW_H + 2.7 * _mm, str(i + 1))
        c.setFont("Helvetica", 9)
        c.drawCentredString(col_x[1] + col_widths[1] / 2, y - ROW_H + 2.7 * _mm,
                            str(int(row["CODIGO"])) if pd.notna(row["CODIGO"]) else "—")
        # Descripción truncada
        desc = str(row.get("DESCRIPCION", "")).strip()[:55]
        c.drawString(col_x[2] + 1.5 * _mm, y - ROW_H + 2.7 * _mm, desc)
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(col_x[3] + col_widths[3] - 1.5 * _mm, y - ROW_H + 2.7 * _mm,
                          f"{float(row['BULTOS']):,.0f}".replace(",", "."))
        c.setFont("Helvetica", 9)
        hl_val = float(row.get("HL", 0)) if pd.notna(row.get("HL", 0)) else 0
        c.drawRightString(col_x[4] + col_widths[4] - 1.5 * _mm, y - ROW_H + 2.7 * _mm,
                          f"{hl_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if hl_val else "—")
        rubro = str(row.get("RUBRO", "")).strip() if "RUBRO" in df_top.columns else ""
        # Truncar a 11 chars para que entre cómodamente en la columna
        if len(rubro) > 11:
            rubro = rubro[:11]
        c.drawCentredString(col_x[5] + col_widths[5] / 2, y - ROW_H + 2.7 * _mm, rubro)

        y -= ROW_H

    # Fila TOTAL
    TOT_H = 9 * _mm
    c.setFillColor(TOTAL_BG)
    c.rect(M, y - TOT_H, table_w, TOT_H, fill=1, stroke=1)
    c.setFillColor(_col.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(col_x[0] + 2 * _mm, y - TOT_H + 3 * _mm, "TOTAL TOP " + str(len(df_top)))
    c.drawRightString(col_x[3] + col_widths[3] - 1.5 * _mm, y - TOT_H + 3 * _mm,
                      f"{tot_bultos:,.0f}".replace(",", "."))
    c.drawRightString(col_x[4] + col_widths[4] - 1.5 * _mm, y - TOT_H + 3 * _mm,
                      f"{tot_hl:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    # Footer
    c.setFillColor(_col.HexColor("#555555"))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(M, 10 * _mm, "Universo: SKUs en DDM (almacén). Excluye envases/esqueletos (2776, 2731, 2730, 2780, etc.).")
    c.drawRightString(PW - M, 10 * _mm, f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.save()
    return bio.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# TAB CLASIFICACIÓN (v4.11) — auto-disparo + PDF a color
# ════════════════════════════════════════════════════════════════════════════

# ── Constantes productividad clasificación (Sheets DPO) ────────────────────
TARGET_PALL_OPERARIO = 25   # pallets objetivo por operario
PALL_X_HORA           = 7    # pallets/hora informativo


def _duracion_hhmm(ratio_horas: float) -> str:
    """1.26 → '01:16'. Convierte un decimal de horas a formato HH:MM.
    Replica la fórmula del Sheets: entero(h) y redondeo de la fracción × 60.
    """
    import math as _m
    if ratio_horas is None or ratio_horas <= 0 or not _m.isfinite(ratio_horas):
        return "00:00"
    h = int(_m.floor(ratio_horas))
    m = int(round((ratio_horas - h) * 60))
    if m == 60:
        h += 1
        m = 0
    return f"{h:02d}:{m:02d}"


def render_tab_clasificacion():
    st.subheader("🏷️ Clasificación — Retornables por Camión (Almacén 1 y 3)")
    st.caption(
        "v4.18 — Input: **ANR.xlsx** y **Frescura 3.0** (desde 📁 Archivos). "
        "Filtra los SKUs de **ALMACÉN ∈ {1, 3}** (envases vacíos + esqueletos) y muestra "
        "**cuánto retorna cada camión** en bultos y pallets. Cálculo per‑fila: "
        "`pallets = bultos / BXP(SKU)`."
    )

    # ANR: desde pestaña Archivos
    anr_use = st.session_state.get("tc_anr")

    # Frescura: SOLO desde Planilla de Carga / Archivos
    fr_use = st.session_state.get("t1_fr")

    if not anr_use:
        st.info("⬅️ Subí el **ANR.xlsx** en la pestaña **📁 Archivos** para generar la clasificación.")
        return
    if not fr_use:
        st.warning(
            "⚠️ Falta la **Frescura 3.0**. Subila en la pestaña **📁 Archivos** "
            "(este tab la reusa de ahí, no tiene uploader propio)."
        )
        return

    try:
        df_cl, tot = _build_clasificacion_anr(
            anr_use.getvalue(), fr_use.getvalue()
        )
    except Exception as e:
        st.error(f"❌ Error construyendo Clasificación: {e}")
        with st.expander("Stack trace"):
            import traceback
            st.code(traceback.format_exc())
        return

    if df_cl.empty:
        st.warning(
            "No se encontraron SKUs de Almacén 1 o 3 con ventas en el ANR. "
            "Verificá que la Frescura tenga la hoja DDM y que ANR tenga la col AP (Transporte)."
        )
        return

    # ── Métricas top ──────────────────────────────────────────────────────
    m1, m2, m3 = st.columns(3)
    m1.metric("🚛 Camiones con retorno", tot["camiones"])
    m2.metric("📦 Bultos a clasificar", f"{tot['bultos']:,.0f}".replace(",", "."))
    m3.metric("🟫 Pallets a clasificar",
              f"{tot['pallets']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
              help=f"Σ (bultos_fila / BXP_SKU) — desde {tot['skus']} SKUs retornables distintos")

    # ── Tabla por CAMIÓN ─────────────────────────────────────────────────
    df_disp = pd.DataFrame({
        "Camión":  df_cl["CAMION"].astype(int).astype(str),
        "Bultos":  df_cl["BULTOS"].map(lambda x: f"{x:,.0f}".replace(",", ".")),
        "Pallets": df_cl["PALLETS"].map(
            lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        ),
    })
    df_disp.loc[len(df_disp)] = [
        "TOTAL",
        f"{tot['bultos']:,.0f}".replace(",", "."),
        f"{tot['pallets']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
    ]
    st.dataframe(
        df_disp,
        width="stretch",
        height=min(440, 38 * (len(df_disp) + 1) + 38),
        hide_index=True,
        column_config={
            "Camión":  st.column_config.TextColumn(width="small"),
            "Bultos":  st.column_config.TextColumn(width="small"),
            "Pallets": st.column_config.TextColumn(width="small"),
        },
    )

    # ── Bloque Productividad Estimada ──────────────────────────────────────
    import math as _m
    st.markdown("### ⏱️ Productividad Estimada de Clasificación")

    pallets_total = float(tot["pallets"])
    ratio         = pallets_total / TARGET_PALL_OPERARIO if TARGET_PALL_OPERARIO else 0
    # Fórmula del Sheets: ENTERO(ratio) + 1 (siempre redondea agregando uno)
    personal_nec  = (int(_m.floor(ratio)) + 1) if ratio > 0 else 0
    duracion      = _duracion_hhmm(ratio)

    p1, p2, p3, p4, p5 = st.columns(5)
    p1.metric("Pallets a clasificar", f"{pallets_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    p2.metric("Target × operario",    f"{TARGET_PALL_OPERARIO}")
    p3.metric("Pallets / hora",       f"{PALL_X_HORA}")
    p4.metric("Personal necesario",   f"{personal_nec}",
              help=f"ratio crudo = {ratio:.4f} → ENTERO + 1 (réplica Sheets)")
    p5.metric("Duración estimada",    duracion,
              help=f"{pallets_total:.2f} pall ÷ {TARGET_PALL_OPERARIO} target = {ratio:.4f} hs → HH:MM")

    st.caption(
        f"📌 Próximo turno: **{personal_nec} operarios** para clasificar "
        f"**{pallets_total:,.2f} pallets** en **{duracion} hs** estimadas."
    )

    # Guardar para el PDF
    tot["productividad"] = {
        "pallets_total":  pallets_total,
        "target_oper":    TARGET_PALL_OPERARIO,
        "pall_hora":      PALL_X_HORA,
        "personal_nec":   personal_nec,
        "ratio":          ratio,
        "duracion":       duracion,
    }

    # ── PDF + XLSX + TSV ──────────────────────────────────────────────────
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    try:
        pdf_bytes = build_clasificacion_anr_pdf(df_cl, tot, fecha_str=fecha_hoy)
    except Exception as e:
        pdf_bytes = None
        st.error(f"❌ Error generando PDF Clasificación: {e}")

    try:
        xlsx_bytes_cl = build_clasificacion_anr_xlsx(df_cl, tot, fecha_str=fecha_hoy)
    except Exception as e:
        xlsx_bytes_cl = None
        st.error(f"❌ Error generando Excel Clasificación: {e}")

    col_dl1, col_dl2, col_dl3 = st.columns(3)
    if pdf_bytes:
        col_dl1.download_button(
            "⬇ PDF Clasificación",
            data=pdf_bytes,
            file_name=_stamp("Clasificacion_PorCamion", "pdf"),
            mime="application/pdf",
            type="primary",
            width="stretch",
            key="clasif_pdf_dl",
        )
    if xlsx_bytes_cl:
        col_dl2.download_button(
            "📊 Excel Clasificación",
            data=xlsx_bytes_cl,
            file_name=_stamp("Clasificacion_PorCamion", "xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            width="stretch",
            key="clasif_xlsx_dl",
        )
    tsv = df_cl.to_csv(sep="\t", index=False)
    col_dl3.download_button(
        "📋 TSV (Sheets)",
        data=tsv.encode("utf-8"),
        file_name="clasificacion_por_camion.tsv",
        mime="text/tab-separated-values",
        width="stretch",
        key="clasif_tsv_dl",
    )

    with st.expander("ℹ️ Metodología"):
        st.markdown(
            "- **Input**: ANR.xlsx hoja **BASE** (header fila 2): col H=ARTÍCULO, K=BULTOS, "
            "**AP=TRANSPORTE (camión)**.\n"
            "- **Universo retornables**: Frescura hoja **DDM** filtrando `ALMACÉN ∈ {1, 3}`, "
            "BXP por SKU desde col G.\n"
            "- **Cálculo per‑fila**: `pallets_fila = bultos_fila / BXP_SKU` "
            "(no se promedia BXP, se usa el del SKU exacto).\n"
            "- **Agregación**: Σ bultos y Σ pallets por **CAMIÓN**.\n"
            "- **Personal necesario**: `ENTERO(total_pallets / target_x_operario) + 1` "
            "(idéntico a `=+ENTERO(E43)+1` del Sheets).\n"
            f"- **Constantes**: target × operario = {TARGET_PALL_OPERARIO}, pall/hora = {PALL_X_HORA}."
        )


# ════════════════════════════════════════════════════════════════════════════
# TAB TOP SKUs (v4.11) — tab propia con auto-disparo
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# TAB TOPS desde ANR (v4.12) — Top SKUs + Top Clientes
# Lee Excel ANR (Análisis de Rechazos / ventas próximo día) y arma dos tops.
# ════════════════════════════════════════════════════════════════════════════

def _build_top10_skus_anr(anr_bytes: bytes, n: int = 10) -> pd.DataFrame:
    """Top N SKUs por BULTOS desc, agregando desde hoja 'BASE' del ANR.
    Cols BASE: H=ARTÍCULO, I=DESCRIPCIÓN ARTÍCULO, K=BULTOS, M=IMPORTE NETO, O=UNIDAD DE MEDIDA (HL).
    Devuelve DataFrame con columnas: CODIGO, DESCRIPCION, BULTOS, IMPORTE, HL.
    """
    df = pd.read_excel(io.BytesIO(anr_bytes), sheet_name="BASE", header=1)
    if df.empty or df.shape[1] < 15:
        return pd.DataFrame(columns=["CODIGO", "DESCRIPCION", "BULTOS", "IMPORTE", "HL"])

    # Columnas posicionales H=7, I=8, K=10, M=12, O=14
    sub = df.iloc[:, [7, 8, 10, 12, 14]].copy()
    sub.columns = ["CODIGO", "DESCRIPCION", "BULTOS", "IMPORTE", "HL"]

    sub["CODIGO"]  = pd.to_numeric(sub["CODIGO"],  errors="coerce")
    sub["BULTOS"]  = pd.to_numeric(sub["BULTOS"],  errors="coerce").fillna(0)
    sub["IMPORTE"] = pd.to_numeric(sub["IMPORTE"], errors="coerce").fillna(0)
    sub["HL"]      = pd.to_numeric(sub["HL"],      errors="coerce").fillna(0)
    sub = sub[sub["CODIGO"].notna() & (sub["BULTOS"] > 0)]

    # Tomar la descripción más larga como representativa (más informativa)
    sub["DESCRIPCION"] = sub["DESCRIPCION"].astype(str).fillna("")
    agg = sub.groupby("CODIGO", as_index=False).agg(
        DESCRIPCION=("DESCRIPCION", lambda x: max(x, key=len) if len(x) else ""),
        BULTOS=("BULTOS", "sum"),
        IMPORTE=("IMPORTE", "sum"),
        HL=("HL", "sum"),
    )
    agg["CODIGO"] = agg["CODIGO"].astype(int)
    agg = agg.sort_values("BULTOS", ascending=False).reset_index(drop=True)
    return agg.head(n)


def _build_division_summary_anr(anr_bytes: bytes) -> pd.DataFrame:
    """Resumen por DESCRIPCIÓN DIVISION (col AT) desde hoja BASE.
    Devuelve: DIVISION, BULTOS, IMPORTE, HL — ordenado por BULTOS desc.
    """
    df = pd.read_excel(io.BytesIO(anr_bytes), sheet_name="BASE", header=1)
    if df.empty or df.shape[1] < 46:
        return pd.DataFrame(columns=["DIVISION", "BULTOS", "IMPORTE", "HL"])

    # K=10, M=12, O=14, AT=45
    sub = df.iloc[:, [45, 10, 12, 14]].copy()
    sub.columns = ["DIVISION", "BULTOS", "IMPORTE", "HL"]
    sub["BULTOS"]  = pd.to_numeric(sub["BULTOS"],  errors="coerce").fillna(0)
    sub["IMPORTE"] = pd.to_numeric(sub["IMPORTE"], errors="coerce").fillna(0)
    sub["HL"]      = pd.to_numeric(sub["HL"],      errors="coerce").fillna(0)
    sub = sub[sub["DIVISION"].notna() & (sub["DIVISION"].astype(str).str.strip() != "")]

    agg = sub.groupby("DIVISION", as_index=False).agg(
        BULTOS=("BULTOS", "sum"),
        IMPORTE=("IMPORTE", "sum"),
        HL=("HL", "sum"),
    )
    agg = agg.sort_values("BULTOS", ascending=False).reset_index(drop=True)
    return agg


def _build_top10_clientes_anr(anr_bytes: bytes, n: int = 10) -> pd.DataFrame:
    """Top N clientes por BULTOS desc desde hoja 'CLIENTES' (ya preagregada).
    Cols: A=Cliente, B=Bultos venta, E=Importe neto, H=Unidad medida (HL).
    Devuelve: CLIENTE, BULTOS, IMPORTE, HL.
    """
    df = pd.read_excel(io.BytesIO(anr_bytes), sheet_name="CLIENTES", header=1)
    if df.empty or df.shape[1] < 8:
        return pd.DataFrame(columns=["CLIENTE", "BULTOS", "IMPORTE", "HL"])

    sub = df.iloc[:, [0, 1, 4, 7]].copy()
    sub.columns = ["CLIENTE", "BULTOS", "IMPORTE", "HL"]

    # Filtrar filas válidas: cliente no nulo, no "-", no "Total general"
    sub = sub[sub["CLIENTE"].notna()]
    sub["CLIENTE"] = sub["CLIENTE"].astype(str).str.strip()
    sub = sub[~sub["CLIENTE"].isin(["-", "Total general", "(en blanco)", ""])]
    sub = sub[~sub["CLIENTE"].str.lower().str.startswith("total")]

    sub["BULTOS"]  = pd.to_numeric(sub["BULTOS"],  errors="coerce").fillna(0)
    sub["IMPORTE"] = pd.to_numeric(sub["IMPORTE"], errors="coerce").fillna(0)
    sub["HL"]      = pd.to_numeric(sub["HL"],      errors="coerce").fillna(0)
    sub = sub[sub["BULTOS"] > 0]

    sub = sub.sort_values("BULTOS", ascending=False).reset_index(drop=True)
    return sub.head(n)


def _fmt_money(v: float) -> str:
    """Formato $ Argentina: $ 12.345.678,90"""
    if v is None or pd.isna(v):
        return "—"
    s = f"{float(v):,.2f}"
    return "$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_num(v: float, dec: int = 2) -> str:
    if v is None or pd.isna(v):
        return "—"
    s = f"{float(v):,.{dec}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _pdf_to_jpeg(pdf_bytes: bytes, dpi: int = 180) -> bytes:
    """
    Convierte un PDF (una o varias páginas) en un JPEG único.
    Si hay varias páginas las apila verticalmente.
    Estrategia en cascada sin dependencias externas obligatorias:
      1. pdf2image + poppler (disponible en Streamlit Cloud)
      2. PyMuPDF / fitz (alternativa sin poppler)
    """
    import io as _io

    # Intento 1: pdf2image (requiere poppler — disponible en Streamlit Cloud)
    try:
        from pdf2image import convert_from_bytes
        pages = convert_from_bytes(pdf_bytes, dpi=dpi, fmt="RGB")
        if not pages:
            raise ValueError("pdf2image: sin páginas")
        from PIL import Image as _PILImage
        if len(pages) == 1:
            img = pages[0]
        else:
            total_h = sum(p.height for p in pages)
            max_w   = max(p.width for p in pages)
            img = _PILImage.new("RGB", (max_w, total_h), "white")
            y_off = 0
            for p in pages:
                img.paste(p, (0, y_off))
                y_off += p.height
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=92, optimize=True)
        return buf.getvalue()
    except Exception:
        pass

    # Intento 2: PyMuPDF / fitz (sin poppler)
    try:
        import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        from PIL import Image as _PILImage
        imgs = []
        for page in doc:
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = _PILImage.frombytes("RGB", (pix.width, pix.height), pix.samples)
            imgs.append(img)
        doc.close()
        if not imgs:
            raise ValueError("fitz: sin páginas")
        if len(imgs) == 1:
            final = imgs[0]
        else:
            total_h = sum(i.height for i in imgs)
            max_w   = max(i.width for i in imgs)
            final = _PILImage.new("RGB", (max_w, total_h), "white")
            y_off = 0
            for i in imgs:
                final.paste(i, (0, y_off))
                y_off += i.height
        buf = _io.BytesIO()
        final.save(buf, format="JPEG", quality=92, optimize=True)
        return buf.getvalue()
    except Exception:
        pass

    raise RuntimeError(
        "No se pudo convertir el PDF a JPEG. "
        "Instalá pdf2image+poppler o PyMuPDF (pip install pymupdf) en el entorno."
    )


def build_top_skus_anr_pdf(df_top: pd.DataFrame, df_div: pd.DataFrame,
                           fecha_str: str = "") -> bytes:
    """PDF a color: Top SKUs (arriba) + Resumen por División (abajo)."""
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib import colors as _col
    from reportlab.lib.units import mm as _mm
    from reportlab.pdfgen import canvas as _cv
    import io as _io

    PW, PH = _A4
    M = 12 * _mm

    DARK = _col.HexColor("#1a3a6b")
    MED  = _col.HexColor("#2e5fa3")
    GOLD = _col.HexColor("#f8d772")
    BAND_LIGHT = _col.HexColor("#f4f6fa")
    BAND_WHITE = _col.white
    BORDER = _col.HexColor("#8c8c8c")
    TOTAL_BG = _col.HexColor("#1a3a6b")
    DIV_HDR = _col.HexColor("#2e5fa3")

    bio = _io.BytesIO()
    c = _cv.Canvas(bio, pagesize=_A4)

    # ─ Header banner ─
    HDR_H = 22 * _mm
    c.setFillColor(DARK)
    c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(M, PH - 11 * _mm, "TOP SKUs — Venta del Día (ANR)")
    c.setFillColor(_col.white)
    c.setFont("Helvetica", 10)
    c.drawString(M, PH - 17 * _mm, "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(PW - M, PH - 11 * _mm, f"Fecha: {fecha_str}")

    # ─ Tabla Top SKUs ─
    y = PH - HDR_H - 8 * _mm
    table_w = PW - 2 * M
    # # | Código | Descripción | Bultos | HL
    col_widths = [
        table_w * 0.05,   # #
        table_w * 0.12,   # Código
        table_w * 0.63,   # Descripción
        table_w * 0.11,   # Bultos
        table_w * 0.09,   # HL
    ]
    col_x = [M]
    for w in col_widths[:-1]:
        col_x.append(col_x[-1] + w)

    THDR_H = 8 * _mm
    c.setFillColor(GOLD)
    c.rect(M, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
    c.setStrokeColor(BORDER)
    c.setFillColor(DARK)
    c.setFont("Helvetica-Bold", 9.5)
    headers = ["#", "CÓDIGO", "DESCRIPCIÓN", "BULTOS", "HL"]
    for i, h in enumerate(headers):
        c.drawCentredString(col_x[i] + col_widths[i] / 2, y - THDR_H + 2.5 * _mm, h)
    y -= THDR_H

    ROW_H = 8 * _mm
    for i, row in df_top.reset_index(drop=True).iterrows():
        bg = BAND_LIGHT if i % 2 == 0 else BAND_WHITE
        c.setFillColor(bg)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=1, stroke=0)
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.3)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=0, stroke=1)
        for cx in col_x[1:]:
            c.line(cx, y - ROW_H, cx, y)

        c.setFillColor(_col.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(col_x[0] + col_widths[0] / 2, y - ROW_H + 2.7 * _mm, str(i + 1))
        c.setFont("Helvetica", 9)
        cod = str(int(row["CODIGO"])) if pd.notna(row["CODIGO"]) else "—"
        c.drawCentredString(col_x[1] + col_widths[1] / 2, y - ROW_H + 2.7 * _mm, cod)
        desc = str(row.get("DESCRIPCION", "")).strip()[:72]
        c.drawString(col_x[2] + 1.5 * _mm, y - ROW_H + 2.7 * _mm, desc)
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(col_x[3] + col_widths[3] - 1.5 * _mm, y - ROW_H + 2.7 * _mm,
                          _fmt_num(row['BULTOS'], 0))
        c.setFont("Helvetica", 9)
        c.drawRightString(col_x[4] + col_widths[4] - 1.5 * _mm, y - ROW_H + 2.7 * _mm,
                          _fmt_num(row.get('HL', 0)))
        y -= ROW_H

    # (sin fila TOTAL — no se suma en el top)

    # ─ Resumen por DIVISIÓN ──────────────────────────────────────────────
    y -= 8 * _mm
    c.setFillColor(DIV_HDR)
    c.rect(M, y - 9 * _mm, table_w, 9 * _mm, fill=1, stroke=0)
    c.setFillColor(_col.white)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(M + 3 * _mm, y - 6 * _mm, "📊  TOTALES POR DIVISIÓN — Venta del Día")
    y -= 9 * _mm

    # Tabla división: Division | Bultos | HL
    cw_d = [
        table_w * 0.55,
        table_w * 0.22,
        table_w * 0.23,
    ]
    cx_d = [M]
    for w in cw_d[:-1]:
        cx_d.append(cx_d[-1] + w)

    # Header
    DHDR_H = 7 * _mm
    c.setFillColor(GOLD)
    c.rect(M, y - DHDR_H, table_w, DHDR_H, fill=1, stroke=1)
    c.setFillColor(DARK)
    c.setFont("Helvetica-Bold", 9)
    for i, h in enumerate(["DIVISIÓN", "BULTOS", "HL"]):
        c.drawCentredString(cx_d[i] + cw_d[i] / 2, y - DHDR_H + 2 * _mm, h)
    y -= DHDR_H

    DROW_H = 7 * _mm
    for i, row in df_div.reset_index(drop=True).iterrows():
        if y < 25 * _mm:
            c.showPage()
            y = PH - 20 * _mm
        bg = BAND_LIGHT if i % 2 == 0 else BAND_WHITE
        c.setFillColor(bg)
        c.rect(M, y - DROW_H, table_w, DROW_H, fill=1, stroke=0)
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.3)
        c.rect(M, y - DROW_H, table_w, DROW_H, fill=0, stroke=1)
        for cx in cx_d[1:]:
            c.line(cx, y - DROW_H, cx, y)

        c.setFillColor(_col.black)
        c.setFont("Helvetica", 8.5)
        div_name = str(row["DIVISION"])[:48]
        c.drawString(cx_d[0] + 1.5 * _mm, y - DROW_H + 2.2 * _mm, div_name)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawRightString(cx_d[1] + cw_d[1] - 1.5 * _mm, y - DROW_H + 2.2 * _mm, _fmt_num(row["BULTOS"], 0))
        c.setFont("Helvetica", 8.5)
        c.drawRightString(cx_d[2] + cw_d[2] - 1.5 * _mm, y - DROW_H + 2.2 * _mm, _fmt_num(row["HL"]))
        y -= DROW_H

    # (sin fila TOTAL GENERAL)

    # Footer
    c.setFillColor(_col.HexColor("#555555"))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(M, 10 * _mm, "Fuente: ANR.xlsx — Hoja BASE (cols H,I,K,M,O,AT). Top 10 por bultos desc.")
    c.drawRightString(PW - M, 10 * _mm,
                      f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.save()
    return bio.getvalue()


def build_top_clientes_anr_pdf(df_top: pd.DataFrame, fecha_str: str = "") -> bytes:
    """PDF a color: Top Clientes (hoja CLIENTES del ANR)."""
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib import colors as _col
    from reportlab.lib.units import mm as _mm
    from reportlab.pdfgen import canvas as _cv
    import io as _io

    PW, PH = _A4
    M = 12 * _mm

    DARK = _col.HexColor("#1a3a6b")
    GOLD = _col.HexColor("#f8d772")
    BAND_LIGHT = _col.HexColor("#f4f6fa")
    BAND_WHITE = _col.white
    BORDER = _col.HexColor("#8c8c8c")
    TOTAL_BG = _col.HexColor("#1a3a6b")

    bio = _io.BytesIO()
    c = _cv.Canvas(bio, pagesize=_A4)

    HDR_H = 22 * _mm
    c.setFillColor(DARK)
    c.rect(0, PH - HDR_H, PW, HDR_H, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(M, PH - 11 * _mm, "TOP CLIENTES — Venta del Día (ANR)")
    c.setFillColor(_col.white)
    c.setFont("Helvetica", 10)
    c.drawString(M, PH - 17 * _mm, "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(PW - M, PH - 11 * _mm, f"Fecha: {fecha_str}")

    y = PH - HDR_H - 8 * _mm
    table_w = PW - 2 * M
    # # | Cliente | Bultos | HL
    col_widths = [
        table_w * 0.06,
        table_w * 0.72,
        table_w * 0.13,
        table_w * 0.09,
    ]
    col_x = [M]
    for w in col_widths[:-1]:
        col_x.append(col_x[-1] + w)

    THDR_H = 8 * _mm
    c.setFillColor(GOLD)
    c.rect(M, y - THDR_H, table_w, THDR_H, fill=1, stroke=1)
    c.setStrokeColor(BORDER)
    c.setFillColor(DARK)
    c.setFont("Helvetica-Bold", 9.5)
    for i, h in enumerate(["#", "CLIENTE", "BULTOS", "HL"]):
        c.drawCentredString(col_x[i] + col_widths[i] / 2, y - THDR_H + 2.5 * _mm, h)
    y -= THDR_H

    ROW_H = 8 * _mm
    for i, row in df_top.reset_index(drop=True).iterrows():
        bg = BAND_LIGHT if i % 2 == 0 else BAND_WHITE
        c.setFillColor(bg)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=1, stroke=0)
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.3)
        c.rect(M, y - ROW_H, table_w, ROW_H, fill=0, stroke=1)
        for cx in col_x[1:]:
            c.line(cx, y - ROW_H, cx, y)

        c.setFillColor(_col.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(col_x[0] + col_widths[0] / 2, y - ROW_H + 2.7 * _mm, str(i + 1))
        c.setFont("Helvetica", 9)
        cli = str(row["CLIENTE"])[:80]
        c.drawString(col_x[1] + 1.5 * _mm, y - ROW_H + 2.7 * _mm, cli)
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(col_x[2] + col_widths[2] - 1.5 * _mm, y - ROW_H + 2.7 * _mm,
                          _fmt_num(row["BULTOS"], 0))
        c.setFont("Helvetica", 9)
        c.drawRightString(col_x[3] + col_widths[3] - 1.5 * _mm, y - ROW_H + 2.7 * _mm,
                          _fmt_num(row["HL"]))
        y -= ROW_H

    # (sin fila TOTAL — el top no representa el universo)

    # Footer
    c.setFillColor(_col.HexColor("#555555"))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(M, 10 * _mm, "Fuente: ANR.xlsx — Hoja CLIENTES (cols A,B,E,H). Top 10 por bultos desc.")
    c.drawRightString(PW - M, 10 * _mm,
                      f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.save()
    return bio.getvalue()



# ── JPEG directo via Pillow (sin poppler, sin PyMuPDF) ───────────────────────
_FONT_REG  = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
_FONT_BOLD = "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"

def _pil_font(size: int, bold: bool = False):
    from PIL import ImageFont
    try:
        return ImageFont.truetype(_FONT_BOLD if bold else _FONT_REG, size)
    except Exception:
        return ImageFont.load_default()


def _draw_tops_table_pil(draw, img_w, y, rows, col_widths, col_x,
                         headers, row_h=38, hdr_h=42):
    """Dibuja header + filas de tabla con Pillow."""
    from PIL import ImageDraw
    DARK  = "#1a3a6b"
    GOLD  = "#f8d772"
    LIGHT = "#f4f6fa"
    WHITE = "#ffffff"
    BORD  = "#b0b0b0"

    # Header de la tabla
    draw.rectangle([col_x[0], y, col_x[0] + sum(col_widths), y + hdr_h],
                   fill=GOLD, outline=BORD)
    fh = _pil_font(18, bold=True)
    for i, h in enumerate(headers):
        cx = col_x[i] + col_widths[i] // 2
        bb = draw.textbbox((0, 0), h, font=fh)
        tw = bb[2] - bb[0]
        draw.text((cx - tw // 2, y + (hdr_h - 20) // 2), h, font=fh, fill=DARK)
    y += hdr_h

    fr = _pil_font(17)
    fb = _pil_font(17, bold=True)
    for ri, cells in enumerate(rows):
        bg = LIGHT if ri % 2 == 0 else WHITE
        draw.rectangle([col_x[0], y, col_x[0] + sum(col_widths), y + row_h],
                       fill=bg, outline=BORD)
        for ci, (text, align, bold) in enumerate(cells):
            font = fb if bold else fr
            bb = draw.textbbox((0, 0), text, font=font)
            tw = bb[2] - bb[0]
            if align == "center":
                tx = col_x[ci] + col_widths[ci] // 2 - tw // 2
            elif align == "right":
                tx = col_x[ci] + col_widths[ci] - tw - 8
            else:
                tx = col_x[ci] + 8
            draw.text((tx, y + (row_h - 18) // 2), text, font=font, fill="#111111")
        y += row_h
    return y


def build_top_skus_anr_jpeg(df_top, df_div, fecha_str="") -> bytes:
    """JPEG directo (sin PDF): Top SKUs + División. Pixel-perfect vs PDF A4."""
    from PIL import Image, ImageDraw
    import io as _io

    # ── Dimensiones: A4 a 200 DPI ──────────────────────────────────────────
    # A4 = 210 × 297 mm → a 200 DPI = 1654 × 2339 px (portrait)
    # Usamos ancho fijo 1654px para fidelidad con el PDF
    DPI     = 200
    W       = 1654          # A4 ancho a 200 DPI
    M       = 96            # 12 mm × (200/25.4) ≈ 94 px → 96 px
    DARK    = "#1a3a6b"
    GOLD    = "#f8d772"
    LIGHT   = "#f4f6fa"
    WHITE   = "#ffffff"
    BORD    = "#8c8c8c"
    GRAY    = "#555555"
    MED     = "#2e5fa3"

    table_w = W - 2 * M

    # Proporciones idénticas al PDF
    cw = [int(table_w * r) for r in [0.05, 0.12, 0.63, 0.11, 0.09]]
    cx = [M]
    for w in cw[:-1]:
        cx.append(cx[-1] + w)

    # Alturas escaladas de mm → px (200 DPI)
    # PDF usa: HDR_H=22mm, ROW_H=8mm, THDR_H=8mm, DHDR_H=7mm, DROW_H=7mm
    def mm2px(mm_val): return int(mm_val * DPI / 25.4)

    HDR_H       = mm2px(22)   # 173 px — igual que el PDF
    THDR_H      = mm2px(8)    # 63 px
    ROW_H       = mm2px(8)    # 63 px
    GAP_TOP     = mm2px(8)    # espacio header → tabla
    GAP_MID     = mm2px(8)    # espacio tabla → div
    DIV_TITLE_H = mm2px(9)    # 71 px — bloque azul título división
    DHDR_H      = mm2px(7)    # 55 px
    DROW_H      = mm2px(7)    # 55 px
    FOOT_H      = mm2px(10)   # 79 px
    PAD         = 16          # padding interno horizontal

    n_skus = len(df_top)
    n_div  = len(df_div)
    total_h = (HDR_H + GAP_TOP +
               THDR_H + n_skus * ROW_H + GAP_MID +
               DIV_TITLE_H + DHDR_H + n_div * DROW_H + GAP_MID +
               FOOT_H + 24)

    img  = Image.new("RGB", (W, total_h), "white")
    draw = ImageDraw.Draw(img)

    # ── Fuentes escaladas al DPI (equivalentes exactas al PDF) ──────────────
    # PDF: título 17pt, subtítulo 10pt, fecha 11pt → × (200/72) factor DPI→pt
    f_title   = _pil_font(int(17 * DPI / 72), bold=True)   # ~47px ≈ 17pt
    f_sub     = _pil_font(int(10 * DPI / 72), bold=False)  # ~28px ≈ 10pt
    f_fecha   = _pil_font(int(11 * DPI / 72), bold=True)   # ~31px
    f_thdr    = _pil_font(int(9.5 * DPI / 72), bold=True)  # ~26px col header
    f_num     = _pil_font(int(9 * DPI / 72), bold=False)   # ~25px
    f_numB    = _pil_font(int(9 * DPI / 72), bold=True)
    f_divhdr  = _pil_font(int(12 * DPI / 72), bold=True)   # título división
    f_dhdr    = _pil_font(int(9 * DPI / 72), bold=True)
    f_foot    = _pil_font(int(7 * DPI / 72), bold=False)   # ~19px

    def text_w(txt, font):
        bb = draw.textbbox((0, 0), txt, font=font)
        return bb[2] - bb[0]

    def text_h(font):
        bb = draw.textbbox((0, 0), "Ag", font=font)
        return bb[3] - bb[1]

    def draw_text_vcenter(x, y_top, h, txt, font, fill, align="left", x_end=None):
        th = text_h(font)
        ty = y_top + (h - th) // 2
        if align == "center":
            mid = (x + (x_end or x)) // 2 if x_end else x
            tx = mid - text_w(txt, font) // 2
        elif align == "right":
            tx = (x_end or x) - text_w(txt, font) - PAD
        else:
            tx = x + PAD
        draw.text((tx, ty), txt, font=font, fill=fill)

    # ── Header banner ─────────────────────────────────────────────────────
    draw.rectangle([0, 0, W, HDR_H], fill=DARK)

    # Título dorado (alineado izquierda con margen)
    th_title = text_h(f_title)
    draw.text((M, int(HDR_H * 0.18)), "TOP SKUs — Venta del Día (ANR)",
              font=f_title, fill=GOLD)
    # Subtítulo blanco debajo
    draw.text((M, int(HDR_H * 0.18) + th_title + 6),
              "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén",
              font=f_sub, fill="white")
    # Fecha alineada a la derecha
    fecha_txt = f"Fecha: {fecha_str}"
    draw.text((W - M - text_w(fecha_txt, f_fecha), int(HDR_H * 0.18)),
              fecha_txt, font=f_fecha, fill="white")

    y = HDR_H + GAP_TOP

    # ── Helper: dibuja una tabla completa ─────────────────────────────────
    def draw_table(y_start, col_w, col_x, headers, rows, hdr_h, row_h,
                   fhdr, frow, fbold):
        y = y_start
        tw_total = sum(col_w)
        # Header dorado
        draw.rectangle([col_x[0], y, col_x[0] + tw_total, y + hdr_h],
                       fill=GOLD, outline=BORD)
        for i, h in enumerate(headers):
            draw_text_vcenter(col_x[i], y, hdr_h, h, fhdr, DARK,
                              align="center", x_end=col_x[i] + col_w[i])
        y += hdr_h

        for ri, cells in enumerate(rows):
            bg = LIGHT if ri % 2 == 0 else WHITE
            draw.rectangle([col_x[0], y, col_x[0] + tw_total, y + row_h],
                           fill=bg, outline=BORD)
            # líneas verticales internas
            for ci in range(1, len(col_x)):
                draw.line([(col_x[ci], y), (col_x[ci], y + row_h)],
                          fill=BORD, width=1)
            for ci, (txt, align, bold) in enumerate(cells):
                font = fbold if bold else frow
                draw_text_vcenter(col_x[ci], y, row_h, txt, font, "#111111",
                                  align=align, x_end=col_x[ci] + col_w[ci])
            y += row_h
        return y

    # ── Tabla Top SKUs ─────────────────────────────────────────────────────
    rows_skus = []
    for i, row in df_top.reset_index(drop=True).iterrows():
        try:
            cod = str(int(float(row["CODIGO"])))
        except Exception:
            cod = str(row["CODIGO"])
        desc   = str(row.get("DESCRIPCION", ""))[:72]
        bultos = _fmt_num(row["BULTOS"], 0)
        hl     = _fmt_num(row.get("HL", 0))
        rows_skus.append([
            (str(i + 1), "center", True),
            (cod,        "center", False),
            (desc,       "left",   False),
            (bultos,     "right",  True),
            (hl,         "right",  False),
        ])

    y = draw_table(y, cw, cx,
                   ["#", "CÓDIGO", "DESCRIPCIÓN", "BULTOS", "HL"],
                   rows_skus, THDR_H, ROW_H,
                   f_thdr, f_num, f_numB)
    y += GAP_MID

    # ── Bloque título División (azul medio, igual al PDF) ──────────────────
    draw.rectangle([M, y, M + table_w, y + DIV_TITLE_H], fill=MED)
    # "📊" puede fallar en PIL sin emoji font → usamos cuadrado + texto
    draw.rectangle([M + 14, y + DIV_TITLE_H // 2 - 10,
                    M + 14 + 20, y + DIV_TITLE_H // 2 + 10], fill="white")
    draw_text_vcenter(M + 44, y, DIV_TITLE_H,
                      "TOTALES POR DIVISIÓN — Venta del Día",
                      f_divhdr, "white", align="left")
    y += DIV_TITLE_H

    cw_d = [int(table_w * r) for r in [0.55, 0.22, 0.23]]
    cx_d = [M]
    for w in cw_d[:-1]:
        cx_d.append(cx_d[-1] + w)

    rows_div = []
    for _, row in df_div.reset_index(drop=True).iterrows():
        rows_div.append([
            (str(row["DIVISION"])[:55], "left",  False),
            (_fmt_num(row["BULTOS"], 0), "right", True),
            (_fmt_num(row["HL"]),        "right", False),
        ])

    y = draw_table(y, cw_d, cx_d,
                   ["DIVISIÓN", "BULTOS", "HL"],
                   rows_div, DHDR_H, DROW_H,
                   f_dhdr, f_num, f_numB)
    y += GAP_MID

    # ── Footer ─────────────────────────────────────────────────────────────
    footer_l = "Fuente: ANR.xlsx — Hoja BASE · Picking Orchestrator v" + APP_VERSION
    footer_r = datetime.now().strftime('%d/%m/%Y %H:%M')
    draw.text((M, y + 6), footer_l, font=f_foot, fill=GRAY)
    draw.text((W - M - text_w(footer_r, f_foot), y + 6),
              footer_r, font=f_foot, fill=GRAY)

    buf = _io.BytesIO()
    img.save(buf, format="JPEG", quality=94, optimize=True)
    return buf.getvalue()


def build_top_clientes_anr_jpeg(df_top, fecha_str="") -> bytes:
    """JPEG directo (sin PDF): Top Clientes. Mismo diseño que el PDF."""
    from PIL import Image, ImageDraw
    import io as _io

    W     = 1240
    M     = 48
    DARK  = "#1a3a6b"
    GOLD  = "#f8d772"
    GRAY  = "#555555"

    table_w = W - 2 * M
    cw = [int(table_w * r) for r in [0.06, 0.72, 0.13, 0.09]]
    cx = [M]
    for w in cw[:-1]:
        cx.append(cx[-1] + w)

    HDR_H  = 90
    ROW_H  = 38
    THDR_H = 42
    FOOT_H = 40
    GAP    = 24

    n = len(df_top)
    total_h = HDR_H + GAP + THDR_H + n * ROW_H + GAP + FOOT_H + 20

    img  = Image.new("RGB", (W, total_h), "white")
    draw = ImageDraw.Draw(img)

    # ── Header ──
    draw.rectangle([0, 0, W, HDR_H], fill=DARK)
    draw.text((M, 12), "TOP CLIENTES — Venta del Día (ANR)",
              font=_pil_font(32, True), fill=GOLD)
    draw.text((M, 52), "Beccacece Hnos SA · DPO 2.1 — Pilar Almacén",
              font=_pil_font(20), fill="white")
    fecha_txt = f"Fecha: {fecha_str}"
    bb = draw.textbbox((0, 0), fecha_txt, font=_pil_font(22, True))
    draw.text((W - M - (bb[2] - bb[0]), 14), fecha_txt,
              font=_pil_font(22, True), fill="white")

    y = HDR_H + GAP

    rows = []
    for i, row in df_top.reset_index(drop=True).iterrows():
        cli    = str(row["CLIENTE"])[:80]
        bultos = _fmt_num(row["BULTOS"], 0)
        hl     = _fmt_num(row.get("HL", 0))
        rows.append([
            (str(i + 1), "center", True),
            (cli,        "left",   False),
            (bultos,     "right",  True),
            (hl,         "right",  False),
        ])

    y = _draw_tops_table_pil(draw, W, y, rows, cw, cx,
                              ["#", "CLIENTE", "BULTOS", "HL"],
                              row_h=ROW_H, hdr_h=THDR_H)
    y += GAP

    footer = (f"Fuente: ANR.xlsx — Hoja CLIENTES  ·  "
              f"Picking Orchestrator v{APP_VERSION} · {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    draw.text((M, y + 8), footer, font=_pil_font(15), fill=GRAY)

    buf = _io.BytesIO()
    img.save(buf, format="JPEG", quality=92, optimize=True)
    return buf.getvalue()

def render_tab_top_skus():
    st.subheader("🏆 Tops del Día — ANR (Análisis de Rechazos / Venta)")
    st.caption(
        "Input: **ANR.xlsx** (cargado en 📁 Archivos). "
        "Genera **Venta total**, **Top 10 SKUs** y **Top 10 Clientes**."
    )

    # ANR: reusa el cargado en 📁 Archivos
    anr_use = st.session_state.get("tc_anr")
    if not anr_use:
        st.info("⬅️ Subí el **ANR.xlsx** en la pestaña **📁 Archivos** para generar los tops.")
        return

    try:
        anr_bytes = anr_use.getvalue()
        df_top_skus = _build_top10_skus_anr(anr_bytes, n=10)
        df_div      = _build_division_summary_anr(anr_bytes)
        df_top_cli  = _build_top10_clientes_anr(anr_bytes, n=10)
    except Exception as e:
        st.error(f"❌ Error procesando ANR: {e}")
        with st.expander("Stack trace"):
            import traceback
            st.code(traceback.format_exc())
        return

    fecha_hoy = datetime.now().strftime("%d/%m/%Y")

    # ════════════════════════════════════════════════════════════════════
    # SECCIÓN 1 — VENTA TOTAL DEL DÍA
    # ════════════════════════════════════════════════════════════════════
    st.markdown("### 📈 Sección 1 — Venta total del día")

    if not df_div.empty:
        tot_b_day = float(df_div["BULTOS"].sum())
        tot_h_day = float(df_div["HL"].sum())
    else:
        tot_b_day = float(df_top_skus["BULTOS"].sum()) if not df_top_skus.empty else 0
        tot_h_day = float(df_top_skus["HL"].sum())     if not df_top_skus.empty else 0

    cc1, cc2 = st.columns(2)
    cc1.metric("Bultos totales vendidos", _fmt_num(tot_b_day, 0))
    cc2.metric("HL totales vendido",      _fmt_num(tot_h_day))

    # División como complemento opcional (compacto)
    if not df_div.empty:
        with st.expander("📊 Totales por División (complemento)", expanded=False):
            df_div_disp = df_div.copy()
            df_div_disp = df_div_disp[["DIVISION", "BULTOS", "HL"]]
            df_div_disp["BULTOS"]  = df_div_disp["BULTOS"].map(lambda x: _fmt_num(x, 0))
            df_div_disp["HL"]      = df_div_disp["HL"].map(lambda x: _fmt_num(x))
            df_div_disp.rename(columns={
                "DIVISION": "División", "BULTOS": "Bultos", "HL": "HL",
            }, inplace=True)
            st.dataframe(
                df_div_disp, width="stretch", height=280, hide_index=True,
                column_config={
                    "División": st.column_config.TextColumn(width="medium"),
                    "Bultos":   st.column_config.TextColumn(width="small"),
                    "HL":       st.column_config.TextColumn(width="small"),
                },
            )

    st.divider()

    # ════════════════════════════════════════════════════════════════════
    # SECCIÓN 2 — TOP 10 SKUs
    # ════════════════════════════════════════════════════════════════════
    st.markdown("### 🏆 Sección 2 — Top 10 SKUs")

    if df_top_skus.empty:
        st.warning("No se encontraron SKUs con ventas > 0 en el ANR.")
    else:
        df_disp = df_top_skus.copy()
        df_disp["CODIGO"]  = df_disp["CODIGO"].astype(int).astype(str)
        df_disp["BULTOS"]  = df_disp["BULTOS"].map(lambda x: _fmt_num(x, 0))
        df_disp["HL"]      = df_disp["HL"].map(lambda x: _fmt_num(x))
        df_disp.index = range(1, len(df_disp) + 1)
        df_disp.rename(columns={
            "CODIGO": "Código", "DESCRIPCION": "Descripción",
            "BULTOS": "Bultos", "HL": "HL",
        }, inplace=True)
        df_disp = df_disp[["Código", "Descripción", "Bultos", "HL"]]
        st.dataframe(
            df_disp, width="stretch", height=395,
            column_config={
                "Código":      st.column_config.TextColumn(width="small"),
                "Descripción": st.column_config.TextColumn(width="large"),
                "Bultos":      st.column_config.TextColumn(width="small"),
                "HL":          st.column_config.TextColumn(width="small"),
            },
        )

        # PDF Top SKUs (hoja 1: venta general + top 10 SKUs) — v4.37: solo PDF, sin JPEG
        if not df_div.empty:
            try:
                pdf_skus = build_top_skus_anr_pdf(df_top_skus, df_div, fecha_str=fecha_hoy)
                st.download_button(
                    "⬇ PDF — Venta total + Top 10 SKUs",
                    data=pdf_skus,
                    file_name=_stamp("Top_SKUs_ANR", "pdf"),
                    mime="application/pdf",
                    type="primary",
                    width="stretch",
                    key="topskus_anr_pdf_dl",
                )
            except Exception as e:
                st.error(f"❌ Error generando PDF Top SKUs: {e}")

    st.divider()

    # ════════════════════════════════════════════════════════════════════
    # SECCIÓN 3 — TOP 10 CLIENTES
    # ════════════════════════════════════════════════════════════════════
    st.markdown("### 👥 Sección 3 — Top 10 Clientes")
    if df_top_cli.empty:
        st.warning("No se encontraron clientes con ventas > 0 en el ANR.")
    else:
        df_cli_disp = df_top_cli.copy()
        df_cli_disp = df_cli_disp[["CLIENTE", "BULTOS", "HL"]]
        df_cli_disp["BULTOS"]  = df_cli_disp["BULTOS"].map(lambda x: _fmt_num(x, 0))
        df_cli_disp["HL"]      = df_cli_disp["HL"].map(lambda x: _fmt_num(x))
        df_cli_disp.index = range(1, len(df_cli_disp) + 1)
        df_cli_disp.rename(columns={
            "CLIENTE": "Cliente", "BULTOS": "Bultos", "HL": "HL",
        }, inplace=True)
        st.dataframe(
            df_cli_disp, width="stretch", height=395,
            column_config={
                "Cliente": st.column_config.TextColumn(width="large"),
                "Bultos":  st.column_config.TextColumn(width="small"),
                "HL":      st.column_config.TextColumn(width="small"),
            },
        )

        # PDF Top Clientes — v4.37: solo PDF, sin JPEG
        try:
            pdf_cli = build_top_clientes_anr_pdf(df_top_cli, fecha_str=fecha_hoy)
            st.download_button(
                "⬇ PDF — Top 10 Clientes",
                data=pdf_cli,
                file_name=_stamp("Top_Clientes_ANR", "pdf"),
                mime="application/pdf",
                type="primary",
                width="stretch",
                key="topcli_anr_pdf_dl",
            )
        except Exception as e:
            st.error(f"❌ Error generando PDF Top Clientes: {e}")

    with st.expander("ℹ️ Metodología"):
        st.markdown(
            "- **Fuente**: ANR.xlsx (Análisis de Rechazos / venta del próximo día).\n"
            "- **Venta total del día**: suma de TODA la hoja BASE — no solo del top.\n"
            "- **Top SKUs**: hoja `BASE`, cols H/I/K/M/O. Agrupado por código, ordenado por bultos desc.\n"
            "- **División** (expandible): col AT, complemento opcional.\n"
            "- **Top Clientes**: hoja `CLIENTES` (preagregada), cols A/B/E/H. Sin fila de totales en el PDF.\n"
            "- **PDFs**: 2 archivos separados — PDF 1 (1 hoja): venta general + top 10 SKUs · "
            "PDF 2 (1 hoja): top clientes."
        )


# ── TAB BOLETAS v4.19 — Agrupación por camión con fallback ANR ──────────────

def _build_anr_lookup() -> dict:
    """
    Construye dos diccionarios desde ANR en session_state:
      - cliente_to_camion: {nombre_cliente_normalizado: num_camion_int}
      - reparto_to_camion: {num_reparto_int: num_camion_int}  (vacío por ahora, placeholder)
      - transporte_to_nombre: {num_transporte_int: desc_transporte_str}
    Devuelve dict con esas tres claves. Si no hay ANR, devuelve dicts vacíos.
    """
    result = {
        "cliente_to_camion": {},
        "transporte_to_nombre": {},
    }
    df_anr = st.session_state.get("anr_df")
    if df_anr is None:
        return result

    try:
        # La hoja BASE tiene la fila 0 como header real después de leer con header=0
        df = df_anr.copy()
        # Si la primera fila tiene los nombres de columna como valores (caso raw), reparar
        if "EMPRESA" not in df.columns and df.iloc[0].astype(str).str.contains("EMPRESA").any():
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)

        # Normalizar nombres de columnas
        df.columns = [str(c).strip().upper() for c in df.columns]

        desc_cli_col = next((c for c in df.columns if "DESCRIPCIÓN CLIENTE" in c and "DETALLADA" not in c and "DOMICILIO" not in c), None)
        trans_col = next((c for c in df.columns if c == "TRANSPORTE"), None)
        desc_trans_col = next((c for c in df.columns if "DESCRIPCIÓN TRANSPORTE" in c and "DETALLADA" not in c), None)

        if desc_cli_col and trans_col:
            for _, row in df[[desc_cli_col, trans_col]].dropna().iterrows():
                cli_raw = str(row[desc_cli_col]).strip()
                try:
                    cam = int(row[trans_col])
                except (ValueError, TypeError):
                    continue
                # Guardar versión normalizada (sin tildes, lowercase)
                cli_norm = _norm(cli_raw)
                result["cliente_to_camion"][cli_norm] = cam

        if trans_col and desc_trans_col:
            for _, row in df[[trans_col, desc_trans_col]].dropna().drop_duplicates().iterrows():
                try:
                    cam = int(row[trans_col])
                except (ValueError, TypeError):
                    continue
                result["transporte_to_nombre"][cam] = str(row[desc_trans_col]).strip()

    except Exception as e:
        pass  # Si falla, devolvemos dicts vacíos

    return result


def _extract_pdf_text_all_pages(pdf_bytes: bytes) -> str:
    """Extrae todo el texto de todas las páginas de un PDF como string único."""
    import re
    try:
        import pdfplumber
        texts = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                texts.append(t)
        return "\n".join(texts)
    except Exception:
        pass
    # Fallback a pypdf
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception:
        return ""


def _detect_camion_for_pdf(pdf_bytes: bytes, anr_lookup: dict) -> tuple[int | None, str]:
    """
    Detecta el número de camión (TRANSPORTE en ANR) para un PDF dado.
    Aplica detección en capas:
      Capa 1: "Reparto: XXXXXXXX" → busca en ANR por reparto (si hay mapeo) → si no, usa directo
      Capa 2: "Transporte: NNN" o "NNN - NOMBRE" en Composición de Carga
      Capa 3: "Fletero: NOMBRE" → busca nombre en transporte_to_nombre invertido
      Capa 4: Nombre del cliente → busca en cliente_to_camion del ANR
    Devuelve (num_camion_int_or_None, metodo_str).
    """
    import re

    text = _extract_pdf_text_all_pages(pdf_bytes)
    if not text:
        return None, "sin texto"

    cliente_to_camion = anr_lookup.get("cliente_to_camion", {})
    transporte_to_nombre = anr_lookup.get("transporte_to_nombre", {})
    # Invertido: nombre_normalizado → num_camion
    nombre_to_camion = {_norm(v): k for k, v in transporte_to_nombre.items()}

    # ── Capa 1: campo Reparto ────────────────────────────────────────────────
    # "Reparto: 00092046" o "Reparto  00092046"
    m = re.search(r"Reparto[:\s]+0*(\d{3,})", text, re.IGNORECASE)
    if m:
        reparto_num = int(m.group(1))
        # El reparto en sí es el número de transporte/camión en el sistema
        # Verificar si ese número existe como TRANSPORTE en ANR
        if reparto_num in transporte_to_nombre:
            return reparto_num, f"Reparto→ANR ({reparto_num})"
        # Si no, devolvemos el reparto como número de orden (comportamiento anterior)
        return reparto_num, f"Reparto ({reparto_num})"

    # ── Capa 2: Composición de Carga — "Transporte: NNN - NOMBRE" ───────────
    # Patrón: "128 - VILLALOBO" o "Transporte: 128 - VILLALOBO"
    m = re.search(r"Transporte[:\s]+(\d+)\s*[-–]\s*(\w+)", text, re.IGNORECASE)
    if m:
        cam = int(m.group(1))
        return cam, f"Composición Carga ({cam})"

    # También puede aparecer solo el número de transporte en la línea de cabecera
    m = re.search(r"Dep[oó]sito[:\s]+\d+[^\n]*Transporte[:\s]*(\d+)", text, re.IGNORECASE)
    if m:
        cam = int(m.group(1))
        return cam, f"Comp.Carga/Depósito ({cam})"

    # ── Capa 3: campo Fletero → nombre del chofer/transporte ────────────────
    # "Fletero: ARANDA" o "Fletero:  COLOMBO S"
    m = re.search(r"Fletero[:\s]+([A-ZÁ-Úa-zá-ú\s]+?)(?:\n|&|$)", text, re.IGNORECASE)
    if m:
        fletero_raw = m.group(1).strip()
        fletero_norm = _norm(fletero_raw)
        # Buscar match parcial en nombre_to_camion
        for nombre_norm, cam in nombre_to_camion.items():
            if fletero_norm in nombre_norm or nombre_norm in fletero_norm:
                return cam, f"Fletero ({fletero_raw}→{cam})"
        # Si no matchea exacto, buscar en cliente_to_camion con el nombre del fletero
        # (a veces el fletero tiene el apellido del transporte)
        for nombre_norm, cam in nombre_to_camion.items():
            words_fletero = set(fletero_norm.split())
            words_nombre = set(nombre_norm.split())
            if words_fletero & words_nombre:  # intersección de palabras
                return cam, f"Fletero~match ({fletero_raw}→{cam})"

    # ── Capa 4: nombre del cliente → ANR lookup ──────────────────────────────
    # Buscar el campo "Cliente:" en la boleta (Factura A / Presupuesto / FAC.PRES)
    # Patrones posibles:
    #   "Cliente:   (016345) ELI SOCIEDAD ANONIMA"
    #   "MOLINAS JONATHAM ALEXIS"  (FAC.PRES sin etiqueta clara)
    clientes_encontrados = []

    # Patrón Factura A/B: "Cliente:   (XXXXXX) NOMBRE APELLIDO"
    for m in re.finditer(r"Cliente[:\s]+\(\d+\)\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)(?:\n|Domicilio|$)", text, re.IGNORECASE):
        clientes_encontrados.append(m.group(1).strip())

    # Patrón FAC.PRES / texto suelto: buscar líneas que sean nombres de clientes del ANR
    # (heurística: línea de 2-4 palabras en mayúsculas, sin números, que matchee en ANR)
    if not clientes_encontrados:
        lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
        for ln in lines:
            # Solo líneas que parezcan nombres propios (2-4 palabras, mayúsculas, sin números)
            if re.match(r"^[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{5,60}$", ln) and not any(c.isdigit() for c in ln):
                words = ln.split()
                if 2 <= len(words) <= 5:
                    clientes_encontrados.append(ln)

    for cliente_raw in clientes_encontrados:
        cli_norm = _norm(cliente_raw)
        # Match exacto
        if cli_norm in cliente_to_camion:
            return cliente_to_camion[cli_norm], f"Cliente→ANR ({cliente_raw})"
        # Match parcial: al menos 2 palabras en común
        words_pdf = set(cli_norm.split())
        for cli_anr_norm, cam in cliente_to_camion.items():
            words_anr = set(cli_anr_norm.split())
            comunes = words_pdf & words_anr
            if len(comunes) >= 2:
                return cam, f"Cliente~match ({cliente_raw}→{cam})"

    return None, "no detectado"


def render_tab_boletas():
    """
    Tab de impresión de boletas v4.19.
    Agrupa PDFs por camión usando detección multicapa + ANR como fallback.
    No altera los documentos: solo los reordena antes de combinar.
    """
    st.subheader("🖨️ Impresión de Boletas")
    st.caption(
        "Subí todos los PDFs de boletas del próximo día. "
        "La app los agrupa automáticamente por camión (usando ANR como referencia) "
        "y genera un único PDF listo para mandar a imprimir."
    )

    # ── Verificar si hay ANR cargado ─────────────────────────────────────────
    anr_df = st.session_state.get("anr_df")
    if anr_df is None:
        st.warning(
            "⚠️ **ANR no cargado.** Para el cross-reference cliente→camión, "
            "cargá el ANR.xlsx en la pestaña 📁 **Archivos** primero. "
            "Sin ANR, solo se detecta el campo `Reparto` del PDF."
        )
        anr_lookup = {"cliente_to_camion": {}, "transporte_to_nombre": {}}
    else:
        anr_lookup = _build_anr_lookup()
        n_clientes = len(anr_lookup["cliente_to_camion"])
        n_transportes = len(anr_lookup["transporte_to_nombre"])
        st.success(
            f"✅ ANR cargado — {n_clientes} clientes mapeados · {n_transportes} camiones registrados"
        )

    uploaded_files = st.file_uploader(
        "PDFs de boletas (facturas A/B/C, remitos, NC, presupuestos, composición de carga) *",
        type=["pdf"],
        accept_multiple_files=True,
        key="boletas_pdfs",
    )

    if not uploaded_files:
        st.info("📄 Subí uno o más PDFs para continuar.")
        return

    st.divider()

    # ── Procesar cada PDF ────────────────────────────────────────────────────
    with st.spinner("Analizando PDFs y detectando camiones..."):
        file_data = []
        no_detectados = []

        for uf in uploaded_files:
            raw = uf.read()
            camion, metodo = _detect_camion_for_pdf(raw, anr_lookup)
            entry = {
                "name": uf.name,
                "camion": camion,
                "metodo": metodo,
                "bytes": raw,
            }
            file_data.append(entry)
            if camion is None:
                no_detectados.append(uf.name)

    # ── Advertencias ─────────────────────────────────────────────────────────
    if no_detectados:
        with st.expander(f"⚠️ {len(no_detectados)} archivo(s) sin camión detectado", expanded=True):
            for nd in no_detectados:
                st.markdown(f"- **{nd}**")
            st.caption(
                "Estos archivos se incluirán al final del PDF combinado. "
                "Revisá que el ANR esté cargado y actualizado para mejorar la detección."
            )

    # ── Separar detectados / no detectados y ordenar ─────────────────────────
    con_camion = sorted(
        [f for f in file_data if f["camion"] is not None],
        key=lambda x: x["camion"],
    )
    sin_camion = [f for f in file_data if f["camion"] is None]

    # ── Tabla resumen ────────────────────────────────────────────────────────
    from pypdf import PdfReader as _PRcount

    def _count_pages(pdf_bytes):
        try:
            return len(_PRcount(io.BytesIO(pdf_bytes)).pages)
        except Exception:
            return "?"

    transporte_to_nombre = anr_lookup.get("transporte_to_nombre", {})

    resumen_data = []
    camiones_unicos = {}

    for f in con_camion:
        cam = f["camion"]
        nombre_transporte = transporte_to_nombre.get(cam, "—")
        resumen_data.append({
            "N° Camión": cam,
            "Transporte": nombre_transporte,
            "Método": f["metodo"],
            "Archivo": f["name"],
            "Páginas": _count_pages(f["bytes"]),
        })
        camiones_unicos[cam] = nombre_transporte

    for f in sin_camion:
        resumen_data.append({
            "N° Camión": "—",
            "Transporte": "SIN CAMIÓN",
            "Método": f["metodo"],
            "Archivo": f["name"],
            "Páginas": _count_pages(f["bytes"]),
        })

    df_resumen = pd.DataFrame(resumen_data)

    col_info, col_metrics = st.columns([3, 1])
    with col_info:
        st.markdown("#### 📋 Agrupación por camión")
        st.dataframe(df_resumen, use_container_width=True, hide_index=True)
    with col_metrics:
        st.metric("Total PDFs", len(file_data))
        st.metric("Camiones detectados", len(camiones_unicos))
        st.metric("Sin camión", len(sin_camion))
        total_pags = sum(
            _count_pages(f["bytes"]) for f in file_data
            if isinstance(_count_pages(f["bytes"]), int)
        )
        st.metric("Total páginas", total_pags)

    st.divider()

    # ── Selector de modo de salida ────────────────────────────────────────────
    modo = st.radio(
        "Modo de salida:",
        ["📄 Un PDF unificado (todos los camiones)", "🚛 Un PDF por camión (ZIP)"],
        horizontal=True,
        key="boletas_modo",
    )

    # ── Botón de generación ──────────────────────────────────────────────────
    if st.button("🖨️ Generar PDF(s)", type="primary", use_container_width=True):
        with st.spinner("Combinando PDFs..."):
            try:
                from pypdf import PdfWriter, PdfReader as _PR

                orden_final = con_camion + sin_camion

                if "unificado" in modo:
                    # ── Modo 1: PDF único ────────────────────────────────────
                    writer = PdfWriter()
                    for f in orden_final:
                        reader = _PR(io.BytesIO(f["bytes"]))
                        for page in reader.pages:
                            writer.add_page(page)

                    output_buf = io.BytesIO()
                    writer.write(output_buf)
                    output_buf.seek(0)
                    pdf_combinado = output_buf.read()

                    fecha_str = datetime.now().strftime("%Y%m%d_%H%M")
                    filename = f"Boletas_Impresion_{fecha_str}.pdf"

                    st.success(
                        f"✅ PDF generado — {len(orden_final)} archivos · {total_pags} páginas · "
                        f"{len(camiones_unicos)} camión(es)"
                    )
                    st.download_button(
                        label="⬇️ Descargar PDF unificado",
                        data=pdf_combinado,
                        file_name=filename,
                        mime="application/pdf",
                        type="primary",
                        use_container_width=True,
                        key="boletas_dl_unificado",
                    )

                else:
                    # ── Modo 2: ZIP con un PDF por camión ────────────────────
                    import zipfile

                    # Agrupar por camión
                    grupos: dict = {}
                    for f in con_camion:
                        cam = f["camion"]
                        grupos.setdefault(cam, []).append(f)
                    if sin_camion:
                        grupos["SIN_CAMION"] = sin_camion

                    zip_buf = io.BytesIO()
                    total_archivos = 0
                    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                        for cam_key in sorted(
                            grupos.keys(),
                            key=lambda x: (0, x) if isinstance(x, int) else (1, str(x))
                        ):
                            archivos_cam = grupos[cam_key]
                            writer_cam = PdfWriter()
                            for f in archivos_cam:
                                reader = _PR(io.BytesIO(f["bytes"]))
                                for page in reader.pages:
                                    writer_cam.add_page(page)

                            buf_cam = io.BytesIO()
                            writer_cam.write(buf_cam)
                            buf_cam.seek(0)

                            nombre_trans = transporte_to_nombre.get(cam_key, str(cam_key)) if isinstance(cam_key, int) else "SIN_CAMION"
                            nombre_trans_safe = nombre_trans.replace(" ", "_").replace("/", "-")
                            pdf_name = f"CAM_{cam_key}_{nombre_trans_safe}.pdf"
                            zf.writestr(pdf_name, buf_cam.read())
                            total_archivos += 1

                    zip_buf.seek(0)
                    fecha_str = datetime.now().strftime("%Y%m%d_%H%M")
                    zip_name = f"Boletas_PorCamion_{fecha_str}.zip"

                    st.success(
                        f"✅ ZIP generado — {total_archivos} PDF(s) · "
                        f"{len(camiones_unicos)} camión(es)"
                    )
                    st.download_button(
                        label="⬇️ Descargar ZIP (un PDF por camión)",
                        data=zip_buf.read(),
                        file_name=zip_name,
                        mime="application/zip",
                        type="primary",
                        use_container_width=True,
                        key="boletas_dl_zip",
                    )

            except Exception as e:
                st.error(f"❌ Error al generar PDF(s): {e}")
                with st.expander("Stack trace"):
                    import traceback
                    st.code(traceback.format_exc())

    with st.expander("ℹ️ ¿Cómo funciona? (detección multicapa)"):
        st.markdown(
            "La detección del camión se aplica en capas, en orden:\n\n"
            "1. **Reparto** → campo `Reparto: XXXXXXXX` en el PDF. Si el número coincide con un camión del ANR, lo asigna directamente.\n"
            "2. **Composición de Carga** → campo `Transporte: NNN - NOMBRE` en la cabecera del documento.\n"
            "3. **Fletero** → campo `Fletero: NOMBRE` en facturas A/B/C. Busca el nombre en el listado de transportes del ANR.\n"
            "4. **Cliente → ANR** → extrae el nombre del cliente del PDF y lo cruza con la hoja BASE del ANR para obtener su camión asignado.\n\n"
            "**Modos de salida:**\n"
            "- *PDF unificado*: todos los camiones en un solo archivo, ordenados de menor a mayor.\n"
            "- *ZIP por camión*: un PDF separado por cada camión (ideal para imprimir por separado).\n\n"
            "Los documentos **no se modifican**: solo se reordenan y combinan.\n\n"
            "**Formatos soportados**: Factura A/B/C, Remito, Nota de Crédito, "
            "Presupuesto, Composición de Carga, FAC.PRES — cualquier PDF del sistema Beccacece."
        )



# ════════════════════════════════════════════════════════════════════════════
# TAB TABLERO RUTEADOR (v4.42.0)
# ════════════════════════════════════════════════════════════════════════════

_TR_CAMIONES_DEFAULT = {
    101: "AE785NJ", 102: "KDC500",  103: "",       104: "AE785NK",
    105: "AE962OJ", 106: "AC354KB", 107: "",        108: "AC354KD",
    109: "UFH587",  110: "DMZ947",  111: "LGH463",  112: "KCT549",
    113: "AF483IS", 114: "DMZ947",  115: "KDC500",  117: "DJE621",
    118: "AE962OJ", 119: "",        120: "SOB215",  121: "",
    122: "MRP648",  123: "",        125: "DJE621",  128: "AST830",
    129: "",
}

_TR_CAPACIDADES_DEFAULT = {
    "AE785NJ": 10450, "KDC500": 9480,  "AE785NK": 9460, "AC354KB": 9480,
    "AC354KD": 9440,  "UFH587": 9470,  "DMZ947":  9480, "LGH463":  9600,
    "KCT549":  9470,  "AF483IS": 9480, "AE962OJ": 9480, "DJE621":  9470,
    "SOB215":  9600,  "MRP648":  9630, "AST830":  10470,
}

_TR_TARGET_OCUP_BODEGA  = 300
_TR_TARGET_FUERA_ZONA   = 0.03
_TR_TARGET_UTIL_VEH     = 0.80
_TR_TARGET_EFICIENCIA   = 0.70


def _tr_parse_time_input(val_str):
    """Convierte 'HH:MM' en datetime.time, None si vacío."""
    import datetime
    if not val_str or str(val_str).strip() in ("", "nan", "None"):
        return None
    s = str(val_str).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _tr_time_to_min(t):
    if t is None:
        return 0.0
    return t.hour * 60 + t.minute + t.second / 60


# TAB TABLERO RUTEADOR (v4.48.0)
# ════════════════════════════════════════════════════════════════════════════

_TR_CAMIONES_DEFAULT = {
    101: "AE785NJ", 102: "KDC500",  103: "",       104: "AE785NK",
    105: "AE962OJ", 106: "AC354KB", 107: "",        108: "AC354KD",
    109: "UFH587",  110: "DMZ947",  111: "LGH463",  112: "KCT549",
    113: "AF483IS", 114: "DMZ947",  115: "KDC500",  117: "DJE621",
    118: "AE962OJ", 119: "",        120: "SOB215",  121: "",
    122: "MRP648",  123: "",        125: "DJE621",  128: "AST830",
    129: "",
}

_TR_CAPACIDADES_DEFAULT = {
    "AE785NJ": 10450, "KDC500": 9480,  "AE785NK": 9460, "AC354KB": 9480,
    "AC354KD": 9440,  "UFH587": 9470,  "DMZ947":  9480, "LGH463":  9600,
    "KCT549":  9470,  "AF483IS": 9480, "AE962OJ": 9480, "DJE621":  9470,
    "SOB215":  9600,  "MRP648":  9630, "AST830":  10470,
}

_TR_TARGET_OCUP_BODEGA  = 300
_TR_TARGET_FUERA_ZONA   = 0.03
_TR_TARGET_UTIL_VEH     = 0.80
_TR_TARGET_EFICIENCIA   = 0.70


def _tr_parse_time_input(val_str):
    """Convierte 'HH:MM' en datetime.time, None si vacío."""
    import datetime
    if not val_str or str(val_str).strip() in ("", "nan", "None"):
        return None
    s = str(val_str).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _tr_time_to_min(t):
    if t is None:
        return 0.0
    return t.hour * 60 + t.minute + t.second / 60


def _safe_int(v):
    try: return int(float(str(v).replace(",", ".")))
    except Exception: return 0

def _safe_float(v):
    try: return float(str(v).replace(",", ".").replace("%", ""))
    except Exception: return 0.0

def render_tab_tablero():
    import io
    _tr_subtabs = st.tabs(["📊 Tablero del día", "📅 Histórico Mensual", "📆 Histórico Anual"])

    # ══════════════════════════════════════════════════════════════════
    # TAB 1 — TABLERO DEL DÍA
    # ══════════════════════════════════════════════════════════════════
    with _tr_subtabs[0]:
        import pandas as pd
        import datetime
        import io

        st.markdown("## 🚚 Tablero Ruteador")
        st.caption("Reemplaza el Excel 05-Tablero Ruteador. Fuentes: ANR (Chess) + Ventas Especiales + Licencias.")

        ss = st.session_state
        anr_raw = ss.get("anr_df")
        tc_anr  = ss.get("tc_anr")
        ddm     = ss.get("df_ddm")

        # v4.67: usar df_ddm_dict (cargado al subir Frescura) como fuente principal
        _ddm_dict = ss.get("df_ddm_dict", {})
        if _ddm_dict:
            # Construir lookups directamente desde el dict precalculado
            sku_bxp_pre   = {k: v["bxp"]     for k, v in _ddm_dict.items() if v["bxp"]     > 0}
            sku_hl_pre    = {k: v["hl_unit"]  for k, v in _ddm_dict.items() if v["hl_unit"] > 0}
            sku_peso_pre  = {k: v["kg_unit"]  for k, v in _ddm_dict.items() if v["kg_unit"] > 0}
        else:
            sku_bxp_pre = {}; sku_hl_pre = {}; sku_peso_pre = {}

        # ── CONFIGURACIÓN FIJA (expander colapsado por defecto) ─────────────────
        with st.expander("⚙️ Configuración fija (no cambia a diario)", expanded=False):
            st.caption("Editá estos valores cuando haya cambios de flota, targets o SLA operativo.")
            _cf1, _cf2, _cf3, _cf4, _cf5 = st.columns(5)
            with _cf1:
                total_camiones_flota = st.number_input(
                    "Flota total (camiones)", value=ss.get("tr_cfg_flota", 15),
                    step=1, key="tr_cfg_flota")
            with _cf2:
                sla_str = st.text_input(
                    "SLA entrega target (HH:MM)",
                    value=ss.get("tr_cfg_sla", "16:15"), key="tr_cfg_sla")
            with _cf3:
                ruteo_str = st.text_input(
                    "T. ruteo target (HH:MM)",
                    value=ss.get("tr_cfg_ruteo", "01:10"), key="tr_cfg_ruteo")
            with _cf4:
                ocup_target = st.number_input(
                    "Ocup. bodega target (UP)",
                    value=ss.get("tr_cfg_ocup", _TR_TARGET_OCUP_BODEGA),
                    step=10, key="tr_cfg_ocup")
            with _cf5:
                fz_target_pct = st.number_input(
                    "Fuera de zona target (%)",
                    value=ss.get("tr_cfg_fz", 3.0),
                    step=0.5, format="%.1f", key="tr_cfg_fz")
            fz_target = fz_target_pct / 100

            # Patentes y capacidades
            st.markdown("**Patentes y capacidades de camiones**")
            cam_data = [{"camion": k, "patente": v,
                         "capacidad_kg": _TR_CAPACIDADES_DEFAULT.get(v, 9480)}
                        for k, v in _TR_CAMIONES_DEFAULT.items()]
            df_cam_edit = st.data_editor(pd.DataFrame(cam_data),
                use_container_width=True, hide_index=True, key="tr_cam_cfg",
                column_config={
                    "camion": st.column_config.NumberColumn("N° Camión", disabled=True),
                    "patente": st.column_config.TextColumn("Patente"),
                    "capacidad_kg": st.column_config.NumberColumn("Cap. máx (kg)", step=10),
                })
            cam_patente = dict(zip(df_cam_edit["camion"], df_cam_edit["patente"]))
            cam_capkg   = dict(zip(df_cam_edit["patente"], df_cam_edit["capacidad_kg"]))

        # ── CARGA DEL DÍA ──────────────────────────────────────────────────────
        # v4.64.0: fecha por defecto = fecha detectada en el ANR (si está cargado)
        _fecha_default_tr = datetime.date.today()
        try:
            _anr_for_date = ss.get("anr_df") or ss.get("tc_anr")
            if _anr_for_date is not None:
                if hasattr(_anr_for_date, "copy"):
                    _anr_d = _anr_for_date.copy()
                else:
                    _anr_for_date.seek(0)
                    import pandas as _pd_date
                    _xl_date = _pd_date.ExcelFile(_anr_for_date)
                    _sh_date = next((s for s in _xl_date.sheet_names if s.upper() == "BASE"), _xl_date.sheet_names[0])
                    _anr_for_date.seek(0)
                    _anr_d = _pd_date.read_excel(_anr_for_date, sheet_name=_sh_date, header=1)
                    _anr_for_date.seek(0)
                _cols_up = [str(c).upper().strip() for c in _anr_d.columns]
                _fcol = None
                for _fc_cand in ("FECHA", "S"):
                    if _fc_cand in _cols_up:
                        _fcol = _anr_d.columns[_cols_up.index(_fc_cand)]
                        break
                if _fcol:
                    _fechas_anr = pd.to_datetime(_anr_d[_fcol], errors="coerce").dropna().dt.date.unique()
                    if len(_fechas_anr):
                        _fecha_default_tr = sorted(_fechas_anr)[-1]
        except Exception:
            pass  # Fallback a hoy si cualquier lectura falla

        st.markdown("### 📅 Carga del día")
        _d1, _d2, _d3, _d4, _d5, _d6, _d7 = st.columns(7)
        with _d1:
            fecha_dia = st.date_input("Fecha", value=_fecha_default_tr, key="tr_fecha")  # v4.64: default desde ANR
        with _d2:
            hora_real_str = st.text_input("Hora entrega real (HH:MM)",
                value="", placeholder="ej: 16:25", key="tr_hora_real")
        with _d3:
            ruteo_real_str = st.text_input("T. ruteo real (HH:MM)",
                value="", placeholder="ej: 00:20", key="tr_truteo_real")
        with _d4:
            fuera_zona_real = st.number_input("Fuera de zona (PDV)",
                value=0, step=1, key="tr_fz_real",
                help="Cantidad PDV fuera de zona")
        with _d5:
            segundas_vueltas = st.number_input("Segundas vueltas",
                value=0, step=1, key="tr_seg_vueltas")
        with _d6:
            visitas_teoricas = st.number_input("Visitas teóricas",
                value=650, step=10, key="tr_vis")
        with _d7:
            causa_demora = st.selectbox("Causa demora",
                options=["", "Chess", "Bees", "Cierre de venta",
                         "Problema sindical", "Corte de luz", "Otro"],
                key="tr_causa_demora")

        es_sabado = fecha_dia.weekday() == 5

        # ── LICENCIAS — tabla editable persistente (v4.70) ────────────────────
        st.markdown("### 🪪 Licencias conductores")
        st.caption("Editá directamente la tabla. Los datos persisten durante la sesión (no se borran con F5 mientras no se recargue la página).")

        # Inicializar tabla de licencias desde datos hardcoded si no existe
        _LICENCIAS_DEFAULT = [
            {"Chofer": "ALMARAZ",     "N° Licencia": "29468526", "Venc. Licencia": "15/10/2026", "Clases": "A.1.2 B.1 B.2"},
            {"Chofer": "ARANDA",      "N° Licencia": "28817797", "Venc. Licencia": "9/4/2027",   "Clases": "A.1.4 B.2 C.1"},
            {"Chofer": "BARRETO",     "N° Licencia": "29691454", "Venc. Licencia": "1/12/2027",  "Clases": "A.1.2 B.1 B.2"},
            {"Chofer": "BELLOCQ J",   "N° Licencia": "27547515", "Venc. Licencia": "7/1/2026",   "Clases": "B.1 C.1 C.3"},
            {"Chofer": "BELLOCQ P",   "N° Licencia": "31435975", "Venc. Licencia": "21/2/2026",  "Clases": "A.1.3 B.1 B.2"},
            {"Chofer": "CANTERO B.",  "N° Licencia": "92174412", "Venc. Licencia": "29/12/2026", "Clases": "A.1.2 B.1 B.2"},
            {"Chofer": "CEBALLOS",    "N° Licencia": "26597201", "Venc. Licencia": "21/8/2026",  "Clases": "A.1.2 B.1 B."},
            {"Chofer": "CHAVES",      "N° Licencia": "21161122", "Venc. Licencia": "13/1/2027",  "Clases": "A.1.2 B.2 C.1"},
            {"Chofer": "COLOMBO F.",  "N° Licencia": "26060015", "Venc. Licencia": "1/12/2026",  "Clases": "A.1.2 B.1 B.2"},
            {"Chofer": "COLOMBO S",   "N° Licencia": "32757797", "Venc. Licencia": "23/9/2026",  "Clases": "A.1.4 B.1"},
            {"Chofer": "DIAZ JAVIER", "N° Licencia": "28894519", "Venc. Licencia": "12/6/2026",  "Clases": "A.1.2 B.1 C.1"},
            {"Chofer": "GARCIA",      "N° Licencia": "21597244", "Venc. Licencia": "14/11/2026", "Clases": "A.1.2 B.2 C.1"},
            {"Chofer": "IBAX",        "N° Licencia": "36577450", "Venc. Licencia": "10/3/2028",  "Clases": "A.1.2 D.1 E.1"},
            {"Chofer": "KARCZEWSKI",  "N° Licencia": "33706322", "Venc. Licencia": "8/10/2026",  "Clases": "B.2 C.1 C"},
            {"Chofer": "PERALTA",     "N° Licencia": "33195760", "Venc. Licencia": "13/5/2026",  "Clases": "A.1.3 B.2 C.1"},
            {"Chofer": "PEREYRA",     "N° Licencia": "28545116", "Venc. Licencia": "17/11/2027", "Clases": "B.1 B.2 C.1 C"},
            {"Chofer": "PONCE",       "N° Licencia": "38361533", "Venc. Licencia": "23/4/2027",  "Clases": "A.1.2 B.1 C.3"},
            {"Chofer": "PRADIE",      "N° Licencia": "24907184", "Venc. Licencia": "6/5/2026",   "Clases": ""},
            {"Chofer": "QUIROGA",     "N° Licencia": "24390067", "Venc. Licencia": "6/5/2027",   "Clases": "A.1.4 B.1 B.2"},
            {"Chofer": "ROBLES",      "N° Licencia": "32946103", "Venc. Licencia": "6/1/2027",   "Clases": "A.1.2 B.1 B.2"},
            {"Chofer": "SCARAFIA",    "N° Licencia": "29468871", "Venc. Licencia": "10/12/2026", "Clases": "A.1.4 B.2 C.2"},
            {"Chofer": "TOTH",        "N° Licencia": "32101353", "Venc. Licencia": "16/4/2026",  "Clases": "A.1.2 B.1 B.2"},
            {"Chofer": "TREJO",       "N° Licencia": "32342194", "Venc. Licencia": "2/7/2027",   "Clases": "B.1 B.2 C.1 C"},
            {"Chofer": "VALDEZ M",    "N° Licencia": "34214857", "Venc. Licencia": "12/11/2027", "Clases": "A.1.3 B.2 C.1"},
            {"Chofer": "VILLAN",      "N° Licencia": "23720949", "Venc. Licencia": "30/1/2027",  "Clases": "A.1.2 B.2 C.3"},
            {"Chofer": "VILLALOBOS",  "N° Licencia": "30874160", "Venc. Licencia": "3/7/2027",   "Clases": "B.1 C.3"},
            {"Chofer": "VESPA",       "N° Licencia": "23874258", "Venc. Licencia": "21/5/2026",  "Clases": "A.1.2 B.1 B."},
            {"Chofer": "GARABATO",    "N° Licencia": "25909446", "Venc. Licencia": "27/5/2026",  "Clases": "A.2.1 B.1 B.2"},
            {"Chofer": "PETROV",      "N° Licencia": "22466598", "Venc. Licencia": "18/12/2029", "Clases": "A.1.4 B.1 B.2"},
            {"Chofer": "SIERRA",      "N° Licencia": "24406077", "Venc. Licencia": "5/9/2026",   "Clases": "A.1.4 B.1 B.2"},
            {"Chofer": "COLQUE",      "N° Licencia": "26691373", "Venc. Licencia": "28/5/2026",  "Clases": "A.1.2 B.2 C.1"},
        ]

        if "tr_licencias_data" not in ss:
            ss["tr_licencias_data"] = _LICENCIAS_DEFAULT.copy()

        _df_lic_edit = pd.DataFrame(ss["tr_licencias_data"])

        _lc_exp = st.expander("🪪 Ver / editar licencias de conductores", expanded=False)
        with _lc_exp:
            st.caption("Doble click en una celda para editar. Los cambios se mantienen durante la sesión.")
            _lcol1, _lcol2 = st.columns([6, 1])
            with _lcol1:
                _df_lic_edited = st.data_editor(
                    _df_lic_edit,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    key="tr_lic_editor",
                    column_config={
                        "Chofer":        st.column_config.TextColumn("Chofer", width="medium"),
                        "N° Licencia":   st.column_config.TextColumn("N° Licencia", width="medium"),
                        "Venc. Licencia": st.column_config.TextColumn("Venc. Licencia (dd/mm/aaaa)", width="medium"),
                        "Clases":        st.column_config.TextColumn("Clases habilitantes", width="large"),
                    }
                )
            with _lcol2:
                if st.button("💾 Guardar", key="tr_lic_save", use_container_width=True, type="primary"):
                    ss["tr_licencias_data"] = _df_lic_edited.to_dict("records")
                    st.success("✅ Guardado")
                    st.rerun()
                if st.button("↩️ Restaurar", key="tr_lic_reset", use_container_width=True):
                    ss["tr_licencias_data"] = _LICENCIAS_DEFAULT.copy()
                    st.rerun()

        # Construir df_lic para uso en la tabla (parsear fechas)
        df_lic = pd.DataFrame(ss["tr_licencias_data"]).copy()
        if not df_lic.empty and "Chofer" in df_lic.columns:
            df_lic.rename(columns={"Chofer": "nombre", "Venc. Licencia": "vencimiento"}, inplace=True)
            df_lic["nombre"] = df_lic["nombre"].astype(str).str.strip().str.upper()
            df_lic["vencimiento"] = pd.to_datetime(df_lic["vencimiento"], dayfirst=True, errors="coerce")
            df_lic = df_lic.dropna(subset=["vencimiento"])
            # Preview alertas
            _bloq_prev = df_lic[df_lic["vencimiento"] <= pd.Timestamp(fecha_dia)]
            _prox_prev = df_lic[
                (df_lic["vencimiento"] > pd.Timestamp(fecha_dia)) &
                (df_lic["vencimiento"] <= pd.Timestamp(fecha_dia) + pd.Timedelta(days=30))
            ]
            if not _bloq_prev.empty:
                st.error(f"🚫 {len(_bloq_prev)} licencia(s) VENCIDA(S): {', '.join(_bloq_prev['nombre'].tolist())}")
            if not _prox_prev.empty:
                st.warning(f"⚠️ {len(_prox_prev)} licencia(s) próximas a vencer (≤30 días): {', '.join(_prox_prev['nombre'].tolist())}")
        else:
            df_lic = pd.DataFrame(columns=["nombre", "vencimiento"])

        # ── VENTAS ESPECIALES (informativas) ──────────────────────────────────
        # v4.68: el CAR cubre TODO el despacho (VE + CHESS).
        # Solo mostramos el bloque VE como preview informativo.
        car_file_tr = ss.get("t1_car") or ss.get("tc_car")
        ve_auto = pd.DataFrame(columns=["camion", "sku", "bultos"])
        if car_file_tr is not None:
            try:
                car_file_tr.seek(0)
                _raw_ve_full = pd.read_excel(car_file_tr, sheet_name=0, header=None)
                car_file_tr.seek(0)
                _hdr_ve = _raw_ve_full.iloc[0].tolist()
                _sep_ve = None
                for _vi in range(1, min(60, len(_raw_ve_full))):
                    _rv = _raw_ve_full.iloc[_vi]
                    if _rv.isna().all():
                        _sep_ve = _vi
                        break
                _end_ve = _sep_ve if _sep_ve is not None else min(41, len(_raw_ve_full))
                _raw_ve = _raw_ve_full.iloc[0:_end_ve]
                _body_ve = _raw_ve.iloc[1:].copy()
                _body_ve.columns = _hdr_ve
                def _fc(df, *opts):
                    up = [str(c).upper().strip() for c in df.columns]
                    for o in opts:
                        if o.upper() in up:
                            return df.columns[up.index(o.upper())]
                    return None
                _c_cam = _fc(_body_ve, "TRANSPORTE", "N° CAMIÓN", "CAMION")
                _c_sku = _fc(_body_ve, "ARTÍCULO", "ARTICULO", "SKU", "CÓDIGO", "CODIGO")
                _c_blt = _fc(_body_ve, "BULTOS", "CANTIDAD")
                if _c_cam and _c_sku and _c_blt:
                    _body_ve["_cam"] = pd.to_numeric(_body_ve[_c_cam], errors="coerce")
                    _body_ve["_sku"] = pd.to_numeric(_body_ve[_c_sku], errors="coerce")
                    _body_ve["_blt"] = pd.to_numeric(_body_ve[_c_blt], errors="coerce").fillna(0)
                    _ve_rows = _body_ve.dropna(subset=["_cam", "_sku"])
                    _ve_rows = _ve_rows[(_ve_rows["_sku"] > 0) & (_ve_rows["_blt"] > 0)]
                    if not _ve_rows.empty:
                        ve_auto = pd.DataFrame({
                            "camion": _ve_rows["_cam"].astype(int),
                            "sku":    _ve_rows["_sku"].astype(int),
                            "bultos": _ve_rows["_blt"],
                        }).reset_index(drop=True)
            except Exception as _ve_err:
                st.caption(f"⚠️ VE no leídas del CAR: {_ve_err}")
        if not ve_auto.empty:
            with st.expander(f"🔵 Ventas Especiales — {len(ve_auto)} líneas (bloque azul CAR)", expanded=False):
                st.dataframe(ve_auto, use_container_width=True, hide_index=True,
                             column_config={
                                 "camion": st.column_config.NumberColumn("N° Camión", format="%d"),
                                 "sku":    st.column_config.NumberColumn("SKU", format="%d"),
                                 "bultos": st.column_config.NumberColumn("Bultos", format="%.2f"),
                             })

        st.markdown("---")

        # ── BOTÓN CALCULAR (v4.70) ──────────────────────────────────────────────
        # El tablero NO calcula automáticamente al cambiar los campos del día.
        # Solo calcula al presionar el botón o si ya hubo un cálculo previo en sesión.
        _btn_calcular = st.button(
            "⚡ Calcular Tablero",
            type="primary",
            use_container_width=False,
            key="tr_btn_calcular",
            help="Presioná después de completar todos los campos del día"
        )
        if _btn_calcular:
            ss["tr_calculado"] = True

        if not ss.get("tr_calculado", False):
            st.info("📋 Completá los campos del día y presioná **⚡ Calcular Tablero** para generar el informe.")
            return

        # ── REQUIERE ANR ────────────────────────────────────────────────────────
        if anr_raw is None and tc_anr is None:
            st.warning("⚠️ Cargá el ANR en la Tab **📁 Archivos** para calcular el tablero.")
            return

        # ── PARSEAR ANR ─────────────────────────────────────────────────────────
        try:
            if anr_raw is not None:
                anr_w = anr_raw.copy()
                first_col = str(anr_w.columns[0]).upper()
                if "UNNAMED" in first_col or first_col.startswith("0"):
                    anr_w.columns = anr_w.iloc[0].astype(str)
                    anr_w = anr_w.iloc[1:].reset_index(drop=True)
            else:
                tc_anr.seek(0)
                xl = pd.ExcelFile(tc_anr)
                sheet = xl.sheet_names[0]
                for s in xl.sheet_names:
                    if s.upper() == "BASE":
                        sheet = s
                        break
                tc_anr.seek(0)
                anr_w = pd.read_excel(tc_anr, sheet_name=sheet, header=1)
        except Exception as e:
            st.error(f"Error leyendo ANR: {e}")
            return

        def _get_col(df, *candidates):
            cols_up = [str(c).upper().strip() for c in df.columns]
            for c in candidates:
                if str(c).upper().strip() in cols_up:
                    return df.columns[cols_up.index(str(c).upper().strip())]
            return None

        col_camion   = _get_col(anr_w, "TRANSPORTE", "AP")
        col_fecha_a  = _get_col(anr_w, "FECHA", "S")
        col_up       = _get_col(anr_w, "UNIDAD PAQUETE", "Q")
        col_clientes = _get_col(anr_w, "TOTAL CLIENTES", "AY")
        # v4.65: ANR tiene "UNIDAD DE MEDIDA" = HL (no "HL" ni "PESO KG")
        col_peso     = _get_col(anr_w, "PESO KG", "AZ")   # generalmente None en Chess ANR
        col_hl       = _get_col(anr_w, "UNIDAD DE MEDIDA", "HL", "BB")  # ← fix v4.65
        col_rechazo  = _get_col(anr_w, "BULTOS RECHAZADOS", "L")
        col_sku_anr  = _get_col(anr_w, "ARTÍCULO", "ARTICULO", "SKU", "CÓDIGO", "CODIGO")
        col_blt_anr  = _get_col(anr_w, "BULTOS", "K")
        col_chofer_d = _get_col(anr_w, "DESCRIPCIÓN TRANSPORTE", "DESCRIPCION TRANSPORTE",
                                "DESCRIPCIÓN CHOFER", "AN", "DESCRIPCION CHOFER")
        col_cliente  = _get_col(anr_w, "CLIENTE", "W")

        if col_camion is None or col_fecha_a is None:
            st.error(f"No se encontraron columnas TRANSPORTE/FECHA en el ANR. Cols: {list(anr_w.columns[:10])}")
            return

        anr_w[col_fecha_a] = pd.to_datetime(anr_w[col_fecha_a], errors="coerce")
        df_dia = anr_w[anr_w[col_fecha_a].dt.date == fecha_dia].copy()

        if df_dia.empty:
            fechas_disp = sorted(anr_w[col_fecha_a].dropna().dt.date.unique())
            if fechas_disp:
                ultima_fecha = fechas_disp[-1]
                df_dia = anr_w[anr_w[col_fecha_a].dt.date == ultima_fecha].copy()
                fecha_dia = ultima_fecha
                st.info(f"ℹ️ Fecha sin datos. Cargando última disponible: **{ultima_fecha.strftime('%d/%m/%Y')}**")
            if df_dia.empty:
                st.warning("No hay datos en el ANR para ninguna fecha.")
                return

        # ── DDM disponibilidad — v4.68: solo flag, los dicts ya están en _ddm_dict ─
        _ddm_disponible = bool(_ddm_dict)

        # ── PDV / CLIENTES — fuente: ANR (único que tiene info de clientes) ────
        # Lógica: deduplicar por cliente dentro de cada camión para evitar
        # contar múltiples líneas del mismo cliente.
        pedidos_por_cam = pd.Series(dtype=float)
        if col_clientes is not None:
            # Columna binaria "TOTAL CLIENTES" (1 = cliente único por fila)
            df_dia[col_clientes] = pd.to_numeric(df_dia[col_clientes], errors="coerce").fillna(0)
            pedidos_por_cam = df_dia[df_dia[col_clientes] == 1].groupby(col_camion).size()
        elif col_cliente is not None:
            # Columna de código cliente: deduplicar con nunique()
            pedidos_por_cam = (
                df_dia.dropna(subset=[col_camion, col_cliente])
                .drop_duplicates(subset=[col_camion, col_cliente])
                .groupby(col_camion)
                .size()
            )
        else:
            pedidos_por_cam = df_dia.groupby(col_camion).size()

        rechazo_por_cam = df_dia.groupby(col_camion)[col_rechazo].sum() if col_rechazo else pd.Series(dtype=float)

        # ── UP · PALETAS · HL · PESO — fuente: CAR completo + DDM (v4.70) ─────
        #
        # Terminología ALINEADA con Proyección Picking (v4.70 — NO cambiar más):
        #   bultos_eq  = bultos_enteros + (unidades_sueltas / unidades_x_bulto)
        #   UP_sku     = bultos_eq / BXP          ← paletas reales (fraccionadas)
        #   AE_pall    = floor(UP_sku) por SKU    ← paletas enteras al AE
        #   pick_up    = UP_sku - AE_pall         ← fracción picking
        #   TOT_UP_cam = Σ pick_up  (suma fracciones = UP picking total camión)
        #   AE_pall_cam= Σ AE_pall  (suma paletas AE completas)
        #   TOT_PALL   = AE_pall_cam + TOT_UP_cam (total paletas equivalentes)
        #   HL         = hl_unit(DDM) × bultos_eq
        #   Peso       = kg_unit(DDM) × bultos_eq
        #
        # Ejemplo SKU BXP=50: 140 blt → UP=2.8 → AE=2 pall, pick=0.8 UP
        # Ejemplo SKU BXP=70: 140 blt → UP=2.0 → AE=2 pall, pick=0.0 UP
        #
        # El CAR contiene TODAS las líneas: VE (bloque azul) + CHESS juntos.

        bult_eq_por_cam = pd.Series(dtype=float)   # bultos equivalentes (informativo)
        up_por_cam      = pd.Series(dtype=float)   # UP picking por camion (fracciones, v4.70)
        paletas_por_cam = pd.Series(dtype=float)   # paletas reales = beq/BXP = UP/50
        hl_por_cam      = pd.Series(dtype=float)
        peso_por_cam    = pd.Series(dtype=float)

        _car_disponible = False
        if car_file_tr is not None and _ddm_dict:
            try:
                import openpyxl as _ox69
                car_file_tr.seek(0)
                _wb69 = _ox69.load_workbook(car_file_tr, read_only=True, data_only=True)
                car_file_tr.seek(0)
                _ws69 = _wb69.worksheets[0]
                _car_rows69 = list(_ws69.iter_rows(min_row=2, values_only=True))
                _wb69.close()

                from collections import defaultdict as _dd69
                import math as _math_tr
                _beq69 = _dd69(float)
                _up69  = _dd69(float)   # UP picking (fracciones)
                _ae69  = _dd69(float)   # paletas AE (enteras)
                _pal69 = _dd69(float)   # TOT PALL = AE + pick_up
                _hl69  = _dd69(float)
                _kg69  = _dd69(float)

                for _r69 in _car_rows69:
                    try:
                        _cam69 = _r69[10]; _sku69 = _r69[17]
                        if _cam69 is None or _sku69 is None: continue
                        _blt69 = float(_r69[19] or 0)
                        _uni69 = float(_r69[21] or 0)
                        _sid69 = int(float(str(_sku69)))
                        _d69   = _ddm_dict.get(_sid69, {})
                        _bxp69 = _d69.get("bxp", 50) or 50
                        _unb69 = _d69.get("un_bulto", 0)
                        _hl_u  = _d69.get("hl_unit",  0)
                        _kg_u  = _d69.get("kg_unit",  0)
                        # bultos equivalentes (enteros + fracción de sueltas)
                        _be69  = _blt69 + (_uni69 / _unb69 if _unb69 > 0 and _uni69 > 0 else 0)
                        _cam_i = int(float(str(_cam69)))
                        # UP = bultos_eq / BXP  (paletas reales, igual que Proyección Picking)
                        _up_sku = _be69 / _bxp69 if _bxp69 > 0 else 0.0
                        _ae_sku = float(_math_tr.floor(_up_sku))   # paletas AE completas
                        _pk_sku = _up_sku - _ae_sku                # fracción picking
                        _beq69[_cam_i] += _be69
                        _up69[_cam_i]  += _pk_sku    # UP picking (fracción)
                        _ae69[_cam_i]  += _ae_sku    # paletas AE (enteras)
                        _pal69[_cam_i] += _up_sku    # TOT PALL = AE + pick_up
                        _hl69[_cam_i]  += _hl_u * _be69
                        _kg69[_cam_i]  += _kg_u * _be69
                    except Exception:
                        pass

                bult_eq_por_cam = pd.Series(dict(_beq69), dtype=float)
                up_por_cam      = pd.Series(dict(_up69),  dtype=float)   # picking UP (fracción)
                ae_pall_por_cam = pd.Series(dict(_ae69),  dtype=float)   # paletas AE
                paletas_por_cam = pd.Series(dict(_pal69), dtype=float)   # TOT PALL
                hl_por_cam      = pd.Series(dict(_hl69),  dtype=float)
                peso_por_cam    = pd.Series(dict(_kg69),  dtype=float)
                _car_disponible = True
            except Exception as _car_err69:
                st.warning(f"⚠️ Error calculando UP desde CAR: {_car_err69}")

        if not _car_disponible:
            # Fallback: columna UNIDAD PAQUETE del ANR
            ae_pall_por_cam = pd.Series(dtype=float)
            if col_up is not None:
                up_por_cam      = df_dia.groupby(col_camion)[col_up].sum()
                bult_eq_por_cam = up_por_cam.copy()
                paletas_por_cam = up_por_cam.copy()
                ae_pall_por_cam = up_por_cam.apply(lambda x: float(int(x)))
                st.warning("⚠️ **UP calculado desde ANR** (cargá CAR + Frescura para mayor precisión).", icon="⚠️")
            if col_hl is not None:
                hl_por_cam = df_dia.groupby(col_camion)[col_hl].apply(
                    lambda x: pd.to_numeric(x, errors="coerce").fillna(0).sum())
            if col_peso is not None:
                peso_por_cam = df_dia.groupby(col_camion)[col_peso].apply(
                    lambda x: pd.to_numeric(x, errors="coerce").fillna(0).sum())
        else:
            ae_pall_por_cam = ss.get("_tr_ae_pall", ae_pall_por_cam) if "_tr_ae_pall" not in dir() else ae_pall_por_cam

        # Caption fuente
        if _car_disponible and _ddm_disponible:
            _fuente_hlkg = "✅ UP=bultos_eq/BXP | Paletas=floor(UP/sku) | HL/Peso=DDM×bultos_eq — fuente CAR+DDM. v4.70"
        elif _car_disponible:
            _fuente_hlkg = "⚠️ UP desde CAR. HL/Peso sin DDM — cargá Frescura 3.0 para activarlos."
        else:
            _fuente_hlkg = "⚠️ UP desde ANR (fallback). Cargá CAR + Frescura para cálculo correcto."
        st.caption(_fuente_hlkg)

        # ── EXCLUSIÓN DE CAMIONES ──────────────────────────────────────────────
        # _todos_cams: unión de camiones con PDV (ANR) + camiones con UP (CAR)
        _cams_pdv = set(pedidos_por_cam.index.tolist()) if not pedidos_por_cam.empty else set()
        _cams_up  = set(up_por_cam.index.tolist())      if not up_por_cam.empty      else set()
        _todos_cams = sorted([c for c in (_cams_pdv | _cams_up | set(cam_patente.keys()))
                              if int(pedidos_por_cam.get(c, 0)) > 0 or float(up_por_cam.get(c, 0)) > 0])
        if "tr_excluidos" not in ss:
            ss["tr_excluidos"] = []
        ss["tr_excluidos"] = [c for c in ss["tr_excluidos"] if c in _todos_cams]

        # ── TOTALES BASE ────────────────────────────────────────────────────────
        def _calc_totales(excl):
            _mask = [c for c in _todos_cams if c not in excl]
            _ped  = pedidos_por_cam[pedidos_por_cam.index.isin(_mask)]  if not pedidos_por_cam.empty  else pd.Series(dtype=float)
            _up   = up_por_cam[up_por_cam.index.isin(_mask)]            if not up_por_cam.empty       else pd.Series(dtype=float)
            _pal  = paletas_por_cam[paletas_por_cam.index.isin(_mask)]  if not paletas_por_cam.empty  else pd.Series(dtype=float)
            _ae   = ae_pall_por_cam[ae_pall_por_cam.index.isin(_mask)]  if not ae_pall_por_cam.empty  else pd.Series(dtype=float)
            _bup  = bult_eq_por_cam[bult_eq_por_cam.index.isin(_mask)]  if not bult_eq_por_cam.empty  else pd.Series(dtype=float)
            _hl_s = hl_por_cam[hl_por_cam.index.isin(_mask)]            if not hl_por_cam.empty       else pd.Series(dtype=float)
            _kg   = peso_por_cam[peso_por_cam.index.isin(_mask)]        if not peso_por_cam.empty     else pd.Series(dtype=float)
            _rec  = rechazo_por_cam[rechazo_por_cam.index.isin(_mask)]  if not rechazo_por_cam.empty  else pd.Series(dtype=float)
            c_rep   = int((_ped > 0).sum())
            t_pdv   = int(_ped.sum())
            t_up    = float(_up.sum())    # UP picking (fracciones)
            t_pal   = float(_pal.sum())   # TOT PALL (AE + pick_up)
            t_ae    = float(_ae.sum())    # paletas AE completas
            t_bup   = float(_bup.sum())   # bultos equivalentes totales
            t_hl    = float(_hl_s.sum())
            t_kg    = float(_kg.sum())
            t_rec   = float(_rec.sum())
            _denom  = c_rep + segundas_vueltas
            t_ocup  = (t_pal - t_rec) / _denom if _denom > 0 else 0.0
            t_fzpct = fuera_zona_real / t_pdv if t_pdv > 0 else 0.0
            t_efic  = t_pdv / visitas_teoricas if visitas_teoricas > 0 else 0.0
            t_util  = c_rep / total_camiones_flota if total_camiones_flota > 0 else 0.0
            t_drop  = t_pal / t_pdv if t_pdv > 0 else 0.0
            t_ppal  = t_pal / c_rep if c_rep > 0 else 0.0
            return dict(c_rep=c_rep, t_pdv=t_pdv, t_up=t_up, t_pal=t_pal, t_ae=t_ae, t_bup=t_bup,
                        t_hl=t_hl, t_kg=t_kg, t_rec=t_rec, t_ocup=t_ocup,
                        t_fzpct=t_fzpct, t_efic=t_efic, t_util=t_util,
                        t_drop=t_drop, t_ppal=t_ppal)

        tv = _calc_totales(set(ss["tr_excluidos"]))
        camiones_en_reparto = tv["c_rep"]; total_pedidos = tv["t_pdv"]
        total_up = tv["t_up"]; total_bup = tv["t_bup"]; total_hl = tv["t_hl"]; total_peso_kg = tv["t_kg"]
        total_rechazos_up = tv["t_rec"]; paletas_total = tv["t_pal"]; ae_total = tv["t_ae"]
        ocup_bodega = tv["t_ocup"]; fuera_zona_pct = tv["t_fzpct"]
        eficiencia = tv["t_efic"]; util_vehiculos = tv["t_util"]
        drop_size = tv["t_drop"]; prom_paletas = tv["t_ppal"]
        conversion_eq = total_bup * 2.4

        sla_t       = _tr_parse_time_input(sla_str)
        ruteo_tgt_t = _tr_parse_time_input(ruteo_str)
        hora_real_t = _tr_parse_time_input(hora_real_str)
        ruteo_rlt   = _tr_parse_time_input(ruteo_real_str)
        sla_min     = _tr_time_to_min(sla_t)
        ruteo_tgt_m = _tr_time_to_min(ruteo_tgt_t)
        hora_rlt_m  = _tr_time_to_min(hora_real_t) if hora_real_t else None
        ruteo_rlt_m = _tr_time_to_min(ruteo_rlt)   if ruteo_rlt  else None
        sla_ok   = (hora_rlt_m <= sla_min) if hora_rlt_m is not None else None
        ruteo_ok = (ruteo_rlt_m <= ruteo_tgt_m) if ruteo_rlt_m is not None else None
        ocup_ok  = ocup_bodega >= ocup_target
        fz_ok    = fuera_zona_pct <= fz_target
        util_ok  = util_vehiculos >= _TR_TARGET_UTIL_VEH
        efic_ok  = eficiencia >= _TR_TARGET_EFICIENCIA
        kpi_vals = [v for v in [sla_ok, ruteo_ok, ocup_ok, fz_ok, util_ok] if v is not None]
        prod_ruteo = sum(kpi_vals) / max(len(kpi_vals), 1)

        # ══ DISPLAY ═══════════════════════════════════════════════════════════
        _sabado_tag = "&nbsp;·&nbsp;🗓 Sábado" if es_sabado else ""
        _hdr_html = (
            '<div style="background:linear-gradient(90deg,#1F3864 0%,#2e5fa3 100%);'
            'padding:14px 22px;border-radius:10px;margin:8px 0 18px 0;'
            'color:white;display:flex;align-items:center;gap:14px;">'
            f'<div style="font-size:22px;font-weight:800;letter-spacing:0.5px;">📊 Tablero — '
            f'{fecha_dia.strftime("%A %d/%m/%Y").capitalize()}{_sabado_tag}</div>'
            f'<div style="margin-left:auto;font-size:15px;font-weight:700;'
            f'background:{"#00C853" if prod_ruteo >= 0.7 else "#FF4B4B"};'
            f'padding:5px 14px;border-radius:20px;">'
            f'Prod. Ruteo: {prod_ruteo:.0%}</div>'
            '</div>'
        )
        st.markdown(_hdr_html, unsafe_allow_html=True)

        # KPI cards
        def _kpi_html(icon, label, real, target, ok):
            if ok is True:   bg, badge, txt = "#d4edda", "✓ OK", "#155724"
            elif ok is False: bg, badge, txt = "#f8d7da", "✗ NO", "#721c24"
            else:             bg, badge, txt = "#fff3cd", "– S/D", "#856404"
            return f"""
            <div style="background:{bg};border-radius:10px;padding:12px 16px;margin:2px;
                        min-height:78px;display:flex;flex-direction:column;justify-content:space-between;
                        border:1px solid {'#c3e6cb' if ok is True else ('#f5c6cb' if ok is False else '#ffeeba')}">
                <div style="font-size:12px;color:{txt};font-weight:700;">{icon} {label}</div>
                <div style="font-size:18px;font-weight:800;color:{txt};">{real}</div>
                <div style="font-size:11px;color:{txt};opacity:0.85;">
                    Target: {target} &nbsp;<span style="font-weight:800;">{badge}</span>
                </div>
            </div>"""

        k_cols = st.columns(6)
        kpi_cards = [
            ("🕐", "SLA Entrega",     hora_real_str or "Sin dato",              sla_str,                             sla_ok),
            ("⏱️", "T. Ruteo",        ruteo_real_str or "Sin dato",             ruteo_str,                            ruteo_ok),
            ("🏠", "Ocup. Bodega",    f"{ocup_bodega:.1f} UP",                  f"≥{ocup_target} UP",                 ocup_ok),
            ("📍", "Fuera Zona",      f"{fuera_zona_real} ({fuera_zona_pct:.1%})", f"≤{fz_target:.0%}",             fz_ok),
            ("🚗", "Util. Vehículos", f"{util_vehiculos:.0%}",                  f"≥{_TR_TARGET_UTIL_VEH:.0%}",       util_ok),
            ("📊", "Eficiencia",      f"{eficiencia:.1%}",                      f"≥{_TR_TARGET_EFICIENCIA:.0%}",     efic_ok),
        ]
        for col, (icon, label, real, target, ok) in zip(k_cols, kpi_cards):
            col.markdown(_kpi_html(icon, label, real, target, ok), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Métricas secundarias
        _ms = st.columns(6)
        _ms[0].metric("🚚 Camiones",        f"{camiones_en_reparto}/{total_camiones_flota}")
        _ms[1].metric("📦 PDV",             total_pedidos)
        _ms[2].metric("🏋 Bultos UP",       f"{paletas_total:.2f}", help="TOT PALL = paletas AE + UP picking (fracciones)")
        _ms[3].metric("📐 Paletas",         f"{paletas_total:.2f}", f"{prom_paletas:.2f}/cam")
        _ms[4].metric("💧 HL",              f"{total_hl:.2f}" if total_hl > 0 else "—")
        _ms[5].metric("⚖️ Peso (kg)",       f"{total_peso_kg:,.0f}")

        _ms2 = st.columns(3)
        _ms2[0].metric("📈 Drop Size",       f"{drop_size:.2f}")
        _ms2[1].metric("🔄 Conversión (kg)", f"{conversion_eq:,.0f}")
        _ms2[2].metric("📈 Prod. Ruteo",     f"{prod_ruteo:.0%}")

        if total_rechazos_up > 0:
            st.warning(f"⚠️ Rechazos del día: **{total_rechazos_up:.1f}** bultos UP")
        if causa_demora:
            st.info(f"📌 Causa de demora: **{causa_demora}**")
        if not _ddm_disponible:
            st.caption("⚠️ Frescura no cargada — HL/Peso desde ANR directo (menos preciso). Cargá Frescura 3.0 para valores exactos.")

        # ── TABLA DETALLE POR CAMIÓN ───────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 🚛 Detalle por Camión")

        # UI exclusión
        _ue1, _ue2 = st.columns([4, 1])
        with _ue1:
            _excl_sel = st.multiselect(
                "🗑️ Excluir camiones del informe (KPIs, tabla, PDF y Excel se recalculan)",
                options=_todos_cams, default=ss["tr_excluidos"],
                format_func=lambda c: f"Cam {c}", key="tr_excl_ms")
        with _ue2:
            if st.button("↩️ Revertir", key="tr_revert_excl",
                         use_container_width=True):
                ss["tr_excluidos"] = []
                st.rerun()
        if set(_excl_sel) != set(ss["tr_excluidos"]):
            ss["tr_excluidos"] = list(_excl_sel)
            st.rerun()

        _excluidos = set(ss["tr_excluidos"])
        if _excluidos:
            tv = _calc_totales(_excluidos)
            camiones_en_reparto = tv["c_rep"]; total_pedidos = tv["t_pdv"]
            total_up = tv["t_up"]; total_bup = tv["t_bup"]; total_hl = tv["t_hl"]; total_peso_kg = tv["t_kg"]
            total_rechazos_up = tv["t_rec"]; paletas_total = tv["t_pal"]; ae_total = tv["t_ae"]
            ocup_bodega = tv["t_ocup"]; fuera_zona_pct = tv["t_fzpct"]
            eficiencia = tv["t_efic"]; util_vehiculos = tv["t_util"]
            drop_size = tv["t_drop"]; prom_paletas = tv["t_ppal"]
            conversion_eq = total_bup * 2.4
            st.info(f"ℹ️ {len(_excluidos)} camión(es) excluido(s): {', '.join(str(c) for c in sorted(_excluidos))}")

        # Construir tabla de camiones
        tabla_rows = []
        for cam in sorted(set(list(pedidos_por_cam.index) + list(up_por_cam.index) + list(cam_patente.keys()))):
            if int(pedidos_por_cam.get(cam, 0)) == 0 and float(up_por_cam.get(cam, 0)) == 0: continue
            if cam in _excluidos: continue
            beq   = float(bult_eq_por_cam.get(cam, 0.0)) if not bult_eq_por_cam.empty else 0.0
            up_c  = float(up_por_cam.get(cam, 0.0))      if not up_por_cam.empty      else 0.0
            pals  = float(paletas_por_cam.get(cam, 0.0)) if not paletas_por_cam.empty else 0.0
            hl_c  = float(hl_por_cam.get(cam, 0.0))      if not hl_por_cam.empty      else 0.0
            kg_c  = float(peso_por_cam.get(cam, 0.0))    if not peso_por_cam.empty    else 0.0
            rec_c = float(rechazo_por_cam.get(cam, 0.0)) if not rechazo_por_cam.empty else 0.0
            pate  = str(cam_patente.get(cam, ""))
            cap_kg = cam_capkg.get(pate, 0) if pate else 0
            # v4.70: semáforo de peso (verde/rojo) en lugar de columna Cap.(kg)
            _peso_ok = (kg_c <= cap_kg) if cap_kg > 0 else None
            _peso_semaforo = ("🟢" if _peso_ok else "🔴") if _peso_ok is not None else "—"

            chofer = ""
            if col_chofer_d is not None:
                mask = df_dia[col_camion] == cam
                if mask.any():
                    vals = df_dia.loc[mask, col_chofer_d].dropna().unique()
                    if len(vals) > 0:
                        ch = str(vals[0]).strip()
                        if " - " in ch:
                            ch = ch.split(" - ", 1)[-1].strip()
                        chofer = ch.title() if ch.upper() == ch else ch
                        if chofer.upper() in ("SIN CHOFER ASIGNADO", "SIN ASIGNAR", ""):
                            chofer = ""

            venc_lic = ""; val_lic = ""; lic_bloqueado = False
            if not df_lic.empty and chofer:
                chofer_up = chofer.upper().strip()
                m = df_lic[df_lic["nombre"].str.upper() == chofer_up]
                if m.empty and len(chofer_up) >= 4:
                    m = df_lic[df_lic["nombre"].str.upper().str.startswith(chofer_up[:6])]
                if m.empty and len(chofer_up) >= 4:
                    m = df_lic[df_lic["nombre"].str.upper().str.contains(chofer_up[:5], na=False)]
                if not m.empty:
                    venc = m.iloc[0]["vencimiento"]
                    if pd.notna(venc):
                        venc_lic = venc.strftime("%d/%m/%Y")
                        lic_bloqueado = pd.Timestamp(fecha_dia) >= venc
                        val_lic = "🚫 BLOQUEADO" if lic_bloqueado else "✅ OK"

            tabla_rows.append({
                "N° Cam":    cam,
                "Chofer":    chofer,
                "PDV":       int(pedidos_por_cam.get(cam, 0)),
                "Bultos":    round(beq,  2),   # bultos equivalentes
                "Bultos UP": round(pals, 2),   # TOT PALL = AE + pick_up (= paletas_por_cam)
                "Paletas":   round(float(ae_pall_por_cam.get(cam, 0.0)) if not ae_pall_por_cam.empty else 0.0, 0),
                "Patente":   pate,
                "Peso(kg)":  round(kg_c, 1),
                "⚖️":         _peso_semaforo,  # semáforo cap. peso
                "HL":        round(hl_c, 2),
                "Rechazos":  round(rec_c, 2),
                "Venc. Lic.": venc_lic,
                "Lic.":      val_lic,
                "_bloqueado": lic_bloqueado,
                "_cap_kg":    int(cap_kg) if cap_kg else 0,
                "_peso_ok":   _peso_ok,
            })

        if tabla_rows:
            df_tab = pd.DataFrame(tabla_rows)

            def _style_tabla(row):
                styles = [""] * len(row)
                if row.get("_bloqueado", False):
                    styles = ["background-color:#fff0f0"] * len(row)
                    try:
                        styles[list(row.index).index("Lic.")] = \
                            "background-color:#ff4b4b;color:white;font-weight:bold;text-align:center"
                    except Exception: pass
                else:
                    try:
                        if row.get("Lic.", "") == "✅ OK":
                            styles[list(row.index).index("Lic.")] = \
                                "background-color:#d4edda;color:#155724;font-weight:bold;text-align:center"
                    except Exception: pass
                # Semáforo peso rojo
                try:
                    if row.get("_peso_ok") is False:
                        styles[list(row.index).index("Peso(kg)")] = \
                            "background-color:#fff0f0;color:#c0392b;font-weight:bold"
                except Exception: pass
                return styles

            df_display = df_tab.drop(columns=["_bloqueado", "_cap_kg", "_peso_ok"], errors="ignore")
            styled = df_display.style.apply(_style_tabla, axis=1)

            st.dataframe(styled, use_container_width=True, hide_index=True,
                column_config={
                    "N° Cam":    st.column_config.NumberColumn("N° Cam",    format="%d"),
                    "PDV":       st.column_config.NumberColumn("PDV",       format="%d"),
                    "Bultos":    st.column_config.NumberColumn("Bultos",    format="%.2f",
                                    help="Bultos equivalentes = bultos enteros + unidades sueltas convertidas"),
                    "Bultos UP": st.column_config.NumberColumn("Bultos UP", format="%.2f",
                                    help="TOT PALL = paletas AE enteras + UP picking (fracción). Igual que Proyección Picking."),
                    "Paletas":   st.column_config.NumberColumn("Pal. AE",   format="%.0f",
                                    help="Paletas completas cargadas por autoelevador (floor de beq/BXP)"),
                    "⚖️":         st.column_config.TextColumn("Peso",       help="🟢 = dentro de capacidad | 🔴 = excede capacidad"),
                    "HL":        st.column_config.NumberColumn("HL",        format="%.2f"),
                    "Peso(kg)":  st.column_config.NumberColumn("Peso(kg)",  format="%.0f"),
                    "Rechazos":  st.column_config.NumberColumn("Rechazos",  format="%.1f"),
                    "Venc. Lic.": st.column_config.TextColumn("Venc. Lic.", help="Vencimiento registro de conducir"),
                    "Lic.":      st.column_config.TextColumn("Lic.",
                                    help="Estado licencia de conducir. 🚫 BLOQUEADO = no puede salir a reparto."),
                },
                height=min(520, 55 + 38 * len(df_display)))

            _hl_tot = df_tab["HL"].sum() if "HL" in df_tab.columns else 0
            _bloc = df_tab[df_tab["_bloqueado"] == True] if "_bloqueado" in df_tab.columns else pd.DataFrame()
            _bloc_str = f"  🚫 <b>{len(_bloc)} LIC. BLOQUEADA(S)</b>" if len(_bloc) > 0 else ""
            st.markdown(f"""
            <div style="background:#1F3864;color:white;border-radius:8px;
                        padding:10px 18px;margin:6px 0;display:flex;gap:18px;flex-wrap:wrap;
                        font-size:13px;font-weight:700;align-items:center;">
                <span>TOTAL</span>
                <span>🚚 {len(df_tab)} cams</span>
                <span>📦 {df_tab['PDV'].sum()} PDV</span>
                <span>📦 {df_tab['Bultos'].sum():.0f} bultos</span>
                <span>🏋 {df_tab['Bultos UP'].sum():.2f} UP</span>
                <span>📐 {df_tab['Paletas'].sum():.0f} pal.AE</span>
                <span>💧 {_hl_tot:.2f} HL</span>
                <span>⚖️ {df_tab['Peso(kg)'].sum():,.0f} kg</span>
                <span>❌ {df_tab['Rechazos'].sum():.1f} rec</span>
                {_bloc_str}
            </div>
            """, unsafe_allow_html=True)
        else:
            st.info("No hay camiones activos para la fecha seleccionada.")
            df_tab = pd.DataFrame()

        # ══ EXPORTACIONES ═════════════════════════════════════════════════════
        st.markdown("---")
        st.markdown("#### 📤 Exportar")

        _ex1, _ex2, _ex3 = st.columns(3)

        # ── Excel 1: Exportación completa ─────────────────────────────────────
        with _ex1:
            st.markdown("**📊 Excel del día** (exportación completa)")
            _buf_xl = io.BytesIO()
            _total_bup_xl = sum(r.get("Bultos",0) for r in tabla_rows)
            with pd.ExcelWriter(_buf_xl, engine="openpyxl") as _wr:
                pd.DataFrame({
                    "KPI": ["Fecha","Camiones reparto","Total PDV","Bultos (equiv.)","Bultos UP","Paletas",
                             "HL","Peso(kg)","Drop Size","Eficiencia","Util. Vehículos",
                             "Fuera de zona (PDV)","Fuera de zona (%)","Ocup. bodega",
                             "Productividad ruteo","Causa demora","Segundas vueltas"],
                    "Valor": [fecha_dia.strftime("%d/%m/%Y"), camiones_en_reparto, total_pedidos,
                               round(total_bup,2), round(total_up,2), round(paletas_total,2),
                               round(total_hl,2), round(total_peso_kg,2), round(drop_size,2),
                               round(eficiencia,4), round(util_vehiculos,4),
                               fuera_zona_real, round(fuera_zona_pct,4),
                               round(ocup_bodega,2), round(prod_ruteo,4),
                               causa_demora, segundas_vueltas],
                    "Target": ["—","—","—","—","—","—","—","—","—",
                                f"≥{_TR_TARGET_EFICIENCIA:.0%}", f"≥{_TR_TARGET_UTIL_VEH:.0%}",
                                "—", f"≤{fz_target:.0%}", f"≥{ocup_target}", "≥70%","—","—"],
                    "Status": ["—","—","—","—","—","—","—","—","—",
                                "OK" if efic_ok else "NO", "OK" if util_ok else "NO",
                                "—", "OK" if fz_ok else "NO",
                                "OK" if ocup_ok else "NO", f"{prod_ruteo:.0%}","—","—"],
                }).to_excel(_wr, sheet_name="KPIs", index=False)
                if tabla_rows:
                    # Detalle completo (sin columnas internas)
                    _df_det = pd.DataFrame(tabla_rows).drop(columns=["_bloqueado","_cap_kg","_peso_ok","⚖️"], errors="ignore")
                    _df_det.to_excel(_wr, sheet_name="Detalle Camiones", index=False)
                    _df_det[["N° Cam","Chofer","PDV","Bultos","Bultos UP","Paletas","HL","Peso(kg)","Rechazos","Venc. Lic.","Lic."]].to_excel(
                        _wr, sheet_name="Resumen Camiones", index=False)
                _ve_cl = ve_auto.dropna(subset=["camion","sku","bultos"]) if not ve_auto.empty else pd.DataFrame()
                if not _ve_cl.empty:
                    _ve_cl.to_excel(_wr, sheet_name="Ventas Especiales", index=False)
                if _excluidos:
                    pd.DataFrame({"Cam excluidos": sorted(_excluidos)}).to_excel(_wr, sheet_name="Excluidos", index=False)
            _buf_xl.seek(0)
            st.download_button("⬇️ Descargar Excel completo", data=_buf_xl.getvalue(),
                file_name=f"{fecha_dia.strftime('%d%m%Y')}_tablero_ruteador_bkcc.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="tr_dl_xl", use_container_width=True)

        # ── Excel 2: Excel Tablero con formato A4 landscape ──────────────────
        with _ex2:
            st.markdown("**📋 Excel Tablero**")
            if st.button("🖨️ Generar Excel Tablero", key="tr_btn_xl2", use_container_width=True):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    from openpyxl.chart import BarChart, Reference
                    from openpyxl.worksheet.page import PageMargins

                    _wb2 = Workbook()
                    _ws2 = _wb2.active
                    _ws2.title = "Tablero Ruteador"

                    # ── Configurar página A4 landscape ──────────────────────────
                    _ws2.page_setup.orientation    = "landscape"
                    _ws2.page_setup.paperSize      = 9   # A4
                    _ws2.page_setup.fitToPage      = True
                    _ws2.page_setup.fitToWidth     = 1
                    _ws2.page_setup.fitToHeight    = 0
                    _ws2.sheet_properties.pageSetUpPr.fitToPage = True
                    _ws2.page_margins = PageMargins(
                        left=0.35, right=0.35, top=0.35, bottom=0.35,
                        header=0.2, footer=0.2
                    )

                    # ── Helpers ──────────────────────────────────────────────────
                    _NX  = "1F3864"; _GRX = "00C853"; _RX = "FF4B4B"
                    _LGX = "F2F2F2"; _BLX = "E8EAF6"; _AX = "FFF3CD"
                    def _fx(h): return PatternFill("solid", fgColor=h)
                    def _fo(bold=False, color="000000", sz=9):
                        return Font(bold=bold, color=color, size=sz, name="Arial")
                    def _al(h="center", v="center", w=False):
                        return Alignment(horizontal=h, vertical=v, wrap_text=w)
                    def _bd():
                        s = Side(style="thin", color="DDDDDD")
                        return Border(left=s, right=s, top=s, bottom=s)
                    def _cell(r, c, val, bold=False, color="000000", sz=9,
                              bg=None, halign="center", wrap=False, border=True):
                        _c = _ws2.cell(r, c, val)
                        _c.font = _fo(bold, color, sz)
                        _c.alignment = _al(halign, "center", wrap)
                        if bg: _c.fill = _fx(bg)
                        if border: _c.border = _bd()
                        return _c

                    row = 1

                    # ── FILA 1: Header título completo ──────────────────────────
                    _ws2.merge_cells(f"A{row}:N{row}")
                    _cell(row, 1,
                        f"TABLERO RUTEADOR — {fecha_dia.strftime('%A %d/%m/%Y').upper()}"
                        f"   |   Prod. Ruteo: {prod_ruteo:.0%}"
                        f"   |   Beccacece Hnos SA · DPO 2.1 · v{APP_VERSION}",
                        bold=True, color="FFFFFF", sz=11, bg=_NX)
                    _ws2.row_dimensions[row].height = 22; row += 1

                    # ── FILA 2: KPI cards (6 KPIs en 6 columnas, 2 columnas cada una) ──
                    _kh = ["SLA Entrega", "T. Ruteo", "Ocup. Bodega",
                           "Fuera Zona", "Util. Veh.", "Eficiencia"]
                    _kr = [hora_real_str or "—", ruteo_real_str or "—",
                           f"{ocup_bodega:.1f} UP",
                           f"{fuera_zona_real} ({fuera_zona_pct:.1%})",
                           f"{util_vehiculos:.0%}", f"{eficiencia:.1%}"]
                    _kt = [sla_str, ruteo_str, f"≥{ocup_target}",
                           f"≤{fz_target:.0%}", f"≥{_TR_TARGET_UTIL_VEH:.0%}",
                           f"≥{_TR_TARGET_EFICIENCIA:.0%}"]
                    _ko = [sla_ok, ruteo_ok, ocup_ok, fz_ok, util_ok, efic_ok]
                    # 6 KPIs → cols A:F (cada KPI ocupa 1 col de ~2 unidades)
                    # Pero tenemos 12 cols para camiones → usamos A:L (2 cols/KPI merged)
                    for _ki, (kh2, kr2, kt2, ko2) in enumerate(zip(_kh,_kr,_kt,_ko)):
                        _bgk = _GRX if ko2 is True else (_RX if ko2 is False else "FFC107")
                        _c2s = _ki * 2 + 1
                        _ws2.merge_cells(f"{get_column_letter(_c2s)}{row}:{get_column_letter(_c2s+1)}{row}")
                        _cell(row, _c2s,
                              f"{kh2}\n{kr2}\nTarget: {kt2}",
                              bold=True, color="FFFFFF" if ko2 is not None else "000000",
                              sz=8, bg=_bgk, wrap=True)
                    _ws2.row_dimensions[row].height = 40; row += 1

                    # ── FILA 3-4: Totales globales ────────────────────────────
                    _tl = ["Camiones","PDV","Bultos","Bultos UP","Paletas","HL","Peso(kg)","Drop Size","Rechazos","SV","FZ PDV","FZ %","Prod.Ruteo"]
                    _tv = [f"{camiones_en_reparto}/{total_camiones_flota}",
                           str(total_pedidos), f"{total_bup:.0f}", f"{total_up:.2f}", f"{paletas_total:.2f}",
                           f"{total_hl:.2f}", f"{total_peso_kg:,.0f}",
                           f"{drop_size:.2f}", f"{total_rechazos_up:.1f}",
                           str(segundas_vueltas), str(fuera_zona_real),
                           f"{fuera_zona_pct:.1%}", f"{prod_ruteo:.0%}"]
                    for _ti, (_tl2, _tv2) in enumerate(zip(_tl, _tv), 1):
                        _cell(row,   _ti, _tl2, bold=True, color=_NX, sz=8, bg=_BLX)
                        _cell(row+1, _ti, _tv2, bold=True, sz=10)
                    _ws2.row_dimensions[row].height   = 14
                    _ws2.row_dimensions[row+1].height = 18
                    row += 3

                    # ── TABLA CAMIONES ────────────────────────────────────────
                    if tabla_rows:
                        # v4.70: sin Cap.kg, semaforo peso, licencias siempre completas
                        _cols_h = ["N°","Chofer","PDV","Bultos\n(eq.)","Bultos\nUP","Pal.\nAE",
                                   "Patente","HL","Peso\n(kg)","Peso\nOK",
                                   "Rech.","Venc.\nLic.","Lic."]
                        for ci5, ch5 in enumerate(_cols_h, 1):
                            _cell(row, ci5, ch5, bold=True, color="FFFFFF", sz=8, bg=_NX, wrap=True)
                        _ws2.row_dimensions[row].height = 28; row += 1

                        for ri2, rr2 in enumerate(tabla_rows):
                            _bg2 = _LGX if ri2 % 2 else "FFFFFF"
                            _lic_txt = rr2.get("Lic.", "—") or "—"
                            _lic_xl  = "BLOQUEADO" if str(_lic_txt).startswith("🚫") else ("OK" if str(_lic_txt).startswith("✅") else "—")
                            _venc_xl = rr2.get("Venc. Lic.", "") or "—"
                            _peso_ok_xl = rr2.get("_peso_ok", None)
                            _peso_semaf = "OK" if _peso_ok_xl is True else ("EXCEDE" if _peso_ok_xl is False else "—")
                            _vv2 = [str(rr2["N° Cam"]), rr2["Chofer"] or "—",
                                    rr2["PDV"],
                                    round(rr2.get("Bultos", 0), 2),
                                    round(rr2.get("Bultos UP", 0), 2),
                                    round(rr2.get("Paletas", 0), 0),
                                    rr2["Patente"],
                                    round(rr2.get("HL", 0), 2),
                                    round(rr2.get("Peso(kg)", 0), 0),
                                    _peso_semaf,
                                    rr2["Rechazos"],
                                    _venc_xl,
                                    _lic_xl]
                            for ci6, vv6 in enumerate(_vv2, 1):
                                _cell(row, ci6, vv6, sz=8, bg=_bg2,
                                      halign="left" if ci6 == 2 else "center")
                            # Semaforo peso (col 10)
                            if _peso_ok_xl is False:
                                _ws2.cell(row, 10).fill = _fx(_RX)
                                _ws2.cell(row, 10).font = _fo(True, "FFFFFF", 8)
                            elif _peso_ok_xl is True:
                                _ws2.cell(row, 10).fill = _fx(_GRX)
                                _ws2.cell(row, 10).font = _fo(True, "FFFFFF", 8)
                            # Colorear columna Lic. (col 13)
                            _valid_col = 13
                            if _lic_xl == "BLOQUEADO":
                                _ws2.cell(row, _valid_col).fill = _fx(_RX)
                                _ws2.cell(row, _valid_col).font = _fo(True, "FFFFFF", 8)
                                for _ci_bl in range(1, len(_vv2)+1):
                                    if _ci_bl not in (_valid_col, 10):
                                        _ws2.cell(row, _ci_bl).fill = _fx("FFE8E8")
                            elif _lic_xl == "OK":
                                _ws2.cell(row, _valid_col).fill = _fx(_GRX)
                                _ws2.cell(row, _valid_col).font = _fo(True, "FFFFFF", 8)
                            _ws2.row_dimensions[row].height = 14; row += 1

                        # Fila TOTAL
                        _totv2 = ["TOTAL", "",
                                  sum(r["PDV"]                    for r in tabla_rows),
                                  round(sum(r.get("Bultos",0)     for r in tabla_rows), 0),
                                  round(sum(r.get("Bultos UP",0)  for r in tabla_rows), 2),
                                  round(sum(r.get("Paletas",0)    for r in tabla_rows), 0),
                                  "", "",
                                  round(sum(r.get("Peso(kg)",0)   for r in tabla_rows), 0),
                                  "",
                                  round(sum(r.get("Rechazos",0)   for r in tabla_rows), 1),
                                  "",""]
                        for ci7, tv7 in enumerate(_totv2, 1):
                            _cell(row, ci7, tv7, bold=True, color=_NX, sz=9, bg=_BLX)
                        _ws2.row_dimensions[row].height = 16; row += 2

                    # ── TABLA ANÁLISIS OPERATIVO ────────────────────────────
                    if tabla_rows:
                        _ws2.merge_cells(f"A{row}:N{row}")
                        _cell(row, 1, "ANÁLISIS OPERATIVO", bold=True, color="FFFFFF", sz=9, bg=_NX)
                        _ws2.row_dimensions[row].height = 16; row += 1

                        _hl_t  = sum(r.get("HL", 0) for r in tabla_rows)
                        _kg_t  = sum(r.get("Peso(kg)", 0) for r in tabla_rows)
                        _bup_t = sum(r.get("Bultos UP", 0) for r in tabla_rows)

                        _ah = ["N° Cam","Chofer","Bultos","Bultos UP","%Carga UP","HL","Peso(kg)","Peso OK","PDV","Drop Size"]
                        for _ai, _ah2 in enumerate(_ah, 1):
                            _cell(row, _ai, _ah2, bold=True, color="FFFFFF", sz=8, bg="2e5fa3")
                        _ws2.row_dimensions[row].height = 14; row += 1

                        for _ri3, _rr3 in enumerate(sorted(tabla_rows, key=lambda x: x.get("Bultos UP",0), reverse=True)):
                            _bg3 = _LGX if _ri3 % 2 else "FFFFFF"
                            _pct_bup = _rr3.get("Bultos UP",0) / _bup_t * 100 if _bup_t > 0 else 0
                            _kg_c2   = _rr3.get("Peso(kg)", 0)
                            _peso_ok_a = _rr3.get("_peso_ok", None)
                            _peso_semaf_a = "OK" if _peso_ok_a is True else ("EXCEDE" if _peso_ok_a is False else "—")
                            _ds2 = _rr3.get("Bultos UP",0) / _rr3["PDV"] if _rr3["PDV"] > 0 else 0
                            _av3 = [str(_rr3["N° Cam"]), _rr3["Chofer"] or "—",
                                    round(_rr3.get("Bultos",0), 0),
                                    round(_rr3.get("Bultos UP",0), 2), f"{_pct_bup:.1f}%",
                                    round(_rr3.get("HL", 0), 2),
                                    round(_kg_c2, 0), _peso_semaf_a,
                                    _rr3["PDV"], round(_ds2, 2)]
                            for _ci8, _vv8 in enumerate(_av3, 1):
                                _cell(row, _ci8, _vv8, sz=8, bg=_bg3,
                                      halign="left" if _ci8 == 2 else "center")
                            # Colorear semaforo peso
                            if _peso_ok_a is False:
                                _ws2.cell(row, 8).fill = _fx(_RX); _ws2.cell(row, 8).font = _fo(True,"FFFFFF",8)
                            elif _peso_ok_a is True:
                                _ws2.cell(row, 8).fill = _fx(_GRX); _ws2.cell(row, 8).font = _fo(True,"FFFFFF",8)
                            _ws2.row_dimensions[row].height = 13; row += 1
                        row += 1

                    # ── GRÁFICOS EMBEBIDOS ────────────────────────────────────
                    if tabla_rows:
                        try:
                            # Datos auxiliares para gráficos (cols O:R)
                            _gd_col  = 15  # col O
                            _gd_row_start = row
                            # Headers
                            for _ghi, _ghn in enumerate(["Camión","Bultos UP","HL","Peso(kg)","PDV"], 1):
                                _ws2.cell(row, _gd_col + _ghi - 1, _ghn).font = _fo(True, _NX, 8)
                            _gdr = row
                            for _gi2, _gr2 in enumerate(tabla_rows, 1):
                                _ws2.cell(_gdr + _gi2, _gd_col,     str(_gr2["N° Cam"]))
                                _ws2.cell(_gdr + _gi2, _gd_col + 1, round(_gr2["Bultos UP"], 1))
                                _ws2.cell(_gdr + _gi2, _gd_col + 2, round(_gr2.get("HL", 0), 2))
                                _ws2.cell(_gdr + _gi2, _gd_col + 3, round(_gr2.get("Peso(kg)", 0), 0))
                                _ws2.cell(_gdr + _gi2, _gd_col + 4, _gr2["PDV"])
                            _gd_end = _gdr + len(tabla_rows)

                            def _make_bar(col_offset, title, chart_anchor, w=14, h=8, color="4472C4"):
                                _bc = BarChart()
                                _bc.type = "col"; _bc.grouping = "clustered"
                                _bc.title = title; _bc.style = 2
                                _bc.width = w; _bc.height = h
                                _bc.dataLabels = None
                                _ref_d = Reference(_ws2, min_col=_gd_col + col_offset,
                                    min_row=_gdr, max_row=_gd_end)
                                _ref_c = Reference(_ws2, min_col=_gd_col,
                                    min_row=_gdr + 1, max_row=_gd_end)
                                _bc.add_data(_ref_d, titles_from_data=True)
                                _bc.set_categories(_ref_c)
                                _bc.series[0].graphicalProperties.solidFill = color
                                _ws2.add_chart(_bc, chart_anchor)

                            # 3 gráficos: Bultos UP | HL | Peso
                            _make_bar(1, f"Bultos UP por Camión — {fecha_dia.strftime('%d/%m/%Y')}",
                                      f"A{row}", w=12, h=9, color="1F3864")
                            _make_bar(2, f"HL por Camión — {fecha_dia.strftime('%d/%m/%Y')}",
                                      f"E{row}", w=12, h=9, color="2e5fa3")
                            _make_bar(3, f"Peso (kg) por Camión — {fecha_dia.strftime('%d/%m/%Y')}",
                                      f"I{row}", w=12, h=9, color="00796B")
                            row += 19
                        except Exception as _ge_xl:
                            _ws2.cell(row, 1, f"[Gráficos no disponibles: {_ge_xl}]")
                            row += 1

                    # ── ALERTAS LICENCIAS ────────────────────────────────────
                    _lic_alerts = [r for r in tabla_rows if str(r.get("Lic.", "")).startswith("🚫")]
                    if _lic_alerts:
                        _ws2.merge_cells(f"A{row}:N{row}")
                        _cell(row, 1, "⚠ ALERTAS — LICENCIAS VENCIDAS (NO PUEDEN SALIR A REPARTO)", bold=True, color="FFFFFF", sz=9, bg=_RX)
                        _ws2.row_dimensions[row].height = 16; row += 1
                        for _la in _lic_alerts:
                            _ws2.merge_cells(f"A{row}:N{row}")
                            _cell(row, 1,
                                  f"🚫 CAM {_la['N° Cam']} — {_la['Chofer'] or 'S/N'}: "
                                  f"LICENCIA VENCIDA ({_la['Venc. Lic.']})",
                                  bold=True, color=_RX, sz=9, bg="FFE8E8")
                            _ws2.row_dimensions[row].height = 14; row += 1
                        row += 1

                    # ── FOOTER ───────────────────────────────────────────────
                    _ws2.merge_cells(f"A{row}:N{row}")
                    _ws2.cell(row, 1,
                        f"Generado: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')} · "
                        f"Picking Orchestrator v{APP_VERSION} · Beccacece Hnos SA — DPO 2.1"
                        + (f" · Causa demora: {causa_demora}" if causa_demora else "")
                    ).font = _fo(False, "888888", 7)
                    _ws2.row_dimensions[row].height = 12

                    # ── ANCHOS DE COLUMNA (13 cols: N°/Chofer/PDV/Bultos/BultosUP/PalAE/Pat/HL/Peso/PesoOK/Rech/VencLic/Lic)
                    _col_widths = {
                        "A": 5,    # N° Cam
                        "B": 14,   # Chofer
                        "C": 5,    # PDV
                        "D": 9,    # Bultos (eq.)
                        "E": 9,    # Bultos UP
                        "F": 7,    # Pal. AE
                        "G": 8,    # Patente
                        "H": 7,    # HL
                        "I": 9,    # Peso (kg)
                        "J": 7,    # Peso OK
                        "K": 6,    # Rechazo
                        "L": 11,   # Venc. Lic
                        "M": 10,   # Lic.
                        "N": 4,    # extra
                    }
                    for _cl, _cw in _col_widths.items():
                        _ws2.column_dimensions[_cl].width = _cw

                    _buf_xl2 = io.BytesIO(); _wb2.save(_buf_xl2); _buf_xl2.seek(0)
                    ss["tr_xl2_bytes"] = _buf_xl2.getvalue()
                    ss["tr_xl2_fname"] = f"{fecha_dia.strftime('%d%m%Y')}_tablero_ruteador_bkcc.xlsx"
                    st.success("✅ Excel Tablero generado — A4 landscape, tabla + 3 gráficos + análisis")
                except Exception as _xe2:
                    import traceback
                    st.error(f"Error: {_xe2}")
                    with st.expander("Detalle"): st.code(traceback.format_exc())
            if ss.get("tr_xl2_bytes"):
                st.download_button("⬇️ Descargar Excel Tablero", data=ss["tr_xl2_bytes"],
                    file_name=ss.get("tr_xl2_fname","tablero_ruteador_bkcc.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="tr_dl_xl2", use_container_width=True)

        # ── PDF Tablero ────────────────────────────────────────────────────────
        with _ex3:
            st.markdown("**📄 PDF Tablero**")
            if st.button("🖨️ Generar PDF Tablero", key="tr_btn_pdf", use_container_width=True):
                try:
                    import base64
                    from reportlab.lib.pagesizes import A4, landscape
                    from reportlab.lib import colors
                    from reportlab.lib.units import mm
                    from reportlab.platypus import (SimpleDocTemplate, Table,
                        TableStyle, Paragraph, Spacer, Image as RLImage)
                    from reportlab.lib.styles import ParagraphStyle
                    from reportlab.graphics.shapes import Drawing, Rect, String, Line

                    _buf_pdf = io.BytesIO()
                    _PW, _PH = landscape(A4)
                    _doc = SimpleDocTemplate(_buf_pdf, pagesize=landscape(A4),
                        topMargin=6*mm, bottomMargin=6*mm,
                        leftMargin=8*mm, rightMargin=8*mm)

                    _NAVY  = colors.HexColor("#1F3864")
                    _GOLD  = colors.HexColor("#C9A84C")
                    _GREEN = colors.HexColor("#00C853")
                    _RED   = colors.HexColor("#FF4B4B")
                    _LGREY = colors.HexColor("#F5F5F5")
                    _BLUE2 = colors.HexColor("#2e5fa3")
                    _AMBER = colors.HexColor("#FFC107")
                    _usable_w = _PW - 16*mm

                    def _p(txt, sz=9, bold=False, color=colors.black):
                        return Paragraph(txt, ParagraphStyle("_px",
                            fontSize=sz, fontName="Helvetica-Bold" if bold else "Helvetica",
                            textColor=color, spaceAfter=0, leading=sz+2))

                    _story = []

                    # ── Header sin logo ──────────────────────────────────────
                    _hdr = Table([[
                        _p(f"TABLERO RUTEADOR — {fecha_dia.strftime('%A %d/%m/%Y').upper()}", 13, True, colors.white),
                        _p(f"Beccacece Hnos SA · DPO 2.1 · v{APP_VERSION}", 8, False, colors.HexColor("#BDD5EA")),
                        _p(f"Prod. Ruteo  <b>{prod_ruteo:.0%}</b>", 11, False,
                           _GREEN if prod_ruteo >= 0.7 else _RED),
                    ]], colWidths=["58%", "28%", "14%"])
                    _hdr.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(1,0),_NAVY),
                        ("BACKGROUND",(2,0),(2,0),colors.HexColor("#1a1a2e")),
                        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                        ("ALIGN",(2,0),(2,0),"CENTER"),
                        ("LEFTPADDING",(0,0),(-1,-1),8),
                        ("TOPPADDING",(0,0),(-1,-1),5),
                        ("BOTTOMPADDING",(0,0),(-1,-1),5),
                    ]))
                    _story.append(_hdr); _story.append(Spacer(1, 2*mm))

                    # KPI cards
                    def _kpi_cell(label, real, target, ok):
                        _bg = _GREEN if ok is True else (_RED if ok is False else colors.HexColor("#FFC107"))
                        _icon = "✓" if ok is True else ("✗" if ok is False else "–")
                        _tc = colors.white if ok is not None else colors.black
                        return (_p(f"<b>{label}</b><br/>{_icon} {real}<br/><font size=8>Target: {target}</font>",
                                   9, color=_tc), _bg)

                    _kpi_data = [
                        _kpi_cell("SLA Entrega",     hora_real_str or "Sin dato",              sla_str,                            sla_ok),
                        _kpi_cell("T. Ruteo",        ruteo_real_str or "Sin dato",             ruteo_str,                           ruteo_ok),
                        _kpi_cell("Ocup. Bodega",    f"{ocup_bodega:.1f} UP",                  f"≥{ocup_target}",                   ocup_ok),
                        _kpi_cell("Fuera Zona",      f"{fuera_zona_real} ({fuera_zona_pct:.1%})", f"≤{fz_target:.0%}",             fz_ok),
                        _kpi_cell("Util. Veh.",      f"{util_vehiculos:.0%}",                  f"≥{_TR_TARGET_UTIL_VEH:.0%}",      util_ok),
                        _kpi_cell("Eficiencia",      f"{eficiencia:.1%}",                      f"≥{_TR_TARGET_EFICIENCIA:.0%}",    efic_ok),
                    ]
                    _kpi_tbl = Table([[c[0] for c in _kpi_data]], colWidths=["16.66%"]*6)
                    _ks = [("GRID",(0,0),(-1,-1),0.5,colors.white),
                           ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                           ("LEFTPADDING",(0,0),(-1,-1),5)]
                    for _i2, _c2 in enumerate(_kpi_data):
                        _ks.append(("BACKGROUND",(_i2,0),(_i2,0),_c2[1]))
                    _kpi_tbl.setStyle(TableStyle(_ks))
                    _story.append(_kpi_tbl); _story.append(Spacer(1, 2*mm))

                    # Totales globales — fila compacta
                    _tot = Table([[
                        _p("Camiones",7,True,_NAVY), _p(f"{camiones_en_reparto}/{total_camiones_flota}",10,True),
                        _p("PDV",7,True,_NAVY),      _p(str(total_pedidos),10,True),
                        _p("Bultos UP",7,True,_NAVY), _p(f"{total_up:.1f}",10,True),
                        _p("Paletas",7,True,_NAVY),  _p(f"{paletas_total:.2f}",10,True),
                        _p("HL",7,True,_NAVY),       _p(f"{total_hl:.2f}",10,True),
                        _p("Peso kg",7,True,_NAVY),  _p(f"{total_peso_kg:,.0f}",10,True),
                        _p("Drop Size",7,True,_NAVY),_p(f"{drop_size:.2f}",10,True),
                        _p("Rechazos",7,True,_RED),  _p(f"{total_rechazos_up:.1f}",10,True,_RED if total_rechazos_up > 0 else colors.black),
                    ]])
                    _tot.setStyle(TableStyle([
                        ("GRID",(0,0),(-1,-1),0.3,colors.lightgrey),
                        *[("BACKGROUND",(i*2,0),(i*2,0),colors.HexColor("#E8EAF6")) for i in range(8)],
                        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                        ("LEFTPADDING",(0,0),(-1,-1),4),
                    ]))
                    _story.append(_tot); _story.append(Spacer(1, 2*mm))

                    # Graficos — barras UP + linea PDV (formato imagen de referencia)
                    if tabla_rows:
                        try:
                            _cam_labels   = [str(r["N° Cam"]) for r in tabla_rows]
                            _cam_bup      = [r.get("Bultos UP", 0) for r in tabla_rows]
                            _cam_pdv      = [r["PDV"] for r in tabla_rows]
                            _n = len(_cam_labels)
                            # Grafico amplio: barras UP (eje izq) + linea PDV (eje der)
                            _chart_full_w = float(_usable_w)
                            _chart_h_full = 38.0 * mm
                            _margin_l = 10*mm; _margin_b = 8*mm
                            _pw = _chart_full_w - _margin_l - 4*mm
                            _ph = _chart_h_full - _margin_b - 6*mm
                            _xs = _pw / max(_n, 1)
                            _mv_up  = max(_cam_bup) if _cam_bup else 1
                            _mv_pdv = max(_cam_pdv) if _cam_pdv else 1

                            _dr_main = Drawing(_chart_full_w, _chart_h_full + 8*mm)
                            # Titulo
                            _dr_main.add(String(_chart_full_w/2, _chart_h_full + 4*mm,
                                f"TABLERO RUTEADOR — {fecha_dia.strftime('%d/%m/%Y')}  |  Bultos UP por Camion / PDV",
                                fontSize=8, fontName="Helvetica-Bold", fillColor=_NAVY, textAnchor="middle"))
                            # Ejes
                            _dr_main.add(Line(_margin_l, _margin_b, _margin_l, _margin_b + _ph,
                                strokeColor=_NAVY, strokeWidth=0.8))
                            _dr_main.add(Line(_margin_l, _margin_b, _margin_l + _pw, _margin_b,
                                strokeColor=_NAVY, strokeWidth=0.8))
                            # Barras UP
                            for _bi, (_lbl, _bup_v) in enumerate(zip(_cam_labels, _cam_bup)):
                                _bx = _margin_l + _bi * _xs + _xs * 0.1
                                _bw = _xs * 0.65
                                _bh = (_bup_v / _mv_up) * _ph if _mv_up > 0 else 0
                                _dr_main.add(Rect(_bx, _margin_b, _bw, max(_bh, 1),
                                    fillColor=_BLUE2, strokeColor=colors.white, strokeWidth=0.3))
                                _dr_main.add(String(_bx + _bw/2, _margin_b + _bh + 1.2*mm,
                                    f"{_bup_v:.1f}", fontSize=5.5, fontName="Helvetica",
                                    fillColor=_NAVY, textAnchor="middle"))
                                _dr_main.add(String(_bx + _bw/2, _margin_b - 4.5*mm,
                                    _lbl, fontSize=5.5, fontName="Helvetica",
                                    fillColor=_NAVY, textAnchor="middle"))
                            # Linea PDV (eje derecho escalado)
                            _pdv_points = []
                            for _bi, _pdv_v in enumerate(_cam_pdv):
                                _cx = _margin_l + _bi * _xs + _xs * 0.425
                                _cy = _margin_b + (_pdv_v / _mv_pdv) * _ph if _mv_pdv > 0 else _margin_b
                                _pdv_points.append((_cx, _cy))
                            for _pi in range(len(_pdv_points) - 1):
                                _dr_main.add(Line(
                                    _pdv_points[_pi][0], _pdv_points[_pi][1],
                                    _pdv_points[_pi+1][0], _pdv_points[_pi+1][1],
                                    strokeColor=colors.HexColor("#00796B"), strokeWidth=1.5))
                            for _cx2, _cy2 in _pdv_points:
                                _dr_main.add(Rect(_cx2-1.5*mm, _cy2-1.5*mm, 3*mm, 3*mm,
                                    fillColor=colors.HexColor("#00796B"), strokeColor=colors.white, strokeWidth=0.3))
                            # Leyenda
                            _dr_main.add(Rect(_margin_l, _chart_h_full - 5*mm, 6*mm, 3*mm,
                                fillColor=_BLUE2, strokeColor=colors.white, strokeWidth=0.3))
                            _dr_main.add(String(_margin_l + 7*mm, _chart_h_full - 3.5*mm,
                                "Bultos UP", fontSize=6, fontName="Helvetica", fillColor=_NAVY))
                            _dr_main.add(Rect(_margin_l + 28*mm, _chart_h_full - 5*mm, 6*mm, 3*mm,
                                fillColor=colors.HexColor("#00796B"), strokeColor=colors.white, strokeWidth=0.3))
                            _dr_main.add(String(_margin_l + 35*mm, _chart_h_full - 3.5*mm,
                                "PDV", fontSize=6, fontName="Helvetica", fillColor=_NAVY))

                            _story.append(_dr_main); _story.append(Spacer(1, 1.5*mm))
                        except Exception as _ge:
                            _story.append(_p(f"[Grafico no disponible: {_ge}]", 7, color=colors.grey))

                    # ── Bloque analisis operativo compacto ──────────────────
                    if tabla_rows and paletas_total > 0:
                        try:
                            _bup_t2 = sum(r.get("Bultos UP",0) for r in tabla_rows)
                            _hl_t2  = sum(r.get("HL", 0) for r in tabla_rows)
                            _best_cam  = max(tabla_rows, key=lambda x: x.get("Bultos UP",0))
                            _worst_cam = min(tabla_rows, key=lambda x: x.get("Bultos UP",0))
                            _avg_bup   = _bup_t2 / len(tabla_rows) if tabla_rows else 0
                            _alertas_peso = [r for r in tabla_rows if r.get("_peso_ok") is False]

                            _an_rows = [
                                [_p("Analisis Operativo", 7, True, colors.white),
                                 _p("", 7), _p("", 7), _p("", 7),
                                 _p("", 7), _p("", 7)],
                                [_p("Prom. Bultos UP/cam",7,True,_NAVY),
                                 _p(f"{_avg_bup:.1f}",8,True),
                                 _p("Mayor carga",7,True,_NAVY),
                                 _p(f"Cam {_best_cam['N° Cam']} — {_best_cam.get('Bultos UP',0):.1f} UP",8),
                                 _p("HL Total",7,True,_NAVY),
                                 _p(f"{_hl_t2:.2f}",8,True)],
                                [_p("Drop Size",7,True,_NAVY),
                                 _p(f"{drop_size:.2f}",8,True),
                                 _p("Menor carga",7,True,_NAVY),
                                 _p(f"Cam {_worst_cam['N° Cam']} — {_worst_cam.get('Bultos UP',0):.1f} UP",8),
                                 _p("Alertas peso excedido",7,True,_NAVY),
                                 _p(f"{len(_alertas_peso)} cam(s)",8,
                                    True if _alertas_peso else False,
                                    _RED if _alertas_peso else colors.black)],
                            ]
                            _an_tbl = Table(_an_rows,
                                colWidths=[_usable_w*0.18, _usable_w*0.14,
                                           _usable_w*0.14, _usable_w*0.22,
                                           _usable_w*0.18, _usable_w*0.14])
                            _an_s = [
                                ("BACKGROUND",(0,0),(-1,0),_NAVY),
                                ("SPAN",(0,0),(-1,0)),
                                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, _LGREY]),
                                ("GRID",(0,0),(-1,-1),0.3,colors.lightgrey),
                                ("TOPPADDING",(0,0),(-1,-1),2),
                                ("BOTTOMPADDING",(0,0),(-1,-1),2),
                                ("LEFTPADDING",(0,0),(-1,-1),4),
                            ]
                            _an_tbl.setStyle(TableStyle(_an_s))
                            _story.append(_an_tbl); _story.append(Spacer(1, 1.5*mm))
                        except Exception as _ae:
                            pass

                    # Tabla detalle por camion — 13 cols sin Cap.kg, con semaforo peso
                    if tabla_rows:
                        _det_hdr = ["N°","Chofer","PDV","Bultos","Blt.\nUP","Pal.\nAE","Patente","HL","Peso(kg)","Peso\nOK","Rech.","Venc.Lic.","Lic."]
                        _alertas_lic = []
                        _det_data = [_det_hdr]
                        for _i3, _r3 in enumerate(tabla_rows, 1):
                            vl = _r3.get("Venc. Lic.","") or "—"
                            va = _r3.get("Lic.","") or "—"
                            va_pdf = "BLOQUEADO" if str(va).startswith("🚫") else ("OK" if str(va).startswith("✅") else "—")
                            _peso_ok_p = _r3.get("_peso_ok", None)
                            _peso_semaf_p = "OK" if _peso_ok_p is True else ("EXCEDE" if _peso_ok_p is False else "—")
                            _det_data.append([
                                str(_r3["N° Cam"]),
                                _r3["Chofer"] or "—",
                                str(_r3["PDV"]),
                                f"{_r3.get('Bultos',0):.0f}",
                                f"{_r3.get('Bultos UP',0):.2f}",
                                f"{_r3.get('Paletas',0):.0f}",
                                _r3["Patente"] or "—",
                                f"{_r3.get('HL',0):.2f}",
                                f"{_r3.get('Peso(kg)',0):,.0f}",
                                _peso_semaf_p,
                                f"{_r3.get('Rechazos',0):.1f}",
                                vl,
                                va_pdf,
                            ])
                            if str(va).startswith("🚫"):
                                _alertas_lic.append(
                                    f"CAM {_r3['N° Cam']} — {_r3['Chofer'] or 'S/N'}: "
                                    f"LICENCIA VENCIDA ({vl}) — NO PUEDE SALIR A REPARTO"
                                )

                        _det_data.append(["TOTAL","",
                            str(sum(r["PDV"] for r in tabla_rows)),
                            f"{sum(r.get('Bultos',0) for r in tabla_rows):.0f}",
                            f"{sum(r.get('Bultos UP',0) for r in tabla_rows):.2f}",
                            f"{sum(r.get('Paletas',0) for r in tabla_rows):.0f}",
                            "","",
                            f"{sum(r.get('Peso(kg)',0) for r in tabla_rows):,.0f}",
                            "",
                            f"{sum(r.get('Rechazos',0) for r in tabla_rows):.1f}","",""])

                        _det_tbl = Table(_det_data, repeatRows=1,
                            colWidths=[
                                _usable_w * 0.038,  # N°
                                _usable_w * 0.110,  # Chofer
                                _usable_w * 0.038,  # PDV
                                _usable_w * 0.062,  # Bultos
                                _usable_w * 0.062,  # Blt.UP
                                _usable_w * 0.048,  # Pal.AE
                                _usable_w * 0.065,  # Patente
                                _usable_w * 0.055,  # HL
                                _usable_w * 0.070,  # Peso(kg)
                                _usable_w * 0.048,  # Peso OK
                                _usable_w * 0.048,  # Rech
                                _usable_w * 0.082,  # Venc.Lic
                                _usable_w * 0.082,  # Lic.
                            ])

                        _ds = [
                            ("BACKGROUND",(0,0),(-1,0),_NAVY),
                            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                            ("FONTSIZE",(0,0),(-1,0),7),
                            ("FONTSIZE",(0,1),(-1,-1),7),
                            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
                            ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white, _LGREY]),
                            ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#E8EAF6")),
                            ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
                            ("FONTSIZE",(0,-1),(-1,-1),7),
                            ("GRID",(0,0),(-1,-1),0.3,colors.lightgrey),
                            ("TOPPADDING",(0,0),(-1,-1),2),
                            ("BOTTOMPADDING",(0,0),(-1,-1),2),
                            ("LEFTPADDING",(0,0),(-1,-1),2),
                            ("RIGHTPADDING",(0,0),(-1,-1),2),
                            ("ALIGN",(0,0),(-1,-1),"CENTER"),
                            ("ALIGN",(1,1),(1,-1),"LEFT"),
                            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                        ]
                        # Colorear columna Lic. (indice 12) y Peso OK (indice 9) segun estado
                        for _ix, _rx in enumerate(tabla_rows, 1):
                            _peso_ok_px = _rx.get("_peso_ok", None)
                            if _rx.get("Lic.","").startswith("🚫"):
                                _ds += [
                                    ("BACKGROUND",(0,_ix),(-1,_ix),colors.HexColor("#FFE8E8")),
                                    ("BACKGROUND",(12,_ix),(12,_ix),_RED),
                                    ("TEXTCOLOR",(12,_ix),(12,_ix),colors.white),
                                    ("FONTNAME",(12,_ix),(12,_ix),"Helvetica-Bold")
                                ]
                            elif _rx.get("Lic.","").startswith("✅"):
                                _ds += [("BACKGROUND",(12,_ix),(12,_ix),_GREEN),
                                        ("TEXTCOLOR",(12,_ix),(12,_ix),colors.white)]
                            # Semaforo peso
                            if _peso_ok_px is False:
                                _ds += [("BACKGROUND",(9,_ix),(9,_ix),_RED),
                                        ("TEXTCOLOR",(9,_ix),(9,_ix),colors.white),
                                        ("FONTNAME",(9,_ix),(9,_ix),"Helvetica-Bold")]
                            elif _peso_ok_px is True:
                                _ds += [("BACKGROUND",(9,_ix),(9,_ix),_GREEN),
                                        ("TEXTCOLOR",(9,_ix),(9,_ix),colors.white)]
                        _det_tbl.setStyle(TableStyle(_ds))
                        _story.append(_det_tbl)

                        # Bloque alertas licencias vencidas
                        if _alertas_lic:
                            _story.append(Spacer(1, 2*mm))
                            _at_rows = [[_p("⚠️ ALERTAS — LICENCIAS VENCIDAS", 8, True, colors.white)]]
                            for _al_txt in _alertas_lic:
                                _at_rows.append([_p(_al_txt, 8, True, _RED)])
                            _at = Table(_at_rows, colWidths=[_usable_w])
                            _at.setStyle(TableStyle([
                                ("BACKGROUND",(0,0),(0,0),_RED),
                                ("BACKGROUND",(0,1),(-1,-1),colors.HexColor("#FFF3CD")),
                                ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                                ("LEFTPADDING",(0,0),(-1,-1),6),("GRID",(0,0),(-1,-1),0.3,_RED),
                            ]))
                            _story.append(_at)

                    # Footer compacto
                    _story.append(Spacer(1, 2*mm))
                    _fp = [f"Generado: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}",
                           f"Picking Orchestrator v{APP_VERSION}",
                           "Beccacece Hnos SA — DPO 2.1"]
                    if causa_demora:
                        _fp.append(f"Causa demora: {causa_demora}")
                    if _excluidos:
                        _fp.append(f"Camiones excluidos: {', '.join(str(c) for c in sorted(_excluidos))}")
                    _story.append(_p("  ·  ".join(_fp), 7, color=colors.grey))

                    _doc.build(_story)
                    _buf_pdf.seek(0)
                    _fname_pdf = f"{fecha_dia.strftime('%d%m%Y')}_tablero_ruteador_bkcc.pdf"
                    ss["tr_pdf_bytes"] = _buf_pdf.getvalue()
                    ss["tr_pdf_fname"] = _fname_pdf
                    st.success(f"✅ PDF listo")
                except Exception as _e_pdf:
                    import traceback
                    st.error(f"Error PDF: {_e_pdf}")
                    with st.expander("Detalle"): st.code(traceback.format_exc())

            if ss.get("tr_pdf_bytes"):
                st.download_button("⬇️ Descargar PDF", data=ss["tr_pdf_bytes"],
                    file_name=ss.get("tr_pdf_fname","tablero.pdf"),
                    mime="application/pdf", key="tr_dl_pdf",
                    use_container_width=True)

    # ══════════════════════════════════════════════════════════════════
    # TAB 2 — HISTÓRICO MENSUAL
    # ══════════════════════════════════════════════════════════════════
    with _tr_subtabs[1]:
        import pandas as pd
        import datetime
        import io
        try:
            import plotly.express as px
            import plotly.graph_objects as go
            from plotly.subplots import make_subplots
            HAS_PLOTLY = True
        except ImportError:
            HAS_PLOTLY = False

        ss = st.session_state
        if "hist_data" not in ss:
            ss["hist_data"] = []

        st.markdown("## 📅 Histórico Mensual — Dashboard")
        st.caption("Subí los Excel diarios generados desde el Tablero Ruteador. Cada archivo agrega un día.")

        # Upload
        _hu1, _hu2 = st.columns([3, 1])
        with _hu1:
            uploaded = st.file_uploader("📁 Excel(s) diarios del Tablero Ruteador",
                type=["xlsx"], accept_multiple_files=True, key="hist_uploader")
        with _hu2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️ Limpiar histórico", key="hist_clear", use_container_width=True):
                ss["hist_data"] = []
                st.rerun()

        if uploaded:
            nuevos = 0
            for f in uploaded:
                try:
                    df_kpi = pd.read_excel(f, sheet_name="KPIs", header=0)
                    kv = dict(zip(df_kpi["KPI"].astype(str), df_kpi["Valor"].astype(str)))
                    fecha_str = kv.get("Fecha","")
                    if not fecha_str: continue
                    fecha = pd.to_datetime(fecha_str, dayfirst=True, errors="coerce")
                    if pd.isna(fecha): continue
                    if fecha in [r["fecha"] for r in ss["hist_data"]]: continue
                    ss["hist_data"].append({
                        "fecha": fecha,
                        "camiones":      _safe_int(kv.get("Camiones reparto",0)),
                        "pedidos":       _safe_int(kv.get("Total PDV",0)),
                        "bultos_up":     _safe_float(kv.get("Bultos UP",0)),
                        "paletas":       _safe_float(kv.get("Paletas",0)),
                        "hl":            _safe_float(kv.get("HL",0)),
                        "peso_kg":       _safe_float(kv.get("Peso(kg)",0)),
                        "drop_size":     _safe_float(kv.get("Drop Size",0)),
                        "eficiencia":    _safe_float(kv.get("Eficiencia",0)),
                        "util_vehiculos":_safe_float(kv.get("Util. Vehículos",0)),
                        "fuera_zona_pct":_safe_float(kv.get("Fuera de zona (%)",0)),
                        "ocup_bodega":   _safe_float(kv.get("Ocup. bodega",0)),
                        "prod_ruteo":    _safe_float(kv.get("Productividad ruteo",0)),
                        "causa_demora":  kv.get("Causa demora",""),
                    })
                    nuevos += 1
                except Exception:
                    pass
            if nuevos > 0:
                ss["hist_data"].sort(key=lambda x: x["fecha"])
                st.success(f"✅ {nuevos} día(s) agregado(s).")

        if not ss["hist_data"]:
            st.info("Aún sin datos. Subí al menos un Excel diario del Tablero Ruteador.")
            st.markdown("""
**¿Cómo usar?**
1. Cada día generá el **Excel del día** desde el Tablero del día.
2. Acumulá los archivos en una carpeta local.
3. Subí todos aquí — el dashboard se arma automáticamente.
4. Exportá el consolidado mensual con PDF y Excel.
""")
            return

        df_h = pd.DataFrame(ss["hist_data"]).sort_values("fecha").reset_index(drop=True)
        df_h["fecha_str"] = df_h["fecha"].dt.strftime("%d/%m")
        df_h["dia_semana"] = df_h["fecha"].dt.day_name()
        n_dias = len(df_h)
        mes_str = df_h["fecha"].dt.strftime("%B %Y").mode()[0] if n_dias > 0 else "—"

        # ── HEADER MENSUAL ─────────────────────────────────────────────────────
        _prod_prom = df_h["prod_ruteo"].mean()
        _efic_prom = df_h["eficiencia"].mean()
        st.markdown(
            f'<div style="background:linear-gradient(90deg,#1F3864 0%,#2e5fa3 100%);'
            f'padding:14px 22px;border-radius:10px;margin:8px 0 18px 0;'
            f'color:white;display:flex;align-items:center;gap:14px;">'
            f'<div style="font-size:22px;font-weight:800;">📅 Dashboard Mensual — {mes_str}</div>'
            f'<div style="margin-left:auto;font-size:14px;font-weight:700;'
            f'background:{"#00C853" if _prod_prom >= 0.7 else "#FF4B4B"};'
            f'padding:5px 14px;border-radius:20px;">'
            f'Prod. Ruteo prom: {_prod_prom:.0%}</div>'
            f'</div>', unsafe_allow_html=True)

        # ── MÉTRICAS RESUMEN ───────────────────────────────────────────────────
        _rm = st.columns(7)
        _rm[0].metric("📆 Días", n_dias)
        _rm[1].metric("📦 PDV prom/día", f"{df_h['pedidos'].mean():.0f}", f"Total: {df_h['pedidos'].sum():.0f}")
        _rm[2].metric("🏋 UP prom/día",  f"{df_h['bultos_up'].mean():.0f}")
        _rm[3].metric("💧 HL prom/día",  f"{df_h['hl'].mean():.2f}")
        _rm[4].metric("📊 Eficiencia",   f"{_efic_prom:.1%}")
        _rm[5].metric("🚗 Util. Veh.",   f"{df_h['util_vehiculos'].mean():.1%}")
        _rm[6].metric("📈 Drop Size",    f"{df_h['drop_size'].mean():.2f}")

        # ── ALERTAS ────────────────────────────────────────────────────────────
        alertas = []
        if _efic_prom < _TR_TARGET_EFICIENCIA:
            alertas.append(f"🔴 Eficiencia promedio **{_efic_prom:.1%}** < target {_TR_TARGET_EFICIENCIA:.0%}")
        if df_h["util_vehiculos"].mean() < _TR_TARGET_UTIL_VEH:
            alertas.append(f"🔴 Util. vehículos **{df_h['util_vehiculos'].mean():.1%}** < target {_TR_TARGET_UTIL_VEH:.0%}")
        if (df_h["fuera_zona_pct"] > _TR_TARGET_FUERA_ZONA).any():
            dias_fz = (df_h["fuera_zona_pct"] > _TR_TARGET_FUERA_ZONA).sum()
            alertas.append(f"⚠️ **{dias_fz}** día(s) con fuera de zona > {_TR_TARGET_FUERA_ZONA:.0%}")
        if (df_h["ocup_bodega"] < _TR_TARGET_OCUP_BODEGA).any():
            alertas.append(f"⚠️ **{(df_h['ocup_bodega'] < _TR_TARGET_OCUP_BODEGA).sum()}** día(s) con ocup. bodega < {_TR_TARGET_OCUP_BODEGA} UP")
        if n_dias >= 5:
            _trend = df_h["drop_size"].iloc[-3:].mean() - df_h["drop_size"].iloc[:3].mean()
            if _trend < -0.5:
                alertas.append(f"📉 Drop Size en caída: últimos 3 días {df_h['drop_size'].iloc[-3:].mean():.2f} vs primeros 3 días {df_h['drop_size'].iloc[:3].mean():.2f}")
            elif _trend > 0.5:
                alertas.append(f"📈 Drop Size en alza: últimos 3 días {df_h['drop_size'].iloc[-3:].mean():.2f} ✅")
        if df_h["causa_demora"].ne("").sum() >= 2:
            _td = df_h[df_h["causa_demora"].ne("")]["causa_demora"].value_counts()
            alertas.append(f"⚠️ Demora recurrente: **{_td.index[0]}** ({_td.iloc[0]} días)")

        if alertas:
            with st.expander("🚨 Alertas del período", expanded=True):
                for a in alertas:
                    st.markdown(f"- {a}")
        else:
            st.success("✅ Sin alertas críticas en el período.")

        if not HAS_PLOTLY:
            st.warning("Plotly no disponible. `pip install plotly`")
            return

        # ── GRÁFICOS ───────────────────────────────────────────────────────────
        NAVY_H = "#1F3864"; GOLD_H = "#C9A84C"; GREEN_H = "#00C853"; RED_H = "#FF4B4B"

        tab_kpi, tab_vol, tab_ops, tab_anal, tab_tabla = st.tabs([
            "📈 KPIs Servicio", "📦 Volumen", "🚚 Operación", "🔍 Análisis", "📋 Tabla"
        ])

        with tab_kpi:
            st.markdown("#### Indicadores de servicio diarios")
            fig = make_subplots(rows=2, cols=2,
                subplot_titles=("Eficiencia (%)", "Utilización Vehículos (%)",
                                "Fuera de Zona (%)", "Productividad de Ruteo (%)"),
                shared_xaxes=False)
            fig.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["eficiencia"]*100,
                marker_color=[GREEN_H if v>=_TR_TARGET_EFICIENCIA*100 else RED_H for v in df_h["eficiencia"]*100],
                text=[f"{v:.0f}%" for v in df_h["eficiencia"]*100], textposition="outside",
                textfont_size=9, name="Eficiencia"), row=1, col=1)
            fig.add_hline(y=_TR_TARGET_EFICIENCIA*100, line_dash="dash", line_color=NAVY_H, row=1, col=1)
            fig.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["util_vehiculos"]*100,
                marker_color=[GREEN_H if v>=_TR_TARGET_UTIL_VEH*100 else RED_H for v in df_h["util_vehiculos"]*100],
                text=[f"{v:.0f}%" for v in df_h["util_vehiculos"]*100], textposition="outside",
                textfont_size=9, name="Util. Veh."), row=1, col=2)
            fig.add_hline(y=_TR_TARGET_UTIL_VEH*100, line_dash="dash", line_color=NAVY_H, row=1, col=2)
            fig.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["fuera_zona_pct"]*100,
                marker_color=[RED_H if v>_TR_TARGET_FUERA_ZONA*100 else GREEN_H for v in df_h["fuera_zona_pct"]*100],
                text=[f"{v:.1f}%" for v in df_h["fuera_zona_pct"]*100], textposition="outside",
                textfont_size=9, name="FZ%"), row=2, col=1)
            fig.add_hline(y=_TR_TARGET_FUERA_ZONA*100, line_dash="dash", line_color=NAVY_H, row=2, col=1)
            fig.add_trace(go.Scatter(x=df_h["fecha_str"], y=df_h["prod_ruteo"]*100,
                mode="lines+markers+text", text=[f"{v:.0f}%" for v in df_h["prod_ruteo"]*100],
                textposition="top center", textfont_size=9,
                marker=dict(color=NAVY_H, size=8), line=dict(color=NAVY_H, width=2),
                name="Prod. Ruteo"), row=2, col=2)
            fig.add_hline(y=70, line_dash="dash", line_color=NAVY_H, row=2, col=2)
            fig.update_layout(height=500, showlegend=False,
                plot_bgcolor="white", paper_bgcolor="white",
                font=dict(family="Arial", size=11),
                margin=dict(t=40, b=20))
            fig.update_yaxes(ticksuffix="%")
            st.plotly_chart(fig, use_container_width=True)

        with tab_vol:
            st.markdown("#### Volumen operativo diario")
            fig2 = make_subplots(rows=2, cols=3,
                subplot_titles=("PDV Ruteados","Bultos UP","HL Ruteados",
                                "Drop Size (UP/PDV)","Paletas","Peso Total (kg)"))
            avg_pdv = df_h["pedidos"].mean()
            fig2.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["pedidos"], marker_color=NAVY_H,
                text=df_h["pedidos"], textposition="outside", textfont_size=9), row=1, col=1)
            fig2.add_hline(y=avg_pdv, line_dash="dot", line_color=GOLD_H,
                annotation_text=f"Prom:{avg_pdv:.0f}", row=1, col=1)
            fig2.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["bultos_up"], marker_color=GOLD_H,
                text=df_h["bultos_up"].round(0), textposition="outside", textfont_size=9), row=1, col=2)
            fig2.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["hl"], marker_color="#4CAF50",
                text=df_h["hl"].round(2), textposition="outside", textfont_size=9), row=1, col=3)
            fig2.add_trace(go.Scatter(x=df_h["fecha_str"], y=df_h["drop_size"],
                mode="lines+markers", marker=dict(color="#2196F3", size=8),
                line=dict(color="#2196F3", width=2)), row=2, col=1)
            fig2.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["paletas"], marker_color="#9C27B0",
                text=df_h["paletas"].round(1), textposition="outside", textfont_size=9), row=2, col=2)
            fig2.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["peso_kg"], marker_color="#FF9800"), row=2, col=3)
            fig2.update_layout(height=520, showlegend=False,
                plot_bgcolor="white", paper_bgcolor="white",
                font=dict(family="Arial", size=11))
            st.plotly_chart(fig2, use_container_width=True)

        with tab_ops:
            st.markdown("#### Operación — camiones y bodega")
            fig3 = make_subplots(rows=1, cols=3,
                subplot_titles=("Camiones en reparto","Ocup. Bodega (UP)","Rechazos"))
            fig3.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["camiones"],
                marker_color=NAVY_H, text=df_h["camiones"],
                textposition="outside", textfont_size=9), row=1, col=1)
            fig3.add_trace(go.Scatter(x=df_h["fecha_str"], y=df_h["ocup_bodega"],
                mode="lines+markers+text", text=df_h["ocup_bodega"].round(1),
                textposition="top center", textfont_size=9,
                marker=dict(color=[GREEN_H if v>=_TR_TARGET_OCUP_BODEGA else RED_H for v in df_h["ocup_bodega"]], size=9),
                line=dict(color=NAVY_H, width=2)), row=1, col=2)
            fig3.add_hline(y=_TR_TARGET_OCUP_BODEGA, line_dash="dash", line_color=RED_H, row=1, col=2)
            if "rechazos" in df_h.columns:
                fig3.add_trace(go.Bar(x=df_h["fecha_str"], y=df_h["rechazos"],
                    marker_color=RED_H), row=1, col=3)
            fig3.update_layout(height=330, showlegend=False,
                plot_bgcolor="white", paper_bgcolor="white",
                font=dict(family="Arial", size=11))
            st.plotly_chart(fig3, use_container_width=True)

            # Causas de demora
            _dm = df_h[df_h["causa_demora"].ne("")]["causa_demora"]
            if len(_dm) > 0:
                st.markdown("##### Causas de demora")
                _dmc = _dm.value_counts().reset_index(); _dmc.columns = ["Causa","Días"]
                _figdm = px.bar(_dmc, x="Causa", y="Días",
                    color_discrete_sequence=[NAVY_H], text="Días",
                    title="Frecuencia de causas de demora")
                _figdm.update_layout(height=250, plot_bgcolor="white",
                    font=dict(family="Arial", size=11), showlegend=False)
                _figdm.update_traces(textposition="outside")
                st.plotly_chart(_figdm, use_container_width=True)

        with tab_anal:
            st.markdown("#### Análisis de tendencias y correlaciones")
            if n_dias >= 3:
                # Tendencia eficiencia con línea de tendencia
                _fig_t = go.Figure()
                _fig_t.add_trace(go.Scatter(x=df_h["fecha_str"], y=df_h["eficiencia"]*100,
                    mode="lines+markers", name="Eficiencia",
                    line=dict(color=NAVY_H, width=2),
                    marker=dict(color=[GREEN_H if v>=_TR_TARGET_EFICIENCIA*100 else RED_H
                                       for v in df_h["eficiencia"]*100], size=10)))
                _fig_t.add_hline(y=_TR_TARGET_EFICIENCIA*100, line_dash="dash",
                    line_color=RED_H, annotation_text=f"Target {_TR_TARGET_EFICIENCIA:.0%}")
                # Media móvil si hay suficientes días
                if n_dias >= 5:
                    _ma = df_h["eficiencia"].rolling(3, min_periods=1).mean()
                    _fig_t.add_trace(go.Scatter(x=df_h["fecha_str"], y=_ma*100,
                        mode="lines", name="Media móvil 3d",
                        line=dict(color=GOLD_H, width=2, dash="dot")))
                _fig_t.update_layout(height=280, plot_bgcolor="white", paper_bgcolor="white",
                    font=dict(family="Arial", size=11), legend=dict(orientation="h", y=1.1),
                    title="Evolución de Eficiencia + Media Móvil")
                _fig_t.update_yaxes(ticksuffix="%")
                st.plotly_chart(_fig_t, use_container_width=True)

                # Drop size vs pedidos
                _fig_c = px.scatter(df_h, x="pedidos", y="drop_size",
                    text="fecha_str", size_max=12,
                    color=[GREEN_H if v>=_TR_TARGET_EFICIENCIA else RED_H for v in df_h["eficiencia"]],
                    labels={"pedidos":"PDV Ruteados","drop_size":"Drop Size (UP/PDV)"},
                    title="Drop Size vs PDV (color = OK/NO eficiencia)")
                _fig_c.update_traces(textposition="top center", textfont_size=8)
                _fig_c.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                    font=dict(family="Arial", size=11), showlegend=False)
                st.plotly_chart(_fig_c, use_container_width=True)

            else:
                st.info("Se necesitan al menos 3 días de datos para análisis de tendencias.")

            # Tabla de promedios por día de semana
            if n_dias >= 5:
                st.markdown("##### Rendimiento por día de semana")
                _dow = df_h.groupby("dia_semana").agg(
                    PDV=("pedidos","mean"), UP=("bultos_up","mean"),
                    Eficiencia=("eficiencia","mean"), Prod_Ruteo=("prod_ruteo","mean"),
                    Dias=("pedidos","count")
                ).round(2)
                _dow["Eficiencia"] = _dow["Eficiencia"].map(lambda x: f"{x:.1%}")
                _dow["Prod_Ruteo"] = _dow["Prod_Ruteo"].map(lambda x: f"{x:.1%}")
                st.dataframe(_dow, use_container_width=True)

        with tab_tabla:
            st.markdown("#### Datos diarios completos")
            _disp_cols = {
                "fecha_str":"Fecha","dia_semana":"Día","camiones":"Cams",
                "pedidos":"PDV","bultos_up":"Bultos UP","paletas":"Paletas",
                "hl":"HL","peso_kg":"Peso(kg)","drop_size":"Drop Size",
                "eficiencia":"Eficiencia","util_vehiculos":"Util.Veh.",
                "fuera_zona_pct":"FZ%","ocup_bodega":"Ocup.Bod.",
                "prod_ruteo":"Prod.Ruteo","causa_demora":"Demora",
            }
            df_disp = df_h[list(_disp_cols.keys())].rename(columns=_disp_cols)
            for _cp in ["Eficiencia","Util.Veh.","FZ%","Prod.Ruteo"]:
                df_disp[_cp] = df_disp[_cp].map(lambda x: f"{x:.1%}")

            def _hl_row(row):
                styles = [""] * len(row)
                try:
                    ef = float(str(row["Eficiencia"]).replace("%","")) / 100
                    if ef < _TR_TARGET_EFICIENCIA:
                        styles[df_disp.columns.get_loc("Eficiencia")] = "background-color:#FFE0E0"
                except Exception:
                    pass
                return styles

            st.dataframe(df_disp.style.apply(_hl_row, axis=1),
                         use_container_width=True, hide_index=True)

        # ── EXPORTAR MENSUAL ───────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📤 Exportar Histórico Mensual")
        _em1, _em2 = st.columns(2)

        with _em1:
            st.markdown("**📊 Excel consolidado mensual**")
            if st.button("📊 Generar Excel mensual", key="hist_gen_xl", use_container_width=True):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    _wbm = Workbook()
                    _wsm = _wbm.active; _wsm.title = "Histórico"
                    _NX2 = "1F3864"; _GRX2 = "00C853"; _RX2 = "FF4B4B"
                    def _fx2(h): return PatternFill("solid", fgColor=h)
                    def _fo2(bold=False, color="000000", sz=10): return Font(bold=bold, color=color, size=sz, name="Arial")
                    def _al2(h="center"): return Alignment(horizontal=h, vertical="center")
                    def _bd2():
                        s = Side(style="thin", color="CCCCCC")
                        return Border(left=s, right=s, top=s, bottom=s)

                    # Header
                    _wsm.merge_cells(f"A1:O1")
                    _hc = _wsm.cell(1,1,f"HISTÓRICO MENSUAL — {mes_str}   |   {n_dias} días   |   Beccacece Hnos SA · DPO 2.1")
                    _hc.fill = _fx2(_NX2); _hc.font = _fo2(True, "FFFFFF", 13); _hc.alignment = _al2()
                    _wsm.row_dimensions[1].height = 22

                    # Headers tabla
                    _hdr_m = ["Fecha","Día","Cams","PDV","Bultos UP","Paletas","HL","Peso(kg)",
                               "Drop Size","Eficiencia","Util.Veh.","FZ%","Ocup.Bod.","Prod.Ruteo","Demora"]
                    for ci, ch in enumerate(_hdr_m, 1):
                        _c = _wsm.cell(2, ci, ch); _c.fill = _fx2("2e5fa3"); _c.font = _fo2(True,"FFFFFF",10); _c.alignment = _al2(); _c.border = _bd2()
                    _wsm.row_dimensions[2].height = 16

                    # Datos
                    for ri, row_d in enumerate(ss["hist_data"], 3):
                        _ef = row_d["eficiencia"]
                        _rfill = _fx2("FFE8E8") if _ef < _TR_TARGET_EFICIENCIA else _fx2("FFFFFF") if ri % 2 == 0 else _fx2("F5F5F5")
                        _vals_m = [
                            row_d["fecha"].strftime("%d/%m/%Y"),
                            row_d["fecha"].strftime("%A"),
                            row_d["camiones"], row_d["pedidos"],
                            round(row_d["bultos_up"],1), round(row_d["paletas"],2),
                            round(row_d["hl"],2), round(row_d["peso_kg"],0),
                            round(row_d["drop_size"],2),
                            f"{row_d['eficiencia']:.1%}", f"{row_d['util_vehiculos']:.1%}",
                            f"{row_d['fuera_zona_pct']:.1%}", round(row_d["ocup_bodega"],1),
                            f"{row_d['prod_ruteo']:.1%}", row_d["causa_demora"] or "—"
                        ]
                        for ci, vv in enumerate(_vals_m, 1):
                            _c = _wsm.cell(ri, ci, vv); _c.fill = _rfill; _c.font = _fo2(sz=10); _c.alignment = _al2(); _c.border = _bd2()
                        _wsm.row_dimensions[ri].height = 14

                    # Col widths
                    for ci4, cw4 in enumerate([12,12,6,6,9,8,7,9,9,10,9,7,9,10,14], 1):
                        _wsm.column_dimensions[get_column_letter(ci4)].width = cw4

                    # Hoja resumen
                    _wsr = _wbm.create_sheet("Resumen")
                    _wsr.merge_cells("A1:C1")
                    _rc = _wsr.cell(1,1,f"RESUMEN MENSUAL — {mes_str}")
                    _rc.fill = _fx2(_NX2); _rc.font = _fo2(True,"FFFFFF",12); _rc.alignment = _al2()
                    _wsr.row_dimensions[1].height = 20
                    _res_data = [
                        ("Indicador","Valor","Target"),
                        ("Días registrados", n_dias, "—"),
                        ("PDV promedio/día", f"{df_h['pedidos'].mean():.0f}", "—"),
                        ("Bultos UP promedio/día", f"{df_h['bultos_up'].mean():.0f}", "—"),
                        ("HL promedio/día", f"{df_h['hl'].mean():.2f}", "—"),
                        ("Peso promedio/día (kg)", f"{df_h['peso_kg'].mean():,.0f}", "—"),
                        ("Drop Size promedio", f"{df_h['drop_size'].mean():.2f}", "—"),
                        ("Eficiencia promedio", f"{df_h['eficiencia'].mean():.1%}", f"≥{_TR_TARGET_EFICIENCIA:.0%}"),
                        ("Util. Vehículos promedio", f"{df_h['util_vehiculos'].mean():.1%}", f"≥{_TR_TARGET_UTIL_VEH:.0%}"),
                        ("Productividad Ruteo promedio", f"{df_h['prod_ruteo'].mean():.1%}", "≥70%"),
                        ("Días con eficiencia < target", str((df_h["eficiencia"] < _TR_TARGET_EFICIENCIA).sum()), "0"),
                        ("Días con fuera de zona > target", str((df_h["fuera_zona_pct"] > _TR_TARGET_FUERA_ZONA).sum()), "0"),
                        ("Días con ocup. bodega < target", str((df_h["ocup_bodega"] < _TR_TARGET_OCUP_BODEGA).sum()), "0"),
                    ]
                    for ri2, rrow2 in enumerate(_res_data, 2):
                        for ci2, vv2 in enumerate(rrow2, 1):
                            _c2 = _wsr.cell(ri2, ci2, vv2)
                            if ri2 == 2:
                                _c2.fill = _fx2("2e5fa3"); _c2.font = _fo2(True,"FFFFFF",10)
                            else:
                                _c2.fill = _fx2("F5F5F5") if ri2 % 2 == 0 else _fx2("FFFFFF")
                                _c2.font = _fo2(sz=10)
                            _c2.alignment = _al2("left" if ci2 == 1 else "center")
                            _c2.border = _bd2()
                        _wsr.row_dimensions[ri2].height = 15
                    for ci3, cw3 in enumerate([32,14,10], 1):
                        _wsr.column_dimensions[get_column_letter(ci3)].width = cw3

                    _buf_hm = io.BytesIO(); _wbm.save(_buf_hm); _buf_hm.seek(0)
                    _mes_nom = df_h["fecha"].dt.strftime("%Y-%m").mode()[0]
                    ss["hist_xl_bytes"] = _buf_hm.getvalue()
                    ss["hist_xl_fname"] = f"{_mes_nom}_historico_mensual_bkcc.xlsx"
                    st.success("✅ Excel mensual generado")
                except Exception as _exm:
                    import traceback; st.error(f"Error: {_exm}")
                    with st.expander("Detalle"): st.code(traceback.format_exc())

            if ss.get("hist_xl_bytes"):
                st.download_button("⬇️ Descargar Excel mensual", data=ss["hist_xl_bytes"],
                    file_name=ss.get("hist_xl_fname","historico.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="hist_dl_xl", use_container_width=True)

        with _em2:
            st.markdown("**📄 PDF resumen mensual**")
            if st.button("📄 Generar PDF mensual", key="hist_gen_pdf", use_container_width=True):
                try:
                    from reportlab.lib.pagesizes import A4, landscape
                    from reportlab.lib import colors
                    from reportlab.lib.units import mm
                    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                        Paragraph, Spacer)
                    from reportlab.lib.styles import ParagraphStyle

                    _buf_pm = io.BytesIO()
                    _PW2, _PH2 = landscape(A4)
                    _docm = SimpleDocTemplate(_buf_pm, pagesize=landscape(A4),
                        topMargin=8*mm, bottomMargin=8*mm,
                        leftMargin=10*mm, rightMargin=10*mm)
                    _usw2 = _PW2 - 20*mm
                    _NAVY2  = colors.HexColor("#1F3864")
                    _GREEN2 = colors.HexColor("#00C853")
                    _RED2   = colors.HexColor("#FF4B4B")
                    _LGREY2 = colors.HexColor("#F5F5F5")

                    def _p2(txt, sz=10, bold=False, color=colors.black):
                        return Paragraph(txt, ParagraphStyle("_p2",
                            fontSize=sz, fontName="Helvetica-Bold" if bold else "Helvetica",
                            textColor=color, spaceAfter=0, leading=sz+3))

                    _sm = []
                    # Header
                    _hdrm = Table([[
                        _p2(f"HISTÓRICO MENSUAL — {mes_str.upper()}   |   {n_dias} días   |   Beccacece Hnos SA · DPO 2.1", 13, True, colors.white),
                        _p2(f"Prod. Ruteo:<br/><b>{df_h['prod_ruteo'].mean():.0%}</b>", 11, False,
                            _GREEN2 if df_h['prod_ruteo'].mean()>=0.7 else _RED2),
                    ]], colWidths=["82%","18%"])
                    _hdrm.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(0,0),_NAVY2),
                        ("BACKGROUND",(1,0),(1,0),colors.HexColor("#1a1a2e")),
                        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                        ("ALIGN",(1,0),(1,0),"CENTER"),
                        ("LEFTPADDING",(0,0),(-1,-1),8),
                        ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                    ]))
                    _sm.append(_hdrm); _sm.append(Spacer(1,4*mm))

                    # Resumen métricas
                    _met_data = [
                        ("Días", str(n_dias)),
                        ("PDV prom/día", f"{df_h['pedidos'].mean():.0f}"),
                        ("UP prom/día", f"{df_h['bultos_up'].mean():.0f}"),
                        ("HL prom/día", f"{df_h['hl'].mean():.2f}"),
                        ("Eficiencia prom", f"{df_h['eficiencia'].mean():.1%}"),
                        ("Util. Veh. prom", f"{df_h['util_vehiculos'].mean():.1%}"),
                        ("Drop Size prom", f"{df_h['drop_size'].mean():.2f}"),
                        ("Prod. Ruteo prom", f"{df_h['prod_ruteo'].mean():.1%}"),
                    ]
                    _met_tbl = Table([[_p2(k,8,True,_NAVY2) for k,_ in _met_data],
                                      [_p2(v,11,True) for _,v in _met_data]])
                    _met_tbl.setStyle(TableStyle([
                        ("GRID",(0,0),(-1,-1),0.3,colors.lightgrey),
                        *[("BACKGROUND",(i,0),(i,0),colors.HexColor("#E8EAF6")) for i in range(len(_met_data))],
                        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                        ("LEFTPADDING",(0,0),(-1,-1),5),("ALIGN",(0,0),(-1,-1),"CENTER"),
                    ]))
                    _sm.append(_met_tbl); _sm.append(Spacer(1,4*mm))

                    # Alertas
                    if alertas:
                        _al_rows = [[_p2("🚨 ALERTAS DEL PERÍODO", 10, True, colors.white)]]
                        for _al2b in alertas:
                            _al2b_clean = _al2b.replace("**","").replace("*","")
                            _al_rows.append([_p2(f"• {_al2b_clean}", 9)])
                        _alt = Table(_al_rows, colWidths=[_usw2])
                        _alt.setStyle(TableStyle([
                            ("BACKGROUND",(0,0),(0,0),_RED2),
                            ("BACKGROUND",(0,1),(-1,-1),colors.HexColor("#FFF3CD")),
                            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                            ("LEFTPADDING",(0,0),(-1,-1),8),("GRID",(0,0),(-1,-1),0.3,_RED2),
                        ]))
                        _sm.append(_alt); _sm.append(Spacer(1,4*mm))

                    # Tabla datos diarios
                    _hdr_td = ["Fecha","Día","Cams","PDV","Bultos UP","Paletas","HL",
                               "Drop Size","Eficiencia","Util.Veh.","FZ%","Ocup.Bod.","Prod.Ruteo"]
                    _td_rows = [_hdr_td]
                    for row_td in ss["hist_data"]:
                        _ef2 = row_td["eficiencia"]
                        _td_rows.append([
                            row_td["fecha"].strftime("%d/%m/%y"),
                            row_td["fecha"].strftime("%a"),
                            str(row_td["camiones"]), str(row_td["pedidos"]),
                            f"{row_td['bultos_up']:.0f}", f"{row_td['paletas']:.1f}",
                            f"{row_td['hl']:.2f}", f"{row_td['drop_size']:.2f}",
                            f"{_ef2:.0%}", f"{row_td['util_vehiculos']:.0%}",
                            f"{row_td['fuera_zona_pct']:.1%}", f"{row_td['ocup_bodega']:.0f}",
                            f"{row_td['prod_ruteo']:.0%}",
                        ])
                    # Fila promedio
                    _td_rows.append([
                        "PROM.", "—", f"{df_h['camiones'].mean():.1f}", f"{df_h['pedidos'].mean():.0f}",
                        f"{df_h['bultos_up'].mean():.0f}", f"{df_h['paletas'].mean():.1f}",
                        f"{df_h['hl'].mean():.2f}", f"{df_h['drop_size'].mean():.2f}",
                        f"{df_h['eficiencia'].mean():.0%}", f"{df_h['util_vehiculos'].mean():.0%}",
                        f"{df_h['fuera_zona_pct'].mean():.1%}", f"{df_h['ocup_bodega'].mean():.0f}",
                        f"{df_h['prod_ruteo'].mean():.0%}",
                    ])

                    _cwm = [_usw2*v for v in [0.08,0.07,0.055,0.055,0.075,0.065,
                                               0.055,0.07,0.08,0.07,0.065,0.07,0.08]]
                    _tbl_td = Table(_td_rows, repeatRows=1, colWidths=_cwm)
                    _ds_td = [
                        ("BACKGROUND",(0,0),(-1,0),_NAVY2),
                        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                        ("FONTSIZE",(0,0),(-1,-1),9),
                        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
                        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white, _LGREY2]),
                        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#E8EAF6")),
                        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
                        ("GRID",(0,0),(-1,-1),0.3,colors.lightgrey),
                        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                        ("LEFTPADDING",(0,0),(-1,-1),4),("ALIGN",(0,0),(-1,-1),"CENTER"),
                    ]
                    # Colorear filas con eficiencia < target
                    for _ri4, _rd4 in enumerate(ss["hist_data"], 1):
                        if _rd4["eficiencia"] < _TR_TARGET_EFICIENCIA:
                            _ds_td.append(("BACKGROUND",(8,_ri4),(8,_ri4),colors.HexColor("#FFE0E0")))
                    _tbl_td.setStyle(TableStyle(_ds_td))
                    _sm.append(_tbl_td)

                    # Footer
                    _sm.append(Spacer(1,3*mm))
                    _sm.append(_p2(f"Generado: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')} · Picking Orchestrator v{APP_VERSION} · Beccacece Hnos SA — DPO 2.1", 7, color=colors.grey))

                    _docm.build(_sm)
                    _buf_pm.seek(0)
                    _mes_nom2 = df_h["fecha"].dt.strftime("%Y-%m").mode()[0]
                    ss["hist_pdf_bytes"] = _buf_pm.getvalue()
                    ss["hist_pdf_fname"] = f"{_mes_nom2}_historico_mensual_bkcc.pdf"
                    st.success("✅ PDF mensual generado")
                except Exception as _epm:
                    import traceback; st.error(f"Error PDF mensual: {_epm}")
                    with st.expander("Detalle"): st.code(traceback.format_exc())

            if ss.get("hist_pdf_bytes"):
                st.download_button("⬇️ Descargar PDF mensual", data=ss["hist_pdf_bytes"],
                    file_name=ss.get("hist_pdf_fname","historico.pdf"),
                    mime="application/pdf", key="hist_dl_pdf", use_container_width=True)

    # ══════════════════════════════════════════════════════════════════
    # TAB 3 — HISTÓRICO ANUAL
    # ══════════════════════════════════════════════════════════════════
    with _tr_subtabs[2]:
        import pandas as pd
        import io
        ss = st.session_state
        if "hist_anual_data" not in ss:
            ss["hist_anual_data"] = []

        st.markdown("## 📆 Histórico Anual — Dashboard")
        st.caption(
            "Subí los Excel de histórico mensual generados desde el Tablero. "
            "Podés subir meses de 2025 y 2026 — cada archivo agrega un mes cerrado."
        )

        _ha1, _ha2 = st.columns([3, 1])
        with _ha1:
            uploaded_anual = st.file_uploader(
                "📁 Excel(s) de Histórico Mensual (uno por mes)",
                type=["xlsx"], accept_multiple_files=True, key="hist_anual_uploader"
            )
        with _ha2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️ Limpiar anual", key="hist_anual_clear", use_container_width=True):
                ss["hist_anual_data"] = []
                st.rerun()

        if uploaded_anual:
            nuevos_a = 0
            for fa in uploaded_anual:
                try:
                    _xl_a = pd.ExcelFile(fa)
                    # Intentar leer hoja KPIs primero (mismo formato que histórico mensual)
                    _sheet_a = "KPIs" if "KPIs" in _xl_a.sheet_names else _xl_a.sheet_names[0]
                    df_ka = pd.read_excel(fa, sheet_name=_sheet_a, header=0)
                    if "KPI" in df_ka.columns and "Valor" in df_ka.columns:
                        kva = dict(zip(df_ka["KPI"].astype(str), df_ka["Valor"].astype(str)))
                        _fecha_a_str = kva.get("Fecha","")
                        if not _fecha_a_str:
                            continue
                        _fecha_a = pd.to_datetime(_fecha_a_str, dayfirst=True, errors="coerce")
                        if pd.isna(_fecha_a):
                            continue
                        _mes_key = _fecha_a.strftime("%Y-%m")
                        if _mes_key in [r["mes_key"] for r in ss["hist_anual_data"]]:
                            continue
                        # Leer hoja Histórico para promediar el mes
                        _df_hist_a = None
                        if "Histórico" in _xl_a.sheet_names:
                            _df_hist_a = pd.read_excel(fa, sheet_name="Histórico", header=0)
                        ss["hist_anual_data"].append({
                            "mes_key":       _mes_key,
                            "mes_str":       _fecha_a.strftime("%b %Y"),
                            "anio":          _fecha_a.year,
                            "mes_num":       _fecha_a.month,
                            "n_dias":        int(kva.get("Días registrados", 0)) or 1,
                            "pedidos_prom":  _safe_float(kva.get("PDV promedio/día", 0)),
                            "bultos_up":     _safe_float(kva.get("Bultos UP promedio/día", 0)),
                            "hl_prom":       _safe_float(kva.get("HL promedio/día", 0)),
                            "drop_size":     _safe_float(kva.get("Drop Size promedio", 0)),
                            "eficiencia":    _safe_float(kva.get("Eficiencia promedio", "0").replace("%","")),
                            "util_veh":      _safe_float(kva.get("Util. Vehículos promedio", "0").replace("%","")),
                            "prod_ruteo":    _safe_float(kva.get("Productividad Ruteo promedio", "0").replace("%","")),
                        })
                        nuevos_a += 1
                except Exception:
                    pass
            if nuevos_a > 0:
                ss["hist_anual_data"].sort(key=lambda x: x["mes_key"])
                st.success(f"✅ {nuevos_a} mes(es) agregado(s).")

        if not ss["hist_anual_data"]:
            st.info("Sin datos anuales. Subí los Excel de histórico mensual de cada mes.")
            st.markdown("""
**¿Cómo usar?**
1. Cada mes, exportá el **Excel mensual** desde Histórico Mensual.
2. Guardalo en una carpeta local por año (ej: `2025/`, `2026/`).
3. Subí todos aquí — el dashboard anual se arma automáticamente.
4. Podés subir meses de 2025 y 2026 juntos.
""")
        else:
            df_a = pd.DataFrame(ss["hist_anual_data"]).sort_values("mes_key").reset_index(drop=True)
            _anios = sorted(df_a["anio"].unique())
            _anio_str = " / ".join(str(a) for a in _anios)
            n_meses = len(df_a)

            # ── HEADER ANUAL ───────────────────────────────────────────────────
            _prod_a = df_a["prod_ruteo"].mean()
            _efic_a = df_a["eficiencia"].mean()
            st.markdown(
                f'<div style="background:linear-gradient(90deg,#1F3864 0%,#2e5fa3 100%);'
                f'padding:14px 22px;border-radius:10px;margin:8px 0 18px 0;'
                f'color:white;display:flex;align-items:center;gap:14px;">'
                f'<div style="font-size:22px;font-weight:800;">📆 Dashboard Anual — {_anio_str}</div>'
                f'<div style="margin-left:auto;font-size:14px;font-weight:700;'
                f'background:{"#00C853" if _prod_a >= 0.7 else "#FF4B4B"};'
                f'padding:5px 14px;border-radius:20px;">'
                f'Prod. Ruteo prom: {_prod_a:.0%}</div>'
                f'</div>', unsafe_allow_html=True)

            # ── MÉTRICAS RESUMEN ──────────────────────────────────────────────
            _ra = st.columns(6)
            _ra[0].metric("📆 Meses", n_meses)
            _ra[1].metric("📦 PDV prom/mes", f"{df_a['pedidos_prom'].mean():.0f}")
            _ra[2].metric("💧 HL prom/mes", f"{df_a['hl_prom'].mean():.2f}")
            _ra[3].metric("📊 Eficiencia prom", f"{_efic_a:.1%}")
            _ra[4].metric("🚗 Util. Veh. prom", f"{df_a['util_veh'].mean():.1%}")
            _ra[5].metric("📈 Drop Size prom", f"{df_a['drop_size'].mean():.2f}")

            # ── ALERTAS ANUALES ───────────────────────────────────────────────
            _alertas_a = []
            if _efic_a < _TR_TARGET_EFICIENCIA:
                _alertas_a.append(f"🔴 Eficiencia anual **{_efic_a:.1%}** < target {_TR_TARGET_EFICIENCIA:.0%}")
            if df_a["util_veh"].mean() < _TR_TARGET_UTIL_VEH:
                _alertas_a.append(f"🔴 Util. vehículos **{df_a['util_veh'].mean():.1%}** < target {_TR_TARGET_UTIL_VEH:.0%}")
            _meses_baja_efic = (df_a["eficiencia"] < _TR_TARGET_EFICIENCIA).sum()
            if _meses_baja_efic > 0:
                _alertas_a.append(f"⚠️ **{_meses_baja_efic}** mes(es) con eficiencia < target")
            if _alertas_a:
                for _aa in _alertas_a:
                    st.warning(_aa)
            else:
                st.success("✅ Todos los KPIs anuales en rango objetivo.")

            # ── TABLA MENSUAL ─────────────────────────────────────────────────
            try:
                import plotly.graph_objects as go
                _fig_a = go.Figure()
                _fig_a.add_trace(go.Bar(
                    x=df_a["mes_str"], y=df_a["eficiencia"]*100,
                    name="Eficiencia (%)", marker_color="#2e5fa3",
                ))
                _fig_a.add_trace(go.Scatter(
                    x=df_a["mes_str"], y=df_a["prod_ruteo"]*100,
                    name="Prod. Ruteo (%)", mode="lines+markers",
                    line=dict(color="#FF8C00", width=2),
                ))
                _fig_a.add_hline(y=_TR_TARGET_EFICIENCIA*100, line_dash="dash",
                                  line_color="red", annotation_text="Target Efic.")
                _fig_a.update_layout(
                    title="Evolución mensual — Eficiencia & Productividad Ruteo",
                    height=320, template="plotly_dark",
                    legend=dict(orientation="h", y=-0.2),
                    margin=dict(l=20,r=20,t=40,b=20),
                )
                st.plotly_chart(_fig_a, use_container_width=True)
            except Exception:
                pass  # Plotly no disponible — solo tabla

            # Tabla detalle
            _df_show_a = df_a[["mes_str","n_dias","pedidos_prom","bultos_up","hl_prom",
                                "drop_size","eficiencia","util_veh","prod_ruteo"]].copy()
            _df_show_a.columns = ["Mes","Días","PDV/día","UP/día","HL/día",
                                   "Drop Size","Eficiencia","Util.Veh.","Prod.Ruteo"]
            for _col_a in ["Eficiencia","Util.Veh.","Prod.Ruteo"]:
                _df_show_a[_col_a] = _df_show_a[_col_a].map(lambda x: f"{x:.1%}" if pd.notna(x) else "—")
            st.dataframe(_df_show_a, use_container_width=True, hide_index=True)

            # ── ANÁLISIS Y RECOMENDACIONES ────────────────────────────────────
            with st.expander("🔍 Análisis y recomendaciones anuales", expanded=True):
                _mejor_mes = df_a.loc[df_a["eficiencia"].idxmax(), "mes_str"]
                _peor_mes  = df_a.loc[df_a["eficiencia"].idxmin(), "mes_str"]
                _tendencia = "📈 en mejora" if len(df_a) >= 3 and df_a["eficiencia"].iloc[-1] > df_a["eficiencia"].iloc[0] else "📉 en baja" if len(df_a) >= 3 else "—"
                st.markdown(f"""
**📊 Análisis automático del período {_anio_str}:**

- **Mejor mes (eficiencia):** {_mejor_mes} — {df_a.loc[df_a['eficiencia'].idxmax(), 'eficiencia']:.1%}
- **Peor mes (eficiencia):** {_peor_mes} — {df_a.loc[df_a['eficiencia'].idxmin(), 'eficiencia']:.1%}
- **Tendencia:** {_tendencia}
- **Meses sobre target ({_TR_TARGET_EFICIENCIA:.0%}):** {(df_a['eficiencia'] >= _TR_TARGET_EFICIENCIA).sum()} / {n_meses}

**🎯 Recomendaciones DPO:**
- Si eficiencia < target en más del 30% de los meses → revisar SLA y proceso de ruteo.
- Si util. vehículos baja consistentemente → analizar carga por camión y mix de clientes.
- Meses de alta temporada (Nov-Feb) naturalmente más exigentes — comparar YoY.
- Cruzar con períodos de rechazos altos (ANR) para identificar meses críticos.
""")

            # ── EXPORTAR ──────────────────────────────────────────────────────
            st.markdown("#### 📤 Exportar Histórico Anual")
            _ea1, _ea2 = st.columns(2)
            with _ea1:
                if st.button("📊 Generar Excel anual", key="hist_anual_xl", use_container_width=True):
                    try:
                        import openpyxl as _oxa
                        from openpyxl.styles import (PatternFill as _PFa, Font as _FOa,
                                                     Alignment as _ALa, Border as _BDa, Side as _SDa)
                        _wba = _oxa.Workbook()
                        _wsa = _wba.active; _wsa.title = "Histórico Anual"
                        _NXa = "1F3864"
                        def _fxa(c): return _PFa("solid", fgColor=c)
                        def _foa(b=False, c="000000", s=10): return _FOa(bold=b, color=c, size=s, name="Calibri")
                        def _ala(h="center"): return _ALa(horizontal=h, vertical="center")
                        def _bda():
                            _s = _SDa(style="thin", color="CCCCCC")
                            return _BDa(left=_s, right=_s, top=_s, bottom=_s)
                        _hdrs_a = ["Mes","Días","PDV/día","UP/día","HL/día","Drop Size",
                                   "Eficiencia","Util.Veh.","Prod.Ruteo"]
                        for ci_a, h_a in enumerate(_hdrs_a, 1):
                            _c_a = _wsa.cell(1, ci_a, h_a)
                            _c_a.fill = _fxa(_NXa); _c_a.font = _foa(True,"FFFFFF",11)
                            _c_a.alignment = _ala(); _c_a.border = _bda()
                        for ri_a, row_a in df_a.iterrows():
                            _vals_a = [
                                row_a["mes_str"], row_a["n_dias"],
                                round(row_a["pedidos_prom"],1), round(row_a["bultos_up"],1),
                                round(row_a["hl_prom"],2), round(row_a["drop_size"],2),
                                f"{row_a['eficiencia']:.1%}", f"{row_a['util_veh']:.1%}",
                                f"{row_a['prod_ruteo']:.1%}",
                            ]
                            for ci_a2, v_a in enumerate(_vals_a, 1):
                                _c_a2 = _wsa.cell(ri_a+2, ci_a2, v_a)
                                _c_a2.fill = _fxa("F5F5F5") if ri_a % 2 == 0 else _fxa("FFFFFF")
                                _c_a2.font = _foa(sz=10); _c_a2.alignment = _ala()
                                _c_a2.border = _bda()
                        for _col_a2 in _wsa.columns:
                            _wsa.column_dimensions[_col_a2[0].column_letter].width = 14
                        _buf_a = io.BytesIO(); _wba.save(_buf_a); _buf_a.seek(0)
                        ss["hist_anual_xl_bytes"] = _buf_a.getvalue()
                        ss["hist_anual_xl_fname"] = f"{_anio_str.replace(' / ','_')}_historico_anual_bkcc.xlsx"
                        st.success("✅ Excel anual generado")
                    except Exception as _exa:
                        st.error(f"Error: {_exa}")
                if ss.get("hist_anual_xl_bytes"):
                    st.download_button("⬇️ Descargar Excel anual",
                        data=ss["hist_anual_xl_bytes"],
                        file_name=ss.get("hist_anual_xl_fname","historico_anual.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="hist_anual_dl_xl", use_container_width=True)


def main():
    st.set_page_config(
        page_title=f"Picking Orchestrator v{APP_VERSION}",
        page_icon="📦",
        layout="wide",
    )

    # v4.16 — Modo oscuro restaurado (sin override de CSS).
    # Streamlit usa el tema configurado en .streamlit/config.toml o el del sistema.

    # Banner header con logo
    st.markdown(
        f"""
        <div style="
            background: linear-gradient(135deg, #1a3a6b 0%, #2e5fa3 100%);
            padding: 14px 22px;
            border-radius: 8px;
            margin-bottom: 18px;
            color: white;
            display: flex;
            align-items: center;
            gap: 16px;
        ">
            <img src="data:image/png;base64,{_LOGO_B64}"
                 style="height: 52px; width: 52px; object-fit: contain; opacity: 0.92; flex-shrink: 0;"
                 alt="Beccacece Hnos" />
            <div>
                <div style="font-size: 22px; font-weight: 600;">
                    📦 Picking Orchestrator
                </div>
                <div style="font-size: 13px; opacity: 0.85; margin-top: 2px;">
                    Beccacece Hnos SA — Distribuidor CMQ &nbsp;·&nbsp; DPO 2.1 — Pilar Almacén
                    &nbsp;·&nbsp; v{APP_VERSION}
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_sidebar()

    tabs = st.tabs([
        "📁 Archivos",
        "📦 Planilla Carga",
        "📋 Resumen Camiones",
        "📊 Proyección Picking ×5",
        "🏷️ Clasificación",
        "🚛 Camiones T2",
        "🏆 Top SKUs",
        "🚚 Tablero Ruteador",
        "🖨️ Boletas",
        "💰 Cierre",
        "✅ Validación + Log",
    ])
    with tabs[0]:  render_tab_archivos()
    with tabs[1]:  render_tab_planilla()
    with tabs[2]:  render_tab_resumen()
    with tabs[3]:  render_tab_proyeccion()
    with tabs[4]:  render_tab_clasificacion()
    with tabs[5]:  render_tab_t2()
    with tabs[6]:  render_tab_top_skus()
    with tabs[7]:  render_tab_tablero()
    with tabs[8]:  render_tab_boletas()
    with tabs[9]:  render_tab_cierre()
    with tabs[10]: render_tab_validacion()


# ════════════════════════════════════════════════════════════════════════════
# TAB CIERRE — Dashboard financiero diario (v4.24)
# Fuente: SR.xlsx (col A=idCns, col B=dsCns, col C='A) TotVal Chess') + ANR.xlsx
# Fix v4.24: SR usa col[2] (TotVal Chess limpio) | ANR usa header=1 (real)
# ════════════════════════════════════════════════════════════════════════════

# Clientes CTA CTE predeterminados (imagen referencia — editables en runtime)
_DEFAULT_CTA_CTE = [
    {"codigo": 158,   "nombre": "AUTO GAS"},
    {"codigo": 2088,  "nombre": "S.A.T"},
    {"codigo": 3652,  "nombre": "EXFALO"},
    {"codigo": 11605, "nombre": "ACEVEDO"},
    {"codigo": 13262, "nombre": "STYL WILL"},
    {"codigo": 16007, "nombre": "EL ABASTECEDOR"},
    {"codigo": 17243, "nombre": "MARRELLA"},
    {"codigo": 17227, "nombre": "BOCA NEGRA"},
    {"codigo": 18428, "nombre": "BRUCE"},
    {"codigo": 8780,  "nombre": "EXFAKO"},
    {"codigo": 17149, "nombre": "LOMFALO SA"},
    {"codigo": 16345, "nombre": "ELI SOCIEDAD ANONIMA"},
    {"codigo": 17869, "nombre": "ELI S.A."},
]
_MUTUAL_CODIGO = 17954
_MUTUAL_NOMBRE = "MUTUAL"


def _cierre_load_sr(sr_file) -> pd.DataFrame:
    """
    Carga SR.xlsx y retorna DataFrame con idCns, dsCns, TotVal.
    Fix v4.24: TotVal toma col[2] = 'A) TotVal Chess' (fuente limpia de Chess),
    no col[5] que es la columna derivada D (= A+B+C, igual cuando B y C son NaN
    pero semánticamente incorrecta). Robusto ante variantes de nombre.
    """
    sr_file.seek(0)
    df = pd.read_excel(sr_file, sheet_name=0, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    cols = list(df.columns)

    # idCns siempre en col[0], dsCns en col[1]
    rename = {}
    if "idCns" not in cols:
        rename[cols[0]] = "idCns"
    if "dsCns" not in cols:
        rename[cols[1]] = "dsCns"

    # TotVal: buscar por nombre primero, fallback a col[2] (A) TotVal Chess)
    totvol_candidates = ["A) TotVal Chess", "TotVal", "TOTVAL", "TOTAL CHESS", "CHESS"]
    totvol_col = None
    for cand in totvol_candidates:
        if cand in cols:
            totvol_col = cand
            break
    if totvol_col is None:
        # fallback posicional: col[2] es siempre 'A) TotVal Chess' en el SR actual
        totvol_col = cols[2]
    if totvol_col != "TotVal":
        rename[totvol_col] = "TotVal"

    if rename:
        df = df.rename(columns=rename)

    df["idCns"] = pd.to_numeric(df["idCns"], errors="coerce")
    df["TotVal"] = pd.to_numeric(df["TotVal"], errors="coerce").fillna(0)
    return df[["idCns", "dsCns", "TotVal"]].dropna(subset=["idCns"])


def _cierre_load_anr(anr_file) -> pd.DataFrame:
    """
    Carga ANR.xlsx y retorna DataFrame normalizado.
    Fix v4.24: el ANR de Chess tiene los headers reales en la fila 1 del Excel
    (índice 1), no en la fila 0. Con header=0 todas las columnas quedan 'Unnamed'.
    Busca la hoja 'BASE' primero (nombre canónico del ANR de Chess).
    """
    anr_file.seek(0)
    xl = pd.ExcelFile(anr_file)

    # Hoja preferida: 'BASE' (ANR Chess) > 'ANR' > la de más columnas
    sheet = xl.sheet_names[0]
    for s in xl.sheet_names:
        if s.upper() == "BASE":
            sheet = s
            break
        if s.upper() == "ANR":
            sheet = s

    anr_file.seek(0)
    # Fix: header=1 porque la fila 0 del Excel son los Unnamed y la fila 1 son los headers reales
    df = pd.read_excel(anr_file, sheet_name=sheet, header=1)
    df.columns = [str(c).strip() for c in df.columns]

    # Mapeo por nombre exacto (los headers reales ya están disponibles con header=1)
    col_map = {}
    for c in df.columns:
        cn = c.upper().replace(" ", "")
        if c == "CLIENTE" and "CLIENTE" not in col_map:
            col_map["CLIENTE"] = c
        if c == "IMPORTE NETO" and "IMPORTE_NETO" not in col_map:
            col_map["IMPORTE_NETO"] = c
        if c == "TRANSPORTE" and "TRANSPORTE" not in col_map:
            col_map["TRANSPORTE"] = c
        if c == "SIN CARGO" and "SIN_CARGO" not in col_map:
            col_map["SIN_CARGO"] = c
        if c == "RECHAZO" and "RECHAZO" not in col_map:
            col_map["RECHAZO"] = c
        if c == "IMPORTE NETO RECHAZADO" and "IMPORTE_RECHAZADO" not in col_map:
            col_map["IMPORTE_RECHAZADO"] = c

    # Fallback por nombre normalizado si no hubo match exacto
    if "CLIENTE" not in col_map or "IMPORTE_NETO" not in col_map:
        for c in df.columns:
            cn = c.upper().replace(" ", "")
            if cn == "CLIENTE" and "CLIENTE" not in col_map:
                col_map["CLIENTE"] = c
            if "IMPORTENETO" in cn and "RECHAZADO" not in cn and "IMPORTE_NETO" not in col_map:
                col_map["IMPORTE_NETO"] = c
            if cn == "TRANSPORTE" and "TRANSPORTE" not in col_map:
                col_map["TRANSPORTE"] = c
            if "SINCARGO" in cn and "SIN_CARGO" not in col_map:
                col_map["SIN_CARGO"] = c
            if cn == "RECHAZO" and "IMPORTE" not in cn and "MOTIVO" not in cn and "RECHAZO" not in col_map:
                col_map["RECHAZO"] = c
            if ("IMPORTENETORECH" in cn or ("IMPORTENETO" in cn and "RECHAZADO" in cn)) \
                    and "IMPORTE_RECHAZADO" not in col_map:
                col_map["IMPORTE_RECHAZADO"] = c

    # Validar columnas mínimas requeridas
    for req in ("CLIENTE", "IMPORTE_NETO", "TRANSPORTE"):
        if req not in col_map:
            raise ValueError(
                f"ANR.xlsx — hoja '{sheet}': no se encontró la columna '{req}'. "
                f"Columnas disponibles: {list(df.columns)[:10]}…"
            )

    df2 = pd.DataFrame()
    df2["CLIENTE"]           = pd.to_numeric(df[col_map["CLIENTE"]], errors="coerce")
    df2["IMPORTE_NETO"]      = pd.to_numeric(df[col_map["IMPORTE_NETO"]], errors="coerce").fillna(0)
    df2["TRANSPORTE"]        = pd.to_numeric(df[col_map["TRANSPORTE"]], errors="coerce")
    df2["SIN_CARGO"]         = df[col_map["SIN_CARGO"]].astype(str).str.strip().str.upper() \
                               if "SIN_CARGO" in col_map else "NO"
    df2["RECHAZO"]           = pd.to_numeric(df[col_map["RECHAZO"]], errors="coerce").fillna(0) \
                               if "RECHAZO" in col_map else 0.0
    df2["IMPORTE_RECHAZADO"] = pd.to_numeric(df[col_map["IMPORTE_RECHAZADO"]], errors="coerce").fillna(0) \
                               if "IMPORTE_RECHAZADO" in col_map else 0.0
    return df2.dropna(subset=["CLIENTE", "TRANSPORTE"])


def _build_cierre_df(df_sr: pd.DataFrame, df_anr: pd.DataFrame,
                     cta_cte_list: list, cobro_anticipado: float,
                     venta_especial_map: dict) -> pd.DataFrame:
    """
    Construye la tabla principal del Cierre por camión.
    v4.24: cobro anticipado es global (no figura por camión).
    v4.49: agrega columna 'CtaCteDetalle' con lista de dicts
           [{nombre, codigo, monto}] por camión (para desglose UI/PDF/Excel).
    Retorna DataFrame con columnas:
      idCns, dsCns, TotVal, CtaCte, VtaEsp, Rechazos, NetoIngresar, CtaCteDetalle
    """
    rows = []
    cta_cte_codigos = {int(c["codigo"]) for c in cta_cte_list}
    nombre_map = {int(c["codigo"]): c["nombre"] for c in cta_cte_list}

    for _, sr_row in df_sr.iterrows():
        cns     = int(sr_row["idCns"])
        nombre  = str(sr_row["dsCns"])
        tot_val = float(sr_row["TotVal"])

        anr_cam = df_anr[df_anr["TRANSPORTE"] == cns] if not df_anr.empty else pd.DataFrame()

        # Detalle cliente × monto CTA CTE para este camión
        cta_detail = []
        if not anr_cam.empty:
            cta_rows = anr_cam[anr_cam["CLIENTE"].isin(cta_cte_codigos)]
            if not cta_rows.empty:
                agg = cta_rows.groupby("CLIENTE")["IMPORTE_NETO"].sum()
                for cod, monto in agg.items():
                    if monto > 0:
                        cta_detail.append({
                            "codigo": int(cod),
                            "nombre": nombre_map.get(int(cod), str(int(cod))),
                            "monto":  float(monto),
                        })

        cta_cte_monto = sum(d["monto"] for d in cta_detail)
        rechazos = float(anr_cam["IMPORTE_RECHAZADO"].sum()) if not anr_cam.empty else 0.0
        vta_esp = float(venta_especial_map.get(cns, 0.0))

        # Neto = TotVal - CtaCte + VtaEsp - Rechazos
        neto = tot_val - cta_cte_monto + vta_esp - rechazos

        rows.append({
            "idCns":         cns,
            "dsCns":         nombre,
            "TotVal":        tot_val,
            "CtaCte":        cta_cte_monto,
            "VtaEsp":        vta_esp,
            "Rechazos":      rechazos,
            "NetoIngresar":  neto,
            "CtaCteDetalle": cta_detail,
        })

    return pd.DataFrame(rows)


def _cierre_pdf(df_main: pd.DataFrame, totales: dict,
                mutual_monto: float, fecha_str: str,
                cobro_anticipado: float) -> bytes:
    """
    Genera PDF del Cierre para gerencia — LANDSCAPE A4 (v4.24).
    Mejor aprovechamiento del ancho: columnas de importes más legibles.
    Sin 'Cobro por otros medios'.
    """
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas

    buf = io.BytesIO()
    # Landscape: ancho = 297mm, alto = 210mm
    PW, PH = rl_landscape(A4)
    M = 14 * mm
    c = rl_canvas.Canvas(buf, pagesize=rl_landscape(A4))

    # ── Paleta ──────────────────────────────────────────────────────────────
    DARK_BLUE  = rl_colors.HexColor("#1a3a6b")
    MED_BLUE   = rl_colors.HexColor("#2e5fa3")
    LIGHT_GRAY = rl_colors.HexColor("#F0F4FA")
    ROW_ALT    = rl_colors.HexColor("#E8EEF8")
    GREEN_OK   = rl_colors.HexColor("#1a7a4a")
    RED_NEG    = rl_colors.HexColor("#b91c1c")
    AMBER_TXT  = rl_colors.HexColor("#92400e")
    GOLD_BG    = rl_colors.HexColor("#d97706")

    def fmt_ars(v):
        s = f"{abs(v):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"$ {s}" if v >= 0 else f"($ {s})"

    def fmt_pct(v):
        return f"{v:.1f}%"

    def new_page():
        c.showPage()
        # Mini-header en páginas siguientes
        c.setFillColor(DARK_BLUE)
        c.rect(M, PH - 16, PW - 2*M, 16, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(M + 4, PH - 11, f"CIERRE FINANCIERO — {fecha_str} (cont.)")
        c.setFont("Helvetica", 7)
        c.drawRightString(PW - M - 4, PH - 11, "Beccacece Hnos SA")
        return PH - 24

    # ── HEADER ───────────────────────────────────────────────────────────────
    y = PH - M
    c.setFillColor(DARK_BLUE)
    c.rect(M, y - 32, PW - 2*M, 34, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(M + 10, y - 20, f"CIERRE FINANCIERO — {fecha_str}")
    c.setFont("Helvetica", 9)
    c.drawRightString(PW - M - 8, y - 12, f"Beccacece Hnos SA  |  Reparto: {fecha_str}")
    c.drawRightString(PW - M - 8, y - 23, "Distribuidor CMQ — ABInBev")
    y -= 46

    # ── TABLA PRINCIPAL ───────────────────────────────────────────────────────
    # Landscape: 297mm disponible → columnas más anchas
    usable_w = PW - 2 * M
    col_labels = ["Camion",      "Total Chess",  "CTA CTE",   "Vta Especial", "Rechazos",   "Neto a Ingresar"]
    col_w      = [
        usable_w * 0.28,   # Camión (ancho — nombres largos)
        usable_w * 0.155,  # Total Chess
        usable_w * 0.135,  # CTA CTE
        usable_w * 0.135,  # Vta Especial
        usable_w * 0.135,  # Rechazos
        usable_w * 0.155,  # Neto a Ingresar
    ]
    row_h = 13
    hdr_h = 16

    # Header tabla
    c.setFillColor(MED_BLUE)
    c.rect(M, y - hdr_h, usable_w, hdr_h, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 9)
    x_cur = M + 5
    for lbl, cw in zip(col_labels, col_w):
        c.drawString(x_cur, y - hdr_h + 5, lbl)
        x_cur += cw
    y -= hdr_h + 2

    SUB_H    = 11   # altura sub-fila CTA CTE
    SUB_FILL = rl_colors.HexColor("#FFF3F3")   # fondo rosado suave
    INDENT   = 16   # sangría izquierda sub-fila

    for i, row in df_main.iterrows():
        bg = LIGHT_GRAY if i % 2 == 0 else rl_colors.white
        c.setFillColor(bg)
        c.rect(M, y - row_h, usable_w, row_h, fill=1, stroke=0)

        vals = [
            f"{int(row['idCns'])} — {row['dsCns']}",
            fmt_ars(row["TotVal"]),
            fmt_ars(row["CtaCte"])  if row["CtaCte"]  > 0 else "—",
            fmt_ars(row["VtaEsp"])  if row["VtaEsp"]  > 0 else "—",
            fmt_ars(row["Rechazos"]) if row["Rechazos"] > 0 else "—",
            fmt_ars(row["NetoIngresar"]),
        ]
        x_cur = M + 5
        for vi, (val, cw) in enumerate(zip(vals, col_w)):
            if vi == 5:
                c.setFillColor(GREEN_OK if row["NetoIngresar"] >= 0 else RED_NEG)
                c.setFont("Helvetica-Bold", 8.5)
            elif vi in (2, 3, 4) and val != "—":
                c.setFillColor(RED_NEG)
                c.setFont("Helvetica", 8.5)
            else:
                c.setFillColor(rl_colors.black)
                c.setFont("Helvetica", 8.5)
            c.drawString(x_cur, y - row_h + 4, val)
            x_cur += cw
        y -= row_h

        # Sub-filas CTA CTE integradas
        detalle_cte = row.get("CtaCteDetalle") or []
        if isinstance(detalle_cte, list) and detalle_cte:
            for d in detalle_cte:
                if y < M + 60:
                    y = new_page()
                c.setFillColor(SUB_FILL)
                c.rect(M, y - SUB_H, usable_w, SUB_H, fill=1, stroke=0)
                # Línea izquierda roja como indicador
                c.setFillColor(RED_NEG)
                c.rect(M, y - SUB_H, 3, SUB_H, fill=1, stroke=0)
                # Texto: "  ↳ Cód · Nombre"
                c.setFillColor(RED_NEG)
                c.setFont("Helvetica-Bold", 7.5)
                nombre_cte = d.get("nombre", str(d.get("codigo", "")))
                cod_cte    = d.get("codigo", "")
                c.drawString(M + INDENT, y - SUB_H + 3,
                             f"\u21b3 {cod_cte}  {nombre_cte}")
                # Monto alineado a la columna CTA CTE
                x_cta = M + col_w[0] + col_w[1]
                c.drawString(x_cta + 3, y - SUB_H + 3,
                             fmt_ars(float(d.get("monto", 0))))
                y -= SUB_H

        if y < M + 100:
            y = new_page()

    # ── FILA TOTAL ────────────────────────────────────────────────────────────
    y -= 4
    c.setFillColor(DARK_BLUE)
    c.rect(M, y - 16, usable_w, 16, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 9)
    tot_vals = [
        "TOTAL",
        fmt_ars(totales["tot_chess"]),
        fmt_ars(totales["tot_cta_cte"]) if totales["tot_cta_cte"] > 0 else "—",
        fmt_ars(totales["tot_vta_esp"]) if totales["tot_vta_esp"] > 0 else "—",
        fmt_ars(totales["tot_rechazos"]) if totales["tot_rechazos"] > 0 else "—",
        fmt_ars(totales["tot_neto"]),
    ]
    x_cur = M + 5
    for tv, cw in zip(tot_vals, col_w):
        c.drawString(x_cur, y - 12, tv)
        x_cur += cw
    y -= 24

    # ── BLOQUE RESUMEN GLOBAL — dos columnas en landscape ─────────────────────
    pct = (totales["tot_rechazos"] / totales["tot_chess"] * 100) if totales["tot_chess"] > 0 else 0

    resumen_h = 115
    c.setFillColor(rl_colors.HexColor("#F0F4FA"))
    c.rect(M, y - resumen_h, usable_w, resumen_h + 4, fill=1, stroke=0)

    c.setFillColor(DARK_BLUE)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(M + 8, y - 14, "RESUMEN DEL DIA")

    resumen_items = [
        ("D) Total Preventa del dia (Mercaderia que sale a reparto)", totales["tot_chess"],       DARK_BLUE),
        ("E) CTA CTE (total a descontar)",                           totales["tot_cta_cte"],     RED_NEG if totales["tot_cta_cte"] > 0 else rl_colors.black),
        ("F) Venta especial",                                        totales["tot_vta_esp"],      rl_colors.black),
        ("F) Total neto a ingresar repartos (D - E + F)",            totales["tot_neto_global"],  GREEN_OK),
        ("G) Rechazos",                                              totales["tot_rechazos"],      RED_NEG if totales["tot_rechazos"] > 0 else rl_colors.black),
        ("H) Total del dia neto rechazos",                           totales["tot_dia_neto"],     GREEN_OK),
    ]

    # Dividir en dos columnas
    left_items  = resumen_items[:4]
    right_items = resumen_items[4:]
    half_w = usable_w / 2 - 10

    yi_l = y - 30
    for label, val, col_txt in left_items:
        c.setFillColor(rl_colors.black)
        c.setFont("Helvetica", 8.5)
        c.drawString(M + 8, yi_l, label)
        c.setFillColor(col_txt)
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(M + half_w, yi_l, fmt_ars(val))
        yi_l -= 13

    yi_r = y - 30
    for label, val, col_txt in right_items:
        c.setFillColor(rl_colors.black)
        c.setFont("Helvetica", 8.5)
        c.drawString(M + half_w + 14, yi_r, label)
        c.setFillColor(col_txt)
        c.setFont("Helvetica-Bold", 9)
        c.drawRightString(PW - M - 8, yi_r, fmt_ars(val))
        yi_r -= 13

    # % Rechazo (debajo columna derecha)
    yi_r -= 2
    c.setFillColor(rl_colors.black)
    c.setFont("Helvetica", 8.5)
    c.drawString(M + half_w + 14, yi_r, "% Rechazo")
    c.setFillColor(RED_NEG if pct > 5 else GREEN_OK)
    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(PW - M - 8, yi_r, fmt_pct(pct))

    # ── MUTUAL ────────────────────────────────────────────────────────────────
    if mutual_monto > 0:
        y_mut = y - resumen_h - 6
        c.setFillColor(rl_colors.HexColor("#FEF3C7"))
        c.rect(M, y_mut - 16, usable_w, 18, fill=1, stroke=0)
        c.setFillColor(AMBER_TXT)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(M + 6, y_mut - 11,
                     f"MUTUAL ({_MUTUAL_CODIGO}) — Monitoreo especial (no es CTA CTE):")
        c.setFont("Helvetica-Bold", 9.5)
        c.drawRightString(PW - M - 6, y_mut - 11, fmt_ars(mutual_monto))

    # ── FOOTER ────────────────────────────────────────────────────────────────
    c.setFillColor(rl_colors.HexColor("#555555"))
    c.setFont("Helvetica", 6.5)
    c.drawCentredString(
        PW / 2, M - 4,
        f"Beccacece Hnos SA  ·  Cierre generado: {fecha_str}  ·  Picking Orchestrator v{APP_VERSION}"
    )

    # ── PÁGINA 2: DETALLE CTA CTE POR CAMIÓN ─────────────────────────────────
    # Armar lista de camiones con detalle CTA CTE
    cta_rows_pdf = []
    if "CtaCteDetalle" in df_main.columns:
        for _, row in df_main.iterrows():
            detalle = row.get("CtaCteDetalle") or []
            if isinstance(detalle, list) and detalle:
                for d in detalle:
                    cta_rows_pdf.append({
                        "camion": f"{int(row['idCns'])} — {row['dsCns']}",
                        "nombre": d["nombre"],
                        "codigo": d["codigo"],
                        "monto":  d["monto"],
                    })

    if cta_rows_pdf:
        c.showPage()
        y2 = PH - M

        # Header página 2
        c.setFillColor(DARK_BLUE)
        c.rect(M, y2 - 32, PW - 2*M, 34, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 13)
        c.drawString(M + 10, y2 - 20, f"DETALLE CTA CTE — {fecha_str}")
        c.setFont("Helvetica", 8)
        c.drawRightString(PW - M - 8, y2 - 20, "Beccacece Hnos SA  |  Desglose cliente × camión")
        y2 -= 46

        # Tabla detalle
        usable_w2 = PW - 2 * M
        col_labels2 = ["Camión",        "Cód.",    "Cliente (CTA CTE)",            "Monto ($)"]
        col_w2      = [usable_w2 * 0.30, usable_w2 * 0.07, usable_w2 * 0.42, usable_w2 * 0.21]
        row_h2 = 13
        hdr_h2 = 15

        c.setFillColor(MED_BLUE)
        c.rect(M, y2 - hdr_h2, usable_w2, hdr_h2, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 8.5)
        x2 = M + 5
        for lbl, cw in zip(col_labels2, col_w2):
            c.drawString(x2, y2 - hdr_h2 + 4, lbl)
            x2 += cw
        y2 -= hdr_h2 + 2

        prev_cam = None
        cam_total = 0.0
        grand_total = 0.0
        alt = False

        for ii, dr in enumerate(cta_rows_pdf):
            # Separador de camión
            if dr["camion"] != prev_cam:
                if prev_cam is not None:
                    # fila subtotal camión anterior
                    c.setFillColor(rl_colors.HexColor("#D9E4F5"))
                    c.rect(M, y2 - row_h2, usable_w2, row_h2, fill=1, stroke=0)
                    c.setFillColor(DARK_BLUE)
                    c.setFont("Helvetica-Bold", 8)
                    sub_x = M + 5 + col_w2[0] + col_w2[1]
                    c.drawString(sub_x, y2 - row_h2 + 4, "SUBTOTAL")
                    c.drawRightString(M + usable_w2 - 4, y2 - row_h2 + 4, fmt_ars(cam_total))
                    y2 -= row_h2 + 2
                    cam_total = 0.0
                    if y2 < M + 40:
                        y2 = new_page()
                prev_cam = dr["camion"]
                alt = False

            bg2 = rl_colors.HexColor("#F5F8FF") if alt else rl_colors.white
            c.setFillColor(bg2)
            c.rect(M, y2 - row_h2, usable_w2, row_h2, fill=1, stroke=0)

            vals2 = [dr["camion"], str(dr["codigo"]), dr["nombre"], fmt_ars(dr["monto"])]
            x2 = M + 5
            for vi2, (val2, cw2) in enumerate(zip(vals2, col_w2)):
                if vi2 == 3:
                    c.setFillColor(RED_NEG)
                    c.setFont("Helvetica-Bold", 8)
                    c.drawRightString(M + usable_w2 - 4, y2 - row_h2 + 4, val2)
                else:
                    c.setFillColor(rl_colors.black)
                    c.setFont("Helvetica", 8)
                    c.drawString(x2, y2 - row_h2 + 4, val2)
                x2 += cw2

            cam_total  += dr["monto"]
            grand_total += dr["monto"]
            alt = not alt
            y2 -= row_h2

            if y2 < M + 40:
                y2 = new_page()

        # Subtotal último camión
        if prev_cam is not None:
            c.setFillColor(rl_colors.HexColor("#D9E4F5"))
            c.rect(M, y2 - row_h2, usable_w2, row_h2, fill=1, stroke=0)
            c.setFillColor(DARK_BLUE)
            c.setFont("Helvetica-Bold", 8)
            sub_x = M + 5 + col_w2[0] + col_w2[1]
            c.drawString(sub_x, y2 - row_h2 + 4, "SUBTOTAL")
            c.drawRightString(M + usable_w2 - 4, y2 - row_h2 + 4, fmt_ars(cam_total))
            y2 -= row_h2 + 4

        # Fila TOTAL GENERAL CTA CTE
        c.setFillColor(DARK_BLUE)
        c.rect(M, y2 - 16, usable_w2, 16, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(M + 6, y2 - 11, "TOTAL GENERAL CTA CTE")
        c.drawRightString(M + usable_w2 - 4, y2 - 11, fmt_ars(grand_total))

        # Footer pág 2
        c.setFillColor(rl_colors.HexColor("#555555"))
        c.setFont("Helvetica", 6.5)
        c.drawCentredString(
            PW / 2, M - 4,
            f"Beccacece Hnos SA  ·  Cierre generado: {fecha_str}  ·  Picking Orchestrator v{APP_VERSION}"
        )

    c.save()
    return buf.getvalue()


def _cierre_excel(df_main: pd.DataFrame, totales: dict,
                  mutual_monto: float, fecha_str: str,
                  cobro_anticipado: float) -> bytes:
    """
    Genera Excel del Cierre con openpyxl.
    v4.24: sin columna 'Cobro por otros medios', 6 columnas (A:F).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cierre"

    hdr_fill   = PatternFill("solid", fgColor="1a3a6b")
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    tot_fill   = PatternFill("solid", fgColor="2e5fa3")
    tot_font   = Font(color="FFFFFF", bold=True, size=10)
    alt_fill   = PatternFill("solid", fgColor="EEF3FB")
    green_font = Font(color="1a7a4a", bold=True)
    red_font   = Font(color="b91c1c", bold=True)
    bold_font  = Font(bold=True)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_al  = Alignment(horizontal="center", vertical="center")
    right_al   = Alignment(horizontal="right", vertical="center")
    left_al    = Alignment(horizontal="left", vertical="center")
    money_fmt  = '#,##0'

    # ── TÍTULO (6 cols: A:F) ─────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"CIERRE FINANCIERO DEL DÍA — {fecha_str}"
    ws["A1"].font = Font(bold=True, size=13, color="1a3a6b")
    ws["A1"].alignment = center_al
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws["A2"] = "Beccacece Hnos SA — Distribuidor CMQ / ABInBev"
    ws["A2"].font = Font(size=9, color="555555", italic=True)
    ws["A2"].alignment = center_al

    # ── HEADER TABLA ─────────────────────────────────────────────────────────
    headers = ["ID", "Camión", "Total Chess ($)", "CTA CTE ($)", "Rechazos ($)", "Neto a Ingresar ($)"]
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center_al
        cell.border = border
    ws.row_dimensions[4].height = 16

    sub_fill   = PatternFill("solid", fgColor="FFF0F0")   # rosado suave
    red_left   = Font(color="b91c1c", size=8, italic=True)

    # ── FILAS ─────────────────────────────────────────────────────────────────
    r = 5   # fila inicial en hoja
    for ri, row in df_main.iterrows():
        vals = [
            int(row["idCns"]),
            row["dsCns"],
            row["TotVal"],
            row["CtaCte"]   if row["CtaCte"]   > 0 else 0,
            row["Rechazos"] if row["Rechazos"] > 0 else 0,
            row["NetoIngresar"],
        ]
        fill_row = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = border
            cell.fill = fill_row
            if ci == 1:
                cell.alignment = center_al
            elif ci == 2:
                cell.alignment = left_al
                cell.font = bold_font
            else:
                cell.alignment = right_al
                cell.number_format = money_fmt
                if ci == 6:
                    cell.font = green_font if row["NetoIngresar"] >= 0 else red_font
                elif ci == 4 and row["CtaCte"] > 0:
                    cell.font = red_font
        ws.row_dimensions[r].height = 14
        r += 1

        # Sub-filas CTA CTE integradas
        detalle_cte = row.get("CtaCteDetalle") or []
        if isinstance(detalle_cte, list):
            for d in detalle_cte:
                monto_d = float(d.get("monto", 0))
                if monto_d <= 0:
                    continue
                # col A: vacío, col B: "  ↳ nombre", col C: vacío, col D: monto, E-F: vacío
                ws.cell(row=r, column=1, value="").fill = sub_fill
                cell_nom = ws.cell(row=r, column=2,
                                   value=f"\u21b3 {d.get('codigo','')}  {d.get('nombre','')}")
                cell_nom.font  = red_left
                cell_nom.fill  = sub_fill
                cell_nom.alignment = left_al
                for ci in range(3, 7):
                    ws.cell(row=r, column=ci).fill = sub_fill
                cell_m = ws.cell(row=r, column=4, value=monto_d)
                cell_m.font          = Font(color="b91c1c", size=8, bold=True)
                cell_m.fill          = sub_fill
                cell_m.alignment     = right_al
                cell_m.number_format = money_fmt
                ws.row_dimensions[r].height = 12
                r += 1

    # ── FILA TOTAL ────────────────────────────────────────────────────────────
    tr = r   # r ya apunta a la siguiente fila libre
    tot_vals = ["", "TOTAL",
                totales["tot_chess"], totales["tot_cta_cte"],
                totales["tot_rechazos"], totales["tot_neto"]]
    for ci, val in enumerate(tot_vals, 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.fill = tot_fill
        cell.font = tot_font
        cell.border = border
        cell.alignment = right_al if ci > 2 else center_al
        if ci > 2:
            cell.number_format = money_fmt
    ws.row_dimensions[tr].height = 16

    # ── RESUMEN GLOBAL ────────────────────────────────────────────────────────
    rs = tr + 2
    pct = (totales["tot_rechazos"] / totales["tot_chess"] * 100) if totales["tot_chess"] > 0 else 0
    resumen_rows = [
        ("D) Total Preventa del día",                              totales["tot_chess"],       None),
        ("E) CTA CTE (total a descontar)",                         totales["tot_cta_cte"],     "b91c1c" if totales["tot_cta_cte"] > 0 else None),
        ("F) Venta especial",                                      totales["tot_vta_esp"],     None),
        ("F) Total neto a ingresar repartos (D - E + F)",          totales["tot_neto_global"], "1a7a4a"),
        ("G) Rechazos",                                            totales["tot_rechazos"],    "b91c1c" if totales["tot_rechazos"] > 0 else None),
        ("H) Total del día neto rechazos",                         totales["tot_dia_neto"],    "1a7a4a"),
        ("% Rechazo",                                              pct / 100,                  "b91c1c" if pct > 5 else "1a7a4a"),
    ]

    ws.merge_cells(f"A{rs}:F{rs}")
    ws[f"A{rs}"] = "RESUMEN DEL DÍA"
    ws[f"A{rs}"].font = Font(bold=True, size=10, color="1a3a6b")
    ws[f"A{rs}"].fill = PatternFill("solid", fgColor="D9E4F5")
    ws[f"A{rs}"].alignment = center_al
    rs += 1

    for label, val, color in resumen_rows:
        ws.merge_cells(f"A{rs}:D{rs}")
        ws[f"A{rs}"] = label
        ws[f"A{rs}"].font = Font(size=9)
        ws[f"A{rs}"].alignment = left_al
        ws.merge_cells(f"E{rs}:F{rs}")
        cell_val = ws[f"E{rs}"]
        cell_val.value = val
        cell_val.alignment = right_al
        if label.startswith("%"):
            cell_val.number_format = "0.0%"
        else:
            cell_val.number_format = money_fmt
        if color:
            cell_val.font = Font(bold=True, color=color, size=9)
        else:
            cell_val.font = Font(bold=True, size=9)
        ws.row_dimensions[rs].height = 13
        rs += 1

    # ── MUTUAL ────────────────────────────────────────────────────────────────
    if mutual_monto > 0:
        rs += 1
        ws.merge_cells(f"A{rs}:D{rs}")
        ws[f"A{rs}"] = f"⚠ MUTUAL ({_MUTUAL_CODIGO}) — Monitoreo especial"
        ws[f"A{rs}"].font = Font(bold=True, color="92400e", size=9)
        ws[f"A{rs}"].fill = PatternFill("solid", fgColor="FEF3C7")
        ws[f"A{rs}"].alignment = left_al
        ws.merge_cells(f"E{rs}:F{rs}")
        ws[f"E{rs}"] = mutual_monto
        ws[f"E{rs}"].number_format = money_fmt
        ws[f"E{rs}"].font = Font(bold=True, color="92400e")
        ws[f"E{rs}"].fill = PatternFill("solid", fgColor="FEF3C7")
        ws[f"E{rs}"].alignment = right_al

    col_widths = [8, 24, 22, 18, 18, 26]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── HOJA 2: CTA CTE DETALLE ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title="CTA CTE Detalle")

    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"DETALLE CTA CTE — {fecha_str}"
    ws2["A1"].font = Font(bold=True, size=12, color="1a3a6b")
    ws2["A1"].alignment = center_al

    ws2.merge_cells("A2:E2")
    ws2["A2"] = "Beccacece Hnos SA — Desglose por camión y cliente"
    ws2["A2"].font = Font(size=9, color="555555", italic=True)
    ws2["A2"].alignment = center_al

    hdr2 = ["ID Camión", "Camión", "Cód. Cliente", "Cliente (CTA CTE)", "Monto ($)"]
    for ci, h in enumerate(hdr2, 1):
        cell = ws2.cell(row=4, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center_al
        cell.border = border
    ws2.row_dimensions[4].height = 15

    r2 = 5
    grand_total_cta = 0.0
    if "CtaCteDetalle" in df_main.columns:
        for _, mrow in df_main.iterrows():
            detalle = mrow.get("CtaCteDetalle") or []
            if not isinstance(detalle, list) or not detalle:
                continue
            cam_total2 = 0.0
            for di, d in enumerate(detalle):
                bg_r = alt_fill if di % 2 == 0 else PatternFill()
                vals2 = [int(mrow["idCns"]), mrow["dsCns"], d["codigo"], d["nombre"], d["monto"]]
                for ci, val in enumerate(vals2, 1):
                    cell = ws2.cell(row=r2, column=ci, value=val)
                    cell.border = border
                    cell.fill = bg_r
                    if ci == 5:
                        cell.alignment = right_al
                        cell.number_format = money_fmt
                        cell.font = Font(color="b91c1c", bold=True)
                    elif ci in (1, 3):
                        cell.alignment = center_al
                    else:
                        cell.alignment = left_al
                ws2.row_dimensions[r2].height = 13
                cam_total2  += d["monto"]
                grand_total_cta += d["monto"]
                r2 += 1

            # Subtotal por camión
            sub_fill = PatternFill("solid", fgColor="D9E4F5")
            ws2.merge_cells(f"A{r2}:D{r2}")
            ws2[f"A{r2}"] = f"Subtotal — {mrow['dsCns']}"
            ws2[f"A{r2}"].font = Font(bold=True, color="1a3a6b", size=9)
            ws2[f"A{r2}"].fill = sub_fill
            ws2[f"A{r2}"].alignment = left_al
            ws2[f"A{r2}"].border = border
            ws2[f"E{r2}"] = cam_total2
            ws2[f"E{r2}"].number_format = money_fmt
            ws2[f"E{r2}"].font = Font(bold=True, color="1a3a6b")
            ws2[f"E{r2}"].fill = sub_fill
            ws2[f"E{r2}"].alignment = right_al
            ws2[f"E{r2}"].border = border
            ws2.row_dimensions[r2].height = 13
            r2 += 1

    # Total general
    r2 += 1
    ws2.merge_cells(f"A{r2}:D{r2}")
    ws2[f"A{r2}"] = "TOTAL GENERAL CTA CTE"
    ws2[f"A{r2}"].font = Font(bold=True, size=10, color="FFFFFF")
    ws2[f"A{r2}"].fill = hdr_fill
    ws2[f"A{r2}"].alignment = left_al
    ws2[f"A{r2}"].border = border
    ws2[f"E{r2}"] = grand_total_cta
    ws2[f"E{r2}"].number_format = money_fmt
    ws2[f"E{r2}"].font = Font(bold=True, color="FFFFFF", size=10)
    ws2[f"E{r2}"].fill = hdr_fill
    ws2[f"E{r2}"].alignment = right_al
    ws2[f"E{r2}"].border = border
    ws2.row_dimensions[r2].height = 16

    # Anchos hoja 2
    for ci, w in zip(range(1, 6), [10, 26, 12, 38, 22]):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_tab_cierre():
    """
    💰 Cierre Financiero del Día — v4.24.
    Fuente: SR.xlsx (col A TotVal Chess) + ANR.xlsx (detalle cliente/camión).
    Sin 'Cobro por otros medios'. PDF en landscape.
    """
    st.subheader("💰 Cierre Financiero del Día")
    st.caption(
        "Dashboard para gerencia. Fuente: **SR.xlsx** (TotVal Chess por camión) + "
        "**ANR.xlsx** (CTA CTE y rechazos). PDF exportado en **A4 horizontal**."
    )

    sr_file  = st.session_state.get("cierre_sr")
    anr_file = st.session_state.get("tc_anr")

    if not sr_file:
        st.warning("⬅️ Subí el **SR.xlsx** en la pestaña **📁 Archivos** para generar el Cierre.")
        return
    if not anr_file:
        st.info("ℹ️ Sin ANR.xlsx los descuentos CTA CTE y rechazos serán 0. "
                "Subilo en **📁 Archivos** para cálculo completo.")

    # ── Inicializar CTA CTE en session_state ─────────────────────────────────
    if "cierre_cta_cte" not in st.session_state:
        st.session_state["cierre_cta_cte"] = list(_DEFAULT_CTA_CTE)

    # ── Cargar SR ─────────────────────────────────────────────────────────────
    try:
        df_sr = _cierre_load_sr(sr_file)
    except Exception as e:
        st.error(f"❌ Error leyendo SR.xlsx: {e}")
        return

    # ── Cargar ANR ────────────────────────────────────────────────────────────
    df_anr = pd.DataFrame()
    if anr_file:
        try:
            df_anr = _cierre_load_anr(anr_file)
        except Exception as e:
            st.warning(f"⚠️ Error leyendo ANR.xlsx: {e}. Continuando sin detalle.")

    # ── v4.39: Fecha del día de REPARTO (desde ANR o CAR) — dd/mm/yyyy ────────
    # Prioridad: col FECHA del ANR → col 'Fecha Mvto' del CAR → hoy
    _fecha_reparto = None
    try:
        # Intentar leer fecha del ANR (hoja BASE, col que contenga 'FECHA')
        if anr_file:
            anr_file.seek(0)
            _xl_anr = pd.ExcelFile(anr_file)
            _sh = next((s for s in _xl_anr.sheet_names if s.upper() == "BASE"), _xl_anr.sheet_names[0])
            anr_file.seek(0)
            _df_anr_raw = pd.read_excel(anr_file, sheet_name=_sh, header=1)
            _col_fecha_anr = next((c for c in _df_anr_raw.columns
                                   if "FECHA" in str(c).upper() and "VENC" not in str(c).upper()), None)
            if _col_fecha_anr:
                _vals_f = pd.to_datetime(_df_anr_raw[_col_fecha_anr], errors="coerce").dropna()
                if not _vals_f.empty:
                    _fecha_reparto = _vals_f.iloc[0].date()
            anr_file.seek(0)
    except Exception:
        pass

    if _fecha_reparto is None:
        # Fallback: col 'Fecha Mvto' del CAR si disponible
        try:
            _car_f = st.session_state.get("t1_car") or st.session_state.get("t4_car")
            if _car_f:
                _car_f.seek(0)
                _df_car_f = pd.read_excel(_car_f, header=0)
                _col_fm = next((c for c in _df_car_f.columns if "fecha" in str(c).lower()), None)
                if _col_fm:
                    _vals_car_f = pd.to_datetime(_df_car_f[_col_fm], errors="coerce").dropna()
                    if not _vals_car_f.empty:
                        _fecha_reparto = _vals_car_f.iloc[0].date()
                _car_f.seek(0)
        except Exception:
            pass

    if _fecha_reparto is None:
        import datetime as _dt_cierre
        _fecha_reparto = _dt_cierre.date.today()

    # fecha_str para mostrar en UI y PDF header
    fecha_str      = _fecha_reparto.strftime("%d/%m/%Y")
    # slug para nombre de archivos: DD-MM-YYYY
    _fecha_slug    = _fecha_reparto.strftime("%d-%m-%Y")

    # ═══════════════════════════════════════════════════════════════════════════
    # PANEL LATERAL: Clientes CTA CTE + controles
    # ═══════════════════════════════════════════════════════════════════════════
    with st.expander("⚙️ Configuración — Clientes CTA CTE", expanded=False):
        st.markdown("##### Lista de clientes con Cuenta Corriente")
        st.caption("Estos clientes se descuentan del total de cada camión que los lleve.")

        cta_list = st.session_state["cierre_cta_cte"]

        # Tabla editable
        if cta_list:
            df_edit = pd.DataFrame(cta_list)
            edited = st.data_editor(
                df_edit,
                column_config={
                    "codigo": st.column_config.NumberColumn("Código", min_value=1, step=1),
                    "nombre": st.column_config.TextColumn("Nombre / Razón Social"),
                },
                num_rows="dynamic",
                use_container_width=True,
                key="cta_cte_editor",
            )
            # Persistir cambios
            st.session_state["cierre_cta_cte"] = edited.to_dict("records")
        else:
            st.info("Lista vacía. Agregá clientes con el botón de abajo.")
            if st.button("➕ Cargar lista predeterminada"):
                st.session_state["cierre_cta_cte"] = list(_DEFAULT_CTA_CTE)
                st.rerun()

        col_r1, col_r2 = st.columns(2)
        with col_r1:
            if st.button("🔄 Restaurar lista predeterminada", help="Vuelve a la lista original"):
                st.session_state["cierre_cta_cte"] = list(_DEFAULT_CTA_CTE)
                st.rerun()
        with col_r2:
            if st.button("🗑️ Limpiar lista"):
                st.session_state["cierre_cta_cte"] = []
                st.rerun()

        st.divider()
        st.markdown(f"**⚠️ MUTUAL ({_MUTUAL_CODIGO})** — Cliente de monitoreo especial (no es CTA CTE). "
                    "Su monto se muestra separado en el resumen.")

    # ── Controles adicionales ─────────────────────────────────────────────────
    col_ctrl1, col_ctrl2 = st.columns(2)
    with col_ctrl1:
        st.markdown("**Venta Especial por camión** — ingresá en la tabla de abajo si aplica.")
    with col_ctrl2:
        if st.button("🔄 Recalcular", type="primary"):
            st.rerun()
    cobro_anticipado = 0.0

    # Venta especial por camión (opcional)
    with st.expander("📝 Venta Especial por camión (opcional)", expanded=False):
        vta_esp_data = []
        for _, row in df_sr.iterrows():
            vta_esp_data.append({"idCns": int(row["idCns"]), "dsCns": row["dsCns"], "VtaEsp": 0.0})
        df_vta_esp = pd.DataFrame(vta_esp_data)
        edited_vta = st.data_editor(
            df_vta_esp,
            column_config={
                "idCns":  st.column_config.NumberColumn("ID", disabled=True),
                "dsCns":  st.column_config.TextColumn("Camión", disabled=True),
                "VtaEsp": st.column_config.NumberColumn("Venta Especial ($)", min_value=0.0, step=100.0, format="%.0f"),
            },
            use_container_width=True,
            key="vta_esp_editor",
        )
        venta_especial_map = {int(r["idCns"]): float(r["VtaEsp"]) for _, r in edited_vta.iterrows()}

    st.divider()

    # ═══════════════════════════════════════════════════════════════════════════
    # CONSTRUCCIÓN DEL DASHBOARD
    # ═══════════════════════════════════════════════════════════════════════════
    cta_cte_list = st.session_state["cierre_cta_cte"]

    cta_cte_list_clean = [
        c for c in cta_cte_list
        if c.get("codigo") and str(c.get("codigo")).strip() not in ("", "nan")
    ]

    df_main = _build_cierre_df(df_sr, df_anr, cta_cte_list_clean,
                                cobro_anticipado, venta_especial_map if "venta_especial_map" in dir() else {})

    # MUTUAL
    mutual_monto = 0.0
    if not df_anr.empty:
        mutual_monto = float(df_anr[df_anr["CLIENTE"] == _MUTUAL_CODIGO]["IMPORTE_NETO"].sum())

    # Totales
    tot_chess       = df_main["TotVal"].sum()
    tot_cta_cte     = df_main["CtaCte"].sum()
    tot_vta_esp     = df_main["VtaEsp"].sum()
    tot_rechazos    = df_main["Rechazos"].sum()
    tot_neto        = df_main["NetoIngresar"].sum()
    tot_neto_global = tot_chess - tot_cta_cte + tot_vta_esp
    tot_dia_neto    = tot_neto_global - tot_rechazos
    pct_rechazo     = (tot_rechazos / tot_chess * 100) if tot_chess > 0 else 0.0

    totales = {
        "tot_chess":       tot_chess,
        "tot_cta_cte":     tot_cta_cte,
        "tot_vta_esp":     tot_vta_esp,
        "tot_rechazos":    tot_rechazos,
        "tot_neto":        tot_neto,
        "tot_neto_global": tot_neto_global,
        "tot_dia_neto":    tot_dia_neto,
    }

    # ── KPIs ──────────────────────────────────────────────────────────────────
    def ars(v):
        s = f"$ {abs(v):,.0f}".replace(",", ".")
        return f"-{s}" if v < 0 else s

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("📦 Total Preventa",  ars(tot_chess))
    k2.metric("🔴 CTA CTE total",   ars(tot_cta_cte),
              delta=f"-{ars(tot_cta_cte)}" if tot_cta_cte > 0 else None, delta_color="inverse")
    k3.metric("✅ Neto Repartos",   ars(tot_neto_global))
    k4.metric("❌ Rechazos",        ars(tot_rechazos),
              delta=f"{pct_rechazo:.1f}%" if tot_rechazos > 0 else None, delta_color="inverse")
    k5.metric("💰 Total Día Neto",  ars(tot_dia_neto))

    if mutual_monto > 0:
        st.warning(
            f"⚠️ **MUTUAL ({_MUTUAL_CODIGO})** — Monto en reparto: **{ars(mutual_monto)}** "
            "(monitoreo especial, no es CTA CTE)"
        )

    st.divider()

    # ── TABLA PRINCIPAL ───────────────────────────────────────────────────────
    st.markdown("#### 🚛 Detalle por Camión")

    def style_neto(val):
        color = "#1a7a4a" if val >= 0 else "#b91c1c"
        return f"color: {color}; font-weight: bold"

    def style_cta(val):
        return "color: #b91c1c; font-weight: bold" if val > 0 else ""

    def style_rechazo(val):
        return "color: #b91c1c;" if val > 0 else ""

    df_display = df_main.copy()
    df_display["Camión"]          = df_display.apply(lambda r: f"{int(r['idCns'])} — {r['dsCns']}", axis=1)
    df_display["Total Chess"]     = df_display["TotVal"]
    df_display["CTA CTE"]         = df_display["CtaCte"]
    df_display["Rechazos"]        = df_display["Rechazos"]
    df_display["Neto a Ingresar"] = df_display["NetoIngresar"]
    df_show = df_display[["Camión", "Total Chess", "CTA CTE", "Rechazos", "Neto a Ingresar"]]

    # pandas >= 2.1: applymap fue reemplazado por map
    _use_map = hasattr(pd.io.formats.style.Styler, "map")
    _styled_base = df_show.style.format({
        "Total Chess":     "$ {:,.0f}",
        "CTA CTE":         "$ {:,.0f}",
        "Rechazos":        "$ {:,.0f}",
        "Neto a Ingresar": "$ {:,.0f}",
    })
    if _use_map:
        styled = (_styled_base
            .map(style_neto,    subset=["Neto a Ingresar"])
            .map(style_cta,     subset=["CTA CTE"])
            .map(style_rechazo, subset=["Rechazos"])
            .set_properties(**{"text-align": "right"},
                            subset=["Total Chess", "CTA CTE", "Rechazos", "Neto a Ingresar"]))
    else:
        styled = (_styled_base
            .applymap(style_neto,    subset=["Neto a Ingresar"])
            .applymap(style_cta,     subset=["CTA CTE"])
            .applymap(style_rechazo, subset=["Rechazos"])
            .set_properties(**{"text-align": "right"},
                            subset=["Total Chess", "CTA CTE", "Rechazos", "Neto a Ingresar"]))

    st.dataframe(styled, use_container_width=True, hide_index=True, height=460)

    # ── RESUMEN GLOBAL ────────────────────────────────────────────────────────
    st.markdown("#### 📊 Resumen del Día")
    rc1, rc2 = st.columns([1.6, 1])

    with rc1:
        res_data = {
            "Concepto": [
                "D) Total Preventa del día",
                "E) CTA CTE (descuento)",
                "F) Venta especial",
                "F) Total neto a ingresar repartos",
                "G) Rechazos",
                "H) Total del día neto rechazos",
                "% Rechazo",
            ],
            "Monto": [
                f"$ {tot_chess:,.0f}",
                f"$ {tot_cta_cte:,.0f}",
                f"$ {tot_vta_esp:,.0f}",
                f"$ {tot_neto_global:,.0f}",
                f"$ {tot_rechazos:,.0f}",
                f"$ {tot_dia_neto:,.0f}",
                f"{pct_rechazo:.1f}%",
            ],
        }
        df_res = pd.DataFrame(res_data)
        st.dataframe(df_res, use_container_width=True, hide_index=True)

    with rc2:
        st.markdown("**CTA CTE detectados en el reparto de hoy:**")
        if cta_cte_list_clean and not df_anr.empty:
            cta_codigos   = {int(c["codigo"]) for c in cta_cte_list_clean}
            clientes_hoy  = df_anr[df_anr["CLIENTE"].isin(cta_codigos)].groupby("CLIENTE")["IMPORTE_NETO"].sum()
            nombre_map    = {int(c["codigo"]): c["nombre"] for c in cta_cte_list_clean}
            for cod, monto in clientes_hoy[clientes_hoy > 0].items():
                st.markdown(f"• **{nombre_map.get(int(cod), cod)}** — {ars(monto)}")
            if clientes_hoy[clientes_hoy > 0].empty:
                st.info("Ningún cliente CTA CTE tiene venta hoy.")
        else:
            st.info("Sin datos ANR o lista CTA CTE vacía.")

    # ── DETALLE CTA CTE POR CAMIÓN (expandible) ───────────────────────────────
    if "CtaCteDetalle" in df_main.columns and tot_cta_cte > 0:
        with st.expander(
            f"💳 Detalle CTA CTE por camión — Total: {ars(tot_cta_cte)}",
            expanded=False
        ):
            st.caption(
                "Desglose de clientes CTA CTE por camión. "
                "La suma de todos los montos debe coincidir con el total E)."
            )
            grand_check = 0.0
            has_detail  = False
            for _, mrow in df_main.iterrows():
                detalle = mrow.get("CtaCteDetalle") or []
                if not isinstance(detalle, list) or not detalle:
                    continue
                has_detail = True
                cam_lbl = f"🚛 {int(mrow['idCns'])} — {mrow['dsCns']}"
                cam_tot = sum(d["monto"] for d in detalle)
                st.markdown(f"**{cam_lbl}** — subtotal CTA CTE: **{ars(cam_tot)}**")
                df_det = pd.DataFrame(detalle)[["codigo", "nombre", "monto"]]
                df_det.columns = ["Código", "Cliente", "Monto ($)"]
                df_det["Monto ($)"] = df_det["Monto ($)"].apply(lambda v: f"$ {v:,.0f}".replace(",", "."))
                st.dataframe(df_det, use_container_width=True, hide_index=True)
                grand_check += cam_tot

            if has_detail:
                st.markdown(f"---")
                st.markdown(
                    f"**✅ Total CTA CTE** (suma clientes): **{ars(grand_check)}** "
                    f"| Total E): **{ars(tot_cta_cte)}** "
                    + ("✅ Coinciden" if abs(grand_check - tot_cta_cte) < 1 else "⚠️ Diferencia detectada")
                )
            else:
                st.info("Sin detalle disponible — verificá que el ANR esté cargado.")

    st.divider()

    # ── EXPORTAR ──────────────────────────────────────────────────────────────
    st.markdown("#### 📥 Exportar")
    exp1, exp2 = st.columns(2)

    with exp1:
        if st.button("📊 Generar Excel", type="primary", key="cierre_xlsx"):
            try:
                xls_bytes = _cierre_excel(df_main, totales, mutual_monto, fecha_str, cobro_anticipado)
                fname = f"{_fecha_slug}_cierre_financiero_bkcc.xlsx"
                st.download_button(
                    label=f"⬇️ Descargar {fname}",
                    data=xls_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cierre_dl_xlsx",
                )
            except Exception as e:
                st.error(f"❌ Error generando Excel: {e}")
                with st.expander("Stack trace"):
                    import traceback
                    st.code(traceback.format_exc())

    with exp2:
        if st.button("📄 Generar PDF (landscape)", type="primary", key="cierre_pdf_btn"):
            try:
                pdf_bytes = _cierre_pdf(df_main, totales, mutual_monto, fecha_str, cobro_anticipado)
                fname = f"{_fecha_slug}_cierre_financiero_bkcc.pdf"
                st.download_button(
                    label=f"⬇️ Descargar {fname}",
                    data=pdf_bytes,
                    file_name=fname,
                    mime="application/pdf",
                    key="cierre_dl_pdf",
                )
            except Exception as e:
                st.error(f"❌ Error generando PDF: {e}")
                with st.expander("Stack trace"):
                    import traceback
                    st.code(traceback.format_exc())

    # ═══════════════════════════════════════════════════════════════════════════
    # SECCIÓN D+1 — CIERRE ACTUALIZADO CON RECHAZOS REALES
    # ═══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 📅 Cierre D+1 — Actualizado con Rechazos Reales")
    st.caption(
        "Al día siguiente del reparto, subí el **SR Actualizado** (col F = TotVal real cobrado "
        "por el chofer) y el **ANR -1** (ANR del día anterior, para CTA CTE). "
        "La app recalcula el cierre con lo que efectivamente ingresó."
    )

    sr_d1_file   = st.session_state.get("cierre_sr_d1")
    anr_m1_file  = st.session_state.get("cierre_anr_m1")

    if not sr_d1_file:
        st.info("⬅️ Subí el **SR D+1.xlsx** en la pestaña **📁 Archivos** para generar el Cierre Actualizado.")
    else:
        # Indicador visual de fuente CTA CTE
        if anr_m1_file:
            st.success("✅ **ANR -1 cargado** — CTA CTE se recalcula desde el ANR del día anterior.")
        else:
            st.warning(
                "⚠️ **Sin ANR -1** — CTA CTE tomado del cierre del día original. "
                "Para mayor precisión, subí el **ANR -1.xlsx** en 📁 Archivos."
            )

        try:
            df_sr_d1 = _cierre_load_sr(sr_d1_file)
        except Exception as e:
            st.error(f"❌ Error leyendo SR D+1.xlsx: {e}")
            df_sr_d1 = pd.DataFrame()

        # Cargar ANR -1 si está disponible
        df_anr_m1 = pd.DataFrame()
        if anr_m1_file:
            try:
                df_anr_m1 = _cierre_load_anr(anr_m1_file)
            except Exception as e:
                st.warning(f"⚠️ Error leyendo ANR -1.xlsx: {e}. CTA CTE tomado del cierre original.")

        if not df_sr_d1.empty:
            fecha_d1_str = f"{fecha_str} (D+1)"
            _fecha_slug_d1 = _fecha_slug

            # Construir cierre D+1 con ANR -1 para CTA CTE
            df_d1 = _build_cierre_d1(df_main, df_sr_d1, cta_cte_list_clean, df_anr, df_anr_m1)

            # Totales D+1
            d1_chess_orig   = df_d1["TotValOrig"].sum()
            d1_real         = df_d1["TotValReal"].sum()
            d1_cta_cte      = df_d1["CtaCte"].sum()
            d1_rechazos     = df_d1["RechazoReal"].sum()
            d1_neto         = df_d1["NetoReal"].sum()
            d1_pct_rech     = (d1_rechazos / d1_chess_orig * 100) if d1_chess_orig > 0 else 0.0

            # KPIs
            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric("📦 Preventa Original",  ars(d1_chess_orig))
            k2.metric("💵 Cobrado Real",        ars(d1_real))
            k3.metric("🔴 CTA CTE",             ars(d1_cta_cte))
            k4.metric("❌ Rechazos Reales",     ars(d1_rechazos),
                      delta=f"{d1_pct_rech:.1f}%" if d1_rechazos > 0 else None, delta_color="inverse")
            k5.metric("💰 Neto Real a Ingresar", ars(d1_neto))

            st.markdown("#### 🚛 Detalle D+1 por Camión")

            def style_neto_d1(val):
                color = "#1a7a4a" if val >= 0 else "#b91c1c"
                return f"color: {color}; font-weight: bold"

            def style_rech_d1(val):
                return "color: #b91c1c;" if val > 0 else ""

            df_d1_show = pd.DataFrame({
                "Camión":          df_d1.apply(lambda r: f"{int(r['idCns'])} — {r['dsCns']}", axis=1),
                "Preventa":        df_d1["TotValOrig"],
                "Cobrado Real":    df_d1["TotValReal"],
                "CTA CTE":         df_d1["CtaCte"],
                "Rechazo Real":    df_d1["RechazoReal"],
                "Neto Real":       df_d1["NetoReal"],
            })

            def style_cta_d1(val):
                return "color: #b91c1c; font-weight: bold" if val > 0 else ""

            _use_map_d1 = hasattr(pd.io.formats.style.Styler, "map")
            _base_d1 = df_d1_show.style.format({
                "Preventa":     "$ {:,.0f}",
                "Cobrado Real": "$ {:,.0f}",
                "CTA CTE":      "$ {:,.0f}",
                "Rechazo Real": "$ {:,.0f}",
                "Neto Real":    "$ {:,.0f}",
            })
            if _use_map_d1:
                styled_d1 = (_base_d1
                    .map(style_neto_d1, subset=["Neto Real"])
                    .map(style_rech_d1, subset=["Rechazo Real"])
                    .map(style_cta_d1,  subset=["CTA CTE"])
                    .set_properties(**{"text-align": "right"},
                                    subset=["Preventa", "Cobrado Real", "CTA CTE", "Rechazo Real", "Neto Real"]))
            else:
                styled_d1 = (_base_d1
                    .applymap(style_neto_d1, subset=["Neto Real"])
                    .applymap(style_rech_d1, subset=["Rechazo Real"])
                    .applymap(style_cta_d1,  subset=["CTA CTE"])
                    .set_properties(**{"text-align": "right"},
                                    subset=["Preventa", "Cobrado Real", "CTA CTE", "Rechazo Real", "Neto Real"]))

            st.dataframe(styled_d1, use_container_width=True, hide_index=True, height=460)

            # Resumen D+1
            st.markdown("#### 📊 Resumen D+1")
            res_d1 = {
                "Concepto": [
                    "D) Preventa original del día",
                    "D') Cobrado real por choferes",
                    "E) CTA CTE (descuento)",
                    "G) Rechazos reales",
                    "H) Neto real a ingresar",
                    "% Rechazo real",
                ],
                "Monto": [
                    f"$ {d1_chess_orig:,.0f}",
                    f"$ {d1_real:,.0f}",
                    f"$ {d1_cta_cte:,.0f}",
                    f"$ {d1_rechazos:,.0f}",
                    f"$ {d1_neto:,.0f}",
                    f"{d1_pct_rech:.1f}%",
                ],
            }
            st.dataframe(pd.DataFrame(res_d1), use_container_width=True, hide_index=True)

            # Exportar D+1
            st.markdown("#### 📥 Exportar D+1")
            d1_exp1, d1_exp2 = st.columns(2)

            totales_d1 = {
                "tot_chess_orig": d1_chess_orig,
                "tot_real":       d1_real,
                "tot_cta_cte":    d1_cta_cte,
                "tot_rechazos":   d1_rechazos,
                "tot_neto":       d1_neto,
                "pct_rech":       d1_pct_rech,
            }

            with d1_exp1:
                if st.button("📊 Generar Excel D+1", type="primary", key="d1_xlsx_btn"):
                    try:
                        xls_d1 = _cierre_d1_excel(df_d1, totales_d1, fecha_d1_str)
                        fname_d1 = f"{_fecha_slug_d1}_cierre_financiero_bkcc_d1.xlsx"
                        st.download_button(
                            label=f"⬇️ Descargar {fname_d1}",
                            data=xls_d1,
                            file_name=fname_d1,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="d1_dl_xlsx",
                        )
                    except Exception as e:
                        st.error(f"❌ Error generando Excel D+1: {e}")
                        with st.expander("Stack trace"):
                            import traceback
                            st.code(traceback.format_exc())

            with d1_exp2:
                if st.button("📄 Generar PDF D+1 (landscape)", type="primary", key="d1_pdf_btn"):
                    try:
                        pdf_d1 = _cierre_d1_pdf(df_d1, totales_d1, fecha_d1_str)
                        fname_d1 = f"{_fecha_slug_d1}_cierre_financiero_bkcc_d1.pdf"
                        st.download_button(
                            label=f"⬇️ Descargar {fname_d1}",
                            data=pdf_d1,
                            file_name=fname_d1,
                            mime="application/pdf",
                            key="d1_dl_pdf",
                        )
                    except Exception as e:
                        st.error(f"❌ Error generando PDF D+1: {e}")
                        with st.expander("Stack trace"):
                            import traceback
                            st.code(traceback.format_exc())


def _build_cierre_d1(df_orig: pd.DataFrame, df_sr_d1: pd.DataFrame,
                     cta_cte_list: list, df_anr: pd.DataFrame,
                     df_anr_m1: pd.DataFrame = None) -> pd.DataFrame:
    """
    Construye la tabla de Cierre D+1.
    df_orig:    DataFrame del cierre original (idCns, dsCns, TotVal, CtaCte, ...)
    df_sr_d1:   SR del día siguiente con TotVal real cobrado por el chofer.
    df_anr_m1:  ANR del día anterior (D-1 desde la perspectiva del cierre D+1).
                Si se provee, la CTA CTE se recalcula desde ahí (cliente × TRANSPORTE).
                Si es None o vacío, cae back al valor del cierre original.
    Los rechazos reales = TotValOrig - TotValReal (lo que no se cobró).
    """
    cta_cte_codigos = {int(c["codigo"]) for c in cta_cte_list}
    nombre_map      = {int(c["codigo"]): c["nombre"] for c in cta_cte_list}

    # Índice del SR D+1 por idCns
    sr_d1_idx = {int(r["idCns"]): float(r["TotVal"]) for _, r in df_sr_d1.iterrows()}

    # Bandera: ¿tenemos ANR -1 con datos utilizables?
    usar_anr_m1 = (
        df_anr_m1 is not None
        and not df_anr_m1.empty
        and "CLIENTE" in df_anr_m1.columns
        and "TRANSPORTE" in df_anr_m1.columns
        and "IMPORTE_NETO" in df_anr_m1.columns
    )

    rows = []
    for _, row in df_orig.iterrows():
        cns      = int(row["idCns"])
        nombre   = str(row["dsCns"])
        tot_orig = float(row["TotVal"])

        # CTA CTE: recalcular desde ANR -1 si está disponible
        if usar_anr_m1:
            cam_rows = df_anr_m1[df_anr_m1["TRANSPORTE"] == cns]
            cta_detail_d1 = []
            if not cam_rows.empty:
                cta_rows_m1 = cam_rows[cam_rows["CLIENTE"].isin(cta_cte_codigos)]
                if not cta_rows_m1.empty:
                    agg = cta_rows_m1.groupby("CLIENTE")["IMPORTE_NETO"].sum()
                    for cod, monto in agg.items():
                        if monto > 0:
                            cta_detail_d1.append({
                                "codigo": int(cod),
                                "nombre": nombre_map.get(int(cod), str(int(cod))),
                                "monto":  float(monto),
                            })
            cta_cte_m = sum(d["monto"] for d in cta_detail_d1)
        else:
            # Fallback: reutilizar CTA CTE del cierre original
            cta_cte_m     = float(row["CtaCte"])
            cta_detail_d1 = (row.get("CtaCteDetalle") or []) if "CtaCteDetalle" in row.index else []

        # TotVal real: si el camión no aparece en D+1, se asume cobró todo
        tot_real     = sr_d1_idx.get(cns, tot_orig)

        # Rechazo real: diferencia; si cobró más (redondeos), rechazo = 0
        rechazo_real = max(0.0, tot_orig - tot_real)

        # Neto real = cobrado - CTA CTE
        neto_real = tot_real - cta_cte_m

        rows.append({
            "idCns":           cns,
            "dsCns":           nombre,
            "TotValOrig":      tot_orig,
            "TotValReal":      tot_real,
            "CtaCte":          cta_cte_m,
            "RechazoReal":     rechazo_real,
            "NetoReal":        neto_real,
            "CtaCteDetalle":   cta_detail_d1,
            "FuenteCtaCte":    "ANR -1" if usar_anr_m1 else "Cierre original",
        })

    return pd.DataFrame(rows)


def _cierre_d1_pdf(df_d1: pd.DataFrame, totales: dict, fecha_str: str) -> bytes:
    """PDF landscape del Cierre D+1."""
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas

    buf = io.BytesIO()
    PW, PH = rl_landscape(A4)
    M = 14 * mm
    c = rl_canvas.Canvas(buf, pagesize=rl_landscape(A4))

    DARK_BLUE  = rl_colors.HexColor("#1a3a6b")
    MED_BLUE   = rl_colors.HexColor("#2e5fa3")
    LIGHT_GRAY = rl_colors.HexColor("#F0F4FA")
    GREEN_OK   = rl_colors.HexColor("#1a7a4a")
    RED_NEG    = rl_colors.HexColor("#b91c1c")
    AMBER_BG   = rl_colors.HexColor("#FEF3C7")
    AMBER_TXT  = rl_colors.HexColor("#92400e")

    def fmt_ars(v):
        s = f"{abs(v):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"$ {s}" if v >= 0 else f"($ {s})"

    # HEADER
    y = PH - M
    c.setFillColor(DARK_BLUE)
    c.rect(M, y - 32, PW - 2*M, 34, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(M + 10, y - 20, "CIERRE FINANCIERO D+1 — RECHAZOS REALES")
    c.setFont("Helvetica", 9)
    c.drawRightString(PW - M - 8, y - 12, f"Beccacece Hnos SA  |  {fecha_str}")
    c.drawRightString(PW - M - 8, y - 23, "Distribuidor CMQ — ABInBev")
    y -= 46

    # TABLA
    usable_w = PW - 2 * M
    col_labels = ["Camión", "Preventa", "Cobrado Real", "CTA CTE", "Rechazo Real", "Neto Real"]
    col_w = [
        usable_w * 0.27,
        usable_w * 0.145,
        usable_w * 0.145,
        usable_w * 0.12,
        usable_w * 0.145,
        usable_w * 0.175,
    ]
    row_h = 13
    hdr_h = 16

    c.setFillColor(MED_BLUE)
    c.rect(M, y - hdr_h, usable_w, hdr_h, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 9)
    x_cur = M + 5
    for lbl, cw in zip(col_labels, col_w):
        c.drawString(x_cur, y - hdr_h + 5, lbl)
        x_cur += cw
    y -= hdr_h + 2

    SUB_H_D1  = 11
    SUB_FILL_D1 = rl_colors.HexColor("#FFF3F3")
    INDENT_D1   = 16

    for i, row in df_d1.iterrows():
        bg = LIGHT_GRAY if i % 2 == 0 else rl_colors.white
        c.setFillColor(bg)
        c.rect(M, y - row_h, usable_w, row_h, fill=1, stroke=0)

        vals = [
            f"{int(row['idCns'])} — {row['dsCns']}",
            fmt_ars(row["TotValOrig"]),
            fmt_ars(row["TotValReal"]),
            fmt_ars(row["CtaCte"]) if row["CtaCte"] > 0 else "—",
            fmt_ars(row["RechazoReal"]) if row["RechazoReal"] > 0 else "—",
            fmt_ars(row["NetoReal"]),
        ]
        x_cur = M + 5
        for vi, (val, cw) in enumerate(zip(vals, col_w)):
            if vi == 5:
                c.setFillColor(GREEN_OK if row["NetoReal"] >= 0 else RED_NEG)
                c.setFont("Helvetica-Bold", 8.5)
            elif vi == 4 and val != "—":
                c.setFillColor(RED_NEG)
                c.setFont("Helvetica", 8.5)
            elif vi == 3 and val != "—":
                c.setFillColor(RED_NEG)
                c.setFont("Helvetica", 8.5)
            else:
                c.setFillColor(rl_colors.black)
                c.setFont("Helvetica", 8.5)
            c.drawString(x_cur, y - row_h + 4, val)
            x_cur += cw
        y -= row_h

        # Sub-filas CTA CTE integradas
        detalle_cte_d1 = row.get("CtaCteDetalle") or []
        if isinstance(detalle_cte_d1, list) and detalle_cte_d1:
            for d in detalle_cte_d1:
                if y < M + 60:
                    c.showPage()
                    c.setFillColor(DARK_BLUE)
                    c.rect(M, PH - 16, PW - 2*M, 16, fill=1, stroke=0)
                    c.setFillColor(rl_colors.white)
                    c.setFont("Helvetica-Bold", 8)
                    c.drawString(M + 4, PH - 11, f"CIERRE FINANCIERO D+1 — {fecha_str} (cont.)")
                    y = PH - 26
                c.setFillColor(SUB_FILL_D1)
                c.rect(M, y - SUB_H_D1, usable_w, SUB_H_D1, fill=1, stroke=0)
                c.setFillColor(RED_NEG)
                c.rect(M, y - SUB_H_D1, 3, SUB_H_D1, fill=1, stroke=0)
                nombre_d1 = d.get("nombre", str(d.get("codigo", "")))
                cod_d1    = d.get("codigo", "")
                c.setFillColor(RED_NEG)
                c.setFont("Helvetica-Bold", 7.5)
                c.drawString(M + INDENT_D1, y - SUB_H_D1 + 3,
                             f"\u21b3 {cod_d1}  {nombre_d1}")
                # Columna CTA CTE en D+1 es la 4ta (idx 3): col_w[0]+col_w[1]+col_w[2]
                x_cta_d1 = M + col_w[0] + col_w[1] + col_w[2]
                c.drawString(x_cta_d1 + 3, y - SUB_H_D1 + 3,
                             fmt_ars(float(d.get("monto", 0))))
                y -= SUB_H_D1

    # FILA TOTAL
    y -= 4
    c.setFillColor(DARK_BLUE)
    c.rect(M, y - 16, usable_w, 16, fill=1, stroke=0)
    c.setFillColor(rl_colors.white)
    c.setFont("Helvetica-Bold", 9)
    tot_vals = [
        "TOTAL",
        fmt_ars(totales["tot_chess_orig"]),
        fmt_ars(totales["tot_real"]),
        fmt_ars(totales["tot_cta_cte"]) if totales["tot_cta_cte"] > 0 else "—",
        fmt_ars(totales["tot_rechazos"]) if totales["tot_rechazos"] > 0 else "—",
        fmt_ars(totales["tot_neto"]),
    ]
    x_cur = M + 5
    for tv, cw in zip(tot_vals, col_w):
        c.drawString(x_cur, y - 12, tv)
        x_cur += cw
    y -= 24

    # RESUMEN
    resumen_h = 90
    c.setFillColor(rl_colors.HexColor("#F0F4FA"))
    c.rect(M, y - resumen_h, usable_w, resumen_h + 4, fill=1, stroke=0)
    c.setFillColor(DARK_BLUE)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(M + 8, y - 14, "RESUMEN D+1")

    resumen_items = [
        ("D) Preventa original del dia",     totales["tot_chess_orig"], DARK_BLUE),
        ("D') Cobrado real por choferes",    totales["tot_real"],        GREEN_OK),
        ("E) CTA CTE (descuento)",           totales["tot_cta_cte"],     RED_NEG if totales["tot_cta_cte"] > 0 else rl_colors.black),
        ("G) Rechazos reales",               totales["tot_rechazos"],    RED_NEG if totales["tot_rechazos"] > 0 else rl_colors.black),
        ("H) Neto real a ingresar",          totales["tot_neto"],        GREEN_OK),
        (f"% Rechazo real",                  None,                       RED_NEG if totales["pct_rech"] > 5 else GREEN_OK),
    ]

    half_w = usable_w / 2 - 10
    yi = y - 30
    for label, val, col_txt in resumen_items:
        c.setFillColor(rl_colors.black)
        c.setFont("Helvetica", 8.5)
        c.drawString(M + 8, yi, label)
        c.setFillColor(col_txt)
        c.setFont("Helvetica-Bold", 9)
        if val is None:
            c.drawRightString(M + half_w, yi, f"{totales['pct_rech']:.1f}%")
        else:
            c.drawRightString(M + half_w, yi, fmt_ars(val))
        yi -= 13

    # FOOTER pág 1
    c.setFillColor(rl_colors.HexColor("#555555"))
    c.setFont("Helvetica", 6.5)
    c.drawCentredString(
        PW / 2, M - 4,
        f"Beccacece Hnos SA  ·  Cierre D+1 generado: {fecha_str}  ·  Picking Orchestrator v{APP_VERSION}"
    )

    # ── PÁGINA 2: DETALLE CTA CTE POR CAMIÓN ─────────────────────────────────
    cta_rows_pdf = []
    if "CtaCteDetalle" in df_d1.columns:
        for _, row in df_d1.iterrows():
            detalle = row.get("CtaCteDetalle") or []
            if isinstance(detalle, list) and detalle:
                for d in detalle:
                    cta_rows_pdf.append({
                        "camion_id":   int(row["idCns"]),
                        "camion_ds":   str(row["dsCns"]),
                        "codigo":      d.get("codigo", ""),
                        "nombre":      d.get("nombre", ""),
                        "monto":       float(d.get("monto", 0)),
                    })

    if cta_rows_pdf:
        c.showPage()
        # Mini-header pág 2
        c.setFillColor(DARK_BLUE)
        c.rect(M, PH - 16, PW - 2*M, 16, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(M + 4, PH - 11, f"CIERRE FINANCIERO D+1 — DETALLE CTA CTE — {fecha_str}")
        c.setFont("Helvetica", 7)
        c.drawRightString(PW - M - 4, PH - 11, "Beccacece Hnos SA")
        y2 = PH - 26

        col_labels2 = ["Camión",        "Cód.",    "Cliente (CTA CTE)",            "Monto ($)"]
        col_w2      = [
            usable_w * 0.28,
            usable_w * 0.10,
            usable_w * 0.42,
            usable_w * 0.20,
        ]
        hdr2_h = 15
        c.setFillColor(MED_BLUE)
        c.rect(M, y2 - hdr2_h, usable_w, hdr2_h, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 9)
        x2 = M + 5
        for lbl2, cw2 in zip(col_labels2, col_w2):
            c.drawString(x2, y2 - hdr2_h + 4, lbl2)
            x2 += cw2
        y2 -= hdr2_h + 2

        current_cam = None
        cam_subtotal = 0.0
        grand_total  = 0.0
        row_idx      = 0

        def _flush_cam_subtotal(y_pos, cam_label, subtot):
            """Dibuja fila de subtotal de camión."""
            c.setFillColor(rl_colors.HexColor("#D9E4F5"))
            c.rect(M, y_pos - 12, usable_w, 12, fill=1, stroke=0)
            c.setFillColor(DARK_BLUE)
            c.setFont("Helvetica-Bold", 8)
            c.drawString(M + 5, y_pos - 9, f"Subtotal {cam_label}")
            c.drawRightString(PW - M - 5, y_pos - 9, fmt_ars(subtot))
            return y_pos - 16

        for cr in cta_rows_pdf:
            cam_label = f"{cr['camion_id']} — {cr['camion_ds']}"
            if cam_label != current_cam:
                if current_cam is not None:
                    # Subtotal camión anterior
                    if y2 < M + 50:
                        c.showPage()
                        c.setFillColor(DARK_BLUE)
                        c.rect(M, PH - 16, PW - 2*M, 16, fill=1, stroke=0)
                        c.setFillColor(rl_colors.white)
                        c.setFont("Helvetica-Bold", 8)
                        c.drawString(M + 4, PH - 11, f"DETALLE CTA CTE D+1 — {fecha_str} (cont.)")
                        y2 = PH - 26
                    y2 = _flush_cam_subtotal(y2, current_cam, cam_subtotal)
                    cam_subtotal = 0.0
                current_cam = cam_label
                # Encabezado de camión
                if y2 < M + 50:
                    c.showPage()
                    c.setFillColor(DARK_BLUE)
                    c.rect(M, PH - 16, PW - 2*M, 16, fill=1, stroke=0)
                    c.setFillColor(rl_colors.white)
                    c.setFont("Helvetica-Bold", 8)
                    c.drawString(M + 4, PH - 11, f"DETALLE CTA CTE D+1 — {fecha_str} (cont.)")
                    y2 = PH - 26
                c.setFillColor(rl_colors.HexColor("#EEF3FB"))
                c.rect(M, y2 - 13, usable_w, 13, fill=1, stroke=0)
                c.setFillColor(MED_BLUE)
                c.setFont("Helvetica-Bold", 8.5)
                c.drawString(M + 5, y2 - 10, cam_label)
                y2 -= 15

            # Fila cliente
            bg2 = LIGHT_GRAY if row_idx % 2 == 0 else rl_colors.white
            c.setFillColor(bg2)
            c.rect(M, y2 - 12, usable_w, 12, fill=1, stroke=0)
            c.setFillColor(rl_colors.black)
            c.setFont("Helvetica", 8)
            x2 = M + 5
            sub_vals = ["", str(cr["codigo"]), cr["nombre"], fmt_ars(cr["monto"])]
            for svi, (sv, scw) in enumerate(zip(sub_vals, col_w2)):
                if svi == 3:
                    c.setFillColor(RED_NEG)
                    c.setFont("Helvetica-Bold", 8)
                    c.drawRightString(x2 + scw - 5, y2 - 9, sv)
                else:
                    c.setFillColor(rl_colors.black)
                    c.setFont("Helvetica", 8)
                    c.drawString(x2 + 2, y2 - 9, sv)
                x2 += scw
            y2 -= 13
            cam_subtotal += cr["monto"]
            grand_total  += cr["monto"]
            row_idx      += 1

            if y2 < M + 50:
                c.showPage()
                c.setFillColor(DARK_BLUE)
                c.rect(M, PH - 16, PW - 2*M, 16, fill=1, stroke=0)
                c.setFillColor(rl_colors.white)
                c.setFont("Helvetica-Bold", 8)
                c.drawString(M + 4, PH - 11, f"DETALLE CTA CTE D+1 — {fecha_str} (cont.)")
                y2 = PH - 26

        # Último subtotal
        if current_cam:
            y2 = _flush_cam_subtotal(y2, current_cam, cam_subtotal)

        # Total general CTA CTE
        y2 -= 4
        c.setFillColor(DARK_BLUE)
        c.rect(M, y2 - 14, usable_w, 14, fill=1, stroke=0)
        c.setFillColor(rl_colors.white)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(M + 6, y2 - 10, "TOTAL GENERAL CTA CTE")
        c.drawRightString(PW - M - 5, y2 - 10, fmt_ars(grand_total))

        # Footer pág 2
        c.setFillColor(rl_colors.HexColor("#555555"))
        c.setFont("Helvetica", 6.5)
        c.drawCentredString(
            PW / 2, M - 4,
            f"Beccacece Hnos SA  ·  Cierre D+1 generado: {fecha_str}  ·  Picking Orchestrator v{APP_VERSION}"
        )

    c.save()
    return buf.getvalue()


def _cierre_d1_excel(df_d1: pd.DataFrame, totales: dict, fecha_str: str) -> bytes:
    """Excel del Cierre D+1."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cierre D+1"

    hdr_fill   = PatternFill("solid", fgColor="1a3a6b")
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    tot_fill   = PatternFill("solid", fgColor="2e5fa3")
    tot_font   = Font(color="FFFFFF", bold=True, size=10)
    alt_fill   = PatternFill("solid", fgColor="EEF3FB")
    green_font = Font(color="1a7a4a", bold=True)
    red_font   = Font(color="b91c1c", bold=True)
    bold_font  = Font(bold=True)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_al  = Alignment(horizontal="center", vertical="center")
    right_al   = Alignment(horizontal="right", vertical="center")
    left_al    = Alignment(horizontal="left", vertical="center")
    money_fmt  = '#,##0'

    ws.merge_cells("A1:G1")
    ws["A1"] = f"CIERRE FINANCIERO D+1 — RECHAZOS REALES — {fecha_str}"
    ws["A1"].font = Font(bold=True, size=13, color="1a3a6b")
    ws["A1"].alignment = center_al
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    ws["A2"] = "Beccacece Hnos SA — Distribuidor CMQ / ABInBev"
    ws["A2"].font = Font(size=9, color="555555", italic=True)
    ws["A2"].alignment = center_al

    headers = ["ID", "Camión", "Preventa ($)", "Cobrado Real ($)", "CTA CTE ($)", "Rechazo Real ($)", "Neto Real ($)"]
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center_al
        cell.border = border
    ws.row_dimensions[4].height = 16

    sub_fill_d1 = PatternFill("solid", fgColor="FFF0F0")
    red_left_d1 = Font(color="b91c1c", size=8, italic=True)

    r_d1 = 5   # fila dinámica
    for ri, row in df_d1.iterrows():
        vals = [
            int(row["idCns"]),
            row["dsCns"],
            row["TotValOrig"],
            row["TotValReal"],
            row["CtaCte"] if row["CtaCte"] > 0 else 0,
            row["RechazoReal"] if row["RechazoReal"] > 0 else 0,
            row["NetoReal"],
        ]
        fill_row = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r_d1, column=ci, value=val)
            cell.border = border
            cell.fill = fill_row
            if ci == 1:
                cell.alignment = center_al
            elif ci == 2:
                cell.alignment = left_al
                cell.font = bold_font
            else:
                cell.alignment = right_al
                cell.number_format = money_fmt
                if ci == 7:
                    cell.font = green_font if row["NetoReal"] >= 0 else red_font
                elif ci == 6 and row["RechazoReal"] > 0:
                    cell.font = red_font
                elif ci == 5 and row["CtaCte"] > 0:
                    cell.font = red_font
        ws.row_dimensions[r_d1].height = 14
        r_d1 += 1

        # Sub-filas CTA CTE integradas
        detalle_d1 = row.get("CtaCteDetalle") or []
        if isinstance(detalle_d1, list):
            for d in detalle_d1:
                monto_dd = float(d.get("monto", 0))
                if monto_dd <= 0:
                    continue
                ws.cell(row=r_d1, column=1, value="").fill = sub_fill_d1
                cell_nom = ws.cell(row=r_d1, column=2,
                                   value=f"\u21b3 {d.get('codigo','')}  {d.get('nombre','')}")
                cell_nom.font      = red_left_d1
                cell_nom.fill      = sub_fill_d1
                cell_nom.alignment = left_al
                for ci in range(3, 8):
                    ws.cell(row=r_d1, column=ci).fill = sub_fill_d1
                cell_m = ws.cell(row=r_d1, column=5, value=monto_dd)
                cell_m.font          = Font(color="b91c1c", size=8, bold=True)
                cell_m.fill          = sub_fill_d1
                cell_m.alignment     = right_al
                cell_m.number_format = money_fmt
                ws.row_dimensions[r_d1].height = 12
                r_d1 += 1

    tr = r_d1   # r_d1 ya apunta a la siguiente fila libre
    tot_vals = ["", "TOTAL",
                totales["tot_chess_orig"], totales["tot_real"],
                totales["tot_cta_cte"], totales["tot_rechazos"], totales["tot_neto"]]
    for ci, val in enumerate(tot_vals, 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.fill = tot_fill
        cell.font = tot_font
        cell.border = border
        cell.alignment = right_al if ci > 2 else center_al
        if ci > 2:
            cell.number_format = money_fmt
    ws.row_dimensions[tr].height = 16

    # Resumen
    rs = tr + 2
    pct = totales["pct_rech"]
    resumen_rows = [
        ("D) Preventa original del día",      totales["tot_chess_orig"], None),
        ("D') Cobrado real por choferes",     totales["tot_real"],       "1a7a4a"),
        ("E) CTA CTE (total a descontar)",    totales["tot_cta_cte"],    "b91c1c" if totales["tot_cta_cte"] > 0 else None),
        ("G) Rechazos reales",                totales["tot_rechazos"],   "b91c1c" if totales["tot_rechazos"] > 0 else None),
        ("H) Neto real a ingresar",           totales["tot_neto"],       "1a7a4a"),
        ("% Rechazo real",                    pct / 100,                 "b91c1c" if pct > 5 else "1a7a4a"),
    ]

    ws.merge_cells(f"A{rs}:G{rs}")
    ws[f"A{rs}"] = "RESUMEN D+1"
    ws[f"A{rs}"].font = Font(bold=True, size=10, color="1a3a6b")
    ws[f"A{rs}"].fill = PatternFill("solid", fgColor="D9E4F5")
    ws[f"A{rs}"].alignment = center_al
    rs += 1

    for label, val, color in resumen_rows:
        ws.merge_cells(f"A{rs}:E{rs}")
        ws[f"A{rs}"] = label
        ws[f"A{rs}"].font = Font(size=9)
        ws[f"A{rs}"].alignment = left_al
        ws.merge_cells(f"F{rs}:G{rs}")
        cell_val = ws[f"F{rs}"]
        cell_val.value = val
        cell_val.alignment = right_al
        if label.startswith("%"):
            cell_val.number_format = "0.0%"
        else:
            cell_val.number_format = money_fmt
        cell_val.font = Font(bold=True, color=color, size=9) if color else Font(bold=True, size=9)
        ws.row_dimensions[rs].height = 13
        rs += 1

    col_widths = [8, 24, 22, 22, 18, 22, 26]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── HOJA 2: CTA CTE DETALLE ───────────────────────────────────────────────
    if "CtaCteDetalle" in df_d1.columns:
        ws2 = wb.create_sheet(title="CTA CTE Detalle")

        ws2.merge_cells("A1:E1")
        ws2["A1"] = f"DETALLE CTA CTE D+1 — {fecha_str}"
        ws2["A1"].font = Font(bold=True, size=12, color="1a3a6b")
        ws2["A1"].alignment = center_al
        ws2.row_dimensions[1].height = 20

        ws2.merge_cells("A2:E2")
        ws2["A2"] = "Beccacece Hnos SA — Distribuidor CMQ / ABInBev"
        ws2["A2"].font = Font(size=9, color="555555", italic=True)
        ws2["A2"].alignment = center_al

        hdr2 = ["ID Camión", "Camión", "Cód. Cliente", "Cliente (CTA CTE)", "Monto ($)"]
        for ci2, h2 in enumerate(hdr2, 1):
            cell2 = ws2.cell(row=4, column=ci2, value=h2)
            cell2.fill = hdr_fill
            cell2.font = hdr_font
            cell2.alignment = center_al
            cell2.border = border
        ws2.row_dimensions[4].height = 15

        r2 = 5
        grand_cte = 0.0
        cam_fill_h  = PatternFill("solid", fgColor="D9E4F5")   # sub-header camión
        cam_sub_fill = PatternFill("solid", fgColor="EEF3FB")  # subtotal camión

        for _, mrow in df_d1.iterrows():
            detalle2 = mrow.get("CtaCteDetalle") or []
            if not isinstance(detalle2, list) or not detalle2:
                continue
            cam_id  = int(mrow["idCns"])
            cam_ds  = str(mrow["dsCns"])
            cam_tot = sum(float(d.get("monto", 0)) for d in detalle2)

            # Fila header de camión
            ws2.merge_cells(f"A{r2}:B{r2}")
            ws2[f"A{r2}"] = f"{cam_id} — {cam_ds}"
            ws2[f"A{r2}"].font = Font(bold=True, size=9, color="1a3a6b")
            ws2[f"A{r2}"].fill = cam_fill_h
            ws2[f"A{r2}"].alignment = left_al
            for ci2 in range(3, 6):
                ws2.cell(row=r2, column=ci2).fill = cam_fill_h
                ws2.cell(row=r2, column=ci2).border = border
            ws2[f"A{r2}"].border = border
            ws2.row_dimensions[r2].height = 14
            r2 += 1

            for d in detalle2:
                monto_d = float(d.get("monto", 0))
                row2_vals = [cam_id, cam_ds, d.get("codigo", ""), d.get("nombre", ""), monto_d]
                row2_fill = alt_fill if (r2 % 2 == 0) else PatternFill()
                for ci2, val2 in enumerate(row2_vals, 1):
                    cell2 = ws2.cell(row=r2, column=ci2, value=val2)
                    cell2.fill = row2_fill
                    cell2.border = border
                    if ci2 == 5:
                        cell2.number_format = money_fmt
                        cell2.font = red_font
                        cell2.alignment = right_al
                    elif ci2 == 1:
                        cell2.alignment = center_al
                    elif ci2 == 3:
                        cell2.alignment = center_al
                    else:
                        cell2.alignment = left_al
                ws2.row_dimensions[r2].height = 13
                r2 += 1
                grand_cte += monto_d

            # Subtotal camión
            ws2.merge_cells(f"A{r2}:D{r2}")
            ws2[f"A{r2}"] = f"Subtotal {cam_id} — {cam_ds}"
            ws2[f"A{r2}"].font = Font(bold=True, size=9, color="1a3a6b")
            ws2[f"A{r2}"].fill = cam_sub_fill
            ws2[f"A{r2}"].alignment = right_al
            ws2[f"A{r2}"].border = border
            sub_cell = ws2.cell(row=r2, column=5, value=cam_tot)
            sub_cell.font = Font(bold=True, color="b91c1c", size=9)
            sub_cell.number_format = money_fmt
            sub_cell.fill = cam_sub_fill
            sub_cell.alignment = right_al
            sub_cell.border = border
            ws2.row_dimensions[r2].height = 13
            r2 += 1

        # Total general
        ws2.merge_cells(f"A{r2}:D{r2}")
        ws2[f"A{r2}"] = "TOTAL GENERAL CTA CTE"
        ws2[f"A{r2}"].font = Font(bold=True, size=10, color="FFFFFF")
        ws2[f"A{r2}"].fill = hdr_fill
        ws2[f"A{r2}"].alignment = right_al
        ws2[f"A{r2}"].border = border
        gtot_cell = ws2.cell(row=r2, column=5, value=grand_cte)
        gtot_cell.font = Font(bold=True, color="FFFFFF", size=10)
        gtot_cell.number_format = money_fmt
        gtot_cell.fill = hdr_fill
        gtot_cell.alignment = right_al
        gtot_cell.border = border
        ws2.row_dimensions[r2].height = 16

        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 24
        ws2.column_dimensions["C"].width = 14
        ws2.column_dimensions["D"].width = 38
        ws2.column_dimensions["E"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    main()


def _render_historico_tablas(df_h):
    import pandas as pd
    st.dataframe(df_h, use_container_width=True, hide_index=True)

